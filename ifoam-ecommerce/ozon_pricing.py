"""
ozon_pricing.py — Batch Ozon FBS price calculator

Usage:
  uv run --with openpyxl ozon_pricing.py                               # folder: Desktop/На сортировку 08.05
  uv run --with openpyxl ozon_pricing.py --folder "PATH/TO/FOLDER"
  uv run --with openpyxl ozon_pricing.py --file   "PATH/TO/FILE.xlsx"

For each row with weight/length/width/height filled:
  - calculates FBS logistics cost from package dimensions
  - finds recommended selling price at 12% margin
  - inserts "Цена продажи" and "Прибыль с единицы" right after "Цена закупки"

Output: subfolder "Цены" next to the source folder (or source file's parent).
"""

import argparse
import math
import sys
from pathlib import Path

import openpyxl

# ── Constants matching ozon_calculator.html ─────────────────────────────────

FBS_TIERS = [100, 300, 1500, 5000, 10000]
FBS_RATES = [0.14, 0.20, 0.44, 0.44, 0.44, 0.44]

LOG_FBS = [(0.5, 75), (1, 90), (2, 115), (5, 155), (10, 210),
           (15, 265), (20, 315), (25, 365), (30, 420), (50, 620)]

ACQ_PCT  = 0.015   # эквайринг
TAX_PCT  = 0.06    # УСН
RET_RATE = 0.03    # % возвратов
REVERSE  = 80      # стоимость обратной логистики, ₽
STORAGE  = 0       # хранение FBS = 0
OTHER    = 30      # упаковка/прочее, ₽
TARGET   = 0.12    # целевая маржа 12%


# ── Commission & logistics helpers ─────────────────────────────────────────

def get_fbs_rate(sell: float) -> float:
    for thresh, rate in zip(FBS_TIERS, FBS_RATES):
        if sell < thresh:
            return rate
    return FBS_RATES[-1]


def log_cost(weight_kg: float) -> float:
    for lim, cost in LOG_FBS:
        if weight_kg <= lim:
            return cost
    lim_last, cost_last = LOG_FBS[-1]
    return cost_last + math.ceil(weight_kg - lim_last) * 15


def calc_logistics(weight_g: float, length_mm: float, width_mm: float, height_mm: float) -> float:
    actual_kg = weight_g / 1000
    vol_kg = length_mm * width_mm * height_mm / 5_000_000
    bill_kg = max(actual_kg, vol_kg)
    return log_cost(bill_kg)


# ── Profit at given sell price ──────────────────────────────────────────────

def calc_profit(purchase: float, sell: float, logistics: float) -> float:
    comm_rate  = get_fbs_rate(sell)
    commission = sell * comm_rate
    acquiring  = sell * ACQ_PCT
    return_loss = RET_RATE * (logistics + REVERSE)
    proceeds   = sell - commission - acquiring - logistics
    tax        = max(0.0, proceeds) * TAX_PCT
    total_cost = purchase + commission + acquiring + logistics + return_loss + STORAGE + OTHER + tax
    return sell - total_cost


# ── Find recommended price at target margin ─────────────────────────────────

def find_rec_price(purchase: float, logistics: float, target: float = TARGET) -> int | None:
    for s in range(50, 500_001):
        profit = calc_profit(purchase, s, logistics)
        if s > 0 and profit / s >= target - 1e-6:
            return s
    return None


# ── Column detection ────────────────────────────────────────────────────────

def find_col(headers: list, keywords: list[str]) -> int | None:
    """Return 1-based column index of the first header matching any keyword (case-insensitive)."""
    for idx, h in enumerate(headers):
        if h is None:
            continue
        h_lower = str(h).lower()
        if any(kw.lower() in h_lower for kw in keywords):
            return idx + 1
    return None


# ── Process single file ─────────────────────────────────────────────────────

def process_file(src_path: Path, out_dir: Path) -> tuple[int, int]:
    """Returns (processed_rows, skipped_rows)."""
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    col_price  = find_col(headers, ['цена закупки', 'цена покупки', 'цена'])
    col_weight = find_col(headers, ['вес'])
    col_len    = find_col(headers, ['длина'])
    col_width  = find_col(headers, ['ширина'])
    col_height = find_col(headers, ['высота'])

    if not col_price:
        print(f"  [SKIP] no price col: {src_path.name}", flush=True)
        return 0, 0

    missing_dim = [n for n, c in [('weight', col_weight), ('length', col_len),
                                   ('width', col_width), ('height', col_height)] if not c]
    if missing_dim:
        print(f"  [SKIP] no dim cols {missing_dim}: {src_path.name}", flush=True)
        return 0, 0

    # Insert two new columns right after col_price
    insert_at = col_price + 1
    ws.insert_cols(insert_at, 2)

    ws.cell(row=1, column=insert_at).value     = 'Цена продажи'
    ws.cell(row=1, column=insert_at + 1).value = 'Прибыль с единицы'

    # Dimension columns shifted right by 2 after insert
    col_weight += 2
    col_len    += 2
    col_width  += 2
    col_height += 2

    processed = skipped = 0
    for row in range(2, ws.max_row + 1):
        purchase = ws.cell(row=row, column=col_price).value
        weight   = ws.cell(row=row, column=col_weight).value
        length   = ws.cell(row=row, column=col_len).value
        width    = ws.cell(row=row, column=col_width).value
        height   = ws.cell(row=row, column=col_height).value

        # Only process rows with all five values present and numeric
        try:
            purchase = float(purchase)
            weight   = float(weight)
            length   = float(length)
            width    = float(width)
            height   = float(height)
            if any(v <= 0 for v in [weight, length, width, height]):
                raise ValueError
        except (TypeError, ValueError):
            skipped += 1
            continue

        logistics  = calc_logistics(weight, length, width, height)
        sell_price = find_rec_price(purchase, logistics)

        if sell_price is None:
            skipped += 1
            continue

        profit_per_unit = round(calc_profit(purchase, sell_price, logistics), 2)

        ws.cell(row=row, column=insert_at).value     = sell_price
        ws.cell(row=row, column=insert_at + 1).value = profit_per_unit
        processed += 1

    out_path = out_dir / src_path.name
    wb.save(out_path)
    print(f"  OK  {processed:4d} prices | {src_path.name}", flush=True)
    return processed, skipped


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    default_folder = Path.home() / "Desktop" / "На сортировку 08.05"

    parser = argparse.ArgumentParser(description='Ozon FBS batch pricer')
    group = parser.add_mutually_exclusive_group()
    group.add_argument('--folder', type=Path, default=None,
                       help=f'Папка с Excel-файлами (по умолч. {default_folder})')
    group.add_argument('--file',   type=Path, default=None,
                       help='Путь к одному Excel-файлу')
    args = parser.parse_args()

    if args.file:
        src_path = args.file.resolve()
        if not src_path.exists():
            sys.exit(f'Файл не найден: {src_path}')
        files  = [src_path]
        out_dir = src_path.parent / 'Цены'
    else:
        folder = (args.folder or default_folder).resolve()
        if not folder.exists():
            sys.exit(f'Папка не найдена: {folder}')
        files   = sorted(folder.glob('*.xlsx'))
        out_dir = folder.parent / (folder.name + ' — Цены')

    out_dir.mkdir(parents=True, exist_ok=True)
    print(f'Output -> {out_dir}', flush=True)
    print(f'Files:  {len(files)}\n', flush=True)

    total_proc = total_skip = 0
    for f in files:
        if f.name.startswith('~$'):
            continue
        p, s = process_file(f, out_dir)
        total_proc += p
        total_skip += s

    print(f'\nDone: {total_proc} prices in {len(files)} files.', flush=True)


if __name__ == '__main__':
    main()
