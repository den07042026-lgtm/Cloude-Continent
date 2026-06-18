import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'Orange Standard 5кг'

YELLOW   = PatternFill('solid', fgColor='FFD966')
GREEN    = PatternFill('solid', fgColor='C6EFCE')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
BLUE     = PatternFill('solid', fgColor='BDD7EE')
GRAY     = PatternFill('solid', fgColor='F2F2F2')
ORANGE_F = PatternFill('solid', fgColor='F4B942')

def brd():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def c(ws, row, col, value, fill=None, font=None, align=None, nf=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.alignment = align or Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = brd()
    if nf: cell.number_format = nf
    return cell

left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
bold = Font(bold=True)

# ── Блок 1: Исходные данные ──
c(ws,1,1,'ИСХОДНЫЕ ДАННЫЕ — Orange Standard 5кг (Art. 119105)', fill=ORANGE_F, font=Font(bold=True, size=12))
ws.merge_cells('A1:M1')

for col, h in enumerate(['Параметр','Значение','Примечание'], 1):
    c(ws, 2, col, h, fill=BLUE, font=bold)

params = [
    ('Себестоимость, руб',          496,          'производство + упаковка'),
    ('Объём, кг',                     5,          'концентрат'),
    ('Разбавление',           '1:30-1:40',        'для пеногенератора'),
    ('Логистика FBO, руб',          250,          'хранение + доставка Ozon'),
    ('Комиссия Ozon (факт), %', '14-18%',         'зависит от цены и акций (по истории)'),
    ('Реклама (ДРР), %',              5,          'если используется'),
    ('Текущая цена (без Ozon Банка)',1340,         'актуальная цена на площадке'),
    ('Текущая цена (с Ozon Банком)', 1213,        'цена с учётом кешбека банка'),
]
for i, (name, val, note) in enumerate(params, 3):
    c(ws, i, 1, name, fill=GRAY, align=left_al)
    c(ws, i, 2, val,  fill=GRAY)
    c(ws, i, 3, note, fill=GRAY, align=left_al)

# ── Блок 2: Сценарии ──
c(ws,10,1,'СЦЕНАРИИ ЦЕНООБРАЗОВАНИЯ', fill=ORANGE_F, font=Font(bold=True, size=12))
ws.merge_cells('A10:M10')

hdrs = ['Сценарий','Цена, руб','Комиссия Ozon, руб','Логистика, руб','Реклама, руб',
        'Итого расходы, руб','Себест., руб','Прибыль с ед., руб','Маржа, %',
        'Цена/л конц., руб','Цена/л р-ра, руб','Позиция vs конкурент','Рекомендация']
for col, h in enumerate(hdrs, 1):
    c(ws, 11, col, h, fill=BLUE, font=bold)

# Комиссия Ozon по истории: ~18% при 1100-1350, ~15% при 1400-1600, ~14% при 1700+
# Текущая цена: 1340 руб (без Ozon Банка) / 1213 руб (с Ozon Банком)
scenarios = [
    ('❌ Ниже текущей — тест минимума',  1100, False,
     'Ниже текущей (1340). ИСТОРИЯ: май-июн 2025 @ 1174-1203 -> 3 шт/мес, +177..+246 руб/шт.',
     'ГИПОТЕЗА: объём вырастет до 3-4 шт/мес, но маржа минимальна. Только для теста эластичности.'),
    ('❌ Ниже текущей + реклама',         1100, True,
     'Тот же уровень с рекламой 5%.',
     'ГИПОТЕЗА: реклама съест прибыль. Маржа уйдёт в ноль или минус. Не рекомендуется.'),
    ('⚡ Текущая цена (без Ozon Банка)',   1340, False,
     'ТЕКУЩАЯ ЦЕНА. ИСТОРИЯ: июл 2025 @ 1260 -> 2 шт/мес, -4 руб/шт (почти ноль). Выше = лучше.',
     'ГИПОТЕЗА: ~2-3 шт/мес. Маржа слабая (~8%). Хранение съедает прибыль. Нужно менять.'),
    ('⚡ Текущая + реклама 5%',            1340, True,
     'Текущая цена с платным продвижением.',
     'ГИПОТЕЗА: реклама увеличит объём до 4-6 шт/мес, но маржа ~3%. Работает только при высокой конверсии.'),
    ('✅ Гипотеза 1 — шаг +12%',          1500, False,
     'На 160 руб выше текущей. ИСТОРИЯ: авг 2025 @ 1480 -> 2 шт/мес, +304 руб/шт (20.5% маржа).',
     'ГИПОТЕЗА: объём не упадёт (конкуренты выше), маржа значительно вырастет. Тестировать первым.'),
    ('✅ Гипотеза 1 + реклама',            1500, True,
     'Тот же уровень с рекламой. Исторически комиссия ~15% в этом диапазоне.',
     'ГИПОТЕЗА: 3-5 шт/мес при активной рекламе, маржа ~14%. Оптимальный баланс.'),
    ('✅ Гипотеза 2 — шаг +27%',          1700, False,
     'ИСТОРИЯ: сент-окт 2025 @ 1700 -> 3-4 шт/мес, +394..+491 руб/шт, комиссия 14%.',
     'ГИПОТЕЗА: лучшая цена по истории. Маржа 22%+. Нужна улучшенная карточка для конверсии.'),
    ('✅ Гипотеза 2 + реклама',            1700, True,
     'Исторически лучший результат + продвижение.',
     'ГИПОТЕЗА: 5-8 шт/мес. Целевой сценарий при обновлённой карточке и активной рекламе.'),
    ('⚠️ Гипотеза 3 — как DR.BERG',        1900, False,
     'DR.BERG ACE: 1903 руб, 562 заказа/28 дн. ИСТОРИЯ: 2100 руб -> 1 шт/мес (слишком высоко).',
     'ГИПОТЕЗА: объём 1-2 шт/мес. Оправдано только при профессиональном позиционировании и отзывах.'),
]

cost = 496
fbo  = 250
adv_pct = 0.05

# Комиссия по истории: <900->30%, 900-1300->18%, 1300-1600->15%, >1600->14%
def get_comm(price):
    if price < 900:   return 0.30
    if price < 1300:  return 0.18
    if price < 1600:  return 0.15
    return 0.14

for i, (name, price, with_adv, position, rec) in enumerate(scenarios, 12):
    comm_rub = round(price * get_comm(price))
    adv_rub  = round(price * adv_pct) if with_adv else 0
    expenses = comm_rub + fbo + adv_rub
    profit   = price - expenses - cost
    margin   = round(profit / price * 100, 1) if price > 0 else 0
    ppl_conc = round(price / 5, 0)
    ppl_ready = round(price / (5 * 35), 2)

    f = GREEN if profit > 150 else (YELLOW if profit > 0 else RED_FILL)

    c(ws, i, 1,  name,       fill=GRAY, align=left_al)
    c(ws, i, 2,  price,      fill=f,    nf='#,##0')
    c(ws, i, 3,  comm_rub,   fill=f,    nf='#,##0')
    c(ws, i, 4,  fbo,        fill=f,    nf='#,##0')
    c(ws, i, 5,  adv_rub,    fill=f,    nf='#,##0')
    c(ws, i, 6,  expenses,   fill=f,    nf='#,##0')
    c(ws, i, 7,  cost,       fill=f,    nf='#,##0')
    c(ws, i, 8,  profit,     fill=f,    nf='#,##0')
    c(ws, i, 9,  margin,     fill=f,    nf='0.0')
    c(ws, i, 10, ppl_conc,   fill=GRAY, nf='#,##0')
    c(ws, i, 11, ppl_ready,  fill=GRAY, nf='0.00')
    c(ws, i, 12, position,   fill=GRAY, align=left_al)
    c(ws, i, 13, rec,        fill=GRAY, align=left_al)

# ── Блок 3: Конкуренты ──
start = 23
c(ws, start, 1, 'КОНКУРЕНТЫ — бесконтактная мойка 5кг (топ-10 по выручке, 28 дней)', fill=ORANGE_F, font=Font(bold=True, size=12))
ws.merge_cells(f'A{start}:M{start}')

ch = ['#','Товар','Бренд','Объём, кг','Ср. цена, руб','Цена экв. 5кг, руб',
      'Выручка 28д','Заказов','Дней без ост.','Динамика %','Дней в акциях','Дней реклама','ДРР %']
for col, h in enumerate(ch, 1):
    c(ws, start+1, col, h, fill=BLUE, font=bold)

# num, name, brand, volume_kg, price, revenue, orders, oos, dyn, promo, adv, ddr
competitors = [
    (1,  'GLITTER G5, 5л',                  'GLITTER',   5.0,  418,   5034446, 12038, '15/28', 239, '28/28', '28/28', 11.5),
    (2,  'GRASS Active Foam Balance, 5кг',   'Grass',     5.0,  611,   4812674,  7883, '13/28', 390, '27/28', '0/28',   7.0),
    (3,  'GRASS Active Foam Pink, 6кг',      'Grass',     6.0, 1620,   3460179,  2136, '-',     465, '0/28',  '0/28',   0.8),
    (4,  'GRASS Active Foam Red, 5.8кг',     'Grass',     5.8, 1252,   3159342,  2523, '-',     221, '0/28',  '0/28',   1.1),
    (5,  'GRASS Active Foam Red, 5.8кг (SKU2)','Grass',   5.8, 1256,   2893082,  2303, '7/28',  397, '18/28', '0/28',   4.7),
    (6,  'GLITTER G10, 5л',                  'GLITTER',   5.0,  554,   2728526,  4923, '17/28', 276, '13/28', '28/28', 10.7),
    (7,  'GRASS Active Foam Pink, 6кг (SKU2)','Grass',    6.0, 1744,   2527411,  1449, '5/28',  158, '18/28', '0/28',   4.1),
    (8,  'DR.BERG ACE, 5кг',                 'DR. BERG',  5.0, 1903,   1069666,   562, '14/28', 281, '28/28', '0/28',  10.3),
    (9,  'BIGHIM Active Foam Expert, 5кг',   'BIGHIM',    5.0,  648,   1027756,  1585, '25/28', 770, '28/28', '28/28',  1.0),
    (10, 'GRASS Active Foam Effect, 6кг',    'Grass',     6.0, 1422,    935861,   658, '-',     386, '0/28',  '0/28',   0.2),
]

for i, row_data in enumerate(competitors, start+2):
    num, name, brand, vol, price, rev, orders, oos, dyn, promo, adv, ddr = row_data
    price_5kg = round(price * 5 / vol)
    f = GREEN if num <= 2 else GRAY
    c(ws, i, 1,  num,       fill=f)
    c(ws, i, 2,  name,      fill=f, align=left_al)
    c(ws, i, 3,  brand,     fill=f, align=left_al)
    c(ws, i, 4,  vol,       fill=f, nf='0.0')
    c(ws, i, 5,  price,     fill=f, nf='#,##0')
    c(ws, i, 6,  price_5kg, fill=YELLOW if price != price_5kg else f, nf='#,##0')
    c(ws, i, 7,  rev,       fill=f, nf='#,##0')
    c(ws, i, 8,  orders,    fill=f, nf='#,##0')
    c(ws, i, 9,  oos,       fill=f)
    c(ws, i, 10, dyn,       fill=f)
    c(ws, i, 11, promo,     fill=f)
    c(ws, i, 12, adv,       fill=f)
    c(ws, i, 13, ddr,       fill=f)

# Ширина колонок
widths = [30, 16, 16, 16, 14, 16, 14, 16, 10, 18, 16, 35, 35]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for r in [1, 10, 11, 23, 24]:
    ws.row_dimensions[r].height = 32
for r in range(12, 22):
    ws.row_dimensions[r].height = 22
for r in range(25, 36):
    ws.row_dimensions[r].height = 22

# Заморозить первую строку
ws.freeze_panes = 'A12'

# ── Лист 2: История продаж ──
ws2 = wb.create_sheet('История продаж')

c(ws2,1,1,'ИСТОРИЯ ПРОДАЖ: Orange Standard 5кг (Art. 119105) | 14 месяцев', fill=ORANGE_F, font=Font(bold=True, size=12))
ws2.merge_cells('A1:J1')

hist_headers = ['Период','Продано, шт','Ср. цена, руб','Комиссия Ozon %','Хранение, руб',
                'Прибыль/шт, руб','Маржа %','Выручка gross','Приход факт','Примечание']
for col, h in enumerate(hist_headers, 1):
    c(ws2, 2, col, h, fill=BLUE, font=bold)

history = [
    ('2025-01',  1,   837, 26,  -8,   -37, -4.5,   837,   459, 'Убыток. Низкая цена + высокая комиссия.'),
    ('2025-02',  3,   862, 23, -10,   +12,  1.4,  2586,  1525, 'Безубыточно. Минимум.'),
    ('2025-03',  2,  1030, 39, -14,   -63, -6.1,  2060,   866, 'УБЫТОК. Комиссия 39% — аномалия.'),
    ('2025-04',  5,   819, 37, -56,  -128,-15.6,  4094,  1840, 'УБЫТОК. Макс объём, мин маржа. Акции.'),
    ('2025-05',  3,  1203, 18, -46,  +246, 20.5,  3609,  2227, 'Прибыльно. Хорошая цена.'),
    ('2025-06',  3,  1174, 19, -49,  +177, 15.1,  3522,  2018, 'Прибыльно.'),
    ('2025-07',  2,  1260, 18, -11,    -4, -0.3,  2520,   984, 'Около нуля. Хранение съедает.'),
    ('2025-08',  2,  1480, 15, -34,  +304, 20.5,  2960,  1599, 'Прибыльно. Цена растёт.'),
    ('2025-09',  3,  1700, 14, -59,  +491, 28.9,  5100,  2962, 'ТОП МАРЖА. Лучший месяц.'),
    ('2025-10',  4,  1700, 14, -20,  +394, 23.2,  6800,  3562, 'ТОП ОБЪЁМ+МАРЖА. Целевая цена.'),
    ('2025-11',  0,     0,  0,   0,     0,  0.0,     0,  -364, 'Нет продаж. Хранение = убыток.'),
    ('2025-12',  0,     0,  0,   0,     0,  0.0,     0,  -251, 'Нет продаж. Хранение = убыток.'),
    ('2026-02',  0,     0,  0,   0,     0,  0.0,     0,  -319, 'Нет продаж. Хранение = убыток.'),
    ('2026-03',  1,  2100, 10,   0,  +440, 20.9,  2100,   936, 'Высокая цена -> 1 шт/мес. Нет объёма.'),
]

for i, row_data in enumerate(history, 3):
    period, sold, price, comm_pct, stor, profit_u, margin, gross, income, note = row_data
    if sold > 0:
        f = GREEN if profit_u > 200 else (YELLOW if profit_u > 0 else RED_FILL)
    else:
        f = RED_FILL
    c(ws2, i, 1,  period,    fill=f)
    c(ws2, i, 2,  sold,      fill=f, nf='#,##0')
    c(ws2, i, 3,  price,     fill=f, nf='#,##0')
    c(ws2, i, 4,  comm_pct,  fill=f, nf='0"%"')
    c(ws2, i, 5,  stor,      fill=f, nf='#,##0')
    c(ws2, i, 6,  profit_u,  fill=f, nf='+#,##0;-#,##0;0')
    c(ws2, i, 7,  margin,    fill=f, nf='0.0"%"')
    c(ws2, i, 8,  gross,     fill=f, nf='#,##0')
    c(ws2, i, 9,  income,    fill=f, nf='#,##0')
    c(ws2, i, 10, note,      fill=f, align=left_al)

# Итого
r_tot = len(history) + 3
c(ws2, r_tot, 1, 'ИТОГО / СРЕДНЕЕ', fill=ORANGE_F, font=bold)
c(ws2, r_tot, 2, 29,      fill=ORANGE_F, font=bold, nf='#,##0')
c(ws2, r_tot, 3, 1247,    fill=ORANGE_F, font=bold, nf='#,##0')
c(ws2, r_tot, 4, '',      fill=ORANGE_F)
c(ws2, r_tot, 5, '',      fill=ORANGE_F)
c(ws2, r_tot, 6, 126,     fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0;0')
c(ws2, r_tot, 7, 10.1,    fill=ORANGE_F, font=bold, nf='0.0"%"')
c(ws2, r_tot, 8, 36188,   fill=ORANGE_F, font=bold, nf='#,##0')
c(ws2, r_tot, 9, 18043,   fill=ORANGE_F, font=bold, nf='#,##0')
c(ws2, r_tot, 10,'2.1 шт/мес за 14 мес. Оптимальная цена: 1700 руб.', fill=ORANGE_F, font=bold, align=left_al)

# Ширина
for i, w in enumerate([12,12,14,14,12,14,10,14,12,45], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 30

# ── Лист 3: WB vs Ozon сравнение ──
ws3 = wb.create_sheet('WB vs Ozon')

c(ws3,1,1,'СРАВНЕНИЕ: Wildberries vs Ozon — Orange Standard 5кг (15 месяцев)', fill=ORANGE_F, font=Font(bold=True, size=12))
ws3.merge_cells('A1:N1')

# Заголовки
hdrs3 = ['Период',
         'WB — шт','WB — цена, руб','WB — маржа %','WB — прибыль/шт','WB — прибыль всего',
         '',
         'Ozon — шт','Ozon — цена, руб','Ozon — маржа %','Ozon — прибыль/шт','Ozon — прибыль всего',
         '',
         'Комментарий']
for col, h in enumerate(hdrs3, 1):
    fill = BLUE if h and h != '' else PatternFill()
    c(ws3, 2, col, h, fill=fill, font=bold if h else None)

# Данные: период | wb_sold wb_price wb_marg wb_prof_u wb_prof | | oz_sold oz_price oz_marg oz_prof_u oz_prof | | note
rows3 = [
    ('2025-01', 4,  1000, 18.7, +187,  +746,  1,   837,  -4.5,   -37,   -37, 'WB прибыльно, Ozon убыток'),
    ('2025-02', 8,   977, 20.2, +197, +1576,  3,   862,   1.4,   +12,   +37, 'WB объём растёт'),
    ('2025-03',47,   763, -2.4,  -18,  -864,  2,  1030,  -6.1,   -63,  -126, '❌ Оба убыточны. WB акция?'),
    ('2025-04',61,   855,  8.2,  +70, +4281,  5,   819, -15.6,  -128,  -640, 'WB в плюсе, Ozon убыток при 819'),
    ('2025-05',83,   928, 10.8, +100, +8295,  3,  1203,  20.5,  +246,  +739, '✅ Оба прибыльны. Лучший мес WB по объёму'),
    ('2025-06',99,   867,  4.6,  +40, +3938,  3,  1174,  15.1,  +177,  +530, 'WB макс объём, низкая маржа'),
    ('2025-07',31,  1420, 29.7, +421,+13062,  2,  1260,  -0.3,    -4,    -8, '✅ WB лучший по марже. Ozon почти 0'),
    ('2025-08',19,  1313, 23.5, +309, +5871,  2,  1480,  20.5,  +304,  +607, '✅ Оба прибыльны при цене 1300-1480'),
    ('2025-09',18,  1270, 17.8, +226, +4072,  3,  1700,  28.9,  +491, +1474, '✅ Ozon лучший месяц по марже/шт'),
    ('2025-10',13,  1202, -2.3,  -27,  -355,  4,  1700,  23.2,  +394, +1578, 'Ozon отлично. WB убыток — акция?'),
    ('2025-11',10,  1297,  3.5,  +45,  +449,  0,     0,   0.0,     0,    0,  'Ozon — нет продаж'),
    ('2025-12',17,  1267,  8.0, +102, +1728,  0,     0,   0.0,     0,    0,  'Ozon — нет продаж'),
    ('2026-01',17,  1176,  0.8,  +10,  +164,  0,     0,   0.0,     0,    0,  'WB почти 0 маржи'),
    ('2026-02',11,  1169, -5.5,  -64,  -705,  0,     0,   0.0,     0,    0,  '❌ WB убыток. Ozon нет продаж'),
    ('2026-03',147, 1046,-15.6, -163,-23908,  1,  2100,  20.9,  +440,  +440, '❌ WB катастрофа (акция?). Ozon 1 шт высокая маржа'),
]

for i, row_data in enumerate(rows3, 3):
    period, wb_s, wb_p, wb_m, wb_pu, wb_pt, oz_s, oz_p, oz_m, oz_pu, oz_pt, note = row_data

    def fwb(val_u):
        if wb_s == 0: return GRAY
        return GREEN if val_u > 100 else (YELLOW if val_u > 0 else RED_FILL)
    def foz(val_u):
        if oz_s == 0: return GRAY
        return GREEN if val_u > 100 else (YELLOW if val_u > 0 else RED_FILL)

    fw = fwb(wb_pu)
    fo = foz(oz_pu)

    c(ws3, i, 1,  period,                         fill=GRAY)
    c(ws3, i, 2,  wb_s if wb_s > 0 else '-',      fill=fw, nf='#,##0' if wb_s > 0 else None)
    c(ws3, i, 3,  wb_p if wb_s > 0 else '-',      fill=fw, nf='#,##0' if wb_s > 0 else None)
    c(ws3, i, 4,  wb_m if wb_s > 0 else '-',      fill=fw, nf='0.0' if wb_s > 0 else None)
    c(ws3, i, 5,  wb_pu if wb_s > 0 else '-',     fill=fw, nf='+#,##0;-#,##0' if wb_s > 0 else None)
    c(ws3, i, 6,  wb_pt if wb_s > 0 else '-',     fill=fw, nf='+#,##0;-#,##0' if wb_s > 0 else None)
    c(ws3, i, 7,  '',                              fill=PatternFill())
    c(ws3, i, 8,  oz_s if oz_s > 0 else '-',      fill=fo, nf='#,##0' if oz_s > 0 else None)
    c(ws3, i, 9,  oz_p if oz_s > 0 else '-',      fill=fo, nf='#,##0' if oz_s > 0 else None)
    c(ws3, i, 10, oz_m if oz_s > 0 else '-',      fill=fo, nf='0.0' if oz_s > 0 else None)
    c(ws3, i, 11, oz_pu if oz_s > 0 else '-',     fill=fo, nf='+#,##0;-#,##0' if oz_s > 0 else None)
    c(ws3, i, 12, oz_pt if oz_s > 0 else '-',     fill=fo, nf='+#,##0;-#,##0' if oz_s > 0 else None)
    c(ws3, i, 13, '',                              fill=PatternFill())
    c(ws3, i, 14, note,                            fill=GRAY, align=left_al)

# Итого
r_tot3 = len(rows3) + 3
c(ws3, r_tot3, 1, 'ИТОГО', fill=ORANGE_F, font=bold)
c(ws3, r_tot3, 2, 585,     fill=ORANGE_F, font=bold, nf='#,##0')
c(ws3, r_tot3, 3, 1010,    fill=ORANGE_F, font=bold, nf='#,##0')
c(ws3, r_tot3, 4, '',      fill=ORANGE_F)
c(ws3, r_tot3, 5, 31,      fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0')
c(ws3, r_tot3, 6, 18130,   fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0')
c(ws3, r_tot3, 7, '',      fill=PatternFill())
c(ws3, r_tot3, 8, 29,      fill=ORANGE_F, font=bold, nf='#,##0')
c(ws3, r_tot3, 9, 1247,    fill=ORANGE_F, font=bold, nf='#,##0')
c(ws3, r_tot3, 10,'',      fill=ORANGE_F)
c(ws3, r_tot3, 11, 126,    fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0')
c(ws3, r_tot3, 12, 3659,   fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0')
c(ws3, r_tot3, 13,'',      fill=PatternFill())
c(ws3, r_tot3, 14,'WB: 39 шт/мес avg | Ozon: 2.1 шт/мес avg | WB объём в 19x больше', fill=ORANGE_F, font=bold, align=left_al)

# ── Блок выводов и гипотез ──
r_hyp = r_tot3 + 2
c(ws3, r_hyp, 1, 'ГИПОТЕЗЫ ДЛЯ OZON НА ОСНОВЕ ОПЫТА WB', fill=ORANGE_F, font=Font(bold=True, size=12))
ws3.merge_cells(f'A{r_hyp}:N{r_hyp}')

hyp_hdrs = ['Гипотеза','Основание (WB данные)','Прогноз объёма Ozon, шт/мес','Прогноз маржи Ozon','Риск','Действие']
for col, h in enumerate(hyp_hdrs, 1):
    c(ws3, r_hyp+1, col, h, fill=BLUE, font=bold)
ws3.merge_cells(f'B{r_hyp+1}:C{r_hyp+1}')
ws3.merge_cells(f'D{r_hyp+1}:E{r_hyp+1}')

hypotheses = [
    ('Цена ~1 100 ₽ даёт объём, но не прибыль',
     'WB при 855-928 руб: 60-83 шт/мес, маржа 8-10%. Ozon комиссия выше — маржа исчезнет.',
     '5-10 шт/мес (Ozon трафика меньше в 19x)',
     '0-5% — на грани',
     'Высокий: Ozon комиссия 18% съедает разницу',
     'Не рекомендуется без рекламного бюджета'),
    ('Цена ~1 340 ₽ (текущая) = зона неопределённости',
     'WB при 1 200-1 300 руб: 13-17 шт/мес, маржа 0-8%. Ozon должен быть лучше по марже.',
     '2-4 шт/мес (текущий факт)',
     '8-12%',
     'Средний: нет рекламы = нет трафика',
     'Улучшить карточку + тест рекламы'),
    ('Цена ~1 420 ₽ = "золотая точка" WB',
     'WB июл 2025 @ 1 420 руб: 31 шт/мес, маржа 29.7%, прибыль +421/шт. Лучший месяц.',
     '3-6 шт/мес',
     '18-22%',
     'Низкий по марже, средний по объёму',
     '✅ Тестировать первым. Поднять с 1340 до 1420.'),
    ('Цена ~1 700 ₽ = исторический максимум Ozon',
     'Ozon сен-окт 2025 @ 1 700: 3-4 шт/мес, маржа 23-29%. WB при такой цене почти нет продаж.',
     '3-5 шт/мес при хорошей карточке',
     '22-28%',
     'Средний: нужна конвертирующая карточка',
     '✅ Целевая цена после теста 1420.'),
    ('Март 2026 WB — катастрофа при 1 046 ₽',
     '147 шт проданы с убытком -163/шт = -23 908 руб. Вероятно: принудительная акция WB или ошибка.',
     'N/A',
     '-15.6%',
     'Критический: акции WB/Ozon могут уронить цену ниже себестоимости',
     '⚠️ Проверить правила участия в акциях. Установить минимальную цену.'),
]

for i, (hyp, basis, vol, margin, risk, action) in enumerate(hypotheses, r_hyp+2):
    f = GREEN if '✅' in action else (RED_FILL if '⚠️' in action else GRAY)
    c(ws3, i, 1, hyp,    fill=f, align=left_al)
    c(ws3, i, 2, basis,  fill=f, align=left_al)
    ws3.merge_cells(f'B{i}:C{i}')
    c(ws3, i, 4, vol,    fill=f, align=left_al)
    ws3.merge_cells(f'D{i}:E{i}')
    c(ws3, i, 6, margin, fill=f)
    c(ws3, i, 7, risk,   fill=f, align=left_al)
    ws3.merge_cells(f'G{i}:J{i}')
    c(ws3, i, 11, action, fill=f, align=left_al)
    ws3.merge_cells(f'K{i}:N{i}')

# Ширина колонок
ws3_widths = [10, 8, 10, 9, 12, 14, 2, 8, 10, 9, 12, 14, 2, 40]
for i, w in enumerate(ws3_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

for r in range(1, r_hyp+8):
    ws3.row_dimensions[r].height = 22
ws3.row_dimensions[1].height = 30
ws3.row_dimensions[r_hyp].height = 30

ws3.freeze_panes = 'A3'

wb.save('C:/Users/1/Downloads/orange_standard_pricing_v4.xlsx')
print('Готово: C:/Users/1/Downloads/orange_standard_pricing_v4.xlsx')
