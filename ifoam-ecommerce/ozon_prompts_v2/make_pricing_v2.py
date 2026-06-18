"""
Промпт 2 — Профиль продаж + таблица ценообразования
Источник: DataExport_2026-03-24_09-37-33.xlsx (сырые данные)
Выход:    orange_pricing_v5.xlsx
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Цвета ──
ORANGE_F = PatternFill('solid', fgColor='F4B942')
BLUE     = PatternFill('solid', fgColor='BDD7EE')
GREEN    = PatternFill('solid', fgColor='C6EFCE')
YELLOW   = PatternFill('solid', fgColor='FFD966')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
GRAY     = PatternFill('solid', fgColor='F2F2F2')
DGRAY    = PatternFill('solid', fgColor='D9D9D9')

def brd():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value, fill=None, font=None, align=None, nf=None, merge_to=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:  c.fill  = fill
    if font:  c.font  = font
    c.alignment = align or Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = brd()
    if nf:    c.number_format = nf
    return c

left  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
right = Alignment(horizontal='right',  vertical='center')
bold  = Font(bold=True)
bold12 = Font(bold=True, size=12)

# ── Загрузка сырых данных ──
RAW = 'C:/Users/1/Downloads/DataExport_2026-03-24_09-37-33.xlsx'
df  = pd.read_excel(RAW, sheet_name='Chart data', header=0)
df.columns = [
    'Период','Маркетплейс','Продано','Маржа_pct','Прибыль','Приход',
    'Выручка','Скидки','Себест_total','Комиссия_mp','C10','C11','C12',
    'Комиссия2','Хранение','Доставка','C16','C17','C18','Себест_шт'
]

def period_start(p):
    return str(p).strip()[:7]   # "2025-01"

df['Период_str'] = df['Период'].apply(period_start)
df['Ср_цена'] = df.apply(
    lambda r: round(r['Выручка'] / r['Продано']) if r['Продано'] > 0 else 0, axis=1)
df['Прибыль_шт'] = df.apply(
    lambda r: round(r['Прибыль'] / r['Продано']) if r['Продано'] > 0 else 0, axis=1)
df['Комис_pct'] = df.apply(
    lambda r: round(abs(r['Комиссия_mp']) / r['Выручка'] * 100, 1) if r['Продано'] > 0 else 0, axis=1)

ozon = df[df['Маркетплейс'] == 'Ozon'].copy().sort_values('Период_str').reset_index(drop=True)
wb_d = df[df['Маркетплейс'] == 'Wildberries'].copy().sort_values('Период_str').reset_index(drop=True)

COST = int(df['Себест_шт'].iloc[0])   # 496

# ══════════════════════════════════════════════════════
# ЛИСТ 1 — ЦЕНООБРАЗОВАНИЕ
# ══════════════════════════════════════════════════════
wb_out = Workbook()
ws1 = wb_out.active
ws1.title = 'Ценообразование'

# ── Блок: Исходные данные ──
cell(ws1, 1, 1, 'ИСХОДНЫЕ ДАННЫЕ — Orange Standard 5кг (Art. 119105)',
     fill=ORANGE_F, font=bold12)
ws1.merge_cells('A1:M1')

for col, h in enumerate(['Параметр', 'Значение', 'Примечание'], 1):
    cell(ws1, 2, col, h, fill=BLUE, font=bold)

params = [
    ('Себестоимость, руб',           COST,           'из данных выгрузки'),
    ('Объём, кг',                    5,              'концентрат'),
    ('Разбавление',                  '1:30–1:40',    'для пеногенератора'),
    ('Логистика FBO, руб',           250,            'хранение + доставка Ozon (оценка)'),
    ('Комиссия Ozon — факт %',       '10–39%',       'диапазон по истории; зависит от цены и акций'),
    ('Реклама (ДРР), %',             5,              'если используется'),
    ('Текущая цена (без Ozon Банка)', 1340,          'актуальная цена на площадке'),
    ('Текущая цена (с Ozon Банком)',  1213,          'с учётом кешбэка банка'),
]
for i, (name, val, note) in enumerate(params, 3):
    cell(ws1, i, 1, name, fill=GRAY, align=left)
    cell(ws1, i, 2, val,  fill=GRAY)
    cell(ws1, i, 3, note, fill=GRAY, align=left)

# ── Блок: Сценарии ──
cell(ws1, 12, 1, 'СЦЕНАРИИ ЦЕНООБРАЗОВАНИЯ (на основе реальной истории)',
     fill=ORANGE_F, font=bold12)
ws1.merge_cells('A12:M12')

hdrs = ['Сценарий', 'Цена, руб', 'Комиссия, руб', 'Логистика, руб', 'Реклама, руб',
        'Итого расходы', 'Себест., руб', 'Прибыль/шт', 'Маржа %',
        'Руб/кг конц.', 'Руб/л р-ра', 'История / позиция', 'Рекомендация']
for col, h in enumerate(hdrs, 1):
    cell(ws1, 13, col, h, fill=BLUE, font=bold)

def get_comm(price):
    if price < 900:   return 0.30
    if price < 1300:  return 0.18
    if price < 1600:  return 0.15
    return 0.14

scenarios = [
    ('❌ Ниже текущей — тест минимума', 1100, False,
     'Ниже текущей (1340). История: май–июн 2025 @ 1174–1203 → 3 шт/мес, +177…+246 руб/шт.',
     'ГИПОТЕЗА: объём 3–4 шт/мес, маржа минимальна. Только для теста эластичности.'),
    ('❌ Ниже текущей + реклама 5%', 1100, True,
     'Тот же уровень + реклама 5%.',
     'ГИПОТЕЗА: реклама съест прибыль. Маржа уйдёт в ноль/минус. Не рекомендуется.'),
    ('⚡ Текущая цена (без Ozon Банка)', 1340, False,
     'Текущая цена. История: июл 2025 @ 1260 → 2 шт, –4 руб/шт. Чуть выше — уже лучше.',
     'ГИПОТЕЗА: 2–3 шт/мес, маржа ~8%. Хранение съедает прибыль. Нужно менять.'),
    ('⚡ Текущая + реклама 5%', 1340, True,
     'Текущая цена с платным продвижением.',
     'ГИПОТЕЗА: 4–6 шт/мес, маржа ~3%. Работает только при высокой конверсии карточки.'),
    ('✅ Гипотеза 1 — шаг +12%', 1500, False,
     'На 160 руб выше текущей. История: авг 2025 @ 1480 → 2 шт, +304 руб/шт (20.5%).',
     'ГИПОТЕЗА: объём не упадёт, маржа вырастет значительно. Тестировать первым.'),
    ('✅ Гипотеза 1 + реклама 5%', 1500, True,
     'Тот же уровень + реклама. Комиссия ~15% в этом диапазоне.',
     'ГИПОТЕЗА: 3–5 шт/мес при рекламе, маржа ~14%. Оптимальный баланс цена/объём.'),
    ('✅ Гипотеза 2 — исторический оптимум', 1700, False,
     'История: сент–окт 2025 @ 1700 → 3–4 шт, +394…+491 руб/шт, комиссия 14%.',
     'ГИПОТЕЗА: лучшая цена по истории. Маржа 22%+. Нужна улучшенная карточка.'),
    ('✅ Гипотеза 2 + реклама 5%', 1700, True,
     'Исторически лучший результат + продвижение.',
     'ГИПОТЕЗА: 5–8 шт/мес. Целевой сценарий при обновлённой карточке + реклама.'),
    ('⚠️ Гипотеза 3 — премиум', 1900, False,
     'DR.BERG ACE: 1903 руб, 562 заказа/28 дн. История: 2100 руб → 1 шт/мес.',
     'ГИПОТЕЗА: 1–2 шт/мес. Оправдано только при сильном бренде, отзывах и карточке.'),
]

FBO = 250
for i, (name, price, with_adv, position, rec) in enumerate(scenarios, 14):
    comm_rub = round(price * get_comm(price))
    adv_rub  = round(price * 0.05) if with_adv else 0
    expenses = comm_rub + FBO + adv_rub
    profit   = price - expenses - COST
    margin   = round(profit / price * 100, 1)
    ppl_conc  = round(price / 5)
    ppl_ready = round(price / (5 * 35), 2)

    f = GREEN if profit > 150 else (YELLOW if profit > 0 else RED_FILL)
    cell(ws1, i,  1, name,      fill=GRAY, align=left)
    cell(ws1, i,  2, price,     fill=f,    nf='#,##0')
    cell(ws1, i,  3, comm_rub,  fill=f,    nf='#,##0')
    cell(ws1, i,  4, FBO,       fill=f,    nf='#,##0')
    cell(ws1, i,  5, adv_rub,   fill=f,    nf='#,##0')
    cell(ws1, i,  6, expenses,  fill=f,    nf='#,##0')
    cell(ws1, i,  7, COST,      fill=f,    nf='#,##0')
    cell(ws1, i,  8, profit,    fill=f,    nf='+#,##0;-#,##0;0')
    cell(ws1, i,  9, margin,    fill=f,    nf='0.0"%"')
    cell(ws1, i, 10, ppl_conc,  fill=GRAY, nf='#,##0')
    cell(ws1, i, 11, ppl_ready, fill=GRAY, nf='0.00')
    cell(ws1, i, 12, position,  fill=GRAY, align=left)
    cell(ws1, i, 13, rec,       fill=GRAY, align=left)

# ширина колонок
widths1 = [30, 14, 15, 14, 13, 14, 12, 14, 10, 12, 12, 38, 38]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[1].height  = 28
ws1.row_dimensions[12].height = 28
ws1.row_dimensions[13].height = 32
for r in range(14, 24):
    ws1.row_dimensions[r].height = 40
ws1.freeze_panes = 'A14'

# ══════════════════════════════════════════════════════
# ЛИСТ 2 — ИСТОРИЯ OZON (из сырых данных)
# ══════════════════════════════════════════════════════
ws2 = wb_out.create_sheet('История Ozon')

cell(ws2, 1, 1, 'ИСТОРИЯ ПРОДАЖ OZON: Orange Standard 5кг — из сырых данных',
     fill=ORANGE_F, font=bold12)
ws2.merge_cells('A1:K1')

h2 = ['Период', 'Продано, шт', 'Ср. цена, руб', 'Комиссия %', 'Скидки, руб',
      'Хранение, руб', 'Прибыль/шт', 'Маржа %', 'Выручка gross', 'Приход факт', 'Примечание']
for col, h in enumerate(h2, 1):
    cell(ws2, 2, col, h, fill=BLUE, font=bold)

notes_oz = {
    '2025-01': 'Убыток. Низкая цена + высокая комиссия.',
    '2025-02': 'Безубыточно. Минимальная прибыль.',
    '2025-03': 'УБЫТОК. Комиссия 39% — аномалия.',
    '2025-04': 'УБЫТОК. Макс объём, мин маржа. Акция.',
    '2025-05': 'Прибыльно. Хорошая цена.',
    '2025-06': 'Прибыльно.',
    '2025-07': 'Около нуля. Хранение съедает.',
    '2025-08': 'Прибыльно. Цена растёт.',
    '2025-09': 'ТОП МАРЖА. Лучший месяц по %.',
    '2025-10': 'ТОП ОБЪЁМ+МАРЖА. Целевая цена.',
    '2025-11': 'Нет продаж. Только хранение.',
    '2025-12': 'Нет продаж. Только хранение.',
    '2026-02': 'Нет продаж. Только хранение.',
    '2026-03': 'Высокая цена → 1 шт/мес. Нет объёма.',
}

for i, row in ozon.iterrows():
    r = i + 3
    sold  = int(row['Продано'])
    price = int(row['Ср_цена']) if sold > 0 else 0
    comm  = row['Комис_pct']
    skid  = round(row['Скидки'])  if sold > 0 else 0
    stor  = round(row['Хранение'])
    prof  = int(row['Прибыль_шт']) if sold > 0 else 0
    marg  = round(row['Маржа_pct'] * 100, 1)
    gross = round(row['Выручка'])
    inc   = round(row['Приход'])
    note  = notes_oz.get(row['Период_str'], '')

    if sold == 0:
        f = DGRAY
    elif prof > 200:
        f = GREEN
    elif prof > 0:
        f = YELLOW
    else:
        f = RED_FILL

    cell(ws2, r,  1, row['Период_str'], fill=f)
    cell(ws2, r,  2, sold if sold > 0 else '-', fill=f, nf='#,##0' if sold > 0 else '@')
    cell(ws2, r,  3, price if price > 0 else '-', fill=f, nf='#,##0' if price > 0 else '@')
    cell(ws2, r,  4, f'{comm:.0f}%' if sold > 0 else '-', fill=f)
    cell(ws2, r,  5, skid  if sold > 0 else '-', fill=f, nf='#,##0' if sold > 0 else '@')
    cell(ws2, r,  6, stor  if stor != 0 else '-', fill=f, nf='#,##0' if stor != 0 else '@')
    cell(ws2, r,  7, prof  if sold > 0 else '-', fill=f, nf='+#,##0;-#,##0;0' if sold > 0 else '@')
    cell(ws2, r,  8, f'{marg:.1f}%' if sold > 0 else '-', fill=f)
    cell(ws2, r,  9, gross if gross > 0 else '-', fill=f, nf='#,##0' if gross > 0 else '@')
    cell(ws2, r, 10, inc,   fill=f, nf='#,##0')
    cell(ws2, r, 11, note,  fill=f, align=left)

# Итоги
total_r = len(ozon) + 3
oz_sell = ozon[ozon['Продано'] > 0]
cell(ws2, total_r, 1, 'ИТОГО', fill=ORANGE_F, font=bold)
cell(ws2, total_r, 2, int(ozon['Продано'].sum()), fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws2, total_r, 3, int(oz_sell['Ср_цена'].mean()) if len(oz_sell) > 0 else 0,
     fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws2, total_r, 7,
     int(ozon['Прибыль'].sum() / ozon[ozon['Продано']>0]['Продано'].sum()),
     fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0;0')
cell(ws2, total_r, 9, int(ozon['Выручка'].sum()), fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws2, total_r, 10, int(ozon['Приход'].sum()), fill=ORANGE_F, font=bold, nf='#,##0')
for col in [4, 5, 6, 8, 11]:
    cell(ws2, total_r, col, '', fill=ORANGE_F)

widths2 = [12, 13, 14, 12, 13, 14, 14, 10, 14, 13, 42]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 28
ws2.row_dimensions[2].height = 32
ws2.freeze_panes = 'A3'

# ══════════════════════════════════════════════════════
# ЛИСТ 3 — ИСТОРИЯ WB + СРАВНЕНИЕ
# ══════════════════════════════════════════════════════
ws3 = wb_out.create_sheet('WB vs Ozon')

cell(ws3, 1, 1, 'СРАВНЕНИЕ WB vs OZON: Orange Standard 5кг — по месяцам',
     fill=ORANGE_F, font=bold12)
ws3.merge_cells('A1:N1')

h3 = ['Период',
      'WB шт', 'WB ср.цена', 'WB маржа%', 'WB прибыль/шт', 'WB выручка',
      '│',
      'OZ шт', 'OZ ср.цена', 'OZ маржа%', 'OZ прибыль/шт', 'OZ выручка',
      '│',
      'Разница объём']
for col, h in enumerate(h3, 1):
    cell(ws3, 2, col, h, fill=BLUE, font=bold)

periods = sorted(set(list(ozon['Период_str']) + list(wb_d['Период_str'])))
for i, p in enumerate(periods, 3):
    wb_r = wb_d[wb_d['Период_str'] == p]
    oz_r = ozon[ozon['Период_str'] == p]

    def get_vals(sub):
        if sub.empty: return 0, 0, 0.0, 0, 0
        r = sub.iloc[0]
        sold = int(r['Продано'])
        pr   = int(r['Ср_цена'])    if sold > 0 else 0
        mg   = round(r['Маржа_pct']*100, 1) if sold > 0 else 0.0
        pf   = int(r['Прибыль_шт']) if sold > 0 else 0
        gr   = int(r['Выручка'])
        return sold, pr, mg, pf, gr

    ws, wp, wm, wpr, wg = get_vals(wb_r)
    os, op, om, opr, og = get_vals(oz_r)
    diff = ws - os

    fw = GREEN if wm > 15 else (YELLOW if wm > 0 else (RED_FILL if ws > 0 else DGRAY))
    fo = GREEN if om > 15 else (YELLOW if om > 0 else (RED_FILL if os > 0 else DGRAY))

    cell(ws3, i,  1, p,   fill=GRAY)
    cell(ws3, i,  2, ws if ws > 0 else '-',  fill=fw, nf='#,##0' if ws > 0 else '@')
    cell(ws3, i,  3, wp if wp > 0 else '-',  fill=fw, nf='#,##0' if wp > 0 else '@')
    cell(ws3, i,  4, f'{wm:.1f}%' if ws > 0 else '-', fill=fw)
    cell(ws3, i,  5, wpr if ws > 0 else '-', fill=fw, nf='+#,##0;-#,##0;0' if ws > 0 else '@')
    cell(ws3, i,  6, wg  if wg > 0 else '-', fill=fw, nf='#,##0' if wg > 0 else '@')
    cell(ws3, i,  7, '│', fill=GRAY)
    cell(ws3, i,  8, os if os > 0 else '-',  fill=fo, nf='#,##0' if os > 0 else '@')
    cell(ws3, i,  9, op if op > 0 else '-',  fill=fo, nf='#,##0' if op > 0 else '@')
    cell(ws3, i, 10, f'{om:.1f}%' if os > 0 else '-', fill=fo)
    cell(ws3, i, 11, opr if os > 0 else '-', fill=fo, nf='+#,##0;-#,##0;0' if os > 0 else '@')
    cell(ws3, i, 12, og  if og > 0 else '-', fill=fo, nf='#,##0' if og > 0 else '@')
    cell(ws3, i, 13, '│', fill=GRAY)
    cell(ws3, i, 14, f'WB ×{diff//max(os,1)}' if os > 0 and ws > 0 else '-', fill=GRAY)

# Итоги WB vs OZ
total_r3 = len(periods) + 3
wb_sell = wb_d[wb_d['Продано'] > 0]
oz_sell2 = ozon[ozon['Продано'] > 0]

for col in range(1, 15):
    cell(ws3, total_r3, col, '', fill=ORANGE_F)

cell(ws3, total_r3, 1,  'ИТОГО',                                           fill=ORANGE_F, font=bold)
cell(ws3, total_r3, 2,  int(wb_d['Продано'].sum()),                        fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws3, total_r3, 3,  int(wb_d['Выручка'].sum()/wb_d['Продано'].sum()),  fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws3, total_r3, 5,  int(wb_d['Прибыль'].sum()/wb_d[wb_d['Продано']>0]['Продано'].sum()),
     fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0;0')
cell(ws3, total_r3, 6,  int(wb_d['Выручка'].sum()),                        fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws3, total_r3, 8,  int(ozon['Продано'].sum()),                        fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws3, total_r3, 9,  int(ozon['Выручка'].sum()/ozon[ozon['Продано']>0]['Продано'].sum()),
     fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws3, total_r3, 11, int(ozon['Прибыль'].sum()/ozon[ozon['Продано']>0]['Продано'].sum()),
     fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0;0')
cell(ws3, total_r3, 12, int(ozon['Выручка'].sum()),                        fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws3, total_r3, 14,
     f'WB ×{round(wb_d["Продано"].sum()/max(ozon["Продано"].sum(),1))}',
     fill=ORANGE_F, font=bold)

widths3 = [10, 8, 10, 10, 14, 12, 3, 8, 10, 10, 14, 12, 3, 13]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[1].height = 28
ws3.row_dimensions[2].height = 32
ws3.freeze_panes = 'A3'

# ── Сохранение ──
OUT = 'C:/Users/1/Downloads/orange_pricing_v5.xlsx'
wb_out.save(OUT)
print('Saved:', OUT)
