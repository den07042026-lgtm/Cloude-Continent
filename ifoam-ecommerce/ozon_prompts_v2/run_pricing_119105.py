# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, os
sys.path.insert(0, "C:/Users/1/Downloads/ozon_prompts_v2")
from save_helpers import get_product_folder, make_docx

# ── Цвета ──
ORANGE_F = PatternFill('solid', fgColor='F4B942')
BLUE     = PatternFill('solid', fgColor='BDD7EE')
GREEN    = PatternFill('solid', fgColor='C6EFCE')
YELLOW   = PatternFill('solid', fgColor='FFD966')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
GRAY     = PatternFill('solid', fgColor='F2F2F2')
DGRAY    = PatternFill('solid', fgColor='D9D9D9')

def brd():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value, fill=None, font=None, align=None, nf=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:  c.fill  = fill
    if font:  c.font  = font
    c.alignment = align or Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = brd()
    if nf:    c.number_format = nf
    return c

left   = Alignment(horizontal='left',  vertical='center', wrap_text=True)
bold   = Font(bold=True)
bold12 = Font(bold=True, size=12)

# ── Загрузка данных ──
RAW = 'C:/Users/1/Downloads/DataExport_2026-03-24_13-35-51.xlsx'
df  = pd.read_excel(RAW, sheet_name='Chart data', header=0)
df.columns = [
    'Период','Маркетплейс','Продано','Маржа_pct','Прибыль','Приход',
    'Выручка','Скидки','Себест_total','Комиссия_mp','C10','C11','C12',
    'Комиссия2','Хранение','Доставка','C16','C17','C18','Себест_шт'
]
df['Период_str'] = df['Период'].apply(lambda p: str(p).strip()[:7])
df['Ср_цена']    = df.apply(lambda r: round(r['Выручка'] / r['Продано']) if r['Продано'] > 0 else 0, axis=1)
df['Прибыль_шт'] = df.apply(lambda r: round(r['Прибыль'] / r['Продано']) if r['Продано'] > 0 else 0, axis=1)
df['Комис_pct']  = df.apply(lambda r: round(abs(r['Комиссия_mp']) / r['Выручка'] * 100, 1) if r['Продано'] > 0 else 0, axis=1)

ozon = df[df['Маркетплейс'] == 'Ozon'].copy().sort_values('Период_str').reset_index(drop=True)
wb_d = df[df['Маркетплейс'] == 'Wildberries'].copy().sort_values('Период_str').reset_index(drop=True)

COST = 496
FBO  = 250

def get_comm(price):
    if price < 900:   return 0.30
    if price < 1300:  return 0.18
    if price < 1600:  return 0.15
    return 0.14

# ══════════════════════════════════════════════════════
# ЛИСТ 1 — ЦЕНООБРАЗОВАНИЕ
# ══════════════════════════════════════════════════════
wb_out = Workbook()
ws1 = wb_out.active
ws1.title = 'Ценообразование'

cell(ws1, 1, 1, 'ИСХОДНЫЕ ДАННЫЕ — Orange Standard 5кг (Art. 119105)', fill=ORANGE_F, font=bold12)
ws1.merge_cells('A1:M1')

for col, h in enumerate(['Параметр', 'Значение', 'Примечание'], 1):
    cell(ws1, 2, col, h, fill=BLUE, font=bold)

params = [
    ('Себестоимость, руб',               COST,        'из справочника combined_reference_normalized'),
    ('Объём, кг',                        5,           'концентрат'),
    ('Разбавление',                      '1:30–1:40', 'для пеногенератора'),
    ('Логистика FBO, руб',               FBO,         'хранение + доставка Ozon (оценка)'),
    ('Комиссия Ozon — факт %',           '14–39%',    'в норме 14% при цене 1600+; аномалии в акциях'),
    ('Реклама (ДРР), %',                 5,           'если используется'),
    ('Текущая цена (зачёркнутая)',       2900,        'old_price в Ozon Seller на 24.03.2026'),
    ('Текущая цена (продажная)',          2100,        'current_price в Ozon Seller на 24.03.2026'),
    ('Минимальная цена / Ozon Банк',     1900,        'min_price — нижний предел'),
]
for i, (name, val, note) in enumerate(params, 3):
    cell(ws1, i, 1, name, fill=GRAY, align=left)
    cell(ws1, i, 2, val,  fill=GRAY)
    cell(ws1, i, 3, note, fill=GRAY, align=left)

cell(ws1, 13, 1, 'СЦЕНАРИИ ЦЕНООБРАЗОВАНИЯ (на основе реальной истории Ozon)', fill=ORANGE_F, font=bold12)
ws1.merge_cells('A13:M13')

hdrs = ['Сценарий', 'Цена, руб', 'Комиссия, руб', 'Логистика, руб', 'Реклама, руб',
        'Итого расходы', 'Себест., руб', 'Прибыль/шт', 'Маржа %',
        'Руб/кг конц.', 'Руб/л р-ра', 'История / позиция', 'Рекомендация']
for col, h in enumerate(hdrs, 1):
    cell(ws1, 14, col, h, fill=BLUE, font=bold)

scenarios = [
    ('❌ Тест нижней границы — 1500', 1500, False,
     'Ниже исторического оптимума. Нет данных по этой цене на Ozon.',
     'Только для теста эластичности. Маржа минимальная.'),
    ('❌ 1500 + реклама 5%', 1500, True,
     'Нижний порог + продвижение.',
     'Маржа под угрозой. Не рекомендуется без теста объёма.'),
    ('✅ Исторический оптимум — 1700', 1700, False,
     'История: сент–окт 2025 @ 1700 → 3–4 шт/мес, +394…+491 руб/шт, комис. 14%.',
     'РЕКОМЕНДУЕТСЯ для теста: лучшая цена по истории, хорошая маржа.'),
    ('✅ 1700 + реклама 5%', 1700, True,
     'Исторически лучший результат + продвижение.',
     'ГИПОТЕЗА: 5–8 шт/мес. Целевой сценарий при обновлённой карточке.'),
    ('✅ Средняя точка — 1900', 1900, False,
     'Равна min_price. Нет истории на этом уровне. Выше исторического max Ozon.',
     'Хорошая маржа. Протестировать: конвертирует ли трафик.'),
    ('✅ 1900 + реклама 5%', 1900, True,
     'Промежуточная цена + продвижение.',
     'ГИПОТЕЗА: баланс маржи и объёма. Тестировать до 2100.'),
    ('⚡ Текущая цена — 2100', 2100, False,
     'Текущая цена в Ozon Seller. Продаж нет (0 шт в фев 2026).',
     'ПРОБЛЕМА: нет продаж. Цена выше рынка или слабая карточка.'),
    ('⚡ 2100 + реклама 5%', 2100, True,
     'Текущая цена + платное продвижение.',
     'ГИПОТЕЗА: 1–2 шт/мес при рекламе. Маржа есть, объём низкий.'),
    ('⚠️ Премиум — 2500', 2500, False,
     'Выше текущей. Нет исторических данных на Ozon.',
     'ГИПОТЕЗА: 0–1 шт/мес. Только при сильном бренде и отзывах.'),
]

for i, (name, price, with_adv, position, rec) in enumerate(scenarios, 15):
    comm_rub = round(price * get_comm(price))
    adv_rub  = round(price * 0.05) if with_adv else 0
    expenses = comm_rub + FBO + adv_rub
    profit   = price - expenses - COST
    margin   = round(profit / price * 100, 1)
    ppl_conc  = round(price / 5)
    ppl_ready = round(price / (5 * 35), 2)

    f = GREEN if profit > 200 else (YELLOW if profit > 0 else RED_FILL)
    cell(ws1, i,  1, name,      fill=GRAY, align=left)
    cell(ws1, i,  2, price,     fill=f,    nf='#,##0')
    cell(ws1, i,  3, comm_rub,  fill=f,    nf='#,##0')
    cell(ws1, i,  4, FBO,       fill=f,    nf='#,##0')
    cell(ws1, i,  5, adv_rub,   fill=f,    nf='#,##0')
    cell(ws1, i,  6, expenses,  fill=f,    nf='#,##0')
    cell(ws1, i,  7, COST,      fill=f,    nf='#,##0')
    cell(ws1, i,  8, profit,    fill=f,    nf='+#,##0;-#,##0;0')
    cell(ws1, i,  9, margin,    fill=f,    nf='0.0"%"')
    cell(ws1, i, 10, ppl_conc,  fill=GRAY, nf='#,##0')
    cell(ws1, i, 11, ppl_ready, fill=GRAY, nf='0.00')
    cell(ws1, i, 12, position,  fill=GRAY, align=left)
    cell(ws1, i, 13, rec,       fill=GRAY, align=left)

widths1 = [30, 14, 15, 14, 13, 14, 12, 14, 10, 12, 12, 40, 40]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[1].height  = 28
ws1.row_dimensions[13].height = 28
ws1.row_dimensions[14].height = 32
for r in range(15, 25):
    ws1.row_dimensions[r].height = 40
ws1.freeze_panes = 'A15'

# ══════════════════════════════════════════════════════
# ЛИСТ 2 — ИСТОРИЯ OZON
# ══════════════════════════════════════════════════════
ws2 = wb_out.create_sheet('История Ozon')
cell(ws2, 1, 1, 'ИСТОРИЯ ПРОДАЖ OZON: Orange Standard 5кг (Art. 119105)', fill=ORANGE_F, font=bold12)
ws2.merge_cells('A1:K1')

h2 = ['Период', 'Продано, шт', 'Ср. цена, руб', 'Комиссия %', 'Скидки, руб',
      'Хранение, руб', 'Прибыль/шт', 'Маржа %', 'Выручка gross', 'Приход факт', 'Примечание']
for col, h in enumerate(h2, 1):
    cell(ws2, 2, col, h, fill=BLUE, font=bold)

notes_oz = {
    '2025-01': 'Убыток. Цена 837 + высокая комиссия 26%.',
    '2025-02': 'Минимальная прибыль +12 руб/шт. Цена 862.',
    '2025-03': 'УБЫТОК -63 руб/шт. Комиссия 39% — аномалия.',
    '2025-04': 'УБЫТОК -128 руб/шт. Комиссия 37% — аномалия.',
    '2025-05': 'Прибыльно +246 руб/шт. Цена 1203.',
    '2025-06': 'Прибыльно +177 руб/шт. Цена 1174.',
    '2025-07': 'Почти ноль -4 руб/шт. Хранение съедает.',
    '2025-08': 'Прибыльно +304 руб/шт. Цена 1480.',
    '2025-09': 'ТОП МАРЖА +491 руб/шт (28.9%). Цена 1700.',
    '2025-10': 'ТОП ОБЪЁМ+МАРЖА: 4 шт @ 1700, +394 руб/шт.',
    '2025-11': 'Нет продаж. Только хранение.',
    '2025-12': 'Нет продаж. Только хранение.',
    '2026-02': 'Нет продаж. Цена 2100 — нет объёма.',
}

for i, row in ozon.iterrows():
    r = i + 3
    sold  = int(row['Продано'])
    price = int(row['Ср_цена']) if sold > 0 else 0
    comm  = row['Комис_pct']
    skid  = round(row['Скидки'])  if sold > 0 else 0
    stor  = round(row['Хранение'])
    prof  = int(row['Прибыль_шт']) if sold > 0 else 0
    marg  = round(row['Маржа_pct'] * 100, 1)
    gross = round(row['Выручка'])
    inc   = round(row['Приход'])
    note  = notes_oz.get(row['Период_str'], '')

    if sold == 0:
        f = DGRAY
    elif prof > 200:
        f = GREEN
    elif prof > 0:
        f = YELLOW
    else:
        f = RED_FILL

    cell(ws2, r,  1, row['Период_str'], fill=f)
    cell(ws2, r,  2, sold  if sold > 0 else '-', fill=f, nf='#,##0' if sold > 0 else '@')
    cell(ws2, r,  3, price if price > 0 else '-', fill=f, nf='#,##0' if price > 0 else '@')
    cell(ws2, r,  4, f'{comm:.0f}%' if sold > 0 else '-', fill=f)
    cell(ws2, r,  5, skid  if sold > 0 else '-', fill=f, nf='#,##0' if sold > 0 else '@')
    cell(ws2, r,  6, stor  if stor != 0 else '-', fill=f, nf='#,##0' if stor != 0 else '@')
    cell(ws2, r,  7, prof  if sold > 0 else '-', fill=f, nf='+#,##0;-#,##0;0' if sold > 0 else '@')
    cell(ws2, r,  8, f'{marg:.1f}%' if sold > 0 else '-', fill=f)
    cell(ws2, r,  9, gross if gross > 0 else '-', fill=f, nf='#,##0' if gross > 0 else '@')
    cell(ws2, r, 10, inc,  fill=f, nf='#,##0')
    cell(ws2, r, 11, note, fill=f, align=left)

total_r = len(ozon) + 3
oz_sell = ozon[ozon['Продано'] > 0]
cell(ws2, total_r, 1, 'ИТОГО', fill=ORANGE_F, font=bold)
cell(ws2, total_r, 2, int(ozon['Продано'].sum()), fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws2, total_r, 3, int(oz_sell['Ср_цена'].mean()) if len(oz_sell) > 0 else 0, fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws2, total_r, 7,
     int(ozon['Прибыль'].sum() / ozon[ozon['Продано']>0]['Продано'].sum()),
     fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0;0')
cell(ws2, total_r,  9, int(ozon['Выручка'].sum()), fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws2, total_r, 10, int(ozon['Приход'].sum()),  fill=ORANGE_F, font=bold, nf='#,##0')
for col in [4, 5, 6, 8, 11]:
    cell(ws2, total_r, col, '', fill=ORANGE_F)

widths2 = [12, 13, 14, 12, 13, 14, 14, 10, 14, 13, 50]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 28
ws2.row_dimensions[2].height = 32
ws2.freeze_panes = 'A3'

# ══════════════════════════════════════════════════════
# ЛИСТ 3 — WB vs OZON
# ══════════════════════════════════════════════════════
ws3 = wb_out.create_sheet('WB vs Ozon')
cell(ws3, 1, 1, 'СРАВНЕНИЕ WB vs OZON: Orange Standard 5кг (Art. 119105) — по месяцам',
     fill=ORANGE_F, font=bold12)
ws3.merge_cells('A1:N1')

h3 = ['Период',
      'WB шт', 'WB ср.цена', 'WB маржа%', 'WB прибыль/шт', 'WB выручка',
      '│',
      'OZ шт', 'OZ ср.цена', 'OZ маржа%', 'OZ прибыль/шт', 'OZ выручка',
      '│', 'Разница объём']
for col, h in enumerate(h3, 1):
    cell(ws3, 2, col, h, fill=BLUE, font=bold)

periods = sorted(set(list(ozon['Период_str']) + list(wb_d['Период_str'])))
for i, p in enumerate(periods, 3):
    wb_r = wb_d[wb_d['Период_str'] == p]
    oz_r = ozon[ozon['Период_str'] == p]

    def get_vals(sub):
        if sub.empty: return 0, 0, 0.0, 0, 0
        rr = sub.iloc[0]
        sold = int(rr['Продано'])
        pr   = int(rr['Ср_цена'])    if sold > 0 else 0
        mg   = round(rr['Маржа_pct']*100, 1) if sold > 0 else 0.0
        pf   = int(rr['Прибыль_шт']) if sold > 0 else 0
        gr   = int(rr['Выручка'])
        return sold, pr, mg, pf, gr

    ws_s, wp, wm, wpr, wg = get_vals(wb_r)
    os_s, op, om, opr, og = get_vals(oz_r)
    diff = ws_s - os_s

    fw = GREEN if wm > 15 else (YELLOW if wm > 0 else (RED_FILL if ws_s > 0 else DGRAY))
    fo = GREEN if om > 15 else (YELLOW if om > 0 else (RED_FILL if os_s > 0 else DGRAY))

    cell(ws3, i,  1, p,    fill=GRAY)
    cell(ws3, i,  2, ws_s if ws_s > 0 else '-', fill=fw, nf='#,##0' if ws_s > 0 else '@')
    cell(ws3, i,  3, wp   if wp   > 0 else '-', fill=fw, nf='#,##0' if wp > 0 else '@')
    cell(ws3, i,  4, f'{wm:.1f}%' if ws_s > 0 else '-', fill=fw)
    cell(ws3, i,  5, wpr  if ws_s > 0 else '-', fill=fw, nf='+#,##0;-#,##0;0' if ws_s > 0 else '@')
    cell(ws3, i,  6, wg   if wg   > 0 else '-', fill=fw, nf='#,##0' if wg > 0 else '@')
    cell(ws3, i,  7, '│',  fill=GRAY)
    cell(ws3, i,  8, os_s if os_s > 0 else '-', fill=fo, nf='#,##0' if os_s > 0 else '@')
    cell(ws3, i,  9, op   if op   > 0 else '-', fill=fo, nf='#,##0' if op > 0 else '@')
    cell(ws3, i, 10, f'{om:.1f}%' if os_s > 0 else '-', fill=fo)
    cell(ws3, i, 11, opr  if os_s > 0 else '-', fill=fo, nf='+#,##0;-#,##0;0' if os_s > 0 else '@')
    cell(ws3, i, 12, og   if og   > 0 else '-', fill=fo, nf='#,##0' if og > 0 else '@')
    cell(ws3, i, 13, '│',  fill=GRAY)
    diff_str = f'WB x{diff//max(os_s,1)}' if os_s > 0 and ws_s > 0 else '-'
    cell(ws3, i, 14, diff_str, fill=GRAY)

total_r3 = len(periods) + 3
for col in range(1, 15):
    cell(ws3, total_r3, col, '', fill=ORANGE_F)

wb_total = int(wb_d['Продано'].sum())
oz_total = int(ozon['Продано'].sum())
cell(ws3, total_r3,  1, 'ИТОГО', fill=ORANGE_F, font=bold)
cell(ws3, total_r3,  2, wb_total, fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws3, total_r3,  3, int(wb_d['Выручка'].sum()/max(wb_total,1)), fill=ORANGE_F, font=bold, nf='#,##0')
wb_sell = wb_d[wb_d['Продано']>0]
cell(ws3, total_r3,  5, int(wb_d['Прибыль'].sum()/max(len(wb_sell),1)/max(wb_sell['Продано'].mean(),1)), fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0;0')
cell(ws3, total_r3,  6, int(wb_d['Выручка'].sum()), fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws3, total_r3,  8, oz_total, fill=ORANGE_F, font=bold, nf='#,##0')
oz_sell2 = ozon[ozon['Продано']>0]
cell(ws3, total_r3,  9, int(ozon['Выручка'].sum()/max(oz_total,1)), fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws3, total_r3, 11, int(ozon['Прибыль'].sum()/max(oz_total,1)), fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0;0')
cell(ws3, total_r3, 12, int(ozon['Выручка'].sum()), fill=ORANGE_F, font=bold, nf='#,##0')
cell(ws3, total_r3, 14, f'WB x{round(wb_total/max(oz_total,1))}', fill=ORANGE_F, font=bold)

widths3 = [10, 8, 10, 10, 14, 12, 3, 8, 10, 10, 14, 12, 3, 13]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[1].height = 28
ws3.row_dimensions[2].height = 32
ws3.freeze_panes = 'A3'

# ── Сохранение Excel ──
OUT = 'C:/Users/1/Downloads/pricing_119105_2026-03-25.xlsx'
wb_out.save(OUT)
print('Excel saved:', OUT)

# ── Сохранение DOCX профиля ──
folder = get_product_folder("ifoam", "OrangeStandard5kg", "119105")
print('Folder:', folder)

oz_sell_rows = ozon[ozon['Продано'] > 0]
wb_sell_rows = wb_d[wb_d['Продано'] > 0]

best_margin_row = oz_sell_rows.loc[oz_sell_rows['Маржа_pct'].idxmax()]
best_vol_row    = oz_sell_rows.loc[oz_sell_rows['Продано'].idxmax()]
worst_row       = oz_sell_rows.loc[oz_sell_rows['Прибыль_шт'].idxmin()]

profile_text = f"""=== ПРОФИЛЬ ПРОДАЖ ===
Найденный файл данных: DataExport_2026-03-24_13-35-51.xlsx (Вариант А — DataLens)
Платформа(ы): Ozon + Wildberries
Период данных: 2025-01 — 2026-02
Всего продано штук (Ozon): {int(ozon['Продано'].sum())} шт
Всего продано штук (WB): {int(wb_d['Продано'].sum())} шт
Среднемесячный объём (Ozon, месяцы с продажами): {round(oz_sell_rows['Продано'].mean(), 1)} шт/мес
Текущая цена: 2100 руб (зачёркн. 2900, min 1900)
Себестоимость: {COST} руб

Ценовой анализ:
  Прибыльные ценовые диапазоны (Ozon):
    1174–1203 руб → маржа 15–20%, +177…+246 руб/шт
    1480 руб       → маржа 20.5%, +304 руб/шт
    1700 руб       → маржа 23–29%, +394…+491 руб/шт (ОПТИМУМ)
  Убыточные ценовые диапазоны:
    819–862 руб    → убыток, комиссия 22–37% (аномалии/акции)
    1030 руб       → убыток, комиссия 39% (аномалия)
    1260 руб       → почти ноль -4 руб/шт
  Оптимальная цена по данным: 1700 руб → маржа ~26%, объём 3–4 шт/мес

Лучший месяц (маржа): {best_margin_row['Период_str']} — {int(best_margin_row['Продано'])} шт @ {int(best_margin_row['Ср_цена'])} руб → маржа {round(best_margin_row['Маржа_pct']*100,1)}%
Лучший месяц (объём): {best_vol_row['Период_str']} — {int(best_vol_row['Продано'])} шт @ {int(best_vol_row['Ср_цена'])} руб → маржа {round(best_vol_row['Маржа_pct']*100,1)}%
Худший месяц: {worst_row['Период_str']} — убыток {int(worst_row['Прибыль_шт'])} руб/шт @ {int(worst_row['Ср_цена'])} руб
Месяцев без продаж (Ozon): 3 (ноябрь, декабрь 2025, февраль 2026) — потери на хранение

Тренд цены (Ozon): растёт — с ~850 руб (янв 2025) до 2100 руб (текущая)
Тренд объёма (Ozon): нестабильно — пик 5 шт (апр 2025 на акции), затем 2–4 шт/мес, сейчас 0

Аномалии:
  - Апрель 2025: комиссия 37% — возможно акция или штраф. Объём 5 шт, убыток -128 руб/шт.
  - Март 2025: комиссия 39% — самая высокая в истории. Причина неизвестна.
  - Ноябрь–декабрь 2025 и февраль 2026: нулевые продажи на Ozon при активных продажах на WB.
  - WB продаёт в 5–10x больше Ozon при более низкой цене (860–1420 руб).

Главный вывод: Исторически оптимальная цена для Ozon — 1700 руб (маржа 23–29%, 3–4 шт/мес).
Текущая цена 2100 руб слишком высока — продажи остановились. Рекомендуется снизить до 1700–1900 руб
и улучшить карточку товара для восстановления органического трафика.
======================"""

docx_path = make_docx(folder, "02_sales_profile_119105.docx", "Профиль продаж — Orange Standard 5кг", profile_text)
print('DOCX saved:', docx_path)
