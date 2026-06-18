import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'Orange Standard 5кг'

YELLOW   = PatternFill('solid', fgColor='FFD966')
GREEN    = PatternFill('solid', fgColor='C6EFCE')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
BLUE     = PatternFill('solid', fgColor='BDD7EE')
GRAY     = PatternFill('solid', fgColor='F2F2F2')
ORANGE_F = PatternFill('solid', fgColor='F4B942')

def brd():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def c(ws, row, col, value, fill=None, font=None, align=None, nf=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.alignment = align or Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = brd()
    if nf: cell.number_format = nf
    return cell

left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
bold = Font(bold=True)

# ── Блок 1: Исходные данные ──
c(ws,1,1,'ИСХОДНЫЕ ДАННЫЕ — Orange Standard 5кг (Art. 119105)', fill=ORANGE_F, font=Font(bold=True, size=12))
ws.merge_cells('A1:M1')

for col, h in enumerate(['Параметр','Значение','Примечание'], 1):
    c(ws, 2, col, h, fill=BLUE, font=bold)

params = [
    ('Себестоимость, руб',  496,  'производство + упаковка'),
    ('Объём, кг',            5,   'концентрат'),
    ('Разбавление',          '1:30-1:40', 'для пеногенератора'),
    ('Логистика FBO, руб',  250,  'хранение + доставка Ozon'),
    ('Комиссия Ozon, %',     30,  '% от цены продажи'),
    ('Реклама (ДРР), %',      5,  'если используется'),
]
for i, (name, val, note) in enumerate(params, 3):
    c(ws, i, 1, name, fill=GRAY, align=left_al)
    c(ws, i, 2, val,  fill=GRAY)
    c(ws, i, 3, note, fill=GRAY, align=left_al)

# ── Блок 2: Сценарии ──
c(ws,10,1,'СЦЕНАРИИ ЦЕНООБРАЗОВАНИЯ', fill=ORANGE_F, font=Font(bold=True, size=12))
ws.merge_cells('A10:M10')

hdrs = ['Сценарий','Цена, руб','Комиссия Ozon, руб','Логистика, руб','Реклама, руб',
        'Итого расходы, руб','Себест., руб','Прибыль с ед., руб','Маржа, %',
        'Цена/л конц., руб','Цена/л р-ра, руб','Позиция vs конкурент','Рекомендация']
for col, h in enumerate(hdrs, 1):
    c(ws, 11, col, h, fill=BLUE, font=bold)

scenarios = [
    ('Агрессивный (как GLITTER G5)',   500,  True,
     'GLITTER G5: 418 руб (акции 28/28 дней). ИСТОРИЯ: при 819 руб был убыток -128 руб/шт.',
     'УБЫТОЧНО однозначно. Комиссия Ozon при низкой цене достигает 37%. Не рассматривать.'),
    ('Нижняя граница, без рекламы',    900,  False,
     'ИСТОРИЯ: при 862-1260 руб комиссия 18-39%, все периоды убыточны или около нуля.',
     'НЕ РЕКОМЕНДУЕТСЯ. Исторически доказано: цены до 1260 руб дают убыток или 0.'),
    ('Нижняя граница + реклама',       900,  True,
     'BIGHIM — 648 руб (770% динамика роста). ИСТОРИЯ: при ~900 руб убыток.',
     'УБЫТОЧНО С РЕКЛАМОЙ. Реклама съест остаток маржи. Не рассматривать.'),
    ('Исторически оптимальная',       1700,  False,
     'ИСТОРИЯ: сент-окт 2025 @ 1700 руб -> 3-4 шт/мес, прибыль +394..+491 руб/шт, комиссия 14%.',
     'РЕКОМЕНДУЕТСЯ как базовая. Доказано историей. Нужно вернуться к этой цене + улучшить карточку.'),
    ('Оптимальная + реклама 5%',      1700,  True,
     'Тот же уровень 1700 руб с платным продвижением.',
     'РЕКОМЕНДУЕТСЯ для роста объёма. При конверсии 0.4% реклама окупится. Целевой сценарий.'),
    ('Выше оптимума',                 1900,  False,
     'DR.BERG ACE — 1903 руб, 562 заказа/28 дн. ИСТОРИЯ: мар 2026 @ 2100 руб -> 1 шт/мес.',
     'Риск потери объёма. Оправдано только при сильном УТП и профессиональной карточке.'),
    ('Премиум (как DR.BERG + реклама)',1900,  True,
     'DR.BERG тратит 10.3% на рекламу при цене 1903 руб.',
     'Маржа приемлема. Требует инвестиций в контент и рекламу одновременно.'),
    ('Текущая цена (мар 2026)',        2100,  False,
     'ИСТОРИЯ: мар 2026 @ 2100 руб -> 1 шт за месяц. Комиссия 10%.',
     'Маржа высокая, но объём нулевой. Товар стоит без движения — убыток от хранения.'),
]

cost = 496
fbo  = 250
comm = 0.30
adv_pct = 0.05

for i, (name, price, with_adv, position, rec) in enumerate(scenarios, 12):
    comm_rub = round(price * comm)
    adv_rub  = round(price * adv_pct) if with_adv else 0
    expenses = comm_rub + fbo + adv_rub
    profit   = price - expenses - cost
    margin   = round(profit / price * 100, 1) if price > 0 else 0
    ppl_conc = round(price / 5, 0)
    ppl_ready = round(price / (5 * 35), 2)

    f = GREEN if profit > 150 else (YELLOW if profit > 0 else RED_FILL)

    c(ws, i, 1,  name,       fill=GRAY, align=left_al)
    c(ws, i, 2,  price,      fill=f,    nf='#,##0')
    c(ws, i, 3,  comm_rub,   fill=f,    nf='#,##0')
    c(ws, i, 4,  fbo,        fill=f,    nf='#,##0')
    c(ws, i, 5,  adv_rub,    fill=f,    nf='#,##0')
    c(ws, i, 6,  expenses,   fill=f,    nf='#,##0')
    c(ws, i, 7,  cost,       fill=f,    nf='#,##0')
    c(ws, i, 8,  profit,     fill=f,    nf='#,##0')
    c(ws, i, 9,  margin,     fill=f,    nf='0.0')
    c(ws, i, 10, ppl_conc,   fill=GRAY, nf='#,##0')
    c(ws, i, 11, ppl_ready,  fill=GRAY, nf='0.00')
    c(ws, i, 12, position,   fill=GRAY, align=left_al)
    c(ws, i, 13, rec,        fill=GRAY, align=left_al)

# ── Блок 3: Конкуренты ──
start = 23
c(ws, start, 1, 'КОНКУРЕНТЫ — бесконтактная мойка 5кг (топ-10 по выручке, 28 дней)', fill=ORANGE_F, font=Font(bold=True, size=12))
ws.merge_cells(f'A{start}:M{start}')

ch = ['#','Товар','Бренд','Объём, кг','Ср. цена, руб','Цена экв. 5кг, руб',
      'Выручка 28д','Заказов','Дней без ост.','Динамика %','Дней в акциях','Дней реклама','ДРР %']
for col, h in enumerate(ch, 1):
    c(ws, start+1, col, h, fill=BLUE, font=bold)

# num, name, brand, volume_kg, price, revenue, orders, oos, dyn, promo, adv, ddr
competitors = [
    (1,  'GLITTER G5, 5л',                  'GLITTER',   5.0,  418,   5034446, 12038, '15/28', 239, '28/28', '28/28', 11.5),
    (2,  'GRASS Active Foam Balance, 5кг',   'Grass',     5.0,  611,   4812674,  7883, '13/28', 390, '27/28', '0/28',   7.0),
    (3,  'GRASS Active Foam Pink, 6кг',      'Grass',     6.0, 1620,   3460179,  2136, '-',     465, '0/28',  '0/28',   0.8),
    (4,  'GRASS Active Foam Red, 5.8кг',     'Grass',     5.8, 1252,   3159342,  2523, '-',     221, '0/28',  '0/28',   1.1),
    (5,  'GRASS Active Foam Red, 5.8кг (SKU2)','Grass',   5.8, 1256,   2893082,  2303, '7/28',  397, '18/28', '0/28',   4.7),
    (6,  'GLITTER G10, 5л',                  'GLITTER',   5.0,  554,   2728526,  4923, '17/28', 276, '13/28', '28/28', 10.7),
    (7,  'GRASS Active Foam Pink, 6кг (SKU2)','Grass',    6.0, 1744,   2527411,  1449, '5/28',  158, '18/28', '0/28',   4.1),
    (8,  'DR.BERG ACE, 5кг',                 'DR. BERG',  5.0, 1903,   1069666,   562, '14/28', 281, '28/28', '0/28',  10.3),
    (9,  'BIGHIM Active Foam Expert, 5кг',   'BIGHIM',    5.0,  648,   1027756,  1585, '25/28', 770, '28/28', '28/28',  1.0),
    (10, 'GRASS Active Foam Effect, 6кг',    'Grass',     6.0, 1422,    935861,   658, '-',     386, '0/28',  '0/28',   0.2),
]

for i, row_data in enumerate(competitors, start+2):
    num, name, brand, vol, price, rev, orders, oos, dyn, promo, adv, ddr = row_data
    price_5kg = round(price * 5 / vol)
    f = GREEN if num <= 2 else GRAY
    c(ws, i, 1,  num,       fill=f)
    c(ws, i, 2,  name,      fill=f, align=left_al)
    c(ws, i, 3,  brand,     fill=f, align=left_al)
    c(ws, i, 4,  vol,       fill=f, nf='0.0')
    c(ws, i, 5,  price,     fill=f, nf='#,##0')
    c(ws, i, 6,  price_5kg, fill=YELLOW if price != price_5kg else f, nf='#,##0')
    c(ws, i, 7,  rev,       fill=f, nf='#,##0')
    c(ws, i, 8,  orders,    fill=f, nf='#,##0')
    c(ws, i, 9,  oos,       fill=f)
    c(ws, i, 10, dyn,       fill=f)
    c(ws, i, 11, promo,     fill=f)
    c(ws, i, 12, adv,       fill=f)
    c(ws, i, 13, ddr,       fill=f)

# Ширина колонок
widths = [30, 16, 16, 16, 14, 16, 14, 16, 10, 18, 16, 35, 35]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for r in [1, 10, 11, 23, 24]:
    ws.row_dimensions[r].height = 32
for r in range(12, 22):
    ws.row_dimensions[r].height = 22
for r in range(25, 36):
    ws.row_dimensions[r].height = 22

# Заморозить первую строку
ws.freeze_panes = 'A12'

# ── Лист 2: История продаж ──
ws2 = wb.create_sheet('История продаж')

c(ws2,1,1,'ИСТОРИЯ ПРОДАЖ: Orange Standard 5кг (Art. 119105) | 14 месяцев', fill=ORANGE_F, font=Font(bold=True, size=12))
ws2.merge_cells('A1:J1')

hist_headers = ['Период','Продано, шт','Ср. цена, руб','Комиссия Ozon %','Хранение, руб',
                'Прибыль/шт, руб','Маржа %','Выручка gross','Приход факт','Примечание']
for col, h in enumerate(hist_headers, 1):
    c(ws2, 2, col, h, fill=BLUE, font=bold)

history = [
    ('2025-01',  1,   837, 26,  -8,   -37, -4.5,   837,   459, 'Убыток. Низкая цена + высокая комиссия.'),
    ('2025-02',  3,   862, 23, -10,   +12,  1.4,  2586,  1525, 'Безубыточно. Минимум.'),
    ('2025-03',  2,  1030, 39, -14,   -63, -6.1,  2060,   866, 'УБЫТОК. Комиссия 39% — аномалия.'),
    ('2025-04',  5,   819, 37, -56,  -128,-15.6,  4094,  1840, 'УБЫТОК. Макс объём, мин маржа. Акции.'),
    ('2025-05',  3,  1203, 18, -46,  +246, 20.5,  3609,  2227, 'Прибыльно. Хорошая цена.'),
    ('2025-06',  3,  1174, 19, -49,  +177, 15.1,  3522,  2018, 'Прибыльно.'),
    ('2025-07',  2,  1260, 18, -11,    -4, -0.3,  2520,   984, 'Около нуля. Хранение съедает.'),
    ('2025-08',  2,  1480, 15, -34,  +304, 20.5,  2960,  1599, 'Прибыльно. Цена растёт.'),
    ('2025-09',  3,  1700, 14, -59,  +491, 28.9,  5100,  2962, 'ТОП МАРЖА. Лучший месяц.'),
    ('2025-10',  4,  1700, 14, -20,  +394, 23.2,  6800,  3562, 'ТОП ОБЪЁМ+МАРЖА. Целевая цена.'),
    ('2025-11',  0,     0,  0,   0,     0,  0.0,     0,  -364, 'Нет продаж. Хранение = убыток.'),
    ('2025-12',  0,     0,  0,   0,     0,  0.0,     0,  -251, 'Нет продаж. Хранение = убыток.'),
    ('2026-02',  0,     0,  0,   0,     0,  0.0,     0,  -319, 'Нет продаж. Хранение = убыток.'),
    ('2026-03',  1,  2100, 10,   0,  +440, 20.9,  2100,   936, 'Высокая цена -> 1 шт/мес. Нет объёма.'),
]

for i, row_data in enumerate(history, 3):
    period, sold, price, comm_pct, stor, profit_u, margin, gross, income, note = row_data
    if sold > 0:
        f = GREEN if profit_u > 200 else (YELLOW if profit_u > 0 else RED_FILL)
    else:
        f = RED_FILL
    c(ws2, i, 1,  period,    fill=f)
    c(ws2, i, 2,  sold,      fill=f, nf='#,##0')
    c(ws2, i, 3,  price,     fill=f, nf='#,##0')
    c(ws2, i, 4,  comm_pct,  fill=f, nf='0"%"')
    c(ws2, i, 5,  stor,      fill=f, nf='#,##0')
    c(ws2, i, 6,  profit_u,  fill=f, nf='+#,##0;-#,##0;0')
    c(ws2, i, 7,  margin,    fill=f, nf='0.0"%"')
    c(ws2, i, 8,  gross,     fill=f, nf='#,##0')
    c(ws2, i, 9,  income,    fill=f, nf='#,##0')
    c(ws2, i, 10, note,      fill=f, align=left_al)

# Итого
r_tot = len(history) + 3
c(ws2, r_tot, 1, 'ИТОГО / СРЕДНЕЕ', fill=ORANGE_F, font=bold)
c(ws2, r_tot, 2, 29,      fill=ORANGE_F, font=bold, nf='#,##0')
c(ws2, r_tot, 3, 1247,    fill=ORANGE_F, font=bold, nf='#,##0')
c(ws2, r_tot, 4, '',      fill=ORANGE_F)
c(ws2, r_tot, 5, '',      fill=ORANGE_F)
c(ws2, r_tot, 6, 126,     fill=ORANGE_F, font=bold, nf='+#,##0;-#,##0;0')
c(ws2, r_tot, 7, 10.1,    fill=ORANGE_F, font=bold, nf='0.0"%"')
c(ws2, r_tot, 8, 36188,   fill=ORANGE_F, font=bold, nf='#,##0')
c(ws2, r_tot, 9, 18043,   fill=ORANGE_F, font=bold, nf='#,##0')
c(ws2, r_tot, 10,'2.1 шт/мес за 14 мес. Оптимальная цена: 1700 руб.', fill=ORANGE_F, font=bold, align=left_al)

# Ширина
for i, w in enumerate([12,12,14,14,12,14,10,14,12,45], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 30

wb.save('C:/Users/1/Downloads/orange_standard_pricing.xlsx')
print('Готово: C:/Users/1/Downloads/orange_standard_pricing.xlsx')
