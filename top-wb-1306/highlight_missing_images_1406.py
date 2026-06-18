import sys, re, openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill
sys.stdout.reconfigure(encoding="utf-8")

XLSX = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1306\Топ-500 ВБ 1406.xlsx")
DST  = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1406\Изображения")

YELLOW = PatternFill("solid", fgColor="FFFF00")
CLEAR  = PatternFill(fill_type=None)

IMG_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".jfif"}

# ── Индексируем что уже есть в папке назначения ──────────────────────────────
existing = {f.name.lower() for f in DST.iterdir() if f.is_file() and f.suffix.lower() in IMG_EXTS}
print(f"В папке назначения: {len(existing)} файлов")

# ── Открываем Excel ───────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX)
ws = wb.active

highlighted = 0
cleared = 0

for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 15).value
    art_cell = ws.cell(r, 1)
    img_cell = ws.cell(r, 15)

    if not v:
        continue

    raw = str(v).replace(";", ",")
    files = [f.strip().lower() for f in raw.split(",") if f.strip()]

    # Есть хотя бы один ненайденный файл?
    missing_any = any(f not in existing for f in files)

    if missing_any:
        art_cell.fill = YELLOW
        img_cell.fill = YELLOW
        highlighted += 1
    else:
        # Снимаем подсветку если была
        if art_cell.fill and art_cell.fill.fill_type not in (None, "none"):
            art_cell.fill = CLEAR
            cleared += 1

wb.save(XLSX)
wb.close()
print(f"Подсвечено строк: {highlighted}")
print(f"Снято с найденных: {cleared}")
print("Готово!")
