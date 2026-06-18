"""
fill_dimensions_1406.py
══════════════════════════════════════════════════════
Заполняет Вес, Длина, Ширина, Высота (упаковка) через GigaChat (браузер)
для строк, где хоть одно из четырёх значений пустое.

Файл: C:\\Users\\Admin\\Desktop\\Топ ВБ 1306\\Топ-500 ВБ 1406.xlsx
Сохраняет поверх исходника.

Первый запуск / протухла сессия:
  uv run --with playwright,openpyxl fill_dimensions_1406.py --login

Основной запуск:
  uv run --with playwright,openpyxl fill_dimensions_1406.py

Тест на 5 строках:
  uv run --with playwright,openpyxl fill_dimensions_1406.py --rows 2-6
"""

import re
import sys
import json
import time
import argparse
import traceback
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("Playwright не установлен.")
    print("  uv run --with playwright python -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("openpyxl не установлен.")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

FILE_PATH    = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1306\Топ-500 ВБ 1406.xlsx")
PROFILE_DIR  = Path.home() / ".gigachat_dimensions_playwright"
GIGACHAT_URL = "https://giga.chat/"

BATCH_SIZE = 10   # строк за один запрос
SAVE_EVERY = 3    # сохранять каждые N батчей

DIM_COLS   = ["Вес, г", "Длина, мм", "Ширина, мм", "Высота, мм"]
PROMPT_MARKER = "Ты эксперт по автозапчастям"


# ══════════════════════════════════════════════════════════════════════════════
#  ПРОМПТ
# ══════════════════════════════════════════════════════════════════════════════

INSTRUCTION = (
    "Ты эксперт по автозапчастям. Для каждой позиции ниже укажи "
    "вес и габариты ТРАНСПОРТНОЙ УПАКОВКИ — внешние размеры коробки или пакета "
    "вместе с самой деталью внутри, именно так, как товар будет отправлен покупателю.\n"
    "ВАЖНО: это НЕ размеры самой детали, а размеры упаковки снаружи.\n"
    "Правила:\n"
    "- Если деталь идёт комплектом (2 шт. и т.п.) — вес/размер всего комплекта в упаковке\n"
    "- Если точных данных в открытых источниках нет — укажи приблизительные теоретические "
    "значения, типичные для данного вида запчасти. Оставлять поля пустыми нельзя.\n"
    "- Верни ТОЛЬКО JSON объект, без пояснений, без markdown блоков:\n"
    '{"items": [{"idx": 0, "вес": 500, "длина": 200, "ширина": 150, "высота": 80}, ...]}\n'
    "вес — граммы (деталь + упаковка), длина/ширина/высота — мм (внешние стороны упаковки), "
    "всё целые числа.\n\n"
)


def build_prompt(items: list[dict]) -> str:
    lines = [INSTRUCTION]
    for item in items:
        lines.append(f"[{item['idx']}]")
        lines.append(f"Артикул: {item['article']}")
        lines.append(f"Наименование: {item['name']}")
        if item.get("brand"):
            lines.append(f"Бренд: {item['brand']}")
        if item.get("params"):
            lines.append(f"Параметры: {str(item['params'])[:300]}")
        if item.get("desc"):
            lines.append(f"Описание: {str(item['desc'])[:200]}")
        # Уже заполненные значения — чтобы ГигаЧат не трогал их (но мы всё равно берём все)
        known = []
        for k, v in [("вес", item.get("cur_w")), ("длина", item.get("cur_l")),
                     ("ширина", item.get("cur_wd")), ("высота", item.get("cur_h"))]:
            if v is not None:
                known.append(f"{k}={v}")
        if known:
            lines.append(f"Уже известно: {', '.join(known)} — остальное нужно заполнить")
        lines.append("")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  ПАРСИНГ ОТВЕТА
# ══════════════════════════════════════════════════════════════════════════════

def extract_json_items(text: str) -> list[dict]:
    clean = re.sub(r"```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    clean = clean.replace("```", "").strip()

    ITEM_KEYS = {"вес", "длина", "ширина", "высота", "idx"}

    def is_valid(lst):
        return bool(lst) and isinstance(lst[0], dict) and bool(ITEM_KEYS & lst[0].keys())

    def try_parse(s):
        try:
            parsed = json.loads(s.strip())
            if isinstance(parsed, dict) and "items" in parsed:
                items = parsed["items"]
                if isinstance(items, list) and is_valid(items):
                    return items
            if isinstance(parsed, list) and is_valid(parsed):
                return parsed
        except Exception:
            pass
        return None

    result = try_parse(clean)
    if result:
        return result

    for start in [i for i, c in enumerate(clean) if c == '{']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == '{':
                depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result:
                        return result
                    break

    for start in [i for i, c in enumerate(clean) if c == '[']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == '[':
                depth += 1
            elif c == ']':
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result:
                        return result
                    break

    return []


# ══════════════════════════════════════════════════════════════════════════════
#  БРАУЗЕР — GIGACHAT
# ══════════════════════════════════════════════════════════════════════════════

def dismiss_modal(page):
    try:
        page.keyboard.press("Escape")
        time.sleep(0.5)
    except Exception:
        pass
    try:
        backdrop = page.locator('[data-backdrop]').first
        if backdrop.is_visible(timeout=500):
            page.mouse.click(5, 5)
            time.sleep(0.5)
    except Exception:
        pass


INPUT_SELECTORS = [
    '#chat-input-textarea',
    'textarea[placeholder*="Спрос"]',
    'textarea[placeholder*="Напиш"]',
    'textarea[placeholder*="Сообщ"]',
    'textarea[placeholder*="Введ"]',
    '[role="textbox"]',
    'div[contenteditable="true"]',
    'textarea',
]

SEND_SELECTORS = [
    'button[aria-label*="Отправить"]',
    'button[aria-label*="отправить"]',
    'button[aria-label*="Send"]',
    'button[type="submit"]',
    '[data-testid="send-button"]',
    'button[class*="send"]',
    'button[class*="Send"]',
]


def find_input(page):
    time.sleep(1)
    for sel in INPUT_SELECTORS:
        try:
            el = page.locator(sel).last
            if el.is_visible(timeout=2000):
                return el, sel
        except Exception:
            continue

    # Диагностика
    try:
        info = page.evaluate("""() => ({
            url:       location.href,
            title:     document.title,
            textareas: document.querySelectorAll('textarea').length,
            contenteditable: document.querySelectorAll('[contenteditable]').length,
            buttons:   document.querySelectorAll('button').length,
        })""")
        print(f"\n    Диагностика страницы: {info}")
    except Exception:
        pass

    # Скриншот для отладки
    try:
        scr = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1306\_debug_gigachat.png")
        page.screenshot(path=str(scr))
        print(f"    Скриншот: {scr}")
    except Exception:
        pass

    return None, None


def type_prompt(page, text: str):
    dismiss_modal(page)
    inp, sel = find_input(page)
    if inp is None:
        raise RuntimeError("Поле ввода GigaChat не найдено")

    is_contenteditable = "contenteditable" in sel or "textbox" in sel

    if is_contenteditable:
        inp.click()
        time.sleep(0.3)
        # Для contenteditable — вставка через clipboard
        page.evaluate("""(text) => {
            const el = document.querySelector('[contenteditable="true"]') ||
                       document.querySelector('[role="textbox"]');
            if (!el) return;
            el.focus();
            el.innerText = text;
            el.dispatchEvent(new Event('input',  { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        }""", text)
    else:
        page.evaluate("""() => {
            const el = document.querySelector('#chat-input-textarea') ||
                       document.querySelector('textarea');
            if (el) el.focus();
        }""")
        time.sleep(0.3)

        page.evaluate("""(text) => {
            const el = document.querySelector('#chat-input-textarea') ||
                       document.querySelector('textarea');
            if (!el) return;
            const setter = Object.getOwnPropertyDescriptor(
                window.HTMLTextAreaElement.prototype, 'value').set;
            setter.call(el, text);
            el.dispatchEvent(new Event('input',  { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        }""", text)
    time.sleep(0.5)


def click_send(page):
    for sel in SEND_SELECTORS:
        try:
            btn = page.locator(sel).last
            if btn.is_visible(timeout=800):
                btn.click()
                return
        except Exception:
            continue
    # Fallback: Enter в любом найденном поле ввода
    try:
        inp, sel = find_input(page)
        if inp:
            inp.press("Enter")
            return
    except Exception:
        pass
    page.keyboard.press("Enter")


def find_json_on_page(page) -> str:
    """Ищет JSON с реальными числами (не шаблон из промпта) прямо в тексте страницы."""
    try:
        return page.evaluate(r"""() => {
            const text = document.body ? (document.body.innerText || '') : '';

            function extractBlock(startStr) {
                // Все вхождения startStr — берём ПОСЛЕДНЕЕ
                let pos = -1, search = 0;
                while (true) {
                    let found = text.indexOf(startStr, search);
                    if (found === -1) break;
                    pos = found;
                    search = found + 1;
                }
                if (pos === -1) return '';
                let depth = 0;
                const open = startStr[0] === '{' ? '{' : '[';
                const close = open === '{' ? '}' : ']';
                for (let i = pos; i < Math.min(pos + 100000, text.length); i++) {
                    if (text[i] === open) depth++;
                    else if (text[i] === close) {
                        depth--;
                        if (depth === 0) {
                            const block = text.slice(pos, i + 1);
                            // Пропускаем блоки-шаблоны (содержат "..." или очень короткие)
                            if (block.includes('...') || block.length < 100) return '';
                            return block;
                        }
                    }
                }
                return '';
            }

            // Вариант 1: {"items": [...]}
            let result = extractBlock('{"items":');
            if (!result) result = extractBlock('{ "items":');
            // Вариант 2: массив [{...}] с ключом "вес"
            if (!result) {
                // Найдём последний "вес": число
                const vesPat = /"вес"\s*:\s*\d+/g;
                let m, lastIdx = -1;
                while ((m = vesPat.exec(text)) !== null) lastIdx = m.index;
                if (lastIdx !== -1) {
                    let start = text.lastIndexOf('[', lastIdx);
                    if (start !== -1) {
                        let depth = 0;
                        for (let i = start; i < Math.min(start + 100000, text.length); i++) {
                            if (text[i] === '[') depth++;
                            else if (text[i] === ']') {
                                depth--;
                                if (depth === 0) {
                                    const block = text.slice(start, i + 1);
                                    if (!block.includes('...') && block.length >= 100)
                                        result = block;
                                    break;
                                }
                            }
                        }
                    }
                }
            }
            return result;
        }""")
    except Exception:
        return ""


def is_generating(page) -> bool:
    try:
        for sel in [
            '[aria-label*="Стоп"]',
            '[aria-label*="стоп"]',
            '[aria-label*="Stop"]',
            'button[aria-label*="остановить"]',
            '[class*="loading"]',
            '[class*="typing"]',
        ]:
            if page.locator(sel).first.is_visible(timeout=300):
                return True
    except Exception:
        pass
    return False


def wait_for_response(page, timeout_sec: int = 240) -> str:
    """Ждёт появления JSON с 'items' или 'вес' на странице."""
    print("    жду", end="", flush=True)
    time.sleep(3)

    prev, stable = "", 0
    for tick in range(timeout_sec):
        # Сначала пробуем найти JSON прямо на странице
        json_found = find_json_on_page(page)
        if json_found:
            if json_found == prev:
                stable += 1
                if stable >= 2 and not is_generating(page):
                    print(f" готово ({len(json_found)} симв.)")
                    return json_found
            else:
                stable = 0
                if tick % 4 == 0:
                    print(".", end="", flush=True)
            prev = json_found
        else:
            stable = 0
            if tick % 6 == 0:
                print(".", end="", flush=True)
        time.sleep(1.5)

    # Таймаут — скриншот для диагностики
    try:
        scr = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1306\_debug_timeout.png")
        page.screenshot(path=str(scr))
        print(f" таймаут, скриншот: {scr.name}")
    except Exception:
        print(" таймаут")
    return ""


def new_chat(page):
    dismiss_modal(page)
    for sel in [
        'a[href="/"][aria-label*="чат"]',
        'button[aria-label*="новый чат"]',
        'button[aria-label*="Новый чат"]',
        '[data-testid="new-chat"]',
        'a[href="/"]',
    ]:
        try:
            btn = page.locator(sel).first
            if btn.is_visible(timeout=800):
                btn.click()
                time.sleep(1.5)
                dismiss_modal(page)
                return
        except Exception:
            continue
    try:
        page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=20000)
    except Exception:
        pass
    time.sleep(2)
    dismiss_modal(page)


# ══════════════════════════════════════════════════════════════════════════════
#  ОСНОВНАЯ ОБРАБОТКА
# ══════════════════════════════════════════════════════════════════════════════

def process(page, row_range=None):
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    # Читаем заголовки
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

    art_col   = headers.get("Код / Артикул")
    name_col  = headers.get("Наименование")
    brand_col = headers.get("Бренд")
    param_col = headers.get("Параметры")
    desc_col  = headers.get("Описание")
    w_col     = headers.get("Вес, г")
    l_col     = headers.get("Длина, мм")
    wd_col    = headers.get("Ширина, мм")
    h_col     = headers.get("Высота, мм")

    if not all([w_col, l_col, wd_col, h_col]):
        missing = [n for n, c in [("Вес, г", w_col), ("Длина, мм", l_col),
                                   ("Ширина, мм", wd_col), ("Высота, мм", h_col)] if not c]
        print(f"Не найдены столбцы: {missing}")
        wb.close()
        return

    # Отбираем строки с хотя бы одним пустым значением из 4 столбцов
    todos = []
    for r in range(2, ws.max_row + 1):
        if row_range and not (row_range[0] <= r <= row_range[1]):
            continue
        # Пропускаем полностью пустые строки
        if not ws.cell(r, art_col or 1).value:
            continue
        vals = [ws.cell(r, c).value for c in [w_col, l_col, wd_col, h_col]]
        if any(v is None for v in vals):
            todos.append(r)

    total = len(todos)
    if not todos:
        print("Нечего обрабатывать — все строки заполнены.")
        wb.close()
        return

    print(f"Строк для заполнения: {total}")

    processed  = 0
    save_timer = 0

    for batch_no, batch_start in enumerate(range(0, total, BATCH_SIZE), start=1):
        batch_rows = todos[batch_start: batch_start + BATCH_SIZE]

        items = []
        for local_idx, r in enumerate(batch_rows):
            items.append({
                "idx":    local_idx,
                "article": ws.cell(r, art_col).value  if art_col   else "",
                "name":    ws.cell(r, name_col).value  if name_col  else "",
                "brand":   ws.cell(r, brand_col).value if brand_col else "",
                "params":  ws.cell(r, param_col).value if param_col else "",
                "desc":    ws.cell(r, desc_col).value  if desc_col  else "",
                "cur_w":   ws.cell(r, w_col).value,
                "cur_l":   ws.cell(r, l_col).value,
                "cur_wd":  ws.cell(r, wd_col).value,
                "cur_h":   ws.cell(r, h_col).value,
            })

        print(f"  Батч {batch_no} (стр. {batch_rows[0]}–{batch_rows[-1]})", end="", flush=True)

        new_chat(page)
        prompt = build_prompt(items)

        try:
            type_prompt(page, prompt)
        except RuntimeError as e:
            print(f"\n  ✗ {e}")
            continue

        click_send(page)
        time.sleep(1)

        results = []
        for attempt in range(1, 4):
            raw = wait_for_response(page)
            if not raw:
                print(f"  ✗ Пустой ответ (попытка {attempt}/3)")
            else:
                results = extract_json_items(raw)
                if results:
                    break
                print(f"  ⚠ JSON не распознан (попытка {attempt}/3)")
                print(f"    Начало ответа: {raw[:300]}")

            if attempt < 3:
                time.sleep(10)
                new_chat(page)
                type_prompt(page, prompt)
                click_send(page)
                time.sleep(1)

        if not results:
            print(f"  ✗ Батч пропущен после 3 попыток")
            continue

        result_map = {r["idx"]: r for r in results}
        filled = 0
        for local_idx, r in enumerate(batch_rows):
            data = result_map.get(local_idx)
            if not data:
                continue
            # Заполняем только пустые ячейки
            if ws.cell(r, w_col).value  is None: ws.cell(r, w_col).value  = data.get("вес")
            if ws.cell(r, l_col).value  is None: ws.cell(r, l_col).value  = data.get("длина")
            if ws.cell(r, wd_col).value is None: ws.cell(r, wd_col).value = data.get("ширина")
            if ws.cell(r, h_col).value  is None: ws.cell(r, h_col).value  = data.get("высота")
            filled += 1

        processed  += filled
        save_timer += 1
        print(f" → заполнено {filled}/{len(batch_rows)}")

        if save_timer >= SAVE_EVERY:
            wb.save(FILE_PATH)
            print(f"    Сохранено (итого: {processed})")
            save_timer = 0

        time.sleep(2)

    wb.save(FILE_PATH)
    print(f"\nИтого: {processed}/{total} строк заполнено.")
    wb.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="Заполнение габаритов через GigaChat")
    ap.add_argument("--rows",  default=None, help="Диапазон строк: напр. 2-50")
    ap.add_argument("--login", action="store_true", help="Войти в GigaChat (браузер)")
    args = ap.parse_args()

    PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    row_range = None
    if args.rows:
        parts = args.rows.split("-")
        row_range = (int(parts[0]), int(parts[1]) if len(parts) > 1 else int(parts[0]))

    print(f"Файл: {FILE_PATH.name}")
    print(f"Профиль: {PROFILE_DIR}")

    with sync_playwright() as pw:
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR),
                channel="chrome",
                headless=False,
                viewport={"width": 1280, "height": 900},
                locale="ru-RU",
                args=["--disable-blink-features=AutomationControlled"],
            )
        except Exception:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR),
                headless=False,
                viewport={"width": 1280, "height": 900},
                locale="ru-RU",
                args=["--disable-blink-features=AutomationControlled"],
            )
        page = context.new_page()

        if args.login:
            print(f"\nОткрываю браузер GigaChat...")
            try:
                page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=15000)
            except Exception:
                pass
            print(f"Войди в аккаунт на {GIGACHAT_URL}")
            print("Когда окажешься в чате — нажми Enter здесь.")
            input("> ")
            context.storage_state(path=str(PROFILE_DIR / "state.json"))
            print("Сессия сохранена.")
            context.close()
            return

        print("\nОткрываю GigaChat...")
        try:
            page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=30000)
        except Exception as e:
            print(f"Не удалось открыть GigaChat: {e}")
            context.close()
            sys.exit(1)

        time.sleep(3)
        if any(k in page.url.lower() for k in ("login", "signin", "auth", "sso", "id.sber")):
            print("\nНе авторизован! Запусти:")
            print("  uv run --with playwright,openpyxl fill_dimensions_1406.py --login")
            context.close()
            sys.exit(1)

        print("Авторизован.\n")

        try:
            process(page, row_range)
        except KeyboardInterrupt:
            print("\nПрерывание — прогресс сохранён.")
        except Exception as e:
            print(f"Ошибка: {e}")
            traceback.print_exc()

        context.close()


if __name__ == "__main__":
    main()
