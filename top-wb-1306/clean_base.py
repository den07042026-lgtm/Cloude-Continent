import openpyxl

folder = "C:/Users/Admin/Desktop/Топ ВБ 1306"
base_path = f"{folder}/Топ-500 ВБ 1306.xlsx"
helper_path = f"{folder}/Топ-500 ВБ_new1306.xlsx"

# Шаг 1: читаем вспомогательный файл, за один проход собираем два списка
print("Читаю вспомогательный файл...")
wb_helper = openpyxl.load_workbook(helper_path, read_only=True)
ws_helper = wb_helper.active

rows_to_clear = []       # артикул пустой → очистить всю строку в основном
rows_kod_only = {}       # артикул есть, наименование пустое → очистить строку + вставить только артикул

for row_idx, row in enumerate(ws_helper.iter_rows(min_row=2), start=2):
    kod  = row[0].value
    name = row[1].value
    kod_empty  = kod  is None or str(kod).strip()  == ""
    name_empty = name is None or str(name).strip() == ""

    if kod_empty:
        rows_to_clear.append(row_idx)
    elif name_empty:
        rows_kod_only[row_idx] = kod

wb_helper.close()

print(f"Строк с пустым артикулом (полная очистка): {len(rows_to_clear)}")
if rows_to_clear:
    print(f"  Позиции: {rows_to_clear}")

print(f"Строк 'только артикул' (без наименования): {len(rows_kod_only)}")
if rows_kod_only:
    print(f"  Позиции: {list(rows_kod_only.keys())}")

if not rows_to_clear and not rows_kod_only:
    print("Нечего делать — выход.")
    exit()

# Шаг 2: применяем оба правила к основному файлу
print("\nОткрываю основной файл...")
wb_base = openpyxl.load_workbook(base_path)
ws_base = wb_base.active

print("Обрабатываю строки с пустым артикулом...")
for row_idx in rows_to_clear:
    for cell in ws_base[row_idx]:
        cell.value = None

print("Обрабатываю строки 'только артикул'...")
for row_idx, kod in rows_kod_only.items():
    for cell in ws_base[row_idx]:
        cell.value = None
    ws_base.cell(row=row_idx, column=1).value = kod

print(f"\nВсего строк в файле (не изменилось): {ws_base.max_row}")

wb_base.save(base_path)
wb_base.close()
print(f"Готово! Файл сохранён: {base_path}")
