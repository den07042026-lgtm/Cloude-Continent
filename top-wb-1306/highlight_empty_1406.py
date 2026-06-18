import sys, openpyxl
from openpyxl.styles import PatternFill
sys.stdout.reconfigure(encoding="utf-8")

path = r"C:\Users\Admin\Desktop\Топ ВБ 1306\Топ-500 ВБ 1406.xlsx"
YELLOW = PatternFill("solid", fgColor="FFFF00")
CLEAR  = PatternFill(fill_type=None)

wb = openpyxl.load_workbook(path)
ws = wb.active

max_col = ws.max_column
max_row = ws.max_row

# Шаг 1: снимаем все цветовые отметки
cleared = 0
for row in ws.iter_rows(min_row=1, max_row=max_row):
    for cell in row:
        if cell.fill and cell.fill.fill_type not in (None, "none"):
            cell.fill = CLEAR
            cleared += 1
print(f"Снято цветовых отметок: {cleared}")

# Шаг 2: подсвечиваем строки с незаполненными ячейками
highlighted_rows = 0
highlighted_cells = 0

for r in range(2, max_row + 1):
    row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]

    filled = sum(1 for v in row_vals if v is not None and str(v).strip() != "")
    empty  = sum(1 for v in row_vals if v is None or str(v).strip() == "")

    # Полностью пустая строка — пропускаем
    if filled == 0:
        continue

    # Есть хотя бы одна пустая — подсвечиваем
    if empty > 0:
        ws.cell(r, 1).fill = YELLOW   # Код/Артикул
        highlighted_cells += 1
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is None or str(v).strip() == "":
                ws.cell(r, c).fill = YELLOW
                highlighted_cells += 1
        highlighted_rows += 1

print(f"Подсвечено строк: {highlighted_rows}, ячеек: {highlighted_cells}")

wb.save(path)
wb.close()
print("Готово!")
