"""
gen_descriptions.py
Генерирует описания для строк Mikado (col10 пустая) через Gemini API,
в том же стиле, что у соседних строк с описаниями.
"""
import sys
import re
import time
import json
import requests
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

BASE_FILE = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1306\Топ-500 ВБ 1306.xlsx")
ENV_FILE  = Path(r"C:\Users\Admin\Documents\Autoparts_Ecommerce\.env")

# ── Читаем API-ключ из .env ───────────────────────────────────────────────────
def load_env(path: Path) -> dict:
    env = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env

env = load_env(ENV_FILE)
GEMINI_KEY = env.get("GEMINI_API_KEY", "")
if not GEMINI_KEY:
    print("ОШИБКА: GEMINI_API_KEY не найден в .env")
    sys.exit(1)

GEMINI_URL = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "gemini-2.5-flash:generateContent"
    f"?key={GEMINI_KEY}"
)

# ── Читаем файл ───────────────────────────────────────────────────────────────
print("Открываю базовый файл...")
wb = openpyxl.load_workbook(BASE_FILE)
ws = wb.active

# Сбор примеров описаний (до 5 строк с непустым col10)
examples = []
target_rows = []  # [(row_idx, dict с данными)]

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    name     = row[1] if len(row) > 1 else None
    brand    = row[2] if len(row) > 2 else None
    supplier = row[4] if len(row) > 4 else None
    params   = row[5] if len(row) > 5 else None
    oem      = row[6] if len(row) > 6 else None
    compat   = row[7] if len(row) > 7 else None
    alts     = row[8] if len(row) > 8 else None
    descr    = row[9] if len(row) > 9 else None

    # Собираем примеры из строк с описанием
    if len(examples) < 5 and descr and str(descr).strip():
        examples.append({
            "name": str(name or ""),
            "brand": str(brand or ""),
            "params": str(params or ""),
            "oem": str(oem or ""),
            "compat": str(compat or ""),
            "alts": str(alts or ""),
            "descr": str(descr).strip(),
        })

    # Находим целевые строки: Mikado, есть наименование, пустое описание
    is_mikado = supplier and str(supplier).strip().lower() == "mikado"
    has_name = name and str(name).strip()
    no_descr = not descr or not str(descr).strip()
    if is_mikado and has_name and no_descr:
        target_rows.append((i, {
            "name": str(name).strip(),
            "brand": str(brand or "Mikado").strip(),
            "params": str(params or "").strip(),
            "oem": str(oem or "").strip(),
            "compat": str(compat or "").strip(),
            "alts": str(alts or "").strip(),
        }))

print(f"Примеров описаний найдено: {len(examples)}")
print(f"Строк Mikado без описания: {len(target_rows)}")
for ridx, d in target_rows:
    print(f"  Строка {ridx}: {d['name'][:60]}")

if not target_rows:
    print("Нечего делать — выход.")
    sys.exit(0)

# ── Формируем prompt ──────────────────────────────────────────────────────────
def build_prompt(examples: list, product: dict) -> str:
    ex_block = ""
    for ex in examples:
        ex_block += f"""
---
ТОВАР: {ex['name']}
БРЕНД: {ex['brand']}
ПАРАМЕТРЫ: {ex['params'][:200]}
OEM: {ex['oem'][:200]}
ПРИМЕНЯЕМОСТЬ: {ex['compat'][:300]}
ОПИСАНИЕ: {ex['descr']}
"""

    prompt = f"""Ты маркетолог интернет-магазина автозапчастей. Напиши описание товара для маркетплейса WildBerries.

СТИЛЬ И СТРУКТУРА ОПИСАНИЯ:
Ниже приведены примеры описаний из нашего каталога — строго следуй их стилю:
{ex_block}

ПРАВИЛА:
1. Объём: 3–5 абзацев, 400–700 символов.
2. Первый абзац — что это за деталь, для каких автомобилей (из ПРИМЕНЯЕМОСТЬ).
3. Второй абзац — технические характеристики из ПАРАМЕТРЫ (тип фильтра, размеры).
4. Третий абзац — совместимость по OEM, взаимозаменяемость.
5. Без маркдауна, без списков через * или -, только сплошной текст с абзацами.
6. Не упоминай WB или магазин, не придумывай несуществующих характеристик.
7. Пиши профессионально и конкретно.

ТЕПЕРЬ НАПИШИ ОПИСАНИЕ ДЛЯ ЭТОГО ТОВАРА:
НАИМЕНОВАНИЕ: {product['name']}
БРЕНД: {product['brand']}
ПАРАМЕТРЫ: {product['params']}
OEM-НОМЕРА: {product['oem']}
ПРИМЕНЯЕМОСТЬ: {product['compat'][:500]}
АЛЬТЕРНАТИВНЫЕ АРТИКУЛЫ: {product['alts'][:200]}

Выведи только готовое описание, без заголовков и пояснений."""
    return prompt


# ── Вызов Gemini API ──────────────────────────────────────────────────────────
def call_gemini(prompt: str) -> str:
    payload = {
        "contents": [
            {
                "parts": [{"text": prompt}]
            }
        ],
        "generationConfig": {
            "temperature": 0.7,
            "maxOutputTokens": 1024,
        }
    }
    resp = requests.post(GEMINI_URL, json=payload, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    try:
        return data["candidates"][0]["content"]["parts"][0]["text"].strip()
    except (KeyError, IndexError) as e:
        raise RuntimeError(f"Неожиданный ответ Gemini: {data}") from e


# ── Основной цикл ─────────────────────────────────────────────────────────────
ok = 0
err = 0
for idx, (row_idx, product) in enumerate(target_rows, 1):
    print(f"\n[{idx}/{len(target_rows)}] Строка {row_idx} — {product['name'][:55]}")
    try:
        prompt = build_prompt(examples, product)
        descr = call_gemini(prompt)
        ws.cell(row_idx, 10).value = descr
        print(f"  OK  ({len(descr)} симв.)")
        ok += 1
    except Exception as e:
        print(f"  ОШИБКА: {e}")
        err += 1

    if idx < len(target_rows):
        time.sleep(1.0)

# ── Сохраняем ─────────────────────────────────────────────────────────────────
print(f"\nСохраняю файл...")
wb.save(BASE_FILE)
wb.close()
print(f"Готово! Заполнено: {ok}, ошибок: {err}")
print(f"Файл: {BASE_FILE}")
