# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "selenium",
#   "webdriver-manager",
#   "openpyxl",
#   "requests",
# ]
# ///
"""
Скачивание изображений с stparts.ru для строк Топ-500 ВБ 1406 где нет фото в папке.
Запуск: uv run stparts_images_1406.py
"""

import re
import sys
import time
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

sys.stdout.reconfigure(encoding="utf-8")

# ──────────────────────────────────────────────────────────
EXCEL_PATH  = r"C:\Users\Admin\Desktop\Топ ВБ 1406\Топ-500 ВБ 1406.xlsx"
IMAGES_DIR  = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1406\СТ")
SITE_URL    = "https://stparts.ru"
LOGIN_EMAIL = os.environ.get("STPARTS_EMAIL", "")
LOGIN_PASS  = os.environ.get("STPARTS_PASS", "")
TIMEOUT     = 20
PAGE_WAIT   = 3.0
IMG_EXTS    = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".jfif"}
# ──────────────────────────────────────────────────────────

RED     = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)

NOT_FOUND = "not_found"
NO_IMAGE  = "no_image"


def brands_match(a: str, b: str) -> bool:
    a, b = a.strip().upper(), b.strip().upper()
    return bool(a) and bool(b) and (a in b or b in a)


def safe_filename(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', s)


def find_col(ws, keyword: str) -> int:
    for cell in ws[1]:
        if cell.value and keyword.lower() in str(cell.value).lower():
            return cell.column
    raise ValueError(f"Столбец с '{keyword}' не найден. Заголовки: {[c.value for c in ws[1]]}")


def safe_save(wb, path: str) -> None:
    p   = Path(path)
    tmp = p.with_suffix(".tmp.xlsx")
    for attempt in range(1, 11):
        try:
            wb.save(tmp); tmp.replace(p); return
        except PermissionError:
            if tmp.exists():
                try: tmp.unlink()
                except Exception: pass
            if attempt == 1:
                print("\n  [!] Файл занят — закройте Excel и нажмите Enter...", end="", flush=True)
                input()
            else:
                print(f"  [!] Ещё занят, жду 5 сек (попытка {attempt}/10)...", flush=True)
                time.sleep(5)
        except Exception as e:
            if tmp.exists():
                try: tmp.unlink()
                except Exception: pass
            raise e
    raise PermissionError(f"Не удалось сохранить {path}")


def img_size(driver, el) -> tuple[int, int]:
    try:
        w = driver.execute_script("return arguments[0].naturalWidth;", el) or 0
        h = driver.execute_script("return arguments[0].naturalHeight;", el) or 0
        return w, h
    except Exception:
        return 0, 0


# ══════════════════════════════════════════════════════════
# Браузер
# ══════════════════════════════════════════════════════════

def make_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=opts)
    driver.set_window_size(1400, 900)
    return driver


# ══════════════════════════════════════════════════════════
# Авторизация (без изменений)
# ══════════════════════════════════════════════════════════

def _is_logged_in(driver: webdriver.Chrome) -> bool:
    try:
        src = driver.page_source
        markers = ["Выход", "Выйти", "выход", "выйти",
                   "подключены как", "Личный кабинет", "личный кабинет",
                   LOGIN_EMAIL, "logout"]
        return any(m in src for m in markers)
    except Exception:
        return False


def do_login(driver: webdriver.Chrome):
    print("Авторизация на stparts.ru...")
    driver.get(SITE_URL)
    time.sleep(2)

    if _is_logged_in(driver):
        print("  Уже авторизован.")
        return

    for xpath in [
        "//*[normalize-space(text())='Войти']",
        "//*[normalize-space(text())='Вход']",
        "//a[contains(@href,'login')]",
        "//a[contains(@href,'auth')]",
    ]:
        try:
            el = driver.find_element(By.XPATH, xpath)
            if el.is_displayed():
                driver.execute_script("arguments[0].click();", el)
                time.sleep(1.5)
                break
        except Exception:
            pass

    for xpath in [
        "//input[@type='email']", "//input[@name='email']",
        "//input[@name='login']", "//input[@name='username']",
        "//input[@id='email']", "//input[@id='login']",
        "//input[contains(@placeholder,'mail') or contains(@placeholder,'огин')]",
    ]:
        try:
            f = WebDriverWait(driver, 8).until(EC.visibility_of_element_located((By.XPATH, xpath)))
            f.clear(); f.send_keys(LOGIN_EMAIL); break
        except Exception:
            pass

    try:
        pw = WebDriverWait(driver, 8).until(
            EC.visibility_of_element_located((By.XPATH, "//input[@type='password']"))
        )
        pw.clear(); pw.send_keys(LOGIN_PASS)
    except Exception:
        driver.save_screenshot(r"C:\Users\Admin\Desktop\debug_stparts_login.png")
        raise RuntimeError("Поле пароля не найдено — см. debug_stparts_login.png")

    for xpath in [
        "//button[contains(normalize-space(.),'Вход')]",
        "//button[contains(normalize-space(.),'Войти')]",
        "//button[@type='submit']", "//input[@type='submit']",
    ]:
        try:
            btn = driver.find_element(By.XPATH, xpath)
            if btn.is_displayed():
                driver.execute_script("arguments[0].click();", btn); break
        except Exception:
            pass

    time.sleep(3)
    if _is_logged_in(driver):
        print("  Авторизован.")
    else:
        driver.save_screenshot(r"C:\Users\Admin\Desktop\debug_stparts_login.png")
        print("  [!] Авторизация не удалась — debug_stparts_login.png")


def ensure_logged_in(driver: webdriver.Chrome):
    if not _is_logged_in(driver):
        print("  Сессия истекла — повторная авторизация...")
        do_login(driver)


# ══════════════════════════════════════════════════════════
# Поиск, разбор, скачивание — без изменений
# ══════════════════════════════════════════════════════════

def do_search(driver: webdriver.Chrome, article: str):
    search_el = None
    for xpath in [
        "//input[@name='search']", "//input[@type='search']",
        "//input[contains(@class,'search')]",
        "//input[contains(@placeholder,'рти') or contains(@placeholder,'оиск')]",
        "//form//input[@type='text']",
    ]:
        try:
            for el in driver.find_elements(By.XPATH, xpath):
                if el.is_displayed(): search_el = el; break
            if search_el: break
        except Exception:
            pass

    if search_el is None:
        driver.save_screenshot(r"C:\Users\Admin\Desktop\debug_stparts_search.png")
        raise RuntimeError("Поле поиска не найдено")

    search_el.clear(); time.sleep(0.3)
    search_el.send_keys(article); time.sleep(0.4)

    submitted = False
    for btn_xpath in ["./following-sibling::button[1]", "./following-sibling::*[@type='submit'][1]",
                      "./parent::*/button", "./parent::*//*[@type='submit']"]:
        try:
            btn = search_el.find_element(By.XPATH, btn_xpath)
            if btn.is_displayed():
                driver.execute_script("arguments[0].click();", btn); submitted = True; break
        except Exception:
            pass

    if not submitted:
        search_el.send_keys(Keys.ENTER)
    time.sleep(PAGE_WAIT)


def _has_product_card(driver: webdriver.Chrome) -> bool:
    skip_words = ("logo", "icon", "banner", "btn", "arrow", "cart", "sprite")
    try:
        for img in driver.find_elements(By.TAG_NAME, "img"):
            src = img.get_attribute("src") or ""
            if not src.startswith("http"): continue
            if any(x in src.lower() for x in skip_words): continue
            if not img.is_displayed(): continue
            w, h = img_size(driver, img)
            if w > 150 and h > 150: return True
    except Exception:
        pass
    return False


def _get_product_card_brand(driver: webdriver.Chrome) -> str | None:
    for xpath in ["//h1", "//h2", "//h3",
                  "//*[contains(@class,'title')]", "//*[contains(@class,'brand')]"]:
        try:
            for el in driver.find_elements(By.XPATH, xpath):
                if not el.is_displayed(): continue
                text = el.text.strip()
                if text and 3 <= len(text) <= 100:
                    return text.split()[0]
        except Exception:
            pass
    return None


def find_brand_in_table(driver: webdriver.Chrome, excel_brand: str):
    try:
        rows = driver.find_elements(By.XPATH, "//table//tr[.//td]")
    except Exception:
        return None
    for row in rows:
        cells = row.find_elements(By.TAG_NAME, "td")
        if len(cells) < 2: continue
        for cell in cells[:3]:
            links = cell.find_elements(By.TAG_NAME, "a")
            txt = (links[0] if links else cell).text.strip().split("\n")[0]
            if txt and 2 <= len(txt) <= 50 and brands_match(excel_brand, txt):
                return cell
    return None


def _get_gallery_urls(driver: webdriver.Chrome) -> list[str]:
    """Ищем прямые ссылки на большие изображения — только в основной карточке товара (не в сайдбаре)."""
    try:
        urls = driver.execute_script("""
            var found = []; var seen = {};
            var IMG_EXT = /\.(jpg|jpeg|png|webp|bmp)(\?|$)/i;
            var pageW = document.documentElement.scrollWidth || 1400;
            // Ищем только в основной карточке (левая половина страницы)
            var containers = [
                document.querySelector('.product-images, .product-photo, .item-photo, .product-card, .catalog-item__image, .good-card__image'),
                document.body
            ];
            var selectors = [
                'a[data-fancybox]','a[data-gallery]','a[data-lightbox]',
                'a.fancybox','a[rel="gallery"]','a[rel="fancybox"]',
                'a[href$=".jpg"]','a[href$=".jpeg"]','a[href$=".png"]','a[href$=".webp"]'
            ];
            for (var c = 0; c < containers.length; c++) {
                var root = containers[c]; if (!root) continue;
                for (var s = 0; s < selectors.length; s++) {
                    var links = root.querySelectorAll(selectors[s]); if (!links.length) continue;
                    for (var i = 0; i < links.length; i++) {
                        var a = links[i];
                        // Пропускаем ссылки из правой части страницы (сайдбар/рекомендации)
                        var rect = a.getBoundingClientRect();
                        if (rect.left > pageW * 0.6) continue;
                        var img = a.querySelector('img');
                        if (!img) continue;
                        var src = img.src || '';
                        if (!src || src.indexOf('logo') >= 0 || src.indexOf('icon') >= 0) continue;
                        var u = a.href || a.getAttribute('href') || src;
                        if (u && !seen[u] && u.indexOf('http') === 0 && (IMG_EXT.test(u) || a.getAttribute('data-fancybox') || a.getAttribute('data-gallery'))) {
                            seen[u] = true; found.push(u);
                        }
                    }
                    if (found.length) return found;
                }
            }
            return found;""")
        return [u for u in (urls or []) if u.startswith("http")]
    except Exception:
        return []


def _get_main_product_image(driver: webdriver.Chrome):
    """Находим главное фото товара — только в левой части страницы, не в сайдбаре."""
    skip = ("logo", "icon", "banner", "btn", "arrow", "cart", "sprite", "adv", "schatz")
    try:
        page_width = driver.execute_script("return document.documentElement.scrollWidth;") or 1400
    except Exception:
        page_width = 1400

    # Сначала ищем img внутри a[data-fancybox] и похожих — только слева
    for xpath in ["//a[@data-fancybox]//img", "//a[@data-gallery]//img",
                  "//a[@data-lightbox]//img", "//a[contains(@class,'fancybox')]//img",
                  "//a[contains(@rel,'gallery') or contains(@rel,'fancybox')]//img",
                  "//a[contains(@href,'.jpg') or contains(@href,'.jpeg') or contains(@href,'.png')]//img"]:
        try:
            for img in driver.find_elements(By.XPATH, xpath):
                if not img.is_displayed(): continue
                src = img.get_attribute("src") or ""
                if not src.startswith("http"): continue
                if any(x in src.lower() for x in skip): continue
                x = img.location.get("x", page_width)
                if x > page_width * 0.6: continue  # пропускаем сайдбар
                w, h = img_size(driver, img)
                if w > 50 and h > 50: return img
        except Exception:
            pass

    # Фолбэк: наибольший img в левой половине страницы
    best_score, best_el = 0, None
    try:
        for img in driver.find_elements(By.TAG_NAME, "img"):
            src = img.get_attribute("src") or ""
            if not src.startswith("http"): continue
            if any(x in src.lower() for x in skip): continue
            if not img.is_displayed(): continue
            x = img.location.get("x", page_width)
            if x > page_width * 0.6: continue  # только левая часть
            w, h = img_size(driver, img)
            if w > 80 and h > 80:
                score = w * h
                if score > best_score:
                    best_score = score; best_el = img
    except Exception:
        pass
    return best_el


def _get_lightbox_image_url(driver: webdriver.Chrome) -> str | None:
    for xpath in ["//*[contains(@class,'fancybox')]//img[@src]",
                  "//*[contains(@class,'lightbox')]//img[@src]",
                  "//*[contains(@id,'lightbox')]//img[@src]",
                  "//*[contains(@class,'modal') and not(contains(@class,'fade'))]//img[@src]",
                  "//*[contains(@class,'popup')]//img[@src]"]:
        try:
            best_area, best_url = 0, ""
            for img in driver.find_elements(By.XPATH, xpath):
                src = img.get_attribute("src") or ""
                if not src.startswith("http"): continue
                w, h = img_size(driver, img)
                if w > 100 and w * h > best_area: best_area, best_url = w * h, src
            if best_url: return best_url
        except Exception:
            pass
    try:
        best_area, best_url = 0, ""
        for img in driver.find_elements(By.TAG_NAME, "img"):
            src = img.get_attribute("src") or ""
            if not src.startswith("http"): continue
            w, h = img_size(driver, img)
            if w > 300 and w * h > best_area: best_area, best_url = w * h, src
        if best_url: return best_url
    except Exception:
        pass
    return None


def _click_lightbox_next(driver: webdriver.Chrome) -> bool:
    for xpath in ["//*[contains(@class,'next') and not(contains(@class,'disabled'))]",
                  "//*[contains(@class,'slick-next') and not(contains(@class,'slick-disabled'))]",
                  "//*[contains(@class,'carousel-control-next')]",
                  "//button[normalize-space(text())='>']"]:
        try:
            for el in driver.find_elements(By.XPATH, xpath):
                if el.is_displayed():
                    driver.execute_script("arguments[0].click();", el); return True
        except Exception:
            pass
    return False


def _close_lightbox(driver: webdriver.Chrome):
    try: driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
    except Exception: pass
    time.sleep(0.5)


def _download(url: str, cookies: dict, save_path: Path) -> bool:
    try:
        resp = requests.get(url, cookies=cookies,
                            headers={"User-Agent": "Mozilla/5.0", "Referer": SITE_URL}, timeout=30)
        resp.raise_for_status()
        if resp.content: save_path.write_bytes(resp.content); return True
    except Exception:
        pass
    return False


def download_all_images(driver: webdriver.Chrome, article: str) -> list[str]:
    base    = safe_filename(article)
    cookies = {c["name"]: c["value"] for c in driver.get_cookies()}
    saved:  list[str] = []

    gallery_urls = _get_gallery_urls(driver)
    if gallery_urls:
        print(f"  Галерея из DOM: {len(gallery_urls)} URL")
        for idx, url in enumerate(gallery_urls, 1):
            filename  = f"{base}_{idx}.jpg"
            save_path = IMAGES_DIR / filename
            if _download(url, cookies, save_path):
                saved.append(filename)
                print(f"    [{idx}] {filename}")
            else:
                print(f"    [{idx}] ошибка: {url[:70]}")
        return saved

    main_img = _get_main_product_image(driver)
    if not main_img: return []

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", main_img)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", main_img)
        time.sleep(2.5)
    except Exception as e:
        print(f"  Ошибка клика по изображению: {e}"); return []

    seen: set[str] = set()
    for idx in range(1, 21):
        time.sleep(1.0)
        url = _get_lightbox_image_url(driver)
        if not url or url in seen: break
        seen.add(url)
        filename  = f"{base}_{idx}.jpg"
        save_path = IMAGES_DIR / filename
        if _download(url, cookies, save_path):
            saved.append(filename)
            print(f"    [{idx}] {filename}")
        else:
            print(f"    [{idx}] ошибка: {url[:70]}")
        if not _click_lightbox_next(driver): break
        time.sleep(1.5)
        new_url = _get_lightbox_image_url(driver)
        if not new_url or new_url == url or new_url in seen: break

    _close_lightbox(driver)
    return saved


def process_article(driver: webdriver.Chrome, article: str, excel_brand: str) -> list[str] | str:
    try: do_search(driver, article)
    except RuntimeError as e: print(f"  Ошибка поиска: {e}"); return NOT_FOUND

    page = driver.page_source or ""
    if len(page) < 500: return NOT_FOUND
    nf_phrases = ["ничего не найдено", "не найдено", "нет результатов", "0 результатов", "not found", "no results"]
    if any(p in page.lower() for p in nf_phrases): return NOT_FOUND

    if _has_product_card(driver):
        card_brand = _get_product_card_brand(driver)
        print(f"  [Карточка] Бренд сайта: '{card_brand}' | Excel: '{excel_brand}'")
        if card_brand and brands_match(excel_brand, card_brand):
            files = download_all_images(driver, article)
            return files if files else NO_IMAGE
        brand_el = find_brand_in_table(driver, excel_brand)
        if not brand_el:
            print("  Бренд не найден ни в карточке, ни в таблице")
            return NOT_FOUND
        print("  Бренд найден в таблице, кликаем...")
        driver.execute_script("arguments[0].click();", brand_el)
        time.sleep(PAGE_WAIT)
    else:
        brand_el = find_brand_in_table(driver, excel_brand)
        print(f"  [Таблица] Бренд найден: {brand_el is not None} | Excel: '{excel_brand}'")
        if not brand_el: return NOT_FOUND
        driver.execute_script("arguments[0].click();", brand_el)
        time.sleep(PAGE_WAIT)

    files = download_all_images(driver, article)
    return files if files else NO_IMAGE


# ══════════════════════════════════════════════════════════
# Главная функция
# ══════════════════════════════════════════════════════════

def main():
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    # Проверяем ОБЕ папки: уже скачанные ранее (Изображения) + новые (СТ)
    existing = set()
    for check_dir in [IMAGES_DIR, Path(r"C:\Users\Admin\Desktop\Топ ВБ 1406\Изображения")]:
        if check_dir.exists():
            for f in check_dir.iterdir():
                if f.is_file() and f.suffix.lower() in IMG_EXTS:
                    existing.add(f.name.lower())
    print(f"Уже есть (в обеих папках): {len(existing)} файлов")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    col_article = find_col(ws, "Код")
    col_brand   = find_col(ws, "Бренд")
    col_images  = find_col(ws, "Изображени")

    # Строки где хотя бы один файл из col_images отсутствует в папке
    rows_todo = []
    for r in range(2, ws.max_row + 1):
        art = ws.cell(r, col_article).value
        v   = ws.cell(r, col_images).value
        if not art or not v: continue
        files = [f.strip().lower() for f in str(v).replace(";", ",").split(",") if f.strip()]
        if any(f not in existing for f in files):
            rows_todo.append(r)

    total = len(rows_todo)
    print(f"Строк для обработки: {total}\n")
    if not total: print("Все изображения уже есть!"); return

    driver = make_driver()
    stats  = {"ok": 0, "red": 0}

    try:
        do_login(driver)

        for i, row_num in enumerate(rows_todo, 1):
            article  = str(ws.cell(row_num, col_article).value).strip()
            brand    = str(ws.cell(row_num, col_brand).value or "").strip()
            art_cell = ws.cell(row_num, col_article)

            ensure_logged_in(driver)
            print(f"\n[{i}/{total}] {article}  [{brand}]")
            result = process_article(driver, article, brand)

            if isinstance(result, list) and result:
                art_cell.fill = NO_FILL
                ws.cell(row_num, col_images).fill = NO_FILL
                ws.cell(row_num, col_images).value = "; ".join(result)
                print(f"  ✓ {len(result)} фото: {', '.join(result)}")
                stats["ok"] += 1
            else:
                art_cell.fill = RED
                reason = result if isinstance(result, str) else NO_IMAGE
                print(f"  ✗ {reason}")
                stats["red"] += 1

            safe_save(wb, EXCEL_PATH)

    except KeyboardInterrupt:
        print("\nОстановлено пользователем")
    finally:
        driver.quit()
        safe_save(wb, EXCEL_PATH)

    print(f"\n{'='*50}")
    print(f"Готово!")
    print(f"  ✓ Скачано : {stats['ok']}")
    print(f"  ✗ Не найдено/нет фото: {stats['red']}")
    print(f"  Папка     : {IMAGES_DIR}")


if __name__ == "__main__":
    main()
