"""
mikado_fill.py
Находит строки в базовом файле где заполнен только "Код / Артикул",
парсит каждый артикул на mikado-parts.ru и заполняет строку данными.
"""

import sys
import re
import time
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(r"C:\Users\Admin\Documents\Autoparts_Ecommerce\scripts")))

from mikado_scraper import (
    login, fetch_product, fetch_oem, fetch_compatibility,
    compat_to_text, fetch_cross_refs, fetch_images, strip_tags, load_env
)
import openpyxl

# ── Пути ──────────────────────────────────────────────────────────────────────
FOLDER      = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1306")
BASE_FILE   = FOLDER / "Топ-500 ВБ 1306.xlsx"
PRICE_FILE  = Path(r"C:\Users\Admin\Documents\Ecommerce\mikado_price_34.xlsx")
IMG_DIR     = FOLDER / "images"
IMG_DIR.mkdir(exist_ok=True)

ENV_PATH  = Path(r"C:\Users\Admin\Documents\Autoparts_Ecommerce\.env")
DELAY     = 1.5   # секунд между артикулами

# Соответствие ключей params → номера столбцов (1-based)
# Вес в граммах, размеры в мм
DIM_KEYS = {
    11: ["вес", "масса", "weight"],
    12: ["длина", "length", "глубина"],
    13: ["ширина", "width"],
    14: ["высота", "height"],
}


def extract_number(value: str) -> float | None:
    """Вытаскивает первое число из строки вида '294 мм' или '0.7 кг'."""
    m = re.search(r"[\d]+(?:[.,]\d+)?", str(value))
    if not m:
        return None
    num = float(m.group(0).replace(",", "."))
    # Если единица — кг, переводим в граммы
    if re.search(r"кг|kg", str(value), re.I):
        num *= 1000
    return num


def find_dim_value(params: dict, col: int) -> float | None:
    """Ищет в словаре params значение для нужной размерной колонки."""
    for key, val in params.items():
        key_low = key.lower()
        if any(alias in key_low for alias in DIM_KEYS[col]):
            return extract_number(val)
    return None


def format_params(params: dict) -> str:
    return "\n".join(f"{k}: {v}" for k, v in params.items())


def format_oem(oem_details: list[dict]) -> str:
    return "; ".join(f"{d['manufacturer']}: {d['oem_code']}" for d in oem_details[:10])


# ── Загружаем прайс Mikado: Code → Prodnum ────────────────────────────────────
print("Загружаю прайс Mikado для lookup Prodnum...")
code_to_prodnum = {}
if PRICE_FILE.exists():
    wb_price = openpyxl.load_workbook(PRICE_FILE, read_only=True, data_only=True)
    ws_price = wb_price.active
    first_row = next(ws_price.iter_rows(min_row=1, max_row=1, values_only=True))
    headers_p = [str(v).strip() if v else f"col{i}" for i, v in enumerate(first_row, 1)]
    prodnum_col = next((i for i, h in enumerate(headers_p) if h.lower() == "prodnum"), None)
    code_col    = next((i for i, h in enumerate(headers_p) if h.lower() == "code"),    None)
    if prodnum_col is not None and code_col is not None:
        for row in ws_price.iter_rows(min_row=2, values_only=True):
            c = row[code_col]
            p = row[prodnum_col]
            if c and p:
                code_to_prodnum[str(c).strip().lower()] = str(p).strip()
    wb_price.close()
    print(f"  Загружено {len(code_to_prodnum)} пар Code→Prodnum")
else:
    print("  Прайс не найден, будет попытка с префиксом xzk-")


def resolve_prodnum(code: str) -> str:
    """Возвращает Prodnum для запроса к Mikado: из прайса или с префиксом xzk-."""
    key = code.lower()
    if key in code_to_prodnum:
        return code_to_prodnum[key]
    return f"xzk-{code}"


# ── Читаем базовый файл, ищем нужные строки ───────────────────────────────────
print("Читаю базовый файл...")
wb = openpyxl.load_workbook(BASE_FILE)
ws = wb.active

target_rows = []  # [(row_idx, article_code)]
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    kod  = row[0].value
    name = row[1].value  # Наименование
    name_empty = name is None or str(name).strip() == ""
    if kod and name_empty:
        target_rows.append((row_idx, str(kod).strip()))

print(f"Строк только с артикулом: {len(target_rows)}")
for ridx, code in target_rows:
    print(f"  Строка {ridx}: {code}")

if not target_rows:
    print("Нечего заполнять — выход.")
    sys.exit(0)

# ── Авторизация ───────────────────────────────────────────────────────────────
env      = load_env(ENV_PATH)
session  = login(env["MIKADO_CODE"], env["MIKADO_PASSWORD"])

# ── Парсим и заполняем ────────────────────────────────────────────────────────
ok_count  = 0
err_count = 0

for i, (row_idx, code) in enumerate(target_rows, 1):
    print(f"\n[{i}/{len(target_rows)}] Строка {row_idx} — {code}")

    try:
        prodnum = resolve_prodnum(code)
        print(f"  Prodnum: {prodnum}")
        _, data = fetch_product(session, prodnum)
        print(f"  {data['name'][:55]}  |  {data['brand']}  |  {data['price']:.0f} руб.")

        artid = data.get("_artid")

        oem_details  = []
        compat_text  = ""
        cross_refs   = []
        images       = []

        if artid:
            time.sleep(0.4)
            oem_details = fetch_oem(session, artid, code)
            print(f"  OEM: {len(oem_details)}")

            time.sleep(0.4)
            compat_rows = fetch_compatibility(session, artid)
            compat_text = compat_to_text(compat_rows)
            print(f"  Применяемость: {len(compat_rows)} строк")

            time.sleep(0.4)
            images = fetch_images(session, artid, code, IMG_DIR)
            print(f"  Фото: {len(images)}")
        else:
            print("  ARTID не найден — OEM/применяемость/фото пропущены")

        time.sleep(0.4)
        qty        = sum(s["qty"] for s in data.get("stock_items", []))
        cross_refs = fetch_cross_refs(session, prodnum, qty, data.get("name", ""))
        print(f"  Перекодировки: {len(cross_refs)}")

        params = data.get("params", {})

        # Записываем в строку базового файла
        ws.cell(row_idx, 2).value  = data.get("name", "")
        ws.cell(row_idx, 3).value  = data.get("brand", "")
        ws.cell(row_idx, 4).value  = data.get("price") or None
        ws.cell(row_idx, 5).value  = "Mikado"
        ws.cell(row_idx, 6).value  = format_params(params) if params else None
        ws.cell(row_idx, 7).value  = format_oem(oem_details) if oem_details else None
        ws.cell(row_idx, 8).value  = compat_text or None
        ws.cell(row_idx, 9).value  = "; ".join(cross_refs) if cross_refs else None
        # col 10 Описание — нет в Mikado, оставляем пустым
        ws.cell(row_idx, 11).value = find_dim_value(params, 11)
        ws.cell(row_idx, 12).value = find_dim_value(params, 12)
        ws.cell(row_idx, 13).value = find_dim_value(params, 13)
        ws.cell(row_idx, 14).value = find_dim_value(params, 14)
        ws.cell(row_idx, 15).value = ", ".join(images) if images else None

        ok_count += 1

    except Exception as e:
        print(f"  ОШИБКА: {e}")
        err_count += 1

    if i < len(target_rows):
        time.sleep(DELAY)

# ── Сохраняем ─────────────────────────────────────────────────────────────────
print(f"\nСохраняю файл...")
wb.save(BASE_FILE)
wb.close()
print(f"Готово! Заполнено: {ok_count}, ошибок: {err_count}")
print(f"Файл: {BASE_FILE}")
