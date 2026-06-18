"""
fill_oem_1406.py
══════════════════════════════════════════════════════
Заполняет «OEM номера» через GigaChat для строк, где стоит прочерк «—».
Если найти не удаётся — прочерк остаётся.

Файл: C:\\Users\\Admin\\Desktop\\Топ ВБ 1306\\Топ-500 ВБ 1406.xlsx
Сохраняет поверх исходника.

Запуск:
  uv run --with playwright,openpyxl fill_oem_1406.py
"""

import re
import sys
import json
import time
import argparse
import traceback
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("Playwright не установлен: uv run --with playwright python -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    sys.exit(1)


FILE_PATH    = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1306\Топ-500 ВБ 1406.xlsx")
PROFILE_DIR  = Path.home() / ".gigachat_dimensions_playwright"
GIGACHAT_URL = "https://giga.chat/"

BATCH_SIZE    = 8
SAVE_EVERY    = 3
PROMPT_MARKER = "Ты эксперт по каталогам автозапчастей"

DASH = "—"


# ══════════════════════════════════════════════════════════════════════════════
#  ПРОМПТ
# ══════════════════════════════════════════════════════════════════════════════

def build_prompt(items: list[dict]) -> str:
    instruction = (
        f"{PROMPT_MARKER}. "
        "Для каждой позиции найди оригинальные OEM-номера (номера деталей производителя автомобиля).\n"
        "Правила:\n"
        "- Возвращай только реальные OEM-номера, которые ты знаешь точно\n"
        "- Если для данной запчасти OEM-номера не известны или позиция — инструмент/расходник без OEM — верни null\n"
        "- Несколько OEM-номеров разделяй точкой с запятой: «A1234567; B9876543»\n"
        "- Не придумывай номера — лучше null, чем неверные данные\n\n"
        "Верни ТОЛЬКО JSON объект без пояснений и без markdown:\n"
        '{"items": [{"idx": 0, "oem": "A1234; B5678"}, {"idx": 1, "oem": null}, ...]}\n\n'
    )

    lines = [instruction]
    for item in items:
        lines.append(f"[{item['idx']}]")
        lines.append(f"Артикул: {item['article']}")
        lines.append(f"Наименование: {item['name']}")
        if item.get("brand"):
            lines.append(f"Бренд: {item['brand']}")
        if item.get("params"):
            lines.append(f"Параметры: {str(item['params'])[:300]}")
        if item.get("compat"):
            lines.append(f"Применяемость: {str(item['compat'])[:200]}")
        if item.get("alt"):
            lines.append(f"Аналоги/альт. артикулы: {str(item['alt'])[:200]}")
        lines.append("")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  ПАРСИНГ ОТВЕТА
# ══════════════════════════════════════════════════════════════════════════════

def extract_items(text: str) -> list[dict]:
    clean = re.sub(r"```(?:json)?\s*", "", text, flags=re.IGNORECASE).replace("```", "").strip()

    def is_valid(lst):
        return bool(lst) and isinstance(lst[0], dict) and "oem" in lst[0] and "idx" in lst[0]

    def try_parse(s):
        try:
            parsed = json.loads(s.strip())
            if isinstance(parsed, dict) and "items" in parsed:
                lst = parsed["items"]
                if isinstance(lst, list) and is_valid(lst):
                    return lst
            if isinstance(parsed, list) and is_valid(parsed):
                return parsed
        except Exception:
            pass
        return None

    result = try_parse(clean)
    if result:
        return result

    for start in [i for i, c in enumerate(clean) if c == '{']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == '{':
                depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result:
                        return result
                    break
    return []


# ══════════════════════════════════════════════════════════════════════════════
#  БРАУЗЕР
# ══════════════════════════════════════════════════════════════════════════════

INPUT_SELECTORS = [
    '#chat-input-textarea',
    'textarea[placeholder*="Спрос"]',
    'textarea[placeholder*="Напиш"]',
    'textarea[placeholder*="Сообщ"]',
    'textarea[placeholder*="Введ"]',
    '[role="textbox"]',
    'div[contenteditable="true"]',
    'textarea',
]

SEND_SELECTORS = [
    'button[aria-label*="Отправить"]',
    'button[aria-label*="отправить"]',
    'button[aria-label*="Send"]',
    'button[type="submit"]',
    '[data-testid="send-button"]',
]


def dismiss_modal(page):
    try:
        page.keyboard.press("Escape")
        time.sleep(0.4)
    except Exception:
        pass


def find_input(page):
    time.sleep(1)
    for sel in INPUT_SELECTORS:
        try:
            el = page.locator(sel).last
            if el.is_visible(timeout=2000):
                return el, sel
        except Exception:
            continue
    return None, None


def type_prompt(page, text: str):
    dismiss_modal(page)
    inp, sel = find_input(page)
    if inp is None:
        raise RuntimeError("Поле ввода GigaChat не найдено")

    is_ce = "contenteditable" in sel or "textbox" in sel
    if is_ce:
        inp.click()
        time.sleep(0.3)
        page.evaluate("""(text) => {
            const el = document.querySelector('[contenteditable="true"]') ||
                       document.querySelector('[role="textbox"]');
            if (!el) return;
            el.focus(); el.innerText = text;
            el.dispatchEvent(new Event('input',  { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        }""", text)
    else:
        page.evaluate("""() => {
            const el = document.querySelector('#chat-input-textarea') ||
                       document.querySelector('textarea');
            if (el) el.focus();
        }""")
        time.sleep(0.3)
        page.evaluate("""(text) => {
            const el = document.querySelector('#chat-input-textarea') ||
                       document.querySelector('textarea');
            if (!el) return;
            const setter = Object.getOwnPropertyDescriptor(
                window.HTMLTextAreaElement.prototype, 'value').set;
            setter.call(el, text);
            el.dispatchEvent(new Event('input',  { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        }""", text)
    time.sleep(0.5)


def click_send(page):
    for sel in SEND_SELECTORS:
        try:
            btn = page.locator(sel).last
            if btn.is_visible(timeout=800):
                btn.click()
                return
        except Exception:
            continue
    try:
        inp, _ = find_input(page)
        if inp:
            inp.press("Enter")
            return
    except Exception:
        pass
    page.keyboard.press("Enter")


def is_generating(page) -> bool:
    try:
        for sel in ['[aria-label*="Стоп"]', '[aria-label*="Stop"]',
                    'button[aria-label*="остановить"]', '[class*="loading"]']:
            if page.locator(sel).first.is_visible(timeout=300):
                return True
    except Exception:
        pass
    return False


def find_json_on_page(page) -> str:
    try:
        return page.evaluate(r"""() => {
            const text = document.body ? (document.body.innerText || '') : '';

            function extractBlock(startStr, openCh, closeCh) {
                let pos = -1, search = 0;
                while (true) {
                    let found = text.indexOf(startStr, search);
                    if (found === -1) break;
                    pos = found; search = found + 1;
                }
                if (pos === -1) return '';
                let depth = 0;
                for (let i = pos; i < Math.min(pos + 200000, text.length); i++) {
                    if (text[i] === openCh) depth++;
                    else if (text[i] === closeCh) {
                        depth--;
                        if (depth === 0) {
                            const block = text.slice(pos, i + 1);
                            if (block.includes('...') || block.length < 30) return '';
                            return block;
                        }
                    }
                }
                return '';
            }

            let result = extractBlock('{"items":', '{', '}');
            if (!result) result = extractBlock('{ "items":', '{', '}');
            if (!result) {
                const pat = /"oem"\s*:/g;
                let m, lastIdx = -1;
                while ((m = pat.exec(text)) !== null) lastIdx = m.index;
                if (lastIdx !== -1) {
                    let start = text.lastIndexOf('[', lastIdx);
                    if (start !== -1) {
                        let depth = 0;
                        for (let i = start; i < Math.min(start + 200000, text.length); i++) {
                            if (text[i] === '[') depth++;
                            else if (text[i] === ']') {
                                depth--;
                                if (depth === 0) {
                                    const block = text.slice(start, i + 1);
                                    if (!block.includes('...') && block.length >= 30)
                                        result = block;
                                    break;
                                }
                            }
                        }
                    }
                }
            }
            return result;
        }""")
    except Exception:
        return ""


def wait_for_response(page, timeout_sec: int = 240) -> str:
    print("    жду", end="", flush=True)
    time.sleep(4)

    prev, stable = "", 0
    for tick in range(timeout_sec):
        json_found = find_json_on_page(page)
        if json_found:
            if json_found == prev:
                stable += 1
                if stable >= 2 and not is_generating(page):
                    print(f" готово ({len(json_found)} симв.)")
                    return json_found
            else:
                stable = 0
                if tick % 4 == 0:
                    print(".", end="", flush=True)
            prev = json_found
        else:
            stable = 0
            if tick % 6 == 0:
                print(".", end="", flush=True)
        time.sleep(1.5)

    try:
        page.screenshot(path=str(FILE_PATH.parent / "_debug_oem_timeout.png"))
    except Exception:
        pass
    print(" таймаут")
    return ""


def new_chat(page):
    dismiss_modal(page)
    for sel in ['a[href="/"][aria-label*="чат"]', 'button[aria-label*="новый чат"]',
                'button[aria-label*="Новый чат"]', '[data-testid="new-chat"]', 'a[href="/"]']:
        try:
            btn = page.locator(sel).first
            if btn.is_visible(timeout=800):
                btn.click()
                time.sleep(2)
                dismiss_modal(page)
                return
        except Exception:
            continue
    try:
        page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=20000)
    except Exception:
        pass
    time.sleep(2)
    dismiss_modal(page)


# ══════════════════════════════════════════════════════════════════════════════
#  ОСНОВНАЯ ОБРАБОТКА
# ══════════════════════════════════════════════════════════════════════════════

def process(page, row_range=None):
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

    art_col    = headers.get("Код / Артикул")
    name_col   = headers.get("Наименование")
    brand_col  = headers.get("Бренд")
    param_col  = headers.get("Параметры")
    oem_col    = headers.get("OEM номера")
    compat_col = headers.get("Применяемость")
    alt_col    = headers.get("Альтернативные артикулы товара")

    if not oem_col:
        print("Столбец «OEM номера» не найден.")
        wb.close()
        return

    todos = []
    for r in range(2, ws.max_row + 1):
        if row_range and not (row_range[0] <= r <= row_range[1]):
            continue
        if not ws.cell(r, art_col or 1).value:
            continue
        v = ws.cell(r, oem_col).value
        if str(v).strip() == DASH:
            todos.append(r)

    total = len(todos)
    if not todos:
        print("Нечего обрабатывать.")
        wb.close()
        return

    print(f"Строк для обработки: {total}")

    processed = 0
    filled_count = 0
    save_timer = 0

    for batch_no, batch_start in enumerate(range(0, total, BATCH_SIZE), start=1):
        batch_rows = todos[batch_start: batch_start + BATCH_SIZE]

        items = []
        for local_idx, r in enumerate(batch_rows):
            items.append({
                "idx":     local_idx,
                "article": ws.cell(r, art_col).value    if art_col    else "",
                "name":    ws.cell(r, name_col).value   if name_col   else "",
                "brand":   ws.cell(r, brand_col).value  if brand_col  else "",
                "params":  ws.cell(r, param_col).value  if param_col  else "",
                "compat":  ws.cell(r, compat_col).value if compat_col else "",
                "alt":     ws.cell(r, alt_col).value    if alt_col    else "",
            })

        print(f"  Батч {batch_no} (стр. {batch_rows[0]}–{batch_rows[-1]})", end="", flush=True)

        new_chat(page)
        prompt = build_prompt(items)

        try:
            type_prompt(page, prompt)
        except RuntimeError as e:
            print(f"\n  ✗ {e}")
            continue

        click_send(page)
        time.sleep(1)

        results = []
        for attempt in range(1, 4):
            raw = wait_for_response(page)
            if not raw:
                print(f"  ✗ Пустой ответ (попытка {attempt}/3)")
            else:
                results = extract_items(raw)
                if results:
                    break
                print(f"  ⚠ JSON не распознан (попытка {attempt}/3)")
                print(f"    Начало: {raw[:150]}")

            if attempt < 3:
                time.sleep(10)
                new_chat(page)
                type_prompt(page, prompt)
                click_send(page)
                time.sleep(1)

        if not results:
            print(f"  ✗ Батч пропущен после 3 попыток")
            continue

        result_map = {r["idx"]: r for r in results}
        batch_filled = 0
        for local_idx, r in enumerate(batch_rows):
            data = result_map.get(local_idx)
            if not data:
                continue
            oem_val = data.get("oem")
            if oem_val and str(oem_val).strip() not in ("", "null", "None", DASH):
                ws.cell(r, oem_col).value = str(oem_val).strip()
                batch_filled += 1

        processed    += len(batch_rows)
        filled_count += batch_filled
        save_timer   += 1
        print(f" → найдено {batch_filled}/{len(batch_rows)}")

        if save_timer >= SAVE_EVERY:
            wb.save(FILE_PATH)
            print(f"    Сохранено (найдено итого: {filled_count}/{processed})")
            save_timer = 0

        time.sleep(2)

    wb.save(FILE_PATH)
    print(f"\nИтого: обработано {processed}/{total}, OEM найдены для {filled_count} строк.")
    wb.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="Поиск OEM-номеров через GigaChat")
    ap.add_argument("--rows",  default=None)
    ap.add_argument("--login", action="store_true")
    args = ap.parse_args()

    PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    row_range = None
    if args.rows:
        parts = args.rows.split("-")
        row_range = (int(parts[0]), int(parts[1]) if len(parts) > 1 else int(parts[0]))

    print(f"Файл: {FILE_PATH.name}")

    with sync_playwright() as pw:
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR), channel="chrome",
                headless=False, viewport={"width": 1280, "height": 900},
                locale="ru-RU", args=["--disable-blink-features=AutomationControlled"],
            )
        except Exception:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR),
                headless=False, viewport={"width": 1280, "height": 900},
                locale="ru-RU", args=["--disable-blink-features=AutomationControlled"],
            )
        page = context.new_page()

        if args.login:
            try:
                page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=15000)
            except Exception:
                pass
            print(f"Войди на {GIGACHAT_URL}, затем нажми Enter.")
            input("> ")
            context.close()
            return

        print("Открываю GigaChat...")
        try:
            page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=30000)
        except Exception as e:
            print(f"Ошибка: {e}")
            context.close()
            sys.exit(1)

        time.sleep(3)
        if any(k in page.url.lower() for k in ("login", "signin", "auth", "sso", "id.sber")):
            print("Не авторизован! Запусти с --login")
            context.close()
            sys.exit(1)

        print("Авторизован.\n")

        try:
            process(page, row_range)
        except KeyboardInterrupt:
            print("\nПрерывание — прогресс сохранён.")
        except Exception as e:
            print(f"Ошибка: {e}")
            traceback.print_exc()

        context.close()


if __name__ == "__main__":
    main()
