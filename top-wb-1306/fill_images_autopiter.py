"""
fill_images_autopiter.py
1. Конвертирует все не-JPG файлы в папке "Изображения Автопитер" → .jpg
2. Соотносит файлы с артикулами из Excel (col 1) по имени файла
3. Заполняет пустые ячейки столбца "Изображения" (col 15)
"""
import sys
import re
import shutil
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from PIL import Image
import openpyxl

FOLDER    = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1306")
IMG_DIR   = FOLDER / "Изображения Автопитер"
BASE_FILE = FOLDER / "Топ-500 ВБ 1306.xlsx"
IMG_COL   = 15   # Изображения (1-based)
KOD_COL   = 1    # Код / Артикул


# ── 1. Конвертируем все не-JPG в .jpg ─────────────────────────────────────────
print("=== Шаг 1: конвертация изображений ===")
convert_count = 0
skip_count = 0

for src in sorted(IMG_DIR.iterdir()):
    if src.suffix.lower() == ".jpg":
        continue                                 # уже JPG
    if src.suffix.lower() not in {".jpeg", ".jfif", ".webp", ".png", ".bmp", ".gif"}:
        continue                                 # не изображение

    dst = src.with_suffix(".jpg")

    if dst.exists():
        print(f"  ПРОПУСК  {src.name}  →  {dst.name} (уже есть)")
        skip_count += 1
        continue

    try:
        with Image.open(src) as img:
            # JPEG не поддерживает прозрачность — конвертируем RGBA→RGB
            if img.mode in ("RGBA", "P", "LA"):
                bg = Image.new("RGB", img.size, (255, 255, 255))
                bg.paste(img, mask=img.split()[-1] if img.mode in ("RGBA", "LA") else None)
                img = bg
            elif img.mode != "RGB":
                img = img.convert("RGB")
            img.save(dst, "JPEG", quality=92, optimize=True)
        print(f"  OK  {src.name}  →  {dst.name}")
        convert_count += 1
    except Exception as e:
        print(f"  ОШИБКА  {src.name}: {e}")

print(f"Конвертировано: {convert_count}, пропущено: {skip_count}\n")


# ── 2. Строим карту: base_code → отсортированный список .jpg-файлов ───────────
print("=== Шаг 2: индексирование JPG-файлов ===")
# Паттерн: {артикул}_{номер}.jpg  или  {артикул}.jpg
STEM_RE = re.compile(r"^(.+?)(?:_(\d+))?$")

code_to_files: dict[str, list[tuple[int, str]]] = {}

for f in IMG_DIR.iterdir():
    if f.suffix.lower() != ".jpg":
        continue
    m = STEM_RE.match(f.stem)
    if not m:
        continue
    base = m.group(1).lower().strip()          # артикул в нижнем регистре
    num  = int(m.group(2)) if m.group(2) else 0
    code_to_files.setdefault(base, []).append((num, f.name))

# Сортируем по номеру фото
for key in code_to_files:
    code_to_files[key].sort()

print(f"Уникальных артикулов в папке: {len(code_to_files)}")
for code, files in list(code_to_files.items())[:10]:
    print(f"  {code}: {', '.join(n for _, n in files)}")
print()


# ── 3. Заполняем Excel ────────────────────────────────────────────────────────
print("=== Шаг 3: заполнение Excel ===")
wb = openpyxl.load_workbook(BASE_FILE)
ws = wb.active

filled = 0
not_found = 0

for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    kod_cell = ws.cell(i, KOD_COL)
    img_cell = ws.cell(i, IMG_COL)

    kod = kod_cell.value
    img = img_cell.value

    if not kod or not str(kod).strip():
        continue                                 # пустая строка
    if img and str(img).strip():
        continue                                 # изображение уже заполнено

    key = str(kod).strip().lower()
    matches = code_to_files.get(key)

    if matches:
        value = ", ".join(name for _, name in matches)
        img_cell.value = value
        print(f"  [{i:>4}]  {str(kod):<20}  →  {value[:70]}")
        filled += 1
    else:
        not_found += 1

print(f"\nЗаполнено строк: {filled}")
print(f"Не найдено изображений: {not_found}")

wb.save(BASE_FILE)
wb.close()
print(f"Файл сохранён: {BASE_FILE}")
