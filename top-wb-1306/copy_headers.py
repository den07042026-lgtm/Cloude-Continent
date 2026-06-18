import sys, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy
sys.stdout.reconfigure(encoding="utf-8")

folder = r"C:\Users\Admin\Desktop\Топ ВБ 1306"
src_path = folder + r"\Топ-500 ВБ 1306_BACKUP.xlsx"
dst_path = folder + r"\Топ-500 ВБ 1306.xlsx"

wb_src = openpyxl.load_workbook(src_path)
ws_src = wb_src.active

wb_dst = openpyxl.load_workbook(dst_path)
ws_dst = wb_dst.active

# Копируем значения и форматирование первой строки из бэкапа
for col in range(1, ws_src.max_column + 1):
    src_cell = ws_src.cell(1, col)
    dst_cell = ws_dst.cell(1, col)

    dst_cell.value = src_cell.value

    if src_cell.has_style:
        dst_cell.font      = src_cell.font.copy()
        dst_cell.fill      = src_cell.fill.copy()
        dst_cell.border    = src_cell.border.copy()
        dst_cell.alignment = src_cell.alignment.copy()
        dst_cell.number_format = src_cell.number_format

    print(f"  col{col}: {repr(src_cell.value)}")

# Высота первой строки
if ws_src.row_dimensions[1].height:
    ws_dst.row_dimensions[1].height = ws_src.row_dimensions[1].height

wb_dst.save(dst_path)
wb_dst.close()
wb_src.close()
print("\nГотово! Заголовки скопированы из бэкапа.")
