"""
normalize_names.py  v2
Приводит столбец "Наименование" к аккуратному виду.
Исправляет в том числе ошибки предыдущего прогона (Iii→III, Amd→AMD, акпп→АКПП).
"""
import sys
import re
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

BASE_FILE = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1306\Топ-500 ВБ 1306.xlsx")
NAME_COL  = 2

# ── Наборы ────────────────────────────────────────────────────────────────────

# Латинские аббревиатуры — всегда в верхнем регистре
KEEP_LATIN = {
    "VW", "BMW", "AMG", "TDI", "FSI", "TSI", "TFSI", "GTI", "SDI",
    "CDI", "HDI", "CRD", "DSG", "ATF", "EGR", "ABS", "ESP", "DPF",
    "FAP", "SCR", "CVT", "AWD", "SUV", "PHEV", "MHEV", "BEV",
    "DCI", "CDTI", "JTD", "JTDM", "THP", "CRDI", "CVVT", "EFI",
    "MPI", "GDI", "VVT", "VTEC", "DOHC", "AMT", "DCT", "LPG", "CNG",
    "ZF", "LSD", "WRC", "DTM", "OEM",
    # Бренд-коды
    "AMD", "SCT",
    # Римские цифры ≥ 3 символов (I/II/IV/V/VI/X — len≤2, покрываются ниже)
    "III", "VII", "VIII", "IX", "XI", "XII", "XIII", "XIV", "XV",
    "XVI", "XVII", "XVIII", "XIX", "XX", "XXI",
}

# Кириллические аббревиатуры — всегда в верхнем регистре
KEEP_CYRILLIC = {
    "ВАЗ", "ГАЗ", "УАЗ", "КАМАЗ", "ЗМЗ", "ЗИЛ", "МАЗ", "КРАЗ",
    "РАФ", "ПАЗ", "ЛАЗ", "ЛИАЗ",
    "ЗИТ", "ДААЗ", "РТПЗ", "ПРАМО", "КЗАТЭ", "КЗТАЭ", "АВТОВАЗ",
    "АКПП", "КПП", "МКПП", "ЭСУД",
    "САРАНСК", "БЕЛЕБЕЙ",
}

# Кириллические собственные имена — Title Case
KEEP_TITLE_CYR = {
    "КАЛИНА", "НИВА", "ПРИОРА", "ГРАНТА", "САМАРА", "ТАВРИЯ",
}

LATIN_VOWELS = set("AEIOU")


def is_roman(s: str) -> bool:
    """True если s (в верхнем регистре) — валидное римское число ≥ I."""
    if not s or not re.match(r"^[IVXLCDM]+$", s):
        return False
    return bool(re.match(
        r"^M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3})$",
        s
    ) and s != "")


def _process_latin_word(word: str) -> str:
    """Одно латинское слово без дефисов."""
    if not word:
        return word
    upper = word.upper()

    # Известная аббревиатура → всегда заглавные (исправляет Amd→AMD)
    if upper in KEEP_LATIN:
        return upper

    # Короткое слово (≤ 2 букв: VW, GT, c, I, II...) → без изменений
    # ВАЖНО: до is_roman, иначе 'c'→'C' (C=100 в римских)
    letters = re.sub(r"[^A-Za-z]", "", word)
    if len(letters) <= 2:
        return word

    # Римское число (len≥3) → всегда заглавные (исправляет Iii→III, Vii→VII)
    if is_roman(upper):
        return upper

    # Без гласных → аббревиатура → без изменений (BMW, GTS...)
    if not any(c in LATIN_VOWELS for c in upper):
        return word

    # Всё в верхнем регистре + есть гласная → Title Case (PASSAT→Passat)
    if word == upper:
        return upper[0].upper() + upper[1:].lower()

    # Уже Mixed/lowercase (Passat, passat) → без изменений
    return word


def _process_latin_token(core: str) -> str:
    """Латинский токен, возможно с дефисами (VI-VIII, MANN-FILTER)."""
    parts = core.split("-")
    return "-".join(_process_latin_word(p) for p in parts)


def _transform_core(core: str, is_first: bool) -> str:
    """Нормализует ядро токена (без обрамляющей пунктуации)."""

    # Содержит цифры → без изменений
    if re.search(r"\d", core):
        return core

    # Содержит встроенную пунктуацию (кроме дефиса) → без изменений
    # например РЕМ.НАБОР, NEXIA,LANOS
    if re.search(r"[^A-Za-zА-ЯЁа-яё\d\-]", core):
        return core

    has_cyr = bool(re.search(r"[А-ЯЁа-яё]", core))
    has_lat = bool(re.search(r"[A-Za-z]", core))

    # Смешанный Cyrillic+Latin → без изменений
    if has_cyr and has_lat:
        return core

    # ── Кириллица ─────────────────────────────────────────────────────────────
    if has_cyr:
        upper = core.upper()
        pure  = upper.strip("-")

        # Известная аббревиатура → всегда ВЕРХНИЙ регистр
        if pure in KEEP_CYRILLIC:
            return upper

        # Известное собственное имя → всегда Title Case
        if pure in KEEP_TITLE_CYR:
            return upper[0].upper() + upper[1:].lower()

        # Уже mixed/lowercase → без изменений (не понижаем то, что уже норм)
        if core != upper:
            return core

        # Всё в верхнем регистре → sentence-case
        low = core.lower()
        if is_first:
            return low[0].upper() + low[1:]
        return low

    # ── Только латиница ────────────────────────────────────────────────────────
    if has_lat:
        return _process_latin_token(core)

    return core


def normalize_token(raw: str, is_first: bool) -> str:
    m = re.match(
        r"^([^A-Za-zА-ЯЁа-яё\d]*)(.+?)([^A-Za-zА-ЯЁа-яё\d]*)$",
        raw
    )
    if not m:
        return raw
    prefix, core, suffix = m.groups()
    return prefix + _transform_core(core, is_first) + suffix


def normalize_name(name: str) -> str:
    s = str(name).strip()
    if not s:
        return name
    parts = re.split(r"(\s+)", s)
    result = []
    word_idx = 0
    for part in parts:
        if re.match(r"^\s+$", part):
            result.append(part)
        else:
            result.append(normalize_token(part, is_first=(word_idx == 0)))
            word_idx += 1
    out = "".join(result)
    if out and out[0].islower():
        out = out[0].upper() + out[1:]
    return out


# ── Обработка ─────────────────────────────────────────────────────────────────
print("Открываю файл...")
wb = openpyxl.load_workbook(BASE_FILE)
ws = wb.active

changed = 0
for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    cell = ws.cell(i, NAME_COL)
    val  = cell.value
    if not val or not str(val).strip():
        continue
    new_val = normalize_name(str(val))
    if new_val != str(val):
        print(f"[{i:>4}]  {str(val)[:65]}")
        print(f"      →  {new_val[:65]}")
        cell.value = new_val
        changed += 1

print(f"\nИзменено строк: {changed}")
wb.save(BASE_FILE)
wb.close()
print(f"Файл сохранён: {BASE_FILE}")
