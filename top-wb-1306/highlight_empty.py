"""
highlight_empty.py
1. Снимает все цветовые отметки ячеек
2. Подсвечивает жёлтым: ячейку "Код / Артикул" + пустые ячейки в строках,
   где есть хоть одна заполненная ячейка и хоть одна пустая.
   Полностью пустые строки не трогает.
"""
import sys
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

BASE_FILE = Path(r"C:\Users\Admin\Desktop\Топ ВБ 1306\Топ-500 ВБ 1306.xlsx")
MAX_COL   = 15   # проверяем столбцы 1..15

NO_FILL     = PatternFill(fill_type=None)
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def is_empty(val) -> bool:
    return val is None or str(val).strip() == ""


print("Открываю файл...")
wb = openpyxl.load_workbook(BASE_FILE)
ws = wb.active

total_rows = ws.max_row
print(f"Строк в файле: {total_rows}")

# ── Шаг 1: снимаем все цвета ──────────────────────────────────────────────────
print("Снимаю все цвета...")
cleared = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.fill and cell.fill.fill_type not in (None, "none"):
            cell.fill = NO_FILL
            cleared += 1
print(f"  Очищено ячеек с цветом: {cleared}")

# ── Шаг 2: подсвечиваем нужные строки ────────────────────────────────────────
print("Расставляю жёлтые отметки...")
highlighted_rows = 0
highlighted_cells = 0

for row_idx in range(2, total_rows + 1):
    vals = [ws.cell(row_idx, c).value for c in range(1, MAX_COL + 1)]

    empty_flags = [is_empty(v) for v in vals]
    has_filled  = any(not e for e in empty_flags)
    has_empty   = any(e for e in empty_flags)

    # Полностью пустая строка → пропускаем
    if not has_filled:
        continue

    # Строка полностью заполнена → пропускаем
    if not has_empty:
        continue

    # Есть и заполненные, и пустые → подсвечиваем
    # 1) ячейка Код / Артикул (col 1)
    ws.cell(row_idx, 1).fill = YELLOW_FILL
    highlighted_cells += 1

    # 2) все пустые ячейки строки
    for col_idx, empty in enumerate(empty_flags, start=1):
        if empty:
            ws.cell(row_idx, col_idx).fill = YELLOW_FILL
            highlighted_cells += 1

    highlighted_rows += 1

print(f"  Строк с неполными данными: {highlighted_rows}")
print(f"  Подсвечено ячеек: {highlighted_cells}")

wb.save(BASE_FILE)
wb.close()
print(f"\nГотово! Файл сохранён: {BASE_FILE}")
