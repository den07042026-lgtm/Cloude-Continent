# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "selenium",
#   "webdriver-manager",
#   "openpyxl",
#   "Pillow",
# ]
# ///
"""
Скачивание изображений с autopiter.ru для строк Топ-500 ВБ без изображений.
Selenium для рендеринга страниц и скачивания изображений (обход антибота).
Запуск:  uv run autopiter_images.py
"""

import base64
import io
import re
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

sys.stdout.reconfigure(encoding="utf-8")

# ──────────────────────────────────────────────────────────
EXCEL_PATH  = r"C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ_new.xlsx"
IMAGES_DIR  = Path(r"C:\Users\Admin\Desktop\Топ-500 ВБ\Изображения Автопитер")
BASE_URL    = "https://autopiter.ru"
DELAY       = 2.0   # пауза между артикулами (сек)
PAGE_WAIT   = 6     # секунды ожидания React-рендеринга после загрузки
# ──────────────────────────────────────────────────────────

RED     = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
NO_FILL = PatternFill(fill_type="none")

# Скачивает текущую страницу браузера через fetch(location.href).
# Работает без CORS — мы уже на нужном origin.
# AbortController (25 сек) + Selenium script_timeout (30 сек) = два слоя защиты от зависания.
_FETCH_THIS_PAGE = """
var done = arguments[0];
var ctrl = new AbortController();
var tid  = setTimeout(function() { ctrl.abort(); done(null); }, 25000);
fetch(location.href, {cache: 'force-cache', signal: ctrl.signal})
    .then(function(r) {
        clearTimeout(tid);
        if (!r.ok) { done(null); return; }
        return r.blob();
    })
    .then(function(blob) {
        if (!blob) { done(null); return; }
        var reader = new FileReader();
        reader.onloadend = function() { done(reader.result); };
        reader.readAsDataURL(blob);
    })
    .catch(function() { clearTimeout(tid); done(null); });
"""


# ══════════════════════════════════════════════════════════
# Вспомогательные функции
# ══════════════════════════════════════════════════════════

def strip_prefix(code: str) -> str:
    """Обрезает Микадо-префикс x{бренд}-  →  чистый артикул."""
    return re.sub(r'^x[a-z]+-', '', str(code).strip().lower())


def norm_brand(b: str) -> str:
    """Нормализация: без пробелов/дефисов/подчёркиваний, верхний регистр."""
    return re.sub(r'[\s\-_]', '', b).upper()


# ══════════════════════════════════════════════════════════
# Selenium-браузер
# ══════════════════════════════════════════════════════════

def make_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=opts)
    driver.set_window_size(1400, 900)
    driver.set_script_timeout(30)
    return driver


def get_rendered_html(driver: webdriver.Chrome, article: str) -> str | None:
    url = f"{BASE_URL}/goods/{article.lower()}"
    try:
        driver.get(url)
        WebDriverWait(driver, 20).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        time.sleep(PAGE_WAIT)
        return driver.page_source
    except Exception:
        return None


# ══════════════════════════════════════════════════════════
# Парсинг изображений из HTML
# ══════════════════════════════════════════════════════════

def get_images_for_brand(html: str, article: str, target_brand: str) -> list[str]:
    target_norm = norm_brand(target_brand)
    if not target_norm:
        return []

    art_upper  = article.upper()
    art_nodash = re.sub(r'[-_]', '', art_upper)

    folder_imgs: dict[str, list[str]] = {}
    for art_esc in dict.fromkeys([re.escape(art_upper), re.escape(art_nodash)]):
        img_re = re.compile(
            r'/static2/detailphoto/w\d+/(P\d+)/(' + art_esc + r'[^"\'>\s]*)',
            re.I,
        )
        for folder, filename in img_re.findall(html):
            full_url = f"{BASE_URL}/static2/detailphoto/w2000/{folder}/{filename}"
            folder_imgs.setdefault(folder, [])
            if full_url not in folder_imgs[folder]:
                folder_imgs[folder].append(full_url)

    if not folder_imgs:
        return []

    for folder, urls in folder_imgs.items():
        folder_str = f"/{folder}/"
        for m in re.finditer(re.escape(folder_str), html, re.I):
            start   = max(0, m.start() - 1200)
            end     = min(len(html), m.end() + 500)
            context = norm_brand(html[start:end])
            if target_norm in context:
                return urls

    return []


# ══════════════════════════════════════════════════════════
# Загрузка изображений через браузер
# ══════════════════════════════════════════════════════════

def download_via_browser(driver: webdriver.Chrome, url: str, save_path: Path) -> bool:
    """
    Открывает URL изображения в новой вкладке, затем скачивает его через
    fetch(location.href) — без CORS-проблем, браузер кэширует файл при открытии.
    """
    original = driver.current_window_handle

    for size in ("w2000", "w300", "w100"):
        try_url  = re.sub(r'/w\d+/', f'/{size}/', url)
        new_tab  = None
        result   = None

        try:
            driver.execute_script("window.open(arguments[0], '_blank');", try_url)
            new_tabs = [h for h in driver.window_handles if h != original]
            if not new_tabs:
                continue
            new_tab = new_tabs[-1]
            driver.switch_to.window(new_tab)

            WebDriverWait(driver, 20).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            time.sleep(0.5)

            result = driver.execute_async_script(_FETCH_THIS_PAGE)

        except Exception:
            pass
        finally:
            if new_tab:
                try:
                    driver.close()
                except Exception:
                    pass
            try:
                driver.switch_to.window(original)
            except Exception:
                pass

        if result and isinstance(result, str) and ',' in result:
            try:
                img_bytes = base64.b64decode(result.split(',', 1)[1])
                if img_bytes:
                    img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
                    img.save(save_path, format="JPEG", quality=95)
                    return True
            except Exception:
                pass

    return False


# ══════════════════════════════════════════════════════════
# Excel helpers
# ══════════════════════════════════════════════════════════

def find_col(ws, keyword: str) -> int:
    for cell in ws[1]:
        if cell.value and keyword.lower() in str(cell.value).lower():
            return cell.column
    raise ValueError(f"Столбец '{keyword}' не найден")


def safe_save(wb, path: str) -> None:
    """Сохраняет через временный файл — оригинал не трогается до успешной записи."""
    p    = Path(path)
    tmp  = p.with_suffix(".tmp.xlsx")
    for attempt in range(1, 11):
        try:
            wb.save(tmp)
            tmp.replace(p)
            return
        except PermissionError:
            if tmp.exists():
                try:
                    tmp.unlink()
                except Exception:
                    pass
            if attempt == 1:
                print("\n  [!] Файл занят — закройте Excel и нажмите Enter...",
                      end="", flush=True)
                input()
            else:
                print(f"  [!] Ещё занят, жду 5 сек (попытка {attempt}/10)...", flush=True)
                time.sleep(5)
        except Exception as e:
            if tmp.exists():
                try:
                    tmp.unlink()
                except Exception:
                    pass
            raise e
    raise PermissionError(f"Не удалось сохранить {path}")


# ══════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════

def main():
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    col_code   = find_col(ws, "Код")
    col_brand  = find_col(ws, "Бренд")
    col_images = find_col(ws, "Изображени")

    rows_todo = [
        r for r in range(2, ws.max_row + 1)
        if not ws.cell(row=r, column=col_images).value
        and ws.cell(row=r, column=col_code).value
    ]
    total = len(rows_todo)
    print(f"Строк без изображений: {total}")

    print("Очищаю цвета ячеек...", end=" ", flush=True)
    for row_num in rows_todo:
        ws.cell(row=row_num, column=col_code).fill = NO_FILL
    safe_save(wb, EXCEL_PATH)
    print("готово.\n")

    print("Запускаю браузер...")
    driver = make_driver()
    stats  = {"ok": 0, "no_img": 0, "skip": 0}

    try:
        for i, row_num in enumerate(rows_todo, 1):
            raw_code = str(ws.cell(row=row_num, column=col_code).value).strip()
            brand    = str(ws.cell(row=row_num, column=col_brand).value or "").strip()
            article  = strip_prefix(raw_code)

            print(f"[{i}/{total}] {raw_code} → {article}  [{brand}]",
                  end=" ... ", flush=True)

            art_cell = ws.cell(row=row_num, column=col_code)

            if "/" in article:
                print("пропуск (слэш) -> red")
                art_cell.fill = RED
                stats["skip"] += 1
                safe_save(wb, EXCEL_PATH)
                continue

            try:
                html = get_rendered_html(driver, article)
                if not html:
                    print("ошибка загрузки -> red")
                    art_cell.fill = RED
                    stats["no_img"] += 1
                    safe_save(wb, EXCEL_PATH)
                    continue

                img_urls = get_images_for_brand(html, article, brand)

                if not img_urls:
                    print("не найдено -> red")
                    art_cell.fill = RED
                    stats["no_img"] += 1
                    safe_save(wb, EXCEL_PATH)
                    time.sleep(DELAY)
                    continue

                print(f"найдено {len(img_urls)} URL", end=" ... ", flush=True)

                saved: list[str] = []
                for idx, url in enumerate(img_urls, 1):
                    filename  = f"{article}_{idx}.jpg"
                    save_path = IMAGES_DIR / filename
                    if download_via_browser(driver, url, save_path):
                        saved.append(filename)
                    time.sleep(0.5)

                if saved:
                    ws.cell(row=row_num, column=col_images).value = "; ".join(saved)
                    art_cell.fill = NO_FILL
                    print(f"сохранено {len(saved)}: {', '.join(saved)}")
                    stats["ok"] += 1
                else:
                    print("ошибка загрузки -> red")
                    art_cell.fill = RED
                    stats["no_img"] += 1

            except Exception as e:
                print(f"ОШИБКА: {e} -> red")
                art_cell.fill = RED
                stats["no_img"] += 1

            safe_save(wb, EXCEL_PATH)
            time.sleep(DELAY)

    finally:
        driver.quit()

    safe_save(wb, EXCEL_PATH)
    print()
    print("=" * 50)
    print(f"Готово!")
    print(f"  Скачано      : {stats['ok']}")
    print(f"  Не найдено   : {stats['no_img']}")
    print(f"  Пропущено    : {stats['skip']}")
    print(f"  Папка        : {IMAGES_DIR}")


if __name__ == "__main__":
    main()
