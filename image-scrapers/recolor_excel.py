# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""
Снимает все цвета с ячеек артикула, затем жёлтым отмечает строки без изображений.
Запуск: uv run --offline recolor_excel.py
"""

import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding="utf-8")

EXCEL_PATH = r"C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ_new.xlsx"

YELLOW  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NO_FILL = PatternFill(fill_type="none")


def safe_save(wb, path: str) -> None:
    p   = Path(path)
    tmp = p.with_suffix(".tmp.xlsx")
    for attempt in range(1, 11):
        try:
            wb.save(tmp)
            tmp.replace(p)
            return
        except PermissionError:
            if tmp.exists():
                try: tmp.unlink()
                except Exception: pass
            if attempt == 1:
                print("\n  [!] Файл занят — закройте Excel и нажмите Enter...",
                      end="", flush=True)
                input()
            else:
                print(f"  [!] Ещё занят, жду 5 сек (попытка {attempt}/10)...", flush=True)
                time.sleep(5)
        except Exception as e:
            if tmp.exists():
                try: tmp.unlink()
                except Exception: pass
            raise e
    raise PermissionError(f"Не удалось сохранить {path}")


def find_col(ws, keyword: str) -> int:
    for cell in ws[1]:
        if cell.value and keyword.lower() in str(cell.value).lower():
            return cell.column
    raise ValueError(f"Столбец с '{keyword}' не найден")


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    col_article = find_col(ws, "Код")
    col_images  = find_col(ws, "Изображени")

    total = ws.max_row - 1
    yellow_count = 0

    for row in range(2, ws.max_row + 1):
        art_cell = ws.cell(row=row, column=col_article)
        has_img  = bool(ws.cell(row=row, column=col_images).value)

        if has_img:
            art_cell.fill = NO_FILL
        else:
            art_cell.fill = YELLOW
            yellow_count += 1

    safe_save(wb, EXCEL_PATH)

    print(f"Всего строк    : {total}")
    print(f"С изображением : {total - yellow_count}")
    print(f"Без изображения: {yellow_count} (жёлтых)")
    print("Готово!")


if __name__ == "__main__":
    main()
