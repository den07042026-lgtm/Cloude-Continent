import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment

source_dir = r"C:\Users\Admin\Desktop\На сортировку 21.04(3)"
dest_dir   = r"C:\Users\Admin\Desktop\На сортировку 24.04"
ROW_HEIGHT = 15  # высота одной строки в пт

os.makedirs(dest_dir, exist_ok=True)

files = [f for f in os.listdir(source_dir) if f.lower().endswith(('.xlsx', '.xlsm'))]
print(f"Файлов для обработки: {len(files)}")

for i, filename in enumerate(files, 1):
    src = os.path.join(source_dir, filename)
    dst = os.path.join(dest_dir, filename)

    try:
        wb = load_workbook(src)
        for ws in wb.worksheets:

            # --- Удаляем столбцы с "Аналог" в заголовке (первая строка) ---
            cols_to_delete = []
            for col in ws.iter_cols(min_row=1, max_row=1):
                cell = col[0]
                if cell.value and "аналог" in str(cell.value).lower():
                    cols_to_delete.append(cell.column)
            # Удаляем с конца, чтобы не сбивать индексы
            for col_idx in sorted(cols_to_delete, reverse=True):
                ws.delete_cols(col_idx)

            # --- Включаем wrap_text=True для всех ячеек ---
            # (предотвращает горизонтальный overflow в соседние ячейки)
            for row in ws.iter_rows():
                for cell in row:
                    al = cell.alignment
                    kwargs = dict(
                        horizontal=al.horizontal,
                        vertical=al.vertical,
                        text_rotation=al.text_rotation,
                        wrap_text=True,   # ключевой параметр — без него текст вытекает
                        shrink_to_fit=False,
                        indent=al.indent,
                    )
                    if hasattr(al, 'reading_order'):
                        kwargs['reading_order'] = al.reading_order
                    cell.alignment = Alignment(**kwargs)

            # --- Фиксируем высоту ВСЕХ строк (не только тех, что уже в row_dimensions) ---
            # Итерируем по номеру строки, а не по values() — иначе новые строки не создаются
            for row_num in range(1, ws.max_row + 1):
                ws.row_dimensions[row_num].height = ROW_HEIGHT

        wb.save(dst)
        print(f"[{i}/{len(files)}] ОК: {filename}")
    except PermissionError:
        print(f"[{i}/{len(files)}] ПРОПУЩЕН (файл открыт в Excel): {filename}")
    except Exception as e:
        print(f"[{i}/{len(files)}] ОШИБКА {filename}: {e}")
        try:
            shutil.copy2(src, dst)
        except Exception:
            pass

print(f"\nГотово! Файлы сохранены в: {dest_dir}")
