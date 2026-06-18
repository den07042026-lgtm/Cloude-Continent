# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""
Поиск и копирование изображений Микадо
Запуск: uv run mikado_images.py
"""

import shutil
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding="utf-8")

# ──────────────────────────────────────────────────────────
EXCEL_PATH = r"C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ_new.xlsx"
IMAGES_SRC = Path(r"C:\Users\Admin\Desktop\На сортировку 08.05\images")
IMAGES_DST = Path(r"C:\Users\Admin\Desktop\Топ-500 ВБ\Изображения Микадо")
# ──────────────────────────────────────────────────────────

RED     = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
NO_FILL = PatternFill(fill_type="none")

IMG_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}


def find_col(ws, keyword: str) -> int:
    for cell in ws[1]:
        if cell.value and keyword.lower() in str(cell.value).lower():
            return cell.column
    raise ValueError(f"Столбец '{keyword}' не найден")


def find_images(article: str, all_files: list[Path]) -> list[Path]:
    """Все файлы, чьё имя (без расширения) содержит артикул (без учёта регистра)."""
    needle = article.lower()
    return sorted(f for f in all_files if needle in f.stem.lower())


def safe_save(wb, path: str) -> None:
    """Сохраняет через временный файл — оригинал не трогается до успешной записи."""
    p   = Path(path)
    tmp = p.with_suffix(".tmp.xlsx")
    for attempt in range(1, 11):
        try:
            wb.save(tmp)
            tmp.replace(p)
            return
        except PermissionError:
            if tmp.exists():
                try:
                    tmp.unlink()
                except Exception:
                    pass
            if attempt == 1:
                print("\n  [!] Файл занят — закройте Excel и нажмите Enter...",
                      end="", flush=True)
                input()
            else:
                print(f"  [!] Ещё занят, жду 5 сек (попытка {attempt}/10)...", flush=True)
                time.sleep(5)
        except Exception as e:
            if tmp.exists():
                try:
                    tmp.unlink()
                except Exception:
                    pass
            raise e
    raise PermissionError(f"Не удалось сохранить {path}")


def main():
    IMAGES_DST.mkdir(parents=True, exist_ok=True)

    all_files = [
        f for f in IMAGES_SRC.iterdir()
        if f.is_file() and f.suffix.lower() in IMG_EXTS
    ]
    print(f"Изображений в источнике: {len(all_files)}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    col_article  = find_col(ws, "Код")
    col_supplier = find_col(ws, "Поставщик")
    col_images   = find_col(ws, "Изображени")

    rows_mikado = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=col_supplier).value
        and "микадо" in str(ws.cell(row=r, column=col_supplier).value).lower()
    ]
    total = len(rows_mikado)
    print(f"Позиций с Микадо: {total}\n")

    stats = {"ok": 0, "red": 0}

    for i, row_num in enumerate(rows_mikado, 1):
        article = ws.cell(row=row_num, column=col_article).value
        if not article:
            continue

        article  = str(article).strip()
        matches  = find_images(article, all_files)
        art_cell = ws.cell(row=row_num, column=col_article)

        if not matches:
            print(f"[{i}/{total}] {article}: не найдено -> красный")
            art_cell.fill = RED
            stats["red"] += 1
        else:
            copied = []
            for src in matches:
                dst = IMAGES_DST / src.name
                shutil.copy2(src, dst)
                copied.append(src.name)

            ws.cell(row=row_num, column=col_images).value = "; ".join(copied)
            art_cell.fill = NO_FILL
            print(f"[{i}/{total}] {article}: {len(copied)} фото -> {', '.join(copied)}")
            stats["ok"] += 1

    safe_save(wb, EXCEL_PATH)

    print(f"\n{'='*50}")
    print(f"Готово!")
    print(f"  Найдено и скопировано: {stats['ok']}")
    print(f"  Не найдено (красные) : {stats['red']}")
    print(f"  Файлы в: {IMAGES_DST}")


if __name__ == "__main__":
    main()
