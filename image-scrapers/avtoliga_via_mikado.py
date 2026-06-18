# /// script
# requires-python = ">=3.10"
# dependencies = ["requests", "openpyxl"]
# ///
"""
Поиск изображений для артикулов Автолиги на mikado-parts.ru
Запуск: uv run avtoliga_via_mikado.py
"""

import re
import sys
import time
from hashlib import md5
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding="utf-8")

# ──────────────────────────────────────────────────────────
EXCEL_PATH  = r"C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ.xlsx"
IMAGES_DIR  = Path(r"C:\Users\Admin\Desktop\Топ-500 ВБ\Изображения Автолига")

BASE_URL    = "https://mikado-parts.ru/office"
LOGIN_URL   = f"{BASE_URL}/SECURE.asp"
PRODUCT_URL = f"{BASE_URL}/galleyp.asp"
AJAX_URL    = f"{BASE_URL}/pp0.asp"

MIKADO_CODE = "35275"
MIKADO_PASS = "KONSHIN1963"

DELAY = 1.5
# ──────────────────────────────────────────────────────────

YELLOW  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NO_FILL = PatternFill(fill_type="none")


def do_login() -> requests.Session:
    session = requests.Session()
    session.headers["User-Agent"] = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"
    )
    resp = session.post(
        LOGIN_URL,
        data={"CODE": MIKADO_CODE, "PASSWORD": MIKADO_PASS, "INSERT": "OK"},
        timeout=20,
    )
    resp.raise_for_status()
    page = resp.content.decode("windows-1251", errors="replace")
    if "Обслуживание клиентов" in page or "Продолжить" in page:
        print("  Login OK")
    else:
        print("  Login FAILED")
    return session


def get_artid(session: requests.Session, code: str) -> str | None:
    resp = session.get(PRODUCT_URL, params={"code": code}, timeout=20)
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")
    m = re.search(r"ARTID=(\d+)", html)
    return m.group(1) if m else None


def download_images(session: requests.Session, code: str, artid: str,
                    save_as: str) -> list[str]:
    """
    Качает изображения через pp0.asp?MODE=PIC.
    Сохраняет в IMAGES_DIR как {save_as}_1.jpg, {save_as}_2.jpg, ...
    """
    resp = session.get(
        AJAX_URL,
        params={"MODE": "PIC", "CODE": code, "ARTID": artid},
        timeout=20,
    )
    resp.raise_for_status()
    pic_html = resp.content.decode("windows-1251", errors="replace")

    img_srcs = re.findall(
        r"<img[^>]+src=[\'\"]?(wi/img\.asp[^\'\"> \t]+)", pic_html, re.I
    )
    if not img_srcs:
        return []

    saved = []
    seen  = set()
    idx   = 1

    for src in img_srcs:
        img_url = f"{BASE_URL}/{src}"
        try:
            r = session.get(img_url, timeout=20)
            r.raise_for_status()
            content = r.content
            if not (content[:2] == b"\xff\xd8" or content[:8] == b"\x89PNG\r\n\x1a\n"):
                continue
            digest = md5(content).hexdigest()
            if digest in seen:
                continue
            seen.add(digest)
            ext      = ".jpg" if content[:2] == b"\xff\xd8" else ".png"
            filename = f"{save_as}_{idx}{ext}"
            (IMAGES_DIR / filename).write_bytes(content)
            saved.append(filename)
            idx += 1
        except Exception:
            continue

    return saved


def find_col(ws, keyword: str) -> int:
    for cell in ws[1]:
        if cell.value and keyword.lower() in str(cell.value).lower():
            return cell.column
    raise ValueError(f"Column '{keyword}' not found")


def safe_name(article: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', article)


def main():
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    col_article  = find_col(ws, "Код")
    col_supplier = find_col(ws, "Поставщик")
    col_images   = find_col(ws, "Изображени")

    # Все строки Автолиги без изображений
    rows = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=col_supplier).value
        and "автолига" in str(ws.cell(row=r, column=col_supplier).value).lower()
        and not ws.cell(row=r, column=col_images).value
    ]
    total = len(rows)
    print(f"Avtoliga rows without images: {total}")

    print("Logging in to mikado-parts.ru...")
    session = do_login()
    print()

    stats = {"ok": 0, "not_found": 0}

    for i, row_num in enumerate(rows, 1):
        article = ws.cell(row=row_num, column=col_article).value
        if not article:
            continue
        article = str(article).strip()

        print(f"[{i}/{total}] {article}", end=" ... ", flush=True)
        art_cell = ws.cell(row=row_num, column=col_article)

        try:
            artid = get_artid(session, article)
            if not artid:
                print("not found")
                art_cell.fill = YELLOW
                stats["not_found"] += 1
                wb.save(EXCEL_PATH)
                time.sleep(DELAY)
                continue

            time.sleep(0.4)
            images = download_images(session, article, artid, save_as=safe_name(article))

            if not images:
                print("no images")
                art_cell.fill = YELLOW
                stats["not_found"] += 1
            else:
                ws.cell(row=row_num, column=col_images).value = "; ".join(images)
                art_cell.fill = NO_FILL
                print(f"{len(images)} images: {', '.join(images)}")
                stats["ok"] += 1

        except Exception as e:
            print(f"ERROR: {e}")
            art_cell.fill = YELLOW
            stats["not_found"] += 1

        wb.save(EXCEL_PATH)
        time.sleep(DELAY)

    wb.save(EXCEL_PATH)
    print()
    print("=" * 50)
    print(f"Done!")
    print(f"  Found & saved : {stats['ok']}")
    print(f"  Not found     : {stats['not_found']}")
    print(f"  Saved to      : {IMAGES_DIR}")


if __name__ == "__main__":
    main()
