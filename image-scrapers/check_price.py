import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

price_file = 'C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx'
wb = openpyxl.load_workbook(price_file, read_only=True, data_only=True)
ws = wb.active

first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
headers = [str(v).strip() if v else f'col{i}' for i, v in enumerate(first_row, 1)]
print('Заголовки:', headers[:10])

targets = {'lf-1076','lf-1391','if-3212p','if-3097k','if-3229p','if-3233k','if-3083k','vsp-0304'}
found = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    for val in row:
        if val and str(val).strip().lower() in targets:
            d = dict(zip(headers, row))
            key = str(val).strip().lower()
            found[key] = d
            break

print(f'\nНайдено: {len(found)}/{len(targets)}')
for art, d in found.items():
    prodnum = d.get('Prodnum')
    code = d.get('Code')
    print(f'  {art} -> Prodnum={prodnum}, Code={code}')

print('\nНе найдены:', targets - set(found.keys()))
wb.close()
