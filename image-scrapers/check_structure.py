import openpyxl

folder = 'C:/Users/Admin/Desktop/Топ ВБ 1306'

for fname in ['Топ-500 ВБ 1306.xlsx', 'Топ-500 ВБ_new1306.xlsx']:
    wb = openpyxl.load_workbook(f'{folder}/{fname}', read_only=True)
    ws = wb.active
    print(f'\n=== {fname} ===')
    print(f'Строк: {ws.max_row} | Столбцов: {ws.max_column}')
    print('Заголовки (строка 1):')
    for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
        print(f'  {i}: {repr(cell.value)}')
    print('Пример строки 2:')
    for i, cell in enumerate(next(ws.iter_rows(min_row=2, max_row=2)), 1):
        print(f'  {i}: {repr(cell.value)}')
    wb.close()
