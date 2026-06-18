# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "selenium",
#   "openpyxl",
#   "webdriver-manager",
#   "requests",
# ]
# ///
"""
Скачивание изображений с b2b.avtoliga.ru
по позициям из Топ-500 ВБ_new.xlsx (Поставщик = Автолига)

Запуск:  uv run avtoliga_images.py
"""

import re
import sys
import time
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

sys.stdout.reconfigure(encoding="utf-8")

# ──────────────────────────────────────────────────────────
EXCEL_PATH   = r"C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ_new.xlsx"
IMAGES_DIR   = Path(r"C:\Users\Admin\Desktop\Топ-500 ВБ\Изображения Автолига")
SITE_URL     = "https://b2b.avtoliga.ru"
LOGIN_EMAIL  = "Control.vlz@gmail.com"
LOGIN_PASS   = "7iXOaxSpU6"
TIMEOUT      = 20
# ──────────────────────────────────────────────────────────

YELLOW  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE  = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
NO_FILL = PatternFill(fill_type="none")

NOT_FOUND      = "not_found"
BRAND_MISMATCH = "brand_mismatch"
NO_IMAGE       = "no_image"


# ══════════════════════════════════════════════════════════
# Вспомогательные функции
# ══════════════════════════════════════════════════════════

def brands_match(excel_brand: str, site_brand: str) -> bool:
    a = excel_brand.strip().upper()
    b = site_brand.strip().upper()
    return a in b or b in a


def safe_filename(article: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', article)


def find_col(ws, keyword: str) -> int:
    for cell in ws[1]:
        if cell.value and keyword.lower() in str(cell.value).lower():
            return cell.column
    raise ValueError(f"Столбец с '{keyword}' не найден. Заголовки: {[c.value for c in ws[1]]}")


def safe_save(wb, path: str) -> None:
    p   = Path(path)
    tmp = p.with_suffix(".tmp.xlsx")
    for attempt in range(1, 11):
        try:
            wb.save(tmp)
            tmp.replace(p)
            return
        except PermissionError:
            if tmp.exists():
                try: tmp.unlink()
                except Exception: pass
            if attempt == 1:
                print("\n  [!] Файл занят — закройте Excel и нажмите Enter...",
                      end="", flush=True)
                input()
            else:
                print(f"  [!] Ещё занят, жду 5 сек (попытка {attempt}/10)...", flush=True)
                time.sleep(5)
        except Exception as e:
            if tmp.exists():
                try: tmp.unlink()
                except Exception: pass
            raise e
    raise PermissionError(f"Не удалось сохранить {path}")


# ══════════════════════════════════════════════════════════
# Браузер
# ══════════════════════════════════════════════════════════

def make_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    driver.set_window_size(1400, 900)
    return driver


# ══════════════════════════════════════════════════════════
# Авторизация
# ══════════════════════════════════════════════════════════

def do_login(driver: webdriver.Chrome):
    print("Авторизация на сайте...")
    wait = WebDriverWait(driver, TIMEOUT)
    driver.get(SITE_URL)
    time.sleep(2)

    for xpath in [
        "//a[contains(@href,'login') or contains(@href,'auth')]",
        "//*[normalize-space(text())='Войти']",
        "//*[normalize-space(text())='Авторизация']",
        "//*[normalize-space(text())='Вход']",
        "//*[contains(@class,'login') or contains(@class,'auth') or contains(@class,'signin')]",
    ]:
        try:
            el = driver.find_element(By.XPATH, xpath)
            if el.is_displayed():
                el.click()
                time.sleep(1.5)
                break
        except Exception:
            pass

    for xpath in [
        "//input[@type='email']", "//input[@name='login']", "//input[@name='email']",
        "//input[@id='login']", "//input[@id='email']",
        "//input[contains(@placeholder,'mail') or contains(@placeholder,'огин')]",
    ]:
        try:
            f = wait.until(EC.visibility_of_element_located((By.XPATH, xpath)))
            f.clear(); f.send_keys(LOGIN_EMAIL)
            break
        except Exception:
            pass

    try:
        pw = wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@type='password']")))
        pw.clear(); pw.send_keys(LOGIN_PASS)
    except Exception:
        driver.save_screenshot(r"C:\Users\Admin\Desktop\debug_login_ERROR.png")
        raise RuntimeError("Поле пароля не найдено")

    for xpath in [
        "//button[normalize-space(.)='Вход']", "//button[normalize-space(.)='Войти']",
        "//button[contains(normalize-space(.),'Вход')]", "//button[contains(normalize-space(.),'Войти')]",
        "//input[@type='submit']",
    ]:
        try:
            btn = driver.find_element(By.XPATH, xpath)
            if btn.is_displayed():
                driver.execute_script("arguments[0].click();", btn)
                break
        except Exception:
            pass

    time.sleep(3)
    print(f"  URL после входа: {driver.current_url}")


def ensure_logged_in(driver: webdriver.Chrome):
    try:
        src = driver.page_source
        if "Выход" in src or "Личный кабинет" in src or "logout" in src.lower():
            return
    except Exception:
        pass
    print("  Сессия истекла — повторная авторизация...")
    do_login(driver)


# ══════════════════════════════════════════════════════════
# Поиск
# ══════════════════════════════════════════════════════════

def do_search(driver: webdriver.Chrome, article: str):
    search_input = None
    for xpath in [
        "//input[@name='pcode']",
        "//input[contains(@class,'ui-autocomplete-input')]",
        "//input[contains(@placeholder,'запчасти') or contains(@placeholder,'Введите')]",
        "//input[@name='article']",
        "//input[@type='text']",
    ]:
        try:
            for el in driver.find_elements(By.XPATH, xpath):
                if el.is_displayed():
                    search_input = el
                    break
            if search_input:
                break
        except Exception:
            pass

    if search_input is None:
        driver.save_screenshot(r"C:\Users\Admin\Desktop\debug_search_fail.png")
        raise RuntimeError("Поле поиска не найдено")

    search_input.clear()
    time.sleep(0.3)
    search_input.send_keys(article)
    time.sleep(0.5)

    find_btn = None
    for btn_xpath in ["./following-sibling::button[1]", "..//button", "../..//button"]:
        try:
            for b in search_input.find_elements(By.XPATH, btn_xpath):
                if b.is_displayed():
                    find_btn = b
                    break
            if find_btn:
                break
        except Exception:
            pass

    if find_btn is None:
        try:
            form = search_input.find_element(By.XPATH, "./ancestor::form[1]")
            for b in form.find_elements(By.XPATH, ".//button | .//input[@type='submit']"):
                if b.is_displayed():
                    find_btn = b
                    break
        except Exception:
            pass

    if find_btn is None:
        for xpath in ["//button[normalize-space(.)='Найти']", "//button[@type='submit']"]:
            try:
                for el in driver.find_elements(By.XPATH, xpath):
                    if el.is_displayed():
                        find_btn = el
                        break
                if find_btn:
                    break
            except Exception:
                pass

    if find_btn is None:
        raise RuntimeError("Кнопка поиска не найдена")

    driver.execute_script("arguments[0].click();", find_btn)
    time.sleep(3)


# ══════════════════════════════════════════════════════════
# Разбор результатов — сценарий 1: «Запрашиваемый артикул»
# ══════════════════════════════════════════════════════════

def get_requested_article_data(driver: webdriver.Chrome) -> tuple[str | None, object | None]:
    """Ищет в секции 'Запрашиваемый артикул'. Возвращает (бренд, миниатюра)."""
    try:
        heading = driver.find_element(By.XPATH, "//*[contains(text(), 'Запрашиваемый артикул')]")
    except NoSuchElementException:
        return None, None

    try:
        rows = heading.find_elements(By.XPATH, "./following::tr[position() <= 10]")
    except Exception:
        rows = []

    for row in rows:
        row_text = row.text.lower()
        if any(kw in row_text for kw in ["рекомендован", "предложени", "аналог", "товары дня"]):
            break

        cells = row.find_elements(By.TAG_NAME, "td")
        if not cells:
            continue

        thumbnail = None
        site_brand = None

        for cell in cells:
            if thumbnail is None:
                for img in cell.find_elements(By.TAG_NAME, "img"):
                    src = img.get_attribute("src") or ""
                    if src and "ico" not in src.lower() and "icon" not in src.lower():
                        thumbnail = img
                        break

            if site_brand is None:
                txt = cell.text.strip().split("\n")[0]
                if txt and 2 <= len(txt) <= 25 and not re.match(r'^\d[\d\s,.₽]*$', txt) and len(txt.split()) <= 4:
                    cls = cell.get_attribute("class") or ""
                    if "brand" in cls.lower() or "бренд" in cls.lower():
                        site_brand = txt
                    elif site_brand is None:
                        site_brand = txt

        if thumbnail is not None or site_brand is not None:
            return site_brand, thumbnail

    return None, None


# ══════════════════════════════════════════════════════════
# Разбор результатов — сценарий 2: общая таблица без заголовка
# ══════════════════════════════════════════════════════════

def find_thumbnail_in_results_table(
    driver: webdriver.Chrome, excel_brand: str
) -> tuple[str | None, object | None]:
    try:
        rows = driver.find_elements(By.XPATH, "//table//tr[.//td]")
    except Exception:
        return None, None

    for row in rows:
        cells = row.find_elements(By.TAG_NAME, "td")
        if len(cells) < 2:
            continue

        site_brand = None
        for cell in cells:
            txt = cell.text.strip().split("\n")[0]
            if txt and 2 <= len(txt) <= 30 and brands_match(excel_brand, txt):
                site_brand = txt
                break

        if site_brand is None:
            continue

        thumbnail = None
        for cell in row.find_elements(By.TAG_NAME, "td"):
            for img in cell.find_elements(By.TAG_NAME, "img"):
                src = img.get_attribute("src") or ""
                if src and "ico" not in src.lower() and "icon" not in src.lower():
                    thumbnail = img
                    break
            if thumbnail:
                break

        return site_brand, thumbnail

    return None, None


# ══════════════════════════════════════════════════════════
# Скачивание изображения
# ══════════════════════════════════════════════════════════

def _get_gallery_urls(driver: webdriver.Chrome) -> list[str]:
    """Все полноразмерные URL из <a data-fancybox-group href=...> в DOM."""
    try:
        urls = driver.execute_script("""
            var found = [];
            var seen  = {};
            var sets  = [
                '.article-image a[data-fancybox-group]',
                '.fancybox-wrap .article-image a[data-fancybox-group]',
                '#fancybox-content a[data-fancybox-group]',
                'a[data-fancybox-group]'
            ];
            for (var s = 0; s < sets.length; s++) {
                var links = document.querySelectorAll(sets[s]);
                if (!links.length) continue;
                for (var i = 0; i < links.length; i++) {
                    var u = links[i].href;
                    if (u && !seen[u]) {
                        seen[u] = true;
                        found.push(u);
                    }
                }
                if (found.length) break;
            }
            return found;
        """)
        return [u for u in (urls or []) if u.startswith("http")]
    except Exception:
        return []


def download_image(driver: webdriver.Chrome, thumbnail_el, article: str) -> list[str]:
    """Кликает миниатюру, ждёт попап, скачивает все изображения."""
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", thumbnail_el)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", thumbnail_el)
        time.sleep(2.5)
    except Exception as e:
        print(f"  Ошибка клика: {e}")
        return []

    base_name = safe_filename(article)
    saved: list[str] = []
    cookies = {c["name"]: c["value"] for c in driver.get_cookies()}

    gallery_urls = _get_gallery_urls(driver)

    if gallery_urls:
        print(f"  Галерея: {len(gallery_urls)} URL")
        for idx, url in enumerate(gallery_urls, 1):
            filename  = f"{base_name}_{idx}.jpg"
            save_path = IMAGES_DIR / filename
            if _download(url, cookies, save_path):
                saved.append(filename)
                print(f"    [{idx}] {filename}")
            else:
                print(f"    [{idx}] ошибка загрузки: {url[:60]}")
    else:
        # Фоллбэк: largest image in popup
        url = _get_big_image_url(driver)
        if url:
            filename  = f"{base_name}_1.jpg"
            save_path = IMAGES_DIR / filename
            if _download(url, cookies, save_path):
                saved.append(filename)
                print(f"    [1] {filename}")
            else:
                print(f"    [1] ошибка загрузки")
        else:
            print("  Изображение в попапе не найдено")

    return saved


def _get_big_image_url(driver: webdriver.Chrome) -> str | None:
    """Самое большое изображение в области попапа, или на странице."""
    # 1. Внутри fancybox / popup-контейнера
    for xpath in [
        "//*[contains(@class,'fancybox') or contains(@id,'fancybox')]//img[@src]",
        "//*[contains(@class,'modal') or contains(@class,'popup') or contains(@class,'lightbox') or contains(@class,'overlay')]//img[@src]",
    ]:
        try:
            best_area, best_url = 0, ""
            for img in driver.find_elements(By.XPATH, xpath):
                src = img.get_attribute("src") or ""
                if not src.startswith("http"):
                    continue
                w = driver.execute_script("return arguments[0].naturalWidth;", img) or 0
                h = driver.execute_script("return arguments[0].naturalHeight;", img) or 0
                if w > 100 and w * h > best_area:
                    best_area, best_url = w * h, src
            if best_url:
                return best_url
        except Exception:
            pass

    # 2. Самое большое на странице (фоллбэк)
    try:
        best_area, best_url = 0, ""
        for img in driver.find_elements(By.TAG_NAME, "img"):
            src = img.get_attribute("src") or ""
            if not src.startswith("http"):
                continue
            w = driver.execute_script("return arguments[0].naturalWidth;", img) or 0
            h = driver.execute_script("return arguments[0].naturalHeight;", img) or 0
            if w > 300 and w * h > best_area:
                best_area, best_url = w * h, src
        if best_url:
            return best_url
    except Exception:
        pass

    return None


def _download(url: str, cookies: dict, save_path: Path) -> bool:
    try:
        resp = requests.get(
            url, cookies=cookies,
            headers={"User-Agent": "Mozilla/5.0", "Referer": SITE_URL},
            timeout=30,
        )
        resp.raise_for_status()
        if resp.content:
            save_path.write_bytes(resp.content)
            return True
    except Exception:
        pass
    return False


# ══════════════════════════════════════════════════════════
# Обработка одного артикула
# ══════════════════════════════════════════════════════════

def process_article(driver: webdriver.Chrome, article: str, excel_brand: str) -> list[str] | str:
    driver.get(SITE_URL)
    time.sleep(2)
    ensure_logged_in(driver)

    if driver.current_url.rstrip("/") != SITE_URL.rstrip("/"):
        driver.get(SITE_URL)
        time.sleep(2)

    try:
        do_search(driver, article)
    except RuntimeError as e:
        print(f"  Ошибка поиска: {e}")
        return NOT_FOUND

    page = driver.page_source

    if "Запрашиваемый артикул" in page:
        site_brand, thumbnail = get_requested_article_data(driver)
        print(f"  [Сц.1] Бренд: '{site_brand}' | Excel: '{excel_brand}'")
        if not site_brand:
            return NOT_FOUND
        if not brands_match(excel_brand, site_brand):
            return BRAND_MISMATCH
        if thumbnail is None:
            return NO_IMAGE
        files = download_image(driver, thumbnail, article)
        return files if files else NO_IMAGE

    site_brand, thumbnail = find_thumbnail_in_results_table(driver, excel_brand)
    print(f"  [Сц.2] Бренд: '{site_brand}' | Excel: '{excel_brand}'")
    if not site_brand:
        return NOT_FOUND
    if thumbnail is None:
        return NO_IMAGE
    files = download_image(driver, thumbnail, article)
    return files if files else NO_IMAGE


# ══════════════════════════════════════════════════════════
# Главная функция
# ══════════════════════════════════════════════════════════

def main():
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    try:
        col_article  = find_col(ws, "Код")
        col_brand    = find_col(ws, "Бренд")
        col_supplier = find_col(ws, "Поставщик")
        col_images   = find_col(ws, "Изображени")
    except ValueError as e:
        print(f"Ошибка определения столбцов: {e}")
        return

    rows_to_process = [
        row for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=col_supplier).value
        and "автолига" in str(ws.cell(row=row, column=col_supplier).value).lower()
        and not ws.cell(row=row, column=col_images).value
    ]
    total = len(rows_to_process)
    print(f"Позиций с Автолигой без фото: {total}\n")

    if not total:
        print("Нет позиций для обработки.")
        return

    driver = make_driver()
    stats = {"ok": 0, "yellow": 0, "orange": 0}

    try:
        do_login(driver)

        for i, row_num in enumerate(rows_to_process, 1):
            article = ws.cell(row=row_num, column=col_article).value
            brand   = ws.cell(row=row_num, column=col_brand).value
            if not article:
                continue
            article = str(article).strip()
            brand   = str(brand or "").strip()

            print(f"\n[{i}/{total}] Артикул: {article}  Бренд: {brand}")
            result   = process_article(driver, article, brand)
            art_cell = ws.cell(row=row_num, column=col_article)

            if isinstance(result, list) and result:
                art_cell.fill = NO_FILL
                ws.cell(row=row_num, column=col_images).value = "; ".join(result)
                print(f"  ✓ {len(result)} фото: {', '.join(result)}")
                stats["ok"] += 1
            elif result == NO_IMAGE:
                art_cell.fill = ORANGE
                stats["orange"] += 1
            else:
                art_cell.fill = YELLOW
                stats["yellow"] += 1

            safe_save(wb, EXCEL_PATH)

    except KeyboardInterrupt:
        print("\nОстановлено пользователем")
    finally:
        driver.quit()
        safe_save(wb, EXCEL_PATH)

    print(f"\n{'='*50}")
    print(f"Готово!")
    print(f"  ✓ Скачано : {stats['ok']}")
    print(f"  ● Оранж.  : {stats['orange']}")
    print(f"  ● Жёлтых  : {stats['yellow']}")
    print(f"  Папка     : {IMAGES_DIR}")


if __name__ == "__main__":
    main()
