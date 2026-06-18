# /// script
# requires-python = ">=3.10"
# dependencies = ["requests", "openpyxl"]
# ///
"""
Скачивание изображений с mikado-parts.ru для позиций Топ-500 ВБ
Запуск: uv run mikado_web_images.py
"""

import re
import sys
import time
from hashlib import md5
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding="utf-8")

# ──────────────────────────────────────────────────────────
EXCEL_PATH  = r"C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ.xlsx"
IMAGES_DIR  = Path(r"C:\Users\Admin\Desktop\Топ-500 ВБ\Изображения Микадо")

BASE_URL    = "https://mikado-parts.ru/office"
LOGIN_URL   = f"{BASE_URL}/SECURE.asp"
PRODUCT_URL = f"{BASE_URL}/galleyp.asp"
AJAX_URL    = f"{BASE_URL}/pp0.asp"

MIKADO_CODE = "35275"
MIKADO_PASS = os.environ.get("MIKADO_PASS", "")

DELAY = 1.5
# ──────────────────────────────────────────────────────────

RED     = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
NO_FILL = PatternFill(fill_type="none")

# Иконки и служебные картинки — исключаем из поиска
_SKIP_SRCS = ("icon", "logo", "blank", "arrow", "button", "pixel", "spacer", "1x1")


def _abs(href: str) -> str:
    """Преобразует относительный URL в абсолютный."""
    if href.startswith("http"):
        return href
    if href.startswith("/"):
        return "https://mikado-parts.ru" + href
    return f"{BASE_URL}/{href}"


# ══════════════════════════════════════════════════════════
# Авторизация
# ══════════════════════════════════════════════════════════

def do_login() -> requests.Session:
    session = requests.Session()
    session.headers["User-Agent"] = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"
    )
    resp = session.post(
        LOGIN_URL,
        data={"CODE": MIKADO_CODE, "PASSWORD": MIKADO_PASS, "INSERT": "OK"},
        timeout=20,
    )
    resp.raise_for_status()
    page = resp.content.decode("windows-1251", errors="replace")
    if "Обслуживание клиентов" in page or "Продолжить" in page:
        print(f"  Login OK (code: {MIKADO_CODE})")
    else:
        print("  Login FAILED — check credentials")
    return session


# ══════════════════════════════════════════════════════════
# Получение ARTID и изображений со страницы товара
# ══════════════════════════════════════════════════════════

def get_product_info(session: requests.Session, code: str) -> tuple[str | None, list[str]]:
    """
    Загружает galleyp.asp?code={code}.
    Возвращает (artid, [img_urls]).

    img_urls собираются двумя способами:
      1. href-ссылки, ведущие на img.asp / bigimg (полноразмерные);
      2. src миниатюр <img src="wi/img.asp..."> — запасной вариант.
    """
    resp = session.get(PRODUCT_URL, params={"code": code}, timeout=20)
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")

    # ARTID
    m = re.search(r"ARTID=(\d+)", html)
    artid = m.group(1) if m else None

    img_urls: list[str] = []
    seen: set[str] = set()

    def _add(u: str) -> None:
        u = _abs(u)
        if u not in seen:
            seen.add(u)
            img_urls.append(u)

    # 1. <a href="...bigimg..."> или <a href="...img.asp..."> — полноразмерные ссылки
    for href in re.findall(
        r'href=["\']([^"\']*(?:bigimg|img\.asp)[^"\']*)["\']', html, re.I
    ):
        _add(href)

    # 2. <img src="wi/img.asp..."> — миниатюры прямо на странице
    for src in re.findall(
        r'<img[^>]+src=["\']([^"\']*wi/img\.asp[^"\']*)["\']', html, re.I
    ):
        _add(src)

    return artid, img_urls


# ══════════════════════════════════════════════════════════
# Получение изображений через pp0.asp (галерея)
# ══════════════════════════════════════════════════════════

def get_pp0_urls(session: requests.Session, code: str, artid: str) -> list[str]:
    """Запрашивает pp0.asp?MODE=PIC, возвращает список URL изображений."""
    resp = session.get(
        AJAX_URL,
        params={"MODE": "PIC", "CODE": code, "ARTID": artid},
        timeout=20,
    )
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")
    return [
        _abs(src)
        for src in re.findall(
            r'<img[^>]+src=["\']?(wi/img\.asp[^"\'> \t]+)', html, re.I
        )
    ]


# ══════════════════════════════════════════════════════════
# Скачивание одного изображения
# ══════════════════════════════════════════════════════════

def fetch_image_bytes(
    session: requests.Session, url: str
) -> tuple[bytes, str] | tuple[None, None]:
    """
    Скачивает изображение по URL.

    img.asp на Микадо отдаёт байты с Content-Type: text/html (или без него).
    Браузер показывает страницу как HTML-документ, но при "Сохранить как .jpg"
    файл оказывается валидным JPEG — потому что байты реально являются JPEG.

    Логика:
      1. Проверяем magic-bytes → JPEG / PNG → сохраняем сразу.
      2. Если magic-bytes нет — сервер вернул HTML-обёртку.
         Ищем в ней тег <img src="..."> и рекурсивно скачиваем оттуда.
      3. Если и там ничего — возвращаем None.
    """
    try:
        r = session.get(url, timeout=20)
        r.raise_for_status()
        content = r.content
        if not content:
            return None, None

        # JPEG
        if content[:2] == b"\xff\xd8":
            return content, ".jpg"
        # PNG
        if content[:8] == b"\x89PNG\r\n\x1a\n":
            return content, ".png"

        # HTML-обёртка — ищем вложенные изображения
        try:
            page = content.decode("windows-1251", errors="replace")
        except Exception:
            page = content.decode("latin-1", errors="replace")

        candidates: list[str] = []

        for src in re.findall(r'<img[^>]+src=["\']([^"\']+)["\']', page, re.I):
            if any(x in src.lower() for x in _SKIP_SRCS):
                continue
            candidates.append(_abs(src))

        # Иногда полноразмерный URL прячется в <a href="...jpg">
        for href in re.findall(
            r'href=["\']([^"\']+\.(?:jpg|jpeg|png))["\']', page, re.I
        ):
            candidates.append(_abs(href))

        for cand in candidates:
            if cand == url:
                continue
            try:
                r2 = session.get(cand, timeout=20)
                r2.raise_for_status()
                c2 = r2.content
                if c2[:2] == b"\xff\xd8":
                    return c2, ".jpg"
                if c2[:8] == b"\x89PNG\r\n\x1a\n":
                    return c2, ".png"
            except Exception:
                continue

        return None, None

    except Exception:
        return None, None


# ══════════════════════════════════════════════════════════
# Скачивание всех изображений товара
# ══════════════════════════════════════════════════════════

def download_images(
    session: requests.Session,
    code: str,
    artid: str,
    page_img_urls: list[str],
) -> list[str]:
    """
    Объединяет URL из pp0.asp (галерея) и со страницы товара,
    скачивает все уникальные изображения.
    Возвращает список имён сохранённых файлов.
    """
    all_urls: list[str] = []
    seen_urls: set[str] = set()

    def _enqueue(u: str) -> None:
        if u not in seen_urls:
            seen_urls.add(u)
            all_urls.append(u)

    # Сначала pp0.asp (обычно самый полный список)
    try:
        for u in get_pp0_urls(session, code, artid):
            _enqueue(u)
    except Exception:
        pass

    # Затем URL прямо со страницы товара (запасной вариант)
    for u in page_img_urls:
        _enqueue(u)

    saved: list[str] = []
    seen_hashes: set[str] = set()
    idx = 1

    for url in all_urls:
        content, ext = fetch_image_bytes(session, url)
        if content is None:
            continue
        digest = md5(content).hexdigest()
        if digest in seen_hashes:
            continue
        seen_hashes.add(digest)
        filename = f"{code}_{idx}{ext}"
        (IMAGES_DIR / filename).write_bytes(content)
        saved.append(filename)
        idx += 1

    return saved


# ══════════════════════════════════════════════════════════
# Excel helpers
# ══════════════════════════════════════════════════════════

def safe_save(wb, path: str) -> None:
    """Сохраняет Excel, при PermissionError ждёт и повторяет (файл открыт в Excel)."""
    for attempt in range(1, 11):
        try:
            wb.save(path)
            return
        except PermissionError:
            if attempt == 1:
                print(f"\n  [!] Файл занят — закройте Excel и нажмите Enter (попытка {attempt}/10)...", end="", flush=True)
                input()
            else:
                print(f"  [!] Ещё занят, жду 5 сек (попытка {attempt}/10)...", flush=True)
                time.sleep(5)
    raise PermissionError(f"Не удалось сохранить {path} после 10 попыток")


def find_col(ws, keyword: str) -> int:
    for cell in ws[1]:
        if cell.value and keyword.lower() in str(cell.value).lower():
            return cell.column
    raise ValueError(f"Column '{keyword}' not found")


# ══════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════

def main():
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    col_article  = find_col(ws, "Код")
    col_supplier = find_col(ws, "Поставщик")
    col_images   = find_col(ws, "Изображени")

    rows_mikado = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=col_supplier).value
        and "микадо" in str(ws.cell(row=r, column=col_supplier).value).lower()
    ]
    total = len(rows_mikado)
    print(f"Mikado rows: {total}")

    print("Logging in to mikado-parts.ru...")
    session = do_login()
    print()

    stats = {"ok": 0, "no_artid": 0, "no_img": 0}

    for i, row_num in enumerate(rows_mikado, 1):
        article = ws.cell(row=row_num, column=col_article).value
        if not article:
            continue
        article = str(article).strip()

        # Пропускаем если изображение уже есть
        existing = ws.cell(row=row_num, column=col_images).value
        if existing:
            print(f"[{i}/{total}] {article}: skip (already has image)")
            continue

        print(f"[{i}/{total}] {article}", end=" ... ", flush=True)
        art_cell = ws.cell(row=row_num, column=col_article)

        try:
            artid, page_img_urls = get_product_info(session, article)

            if not artid:
                print("ARTID not found -> red")
                art_cell.fill = RED
                stats["no_artid"] += 1
                safe_save(wb, EXCEL_PATH)
                time.sleep(DELAY)
                continue

            time.sleep(0.4)
            images = download_images(session, article, artid, page_img_urls)

            if not images:
                print("no images -> red")
                art_cell.fill = RED
                stats["no_img"] += 1
            else:
                ws.cell(row=row_num, column=col_images).value = "; ".join(images)
                art_cell.fill = NO_FILL
                print(f"{len(images)} img: {', '.join(images)}")
                stats["ok"] += 1

        except Exception as e:
            print(f"ERROR: {e} -> red")
            art_cell.fill = RED
            stats["no_artid"] += 1

        safe_save(wb, EXCEL_PATH)
        time.sleep(DELAY)

    safe_save(wb, EXCEL_PATH)
    print()
    print("=" * 50)
    print(f"Done!")
    print(f"  Downloaded : {stats['ok']}")
    print(f"  No images  : {stats['no_img']}")
    print(f"  Not found  : {stats['no_artid']}")
    print(f"  Saved to   : {IMAGES_DIR}")


if __name__ == "__main__":
    main()
