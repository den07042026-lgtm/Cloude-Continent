# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl", "Pillow"]
# ///
"""
Удаляет изображения-заглушки (логотип ОК) из папки Автолига,
чистит ссылки в Excel, ставит жёлтый цвет на ячейки без фото.
Запуск: uv run cleanup_avtoliga.py
"""

import sys
import time
from hashlib import md5
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding="utf-8")

EXCEL_PATH  = r"C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ_new.xlsx"
IMAGES_DIR  = Path(r"C:\Users\Admin\Desktop\Топ-500 ВБ\Изображения Автолига")
REF_IMAGE   = IMAGES_DIR / "77295_1.jpg"

YELLOW  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NO_FILL = PatternFill(fill_type="none")

IMG_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}


def safe_save(wb, path: str) -> None:
    p   = Path(path)
    tmp = p.with_suffix(".tmp.xlsx")
    for attempt in range(1, 11):
        try:
            wb.save(tmp)
            tmp.replace(p)
            return
        except PermissionError:
            if tmp.exists():
                try: tmp.unlink()
                except Exception: pass
            if attempt == 1:
                print("\n  [!] Файл занят — закройте Excel и нажмите Enter...",
                      end="", flush=True)
                input()
            else:
                print(f"  [!] Ещё занят, жду 5 сек (попытка {attempt}/10)...", flush=True)
                time.sleep(5)
        except Exception as e:
            if tmp.exists():
                try: tmp.unlink()
                except Exception: pass
            raise e
    raise PermissionError(f"Не удалось сохранить {path}")


def find_col(ws, keyword: str) -> int:
    for cell in ws[1]:
        if cell.value and keyword.lower() in str(cell.value).lower():
            return cell.column
    raise ValueError(f"Столбец с '{keyword}' не найден")


def main():
    ref_hash = md5(REF_IMAGE.read_bytes()).hexdigest()
    print(f"Хеш заглушки: {ref_hash}")

    # Найти все файлы-заглушки
    bad_files: set[str] = set()
    for f in IMAGES_DIR.iterdir():
        if f.is_file() and f.suffix.lower() in IMG_EXTS:
            if md5(f.read_bytes()).hexdigest() == ref_hash:
                bad_files.add(f.name)

    print(f"Найдено заглушек: {len(bad_files)}")
    for name in sorted(bad_files):
        print(f"  {name}")

    if not bad_files:
        print("Нечего удалять.")
        return

    # Удалить файлы
    for name in bad_files:
        (IMAGES_DIR / name).unlink()
    print(f"\nУдалено файлов: {len(bad_files)}")

    # Обновить Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    col_article = find_col(ws, "Код")
    col_images  = find_col(ws, "Изображени")

    updated = 0
    for row in range(2, ws.max_row + 1):
        cell_img = ws.cell(row=row, column=col_images)
        val = cell_img.value
        if not val:
            continue

        parts    = [p.strip() for p in str(val).split(";")]
        clean    = [p for p in parts if p and p not in bad_files]
        removed  = len(parts) - len(clean)

        if removed == 0:
            continue

        art_cell = ws.cell(row=row, column=col_article)

        if clean:
            cell_img.value = "; ".join(clean)
            print(f"  Строка {row}: убрано {removed} заглушка(ок), осталось {len(clean)} фото")
        else:
            cell_img.value = None
            art_cell.fill  = YELLOW
            print(f"  Строка {row}: все фото были заглушками → жёлтый")

        updated += 1

    safe_save(wb, EXCEL_PATH)
    print(f"\nОбновлено строк в Excel: {updated}")
    print("Готово!")


if __name__ == "__main__":
    main()
