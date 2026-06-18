import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

folder = 'C:/Users/Admin/Desktop/Топ ВБ 1306'
wb = openpyxl.load_workbook(f'{folder}/Топ-500 ВБ 1306.xlsx', read_only=True)
ws = wb.active
print('=== ЗАГОЛОВКИ ===')
for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
    print(f'  {i}: {repr(cell.value)}')
print()
print('=== СТРОКА 2 (пример полной строки) ===')
for i, cell in enumerate(next(ws.iter_rows(min_row=2, max_row=2)), 1):
    val = str(cell.value)[:60] if cell.value else None
    print(f'  {i}: {repr(val)}')
print()
print('=== СТРОКИ ТОЛЬКО С АРТИКУЛОМ (первые 5) ===')
count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    kod = row[0].value
    rest = [row[c].value for c in range(1, 15)]
    if kod and all(v is None for v in rest):
        print(f'  Строка {row_idx}: {repr(kod)}')
        count += 1
        if count >= 5:
            break
wb.close()
