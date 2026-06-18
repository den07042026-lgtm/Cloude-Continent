"""
Обрезает часть артикула до первого дефиса (включительно) в столбцах:
  - Код (Mikado)
  - Альтернативные артикулы товара

Обрабатывает все .xlsx файлы в папке INPUT_DIR и сохраняет их на месте.
"""

import sys
import os
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

INPUT_DIR = r"C:\Users\Admin\Desktop\На сортировку 21.04(3)"

TARGET_COLS = {"Код (Mikado)", "Альтернативные артикулы товара"}


def trim_article(value: str) -> str:
    """Убирает всё до первого дефиса включительно."""
    if not isinstance(value, str):
        return value
    idx = value.find('-')
    if idx == -1:
        return value  # дефиса нет — оставляем как есть
    return value[idx + 1:]


def process_cell(value, col_name: str):
    """Обрабатывает значение ячейки с учётом формата столбца."""
    if not isinstance(value, str):
        return value

    if col_name == "Альтернативные артикулы товара":
        # Несколько артикулов через '; '
        parts = value.split('; ')
        parts = [trim_article(p) for p in parts]
        return '; '.join(parts)
    else:
        return trim_article(value)


def process_file(filepath: str):
    wb = openpyxl.load_workbook(filepath)
    changed = False

    for ws in wb.worksheets:
        # Определяем индексы нужных столбцов по заголовку первой строки
        col_map = {}  # {col_index: col_name}
        for cell in ws[1]:
            if cell.value in TARGET_COLS:
                col_map[cell.column] = cell.value

        if not col_map:
            continue  # нужных столбцов в листе нет

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in col_map:
                    new_val = process_cell(cell.value, col_map[cell.column])
                    if new_val != cell.value:
                        cell.value = new_val
                        changed = True

    if changed:
        wb.save(filepath)
        return True
    return False


def main():
    files = [f for f in os.listdir(INPUT_DIR) if f.lower().endswith('.xlsx')]
    total = len(files)
    updated = 0

    for i, filename in enumerate(files, 1):
        filepath = os.path.join(INPUT_DIR, filename)
        try:
            result = process_file(filepath)
            status = "обновлён" if result else "без изменений"
            print(f"[{i}/{total}] {filename} — {status}")
            if result:
                updated += 1
        except Exception as e:
            print(f"[{i}/{total}] ОШИБКА {filename}: {e}")

    print(f"\nГотово. Обновлено файлов: {updated} из {total}.")


if __name__ == "__main__":
    main()
