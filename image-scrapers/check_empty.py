import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:/Users/Admin/Desktop/Топ ВБ 1306/Топ-500 ВБ 1306.xlsx', read_only=True)
ws = wb.active

print(f"max_column: {ws.max_column}")
print("\nСтроки 21-32 (детально):")
for row_idx, row in enumerate(ws.iter_rows(min_row=21, max_row=32), start=21):
    vals = [repr(cell.value) for cell in row]
    kod = row[0].value
    rest = [row[c].value for c in range(1, ws.max_column)]
    empty = all(v is None for v in rest)
    empty2 = all(v is None or str(v).strip() == "" for v in rest)
    print(f"  Строка {row_idx}: kod={repr(kod)}, all_none={empty}, all_empty={empty2}")
    print(f"    Значения: {vals}")
wb.close()
