"""
wb_deficit_analyzer.py
════════════════════════════════════════════════════════════════════════════
Анализ дефицитных позиций WB + топ-продавцы + проверка нашего каталога.

Фаза 1 — предметы (WB subjects):
  Скоринг предмета: OOS% × log(продажи) × (1/√продавцов)

Фаза 2 — дефицитные товары:
  WB поиск по топ-предметам → OOS из MPStats
  Листы: "Дефицит — в каталоге" | "Дефицит — нет в каталоге"

Фаза 3 — топ-продавцы в наших нишах:
  Только наши бренды с OOS < 20% → "Топ продавцы"

Фаза 4 — наш каталог на WB:
  Для каждого из топ-N брендов нашего прайса — ищет товары на WB,
  проверяет OOS, скорит. Лист: "Наш каталог — топ WB"

Запуск:
  uv run --with requests,openpyxl,xlrd scripts/wb_deficit_analyzer.py
  uv run --with requests,openpyxl,xlrd scripts/wb_deficit_analyzer.py --with-products
  uv run --with requests,openpyxl,xlrd scripts/wb_deficit_analyzer.py --with-products --top-catalog-brands 100
"""

import sys
import math
import time
import json
import logging
import argparse
import sqlite3
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, str(Path(__file__).parent))
sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Установи: uv run --with requests,openpyxl,xlrd scripts/wb_deficit_analyzer.py")
    sys.exit(1)

# ─── Константы ────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.parent
ENV_FILE   = BASE_DIR / ".env"
LOG_FILE   = BASE_DIR / "logs" / "wb_deficit_analyzer.log"
OUTPUT_DIR = BASE_DIR / "data" / "analytics"
CACHE_DIR  = BASE_DIR / "data" / "analytics" / "cache"
DB_PATH    = BASE_DIR / "data" / "analytics" / "wb_index.db"

MPSTATS_BASE = "https://mpstats.io/api"

AUTO_KEYWORDS = [
    "запчаст", "колодк", "амортизатор", "фильтр", "свеч",
    "тормоз", "подвеск", "рулев", "сцеплен", "ремен", "ремни",
    "шаровая", "шаровой", "стойк", "глушитель", "выхлоп",
    "радиатор", "насос", "помп", "генератор", "стартер",
    "катушк", "датчик", "термостат", "сальник", "подшипник",
    "прокладк", "поршн", "кольц", "клапан", "распредвал",
    "коленвал", "грм", "топлив", "форсунк", "инжектор",
    "тяг", "наконечник", "рычаг", "пружин", "зажигани",
    "аккумулятор", "охлаждени", "тосол", "антифриз",
    "трансмисс", "коробк", "привод", "шрус", "ступиц",
]

AUTO_SUBJECT_SECTIONS = {
    "автозапчаст",
    "автоаксессуар",
    "масла и техн",
    "шины и диски",
    "автокосметик",
    "автосвет",
    "аккумулятор",
    "автомобильн",
    "автоэлектрик",
    "автохими",
}

SUBJECT_KEYWORDS = AUTO_KEYWORDS + [
    "автозапчаст", "автодеталь", "автомобиль",
    "двигатель", "кузов", "электрик", "зажиган", "охлажден",
]

MIN_MONTHLY_SALES    = 50     # минимум продаж предмета (дефицит)
MIN_OOS_PCT          = 5.0    # минимальный OOS% (дефицит)
MIN_SALES_TOP_SELLER  = 100   # минимум продаж предмета для фазы 3
MAX_PRODUCT_OOS_TOP   = 20.0  # максимум OOS% конкретного товара в "Топ продавцы"
REQUEST_DELAY        = 0.8
WB_FBS_COMMISSION    = 0.17

# ─── Логирование ──────────────────────────────────────────────────────────────
for d in (LOG_FILE.parent, OUTPUT_DIR, CACHE_DIR):
    d.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ─── .env ─────────────────────────────────────────────────────────────────────
def load_env() -> dict:
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


# ─── MPStats API ──────────────────────────────────────────────────────────────
class MPStatsClient:
    def __init__(self, token: str):
        self.token   = token
        self.session = requests.Session()
        self.session.headers.update({
            "X-Mpstats-TOKEN": token,
            "Content-Type": "application/json",
        })

    def _get(self, path: str, params: dict | None = None, retries: int = 3) -> dict | list | None:
        url = f"{MPSTATS_BASE}{path}"
        for attempt in range(retries):
            try:
                r = self.session.get(url, params=params, timeout=30)
                if r.status_code == 429:
                    wait = 10 * (attempt + 1)
                    log.warning(f"Rate limit, ждём {wait}с...")
                    time.sleep(wait)
                    continue
                if r.status_code == 200:
                    return r.json()
                log.warning(f"MPStats {path}: {r.status_code} {r.text[:150]}")
                return None
            except Exception as e:
                log.warning(f"MPStats {path} попытка {attempt+1}: {e}")
                time.sleep(2)
        return None

    def get_categories(self) -> list[dict]:
        cache = CACHE_DIR / "wb_categories.json"
        if cache.exists() and (datetime.now() - datetime.fromtimestamp(cache.stat().st_mtime)).days < 1:
            return json.loads(cache.read_text(encoding="utf-8"))
        result = self._get("/wb/get/categories") or []
        cache.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")
        return result

    def get_subjects(self, path: str, d1: str, d2: str) -> list[dict]:
        cache_key = path.replace("/", "_").replace(" ", "-")[:80]
        cache = CACHE_DIR / f"subj_{cache_key}_{d1}.json"
        if cache.exists():
            return json.loads(cache.read_text(encoding="utf-8"))
        result = self._get("/wb/get/category/items", {
            "path": path, "d1": d1, "d2": d2,
            "startRow": 0, "endRow": 1000,
        })
        if isinstance(result, list):
            subjects = [x for x in result if isinstance(x, dict) and x.get("id")]
        elif isinstance(result, dict):
            inner = result.get("data") or result.get("items") or result.get("value") or []
            subjects = [x for x in inner if isinstance(x, dict) and x.get("id")]
        else:
            subjects = []
        cache.write_text(json.dumps(subjects, ensure_ascii=False), encoding="utf-8")
        time.sleep(REQUEST_DELAY)
        return subjects

    def get_item_daily(self, nm_id: int, d1: str, d2: str) -> dict | None:
        cache = CACHE_DIR / f"item_{nm_id}_{d1}.json"
        if cache.exists():
            return json.loads(cache.read_text(encoding="utf-8"))
        result = self._get(f"/wb/get/item/{nm_id}/by_category", {"d1": d1, "d2": d2})
        if result:
            cache.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")
        time.sleep(REQUEST_DELAY)
        return result


# ─── WB публичный поиск ───────────────────────────────────────────────────────
def search_wb_products(query: str, subject_id: int = 0, limit: int = 50) -> list[dict]:
    cache_key = f"{subject_id}_{query.replace(' ', '_').replace('/', '-')[:50]}"
    cache = CACHE_DIR / f"wb_search_{cache_key}.json"
    if cache.exists():
        cached = json.loads(cache.read_text(encoding="utf-8"))
        if cached:
            return cached

    search_query = query.split(" / ")[-1] if " / " in query else query

    def _parse(data: dict) -> list[dict]:
        raw = data.get("products") or (data.get("data") or {}).get("products") or []
        out = []
        for p in raw:
            nm_id = p.get("id")
            if not nm_id:
                continue
            if subject_id and p.get("subjectId") != subject_id:
                continue
            out.append({
                "nm_id":     nm_id,
                "name":      p.get("name", ""),
                "brand":     p.get("brand", ""),
                "price_rub": 0,
                "rating":    p.get("rating", 0),
                "feedbacks": p.get("feedbacks", 0),
                "supplier":  p.get("supplier", ""),
            })
            if len(out) >= limit:
                break
        return out

    try:
        r = requests.get(
            "https://search.wb.ru/exactmatch/ru/common/v5/search",
            params={
                "query": search_query, "resultset": "catalog",
                "limit": min(limit * 3, 150), "sort": "popular",
                "page": 1, "dest": -1257786, "appType": 1,
                "curr": "rub", "spp": 27,
            },
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Accept-Language": "ru-RU,ru;q=0.9",
            },
            timeout=15,
        )
        products = []
        if r.status_code == 200:
            products = _parse(r.json())
        elif r.status_code == 429:
            log.warning("WB поиск rate-limit, ждём 30с...")
            time.sleep(30)
            r2 = requests.get(r.url, headers=r.request.headers, timeout=15)
            if r2.status_code == 200:
                products = _parse(r2.json())

        if products:
            cache.write_text(json.dumps(products, ensure_ascii=False), encoding="utf-8")
        time.sleep(2.0)
        return products
    except Exception as e:
        log.warning(f"WB поиск '{search_query}': {e}")
        return []


# ─── OOS и скоринг товара ─────────────────────────────────────────────────────
def calc_oos(daily: dict) -> dict:
    balance       = daily.get("balance", [])
    sales         = daily.get("sales", [])
    final_price   = daily.get("final_price", [])
    period        = len(balance) or 1
    oos_days      = sum(1 for b in balance if b == 0)
    monthly_sales = sum(sales)
    prices        = [p for p in final_price if p and p > 0]
    avg_price     = sum(prices) / len(prices) if prices else 0
    return {
        "oos_days":      oos_days,
        "oos_ratio":     oos_days / period,
        "monthly_sales": monthly_sales,
        "avg_price_mp":  round(avg_price, 0),
        "period":        period,
    }


def score_product(oos_pct: float, monthly_sales: int) -> float:
    """OOS-давление × масштаб спроса. Выше = приоритетнее для входа."""
    return round(oos_pct * math.log1p(monthly_sales), 2)


# ─── WB ценовой калькулятор ───────────────────────────────────────────────────
_WB_COMMISSION  = 0.25    # Автозапчасти FBS, май 2026
_WB_TAX         = 0.06    # УСН 6%
_WB_RET_RATE    = 0.03    # доля возвратов
_WB_PACKAGING   = 30      # упаковка, ₽
_WB_DEL_BASE    = 50.6    # доставка FBS база ₽ (Волгоград)
_WB_DEL_LITER   = 15.4    # доставка FBS ₽/л
_WB_RET_BASE    = 136.0   # возврат через ПВЗ, база ₽
_WB_RET_LITER   = 14.0    # возврат через ПВЗ, ₽/л
_WB_DEFAULT_VOL = 3.0     # дефолт объём ≈ 20×15×10 см
_WB_TARGET_MIN  = 0.0     # безубыточность
_WB_TARGET_OPT  = 0.15    # целевая маржа 15%


def _wb_margin(purchase: float, sell: float, liters: float = _WB_DEFAULT_VOL) -> float:
    if sell <= 0:
        return 0.0
    commission = sell * _WB_COMMISSION
    delivery   = _WB_DEL_BASE + liters * _WB_DEL_LITER
    ret_cost   = _WB_RET_RATE * (_WB_RET_BASE + liters * _WB_RET_LITER)
    proceeds   = sell - commission - delivery
    tax        = max(0.0, proceeds) * _WB_TAX
    profit     = sell - purchase - commission - delivery - ret_cost - tax - _WB_PACKAGING
    return profit / sell


def _wb_find_price(purchase: float, liters: float, target: float) -> int | None:
    if _wb_margin(purchase, 500_000, liters) < target:
        return None
    lo, hi = 50, 500_000
    while lo < hi:
        mid = (lo + hi) // 2
        if _wb_margin(purchase, mid, liters) >= target - 1e-6:
            hi = mid
        else:
            lo = mid + 1
    return lo if _wb_margin(purchase, lo, liters) >= target - 1e-6 else None


def _get_purchase_prices_from_db(nm_ids: list[int]) -> dict[int, float]:
    """nm_id → минимальная закупочная цена из wb_matches."""
    if not DB_PATH.exists() or not nm_ids:
        return {}
    result: dict[int, float] = {}
    conn = _db_open()
    try:
        for i in range(0, len(nm_ids), 999):
            batch = nm_ids[i : i + 999]
            pl = ",".join("?" * len(batch))
            for r in conn.execute(
                f"SELECT nm_id, MIN(our_price) AS min_p "
                f"FROM wb_matches WHERE nm_id IN ({pl}) AND our_price > 0 "
                f"GROUP BY nm_id",
                batch,
            ).fetchall():
                if r["min_p"] and r["min_p"] > 0:
                    result[r["nm_id"]] = float(r["min_p"])
    finally:
        conn.close()
    return result


def _enrich_with_prices(products: list[dict]) -> None:
    """In-place: добавляет purchase_price, sell_min, sell_opt, margin_at_wb."""
    need_lookup = [
        p["nm_id"] for p in products
        if not (p.get("our_min_price") and p["our_min_price"] > 0)
    ]
    purchase_map = _get_purchase_prices_from_db(need_lookup) if need_lookup else {}

    for p in products:
        own = p.get("our_min_price") or 0
        purchase = float(own) if own and own > 0 else purchase_map.get(p["nm_id"], 0.0)

        if purchase <= 0:
            p["purchase_price"] = None
            p["sell_min"]       = None
            p["sell_opt"]       = None
            p["margin_at_wb"]   = None
            continue

        p["purchase_price"] = round(purchase)
        p["sell_min"]       = _wb_find_price(purchase, _WB_DEFAULT_VOL, _WB_TARGET_MIN)
        p["sell_opt"]       = _wb_find_price(purchase, _WB_DEFAULT_VOL, _WB_TARGET_OPT)
        wb_price = p.get("price_rub") or 0
        p["margin_at_wb"] = (
            round(_wb_margin(purchase, wb_price, _WB_DEFAULT_VOL) * 100, 1)
            if wb_price > 0 else None
        )


# ─── Базовый сборщик товаров ──────────────────────────────────────────────────
def _fetch_products_base(
    subjects:             list[dict],
    client:               MPStatsClient,
    our_brands:           set[str],
    d1:                   str,
    d2:                   str,
    label:                str,
    products_per_subject: int  = 30,
    filter_in_catalog:    bool = False,
) -> list[dict]:
    results:  list[dict] = []
    seen_nm:  set[int]   = set()
    total = len(subjects)

    for idx, subj in enumerate(subjects):
        subj_name  = subj["name"]
        subject_id = subj.get("subject_id", 0)
        log.info(f"  {label} {idx+1}/{total}: {subj_name}")

        products = search_wb_products(subj_name, subject_id=subject_id, limit=products_per_subject)
        if not products:
            log.warning(f"    Нет результатов WB (subjectId={subject_id}): {subj_name}")
            continue

        log.info(f"    Найдено {len(products)} товаров, запрашиваем OOS...")

        for prod in products:
            nm_id = prod.get("nm_id")
            if not nm_id or nm_id in seen_nm:
                continue
            seen_nm.add(nm_id)

            daily = client.get_item_daily(nm_id, d1, d2)
            if not daily:
                continue

            oos = calc_oos(daily)
            if oos["monthly_sales"] < 1:
                continue

            brand_lower = prod["brand"].strip().lower()
            in_catalog  = brand_lower in our_brands

            if filter_in_catalog and not in_catalog:
                continue

            oos_pct = round(oos["oos_ratio"] * 100, 1)
            results.append({
                "nm_id":         nm_id,
                "name":          prod["name"],
                "brand":         prod["brand"],
                "subject":       subj_name,
                "subject_score": subj.get("score", 0),
                "price_rub":     oos["avg_price_mp"] or prod["price_rub"],
                "rating":        prod["rating"],
                "feedbacks":     prod["feedbacks"],
                "supplier":      prod["supplier"],
                "oos_days":      oos["oos_days"],
                "oos_pct":       oos_pct,
                "monthly_sales": oos["monthly_sales"],
                "in_catalog":    in_catalog,
                "prod_score":    score_product(oos_pct, oos["monthly_sales"]),
                "wb_link":       f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
            })

    return results


# ─── Фаза 2: дефицитные товары ───────────────────────────────────────────────
def fetch_deficit_products(
    subjects:             list[dict],
    client:               MPStatsClient,
    our_brands:           set[str],
    d1:                   str,
    d2:                   str,
    top_subjects:         int = 20,
    products_per_subject: int = 30,
) -> list[dict]:
    results = _fetch_products_base(
        subjects[:top_subjects], client, our_brands, d1, d2,
        label="Предмет", products_per_subject=products_per_subject,
    )
    results.sort(key=lambda x: x["prod_score"], reverse=True)
    in_cat = sum(1 for r in results if r["in_catalog"])
    log.info(f"Дефицитных товаров: {len(results)}, из них в каталоге: {in_cat}")
    return results


# ─── Фаза 3: топ-продавцы в наших нишах ──────────────────────────────────────
def fetch_top_seller_products(
    subjects:             list[dict],
    client:               MPStatsClient,
    our_brands:           set[str],
    d1:                   str,
    d2:                   str,
    top_subjects:         int = 20,
    products_per_subject: int = 30,
) -> list[dict]:
    raw = _fetch_products_base(
        subjects[:top_subjects], client, our_brands, d1, d2,
        label="Топ-предмет", products_per_subject=products_per_subject,
        filter_in_catalog=True,
    )
    # Только товары с низким индивидуальным OOS — реальные стабильные продавцы
    results = [r for r in raw if r["oos_pct"] < MAX_PRODUCT_OOS_TOP]
    results.sort(key=lambda x: x["monthly_sales"], reverse=True)
    log.info(f"Топ-продавцов в каталоге (OOS <{MAX_PRODUCT_OOS_TOP}%): {len(results)}")
    return results


# ─── Каталог поставщиков ──────────────────────────────────────────────────────
def load_supplier_brands() -> set[str]:
    brands: set[str] = set()

    try:
        import io, openpyxl as ox
        price_url = (
            "https://mikado-parts.ru/api/Price/GetPriceExcel"
            "?StockId=34&Key=BBE2E029-54CF-4D9E-9FAC-9FE25E85B300"
        )
        resp = requests.get(price_url, timeout=60)
        if resp.ok and resp.content[:2] == b"PK":
            wb = ox.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            hdr = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]
            bi = next((i for i, h in enumerate(hdr) if h == "brandname"), None)
            for row in rows:
                brand = str(row[bi]).strip().lower() if bi is not None and len(row) > bi and row[bi] else ""
                if brand:
                    brands.add(brand)
            wb.close()
            log.info(f"Mikado: {len(brands)} брендов")
    except Exception as e:
        log.warning(f"Mikado каталог: {e}")

    try:
        from autoliga_loader import load_autoliga
        al = load_autoliga()
        before = len(brands)
        for v in al.values():
            b = v.get("brand", "").strip().lower()
            if b:
                brands.add(b)
        log.info(f"Автолига добавила {len(brands) - before} новых брендов")
    except Exception as e:
        log.warning(f"Автолига каталог: {e}")

    log.info(f"Итого брендов в каталоге: {len(brands)}")
    return brands


def load_catalog_brand_info() -> dict[str, dict]:
    """
    Возвращает {brand_lower: {count, min_price, display}} из Mikado + Автолига.
    Используется в Фазе 4 для поиска по брендам нашего каталога.
    """
    info: dict[str, dict] = {}

    def _add(brand_raw: str, price: float) -> None:
        bl = brand_raw.strip().lower()
        if not bl or bl in ("none", "nan", ""):
            return
        if bl not in info:
            info[bl] = {"count": 0, "min_price": 999_999, "display": brand_raw.strip()}
        info[bl]["count"] += 1
        if 0 < price < info[bl]["min_price"]:
            info[bl]["min_price"] = price

    try:
        import io, openpyxl as ox
        resp = requests.get(
            "https://mikado-parts.ru/api/Price/GetPriceExcel"
            "?StockId=34&Key=BBE2E029-54CF-4D9E-9FAC-9FE25E85B300",
            timeout=60,
        )
        if resp.ok and resp.content[:2] == b"PK":
            wb_xls = ox.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
            ws = wb_xls.active
            rows = ws.iter_rows(values_only=True)
            hdr = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]
            bi = next((i for i, h in enumerate(hdr) if h == "brandname"), None)
            pi = next((i for i, h in enumerate(hdr) if "price" in h and "rub" not in h), None) \
              or next((i for i, h in enumerate(hdr) if "price" in h), None)
            for row in rows:
                b = str(row[bi]).strip() if bi is not None and len(row) > bi and row[bi] else ""
                try:
                    p = float(row[pi]) if pi is not None and len(row) > pi and row[pi] else 0
                except (ValueError, TypeError):
                    p = 0
                _add(b, p)
            wb_xls.close()
            log.info(f"Mikado catalog_info: {len(info)} брендов")
    except Exception as e:
        log.warning(f"Mikado catalog_info: {e}")

    try:
        from autoliga_loader import load_autoliga
        al = load_autoliga()
        before = len(info)
        for v in al.values():
            b = v.get("brand", "")
            p = float(v.get("price") or v.get("price_rub") or v.get("cost") or 0)
            _add(b, p)
        log.info(f"Автолига catalog_info: добавлено {len(info) - before} брендов, всего {len(info)}")
    except Exception as e:
        log.warning(f"Автолига catalog_info: {e}")

    for b in info:
        if info[b]["min_price"] >= 999_999:
            info[b]["min_price"] = 0

    return info


# ─── Фаза 4: наш каталог → WB поиск ─────────────────────────────────────────
def fetch_catalog_products(
    brand_info:        dict[str, dict],
    client:            MPStatsClient,
    d1:                str,
    d2:                str,
    top_brands:        int = 100,
    products_per_brand: int = 15,
    top_results:       int = 1000,
) -> list[dict]:
    """
    Для топ-N брендов нашего каталога (по количеству позиций) ищет их товары
    на WB, получает OOS из MPStats, скорит и возвращает топ-top_results.
    """
    sorted_brands = sorted(brand_info.items(), key=lambda x: x[1]["count"], reverse=True)
    selected = sorted_brands[:top_brands]
    total = len(selected)

    results:  list[dict] = []
    seen_nm:  set[int]   = set()

    for idx, (brand_lower, binfo) in enumerate(selected):
        display = binfo["display"]
        log.info(f"  Бренд {idx+1}/{total}: {display} ({binfo['count']} поз.)")

        products = search_wb_products(display, subject_id=0, limit=products_per_brand)
        if not products:
            log.warning(f"    Нет результатов WB: {display}")
            continue

        # Оставляем только товары нашего бренда
        matched = [
            p for p in products
            if brand_lower in p["brand"].strip().lower()
            or p["brand"].strip().lower() in brand_lower
        ]
        if not matched:
            matched = products  # транслитерация/сокращение — берём всё

        log.info(f"    {len(matched)} товаров → OOS...")

        for prod in matched[:products_per_brand]:
            nm_id = prod.get("nm_id")
            if not nm_id or nm_id in seen_nm:
                continue
            seen_nm.add(nm_id)

            daily = client.get_item_daily(nm_id, d1, d2)
            if not daily:
                continue

            oos = calc_oos(daily)
            if oos["monthly_sales"] < 1:
                continue

            oos_pct   = round(oos["oos_ratio"] * 100, 1)
            wb_price  = oos["avg_price_mp"] or prod["price_rub"]
            our_price = binfo["min_price"]

            results.append({
                "nm_id":         nm_id,
                "name":          prod["name"],
                "brand":         prod["brand"],
                "catalog_count": binfo["count"],
                "our_min_price": our_price,
                "price_rub":     wb_price,
                "rating":        prod["rating"],
                "feedbacks":     prod["feedbacks"],
                "supplier":      prod["supplier"],
                "oos_days":      oos["oos_days"],
                "oos_pct":       oos_pct,
                "monthly_sales": oos["monthly_sales"],
                "prod_score":    score_product(oos_pct, oos["monthly_sales"]),
                "wb_link":       f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
            })

    results.sort(key=lambda x: x["prod_score"], reverse=True)
    final = results[:top_results]
    log.info(f"Каталог-топ: найдено {len(results)}, в отчёте топ-{len(final)}")
    return final


def subject_in_catalog(subject_name: str, our_brands: set[str]) -> bool:
    name_lower = subject_name.lower()
    return any(kw in name_lower for kw in SUBJECT_KEYWORDS)


# ─── Скоринг предмета ────────────────────────────────────────────────────────
def score_subject(subj: dict) -> float:
    sales   = subj.get("sales", 0)
    sellers = max(1, subj.get("sellers_with_sells", subj.get("sellers", 1)))
    oos_pct = subj.get("lost_profit_percent", 0.0)
    if sales < MIN_MONTHLY_SALES or oos_pct < MIN_OOS_PCT:
        return 0.0
    return (oos_pct / 100.0) * math.log1p(sales) * (1.0 / math.sqrt(sellers))


# ─── Основной анализ ──────────────────────────────────────────────────────────
def run_analysis(
    token:       str,
    days:        int  = 30,
    top:         int  = 200,
    list_only:   bool = False,
    path_filter: str  = "Автотовары/Запчасти",
    max_cats:    int  = 0,
) -> tuple[list[dict], list[dict]]:
    """
    Возвращает (deficit_subjects, top_seller_subjects).
    deficit_subjects    — топ-N предметов с высоким OOS.
    top_seller_subjects — предметы с высокими продажами и низким OOS (для фазы 3).
    """
    client = MPStatsClient(token)
    d2 = datetime.now().strftime("%Y-%m-%d")
    d1 = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    log.info(f"Период анализа: {d1} → {d2}  ({days} дней)")

    log.info("Загружаем категории WB...")
    all_cats = client.get_categories()
    log.info(f"Всего категорий: {len(all_cats)}")

    auto_cats = [
        c for c in all_cats
        if isinstance(c, dict) and any(kw in c.get("path", "").lower() for kw in AUTO_KEYWORDS)
    ]
    if path_filter:
        auto_cats = [
            c for c in auto_cats
            if c.get("path", "") == path_filter or c.get("path", "").startswith(path_filter + "/")
        ]
        log.info(f"Авто-категорий (фильтр '{path_filter}'): {len(auto_cats)}")
    else:
        log.info(f"Авто-категорий: {len(auto_cats)}")

    if list_only:
        for c in sorted(auto_cats, key=lambda x: x.get("path", "")):
            print(f"  {c['path']}")
        return [], []

    if max_cats and len(auto_cats) > max_cats:
        auto_cats = auto_cats[:max_cats]
        log.info(f"Ограничено до {max_cats} категорий (--max-cats)")

    log.info("Загружаем каталог Mikado + Автолига...")
    our_brands = load_supplier_brands()

    all_subjects: dict[int, dict] = {}
    log.info(f"Обходим {len(auto_cats)} категорий...")
    for i, cat in enumerate(auto_cats):
        path = cat.get("path", "")
        if (i + 1) % 20 == 0 or i == 0:
            log.info(f"  Категории: {i+1}/{len(auto_cats)}  текущая: {path}")
        for s in client.get_subjects(path, d1, d2):
            sid = s.get("id")
            if sid and sid not in all_subjects:
                all_subjects[sid] = s

    log.info(f"Уникальных предметов WB: {len(all_subjects)}")

    auto_subjects = {
        sid: s for sid, s in all_subjects.items()
        if any(s.get("name", "").lower().startswith(sec) for sec in AUTO_SUBJECT_SECTIONS)
    }
    log.info(f"Авто-предметов (по названию): {len(auto_subjects)}")

    deficit_results:    list[dict] = []
    top_seller_results: list[dict] = []

    for sid, s in auto_subjects.items():
        sc      = score_subject(s)
        name    = s.get("name", "")
        sales   = s.get("sales", 0)
        sellers = s.get("sellers_with_sells", s.get("sellers", 0))
        oos_pct = s.get("lost_profit_percent", 0.0)

        base = {
            "subject_id":    sid,
            "name":          name,
            "in_scope":      subject_in_catalog(name, our_brands),
            "sales":         sales,
            "sellers":       sellers,
            "oos_pct":       round(oos_pct, 1),
            "lost_profit":   s.get("lost_profit", 0),
            "avg_price":     round(s.get("avg_price", 0.0), 0),
            "balance":       s.get("balance", 0),
            "items_total":   s.get("live_items", s.get("items", 0)),
            "items_sells":   s.get("items_with_sells", 0),
            "commision_fbs": round(s.get("commision_fbs", WB_FBS_COMMISSION * 100), 1),
            "score":         round(sc, 4),
        }

        if sc > 0:
            deficit_results.append(base)
        if sales >= MIN_SALES_TOP_SELLER:  # независимо от дефицита — для фазы 3
            top_seller_results.append(base)

    deficit_results.sort(key=lambda x: x["score"], reverse=True)
    top_seller_results.sort(key=lambda x: x["sales"], reverse=True)

    deficit_top = deficit_results[:top]
    log.info(
        f"Предметов с продажами ≥{MIN_MONTHLY_SALES} и OOS ≥{MIN_OOS_PCT}%: "
        f"{len(deficit_results)} | в топ-{top}: {len(deficit_top)}"
    )
    log.info(f"Топ-предметов для фазы 3 (продажи ≥{MIN_SALES_TOP_SELLER}): {len(top_seller_results)}")
    return deficit_top, top_seller_results


# ─── Excel-отчёт ──────────────────────────────────────────────────────────────
_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="EEEEEE"),
)


def _header_row(ws, headers: list[tuple], fill_color: str) -> None:
    h_fill  = PatternFill("solid", fgColor=fill_color)
    h_font  = Font(bold=True, color="FFFFFF", size=11)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, (title, width) in enumerate(headers, 1):
        c = ws.cell(1, col, title)
        c.font = h_font; c.fill = h_fill
        c.alignment = h_align; c.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"


def export_excel(deficit_subjects: list[dict], d1: str, d2: str) -> tuple:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дефицитные предметы"

    _header_row(ws, [
        ("№",                  5),
        ("Предмет WB",        45),
        ("ID предмета",       12),
        ("Авто-тема?",        10),
        ("Продаж/мес",        12),
        ("Продавцов",         11),
        ("OOS %",              9),
        ("Упущ. выручка, ₽",  18),
        ("Ср.цена WB, ₽",     14),
        ("Остаток",           10),
        ("Товаров всего",     13),
        ("Товаров с прод.",   14),
        ("Комиссия FBS,%",    14),
        ("Скор",               9),
        ("Ссылка WB поиск",   20),
    ], "1E3A5F")

    YES_FILL = PatternFill("solid", fgColor="C6EFCE")
    WRN_FILL = PatternFill("solid", fgColor="FFEB9C")
    RED_FILL = PatternFill("solid", fgColor="FFC7CE")

    for ri, r in enumerate(deficit_subjects, 2):
        name_enc   = r["name"].replace(" / ", " ").replace("/", " ")
        search_url = f"https://www.wildberries.ru/catalog/0/search.aspx?search={name_enc.split()[-1]}"
        row_data   = [
            ri - 1, r["name"], r["subject_id"],
            "Да" if r["in_scope"] else "Нет",
            r["sales"], r["sellers"], r["oos_pct"], r["lost_profit"],
            r["avg_price"], r["balance"], r["items_total"],
            r["items_sells"], r["commision_fbs"], r["score"], search_url,
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(ri, col, val)
            c.border = _BORDER
            c.alignment = Alignment(vertical="center")
        ws.cell(ri, 4).fill = YES_FILL if r["in_scope"] else WRN_FILL
        if r["oos_pct"] >= 20:
            ws.cell(ri, 7).fill = RED_FILL

    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / f"wb_deficit_{ts}.xlsx"
    wb.save(path)
    log.info(f"Отчёт сохранён: {path}")
    return path, wb, ts


def _add_products_sheet(
    wb_file:      openpyxl.Workbook,
    products:     list[dict],
    sheet_name:   str,
    header_color: str,
    position:     int,
    show_score:   bool = True,
) -> None:
    ws = wb_file.create_sheet(sheet_name, position)

    RED_FILL = PatternFill("solid", fgColor="FFC7CE")
    YLW_FILL = PatternFill("solid", fgColor="FFEB9C")
    GRN_FILL = PatternFill("solid", fgColor="C6EFCE")

    if show_score:
        headers = [
            ("№",                    5), ("Скор",                8), ("Артикул WB",         12),
            ("Название",            50), ("Бренд",              18), ("Предмет WB",          35),
            ("Цена WB, ₽",         12), ("Цена закупки, ₽",    13), ("Цена прод. мин., ₽",  15),
            ("Цена прод. опт., ₽", 15), ("Маржа WB, %",        10), ("OOS дней",            10),
            ("OOS %",               9), ("Продаж/мес",         12), ("Рейтинг",              9),
            ("Отзывов",             9), ("Продавец",            25), ("Ссылка WB",           14),
        ]
        oos_col    = 13
        link_col   = 18
        margin_col = 11
    else:
        headers = [
            ("№",                    5), ("Артикул WB",         12), ("Название",            50),
            ("Бренд",              18), ("Предмет WB",          35), ("Цена WB, ₽",          12),
            ("Цена закупки, ₽",   13), ("Цена прод. мин., ₽",  15), ("Цена прод. опт., ₽",  15),
            ("Маржа WB, %",        10), ("OOS %",               9), ("Продаж/мес",           12),
            ("Рейтинг",             9), ("Отзывов",             9), ("Продавец",             25),
            ("Ссылка WB",          14),
        ]
        oos_col    = 11
        link_col   = 16
        margin_col = 10

    _header_row(ws, headers, header_color)

    def _price_cell(v: int | None) -> int | str:
        return v if v is not None else "—"

    def _margin_cell(v: float | None) -> float | str:
        return v if v is not None else "—"

    for ri, p in enumerate(products, 2):
        pur = _price_cell(p.get("purchase_price"))
        mn  = _price_cell(p.get("sell_min"))
        opt = _price_cell(p.get("sell_opt"))
        mrg = _margin_cell(p.get("margin_at_wb"))

        if show_score:
            row_data = [
                ri - 1, p["prod_score"], p["nm_id"], p["name"], p["brand"],
                p["subject"], p["price_rub"],
                pur, mn, opt, mrg,
                p["oos_days"], p["oos_pct"],
                p["monthly_sales"], p["rating"], p["feedbacks"], p["supplier"], "→ WB",
            ]
        else:
            row_data = [
                ri - 1, p["nm_id"], p["name"], p["brand"],
                p["subject"], p["price_rub"],
                pur, mn, opt, mrg,
                p["oos_pct"],
                p["monthly_sales"], p["rating"], p["feedbacks"], p["supplier"], "→ WB",
            ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(ri, col, val)
            c.border = _BORDER
            c.alignment = Alignment(vertical="center")

        lc = ws.cell(ri, link_col)
        lc.hyperlink = p["wb_link"]
        lc.font = Font(color="0563C1", underline="single")

        oos = p["oos_pct"]
        if oos >= 30:
            ws.cell(ri, oos_col).fill = RED_FILL
        elif oos >= 15:
            ws.cell(ri, oos_col).fill = YLW_FILL

        margin_val = p.get("margin_at_wb")
        if margin_val is not None:
            if margin_val >= 15:
                ws.cell(ri, margin_col).fill = GRN_FILL
            elif margin_val >= 5:
                ws.cell(ri, margin_col).fill = YLW_FILL
            else:
                ws.cell(ri, margin_col).fill = RED_FILL

    if not products:
        ws.cell(2, 1, "Нет данных")


def add_all_product_sheets(
    wb_file:         openpyxl.Workbook,
    deficit_products: list[dict],
    top_sellers:     list[dict],
) -> None:
    in_cat  = [p for p in deficit_products if p["in_catalog"]]
    out_cat = [p for p in deficit_products if not p["in_catalog"]]

    _add_products_sheet(wb_file, in_cat,  "Дефицит — в каталоге",     "0D5C0D", 0, show_score=True)
    _add_products_sheet(wb_file, out_cat, "Дефицит — нет в каталоге", "7B1818", 1, show_score=True)
    if top_sellers:
        _add_products_sheet(wb_file, top_sellers, "Топ продавцы",     "1E3A5F", 2, show_score=False)

    log.info(
        f"Листы: в каталоге={len(in_cat)}, "
        f"нет в каталоге={len(out_cat)}, "
        f"топ-продавцы={len(top_sellers)}"
    )


def add_catalog_sheet(wb_file: openpyxl.Workbook, products: list[dict]) -> None:
    """Лист 'Наш каталог — топ WB': бренды из наших прайсов, отсортированы по Баллам."""
    ws = wb_file.create_sheet("Наш каталог — топ WB")   # добавляется последним

    RED_FILL = PatternFill("solid", fgColor="FFC7CE")
    YLW_FILL = PatternFill("solid", fgColor="FFEB9C")
    GRN_FILL = PatternFill("solid", fgColor="C6EFCE")

    _header_row(ws, [
        ("№",                    5),
        ("Баллы",                9),
        ("Бренд",               18),
        ("Артикул WB",          12),
        ("Наш артикул",         28),
        ("Название WB",         50),
        ("Цена WB, ₽",         12),
        ("Цена закупки, ₽",    13),
        ("Цена прод. мин., ₽",  15),
        ("Цена прод. опт., ₽",  15),
        ("Маржа WB, %",         10),
        ("OOS %",                9),
        ("Продаж/мес",          12),
        ("OOS дней",            10),
        ("Поз. в каталоге",     15),
        ("Рейтинг",              9),
        ("Отзывов",              9),
        ("Продавец",            25),
        ("Ссылка WB",           14),
    ], "4A235A")  # тёмно-фиолетовый — отличается от остальных листов

    for ri, p in enumerate(products, 2):
        our = p["our_min_price"]
        mn  = p.get("sell_min")
        opt = p.get("sell_opt")
        mrg = p.get("margin_at_wb")
        row_data = [
            ri - 1,
            p["prod_score"],
            p["brand"],
            p["nm_id"],
            p.get("our_articles", ""),
            p["name"],
            p["price_rub"],
            our if our > 0 else "—",
            mn  if mn  is not None else "—",
            opt if opt is not None else "—",
            mrg if mrg is not None else "—",
            p["oos_pct"],
            p["monthly_sales"],
            p["oos_days"],
            p["catalog_count"],
            p["rating"],
            p["feedbacks"],
            p["supplier"],
            "→ WB",
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(ri, col, val)
            c.border = _BORDER
            c.alignment = Alignment(vertical="center", wrap_text=False)

        lc = ws.cell(ri, 19)
        lc.hyperlink = p["wb_link"]
        lc.font = Font(color="0563C1", underline="single")

        oos = p["oos_pct"]
        if oos >= 30:
            ws.cell(ri, 12).fill = RED_FILL
        elif oos >= 10:
            ws.cell(ri, 12).fill = YLW_FILL
        else:
            ws.cell(ri, 12).fill = GRN_FILL

        if mrg is not None:
            if mrg >= 15:
                ws.cell(ri, 11).fill = GRN_FILL
            elif mrg >= 5:
                ws.cell(ri, 11).fill = YLW_FILL
            else:
                ws.cell(ri, 11).fill = RED_FILL

    if not products:
        ws.cell(2, 1, "Нет данных")

    log.info(f"Лист 'Наш каталог — топ WB': {len(products)} строк")


# ─── wb_index.db интеграция ───────────────────────────────────────────────────

def fetch_subjects_from_db(
    top:       int   = 200,
    min_oos:   float = 5.0,
    min_sales: float = 10.0,
) -> tuple[list[dict], list[dict]]:
    """
    Фаза 1 из wb_index.db: агрегирует предметы из wb_products по subject_id.
    Данные частичные (только наши бренды), но не требуют MPStats.
    Возвращает (deficit_subjects, top_seller_subjects) — тот же формат, что run_analysis().
    """
    conn = _db_open()
    try:
        rows = conn.execute("""
            SELECT subject_id, subject,
                   COUNT(*)                                          AS items_total,
                   COUNT(CASE WHEN sales_30d > 0 THEN 1 END)        AS items_sells,
                   CAST(SUM(sales_30d) AS REAL)                     AS total_sales,
                   AVG(CASE WHEN sales_30d > 0 THEN oos_pct END)    AS avg_oos,
                   AVG(price_rub)                                   AS avg_price,
                   AVG(commission_fbs)                              AS avg_comm,
                   COUNT(DISTINCT brand_norm)                       AS brand_count
            FROM wb_products
            WHERE subject_id IS NOT NULL AND subject_id > 0
            GROUP BY subject_id, subject
        """).fetchall()
    finally:
        conn.close()

    deficit:     list[dict] = []
    top_sellers: list[dict] = []

    for r in rows:
        oos     = float(r["avg_oos"]    or 0)
        sales   = float(r["total_sales"] or 0)
        sellers = max(1, int(r["brand_count"] or 1))

        if sales < min_sales or oos < min_oos:
            continue

        sc = score_subject({
            "sales":               sales,
            "sellers_with_sells":  sellers,
            "lost_profit_percent": oos,
        })
        base = {
            "subject_id":    r["subject_id"],
            "name":          r["subject"]   or "",
            "in_scope":      True,
            "sales":         int(sales),
            "sellers":       sellers,
            "oos_pct":       round(oos, 1),
            "lost_profit":   0,
            "avg_price":     round(float(r["avg_price"] or 0), 0),
            "balance":       0,
            "items_total":   r["items_total"],
            "items_sells":   r["items_sells"],
            "commision_fbs": round(float(r["avg_comm"] or 0.25) * 100, 1),
            "score":         round(sc, 4),
        }
        if sc > 0:
            deficit.append(base)
        if sales >= MIN_SALES_TOP_SELLER:
            top_sellers.append(base)

    deficit.sort(key=lambda x: x["score"], reverse=True)
    top_sellers.sort(key=lambda x: x["sales"], reverse=True)
    log.info(
        f"[DB] Фаза 1: {len(deficit)} дефицитных предметов, "
        f"{len(top_sellers)} топ-предметов (из wb_index.db)"
    )
    return deficit[:top], top_sellers


def _db_open() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH), timeout=10)
    conn.row_factory = sqlite3.Row
    return conn


def _in_catalog_set(conn: sqlite3.Connection, nm_ids: list[int]) -> set[int]:
    """Возвращает подмножество nm_ids, присутствующих в wb_matches."""
    result: set[int] = set()
    for i in range(0, len(nm_ids), 999):
        batch = nm_ids[i:i + 999]
        pl = ",".join("?" * len(batch))
        for r in conn.execute(
            f"SELECT DISTINCT nm_id FROM wb_matches WHERE nm_id IN ({pl})", batch
        ).fetchall():
            result.add(r[0])
    return result


def fetch_deficit_from_db(
    deficit_subjects:  list[dict],
    top_subjects:      int   = 20,
    min_oos:           float = 5.0,
    min_sales:         float = 3.0,
) -> list[dict]:
    """
    Фаза 2 из wb_index.db: товары по subject_id дефицитных предметов.
    Заменяет WB-поиск + MPStats per-item вызовы.
    """
    top_subjs   = deficit_subjects[:top_subjects]
    sid_to_subj = {s["subject_id"]: s for s in top_subjs if s.get("subject_id")}

    if not sid_to_subj:
        log.warning("[DB] Нет subject_id в дефицитных предметах — фаза 2 пропущена")
        return []

    conn = _db_open()
    try:
        pl   = ",".join("?" * len(sid_to_subj))
        rows = conn.execute(f"""
            SELECT p.nm_id, p.brand, p.name, p.subject, p.subject_id,
                   p.price_rub, p.sales_30d, p.oos_pct
            FROM wb_products p
            WHERE p.subject_id IN ({pl})
              AND (p.oos_pct  IS NULL OR p.oos_pct  >= ?)
              AND (p.sales_30d IS NULL OR p.sales_30d >= ?)
        """, list(sid_to_subj.keys()) + [min_oos, min_sales]).fetchall()

        if not rows:
            log.info("[DB] Нет товаров по этим subject_id в wb_index.db")
            return []

        nm_ids     = [r["nm_id"] for r in rows]
        in_cat_nm  = _in_catalog_set(conn, nm_ids)

        results: list[dict] = []
        seen:    set[int]   = set()
        for r in rows:
            nm_id = r["nm_id"]
            if nm_id in seen:
                continue
            seen.add(nm_id)
            oos   = float(r["oos_pct"]   or 0)
            sales = float(r["sales_30d"] or 0)
            results.append({
                "nm_id":         nm_id,
                "name":          r["name"]    or "",
                "brand":         r["brand"]   or "",
                "subject":       r["subject"] or "",
                "subject_score": sid_to_subj.get(r["subject_id"], {}).get("score", 0),
                "price_rub":     r["price_rub"] or 0,
                "rating":        0,
                "feedbacks":     0,
                "supplier":      "",
                "oos_days":      round(oos / 100 * 30),
                "oos_pct":       round(oos, 1),
                "monthly_sales": int(sales),
                "in_catalog":    nm_id in in_cat_nm,
                "prod_score":    score_product(oos, int(sales)),
                "wb_link":       f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
            })
    finally:
        conn.close()

    results.sort(key=lambda x: x["prod_score"], reverse=True)
    in_cat = sum(1 for r in results if r["in_catalog"])
    log.info(f"[DB] Фаза 2: {len(results)} товаров (в каталоге: {in_cat})")
    return results


def fetch_top_sellers_from_db(
    top_seller_subjects: list[dict],
    top_subjects:        int   = 20,
    max_oos:             float = 20.0,
    min_sales:           float = 100.0,
) -> list[dict]:
    """
    Фаза 3 из wb_index.db: наши бренды в топ-нишах с низким OOS.
    Заменяет WB-поиск + MPStats per-item вызовы.
    """
    top_subjs   = top_seller_subjects[:top_subjects]
    sid_to_subj = {s["subject_id"]: s for s in top_subjs if s.get("subject_id")}

    if not sid_to_subj:
        log.warning("[DB] Нет subject_id для топ-продавцов — фаза 3 пропущена")
        return []

    conn = _db_open()
    try:
        pl   = ",".join("?" * len(sid_to_subj))
        rows = conn.execute(f"""
            SELECT DISTINCT
                   p.nm_id, p.brand, p.name, p.subject, p.subject_id,
                   p.price_rub, p.sales_30d, p.oos_pct
            FROM wb_products p
            INNER JOIN wb_matches m ON p.nm_id = m.nm_id
            WHERE p.subject_id IN ({pl})
              AND (p.oos_pct   IS NULL OR p.oos_pct   < ?)
              AND (p.sales_30d IS NULL OR p.sales_30d >= ?)
        """, list(sid_to_subj.keys()) + [max_oos, min_sales]).fetchall()

        if not rows:
            log.info("[DB] Нет топ-продавцов по этим subject_id")
            return []

        results: list[dict] = []
        seen:    set[int]   = set()
        for r in rows:
            nm_id = r["nm_id"]
            if nm_id in seen:
                continue
            seen.add(nm_id)
            oos   = float(r["oos_pct"]   or 0)
            sales = float(r["sales_30d"] or 0)
            results.append({
                "nm_id":         nm_id,
                "name":          r["name"]    or "",
                "brand":         r["brand"]   or "",
                "subject":       r["subject"] or "",
                "subject_score": sid_to_subj.get(r["subject_id"], {}).get("score", 0),
                "price_rub":     r["price_rub"] or 0,
                "rating":        0,
                "feedbacks":     0,
                "supplier":      "",
                "oos_days":      round(oos / 100 * 30),
                "oos_pct":       round(oos, 1),
                "monthly_sales": int(sales),
                "in_catalog":    True,
                "prod_score":    score_product(oos, int(sales)),
                "wb_link":       f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
            })
    finally:
        conn.close()

    results.sort(key=lambda x: x["monthly_sales"], reverse=True)
    log.info(f"[DB] Фаза 3: {len(results)} топ-продавцов (OOS<{max_oos}%, прод≥{min_sales})")
    return results


def fetch_catalog_from_db(
    top_brands:  int = 100,
    top_results: int = 1000,
) -> list[dict]:
    """
    Фаза 4 из wb_index.db: наши бренды (Mikado+Автолига) на WB.
    JOIN wb_matches + wb_products — без единого сетевого запроса.
    """
    conn = _db_open()
    try:
        # Топ-N брендов нашего каталога по числу уникальных артикулов
        brand_rows = conn.execute("""
            SELECT our_brand, COUNT(DISTINCT our_article) AS cnt,
                   MIN(our_price) AS min_price
            FROM wb_matches
            GROUP BY our_brand
            ORDER BY cnt DESC
            LIMIT ?
        """, (top_brands,)).fetchall()

        if not brand_rows:
            log.warning("[DB] wb_matches пуст — фаза 4 невозможна")
            return []

        brand_info = {
            r["our_brand"]: {"count": r["cnt"], "min_price": r["min_price"] or 0}
            for r in brand_rows
        }
        brand_list   = list(brand_info.keys())
        pl           = ",".join("?" * len(brand_list))

        # score >= 0.3 отсекает мусорные brand_name-матчи (avg у плохих ~0.013)
        MIN_SCORE = 0.3
        rows = conn.execute(f"""
            SELECT p.nm_id, p.brand AS wb_brand, p.name,
                   p.subject, p.price_rub, p.sales_30d, p.oos_pct,
                   m.our_brand, m.our_price, m.our_article, m.score
            FROM wb_matches m
            JOIN wb_products p ON m.nm_id = p.nm_id
            WHERE m.our_brand IN ({pl})
              AND (p.sales_30d IS NULL OR p.sales_30d >= 1)
              AND m.score >= ?
            ORDER BY p.nm_id, m.our_price ASC
        """, brand_list + [MIN_SCORE]).fetchall()

        if not rows:
            log.info("[DB] Нет WB-товаров для наших брендов (после фильтра score)")
            return []

        # На каждый nm_id берём артикул с минимальной ценой (первый в ORDER BY our_price)
        # и считаем сколько артикулов прошли порог качества
        nm_data:     dict[int, dict]      = {}
        nm_art_cnt:  dict[int, int]       = {}  # сколько артикулов к этому nm_id

        for r in rows:
            nm_id   = r["nm_id"]
            article = r["our_article"] or ""
            price   = float(r["our_price"] or 0)
            if nm_id not in nm_data:
                oos   = float(r["oos_pct"]   or 0)
                sales = float(r["sales_30d"] or 0)
                binfo = brand_info.get(r["our_brand"], {})
                nm_data[nm_id] = {
                    "nm_id":         nm_id,
                    "name":          r["name"]     or "",
                    "brand":         r["wb_brand"] or "",
                    "catalog_count": binfo.get("count", 0),
                    "our_min_price": price or binfo.get("min_price", 0),
                    "our_article":   article,   # лучший (мин. цена, первый в сортировке)
                    "price_rub":     r["price_rub"] or 0,
                    "rating":        0,
                    "feedbacks":     0,
                    "supplier":      "",
                    "oos_days":      round(oos / 100 * 30),
                    "oos_pct":       round(oos, 1),
                    "monthly_sales": int(sales),
                    "prod_score":    score_product(oos, int(sales)),
                    "wb_link":       f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
                }
                nm_art_cnt[nm_id] = 0
            nm_art_cnt[nm_id] += 1

        results = list(nm_data.values())
        for item in results:
            cnt = nm_art_cnt[item["nm_id"]]
            art = item.pop("our_article", "")
            if cnt <= 1:
                item["our_articles"] = art
            else:
                item["our_articles"] = f"{art} (+{cnt - 1})" if art else f"{cnt} арт."
    finally:
        conn.close()

    results.sort(key=lambda x: x["prod_score"], reverse=True)
    final = results[:top_results]
    log.info(f"[DB] Фаза 4: {len(results)} WB-товаров → топ-{len(final)}")
    return final


# ─── Точка входа ──────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="WB Deficit Analyzer")
    parser.add_argument("--top",                   type=int, default=200,
                        help="Топ-N дефицитных предметов (по умолч. 200)")
    parser.add_argument("--days",                  type=int, default=30,
                        help="Период анализа в днях (по умолч. 30)")
    parser.add_argument("--list-categories",       action="store_true")
    parser.add_argument("--clear-cache",           action="store_true")
    parser.add_argument("--path-filter",           type=str, default="Автотовары/Запчасти")
    parser.add_argument("--max-cats",              type=int, default=0)
    parser.add_argument("--with-products",         action="store_true",
                        help="Фазы 2+3: найти конкретные товары с артикулами WB")
    parser.add_argument("--top-subjects",          type=int, default=20,
                        help="Дефицитных предметов для фазы 2 (по умолч. 20)")
    parser.add_argument("--products-per-subject",  type=int, default=30,
                        help="Товаров на предмет из WB поиска (по умолч. 30)")
    parser.add_argument("--top-seller-subjects",   type=int, default=20,
                        help="Топ-предметов для фазы 3 'Топ продавцы' (по умолч. 20)")
    parser.add_argument("--top-catalog-brands",    type=int, default=0,
                        help="Фаза 4: топ-N брендов каталога для проверки на WB (0 = выкл.)")
    parser.add_argument("--products-per-brand",    type=int, default=15,
                        help="Товаров на бренд в фазе 4 (по умолч. 15)")
    parser.add_argument("--live",                  action="store_true",
                        help="Принудительно использовать live WB-поиск + MPStats per-item "
                             "(игнорировать wb_index.db)")
    parser.add_argument("--db-phase1",             action="store_true",
                        help="Фаза 1 из wb_index.db (агрегация предметов) вместо MPStats. "
                             "Включается автоматически при таймауте MPStats.")
    parser.add_argument("--catalog-only",          action="store_true",
                        help="Только лист 'Наш каталог — топ WB' из wb_index.db, "
                             "без MPStats и WB-запросов. 3000 товаров.")
    args = parser.parse_args()

    env   = load_env()
    token = env.get("MPSTATS_TOKEN", "")

    # ── Режим --catalog-only: только Фаза 4 из DB, MPStats не нужен ──────────
    if args.catalog_only:
        if not DB_PATH.exists():
            log.error(f"wb_index.db не найден: {DB_PATH}")
            sys.exit(1)
        top_brands = args.top_catalog_brands if args.top_catalog_brands > 0 else 200
        log.info("=" * 60)
        log.info(f"WB Deficit Analyzer  |  --catalog-only  |  топ-{top_brands} брендов")
        log.info("=" * 60)
        catalog_products = fetch_catalog_from_db(top_brands=top_brands, top_results=3000)
        if not catalog_products:
            log.error("Нет товаров в wb_index.db")
            sys.exit(1)
        _enrich_with_prices(catalog_products)
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)
        add_catalog_sheet(wb_out, catalog_products)
        ts   = datetime.now().strftime("%Y%m%d_%H%M")
        path = OUTPUT_DIR / f"wb_catalog_{ts}.xlsx"
        wb_out.save(path)
        log.info(f"Файл сохранён: {path}  ({len(catalog_products)} строк)")
        return

    if not token:
        log.error("MPSTATS_TOKEN не задан в .env")
        sys.exit(1)

    if args.clear_cache:
        import shutil
        shutil.rmtree(CACHE_DIR, ignore_errors=True)
        CACHE_DIR.mkdir(parents=True)
        log.info("Кэш очищен")

    d2 = datetime.now().strftime("%Y-%m-%d")
    d1 = (datetime.now() - timedelta(days=args.days)).strftime("%Y-%m-%d")

    log.info("=" * 60)
    log.info(f"WB Deficit Analyzer  |  топ-{args.top}  |  период {args.days} дней")
    use_db_p1 = (args.db_phase1 or False) and DB_PATH.exists() and not args.live
    if args.with_products:
        db_mode = "DB" if (DB_PATH.exists() and not args.live) else "live"
        p1_mode = "DB" if use_db_p1 else "MPStats"
        log.info(f"Фаза 1: предметы [{p1_mode}]")
        log.info(f"Фаза 2: топ-{args.top_subjects} дефицитных предметов [{db_mode}]")
        log.info(f"Фаза 3: топ-{args.top_seller_subjects} продавцов-предметов [{db_mode}]")
    if args.top_catalog_brands:
        db_mode = "DB" if (DB_PATH.exists() and not args.live) else "live"
        log.info(f"Фаза 4: топ-{args.top_catalog_brands} брендов [{db_mode}]")
    log.info("=" * 60)

    if use_db_p1:
        log.info("[DB] Фаза 1 из wb_index.db (--db-phase1)")
        deficit_subjects, top_seller_subjects = fetch_subjects_from_db(top=args.top)
    else:
        deficit_subjects, top_seller_subjects = run_analysis(
            token       = token,
            days        = args.days,
            top         = args.top,
            list_only   = args.list_categories,
            path_filter = args.path_filter,
            max_cats    = args.max_cats,
        )
        # Автофоллбэк: MPStats недоступен, но DB есть
        if not deficit_subjects and DB_PATH.exists() and not args.live:
            log.warning("[DB] MPStats вернул пустой результат — пробуем Фазу 1 из wb_index.db")
            deficit_subjects, top_seller_subjects = fetch_subjects_from_db(top=args.top)

    if not deficit_subjects:
        log.error("Нет дефицитных предметов: MPStats недоступен и wb_index.db не найден")
        return

    path, wb_obj, ts = export_excel(deficit_subjects, d1, d2)
    log.info(f"Предметов в отчёте: {len(deficit_subjects)}")
    log.info(f"Авто-тема (in_scope): {sum(1 for r in deficit_subjects if r['in_scope'])}")

    if args.with_products:
        use_db = DB_PATH.exists() and not args.live
        if use_db:
            log.info(f"[DB] wb_index.db найден → используем SQLite (без WB-запросов)")
        else:
            reason = "--live флаг" if args.live else "wb_index.db не найден"
            log.info(f"[live] {reason} → используем live WB-поиск + MPStats per-item")

        log.info(f"\n{'='*60}")
        log.info(f"ФАЗА 2: дефицитные товары (топ-{args.top_subjects} предметов)...")
        log.info(f"{'='*60}")

        if use_db:
            deficit_products = fetch_deficit_from_db(
                deficit_subjects = deficit_subjects,
                top_subjects     = args.top_subjects,
            )
        else:
            client     = MPStatsClient(token)
            our_brands = load_supplier_brands()
            deficit_products = fetch_deficit_products(
                subjects             = deficit_subjects,
                client               = client,
                our_brands           = our_brands,
                d1                   = d1,
                d2                   = d2,
                top_subjects         = args.top_subjects,
                products_per_subject = args.products_per_subject,
            )

        top_sellers: list[dict] = []
        if top_seller_subjects and args.top_seller_subjects > 0:
            log.info(f"\n{'='*60}")
            log.info(f"ФАЗА 3: топ-продавцы в наших нишах (топ-{args.top_seller_subjects} предметов)...")
            log.info(f"{'='*60}")
            if use_db:
                top_sellers = fetch_top_sellers_from_db(
                    top_seller_subjects = top_seller_subjects,
                    top_subjects        = args.top_seller_subjects,
                )
            else:
                top_sellers = fetch_top_seller_products(
                    subjects             = top_seller_subjects,
                    client               = client,
                    our_brands           = our_brands,
                    d1                   = d1,
                    d2                   = d2,
                    top_subjects         = args.top_seller_subjects,
                    products_per_subject = args.products_per_subject,
                )

        log.info("Обогащаем данные ценами (закупка/мин/опт)...")
        _enrich_with_prices(deficit_products)
        if top_sellers:
            _enrich_with_prices(top_sellers)

        add_all_product_sheets(wb_obj, deficit_products, top_sellers)

        # Фаза 4 — наш каталог на WB
        if args.top_catalog_brands > 0:
            log.info(f"\n{'='*60}")
            log.info(f"ФАЗА 4: проверяем наш каталог на WB (топ-{args.top_catalog_brands} брендов)...")
            log.info(f"{'='*60}")
            if use_db:
                catalog_products = fetch_catalog_from_db(
                    top_brands  = args.top_catalog_brands,
                    top_results = 1000,
                )
            else:
                brand_info = load_catalog_brand_info()
                catalog_products = fetch_catalog_products(
                    brand_info         = brand_info,
                    client             = client,
                    d1                 = d1,
                    d2                 = d2,
                    top_brands         = args.top_catalog_brands,
                    products_per_brand = args.products_per_brand,
                    top_results        = 1000,
                )
            if catalog_products:
                _enrich_with_prices(catalog_products)
                add_catalog_sheet(wb_obj, catalog_products)

        final_path = OUTPUT_DIR / f"wb_deficit_{ts}.xlsx"
        wb_obj.save(final_path)
        log.info(f"Файл сохранён: {final_path}")

        log.info(f"\nТОП-10 дефицитных товаров:")
        for i, p in enumerate(deficit_products[:10], 1):
            cat = "✓" if p["in_catalog"] else " "
            log.info(
                f"  {i:2}. [{cat}] арт.{p['nm_id']}  скор={p['prod_score']}  "
                f"OOS={p['oos_pct']}%  прод={p['monthly_sales']}  "
                f"цена={p['price_rub']}₽  {p['brand']} | {p['name'][:40]}"
            )

        if top_sellers:
            log.info(f"\nТОП-10 продавцов в наших нишах:")
            for i, p in enumerate(top_sellers[:10], 1):
                log.info(
                    f"  {i:2}. арт.{p['nm_id']}  прод={p['monthly_sales']}  "
                    f"OOS={p['oos_pct']}%  цена={p['price_rub']}₽  "
                    f"{p['brand']} | {p['name'][:40]}"
                )
    else:
        log.info(f"\nФайл: {path}")
        log.info("Совет: добавь --with-products для конкретных артикулов WB")

        log.info("\nТОП-10 дефицитных предметов:")
        for i, r in enumerate(deficit_subjects[:10], 1):
            scope = "✓" if r["in_scope"] else " "
            log.info(
                f"  {i:2}. [{scope}] OOS={r['oos_pct']}%  "
                f"продаж={r['sales']}  упущено={r['lost_profit']:,}₽  "
                f"скор={r['score']:.3f}  |  {r['name']}"
            )


if __name__ == "__main__":
    main()
