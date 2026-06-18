"""
Диагностика: смотрим что происходит с GF-1904 в прайсе Mikado.
Запуск: uv run --with requests,openpyxl scripts/_diag_gf1904.py
"""
import io
import sys
import requests
import openpyxl
from pathlib import Path
from collections import Counter

sys.stdout.reconfigure(encoding="utf-8")

MIKADO_PRICE_URL = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=BBE2E029-54CF-4D9E-9FAC-9FE25E85B300"
)
PRICE_FALLBACK = Path("C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx")
TARGET = "GF-1904"

def load_wb():
    try:
        resp = requests.get(MIKADO_PRICE_URL, timeout=60)
        resp.raise_for_status()
        if resp.content[:2] == b"PK":
            print(f"[OK] Прайс скачан с сайта ({len(resp.content):,} байт)")
            return openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    except Exception as e:
        print(f"[!] Онлайн недоступен: {e}")
    if PRICE_FALLBACK.exists():
        print(f"[!] Используем локальный файл: {PRICE_FALLBACK}")
        return openpyxl.load_workbook(PRICE_FALLBACK, read_only=True, data_only=True)
    print("[ERR] Прайс недоступен")
    sys.exit(1)

wb = load_wb()
ws = wb.active
rows_iter = ws.iter_rows(values_only=True)
header_raw = next(rows_iter, [])
header = [str(v).strip().lower() if v else "" for v in header_raw]
print(f"\nЗаголовки ({len(header)} колонок): {header[:10]}")

code_idx = price_idx = name_idx = None
for i, h in enumerate(header):
    if h == "code":        code_idx  = i
    if h == "priceout":    price_idx = i
    if "name" in h or "наименование" in h: name_idx = i

print(f"code_idx={code_idx}, price_idx={price_idx}, name_idx={name_idx}")

all_codes: dict[str, list] = {}  # code → [price, ...]
target_rows = []
similar_rows = []

for row in rows_iter:
    raw = row[code_idx] if code_idx is not None and len(row) > code_idx else None
    if not raw:
        continue
    code = str(raw).strip()
    price = 0.0
    if price_idx is not None and len(row) > price_idx:
        try: price = float(str(row[price_idx] or 0))
        except: pass
    name = ""
    if name_idx is not None and len(row) > name_idx:
        name = str(row[name_idx] or "").strip()

    if code not in all_codes:
        all_codes[code] = []
    all_codes[code].append(price)

    # Точное совпадение
    if code.upper() == TARGET.upper():
        target_rows.append((code, price, name, repr(raw)))

    # Похожие: содержит 1904 или начинается с GF
    t_up = TARGET.upper()
    if "1904" in code.upper() or code.upper().startswith("GF"):
        similar_rows.append((code, price, name))

wb.close()

# ─── 1. Что нашли по GF-1904 ────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"Строки с кодом '{TARGET}' (точное совпадение, регистр не важен):")
if target_rows:
    for code, price, name, raw_repr in target_rows:
        print(f"  код={code!r}  цена={price}  назв={name[:60]}  raw={raw_repr}")
else:
    print("  НЕ НАЙДЕНО")

# ─── 2. Похожие коды (GF-* или *1904*) ──────────────────────────────────────
print(f"\nПохожие коды (содержит 1904 или начинается с GF), всего {len(similar_rows)} строк:")
for code, price, name in similar_rows[:30]:
    dup = " ← ДУБЛЬ" if len(all_codes.get(code, [])) > 1 else ""
    print(f"  {code:<20}  {price:>8.0f} ₽  {name[:50]}{dup}")
if len(similar_rows) > 30:
    print(f"  ... и ещё {len(similar_rows)-30}")

# ─── 3. Дубли ────────────────────────────────────────────────────────────────
dups = {c: prices for c, prices in all_codes.items() if len(prices) > 1}
print(f"\nДублирующихся кодов в прайсе Mikado: {len(dups)}")
if dups:
    print("Первые 20 дублей:")
    for c, prices in list(dups.items())[:20]:
        print(f"  {c:<20}  цены: {prices}")

print(f"\nИтого уникальных кодов: {len(all_codes)}")
print(f"Итого строк с ценой>0:  {sum(1 for prices in all_codes.values() if any(p>0 for p in prices))}")
