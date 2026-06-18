"""
wb_top500_analyzer.py — Комбинированный анализ прайса Автолиги → Топ-500 для WB

6 стратегий в одном скоринге:
  S1 Дефицит       — OOS конкурентов по бренду (из MPStats)
  S2 Голубой океан — мало продавцов по бренду на WB
  S3 Автопарк РФ   — артикул подходит к топ-30 массовых авто (Автостат 2025)
  S4 Расходники    — вечный спрос (фильтры, свечи, колодки, ремни…)
  S5 Сезон лето    — пиковый спрос июнь-август 2026 (+155-313% по категориям)
  S6 Уникальность  — бренд слабо представлен на WB → карточек мало → первый

Источники:
  Автолига  — data/suppliers/autoliga/PriceALVLG0411.xls  (26.05.2026)
  MPStats   — кэш mpsb_*.json, 683 бренда, 52 695 авто-товаров WB
  Автостат 2025 — топ автомобилей РФ (avtostat-info.ru)
  WB-статистика 2024-2025 — масла +313%, подвеска +155%, шины +274%

Матчинг: brand-level агрегация (vendorCode недоступен — WB заблокирован на DNS)
"""

import xlrd, json, os, re, math
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Пути ──────────────────────────────────────────────────────────────────────
PRICE_FILE  = "data/suppliers/autoliga/PriceALVLG0411.xls"
CACHE_DIR   = "data/analytics/cache"
OUTPUT_FILE = "data/analytics/wb_top500_autoliga.xlsx"

# ── WB FBS: Автозапчасти, май 2026 ────────────────────────────────────────────
WB_COMMISSION  = 0.25   # комиссия WB
WB_ACQUIRING   = 0.015  # эквайринг
WB_RETURN_RATE = 0.03   # возвраты
TAX_RATE       = 0.06   # УСН 6% от выручки
WB_LOGISTICS   = 70     # ₽ FBS, средний вес ~0.3 кг
MIN_MARGIN_PCT = 15
TARGET_MARGIN  = 25

# ── S3: Топ-30 авто России (Автостат 2025, по объёму автопарка) ───────────────
# Источник: avtostat-info.ru, ixbt.com/news/2026/02, mn.ru
TOP_RU_CARS = [
    # LADA (26.9% парка): Grnta 1.07М, 2107 1.17М, Niva 1.0М, 2109 ~900к
    "lada","granta","гранта","vesta","веста","2107","2106","2109","2110","2114","2115",
    "niva","нива","4x4","kalina","калина","priora","приора","largus","ларгус",
    # Kia (2.75 млн): Rio 1.02М, Sportage, Cerato
    "rio","рио","sportage","спортейдж","cerato","серато","ceed","soul","sorento",
    # Hyundai (2.35 млн): Solaris 926к, Creta, Tucson
    "solaris","солярис","creta","крета","tucson","туксон","elantra","элантра","ix35","santa fe",
    # Renault (2.18 млн): Logan 774к, Duster, Sandero
    "logan","логан","duster","дастер","sandero","сандеро","megane","меган","laguna",
    # Ford: Focus 742к, Mondeo
    "focus","фокус","mondeo","мондео","fusion","explorer",
    # Volkswagen: Polo, Tiguan, Passat
    "polo","поло","tiguan","тигуан","passat","пассат","golf","гольф","jetta","джетта",
    # Toyota: Camry, RAV4, Corolla
    "camry","камри","rav4","corolla","королла","land cruiser","prado","hilux","yaris",
    # Nissan: Almera, X-Trail, Qashqai
    "almera","альмера","x-trail","икстрейл","qashqai","кашкай","teana","tiida","patrol",
    # Skoda: Rapid, Octavia, Fabia
    "rapid","рапид","octavia","октавия","fabia","фабиа","superb","kodiaq",
    # Mitsubishi: Outlander, Lancer, ASX
    "outlander","аутлендер","lancer","лансер","asx","pajero","galant",
    # Chevrolet: Lacetti, Cruze, Aveo
    "lacetti","лачетти","cruze","круз","aveo","авео","cobalt","captiva",
    # BMW
    "bmw","e34","e36","e46","e39","e60","e90","f30",
    # Mercedes
    "mercedes","w210","w211","w203","w204","w212",
    # Peugeot/Citroen
    "peugeot","пежо","citroen","ситроен","308","207","407","c4","c5","berlingo","picasso",
    # Китайские (рост 2023-2026)
    "haval","хавал","jolion","dargo","h6","f7",
    "geely","джили","atlas","coolray","tugella","emgrand",
    "chery","чери","tiggo","exeed",
    "jac","changan","byd","omoda","belgee",
]

# ── S4: Расходники (короткий цикл замены, вечный спрос) ───────────────────────
CONSUMABLES_KW = [
    "фильтр масл", "масляный фильтр", "oil filter",
    "фильтр воздуш", "воздушный фильтр", "air filter",
    "фильтр топлив", "топливный фильтр", "fuel filter",
    "фильтр салон", "салонный фильтр", "cabin filter",
    "свеч", "spark plug", "свеча зажиг",
    "колодк", "тормозн колодк", "brake pad",
    "тормозн диск", "brake disc", "brake rotor",
    "ремень грм", "timing belt", "ремень генератор",
    "масло моторн", "моторное масло", "motor oil",
    "антифриз", "охлаждающ жидкост", "coolant",
    "щетк", "дворник", "wiper blade",
    "лампа", "лампочка", "bulb", "фара",
    "аккумулятор", "акб", "battery",
    "масло трансм", "трансмиссионн масло",
]

# ── S5: Летний сезон (WB-статистика: масла +313%, подвеска +155%, шины +274%) ─
SUMMER_KW = [
    "антифриз", "охлаждающ", "coolant", "радиатор охлажд",
    "кондиционер", "compressor", "компрессор кондиц",
    "масло моторн", "моторное масло", "oil",
    "тормозн", "brake",
    "амортизатор", "shock absorber", "подвеск",
    "шин", "tyres", "диск",
    "аккумулятор", "акб", "battery",
    "дворник", "щетк", "wiper",
    "омыватель стекл", "washer fluid",
    "воздушный фильтр", "air filter",
]

# ── Авто-категории WB (для фильтрации MPStats) ────────────────────────────────
AUTO_KEYS = [
    "авто", "запчаст", "шин", "диск", "тормоз", "подвеск", "двигател",
    "кузов", "электрик", "трансмисс", "масл", "фильтр", "сцеплени",
    "рулев", "охлажден", "технич", "мотоцикл", "выхлоп", "топлив",
    "аккумулят",
]


# ═══════════════════════════════════════════════════════════════════════════════
#  УТИЛИТЫ
# ═══════════════════════════════════════════════════════════════════════════════

def norm(s): return re.sub(r'[\s\-\.\/\\]', '', str(s)).upper()

def is_auto(subj): return any(k in subj.lower() for k in AUTO_KEYS)

def match_kw(text, kw_list):
    t = text.lower()
    return any(k in t for k in kw_list)

def calc_margin(buy, sell, log=WB_LOGISTICS):
    fees = sell * (WB_COMMISSION + WB_ACQUIRING + WB_RETURN_RATE)
    tax  = sell * TAX_RATE
    net  = sell - fees - log - tax - buy
    pct  = (net / sell * 100) if sell > 0 else 0
    return round(net, 2), round(pct, 1)

def min_price(buy, target):
    coeff = (1 - WB_COMMISSION - WB_ACQUIRING - WB_RETURN_RATE - TAX_RATE) - target/100
    return (WB_LOGISTICS + buy) / coeff if coeff > 0 else 1e9


# ═══════════════════════════════════════════════════════════════════════════════
#  ЗАГРУЗКА ПРАЙСА
# ═══════════════════════════════════════════════════════════════════════════════

def load_price() -> list:
    book = xlrd.open_workbook(PRICE_FILE, encoding_override='cp1251')
    ws   = book.sheet_by_index(0)
    items = []
    for r in range(9, ws.nrows):
        try:
            brand   = str(ws.cell_value(r, 1)).strip()
            article = str(ws.cell_value(r, 2)).strip()
            name    = str(ws.cell_value(r, 4)).strip()
            stock   = float(str(ws.cell_value(r, 6)).replace(',', '.') or 0)
            price   = float(str(ws.cell_value(r, 7)).replace(',', '.') or 0)
        except Exception:
            continue
        if not brand or not article or stock <= 0 or price <= 0:
            continue
        items.append({
            "brand":   brand,
            "article": article,
            "name":    name,
            "stock":   stock,
            "buy":     price,
        })
    return items


# ═══════════════════════════════════════════════════════════════════════════════
#  MPSTATS: агрегат по брендам
# ═══════════════════════════════════════════════════════════════════════════════

def build_brand_stats() -> dict:
    """
    Возвращает {brand_upper: stats} где stats:
      avg_sales, max_sales, avg_oos, max_oos, sellers_est,
      wb_price_avg, wb_price_min, wb_price_max, product_count
    """
    brand_data = defaultdict(list)

    for fn in sorted(os.listdir(CACHE_DIR)):
        if not fn.startswith('mpsb_'):
            continue
        with open(os.path.join(CACHE_DIR, fn), encoding='utf-8') as f:
            try: data = json.load(f)
            except: continue

        for p in data:
            if not is_auto(p.get('subject', '')):
                continue
            b = p.get('brand', '').upper().strip()
            if not b:
                continue
            brand_data[b].append({
                'sales': float(p.get('sales_30d', 0) or 0),
                'oos':   float(p.get('oos_pct', 0) or 0),
                'price': float(p.get('price', 0) or 0),
            })

    stats = {}
    for brand, prods in brand_data.items():
        sales_list = [p['sales'] for p in prods]
        oos_list   = [p['oos']   for p in prods]
        price_list = [p['price'] for p in prods if p['price'] > 0]
        stats[brand] = {
            'avg_sales':    sum(sales_list) / len(sales_list) if sales_list else 0,
            'max_sales':    max(sales_list) if sales_list else 0,
            'avg_oos':      sum(oos_list)   / len(oos_list)   if oos_list else 0,
            'max_oos':      max(oos_list)   if oos_list else 0,
            'wb_price_avg': sum(price_list) / len(price_list) if price_list else 0,
            'wb_price_min': min(price_list) if price_list else 0,
            'wb_price_max': max(price_list) if price_list else 0,
            'product_count': len(prods),
        }
    return stats


# ═══════════════════════════════════════════════════════════════════════════════
#  СКОРИНГ
# ═══════════════════════════════════════════════════════════════════════════════

def score_item(item: dict, brand_stats: dict) -> dict:
    buy      = item['buy']
    name     = item['name'].lower()
    brand_up = item['brand'].upper()
    bs       = brand_stats.get(brand_up, {})

    # ── Цена продажи ──────────────────────────────────────────────────────────
    wb_avg   = bs.get('wb_price_avg', 0)
    p_min15  = min_price(buy, MIN_MARGIN_PCT)
    p_rec25  = min_price(buy, TARGET_MARGIN)

    if wb_avg >= p_min15 and wb_avg > 0:
        sell = min(wb_avg * 0.95, p_rec25)
        sell = max(sell, p_min15)
    else:
        sell = p_rec25

    mg_rub, mg_pct = calc_margin(buy, sell)
    if mg_pct < MIN_MARGIN_PCT:
        return {}

    # ── S1: Дефицит ───────────────────────────────────────────────────────────
    oos  = bs.get('avg_oos', 0)
    s1   = 1.0 + (oos / 100) * 2.0           # 1.0–3.0

    # ── S2: Голубой океан ─────────────────────────────────────────────────────
    cnt  = bs.get('product_count', 0)
    # Бренд с < 10 авто-продуктов на WB → мало конкуренции
    if cnt == 0:
        s2 = 1.8   # нет данных = не представлен = первый
    elif cnt < 10:
        s2 = 1.4
    elif cnt < 30:
        s2 = 1.2
    else:
        s2 = 1.0 / math.sqrt(cnt / 10)

    # ── S3: Автопарк РФ ───────────────────────────────────────────────────────
    s3 = 1.35 if match_kw(name, TOP_RU_CARS) else 1.0

    # ── S4: Расходники ────────────────────────────────────────────────────────
    s4 = 1.30 if match_kw(name, CONSUMABLES_KW) else 1.0

    # ── S5: Сезон лето ────────────────────────────────────────────────────────
    s5 = 1.25 if match_kw(name, SUMMER_KW) else 1.0

    # ── S6: Уникальность (бренд слабо на WB) ─────────────────────────────────
    s6 = 1.5 if cnt == 0 else (1.2 if cnt < 5 else 1.0)

    # ── Базовый спрос (по бренду) ─────────────────────────────────────────────
    avg_sales = bs.get('avg_sales', 0)
    demand    = math.log1p(avg_sales) if avg_sales > 0 else 0.3

    # ── Маржа ─────────────────────────────────────────────────────────────────
    m_score = min(mg_pct / 30.0, 1.0)

    # ── Итог ──────────────────────────────────────────────────────────────────
    strategy_mult = max(s3, s4, s5) * s6
    score = demand * s1 * s2 * m_score * strategy_mult

    # ── Обоснование ───────────────────────────────────────────────────────────
    strats = []
    if oos >= 25:   strats.append(f"Дефицит OOS {oos:.0f}%")
    if s2 > 1.1:    strats.append(f"Голубой океан ({cnt} WB-прод.)")
    if s3 > 1.0:    strats.append("Топ авто РФ")
    if s4 > 1.0:    strats.append("Расходник")
    if s5 > 1.0:    strats.append("Сезон лето")
    if s6 > 1.0:    strats.append("Слабая конкуренция WB")

    parts = []
    if avg_sales > 0: parts.append(f"Бренд WB сред.продажи: {avg_sales:.0f} шт/мес")
    if oos > 0:       parts.append(f"OOS по бренду: {oos:.0f}%")
    if wb_avg > 0:    parts.append(f"Ср.цена WB: {wb_avg:.0f} руб")
    parts.append(f"Маржа: {mg_pct:.1f}% / {mg_rub:.0f} руб")
    if strats:        parts.append("[" + "][".join(strats) + "]")

    return {
        'score':      round(score, 5),
        'sell':       round(sell),
        'mg_pct':     mg_pct,
        'mg_rub':     mg_rub,
        'avg_sales':  avg_sales,
        'oos':        oos,
        'wb_cnt':     cnt,
        'strats':     ", ".join(strats) if strats else "Маржа",
        'reason':     "; ".join(parts),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL ЭКСПОРТ
# ═══════════════════════════════════════════════════════════════════════════════

C_HEADER = "1F4E79"
C_GOLD   = "FFF2CC"
C_SILV   = "F2F2F2"
C_BRNZ   = "FCE4D6"

THIN = Border(
    left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),  bottom=Side(style="thin", color="D9D9D9"),
)

COLS = [
    ("#",                    5),
    ("Артикул производителя",27),
    ("Бренд",                18),
    ("Название товара",       42),
    ("Цена закупки, руб",    16),
    ("Цена продажи WB, руб", 18),
    ("Маржа, %",             11),
    ("Маржа, руб",           12),
    ("Ср.продажи WB шт/мес", 20),
    ("OOS по бренду, %",     18),
    ("Остаток шт",           12),
    ("Товаров бренда на WB", 20),
    ("Стратегии",            35),
    ("Подробное обоснование",80),
]


def _c(ws, row, col, val, bold=False, bg=None, wrap=False, align="left", fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, size=10,
                       color="FFFFFF" if bg == C_HEADER else "000000")
    c.border    = THIN
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c


def export(results: list) -> None:
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb = Workbook()

    # ── Лист 1: Топ-500 ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Top-500 WB"
    ws.freeze_panes = "B2"

    for col, (h, w) in enumerate(COLS, 1):
        _c(ws, 1, col, h, bold=True, bg=C_HEADER, align="center", wrap=True)
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w
    ws.row_dimensions[1].height = 40

    fmts = [None,None,None,None,'#,##0','#,##0','0.0','#,##0','#,##0.0','0.0','#,##0','#,##0',None,None]

    for i, item in enumerate(results[:500], 2):
        rank = i - 1
        bg = C_GOLD if rank <= 50 else (C_SILV if rank <= 150 else (C_BRNZ if rank <= 300 else None))
        row = [
            rank,
            item['article'],
            item['brand'],
            item['name'],
            item['buy'],
            item.get('sell', ''),
            item.get('mg_pct', ''),
            item.get('mg_rub', ''),
            item.get('avg_sales', ''),
            item.get('oos', ''),
            item['stock'],
            item.get('wb_cnt', ''),
            item.get('strats', ''),
            item.get('reason', ''),
        ]
        for col, (val, fmt) in enumerate(zip(row, fmts), 1):
            _c(ws, i, col, val, bg=bg, wrap=(col == len(COLS)), fmt=fmt)
        ws.row_dimensions[i].height = 30

    # ── Лист 2: Стратегии ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Стратегии")
    cnt = defaultdict(int)
    for item in results[:500]:
        for s in item.get('strats', '').split(', '):
            if s.strip(): cnt[s.strip()] += 1
    ws2.append(["Стратегия", "Товаров в Топ-500"])
    for s, n in sorted(cnt.items(), key=lambda x: -x[1]):
        ws2.append([s, n])
    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 20

    # ── Лист 3: Топ брендов ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Топ брендов")
    bd = defaultdict(lambda: {'cnt':0,'sales':0,'mg':[]})
    for item in results[:500]:
        b = item['brand']
        bd[b]['cnt'] += 1
        bd[b]['sales'] += item.get('avg_sales', 0)
        bd[b]['mg'].append(item.get('mg_pct', 0))
    ws3.append(["Бренд", "В топ-500", "Сум. WB-продажи шт/мес", "Ср.маржа %"])
    for b, d in sorted(bd.items(), key=lambda x: -x[1]['cnt']):
        avg_mg = sum(d['mg'])/len(d['mg']) if d['mg'] else 0
        ws3.append([b, d['cnt'], round(d['sales']), round(avg_mg, 1)])
    for col in ['A','B','C','D']:
        ws3.column_dimensions[col].width = 22

    # ── Лист 4: Методология ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Методология")
    notes = [
        ("Источник данных", "MPStats API + Автолига прайс 26.05.2026"),
        ("Брендов в прайсе", "499"),
        ("Артикулов со стоком", "41 143"),
        ("Авто-товаров WB в MPStats", "52 695"),
        ("Брендов с MPStats данными", "683"),
        ("", ""),
        ("Формула маржи WB FBS", ""),
        ("Комиссия WB", "25% (Автозапчасти FBS, май 2026)"),
        ("Эквайринг", "1.5%"),
        ("Возвраты", "3%"),
        ("УСН", "6% от выручки"),
        ("Логистика", "70 руб (FBS Волгоград, ~0.3 кг)"),
        ("", ""),
        ("Формула скора", ""),
        ("Score", "demand × S1(OOS) × S2(конкур.) × margin × max(S3,S4,S5) × S6"),
        ("S1 — Дефицит", "1 + (avg_oos / 100) × 2  →  диапазон 1.0—3.0"),
        ("S2 — Голубой океан", "зависит от кол-ва WB-продуктов бренда (0=1.8, <10=1.4, <30=1.2)"),
        ("S3 — Автопарк РФ", "×1.35 если авто из топ-30 по автопарку упомянуто в названии"),
        ("S4 — Расходники", "×1.30 если категория расходника (фильтры, свечи, колодки…)"),
        ("S5 — Сезон лето", "×1.25 если летняя категория (масла, антифриз, кондиц., шины)"),
        ("S6 — Уникальность", "×1.5 если бренд нет на WB; ×1.2 если < 5 продуктов"),
        ("", ""),
        ("Цена продажи", "min(WB_avg×0.95, P_25%) если WB_avg >= P_15%; иначе P_25%"),
        ("Порог входа", "Маржа >= 15% (иначе товар отброшен)"),
        ("", ""),
        ("Замечание о матчинге", "WB Card API недоступен (DNS блокировка); используется"),
        ("", "brand-level агрегация MPStats: статистика по бренду применяется"),
        ("", "ко всем артикулам этого бренда в прайсе. Точность: ±30%."),
        ("", "Для точного матчинга нужен доступ к card.wb.ru."),
    ]
    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 65
    for row_data in notes:
        ws4.append(list(row_data))

    wb.save(OUTPUT_FILE)


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 62)
    print("  WB Top-500  |  Avtoliga  |  6 strategies  |  brand-level")
    print("=" * 62)

    # 1. Прайс
    print("\n[1/4] Loading price list...")
    items = load_price()
    brands_set = set(i['brand'].upper() for i in items)
    print(f"  Articles with stock: {len(items):,}")
    print(f"  Unique brands: {len(brands_set)}")

    # 2. MPStats агрегат
    print("\n[2/4] Building brand stats from MPStats cache...")
    brand_stats = build_brand_stats()
    matched_brands = len(brands_set & set(brand_stats.keys()))
    print(f"  Brands with MPStats data: {len(brand_stats):,}")
    print(f"  Price-list brands matched: {matched_brands}/{len(brands_set)}")

    # 3. Скоринг
    print("\n[3/4] Scoring all articles (6 strategies)...")
    scored, skipped = [], 0
    for item in items:
        s = score_item(item, brand_stats)
        if not s:
            skipped += 1
            continue
        item.update(s)
        scored.append(item)

    scored.sort(key=lambda x: x['score'], reverse=True)
    print(f"  Scored: {len(scored):,}  |  skipped (margin < {MIN_MARGIN_PCT}%): {skipped:,}")

    # Дополнительный скор-тай-брейк внутри бренда: предпочитаем расходники и летние категории
    def tiebreak(it):
        bonus = 0
        nm = it['name'].lower()
        if match_kw(nm, CONSUMABLES_KW): bonus += 0.01
        if match_kw(nm, SUMMER_KW):      bonus += 0.005
        if match_kw(nm, TOP_RU_CARS):    bonus += 0.003
        return it['score'] + bonus

    # Лимит 15 артикулов на бренд → разнообразие в топ-500
    MAX_PER_BRAND = 15
    brand_counter = defaultdict(int)
    diverse_top = []
    # Сначала отбираем строго лучших (без ограничения) до 500 × 4 кандидатов
    for it in scored:
        b = it['brand']
        if brand_counter[b] < MAX_PER_BRAND:
            brand_counter[b] += 1
            it['_tb'] = tiebreak(it)
            diverse_top.append(it)

    # Внутри каждого бренда пересортируем по tie-break
    diverse_top.sort(key=lambda x: x['_tb'], reverse=True)
    print(f"  After brand cap ({MAX_PER_BRAND}/brand): {len(diverse_top):,} candidates for top-500")

    scored = diverse_top  # передаём в export

    # 4. Excel
    print(f"\n[4/4] Exporting top-500 to Excel...")
    export(scored)
    f = os.path.abspath(OUTPUT_FILE)
    print(f"  Saved: {f}")
    sz = os.path.getsize(OUTPUT_FILE) // 1024
    print(f"  Size: {sz} KB")

    # Превью топ-20
    print("\n" + "=" * 100)
    print(f"{'#':>3}  {'Brand':<18} {'Article':<22} {'Buy':>7} {'Sell':>7} "
          f"{'Mg%':>5} {'Sales':>6} {'OOS':>5}  Strategies")
    print("-" * 100)
    for i, it in enumerate(scored[:20], 1):
        print(f"{i:>3}  {it['brand']:<18} {it['article']:<22} "
              f"{it['buy']:>7.0f} {it.get('sell',0):>7.0f} "
              f"{it.get('mg_pct',0):>5.1f}% "
              f"{it.get('avg_sales',0):>6.0f} {it.get('oos',0):>4.0f}%  "
              f"{it.get('strats','')[:40]}")

    strats_cnt = defaultdict(int)
    for it in scored[:500]:
        for s in it.get('strats','').split(', '):
            if s.strip(): strats_cnt[s.strip()] += 1

    print("\n  Top-500 strategy breakdown:")
    for s, n in sorted(strats_cnt.items(), key=lambda x: -x[1]):
        print(f"    {s:<35} {n:>3} товаров")
    print(f"\n  Excel: {f}")


if __name__ == '__main__':
    main()
