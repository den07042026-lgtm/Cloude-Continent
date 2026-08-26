"""
wb_breakeven_report.py
════════════════════════════════════════════════════════════════════════════
Разовый отчёт: точка безубытка по каждому нашему товару на Wildberries.

Берёт все товары WB (nmID + vendorCode + текущая цена), сопоставляет с
закупочными ценами из СВЕЖИХ прайсов Mikado (live-загрузка) и Автолиги
(файл за сегодня), и считает точку безубытка «нашим» WB-калькулятором
(формула wb_price_recalc.py: комиссия 25%, тариф склада Волгоград).

Точка безубытка = минимальная цена продажи, при которой чистая прибыль = 0.

Результат — Excel на Рабочем столе.

Запуск:
  cd C:\\Users\\Admin\\Documents\\Autoparts_Ecommerce
  uv run --with requests,openpyxl,xlrd scripts/wb_breakeven_report.py
"""

import sys
import json
import logging
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Переиспользуем «наш ВБ калькулятор» и загрузчики из рабочего демона
from wb_price_recalc import (
    load_env,
    load_mikado_price,
    load_autoliga_price,
    load_product_dims,
    get_wb_goods,
    calc_profit,
    _volume_liters,
    _delivery_cost,
    DEFAULT_VOLUME,
    SKIP_CODES,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("breakeven")

DESKTOP = Path.home() / "Desktop"
OUT_FILE = DESKTOP / f"ВБ точка безубытка {datetime.now():%Y-%m-%d}.xlsx"

# Источник габаритов (Д×Ш×В) — Топ-500 ВБ_new
TOPWB_DIMS_FILE = DESKTOP / "Топ-500 ВБ" / "Топ-500 ВБ_new.xlsx"

# Кэш списка товаров WB (на случай rate limit при чтении WB API)
WB_GOODS_CACHE = Path(__file__).parent.parent / "data" / "wb_goods_cache.json"


def get_wb_goods_cached(token: str) -> list[dict]:
    """Товары WB с диска-кэшем: при rate limit используем последний успешный список."""
    goods = get_wb_goods(token)
    if goods:
        try:
            WB_GOODS_CACHE.parent.mkdir(parents=True, exist_ok=True)
            WB_GOODS_CACHE.write_text(
                json.dumps(goods, ensure_ascii=False), encoding="utf-8"
            )
        except Exception as e:
            log.warning(f"Не удалось сохранить кэш товаров WB: {e}")
        return goods
    if WB_GOODS_CACHE.exists():
        cached = json.loads(WB_GOODS_CACHE.read_text(encoding="utf-8"))
        log.warning(
            f"WB API недоступен (rate limit) — использую кэш: {len(cached)} товаров "
            f"({datetime.fromtimestamp(WB_GOODS_CACHE.stat().st_mtime):%Y-%m-%d %H:%M})"
        )
        return cached
    return []


def _norm_art(s: str) -> str:
    """Нормализация артикула: без пробелов/дефисов/точек, верхний регистр."""
    return str(s).replace(" ", "").replace("-", "").replace(".", "").upper().strip()


def load_topwb_dims() -> dict[str, dict]:
    """Габариты из Топ-500 ВБ_new.xlsx → {нормализованный_артикул: {length,width,height}}."""
    dims: dict[str, dict] = {}
    if not TOPWB_DIMS_FILE.exists():
        log.warning(f"Габариты: файл не найден {TOPWB_DIMS_FILE}")
        return dims
    wb = openpyxl.load_workbook(TOPWB_DIMS_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # заголовок: 0=Код/Артикул, 11=Длина, 12=Ширина, 13=Высота
    for row in rows:
        if not row or not row[0]:
            continue
        try:
            l = float(row[11] or 0)
            w = float(row[12] or 0)
            h = float(row[13] or 0)
        except (ValueError, TypeError, IndexError):
            continue
        if all(v > 0 for v in (l, w, h)):
            dims[_norm_art(row[0])] = {"length": l, "width": w, "height": h}
    wb.close()
    log.info(f"Габариты: {len(dims)} позиций из Топ-500 ВБ_new")
    return dims


def find_breakeven(purchase: float, liters: float) -> int | None:
    """Минимальная целая цена продажи, при которой прибыль >= 0 (маржа 0%).

    calc_profit монотонно возрастает по цене → двоичный поиск."""
    lo, hi = 1, 500_000
    if calc_profit(purchase, hi, liters) < 0:
        return None
    while lo < hi:
        mid = (lo + hi) // 2
        if calc_profit(purchase, mid, liters) >= 0:
            hi = mid
        else:
            lo = mid + 1
    return lo


def main() -> None:
    env = load_env()
    token = env.get("WB_API_KEY", "")
    if not token:
        log.error("WB_API_KEY не задан в .env — выход")
        sys.exit(1)

    log.info("1/4 Загружаю свежий прайс Mikado (live)…")
    mikado = load_mikado_price()
    log.info("2/4 Загружаю прайс Автолиги (файл за сегодня)…")
    autoliga = load_autoliga_price()
    log.info("3/4 Загружаю габариты товаров (Топ-500 ВБ_new)…")
    dims_db = load_topwb_dims()

    if not mikado and not autoliga:
        log.error("Оба прайса пусты — отчёт отменён")
        sys.exit(1)

    log.info("4/4 Получаю список товаров с WB…")
    goods = get_wb_goods_cached(token)
    if not goods:
        log.error("WB не вернул товары (возможно, rate limit) — попробуйте позже")
        sys.exit(1)

    rows = []
    matched = no_price = no_breakeven = skipped = 0

    for g in goods:
        vc = g["vendorCode"]
        vc_key = vc.lower()
        base_key = vc_key[:-4] if vc_key.endswith("-con") else vc_key

        if base_key in SKIP_CODES or vc_key in SKIP_CODES:
            skipped += 1
            continue

        al_key = base_key.replace("-", "").replace(" ", "").replace(".", "").upper()
        purchase = mikado.get(base_key)
        source = "Микадо"
        if not purchase:
            purchase = autoliga.get(al_key)
            source = "Автолига"
        if not purchase:
            no_price += 1
            continue

        dims = dims_db.get(_norm_art(base_key)) or dims_db.get(al_key)
        if dims:
            liters = _volume_liters(dims["length"], dims["width"], dims["height"])
            has_dims = "да"
        else:
            liters = DEFAULT_VOLUME
            has_dims = "нет (дефолт 3 л)"

        breakeven = find_breakeven(purchase, liters)
        if breakeven is None:
            no_breakeven += 1
            continue

        cur = g["current_price"]
        cur_profit = calc_profit(purchase, cur, liters) if cur > 0 else None
        cur_margin = (cur_profit / cur * 100) if (cur and cur > 0) else None
        gap = (cur - breakeven) if cur > 0 else None

        if cur <= 0:
            status = "нет цены на WB"
        elif cur < breakeven:
            status = "УБЫТОК"
        elif cur_margin is not None and cur_margin < 12:
            status = "ниже целевой (12%)"
        else:
            status = "OK"

        rows.append({
            "nmID": g["nmID"],
            "vendorCode": vc,
            "source": source,
            "purchase": round(purchase, 2),
            "liters": round(liters, 2),
            "has_dims": has_dims,
            "delivery": round(_delivery_cost(liters), 2),
            "breakeven": breakeven,
            "current": round(cur, 2) if cur else 0,
            "cur_profit": round(cur_profit, 2) if cur_profit is not None else None,
            "cur_margin": round(cur_margin, 2) if cur_margin is not None else None,
            "gap": round(gap, 2) if gap is not None else None,
            "status": status,
        })
        matched += 1

    log.info(
        f"Сопоставлено: {matched}  |  без закупки: {no_price}  |  "
        f"без безубытка: {no_breakeven}  |  пропущено: {skipped}"
    )

    if not rows:
        log.error("Нет строк для отчёта — выход")
        sys.exit(1)

    # Сортировка: сначала убыточные, потом по запасу до безубытка
    rows.sort(key=lambda r: (r["gap"] if r["gap"] is not None else 1e9))

    write_excel(rows, matched, no_price, no_breakeven, skipped, len(goods))
    log.info(f"Готово: {OUT_FILE}")


def write_excel(rows, matched, no_price, no_breakeven, skipped, total_goods):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Точка безубытка WB"

    headers = [
        ("nmID", 12),
        ("Артикул", 18),
        ("Источник закупки", 15),
        ("Цена закупки, ₽", 15),
        ("Объём, л", 10),
        ("Габариты", 16),
        ("Логистика WB, ₽", 15),
        ("Точка безубытка, ₽", 17),
        ("Текущая цена WB, ₽", 17),
        ("Прибыль сейчас, ₽", 16),
        ("Маржа сейчас, %", 14),
        ("Запас до безубытка, ₽", 20),
        ("Статус", 20),
    ]

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    loss_fill = PatternFill("solid", fgColor="F8CBAD")   # убыток
    warn_fill = PatternFill("solid", fgColor="FFF2CC")   # ниже целевой
    ok_fill = PatternFill("solid", fgColor="E2EFDA")     # ок

    for c, (name, width) in enumerate(headers, 1):
        cell = ws.cell(1, c, name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = width

    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, 2):
        vals = [
            row["nmID"], row["vendorCode"], row["source"], row["purchase"],
            row["liters"], row["has_dims"], row["delivery"], row["breakeven"],
            row["current"] or None, row["cur_profit"], row["cur_margin"],
            row["gap"], row["status"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c != 2 else "left", vertical="center"
            )
        # подсветка строки по статусу
        st = row["status"]
        fill = None
        if st == "УБЫТОК":
            fill = loss_fill
        elif st == "ниже целевой (12%)":
            fill = warn_fill
        elif st == "OK":
            fill = ok_fill
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = fill

        # формат чисел
        for c in (4, 7, 8, 9, 10, 12):
            ws.cell(r, c).number_format = "# ##0"
        ws.cell(r, 11).number_format = "0.0"

    # ── Лист «Сводка» ──
    s = wb.create_sheet("Сводка")
    losses = [r for r in rows if r["status"] == "УБЫТОК"]
    below = [r for r in rows if r["status"] == "ниже целевой (12%)"]
    ok = [r for r in rows if r["status"] == "OK"]
    summary = [
        ("Дата отчёта", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Калькулятор", "wb_price_recalc.py (комиссия 25%, склад Волгоград)"),
        ("Всего товаров на WB", total_goods),
        ("Сопоставлено с закупкой", matched),
        ("  — из них по прайсу Микадо", sum(1 for r in rows if r["source"] == "Микадо")),
        ("  — из них по прайсу Автолиги", sum(1 for r in rows if r["source"] == "Автолига")),
        ("  — с реальными габаритами", sum(1 for r in rows if r["has_dims"] == "да")),
        ("  — объём по умолчанию (3 л)", sum(1 for r in rows if r["has_dims"] != "да")),
        ("Не найдена цена закупки", no_price),
        ("Не удалось рассчитать безубыток", no_breakeven),
        ("Пропущено (SKIP_CODES)", skipped),
        ("", ""),
        ("⚠ Продаются В УБЫТОК", len(losses)),
        ("⚠ Ниже целевой маржи 12%", len(below)),
        ("✓ OK (маржа ≥ 12%)", len(ok)),
    ]
    s.column_dimensions["A"].width = 38
    s.column_dimensions["B"].width = 45
    for r, (k, v) in enumerate(summary, 1):
        kc = s.cell(r, 1, k)
        s.cell(r, 2, v)
        if k.startswith("⚠") or k.startswith("✓") or k == "Дата отчёта":
            kc.font = Font(bold=True)

    wb.save(OUT_FILE)


if __name__ == "__main__":
    main()
