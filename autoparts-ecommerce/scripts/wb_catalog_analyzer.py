"""
wb_catalog_analyzer.py
════════════════════════════════════════════════════════════════════════════
Анализ дефицитных позиций WB по нашему каталогу.

Алгоритм:
  1. Загружает позиции Mikado + Автолиги (артикул + закупочная цена)
  2. Для топ-N брендов берёт топ-M позиций по цене
  3. По каждой позиции ищет nmID на WB по артикулу (OEM)
     через публичный WB search API (exactmatch)
  4. Для найденных nmID получает OOS%, продажи/мес, среднюю цену
     через MPStats /wb/get/item/{nm_id}/by_category
  5. WB-калькулятор (25%, Волгоград): мин.цена и маржа
  6. Скор = OOS_ratio × log(продажи+1) × (маржа / 12%)
  7. Топ-500 в Excel

Запуск:
  uv run --with requests,openpyxl,xlrd scripts/wb_catalog_analyzer.py
  uv run --with requests,openpyxl,xlrd scripts/wb_catalog_analyzer.py --top-brands 100 --items-per-brand 30

Аргументы:
  --days             период MPStats (по умолч. 30)
  --top-brands       топ-N брендов каждого поставщика (по умолч. 50)
  --items-per-brand  позиций на бренд для поиска (по умолч. 20)
  --min-sales        мин. продаж/мес (по умолч. 5)
  --min-oos          мин. OOS% (по умолч. 0)
  --top              строк в Excel (по умолч. 500)
  --only-viable      только жизнеспособные
  --skip-autolika / --skip-mikado
  --clear-cache
  --dry-run
"""

import sys
import io
import math
import time
import json
import random
import logging
import argparse
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, str(Path(__file__).parent))
sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Установи: uv run --with requests,openpyxl,xlrd scripts/wb_catalog_analyzer.py")
    sys.exit(1)

# ─── Пути ─────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.parent
ENV_FILE   = BASE_DIR / ".env"
LOG_FILE   = BASE_DIR / "logs" / "wb_catalog_analyzer.log"
OUTPUT_DIR = BASE_DIR / "data" / "analytics"
CACHE_DIR  = BASE_DIR / "data" / "analytics" / "cache"

MIKADO_PRICE_URL = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=BBE2E029-54CF-4D9E-9FAC-9FE25E85B300"
)
MPSTATS_BASE = "https://mpstats.io/api"
WB_SEARCH_URL = "https://search.wb.ru/exactmatch/ru/common/v5/search"

# ─── Тарифы WB FBS (Волгоград, май 2026) ─────────────────────────────────────
WB_COMMISSION  = 0.25
TAX_PCT        = 0.06
RET_RATE       = 0.03
PACKAGING      = 30
DELIVERY_BASE  = 50.6
DELIVERY_LITER = 15.4
RETURN_BASE    = 136.0
RETURN_LITER   = 14.0
DEFAULT_VOLUME = 3.0
TARGET_MARGIN  = 0.12

WB_DELAY       = 3.0              # базовая пауза между запросами к WB (сек)
WB_JITTER      = (0.5, 2.0)      # случайный разброс ±
MP_DELAY       = 0.8             # пауза между запросами к MPStats (сек)
WB_RETRY_WAITS = [60, 180, 360]  # backoff при 429 от WB (сек)

# ─── Логирование ──────────────────────────────────────────────────────────────
for _d in (LOG_FILE.parent, OUTPUT_DIR, CACHE_DIR):
    _d.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


def load_env() -> dict:
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


# ══════════════════════════════════════════════════════════════════════════════
#  WB КАЛЬКУЛЯТОР
# ══════════════════════════════════════════════════════════════════════════════

def calc_margin(purchase: float, sell: float, liters: float = DEFAULT_VOLUME) -> float:
    if sell <= 0:
        return -1.0
    commission = sell * WB_COMMISSION
    delivery   = DELIVERY_BASE + liters * DELIVERY_LITER
    ret        = RET_RATE * (RETURN_BASE + liters * RETURN_LITER)
    proceeds   = sell - commission - delivery
    tax        = max(0.0, proceeds) * TAX_PCT
    profit     = sell - purchase - commission - delivery - ret - tax - PACKAGING
    return profit / sell


def find_min_price(purchase: float, liters: float = DEFAULT_VOLUME,
                   target: float = TARGET_MARGIN) -> int | None:
    for s in range(50, 500_001):
        commission = s * WB_COMMISSION
        delivery   = DELIVERY_BASE + liters * DELIVERY_LITER
        ret        = RET_RATE * (RETURN_BASE + liters * RETURN_LITER)
        proceeds   = s - commission - delivery
        tax        = max(0.0, proceeds) * TAX_PCT
        profit     = s - purchase - commission - delivery - ret - tax - PACKAGING
        if s > 0 and profit / s >= target - 1e-6:
            return s
    return None


def compute_score(oos_ratio: float, monthly_sales: float, margin: float) -> float:
    return round(oos_ratio * math.log1p(monthly_sales) * max(0.0, margin / TARGET_MARGIN), 3)


# ══════════════════════════════════════════════════════════════════════════════
#  WB ПОИСК (OEM → nmID)
# ══════════════════════════════════════════════════════════════════════════════

_WB_SESSION: requests.Session | None = None


def _wb_session() -> requests.Session:
    """Возвращает сессию с куками wb.ru (инициализируется один раз)."""
    global _WB_SESSION
    if _WB_SESSION is not None:
        return _WB_SESSION

    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept":          "*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Origin":          "https://www.wildberries.ru",
        "Referer":         "https://www.wildberries.ru/",
        "Sec-Ch-Ua":       '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
        "Sec-Ch-Ua-Mobile":   "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "Sec-Fetch-Dest":  "empty",
        "Sec-Fetch-Mode":  "cors",
        "Sec-Fetch-Site":  "same-site",
        "Connection":      "keep-alive",
    })
    try:
        s.get("https://www.wildberries.ru/", timeout=12)
        time.sleep(random.uniform(1.5, 2.5))
        log.info("WB сессия инициализирована (куки получены)")
    except Exception as e:
        log.warning(f"WB сессия: не удалось получить куки ({e}), продолжаем без них")

    _WB_SESSION = s
    return s


def search_wb_by_article(article: str, limit: int = 5) -> list[dict]:
    """
    Ищет товары WB по артикулу/OEM через публичный exactmatch API.
    Возвращает [{nm_id, name, brand}].
    Кеш бессрочный (артикул → nmID не меняется).
    """
    safe  = article.replace(" ", "_").replace("/", "-").replace("\\", "-")[:60]
    cache = CACHE_DIR / f"wb_{safe}.json"
    if cache.exists():
        return json.loads(cache.read_text(encoding="utf-8"))

    params = {
        "query":     article,
        "resultset": "catalog",
        "limit":     limit * 3,
        "sort":      "popular",
        "page":      1,
        "dest":      -1257786,
        "appType":   1,
        "curr":      "rub",
        "spp":       27,
    }

    sess = _wb_session()
    r = None
    for attempt, wait in enumerate([0] + WB_RETRY_WAITS):
        if wait:
            log.warning(f"WB rate-limit (попытка {attempt}), ждём {wait}с...")
            time.sleep(wait)
        try:
            r = sess.get(WB_SEARCH_URL, params=params, timeout=15)
            if r.status_code != 429:
                break
        except Exception as e:
            log.warning(f"WB поиск '{article}' попытка {attempt+1}: {e}")
            r = None

    products = []
    if r is not None and r.status_code == 200:
        try:
            data = r.json()
            raw  = data.get("products") or (data.get("data") or {}).get("products") or []
            for p in raw[:limit]:
                nm_id = p.get("id")
                if nm_id:
                    products.append({
                        "nm_id": int(nm_id),
                        "name":  p.get("name", ""),
                        "brand": p.get("brand", ""),
                    })
        except Exception as e:
            log.warning(f"WB парсинг '{article}': {e}")
    elif r is not None:
        log.debug(f"WB search '{article}': HTTP {r.status_code}")

    cache.write_text(json.dumps(products, ensure_ascii=False), encoding="utf-8")
    time.sleep(WB_DELAY + random.uniform(*WB_JITTER))
    return products


# ══════════════════════════════════════════════════════════════════════════════
#  MPSTATS — статистика по nmID
# ══════════════════════════════════════════════════════════════════════════════

class MPStatsClient:
    def __init__(self, token: str):
        self.session = requests.Session()
        self.session.headers.update({
            "X-Mpstats-TOKEN": token,
            "Content-Type":    "application/json",
        })

    def get_item_daily(self, nm_id: int, d1: str, d2: str) -> dict | None:
        cache = CACHE_DIR / f"item_{nm_id}_{d1}.json"
        if cache.exists():
            return json.loads(cache.read_text(encoding="utf-8"))

        url = f"{MPSTATS_BASE}/wb/get/item/{nm_id}/by_category"
        for attempt in range(3):
            try:
                r = self.session.get(url, params={"d1": d1, "d2": d2}, timeout=30)
                if r.status_code == 429:
                    wait = 15 * (attempt + 1)
                    log.warning(f"MPStats rate-limit, ждём {wait}с...")
                    time.sleep(wait)
                    continue
                if r.status_code == 200:
                    data = r.json()
                    cache.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
                    time.sleep(MP_DELAY)
                    return data
                log.debug(f"MPStats item/{nm_id}: HTTP {r.status_code} {r.text[:80]}")
                time.sleep(MP_DELAY)
                return None
            except Exception as e:
                log.warning(f"MPStats item/{nm_id} попытка {attempt+1}: {e}")
                time.sleep(3)
        return None


def parse_item_stats(daily: dict) -> dict:
    """Извлекает OOS, продажи, среднюю цену из ответа MPStats."""
    balance   = daily.get("balance", [])
    sales_arr = daily.get("sales", [])
    price_arr = daily.get("final_price", [])
    period    = len(balance) or 1
    oos_days  = sum(1 for b in balance if b == 0)
    monthly_sales = sum(sales_arr)
    prices    = [p for p in price_arr if p and p > 0]
    avg_price = sum(prices) / len(prices) if prices else 0.0
    return {
        "oos_ratio":     round(oos_days / period, 4),
        "oos_days":      oos_days,
        "monthly_sales": monthly_sales,
        "avg_price_mp":  round(avg_price, 0),
    }


# ══════════════════════════════════════════════════════════════════════════════
#  ЗАГРУЗКА КАТАЛОГОВ
# ══════════════════════════════════════════════════════════════════════════════

def load_mikado_items() -> list[dict]:
    log.info("Mikado: загружаем прайс...")
    try:
        resp = requests.get(MIKADO_PRICE_URL, timeout=60)
        resp.raise_for_status()
        if resp.content[:2] != b"PK":
            log.error("Mikado API вернул не Excel")
            return []
        wb   = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
        ws   = wb.active
        rows = ws.iter_rows(values_only=True)
        hdr  = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]
        ci   = next((i for i, h in enumerate(hdr) if h == "code"),      None)
        bi   = next((i for i, h in enumerate(hdr) if h == "brandname"), None)
        ni   = next((i for i, h in enumerate(hdr) if h == "name"),      None)
        pi   = next((i for i, h in enumerate(hdr) if h in ("priceout", "price")), None)

        items = []
        for row in rows:
            def _s(i): return str(row[i]).strip() if i is not None and len(row) > i and row[i] else ""
            def _f(i):
                try: return float(row[i] or 0) if i is not None and len(row) > i else 0.0
                except: return 0.0
            code = _s(ci); price = _f(pi)
            if not code or price <= 0:
                continue
            items.append({"code": code, "brand": _s(bi), "name": _s(ni), "price": price})
        wb.close()
        log.info(f"Mikado: {len(items):,} позиций")
        return items
    except Exception as e:
        log.error(f"Mikado: {e}")
        return []


def load_autolika_items() -> list[dict]:
    try:
        from autoliga_loader import load_autoliga
        catalog = load_autoliga()
        items = list(catalog.values())
        log.info(f"Автолига: {len(items):,} позиций")
        return items
    except Exception as e:
        log.error(f"Автолига: {e}")
        return []


def deduplicated_items(items: list[dict],
                       article_key: str,
                       price_key:   str = "price",
                       brand_key:   str = "brand") -> list[dict]:
    """
    Все позиции каталога, дедуплицированные по артикулу.
    При дублях сохраняет запись с наименьшей ценой (лучший поставщик).
    """
    by_article: dict[str, dict] = {}
    for item in items:
        art = str(item.get(article_key) or "").strip()
        p   = float(item.get(price_key) or 0)
        if not art or p <= 0:
            continue
        if art not in by_article or p < by_article[art]["our_cost"]:
            by_article[art] = {**item, "our_article": art, "our_cost": p}
    return list(by_article.values())


def _wb_cached(article: str) -> bool:
    """Проверяет, есть ли кэш WB-поиска для данного артикула."""
    safe = article.replace(" ", "_").replace("/", "-").replace("\\", "-")[:60]
    return (CACHE_DIR / f"wb_{safe}.json").exists()


# ══════════════════════════════════════════════════════════════════════════════
#  АНАЛИЗ
# ══════════════════════════════════════════════════════════════════════════════

def analyze_items(
    items:     list[dict],
    client:    MPStatsClient,
    d1: str, d2: str,
    source:    str,
    min_sales: int = 5,
    seen_nm:   set | None = None,
    limit:     int = 0,          # макс. новых WB-запросов за сессию (0 = без лимита)
) -> list[dict]:
    """
    Для каждого артикула: WB exactmatch → nmID → MPStats stats → score.
    Кэшированные артикулы обрабатываются всегда (лимит не считает их).
    """
    if seen_nm is None:
        seen_nm = set()

    results:      list[dict] = []
    total         = len(items)
    wb_found      = 0
    mp_found      = 0
    new_wb_calls  = 0   # реальные сетевые запросы к WB (не кэш)

    for idx, item in enumerate(items):
        article  = item["our_article"]
        our_cost = item["our_cost"]
        brand    = str(item.get("brand") or "")

        cached = _wb_cached(article)
        if not cached and limit and new_wb_calls >= limit:
            continue   # лимит новых запросов за сессию исчерпан

        if (idx + 1) % 100 == 0 or idx == 0:
            log.info(
                f"  {source}: {idx+1}/{total}  "
                f"WB новых: {new_wb_calls}  найдено: {wb_found}  "
                f"MPStats: {mp_found}  в отчёте: {len(results)}"
            )

        # ── Шаг 1: WB поиск по артикулу → nmID ───────────────────────────
        wb_products = search_wb_by_article(article, limit=5)
        if not cached:
            new_wb_calls += 1
        if not wb_products:
            continue
        wb_found += len(wb_products)

        # ── Шаг 2: MPStats per-nmID ───────────────────────────────────────
        for prod in wb_products:
            nm_id = prod["nm_id"]
            if nm_id in seen_nm:
                continue

            daily = client.get_item_daily(nm_id, d1, d2)
            if not daily:
                continue
            mp_found += 1

            stats = parse_item_stats(daily)
            if stats["monthly_sales"] < min_sales:
                continue

            wb_price = stats["avg_price_mp"]
            if wb_price <= 0 or our_cost <= 0:
                continue

            min_price = find_min_price(our_cost)
            if min_price is None:
                continue

            margin = calc_margin(our_cost, wb_price)
            if margin < -0.50:
                continue

            seen_nm.add(nm_id)
            score = compute_score(stats["oos_ratio"], stats["monthly_sales"], margin)

            results.append({
                "score":         score,
                "source":        source,
                "brand":         prod["brand"] or brand,
                "nm_id":         nm_id,
                "name":          prod["name"],
                "our_article":   article,
                "our_cost":      round(our_cost, 0),
                "min_price":     min_price,
                "wb_price":      round(wb_price, 0),
                "margin_pct":    round(margin * 100, 1),
                "viable":        wb_price >= min_price,
                "oos_pct":       round(stats["oos_ratio"] * 100, 1),
                "oos_days":      stats["oos_days"],
                "monthly_sales": stats["monthly_sales"],
                "wb_link":       f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
            })

    log.info(
        f"{source}: обработано {total} арт.  |  WB нашёл: {wb_found}  "
        f"|  MPStats OK: {mp_found}  |  в отчёте: {len(results)}"
    )
    return results


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ══════════════════════════════════════════════════════════════════════════════

_BORDER = Border(
    left=Side(style="thin",  color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="EEEEEE"),
)
HEADERS = [
    ("№",             4),  ("Баллы",       9),  ("Источник",    11),
    ("Бренд WB",     20),  ("Артикул WB", 12),  ("Название WB", 52),
    ("Наш артикул",  24),  ("Закупка, ₽", 12),  ("Мин.цена ₽",  13),
    ("Ср.цена WB ₽", 13),  ("Маржа, %",   11),  ("OOS, %",       9),
    ("OOS, дней",    10),  ("Продаж/мес", 12),  ("Ссылка WB",   14),
]
_RED  = PatternFill("solid", fgColor="FFC7CE")
_YLW  = PatternFill("solid", fgColor="FFEB9C")
_GRN  = PatternFill("solid", fgColor="C6EFCE")
_GREY = PatternFill("solid", fgColor="F2F2F2")


def export_excel(rows: list[dict], d1: str, d2: str, dry_run: bool = False) -> Path | None:
    if not rows:
        log.warning("Нет данных для экспорта")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каталог WB"

    hf = PatternFill("solid", fgColor="1E3A5F")
    for col, (title, width) in enumerate(HEADERS, 1):
        c = ws.cell(1, col, title)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = hf
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    viable_count = 0
    for ri, r in enumerate(rows, 2):
        viable = r["viable"]
        if viable:
            viable_count += 1
        bf = None if viable else _GREY

        row_data = [
            ri - 1, r["score"], r["source"], r["brand"],
            r["nm_id"], r["name"], r["our_article"],
            r["our_cost"], r["min_price"], r["wb_price"],
            r["margin_pct"], r["oos_pct"], r["oos_days"],
            r["monthly_sales"], "→ WB",
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(ri, col, val)
            c.border    = _BORDER
            c.alignment = Alignment(vertical="center")
            if bf:
                c.fill = bf

        lc = ws.cell(ri, 15)
        lc.hyperlink = r["wb_link"]
        lc.font = Font(color="0563C1", underline="single")
        if bf:
            lc.fill = bf

        if viable:
            oc = ws.cell(ri, 12)
            mc = ws.cell(ri, 11)
            if r["oos_pct"] >= 30:    oc.fill = _RED
            elif r["oos_pct"] >= 10:  oc.fill = _YLW
            else:                     oc.fill = _GRN
            if r["margin_pct"] >= TARGET_MARGIN * 100:  mc.fill = _GRN
            elif r["margin_pct"] >= 0:                  mc.fill = _YLW
            else:                                        mc.fill = _RED

    ws2 = wb.create_sheet("Инфо")
    for row, (k, v) in enumerate([
        ("Дата",           datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Период MPStats", f"{d1} → {d2}"),
        ("Комиссия WB",    f"{WB_COMMISSION*100:.0f}%  FBS Автозапчасти"),
        ("Доставка",       f"{DELIVERY_BASE}₽ + {DELIVERY_LITER}₽/л  (Волгоград)"),
        ("Возврат",        f"{RETURN_BASE}₽ + {RETURN_LITER}₽/л × {RET_RATE*100:.0f}%"),
        ("УСН",            f"{TAX_PCT*100:.0f}%"),
        ("Целевая маржа",  f"{TARGET_MARGIN*100:.0f}%"),
        ("Объём по умолч.", f"{DEFAULT_VOLUME} л"),
        ("Всего строк",    str(len(rows))),
        ("Жизнеспособных", f"{viable_count}  (цена WB ≥ мин. цена)"),
        ("",               ""),
        ("Поиск",          "OEM/артикул → WB exactmatch → nmID → MPStats stats"),
        ("Баллы",          "OOS_ratio × log(продажи+1) × (маржа / 12%)"),
        ("Серые строки",   "цена WB ниже расчётного минимума"),
    ], 1):
        ws2.cell(row, 1, k).font = Font(bold=True)
        ws2.cell(row, 2, v)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 50

    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / f"wb_catalog_{ts}.xlsx"

    if dry_run:
        log.info(f"[DRY RUN] {len(rows)} строк, {viable_count} жизнеспособных")
        return None

    wb.save(path)
    log.info(f"Excel: {path}  ({len(rows)} строк, {viable_count} жизнеспособных)")
    return path


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def _show_eta(items: list[dict], source: str, session_limit: int) -> int:
    """Показывает статистику кэша и ETA. Возвращает кол-во новых запросов в этой сессии."""
    cached   = sum(1 for it in items if _wb_cached(it["our_article"]))
    uncached = len(items) - cached
    this_run = min(uncached, session_limit) if session_limit else uncached
    eta_min  = this_run * WB_DELAY / 60
    log.info(
        f"  {source}: всего {len(items):,}  |  кэш: {cached:,}  "
        f"|  новых запросов: {this_run:,}  |  ETA: ~{eta_min:.0f} мин"
    )
    return this_run


def main():
    parser = argparse.ArgumentParser(description="WB Catalog Analyzer — полный каталог → WB → MPStats")
    parser.add_argument("--days",          type=int,   default=30,
                        help="Период MPStats в днях (по умолч. 30)")
    parser.add_argument("--limit",         type=int,   default=0,
                        help="Макс. новых WB-запросов за сессию на источник "
                             "(0 = без лимита). Кэшированные не считаются.")
    parser.add_argument("--min-sales",     type=int,   default=5,
                        help="Мин. продаж/мес для включения в отчёт (по умолч. 5)")
    parser.add_argument("--min-oos",       type=float, default=0.0,
                        help="Мин. OOS%% для включения в отчёт (по умолч. 0)")
    parser.add_argument("--top",           type=int,   default=500,
                        help="Строк в итоговом Excel (по умолч. 500)")
    parser.add_argument("--only-viable",   action="store_true",
                        help="Только позиции где цена WB ≥ мин. цена")
    parser.add_argument("--skip-autolika", action="store_true")
    parser.add_argument("--skip-mikado",   action="store_true")
    parser.add_argument("--clear-cache",   action="store_true",
                        help="Очистить кэш WB и MPStats перед запуском")
    parser.add_argument("--dry-run",       action="store_true",
                        help="Не сохранять Excel, только показать статистику")
    args = parser.parse_args()

    env   = load_env()
    token = env.get("MPSTATS_TOKEN", "")
    if not token:
        log.error("MPSTATS_TOKEN не задан в .env")
        sys.exit(1)

    if args.clear_cache:
        import shutil
        shutil.rmtree(CACHE_DIR, ignore_errors=True)
        CACHE_DIR.mkdir(parents=True)
        log.info("Кэш очищен")

    d2 = datetime.now().strftime("%Y-%m-%d")
    d1 = (datetime.now() - timedelta(days=args.days)).strftime("%Y-%m-%d")

    log.info("═" * 70)
    log.info("WB Catalog Analyzer  |  полный каталог  |  OEM → WB → MPStats")
    log.info(f"Период: {d1} → {d2}")
    if args.limit:
        log.info(f"Лимит новых WB-запросов за сессию: {args.limit:,} на источник")
    else:
        log.info("Лимит: без ограничений (полный прогон)")
    log.info("═" * 70)

    client  = MPStatsClient(token)
    seen_nm: set[int] = set()
    all_results: list[dict] = []

    # ── Автолига: OEM-артикул ─────────────────────────────────────────────────
    if not args.skip_autolika:
        log.info("\n── ФАЗА 1: Автолига (OEM-артикул → WB → MPStats) ───────────────")
        al_raw = load_autolika_items()
        if al_raw:
            al_items = deduplicated_items(al_raw, article_key="article",
                                          price_key="price", brand_key="brand")
            log.info(f"Автолига: {len(al_raw):,} строк → {len(al_items):,} уникальных артикулов")
            _show_eta(al_items, "Автолига", args.limit)
            res = analyze_items(al_items, client, d1, d2,
                                source="Автолига",
                                min_sales=args.min_sales,
                                seen_nm=seen_nm,
                                limit=args.limit)
            all_results.extend(res)

    # ── Mikado: внутренний код ────────────────────────────────────────────────
    if not args.skip_mikado:
        log.info("\n── ФАЗА 2: Mikado (код поставщика → WB → MPStats) ──────────────")
        m_raw = load_mikado_items()
        if m_raw:
            m_items = deduplicated_items(m_raw, article_key="code",
                                         price_key="price", brand_key="brand")
            log.info(f"Mikado: {len(m_raw):,} строк → {len(m_items):,} уникальных кодов")
            _show_eta(m_items, "Mikado", args.limit)
            res = analyze_items(m_items, client, d1, d2,
                                source="Mikado",
                                min_sales=args.min_sales,
                                seen_nm=seen_nm,
                                limit=args.limit)
            all_results.extend(res)

    log.info(f"\nВсего найденных позиций: {len(all_results):,}")

    if not all_results:
        log.warning("Нет данных. Запусти ещё раз — кэш накапливается между сессиями.")
        return

    if args.min_oos > 0:
        before = len(all_results)
        all_results = [r for r in all_results if r["oos_pct"] >= args.min_oos]
        log.info(f"Фильтр OOS ≥ {args.min_oos}%: {before} → {len(all_results)}")

    if args.only_viable:
        before = len(all_results)
        all_results = [r for r in all_results if r["viable"]]
        log.info(f"--only-viable: {before} → {len(all_results)}")

    all_results.sort(key=lambda x: x["score"], reverse=True)
    if args.top and len(all_results) > args.top:
        all_results = all_results[:args.top]

    log.info("\nТОП-15:")
    for i, r in enumerate(all_results[:15], 1):
        v = "✓" if r["viable"] else "✗"
        log.info(
            f"  {i:2}. [{v}] скор={r['score']:.3f}  OOS={r['oos_pct']}%  "
            f"прод={r['monthly_sales']:.0f}  WB={r['wb_price']}₽  "
            f"маржа={r['margin_pct']}%  |  {r['brand']} / {r['name'][:35]}"
        )

    viable = sum(1 for r in all_results if r["viable"])
    log.info(f"Итого: {len(all_results):,} строк, жизнеспособных: {viable:,}")
    export_excel(all_results, d1, d2, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
