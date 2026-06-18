"""
moskvorechie_enricher.py — дополнение Excel данными с portal.moskvorechie.ru
══════════════════════════════════════════════════════════════════════════════

Читает Excel файл (например, mikado_data.xlsx), находит строки с пустыми ячейками
в колонках: Параметры, OEM номера, Применяемость, Альтернативные артикулы товара, Изображения.
Для каждой такой строки ищет товар на portal.moskvorechie.ru и заполняет пустые ячейки.

Запуск:
  uv run --with requests,openpyxl scripts/moskvorechie_enricher.py --file mikado_data.xlsx
  uv run --with requests,openpyxl scripts/moskvorechie_enricher.py --file mikado_data.xlsx --rows 2-50
  uv run --with requests,openpyxl scripts/moskvorechie_enricher.py --file data.xlsx --out result.xlsx --login controlvlz1 --password k0nshin

Аргументы:
  --file      Путь к Excel файлу для обогащения  [обязательный]
  --out       Путь для сохранения результата  [по умолчанию: file_enriched.xlsx рядом с исходным]
  --rows      Диапазон строк, например 2-50  [по умолчанию: все строки]
  --login     Логин на portal.moskvorechie.ru  [или MSK_LOGIN в .env]
  --password  Пароль  [или MSK_PASSWORD в .env]
  --delay     Пауза между запросами в секундах  [по умолчанию: 1.5]
  --env       Путь к .env файлу  [по умолчанию: .env рядом со скриптом]
  --images    Папка для сохранения изображений  [по умолчанию: images/ рядом с файлом]
  --no-images Не скачивать изображения
"""

import re
import sys
import time
import argparse
from pathlib import Path
from hashlib import md5

sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
    from openpyxl.styles import Alignment
except ImportError:
    print("Установи зависимости: uv run --with requests,openpyxl ...")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

BASE_URL       = "https://portal.moskvorechie.ru"
LOGIN_URL      = f"{BASE_URL}/login.lmz"
SEARCH_URL     = f"{BASE_URL}/search.lmz"
PRODUCT_URL    = f"{BASE_URL}/good_info.lmc"
SHOW_PRICE_URL = f"{BASE_URL}/show_price.lmz"

# Колонки которые обогащаем (если пустые)
ENRICH_COLS = [
    "Параметры",
    "OEM номера",
    "Применяемость",
    "Альтернативные артикулы товара",
    "Изображения",
]


# ══════════════════════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ══════════════════════════════════════════════════════════════════════════════

def load_env(env_path: Path) -> dict:
    env = {}
    if env_path and env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


def strip_tags(html: str) -> str:
    html = re.sub(r"<[^>]+>", " ", html)
    html = re.sub(r"&nbsp;", " ", html)
    html = re.sub(r"&gt;",   ">", html)
    html = re.sub(r"&lt;",   "<", html)
    html = re.sub(r"&amp;",  "&", html)
    html = re.sub(r"[ \t]+", " ", html)
    return html.strip()


def normalize_code(mikado_code: str) -> str:
    """
    Нормализует артикул Mikado для поиска на Москворечье.
    f-a22025   → a22025   (однобуквенный префикс)
    ab-1234    → 1234     (двухбуквенный префикс)
    xzk-sg-31  → xzk-sg-31 (длинный префикс — оставляем как есть)
    """
    code = str(mikado_code).strip().lower()
    m = re.match(r"^[a-z]{1,2}-(.+)$", code)
    if m:
        return m.group(1)
    return code


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 1 — АВТОРИЗАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

def login(username: str, password: str) -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"
    })
    resp = session.post(
        LOGIN_URL,
        data={"username": username, "password": password,
              "submit": "Вход", "come_from": "/index.lmz"},
        timeout=20,
    )
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")
    if "exit.lmz" in html or "Выход" in html:
        print(f"  ✓ Авторизация успешна (логин: {username})")
    else:
        print("  ✗ Ошибка авторизации — проверьте логин и пароль")
    return session


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 2 — ПОИСК ТОВАРА
# ══════════════════════════════════════════════════════════════════════════════

def search_product(session: requests.Session, article: str) -> list[dict]:
    """
    Ищет товар по артикулу на Москворечье.
    Возвращает: [{"gid": "...", "brand": "...", "article": "...", "name": "...", "via_oem": bool}]

    Примечание: каждый товар отдаётся несколькими строками (по одной на склад).
    Мы берём первую строку с достаточным количеством ячеек и дедуплицируем по GID.
    """
    resp = session.get(SEARCH_URL, params={"sv": article, "dv": "1"}, timeout=20)
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")

    results = []
    seen_gids: set = set()

    for tr in re.finditer(r"<tr[^>]*>(.*?)</tr>", html, re.S | re.I):
        row_html = tr.group(1)
        if "good_info" not in row_html:
            continue

        m_gid = re.search(r"good_info\.lmc\?[^\"']*gid=(\d+)", row_html)
        if not m_gid:
            continue
        gid = m_gid.group(1)

        # Парсим ячейки ДО дедупликации — первая строка с нужными ячейками побеждает
        cells = [strip_tags(c)
                 for c in re.findall(r"<td[^>]*>(.*?)</td>", row_html, re.S | re.I)]
        cells = [c for c in cells if c]

        # Нужно минимум: бренд + артикул
        if len(cells) < 2:
            continue

        # Первые три ячейки: бренд, артикул, название
        # (последующие ячейки — склады/цены — нам не нужны)
        brand = cells[0]
        art   = cells[1]
        name  = cells[2] if len(cells) > 2 else ""

        # Пропускаем служебные заголовки
        if art.lower() in ("артикул", "код", "article"):
            continue

        if gid in seen_gids:
            continue
        seen_gids.add(gid)

        # Найдено через OEM (не прямое совпадение артикула)
        via_oem = bool(re.search(r"найдено через", " ".join(cells), re.I))

        results.append({
            "gid":     gid,
            "brand":   brand,
            "article": art,
            "name":    name,
            "via_oem": via_oem,
        })

    return results


def select_best_match(results: list[dict], article: str, brand_hint: str = "") -> dict | None:
    """
    Выбирает наиболее подходящий товар из результатов поиска.
    Приоритет:
      1. Точное совпадение артикула + бренд
      2. Точное совпадение артикула (первый прямой, не через OEM)
      3. Первый прямой результат
      4. Первый результат вообще
    """
    if not results:
        return None

    art_norm   = article.upper().replace("-", "").replace(" ", "")
    brand_up   = brand_hint.upper() if brand_hint else ""

    def art_matches(r):
        r_art = r["article"].upper().replace("-", "").replace(" ", "")
        return r_art == art_norm

    # Совпадение артикула + бренд
    if brand_up:
        for r in results:
            if art_matches(r) and brand_up in r["brand"].upper() and not r["via_oem"]:
                return r

    # Совпадение только артикула (прямой результат)
    for r in results:
        if art_matches(r) and not r["via_oem"]:
            return r

    # Первый прямой результат
    for r in results:
        if not r["via_oem"]:
            return r

    return results[0]


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 3 — ПАРСИНГ КАРТОЧКИ ТОВАРА
# ══════════════════════════════════════════════════════════════════════════════

def fetch_product_data(session: requests.Session, gid: str) -> dict:
    """
    Загружает полную страницу товара (types=1 даёт применяемость на той же странице).
    Возвращает словарь с параметрами, OEM, применяемостью, путём к фото.
    """
    resp = session.get(
        PRODUCT_URL,
        params={"cat_id": "2", "types": "1", "gid": gid},
        timeout=20,
    )
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")

    data: dict = {}

    # ── Параметры (Критерии товара) ───────────────────────────────────────────
    params: dict = {}
    m = re.search(
        r"Критерии товара(.*?)(?:Фотографи|Оригинальные|Применение товара|$)",
        html, re.S | re.I,
    )
    if m:
        for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", m.group(1), re.S | re.I):
            cells = [strip_tags(c)
                     for c in re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S | re.I)]
            cells = [c for c in cells if c]
            if len(cells) >= 2:
                key = cells[0].rstrip(": ")
                val = cells[1]
                if key in params:
                    params[key] = f"{params[key]} / {val}"
                else:
                    params[key] = val
    data["params"]      = params
    data["params_text"] = "\n".join(f"{k}: {v}" for k, v in params.items()) if params else ""

    # ── OEM номера ────────────────────────────────────────────────────────────
    oem_list: list[dict] = []
    m = re.search(
        r"Оригинальные номера товара(.*?)(?=<(?:div|table)\b[^>]*(?:id|class)=|Применение товара|$)",
        html, re.S | re.I,
    )
    if m:
        for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", m.group(1), re.S | re.I):
            cells = [strip_tags(c)
                     for c in re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S | re.I)]
            cells = [c for c in cells if c]
            if len(cells) >= 2:
                mfr  = cells[0].strip()
                code = cells[1].strip().replace(" ", "")
                if mfr and code and mfr.lower() not in ("производитель", "марка", "бренд"):
                    oem_list.append({"manufacturer": mfr, "oem_code": code})
    data["oem_details"] = oem_list
    data["oem_text"]    = "; ".join(
        f"{o['manufacturer']}: {o['oem_code']}" for o in oem_list[:10]
    ) if oem_list else ""

    # ── Применяемость ─────────────────────────────────────────────────────────
    # Строки вида: "Audi A3 (8P1) | 1.2 TSI | 2010.04 | 2012.08 | 77 | 105 | ..."
    # Признак строки применяемости: минимум 2 ячейки вида YYYY.MM
    compat_rows: list[dict] = []
    for tr_html in re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S | re.I):
        cells = [strip_tags(c)
                 for c in re.findall(r"<td[^>]*>(.*?)</td>", tr_html, re.S | re.I)]
        cells = [c for c in cells if c]

        year_vals = [c for c in cells if re.match(r"^\d{4}\.\d{2}$", c)]
        if len(year_vals) < 2:
            continue
        if len(cells) < 3:
            continue

        brand_model  = cells[0]
        # modification — ячейка перед первым годом, если она не число
        yr_idx = cells.index(year_vals[0])
        modification = cells[yr_idx - 1] if yr_idx > 1 and not cells[yr_idx - 1].replace(".", "").isdigit() else ""

        compat_rows.append({
            "brand_model":  brand_model,
            "modification": modification,
            "year_from":    year_vals[0][:4],
            "year_to":      year_vals[1][:4],
        })

    data["compat_rows"] = compat_rows
    data["compat_text"] = _compat_to_text(compat_rows)

    # img_path не используем из good_info — там пути /2/2/2/... недоступны.
    # Реальные фото берём через show_price.lmz (см. fetch_images_from_show_price).
    data["img_path"] = ""

    return data


def _compat_to_text(rows: list[dict]) -> str:
    """
    Преобразует список применяемости в строку формата Mikado:
    "Audi A3 (8P1) (2003–2012); VW GOLF V (2004–2008)"
    """
    if not rows:
        return ""

    groups: dict = {}
    for r in rows:
        key = r["brand_model"]
        if key not in groups:
            groups[key] = {"year_from": r["year_from"], "year_to": r["year_to"]}
        else:
            g = groups[key]
            if r["year_from"] and (not g["year_from"] or r["year_from"] < g["year_from"]):
                g["year_from"] = r["year_from"]
            if r["year_to"] and (not g["year_to"] or r["year_to"] > g["year_to"]):
                g["year_to"] = r["year_to"]

    parts = []
    for bm, g in groups.items():
        yf, yt = g["year_from"], g["year_to"]
        line = bm
        if yf or yt:
            line += f" ({yf}–{yt})"
        parts.append(line)
    return "; ".join(parts)


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 4 — СКАЧИВАНИЕ ИЗОБРАЖЕНИЙ
# ══════════════════════════════════════════════════════════════════════════════

def fetch_images_from_show_price(session: requests.Session, gid: str,
                                  article: str, images_dir: Path) -> list[str]:
    """
    Скачивает изображения товара через show_price.lmz.

    Почему show_price, а не good_info:
      - good_info выдаёт пути вида /2/2/2/...  — nginx возвращает 404
      - show_price выдаёт пути вида /2/2/3/...  — доступны и возвращают реальные JPEG

    Сохраняет файлы как {article}_msv_1.jpg, _msv_2.jpg, ...
    Дубли по MD5 пропускает.
    Возвращает список имён сохранённых файлов.
    """
    if not images_dir:
        return []

    try:
        r = session.get(
            SHOW_PRICE_URL,
            params={"cat_id": "2", "gid": gid},
            timeout=20,
        )
        r.raise_for_status()
        html = r.content.decode("windows-1251", errors="replace")
    except Exception:
        return []

    # Пути к фото: /2/2/3/... или /2//2/3/... (двойной слеш — баг в HTML сайта)
    img_paths = re.findall(
        r"""src=['"](/\d+/[^'"]+\.(?:jpg|jpeg|png))['"]""", html, re.I
    )
    # Убираем служебные картинки (иконки из /html/)
    img_paths = [p for p in img_paths if not p.startswith("/html/")]
    # Нормализуем двойной слеш
    img_paths = [re.sub(r"//+", "/", p) for p in img_paths]

    if not img_paths:
        return []

    saved: list[str] = []
    seen:  set        = set()
    idx = 1

    for path in img_paths:
        url = BASE_URL + path
        try:
            ri = session.get(url, timeout=20)
            ri.raise_for_status()
            content = ri.content

            if content[:2] == b"\xff\xd8":
                ext = ".jpg"
            elif content[:8] == b"\x89PNG\r\n\x1a\n":
                ext = ".png"
            else:
                continue  # не изображение

            # Пропускаем дубли и placeholder'ы (< 2KB)
            if len(content) < 2048:
                continue
            digest = md5(content).hexdigest()
            if digest in seen:
                continue
            seen.add(digest)

            fname = f"{article}_msv_{idx}{ext}"
            (images_dir / fname).write_bytes(content)
            saved.append(fname)
            idx += 1

        except Exception:
            continue

    return saved


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 5 — ОБОГАЩЕНИЕ EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def enrich_excel(
    file_path: Path,
    out_path: Path,
    session: requests.Session,
    images_dir: Path | None,
    delay: float = 1.5,
    row_start: int = 2,
    row_end: int | None = None,
) -> tuple[int, int]:
    """
    Читает Excel, находит строки с пустыми целевыми ячейками,
    заполняет данными с Москворечья.
    Возвращает (обработано, обогащено).
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Читаем заголовки
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_idx: dict[str, int] = {}
    for i, h in enumerate(header_row, 1):
        if h:
            col_idx[str(h).strip()] = i

    print(f"  Столбцов в файле: {len(col_idx)}")

    code_col  = col_idx.get("Код (Mikado)")
    brand_col = col_idx.get("Бренд")

    if not code_col:
        print("  ✗ Колонка 'Код (Mikado)' не найдена в файле")
        return 0, 0

    # Позиции целевых колонок
    target_cols: dict[str, int | None] = {name: col_idx.get(name) for name in ENRICH_COLS}
    available = [name for name, ci in target_cols.items() if ci is not None]
    print(f"  Целевые колонки найдены: {available}")
    if not available:
        print("  ✗ Ни одна из целевых колонок не найдена")
        return 0, 0

    effective_end = row_end if row_end else ws.max_row
    total_rows    = effective_end - row_start + 1
    processed = 0
    enriched  = 0

    for row_num in range(row_start, effective_end + 1):
        mikado_code = ws.cell(row_num, code_col).value
        if not mikado_code:
            continue
        mikado_code = str(mikado_code).strip()
        brand_hint  = str(ws.cell(row_num, brand_col).value or "").strip() if brand_col else ""

        # Определяем какие ячейки пустые
        empty: dict[str, int] = {}
        for col_name, ci in target_cols.items():
            if ci and not ws.cell(row_num, ci).value:
                empty[col_name] = ci

        if not empty:
            continue  # все ячейки уже заполнены

        processed += 1
        print(f"\n  [{row_num}] {mikado_code}  (пусто: {', '.join(empty)})")

        search_code = normalize_code(mikado_code)

        try:
            # Поиск
            results = search_product(session, search_code)
            if not results:
                print(f"    ✗ Не найдено (поиск: «{search_code}»)")
                time.sleep(delay)
                continue

            best = select_best_match(results, search_code, brand_hint)
            if not best:
                print(f"    ✗ Нет подходящего совпадения")
                time.sleep(delay)
                continue

            print(f"    → {best['brand']} {best['article']} (gid={best['gid']})")

            # Получаем данные карточки (небольшая пауза перед следующим запросом)
            time.sleep(0.5)
            data = fetch_product_data(session, best["gid"])

            row_enriched = False

            # Параметры
            if "Параметры" in empty and data["params_text"]:
                _write_cell(ws, row_num, empty["Параметры"], data["params_text"])
                print(f"    ✓ Параметры: {len(data['params'])} значений")
                row_enriched = True

            # OEM номера
            if "OEM номера" in empty and data["oem_text"]:
                _write_cell(ws, row_num, empty["OEM номера"], data["oem_text"])
                print(f"    ✓ OEM: {len(data['oem_details'])} номеров")
                row_enriched = True

            # Применяемость
            if "Применяемость" in empty and data["compat_text"]:
                _write_cell(ws, row_num, empty["Применяемость"], data["compat_text"])
                print(f"    ✓ Применяемость: {len(data['compat_rows'])} моделей")
                row_enriched = True

            # Альтернативные артикулы — другие бренды/коды из результатов поиска
            if "Альтернативные артикулы товара" in empty:
                alt_codes = [
                    f"{r['brand']}: {r['article']}"
                    for r in results
                    if r["gid"] != best["gid"]
                ]
                if alt_codes:
                    _write_cell(ws, row_num, empty["Альтернативные артикулы товара"],
                                "; ".join(alt_codes))
                    print(f"    ✓ Альт. артикулы: {len(alt_codes)} кодов")
                    row_enriched = True

            # Изображения
            if "Изображения" in empty and images_dir:
                imgs = fetch_images_from_show_price(session, best["gid"],
                                                    mikado_code, images_dir)
                if imgs:
                    _write_cell(ws, row_num, empty["Изображения"], ", ".join(imgs))
                    print(f"    ✓ Фото: {', '.join(imgs)}")
                    row_enriched = True
                else:
                    print(f"    ○ Фото: не найдено")

            if row_enriched:
                enriched += 1

        except Exception as e:
            print(f"    ✗ Ошибка: {e}")

        time.sleep(delay)

    wb.save(out_path)
    return processed, enriched


def _write_cell(ws, row: int, col: int, value: str):
    ws.cell(row, col).value = value
    ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Дополнение Excel данными с portal.moskvorechie.ru",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--file",      required=True,           help="Путь к Excel файлу")
    ap.add_argument("--out",       default=None,            help="Путь для сохранения результата")
    ap.add_argument("--rows",      default=None,            help="Диапазон строк, например 2-50")
    ap.add_argument("--login",     default=None,            help="Логин на portal.moskvorechie.ru")
    ap.add_argument("--password",  default=None,            help="Пароль")
    ap.add_argument("--delay",     default=1.5, type=float, help="Пауза между запросами (сек)")
    ap.add_argument("--env",       default=None,            help="Путь к .env файлу")
    ap.add_argument("--images",    default=None,            help="Папка для сохранения изображений")
    ap.add_argument("--no-images", action="store_true",     help="Не скачивать изображения")
    args = ap.parse_args()

    # ── Пути ──────────────────────────────────────────────────────────────────
    file_path = Path(args.file)
    if not file_path.exists():
        print(f"Файл не найден: {file_path}")
        sys.exit(1)

    if args.out:
        out_path = Path(args.out)
    else:
        out_path = file_path.parent / (file_path.stem + "_enriched" + file_path.suffix)

    if args.no_images:
        images_dir = None
    elif args.images:
        images_dir = Path(args.images)
        images_dir.mkdir(parents=True, exist_ok=True)
    else:
        images_dir = file_path.parent / "images"
        images_dir.mkdir(exist_ok=True)

    # ── Логин и пароль ────────────────────────────────────────────────────────
    env_path = Path(args.env) if args.env else Path(__file__).parent.parent / ".env"
    env      = load_env(env_path)
    username = args.login    or env.get("MSK_LOGIN",    "")
    password = args.password or env.get("MSK_PASSWORD", "")

    if not username or not password:
        print("Укажите логин и пароль: --login XXXXX --password XXXXX")
        print("  или добавьте MSK_LOGIN и MSK_PASSWORD в .env файл")
        sys.exit(1)

    # ── Диапазон строк ────────────────────────────────────────────────────────
    if args.rows:
        row_start, row_end = map(int, args.rows.split("-"))
    else:
        wb      = openpyxl.load_workbook(file_path, read_only=True)
        row_end = wb.active.max_row or 1000
        wb.close()
        row_start = 2

    # ── Шапка ─────────────────────────────────────────────────────────────────
    print()
    print("═" * 60)
    print("  Moskvorechie Enricher")
    print(f"  Файл:   {file_path.name}  (строки {row_start}–{row_end})")
    print(f"  Результат: {out_path}")
    print("═" * 60)

    # ── Авторизация ───────────────────────────────────────────────────────────
    print("\n[1/3] Авторизация...")
    session = login(username, password)

    # ── Обогащение ───────────────────────────────────────────────────────────
    print(f"\n[2/3] Обогащаю данные...")
    print(f"  Изображения: {'отключены (--no-images)' if images_dir is None else str(images_dir)}")

    processed, enriched = enrich_excel(
        file_path=file_path,
        out_path=out_path,
        session=session,
        images_dir=images_dir,
        delay=args.delay,
        row_start=row_start,
        row_end=row_end,
    )

    # ── Итог ──────────────────────────────────────────────────────────────────
    print()
    print("═" * 60)
    print("  Готово!")
    print(f"  Строк с пустыми ячейками: {processed}")
    print(f"  Строк обогащено:          {enriched}")
    print(f"  Результат: {out_path}")
    if images_dir:
        print(f"  Изображения: {images_dir}")
    print("═" * 60)
    print()


if __name__ == "__main__":
    main()
