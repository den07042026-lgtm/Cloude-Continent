"""
wb_art_matcher.py  (wb_oem_matcher.py v4)
════════════════════════════════════════════════════════════════════════════
Матчинг Автолиги с WB по артикулу производителя.

Термины:
  Автолига "Код завода" (col 2) = артикул производителя аналога
  WB vendorCode                 = "Артикул производителя" в карточке WB
  Матчинг: normalize(Код завода) == normalize(vendorCode WB)

Алгоритм:
  1. Загружаем прайс Автолиги: ключ = Код завода (col 2)
  2. Строим индекс WB vendorCode из таблицы vendor_codes
  3. Точное совпадение после нормализации (убираем спецсимволы, upper)
  4. WB-цены: маржа, мин/опт цена продажи (тарифы FBS Автозапчасти)
  5. Excel: 4 листа (Топ дефицит / В наличии / Нет в наличии / Не найдено)

Запуск:
  cd C:\\Users\\Admin\\Documents\\Autoparts_Ecommerce
  uv run --with openpyxl,xlrd scripts/wb_oem_matcher.py
"""

import sys
import math
import re
import logging
import argparse
import sqlite3
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import xlrd
except ImportError:
    print("Установи: uv run --with openpyxl,xlrd scripts/wb_oem_matcher.py")
    sys.exit(1)

# ─── Пути ────────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent.parent
AUTOLIGA_DIR  = BASE_DIR / "data" / "suppliers" / "autoliga"
# Берём последний скачанный прайс из папки
def _latest_autoliga() -> Path:
    files = sorted(AUTOLIGA_DIR.glob("*.xls*"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"Прайс Автолиги не найден в {AUTOLIGA_DIR}")
    return files[0]
AUTOLIGA_PATH = _latest_autoliga()
DB_PATH       = BASE_DIR / "data" / "analytics" / "wb_index.db"
OUTPUT_DIR    = BASE_DIR / "data" / "analytics"
LOG_FILE      = BASE_DIR / "logs" / "wb_oem_matcher.log"

for d in (OUTPUT_DIR, LOG_FILE.parent):
    d.mkdir(parents=True, exist_ok=True)

# ─── Логирование ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ─── WB тарифы (FBS, Автозапчасти, Волгоград) ───────────────────────────────
_WB_COMMISSION  = 0.25
_WB_TAX         = 0.06
_WB_RET_RATE    = 0.03
_WB_PACKAGING   = 30
_WB_DEL_BASE    = 50.6
_WB_DEL_LITER   = 15.4
_WB_RET_BASE    = 136.0
_WB_RET_LITER   = 14.0
_WB_DEFAULT_VOL = 3.0

MIN_ART_LEN = 4
# Артикулы короче этого требуют совпадения бренда (слишком неуникальны)
MIN_ART_LEN_WITHOUT_BRAND = 9

# Разрешённые WB-категории (subject LIKE prefix%)
_AUTO_SUBJECT_PREFIXES = (
    "Автозапчасти",
    "Масла и технические жидкости",
    "Автоаксессуары и дополнительное оборудование",
    "Автохимия и автокосметика",
    "Аккумуляторы для ТС",
)


def normalize_art(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def normalize_brand(s: str) -> str:
    """Только латиница+цифры, верхний регистр. SKF == SKF, FEBI BILSTEIN → FEBIBILSTEIN."""
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def brands_match(brand_al: str, brand_wb: str) -> bool:
    """True если бренды совпадают или один является префиксом другого (FEBI vs FEBI BILSTEIN)."""
    a = normalize_brand(brand_al)
    b = normalize_brand(brand_wb)
    if not a or not b:
        return False
    return a == b or a.startswith(b) or b.startswith(a)


def split_wb_art(vc: str) -> list[str]:
    """WB vendorCode иногда содержит несколько артикулов через / , ;"""
    return [p.strip() for p in re.split(r"[/,;]+", vc) if p.strip()]


# ─── SQLite ──────────────────────────────────────────────────────────────────
def db_open() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


# ─── Фаза 1: Загрузка Автолиги ──────────────────────────────────────────────
def load_autoliga() -> dict[str, dict]:
    """
    Новый формат (с 25.05.2026):
      col 1 = Бренд
      col 2 = Код завода  ← ключ матчинга с WB vendorCode
      col 3 = Внутренний код Автолиги
      col 4 = Название
      col 5 = МинОбъём
      col 6 = Остаток
      col 7 = Цена VIP
    Возвращает {norm_kod_zavoda: {...}}.
    При дубликатах Кода завода берём строку с наименьшей ненулевой ценой.
    """
    log.info(f"Читаем прайс: {AUTOLIGA_PATH.name}")
    wb  = xlrd.open_workbook(str(AUTOLIGA_PATH), encoding_override="cp1251")
    ws  = wb.sheet_by_index(0)
    out: dict[str, dict] = {}
    skipped = 0

    for r in range(9, ws.nrows):
        row = ws.row_values(r)
        if len(row) < 8:
            continue

        brand   = str(row[1]).strip()
        kod_raw = str(row[2]).strip()
        if kod_raw.endswith(".0"):
            kod_raw = kod_raw[:-2]
        name    = str(row[4]).strip()

        try:
            stock = int(float(row[6]))
        except (ValueError, TypeError):
            stock = 0
        try:
            price = float(row[7])
        except (ValueError, TypeError):
            price = 0.0

        if not kod_raw:
            skipped += 1
            continue

        norm_art = normalize_art(kod_raw)
        if len(norm_art) < MIN_ART_LEN:
            skipped += 1
            continue

        rec = {
            "article":  kod_raw,    # Код завода = артикул производителя
            "art_raw":  kod_raw,
            "norm_art": norm_art,
            "brand":    brand,
            "name":     name,
            "stock":    stock,
            "price":    price,
        }

        if norm_art not in out:
            out[norm_art] = rec
        else:
            ep = out[norm_art]["price"]
            if price > 0 and (ep <= 0 or price < ep):
                out[norm_art] = rec

    log.info(f"Автолига: {len(out)} уникальных артикулов производителя  (пропущено: {skipped})")
    return out


# ─── Матчинг: артикул производителя Автолиги → vendorCode WB ────────────────
def match_by_mfr_art(
    al_data: dict[str, dict],
    conn:    sqlite3.Connection,
) -> list[dict]:
    """
    Точное совпадение: normalize(Код завода Автолиги) == normalize(vendorCode WB).
    al_data: ключ = norm_art (нормализованный артикул производителя).
    """
    cached = conn.execute(
        "SELECT COUNT(*) FROM vendor_codes WHERE fetched=1 AND vc_raw IS NOT NULL AND vc_raw != ''"
    ).fetchone()[0]

    if cached == 0:
        log.warning("vendor_codes пуст — запусти wb_vc_fetcher.py")
        return []

    log.info(f"Матчинг артикул произв. → vendorCode WB  ({cached} записей в кеше)...")

    # Индекс: norm(vendorCode WB) → [nm_id, ...]
    vc_rows = conn.execute(
        "SELECT nm_id, vc_raw FROM vendor_codes WHERE fetched=1 AND vc_raw IS NOT NULL AND vc_raw != ''"
    ).fetchall()

    wb_art_index: dict[str, list[int]] = {}
    for row in vc_rows:
        nm_id = int(row["nm_id"])
        for part in split_wb_art(row["vc_raw"]):
            norm = normalize_art(part)
            if len(norm) >= MIN_ART_LEN:
                wb_art_index.setdefault(norm, []).append(nm_id)

    log.info(f"  Индекс WB vendorCode: {len(wb_art_index)} уникальных артикулов")

    subject_filter = " OR ".join(
        f"subject LIKE '{p}%'" for p in _AUTO_SUBJECT_PREFIXES
    )
    wb_data: dict[int, sqlite3.Row] = {
        int(r["nm_id"]): r
        for r in conn.execute(
            f"SELECT nm_id, name, brand, subject, price_rub, sales_30d, oos_pct FROM wb_products WHERE {subject_filter}"
        ).fetchall()
    }
    log.info(f"  WB товаров в авто-категориях: {len(wb_data)}")
    vc_map: dict[int, str] = {
        int(r[0]): r[1]
        for r in conn.execute(
            "SELECT nm_id, vc_raw FROM vendor_codes WHERE vc_raw IS NOT NULL AND vc_raw != ''"
        ).fetchall()
    }

    skipped_brand = 0
    results: list[dict] = []
    seen: set[tuple[str, int]] = set()

    for norm_art, al in al_data.items():
        nm_ids = wb_art_index.get(norm_art, [])
        for nm_id in nm_ids:
            key = (norm_art, nm_id)
            if key in seen:
                continue
            seen.add(key)

            wb = wb_data.get(nm_id)

            # Короткий артикул (< MIN_ART_LEN_WITHOUT_BRAND) требует совпадения бренда
            if len(norm_art) < MIN_ART_LEN_WITHOUT_BRAND:
                wb_brand = wb["brand"] if wb else ""
                if not brands_match(al["brand"], wb_brand):
                    skipped_brand += 1
                    continue

            results.append({
                "our_article": al["article"],
                "our_art":     al["art_raw"],
                "our_brand":   al["brand"],
                "our_name":    al["name"],
                "our_stock":   al["stock"],
                "our_price":   al["price"],
                "nm_id":       nm_id,
                "wb_art":      vc_map.get(nm_id, ""),
                "match_score": 1.0,
                "wb_name":     wb["name"]      if wb else "",
                "wb_brand":    wb["brand"]     if wb else "",
                "wb_subject":  wb["subject"]   if wb else "",
                "wb_price":    float(wb["price_rub"]  or 0) if wb else 0.0,
                "sales_30d":   int(wb["sales_30d"]    or 0) if wb else 0,
                "oos_pct":     float(wb["oos_pct"]    or 0) if wb else 0.0,
            })

    if skipped_brand:
        log.info(f"  Отфильтровано по бренду (короткий арт.): {skipped_brand}")

    log.info(f"  Совпадений: {len(results)} пар")
    return results


# ─── WB ценовой калькулятор ──────────────────────────────────────────────────
def _wb_margin(purchase: float, sell: float,
               liters: float = _WB_DEFAULT_VOL) -> float:
    if sell <= 0:
        return 0.0
    commission = sell * _WB_COMMISSION
    delivery   = _WB_DEL_BASE + liters * _WB_DEL_LITER
    ret_cost   = _WB_RET_RATE * (_WB_RET_BASE + liters * _WB_RET_LITER)
    tax        = max(0.0, sell - commission - delivery) * _WB_TAX
    profit     = sell - purchase - commission - delivery - ret_cost - tax - _WB_PACKAGING
    return profit / sell


def _wb_find_price(purchase: float, target: float) -> int | None:
    if purchase <= 0:
        return None
    if _wb_margin(purchase, 500_000) < target:
        return None
    lo, hi = 50, 500_000
    while lo < hi:
        mid = (lo + hi) // 2
        if _wb_margin(purchase, mid) >= target - 1e-6:
            hi = mid
        else:
            lo = mid + 1
    return lo if _wb_margin(purchase, lo) >= target - 1e-6 else None


def enrich_pricing(matched: list[dict]) -> None:
    for m in matched:
        p  = m.get("our_price") or 0.0
        wb = m.get("wb_price")  or 0.0
        oos   = m.get("oos_pct")   or 0.0
        sales = m.get("sales_30d") or 0

        m["sell_min"]     = _wb_find_price(p, 0.0)  if p > 0 else None
        m["sell_opt"]     = _wb_find_price(p, 0.15) if p > 0 else None
        m["margin_at_wb"] = (
            round(_wb_margin(p, wb) * 100, 1) if p > 0 and wb > 0 else None
        )
        m["score"] = round(oos * math.log1p(sales), 2)


# ─── Excel ───────────────────────────────────────────────────────────────────
_BORDER = Border(
    left=Side(style="thin",  color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="EEEEEE"),
)
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YLW_FILL = PatternFill("solid", fgColor="FFEB9C")
GRN_FILL = PatternFill("solid", fgColor="C6EFCE")
BLU_FILL = PatternFill("solid", fgColor="BDD7EE")


def _write_header(ws, headers: list[tuple[str, int]], color: str) -> None:
    fill  = PatternFill("solid", fgColor=color)
    font  = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, (title, width) in enumerate(headers, 1):
        c = ws.cell(1, col, title)
        c.font = font; c.fill = fill
        c.alignment = align; c.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "A2"


MATCH_HDR = [
    ("№",                         4),
    ("Арт. произв.\n(Автолига)",  22),
    ("Бренд\n(Автолига)",         16),
    ("Название\n(Автолига)",      35),
    ("Остаток,\nшт.",              8),
    ("Цена\nзакупки, ₽",         13),
    ("nm_id WB",                  12),
    ("Арт. произв.\n(WB)",        22),
    ("Название WB",               45),
    ("Бренд WB",                  16),
    ("Предмет WB",                28),
    ("Цена WB, ₽",               11),
    ("OOS %",                      8),
    ("Продаж/\nмес",              10),
    ("Скор\nдефицит",              9),
    ("Маржа при\nцене WB,%",      14),
    ("Цена мин.\n(0%), ₽",        13),
    ("Цена опт.\n(15%), ₽",       14),
    ("Ссылка WB",                  12),
]

OOS_COL    = 13
MARGIN_COL = 16
LINK_COL   = 19


def _write_match_rows(ws, items: list[dict]) -> None:
    for ri, m in enumerate(items, 2):
        row = [
            ri - 1,
            m["our_article"],           # Арт. произв. Автолиги (Код завода)
            m["our_brand"],
            m["our_name"],
            m["our_stock"],
            m["our_price"] or "—",
            m["nm_id"],
            m.get("wb_art", ""),        # vendorCode WB
            m.get("wb_name", ""),
            m.get("wb_brand", ""),
            m.get("wb_subject", ""),
            m.get("wb_price") or "—",
            m.get("oos_pct", 0),
            m.get("sales_30d", 0),
            m.get("score", 0),
            m["margin_at_wb"] if m.get("margin_at_wb") is not None else "—",
            m["sell_min"]     if m.get("sell_min")     is not None else "—",
            m["sell_opt"]     if m.get("sell_opt")     is not None else "—",
            "→ WB",
        ]
        for col, val in enumerate(row, 1):
            c = ws.cell(ri, col, val)
            c.border = _BORDER
            c.alignment = Alignment(vertical="center")

        lc = ws.cell(ri, LINK_COL)
        lc.hyperlink = f"https://www.wildberries.ru/catalog/{m['nm_id']}/detail.aspx"
        lc.font = Font(color="0563C1", underline="single")

        oos = m.get("oos_pct", 0)
        if oos >= 30:
            ws.cell(ri, OOS_COL).fill = RED_FILL
        elif oos >= 15:
            ws.cell(ri, OOS_COL).fill = YLW_FILL

        mrg = m.get("margin_at_wb")
        if mrg is not None:
            ws.cell(ri, MARGIN_COL).fill = (
                GRN_FILL if mrg >= 15 else
                YLW_FILL if mrg >= 0  else
                RED_FILL
            )

    if not items:
        ws.cell(2, 1, "Нет данных")


def export_excel(
    matched:   list[dict],
    unmatched: list[dict],
) -> Path:
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    in_stock = [m for m in matched if m["our_stock"] > 0]
    no_stock = [m for m in matched if m["our_stock"] <= 0]
    top      = sorted(
        [m for m in in_stock if (m.get("oos_pct") or 0) > 0],
        key=lambda x: x["score"], reverse=True,
    )

    sheets = [
        ("Топ дефицит",            "8B0000", top),
        ("В наличии",              "1A5C1A", in_stock),
        ("Нет в наличии",          "555555", no_stock),
    ]
    for title, color, items in sheets:
        ws = wb_out.create_sheet(title)
        _write_header(ws, MATCH_HDR, color)
        _write_match_rows(ws, items)
        log.info(f"  Лист '{title}': {len(items)} строк")

    # Лист "Не найдено"
    ws_miss = wb_out.create_sheet("Не найдено")
    miss_hdr = [
        ("№", 4), ("Арт. произв.\n(Автолига)", 24),
        ("Бренд", 15), ("Название", 40), ("Остаток", 8), ("Цена, ₽", 12),
    ]
    _write_header(ws_miss, miss_hdr, "7B1818")
    for ri, al in enumerate(unmatched, 2):
        row = [ri - 1, al["article"],
               al["brand"], al["name"], al["stock"], al["price"] or "—"]
        for col, val in enumerate(row, 1):
            c = ws_miss.cell(ri, col, val)
            c.border = _BORDER
            c.alignment = Alignment(vertical="center")
    log.info(f"  Лист 'Не найдено': {len(unmatched)} строк")

    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / f"wb_art_match_{ts}.xlsx"
    wb_out.save(path)
    return path


# ─── Точка входа ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="WB OEM Matcher v3 — Код завода → vendorCode WB")
    args = parser.parse_args()

    if not DB_PATH.exists():
        log.error(f"wb_index.db не найден: {DB_PATH}")
        sys.exit(1)

    log.info("=" * 62)
    log.info("WB Art Matcher v4  |  Арт. произв. Автолиги × vendorCode WB")
    log.info("=" * 62)

    conn = db_open()

    # 1. Автолига: ключ словаря = norm_art
    al_data = load_autoliga()

    # 2. Матчинг: артикул производителя → vendorCode WB
    matched = match_by_mfr_art(al_data, conn)

    # 3. Фильтр: убираем где закупка >= цены WB (продавать убыточно)
    before = len(matched)
    matched = [
        m for m in matched
        if not (m["our_price"] > 0 and m["wb_price"] > 0 and m["our_price"] >= m["wb_price"])
    ]
    if before != len(matched):
        log.info(f"  Отфильтровано убыточных: {before - len(matched)}")

    # 4. Кто не найден
    found_norm = {normalize_art(m["our_article"]) for m in matched}
    unmatched  = [al for norm, al in al_data.items() if norm not in found_norm]

    # 5. Цены и скоры
    log.info("Рассчитываем цены и скоры...")
    enrich_pricing(matched)
    matched.sort(key=lambda x: x["score"], reverse=True)

    # 6. Excel
    log.info("Экспортируем Excel...")
    path = export_excel(matched, unmatched)

    in_stock_cnt = sum(1 for m in matched if m["our_stock"] > 0)

    log.info("")
    log.info("ИТОГО:")
    log.info(f"  Арт. произв. в Автолиге       : {len(al_data)}")
    log.info(f"  Совпадений с WB (vendorCode)  : {len(matched)}")
    log.info(f"  В наличии                     : {in_stock_cnt}")
    log.info(f"  Не найдено                    : {len(unmatched)}")
    log.info(f"  Файл                          : {path}")

    if matched:
        log.info("\nТОП-10 по дефициту (в наличии):")
        top10 = [m for m in matched if m["our_stock"] > 0][:10]
        for i, m in enumerate(top10, 1):
            log.info(
                f"  {i:2}. OOS={m['oos_pct']:5.1f}%  "
                f"прод={m['sales_30d']:4d}  скор={m['score']:7.2f}  "
                f"WB={m['wb_price']:>7.0f}₽  закуп={m['our_price']:>7.0f}₽  "
                f"| {m['wb_brand']} {m['wb_name'][:35]}"
            )

    conn.close()


if __name__ == "__main__":
    main()
