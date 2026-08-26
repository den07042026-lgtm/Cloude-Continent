# /// script
# requires-python = ">=3.10"
# dependencies = ["playwright", "openpyxl"]
# ///
"""
gpt_ozon500_dimensions.py
═══════════════════════════
Заполняет столбцы «Длина, мм», «Ширина, мм», «Высота мм», «Вес, гр» в
"500 обогатитель.xlsx" через ChatGPT (браузер, Playwright, тот же CDP-механизм,
что и gpt_ozon500_descriptions.py — переиспользует его функции).

Габариты и вес — ТОВАРА В УПАКОВКЕ (для логистики/отгрузки), не голой детали.
В отличие от OEM/применяемости, здесь пустых значений быть не должно: если
точных данных по артикулу нет, промпт требует от ChatGPT дать максимально
точную оценку по аналогу или по типовым габаритам для категории/типоразмера.

Уже заполненные строки (все 4 поля непустые) пропускаются — можно продолжать
после прерывания. Сохранение частое + в finally (см. историю гонки записи в
fill_applicability_mikado.py / gpt_ozon500_applicability.py / gpt_ozon500_specs.py).

Требует уже запущенный "тестовый" Chrome с CDP-портом 9222, авторизованный в
ChatGPT (scripts/launch_chatgpt.bat).

Тест:
  uv run --with playwright,openpyxl scripts/gpt_ozon500_dimensions.py --rows 2-4 --debug

Обычный запуск (продолжает с непройденных строк):
  uv run --with playwright,openpyxl scripts/gpt_ozon500_dimensions.py --delay 60
"""

import re
import sys
import time
import argparse
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).parent))

from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Alignment

from gpt_ozon500_descriptions import (
    CHATGPT_URL, CDP_PORT, SAVE_EVERY,
    load_category_map, dismiss_modals, type_message, enable_web_search,
    click_send, wait_for_response, safe_save, _is_rate_limited,
)

XLSX_PATH = Path(r"C:\Users\Admin\Desktop\Озон-500\500 обогатитель.xlsx")

_DIMENSIONS_PROMPT_TEMPLATE = """Ты — специалист по логистике и упаковке автозапчастей. Нужно определить максимально точные габариты и вес ТОВАРА В РОЗНИЧНОЙ УПАКОВКЕ (в которой деталь физически отправляется покупателю почтой/транспортной компанией), а не голой детали без упаковки.

Исходные данные:

Наименование: {name}

Бренд: {brand}

Артикул: {code}

Категория детали: {category}

Найди эти данные по карточкам товара на маркетплейсах (Ozon, Wildberries) и в каталогах (Exist, Emex, Autodoc, сайт производителя {brand}) — там часто указан вес и габариты в упаковке.

Если точных данных по этому конкретному артикулу найти не удалось, ОБЯЗАТЕЛЬНО определи их максимально точно косвенным способом:
- по данным ближайшего аналога (кросс-номер / деталь того же типоразмера у другого производителя);
- либо по типовым габаритам и весу для данной категории детали и конкретного типоразмера, с поправкой на марку/модель автомобиля из наименования (например тормозной диск для малолитражки меньше и легче, чем для кроссовера/внедорожника; масляный фильтр для двигателя 1.4-1.6 л меньше, чем для дизеля 2.0+ и т.п.).

КАТЕГОРИЧЕСКИ ЗАПРЕЩЕНО оставлять поле пустым или писать "не найдено", "неизвестно" и подобное — по каждому из 4 параметров всегда должна быть числовая оценка, даже если это осознанная оценка по аналогии, а не точные данные из источника.

ПРОВЕРКА НА ПРАВДОПОДОБНОСТЬ (обязательно перед финальным ответом): если эта деталь — типовой представитель своей категории и типоразмера (например обычный масляный фильтр АКПП с прокладкой для легкового авто среднего класса), её упаковка должна быть сопоставима по габаритам и весу с другими типичными деталями той же категории и типоразмера, а не отличаться от них в разы. Отклонение в несколько раз от типичного для этой категории размера допустимо, только если для этого есть явная техническая причина (например деталь для грузовика/спецтехники, значительно больший или меньший типоразмер, многокомпонентный комплект и т.п.) — в остальных случаях перепроверь оценку и скорректируй её к правдоподобному диапазону для этой категории.

Единицы: длина/ширина/высота — в миллиметрах, целыми числами. Вес — в граммах, целым числом.

Ответ дай СТРОГО четырьмя строками, только целые числа после двоеточия, без единиц измерения, без пояснений, markdown и текста до или после:

ДЛИНА: число
ШИРИНА: число
ВЫСОТА: число
ВЕС: число"""


def build_prompt(row: dict) -> str:
    return _DIMENSIONS_PROMPT_TEMPLATE.format(
        name=row["name"],
        brand=row["brand"],
        code=row["code"],
        category=row.get("category") or "не указана",
    )


_FIELD_PATTERNS = [
    ("length", re.compile(r'^\s*длина\s*:\s*(.*)$', re.I)),
    ("width", re.compile(r'^\s*ширина\s*:\s*(.*)$', re.I)),
    ("height", re.compile(r'^\s*высота\s*:\s*(.*)$', re.I)),
    ("weight", re.compile(r'^\s*вес\s*:\s*(.*)$', re.I)),
]

_NUMBER_RE = re.compile(r'\d+(?:[.,]\d+)?')


def parse_dimensions(text: str) -> dict | None:
    """Возвращает {length,width,height,weight} как int, либо None если что-то не распознано."""
    text = re.sub(r'\*+', '', text)
    values = {}
    for line in text.splitlines():
        line = line.strip()
        for key, pat in _FIELD_PATTERNS:
            if key in values:
                continue
            m = pat.match(line)
            if m:
                num_m = _NUMBER_RE.search(m.group(1))
                if num_m:
                    values[key] = int(round(float(num_m.group(0).replace(",", "."))))
    if len(values) < 4:
        return None
    return values


def process_row(page, row: dict, debug: bool, retries: int = 2) -> dict | None:
    for attempt in range(1, retries + 1):
        try:
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30_000)
            try:
                page.wait_for_load_state("networkidle", timeout=10_000)
            except Exception:
                pass
            time.sleep(4)
            dismiss_modals(page)

            try:
                page_text = page.locator("body").inner_text(timeout=3000)
            except Exception:
                page_text = ""
            if _is_rate_limited(page_text):
                print("    [!] ChatGPT: рейт-лимит (баннер страницы) — жду 20 мин...")
                time.sleep(1200)
                continue

            prompt = build_prompt(row)
            type_message(page, prompt)

            web_search_on = enable_web_search(page)
            if web_search_on:
                print("    [web search ON]", end=" ", flush=True)

            if debug and attempt == 1:
                print("    [debug] промпт введён, нажми Enter для отправки...")
                input()

            click_send(page)
            time.sleep(2)

            response = wait_for_response(page, timeout_sec=300 if web_search_on else 180)

            if not response:
                if attempt < retries:
                    print(f"    [!] Пустой ответ, попытка {attempt + 1}/{retries}...")
                    time.sleep(15)
                continue

            if _is_rate_limited(response):
                print("    [!] ChatGPT: рейт-лимит — жду 20 мин...")
                time.sleep(1200)
                continue

            values = parse_dimensions(response)
            if values is None:
                print(f"    [!] Не распознан формат ответа, попытка {attempt}")
                if attempt < retries:
                    time.sleep(10)
                continue
            return values

        except RuntimeError as e:
            print(f"    [!] {e}")
            break
        except Exception as e:
            print(f"    [!] Ошибка (попытка {attempt}): {e}")
            if attempt < retries:
                time.sleep(15)
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--debug", action="store_true", help="Пауза перед отправкой (проверка промпта)")
    ap.add_argument("--rows", default=None, help="Диапазон строк Excel: 2-10 или одна строка: 5")
    ap.add_argument("--delay", default=60, type=int, help="Пауза между строками, сек")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb.active
    category_map = load_category_map()

    col_code, col_name, col_brand = 1, 2, 3
    col_len, col_wid, col_hei, col_wei = 10, 11, 12, 13

    max_row = ws.max_row
    row_start, row_end = 2, max_row
    if args.rows:
        if "-" in args.rows:
            a, b = args.rows.split("-")
            row_start, row_end = int(a), int(b)
        else:
            row_start = row_end = int(args.rows)
    row_end = min(row_end, max_row)

    print()
    print("=" * 62)
    print("  GPT Ozon-500 Dimensions [Длина/Ширина/Высота/Вес]")
    print(f"  Файл:   {XLSX_PATH.name}")
    print(f"  Строки: {row_start}-{row_end} ({row_end - row_start + 1} шт.)")
    print("=" * 62)

    with sync_playwright() as pw:
        browser = pw.chromium.connect_over_cdp(f"http://127.0.0.1:{CDP_PORT}")
        context = browser.contexts[0] if browser.contexts else browser.new_context()

        page = None
        for p in context.pages:
            if "chatgpt.com" in p.url:
                page = p
                break
        if page is None:
            page = context.pages[0] if context.pages else context.new_page()

        print("\n  Открываю ChatGPT...")
        try:
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30_000)
        except Exception as e:
            print(f"  Не удалось открыть ChatGPT: {e}")
            sys.exit(1)

        time.sleep(3)
        if any(k in page.url.lower() for k in ("login", "auth", "signin", "sign-in")):
            print("  Не авторизован! Войди в ChatGPT в открытом окне тестового Chrome и перезапусти скрипт.")
            sys.exit(1)

        print("  Авторизован\n")

        stats = {"done": 0, "skipped": 0, "failed": 0}
        since_save = 0

        try:
            for row_num in range(row_start, row_end + 1):
                code = ws.cell(row_num, col_code).value
                if not code:
                    continue
                code = str(code).strip()
                name = str(ws.cell(row_num, col_name).value or "")
                brand = str(ws.cell(row_num, col_brand).value or "")

                if all(ws.cell(row_num, c).value for c in (col_len, col_wid, col_hei, col_wei)):
                    print(f"  [{row_num:3}] {code}: пропуск (уже заполнено)")
                    stats["skipped"] += 1
                    continue

                print(f"\n  [{row_num:3}/{row_end}] {brand} {code} — {name[:45]}")

                category = category_map.get(code.lower(), "")
                row_data = {"code": code, "name": name, "brand": brand, "category": category}
                values = process_row(page, row_data, debug=args.debug)

                if values:
                    ws.cell(row_num, col_len, values["length"]).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row_num, col_wid, values["width"]).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row_num, col_hei, values["height"]).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row_num, col_wei, values["weight"]).alignment = Alignment(horizontal="center", vertical="center")
                    print(f"         + Д{values['length']} x Ш{values['width']} x В{values['height']} мм, {values['weight']} г")
                    stats["done"] += 1
                    since_save += 1
                else:
                    print("         не удалось")
                    stats["failed"] += 1

                if since_save >= SAVE_EVERY:
                    safe_save(wb, XLSX_PATH)
                    since_save = 0

                if row_num < row_end:
                    time.sleep(args.delay)
        finally:
            safe_save(wb, XLSX_PATH)

        # НЕ закрываем context/browser - это подключение по CDP к уже открытому
        # "тестовому" Chrome пользователя, а не браузер, запущенный этим скриптом.

    print()
    print("=" * 62)
    print("  Готово!")
    print(f"  Заполнено: {stats['done']}")
    print(f"  Пропущено: {stats['skipped']}")
    print(f"  Ошибок:    {stats['failed']}")
    print("=" * 62)
    print()


if __name__ == "__main__":
    main()
