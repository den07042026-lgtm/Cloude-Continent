"""
mikado_scraper.py — сбор данных с mikado-parts.ru
══════════════════════════════════════════════════

Что делает:
  1. Авторизуется на сайте по логину и паролю
  2. Читает артикулы из Excel прайса (колонка Prodnum или Code)
  3. По каждому артикулу скачивает:
       • основную карточку — название, бренд, цена, склады
       • параметры — всё что указано в разделе «Параметры» на сайте
       • OEM номера
       • применяемость (марка / модель / годы)
       • изображения
  4. Сохраняет сводную таблицу Excel с результатами
  5. Сохраняет изображения в папку images/
  6. Сохраняет JSON-файл по каждому артикулу

Запуск:
  uv run --with requests,openpyxl scripts/mikado_scraper.py --price прайс.xlsx
  uv run --with requests,openpyxl scripts/mikado_scraper.py --price прайс.xlsx --rows 2-20
  uv run --with requests,openpyxl scripts/mikado_scraper.py --price прайс.xlsx --out C:/output/
  uv run --with requests,openpyxl scripts/mikado_scraper.py --login 35275 --password XXXX --price прайс.xlsx

Аргументы:
  --price      Путь к Excel прайсу поставщика  [обязательный]
  --rows       Диапазон строк прайса, например 2-20  [по умолчанию: все строки]
  --out        Папка для сохранения результатов  [по умолчанию: папка рядом с прайсом]
  --login      Код клиента Mikado  [или MIKADO_CODE в .env]
  --password   Пароль  [или MIKADO_PASSWORD в .env]
  --delay      Пауза между запросами в секундах  [по умолчанию: 1.5]
  --env        Путь к .env файлу  [по умолчанию: .env рядом со скриптом]
"""

import re
import sys
import time
import json
import argparse
from pathlib import Path
from hashlib import md5

sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Установи зависимости: uv run --with requests,openpyxl mikado_scraper.py")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

BASE_URL    = "https://mikado-parts.ru/office"
LOGIN_URL   = f"{BASE_URL}/SECURE.asp"
PRODUCT_URL = f"{BASE_URL}/galleyp.asp"
AJAX_URL    = f"{BASE_URL}/pp0.asp"


# ══════════════════════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ══════════════════════════════════════════════════════════════════════════════

def load_env(env_path: Path) -> dict:
    """Читает переменные из .env файла."""
    env = {}
    if env_path and env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


def strip_tags(html: str) -> str:
    """Убирает HTML-теги, декодирует спецсимволы, нормализует пробелы."""
    html = re.sub(r"<[^>]+>", " ", html)
    html = re.sub(r"&nbsp;",  " ", html)
    html = re.sub(r"&gt;",    ">", html)
    html = re.sub(r"&lt;",    "<", html)
    html = re.sub(r"&amp;",   "&", html)
    html = re.sub(r"[ \t]+",  " ", html)
    return html.strip()


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 1 — АВТОРИЗАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

def login(code: str, password: str) -> requests.Session:
    """
    Авторизуется на mikado-parts.ru.
    Возвращает сессию с активными куками — её используют все последующие запросы.
    """
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"
    })
    resp = session.post(
        LOGIN_URL,
        data={"CODE": code, "PASSWORD": password, "INSERT": "OK"},
        timeout=20
    )
    resp.raise_for_status()
    page = resp.content.decode("windows-1251", errors="replace")

    if "Обслуживание клиентов" in page or "Продолжить" in page:
        print(f"  ✓ Авторизация успешна (код: {code})")
    else:
        print("  ✗ Ошибка авторизации — проверьте логин и пароль")
    return session


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 2 — ЧТЕНИЕ ПРАЙСА
# ══════════════════════════════════════════════════════════════════════════════

def read_price(price_file: Path, row_start: int, row_end: int) -> list[dict]:
    """
    Читает строки из Excel прайса Mikado.
    Ищет колонку Prodnum (артикул для URL: f-a22025).
    Если Prodnum нет — использует Code.
    Возвращает список словарей, по одному на строку прайса.
    """
    wb = openpyxl.load_workbook(price_file, read_only=True, data_only=True)
    ws = wb.active

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(v).strip() if v else f"col{i}" for i, v in enumerate(first_row, 1)]

    prodnum_idx = next((i for i, h in enumerate(headers) if h.lower() == "prodnum"), None)
    code_idx    = next((i for i, h in enumerate(headers) if h.lower() == "code"),    None)
    url_idx     = prodnum_idx if prodnum_idx is not None else code_idx

    if url_idx is None:
        raise ValueError("Колонки 'Prodnum' или 'Code' не найдены в прайсе")

    items = []
    for row_num, row in enumerate(
        ws.iter_rows(min_row=row_start, max_row=row_end, values_only=True),
        start=row_start
    ):
        art_code = row[url_idx] if len(row) > url_idx else None
        if not art_code:
            continue
        item = {"_row": row_num, "_code": str(art_code).strip()}
        for col_i, val in enumerate(row):
            h = headers[col_i] if col_i < len(headers) else f"col{col_i + 1}"
            item[h] = val
        items.append(item)

    wb.close()
    return items


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 3 — ПАРСИНГ ДАННЫХ ПО КАЖДОМУ АРТИКУЛУ
# ══════════════════════════════════════════════════════════════════════════════

def fetch_product(session: requests.Session, code: str) -> tuple[str, dict]:
    """
    Загружает основную страницу товара galleyp.asp?code=...
    Извлекает: название, бренд, цену, склады, параметры, ARTID.
    Возвращает (html_текст, словарь_данных).
    """
    resp = session.get(PRODUCT_URL, params={"code": code}, timeout=20)
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")

    data: dict = {"code": code}

    # Название
    m = re.search(r"[Нн]аименование[^<]*</td>\s*<td[^>]*>([^<]+)</td>", html, re.I | re.S)
    data["name"] = strip_tags(m.group(1)) if m else ""

    # Производитель / бренд
    m = re.search(r"[Пп]роизводитель[^<]*</td>\s*<td[^>]*>(.*?)</td>", html, re.I | re.S)
    data["brand"] = strip_tags(m.group(1)) if m else ""

    # Цена — в HTML хранится как "2&nbsp;365.00р" внутри <b> рядом с id=mypricespan
    m = re.search(
        r'mypricespan[^>]*>.*?<b>\s*(\d[\d&nbsp;\s\xa0\u00a0]*[\.,]\d{2})',
        html, re.I | re.S
    )
    if m:
        price_str = (m.group(1)
                     .replace("&nbsp;", "").replace("\xa0", "")
                     .replace(" ", "").replace(",", "."))
        data["price"] = float(price_str) if price_str else 0.0
    else:
        data["price"] = 0.0

    # Склады: ищем паттерн "ГородNшт." в тексте страницы
    stock_items = []
    seen_keys   = set()
    for m in re.finditer(
        r"([А-ЯЁ][а-яё]+(?:\s[А-ЯЁ][а-яё]+)?)\s+(\d+)\s*шт",
        strip_tags(html)
    ):
        key = f"{m.group(1)}_{m.group(2)}"
        if key not in seen_keys:
            seen_keys.add(key)
            stock_items.append({"warehouse": m.group(1), "qty": int(m.group(2))})
    data["stock_items"] = stock_items[:10]
    data["stock"]       = ", ".join(f"{s['warehouse']} {s['qty']}шт." for s in stock_items)

    # Параметры — таблица <tbody id=params>
    # Ключи, которые встречаются несколько раз, объединяются через " / "
    params: dict = {}
    m = re.search(r'<tbody\s+id=["\']?params["\']?[^>]*>(.*?)</tbody>', html, re.S | re.I)
    if m:
        for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", m.group(1), re.S | re.I):
            cells = [strip_tags(c)
                     for c in re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S | re.I)]
            cells = [c for c in cells if c]
            if len(cells) >= 2:
                key, val = cells[0], cells[1]
                params[key] = f"{params[key]} / {val}" if key in params else val
    data["params"] = params

    # ARTID — нужен для AJAX-запросов (OEM, применяемость, фото)
    m = re.search(r"ARTID=(\d+)", html)
    data["_artid"] = m.group(1) if m else None

    return html, data


def fetch_oem(session: requests.Session, artid: str, code: str) -> list[dict]:
    """
    Загружает OEM номера через pp0.asp?MODE=OEM.
    Возвращает список вида: [{"manufacturer": "VW", "oem_code": "1K0513029JA"}, ...]
    """
    resp = session.get(
        AJAX_URL,
        params={"MODE": "OEM", "ARTID": artid, "code": code},
        timeout=20
    )
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")

    result = []
    for tr in re.finditer(r"<tr[^>]*>(.*?)</tr>", html, re.S | re.I):
        cells = re.findall(r"<td[^>]*>(.*?)</td>", tr.group(1), re.S | re.I)
        if len(cells) < 2:
            continue
        manufacturer = strip_tags(cells[0])
        oem_code     = strip_tags(cells[1])
        if manufacturer and oem_code and \
                manufacturer.lower() not in ("производитель", "марка", "бренд"):
            result.append({"manufacturer": manufacturer, "oem_code": oem_code})
    return result


def fetch_compatibility(session: requests.Session, artid: str) -> list[dict]:
    """
    Загружает применяемость через pp0.asp?MODE=APL.

    HTML-таблица использует rowspan для марки и модели:
      - 2 rowspan-ячейки → начало новой марки И новой модели
      - 1 rowspan-ячейка → новая модель (марка не меняется)
      - 0 rowspan-ячеек  → очередная модификация той же модели

    Возвращает список:
      [{"brand": "AUDI", "model": "A3 (8P1)", "modification": "1.6 TDI",
        "year_from": "05.2003", "year_to": "08.2012"}, ...]
    """
    resp = session.get(AJAX_URL, params={"MODE": "APL", "ARTID": artid}, timeout=20)
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")

    rows = []
    for tbody in re.finditer(r"<tbody[^>]*>(.*?)</tbody>", html, re.S | re.I):
        current_brand = ""
        current_model = ""

        for tr in re.finditer(r"<tr[^>]*>(.*?)</tr>", tbody.group(1), re.S | re.I):
            td_tags = re.findall(r"<td([^>]*)>(.*?)</td>", tr.group(1), re.S | re.I)
            if not td_tags:
                continue

            rowspan_cells = [(strip_tags(v), attrs)
                             for attrs, v in td_tags if "rowspan" in attrs.lower()]

            if len(rowspan_cells) >= 2:
                current_brand = rowspan_cells[0][0]
                current_model = rowspan_cells[1][0]
            elif len(rowspan_cells) == 1:
                current_model = rowspan_cells[0][0]

            if not current_brand:
                continue

            vals  = [strip_tags(v) for _, v in td_tags]
            years = [v for v in vals if re.match(r"\d{2}\.\d{4}$", v)]
            skip  = {current_brand, current_model}
            mods  = [v for v in vals
                     if v and v not in years and v not in skip
                     and not re.match(r"^\d+(\s+\d+)*\s*$", v)]

            rows.append({
                "brand":        current_brand,
                "model":        current_model,
                "modification": mods[0] if mods else "",
                "year_from":    years[0] if len(years) > 0 else "",
                "year_to":      years[1] if len(years) > 1 else "",
            })
    return rows


def compat_to_text(rows: list[dict]) -> str:
    """
    Преобразует список применяемости в читаемую строку.
    Группирует по марке+модели, берёт крайние годы.
    Пример: "AUDI A3 (8P1) (2003–2012); VW GOLF V (1K1) (2004–2008)"
    """
    if not rows:
        return ""

    groups: dict[str, dict] = {}
    for r in rows:
        key = f"{r['brand']}|{r['model']}"
        if key not in groups:
            groups[key] = {
                "brand":     r["brand"],
                "model":     r["model"],
                "year_from": r["year_from"],
                "year_to":   r["year_to"],
            }
        else:
            g = groups[key]
            if r["year_from"] and (not g["year_from"] or r["year_from"] < g["year_from"]):
                g["year_from"] = r["year_from"]
            if r["year_to"] and (not g["year_to"] or r["year_to"] > g["year_to"]):
                g["year_to"] = r["year_to"]

    parts = []
    for g in groups.values():
        line = f"{g['brand']} {g['model']}".strip()
        yf   = re.sub(r"^\d{2}\.", "", g["year_from"]) if g["year_from"] else ""
        yt   = re.sub(r"^\d{2}\.", "", g["year_to"])   if g["year_to"]   else ""
        if yf or yt:
            line += f" ({yf}–{yt})"
        parts.append(line)
    return "; ".join(parts)


def fetch_cross_refs(session: requests.Session, code: str, qty: int, name: str) -> list[str]:
    """
    Загружает таблицу перекодировок через pp0.asp?Mode=Cross.
    Возвращает список альтернативных кодов из колонки «Код».
    """
    resp = session.get(
        AJAX_URL,
        params={"Mode": "Cross", "code": code, "QTY": qty, "Name": name},
        timeout=20
    )
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")

    codes = []
    # Строки таблицы: <tr class='bld'><td>Бренд</td><td nowrap><a href='...'>КОД</a></td>...
    for tr in re.finditer(r"<tr[^>]*class=['\"]bld['\"][^>]*>(.*?)</tr>", html, re.S | re.I):
        # Код — текст первой ссылки во второй ячейке
        tds = re.findall(r"<td[^>]*>(.*?)</td>", tr.group(1), re.S | re.I)
        if len(tds) >= 2:
            link_text = re.search(r"<a[^>]*>([^<]+)</a>", tds[1], re.I)
            if link_text:
                alt_code = link_text.group(1).strip()
                if alt_code:
                    codes.append(alt_code)
    return codes


def fetch_images(session: requests.Session, artid: str, code: str,
                 images_dir: Path) -> list[str]:
    """
    Загружает изображения из вкладки КАРТИНКИ (pp0.asp?MODE=PIC).
    Сохраняет файлы как {code}_1.jpg, {code}_2.jpg, ... в папку images_dir.
    Дубли (одинаковое содержимое) пропускает.
    Возвращает список имён сохранённых файлов.
    """
    resp = session.get(
        AJAX_URL,
        params={"MODE": "PIC", "CODE": code, "ARTID": artid},
        timeout=20
    )
    resp.raise_for_status()
    pic_html = resp.content.decode("windows-1251", errors="replace")

    img_srcs = re.findall(r"<img[^>]+src=[\'\"]?(wi/img\.asp[^\'\"> \t]+)", pic_html, re.I)
    if not img_srcs:
        return []

    saved  = []
    seen   = set()
    idx    = 1
    for src in img_srcs:
        img_url = f"{BASE_URL}/{src}"
        try:
            r = session.get(img_url, timeout=20)
            r.raise_for_status()
            content = r.content
            # Проверяем формат по magic bytes: JPEG = FFD8, PNG = 89504E47
            if not (content[:2] == b'\xff\xd8' or content[:8] == b'\x89PNG\r\n\x1a\n'):
                continue
            # Пропускаем дубли
            digest = md5(content).hexdigest()
            if digest in seen:
                continue
            seen.add(digest)
            ext      = ".jpg" if content[:2] == b'\xff\xd8' else ".png"
            filename = f"{code}_{idx}{ext}"
            (images_dir / filename).write_bytes(content)
            saved.append(filename)
            idx += 1
        except Exception:
            continue
    return saved


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 4 — СОХРАНЕНИЕ В EXCEL
# ══════════════════════════════════════════════════════════════════════════════

# Структура таблицы: (заголовок колонки, ширина колонки)
COLUMNS = [
    ("Код (Mikado)",        15),
    ("Наименование",        35),
    ("Бренд",               12),
    ("Цена закупки, руб.",  16),
    ("Склады",              38),
    ("Всего, шт.",          10),
    ("Параметры",           45),   # все параметры товара одним блоком
    ("OEM номера",          45),
    ("Применяемость",       60),
    ("Аналог 1 — бренд",    14),
    ("Аналог 1 — код",      20),
    ("Аналог 1 — цена",     13),
    ("Аналог 2 — бренд",    14),
    ("Аналог 2 — код",      20),
    ("Аналог 2 — цена",     13),
    ("Аналог 3 — бренд",    14),
    ("Аналог 3 — код",      20),
    ("Аналог 3 — цена",     13),
    ("Альтернативные артикулы товара", 45),
    ("Изображения",                   30),
]

# Номера колонок с ценой аналогов (для подсветки если дешевле нашей цены)
ANALOG_PRICE_COLS = {12, 15, 18}


def _analog_cells(analogs: list[dict], idx: int) -> list:
    """Возвращает [бренд, код, цена] для аналога с индексом idx, или ['','','']."""
    if idx < len(analogs):
        a = analogs[idx]
        return [a.get("brand", ""), a.get("code", ""), a.get("price", 0)]
    return ["", "", ""]


def save_excel(results: list[dict], out_path: Path):
    """
    Создаёт Excel-файл со сводной таблицей.
    Чётные строки — белые, нечётные — светло-сиреневые.
    Цена аналога дешевле нашей цены — красным жирным.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mikado"

    H_FILL   = PatternFill("solid", fgColor="1A1A2E")   # тёмно-синий заголовок
    H_FONT   = Font(bold=True, color="FFFFFF", size=10)
    ROW_EVEN = PatternFill("solid", fgColor="FFFFFF")
    ROW_ODD  = PatternFill("solid", fgColor="F5F0FF")
    RED_FONT = Font(color="CC0000", bold=True)

    # Заголовки
    for col_i, (title, width) in enumerate(COLUMNS, start=1):
        cell           = ws.cell(1, col_i, title)
        cell.font      = H_FONT
        cell.fill      = H_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 32

    # Данные
    for row_i, r in enumerate(results, start=2):
        fill      = ROW_EVEN if row_i % 2 == 0 else ROW_ODD
        p         = r.get("params", {})
        our_price = r.get("price", 0)

        stock_text  = ", ".join(
            f"{s['warehouse']} {s['qty']}шт." for s in r.get("stock_items", [])
        )
        total_stock = sum(s["qty"] for s in r.get("stock_items", []))

        # Параметры: каждая пара "ключ: значение" на отдельной строке
        params_text = "\n".join(f"{k}: {v}" for k, v in p.items()) if p else ""

        # OEM: "AUDI: 1K0513029JA; VW: 1K0513029JA; ..."  (не более 10 штук)
        oem_details = r.get("oem_details", [])
        oem_text    = "; ".join(
            f"{d['manufacturer']}: {d['oem_code']}" for d in oem_details[:10]
        ) if oem_details else "; ".join(r.get("oem_numbers", [])[:10])

        # Аналоги — сортируем по цене, дешевле первые
        analogs = sorted(
            [a for a in r.get("analogs", []) if a.get("price", 0) > 0],
            key=lambda x: x["price"]
        )

        cross_refs_text = "; ".join(r.get("cross_refs", []))

        row_vals = [
            r.get("code",  ""),
            r.get("name",  ""),
            r.get("brand", ""),
            our_price,
            stock_text,
            total_stock,
            params_text,
            oem_text,
            r.get("compatibility", ""),
            *_analog_cells(analogs, 0),
            *_analog_cells(analogs, 1),
            *_analog_cells(analogs, 2),
            cross_refs_text,
            ", ".join(r.get("images", [])),
        ]

        for col_i, val in enumerate(row_vals, start=1):
            cell           = ws.cell(row_i, col_i, val)
            cell.fill      = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Подсветка: аналог дешевле нашей цены
            if col_i in ANALOG_PRICE_COLS and isinstance(val, (int, float)):
                if val > 0 and val < our_price:
                    cell.font = RED_FONT

    ws.freeze_panes = "A2"
    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Сбор данных с mikado-parts.ru",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--price",    required=True,           help="Путь к Excel прайсу")
    ap.add_argument("--rows",     default=None,            help="Диапазон строк, например 2-20")
    ap.add_argument("--out",      default=None,            help="Папка для результатов")
    ap.add_argument("--login",    default=None,            help="Код клиента Mikado")
    ap.add_argument("--password", default=None,            help="Пароль")
    ap.add_argument("--delay",    default=1.5, type=float, help="Пауза между запросами (сек)")
    ap.add_argument("--env",      default=None,            help="Путь к .env файлу")
    args = ap.parse_args()

    # ── Пути ──────────────────────────────────────────────────────────────────
    price_file = Path(args.price)
    if not price_file.exists():
        print(f"Файл не найден: {price_file}")
        sys.exit(1)

    out_dir     = Path(args.out) if args.out else price_file.parent / "mikado_output"
    images_dir  = out_dir / "images"
    result_file = out_dir / "mikado_data.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)
    images_dir.mkdir(exist_ok=True)

    # ── Логин и пароль ────────────────────────────────────────────────────────
    env_path   = Path(args.env) if args.env else Path(__file__).parent.parent / ".env"
    env        = load_env(env_path)
    login_code = args.login    or env.get("MIKADO_CODE",     "")
    password   = args.password or env.get("MIKADO_PASSWORD", "")

    if not login_code or not password:
        print("Укажите логин и пароль: --login XXXXX --password XXXXX")
        print("  или добавьте MIKADO_CODE и MIKADO_PASSWORD в .env файл")
        sys.exit(1)

    # ── Диапазон строк ────────────────────────────────────────────────────────
    if args.rows:
        row_start, row_end = map(int, args.rows.split("-"))
    else:
        wb        = openpyxl.load_workbook(price_file, read_only=True)
        row_end   = wb.active.max_row or 1000
        row_start = 2
        wb.close()

    # ── Шапка ─────────────────────────────────────────────────────────────────
    print()
    print("═" * 60)
    print("  Mikado Scraper")
    print(f"  Прайс:  {price_file.name}  (строки {row_start}–{row_end})")
    print(f"  Папка:  {out_dir}")
    print("═" * 60)

    # ── [1/4] Читаем прайс ───────────────────────────────────────────────────
    print("\n[1/4] Читаю прайс...")
    items = read_price(price_file, row_start, row_end)
    print(f"  Найдено артикулов: {len(items)}")
    for it in items:
        print(f"    {it['_code']}")

    # ── [2/4] Авторизация ─────────────────────────────────────────────────────
    print("\n[2/4] Авторизация...")
    session = login(login_code, password)

    # ── [3/4] Сбор данных ─────────────────────────────────────────────────────
    print(f"\n[3/4] Обрабатываю {len(items)} артикулов...")
    results = []

    for i, item in enumerate(items, start=1):
        code = item["_code"]
        print(f"\n  [{i:2}/{len(items)}] {code}")

        try:
            # Основная карточка
            _, data = fetch_product(session, code)
            data["_price_row"] = item
            print(f"    {data['name'][:45]}")
            print(f"    {data['brand']}  |  {data['price']:.0f} руб.  "
                  f"|  {sum(s['qty'] for s in data['stock_items'])} шт. на складе")
            if data["params"]:
                print(f"    Параметры: {len(data['params'])} полей")

            artid = data.get("_artid")
            if artid:
                # OEM номера
                time.sleep(0.4)
                oem = fetch_oem(session, artid, code)
                data["oem_details"] = oem
                data["oem_numbers"] = [r["oem_code"] for r in oem]
                print(f"    OEM: {len(oem)} номеров")

                # Применяемость
                time.sleep(0.4)
                compat_rows = fetch_compatibility(session, artid)
                data["compatibility_rows"] = compat_rows
                data["compatibility"]      = compat_to_text(compat_rows)
                print(f"    Применяемость: {len(compat_rows)} строк")

                # Изображения
                time.sleep(0.4)
                imgs = fetch_images(session, artid, code, images_dir)
                data["images"] = imgs
                print(f"    Фото: {len(imgs)} шт."
                      + (f"  →  {', '.join(imgs)}" if imgs else ""))

            # Таблица перекодировок — не зависит от ARTID
            time.sleep(0.4)
            qty  = sum(s["qty"] for s in data.get("stock_items", []))
            refs = fetch_cross_refs(session, code, qty, data.get("name", ""))
            data["cross_refs"] = refs
            print(f"    Перекодировки: {len(refs)} кодов")

            if not artid:
                data.setdefault("oem_details",        [])
                data.setdefault("oem_numbers",        [])
                data.setdefault("compatibility_rows", [])
                data.setdefault("compatibility",      "")
                data.setdefault("images",             [])
                print("    ARTID не найден — OEM / применяемость / фото пропущены")

            results.append(data)

            # JSON по каждому артикулу
            json_path = out_dir / f"{code}.json"
            with open(json_path, "w", encoding="utf-8") as f:
                export = {k: v for k, v in data.items() if not k.startswith("_")}
                json.dump(export, f, ensure_ascii=False, indent=2)

        except Exception as e:
            print(f"    ✗ Ошибка: {e}")

        if i < len(items):
            time.sleep(args.delay)

    # ── [4/4] Сохраняем таблицу ───────────────────────────────────────────────
    print(f"\n[4/4] Сохраняю таблицу...")
    save_excel(results, result_file)
    print(f"  ✓ {result_file}")

    # ── Итог ──────────────────────────────────────────────────────────────────
    total_imgs = sum(len(r.get("images", [])) for r in results)
    print()
    print("═" * 60)
    print(f"  Готово!")
    print(f"  Обработано:   {len(results)} / {len(items)} артикулов")
    print(f"  Таблица:      {result_file}")
    print(f"  Изображения:  {images_dir}  ({total_imgs} файлов)")
    print(f"  JSON:         {out_dir}  ({len(results)} файлов)")
    print("═" * 60)
    print()


if __name__ == "__main__":
    main()
