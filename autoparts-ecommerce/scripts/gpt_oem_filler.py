"""
gpt_oem_filler.py
═════════════════
Заполняет «OEM номера» (col 7) и «Альтернативные артикулы» (col 9)
в файле Топ-500 ВБ.xlsx через основной Chrome (уже залогинен в ChatGPT).

ВАЖНО: перед запуском закрой Chrome полностью.

Основной запуск:
  uv run --with playwright,openpyxl scripts/gpt_oem_filler.py

Тест на 5 строках:
  uv run --with playwright,openpyxl scripts/gpt_oem_filler.py --rows 2-6

Продолжение прерванной обработки:
  Запустить снова — пропустит строки где col 7 уже заполнен.
"""

import os
import re
import sys
import time
import argparse
import subprocess
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Playwright не установлен. Запусти:")
    print("  uv run --with playwright python -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side
except ImportError:
    print("openpyxl не установлен.")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_FILE  = Path(r"C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ_new.xlsx")
PROFILE_DIR = Path(os.environ["USERPROFILE"]) / "AppData" / "Local" / "Google" / "Chrome" / "User Data"
CHATGPT_URL = "https://chatgpt.com"

BATCH_SIZE = 3    # артикулов за один запрос к ChatGPT
SAVE_EVERY = 5    # сохранять Excel каждые N батчей (~15 строк)

COL_ARTICLE = 1
COL_NAME    = 2
COL_BRAND   = 3
COL_OEM     = 7
COL_APPLIC  = 8
COL_ALTS    = 9

SEL_INPUT    = '#prompt-textarea'
SEL_SEND_BTN = 'button[data-testid="send-button"]'
SEL_STOP_BTN = 'button[aria-label="Stop streaming"]'
SEL_RESPONSE = '[data-message-author-role="assistant"]'

RATE_LIMIT_PHRASES = [
    "достигнут лимит", "message limit", "you've reached",
    "rate limit", "попробуйте позже", "try again later", "upgrade",
]
SERVER_ERROR_PHRASES = [
    "something went wrong", "if this issue persists",
    "unexpected error", "server error",
]


# ══════════════════════════════════════════════════════════════════════════════
#  ПРОМПТ
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_INSTRUCTION = """Я буду присылать список артикулов автозапчастей с названиями.

Для каждой детали:
- Найди точные OEM-номера (оригинальные номера от производителей автомобилей: Honda, VW, Hyundai, BMW, Renault, Lada, Toyota и т.д.)
- Используй название детали как главную подсказку — там указаны автомобили и двигатели для которых она подходит
- Найди максимально полный список аналогов (кросс-номеров) от других производителей запчастей
- Проверяй, чтобы кроссы относились именно к указанной детали, а не к похожим артикулам
- Удаляй дубликаты
- ВАЖНО: если название явно указывает на автомобиль (напр. "VESTA, LARGUS, X-RAY ДВ.РЕНО") — ищи OEM именно этого производителя

Формат ответа для КАЖДОЙ детали — строго следующий:

[N]
OEM-номера:
номер1; номер2; номер3

Альтернативные артикулы:
артикул1; артикул2; артикул3

Правила:
- Никаких таблиц, пояснений, ссылок, комментариев
- Только блоки [N] с двумя строками
- Номера разделять исключительно точкой с запятой и пробелом
- Сохранять оригинальное написание артикулов
- Если OEM-номер в нескольких форматах (с дефисом и без) — указывать все варианты
- Для аналогов собирать максимально полный список из TecDoc, производителей и OEM-каталогов
- Если OEM-номеров нет — писать: OEM-номера: —
- Если аналогов нет — писать: Альтернативные артикулы: —

Пример ответа:

[1]
OEM-номера:
58101C8A00; 58101-C8A00; 58101C8A50

Альтернативные артикулы:
BD3636; GDB3630; PN0537; 0986494563

[2]
OEM-номера:
—

Альтернативные артикулы:
W68; OC90; HU716X
"""


def build_prompt(items: list[dict]) -> str:
    lines = [SYSTEM_INSTRUCTION.strip(), "", "Список деталей:"]
    for item in items:
        lines.append(f"[{item['idx']}] Артикул: {item['article']} / Бренд: {item['brand']} / Название: {item['name']}")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  ПАРСИНГ ОТВЕТА
# ══════════════════════════════════════════════════════════════════════════════

def parse_response(text: str, n_items: int) -> dict[int, dict]:
    """
    Парсит текст вида:
        [1]
        OEM-номера:
        58101C8A00; 58101-C8A00

        Альтернативные артикулы:
        BD3636; GDB3630

    Возвращает {idx: {"oem": str, "alts": str}}
    """
    result = {}

    # Разбиваем на блоки по [N]
    blocks = re.split(r'\[(\d+)\]', text)
    # blocks = ['pretext', '1', 'block1 text', '2', 'block2 text', ...]

    i = 1
    while i < len(blocks) - 1:
        idx_str = blocks[i].strip()
        block   = blocks[i + 1] if i + 1 < len(blocks) else ''
        i += 2

        try:
            idx = int(idx_str)
        except ValueError:
            continue

        oem  = _extract_field(block, 'OEM-номера')
        alts = _extract_field(block, 'Альтернативные артикулы')

        result[idx] = {'oem': oem, 'alts': alts}

    return result


def _extract_field(block: str, label: str) -> str:
    """Извлекает значение поля по метке. Возвращает '' если не найдено или '—'."""
    pattern = re.compile(
        rf'{re.escape(label)}\s*:\s*\n?(.*?)(?=\n[A-ZА-ЯЁ][^\n]*:|$)',
        re.DOTALL | re.IGNORECASE
    )
    m = pattern.search(block)
    if not m:
        return ''
    value = m.group(1).strip()
    # Убираем маркер "нет данных"
    if value in ('—', '-', '–', 'нет', 'нет данных', 'н/д', 'N/A', 'n/a'):
        return ''
    # Очищаем пробелы, лишние переносы
    value = re.sub(r'\s*\n\s*', ' ', value).strip()
    return value


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ХЕЛПЕРЫ
# ══════════════════════════════════════════════════════════════════════════════

THIN = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def set_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value or None)
    cell.font      = Font(size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border    = THIN


# ══════════════════════════════════════════════════════════════════════════════
#  ВЗАИМОДЕЙСТВИЕ С БРАУЗЕРОМ
# ══════════════════════════════════════════════════════════════════════════════

def is_logged_in(page) -> bool:
    """Проверяет что пользователь залогинен в ChatGPT."""
    try:
        # Кнопка логина видна только для гостей
        login_visible = page.locator('[data-testid="login-button"]').count() > 0
        return not login_visible
    except Exception:
        return False


def wait_for_login(page):
    """Ждёт пока пользователь залогинится вручную."""
    print("\n" + "="*55)
    print("  ChatGPT не залогинен!")
    print("  Войди в аккаунт в открытом окне Chrome.")
    print("  Скрипт продолжит автоматически после входа.")
    print("="*55)
    for i in range(300):  # ждём до 5 минут
        time.sleep(2)
        if is_logged_in(page):
            print("  Вход выполнен! Продолжаю...")
            time.sleep(2)
            return
        if i % 15 == 14:
            remaining = (300 - i - 1) * 2
            print(f"  Жду входа... (осталось {remaining} сек)")
    raise RuntimeError("Вход в ChatGPT не выполнен за 5 минут")


def enable_web_search(page) -> bool:
    """Включает веб-поиск через меню + → 'Найди что-то'. Возвращает True если успешно."""
    # Уже включён?
    if page.locator('button[aria-label*="Поиск, нажмите"]').count() > 0:
        return True

    try:
        # Шаг 1: клик "+"
        plus = page.locator('[data-testid="composer-plus-btn"]')
        if plus.count() == 0:
            return False
        plus.click()
        time.sleep(1)

        # Шаг 2: выбрать "Найди что-то"
        item = page.locator('text="Найди что-то"').first
        if item.count() == 0:
            page.keyboard.press('Escape')
            return False
        item.click()
        time.sleep(0.8)

        # Проверяем что чип "Поиск" появился
        active = page.locator('button[aria-label*="Поиск, нажмите"]').count() > 0
        if active:
            print("  Веб-поиск включён")
        return active
    except Exception as e:
        print(f"  Не удалось включить поиск: {e}")
        return False


def close_modal_if_present(page) -> bool:
    """Закрывает модальное окно (modal-conversation-history-rate-limit) если оно блокирует клики."""
    modal_sel = '[data-testid="modal-conversation-history-rate-limit"]'
    if page.locator(modal_sel).count() == 0:
        return False
    print("  Модальное окно — закрываю...")
    for close_sel in [
        f'{modal_sel} button',
        'button[aria-label="Close"]',
        'button[aria-label="Закрыть"]',
        'button:has-text("OK")',
        'button:has-text("Got it")',
        'button:has-text("Понятно")',
        'button:has-text("Dismiss")',
    ]:
        btn = page.locator(close_sel).first
        if btn.count() > 0:
            try:
                btn.click(timeout=3000, force=True)
                time.sleep(1)
                if page.locator(modal_sel).count() == 0:
                    print("  Модальное окно закрыто.")
                    return True
            except Exception:
                continue
    page.keyboard.press('Escape')
    time.sleep(1)
    if page.locator(modal_sel).count() == 0:
        print("  Модальное окно закрыто (Escape).")
        return True
    return False


def is_rate_limited(page) -> bool:
    try:
        body = page.locator("body").inner_text(timeout=2000).lower()
        return any(p in body for p in RATE_LIMIT_PHRASES)
    except Exception:
        return False


def wait_for_rate_limit_reset(page):
    print(f"\n  Лимит сообщений — жду 65 мин...")
    for remaining in range(65, 0, -1):
        print(f"\r  Осталось: {remaining} мин.  ", end="", flush=True)
        time.sleep(60)
    print("\r  Продолжаю...                    ")
    try:
        page.goto(CHATGPT_URL, wait_until="networkidle", timeout=40000)
        time.sleep(3)
    except Exception:
        pass


def _open_new_chat(page):
    """Гарантированно открывает свежий чат (без истории предыдущих сообщений)."""
    page.goto(CHATGPT_URL, wait_until='domcontentloaded', timeout=30000)
    time.sleep(2)
    close_modal_if_present(page)

    # Если редиректнуло в конкретный чат — ищем «New chat» кнопку
    for attempt in range(3):
        n_responses = page.locator(SEL_RESPONSE).count()
        if n_responses == 0:
            return  # Уже чистый чат

        # Пробуем разные селекторы кнопки «New chat»
        new_chat_sels = [
            'button[aria-label="New chat"]',
            'a[aria-label="New chat"]',
            'button[aria-label*="new chat" i]',
            'a[href="/"]',
            'button[data-testid="create-new-chat-button"]',
            'nav a[href="/"]',
        ]
        clicked = False
        for sel in new_chat_sels:
            el = page.locator(sel).first
            if el.count() > 0:
                try:
                    el.click(timeout=3000)
                    time.sleep(2)
                    clicked = True
                    break
                except Exception:
                    continue

        if not clicked:
            # Последний вариант: принудительная навигация
            page.goto(CHATGPT_URL, wait_until='domcontentloaded', timeout=20000)
            time.sleep(2)
            break


def send_message(page, text: str) -> str:
    """Отправляет сообщение и ждёт полного ответа. Возвращает текст ответа."""
    # Открываем НОВЫЙ чат
    _open_new_chat(page)
    print(f"  URL: {page.url[:60]}")

    # Закрываем модальное окно если оно блокирует интерфейс
    close_modal_if_present(page)

    # Ждём поля ввода
    try:
        page.wait_for_selector(SEL_INPUT, timeout=20000)
    except PWTimeout:
        print(f"  Заголовок страницы: {page.title()}")
        raise RuntimeError("Поле ввода не найдено — возможно, нужен повторный логин")

    # Закрываем ещё раз на случай если появилось после загрузки
    close_modal_if_present(page)

    # Включаем веб-поиск если доступен
    enable_web_search(page)

    # Считаем существующие ответы ДО отправки (в новом чате должно быть 0)
    n_before = page.locator(SEL_RESPONSE).count()
    if n_before > 0:
        print(f"  Внимание: найдено {n_before} старых ответов в чате")

    # Очищаем поле ввода и вставляем текст
    close_modal_if_present(page)
    input_el = page.locator(SEL_INPUT).first
    input_el.click()
    time.sleep(0.3)
    page.keyboard.press('Control+a')
    page.keyboard.press('Delete')
    time.sleep(0.2)

    # Clipboard paste через JS (работает надёжно в CDP)
    page.evaluate(
        """(text) => {
            const el = document.querySelector('#prompt-textarea');
            if (!el) return;
            el.focus();
            const dt = new DataTransfer();
            dt.setData('text/plain', text);
            el.dispatchEvent(new ClipboardEvent('paste', {clipboardData: dt, bubbles: true}));
        }""",
        text
    )
    time.sleep(0.8)

    # Проверяем количество введённых символов
    val = page.evaluate("() => { const el = document.querySelector('#prompt-textarea'); return el ? (el.value || el.innerText || el.textContent) : ''; }")
    print(f"  Введено символов: {len(val or '')}")

    if not val or len(val.strip()) < 10:
        raise RuntimeError("Текст не был введён в поле")

    # Отправляем
    send_btn = page.locator(SEL_SEND_BTN)
    if send_btn.count() > 0:
        send_btn.first.click()
    else:
        print("  Кнопка Send не найдена, нажимаем Enter")
        input_el.press('Enter')
    print("  Сообщение отправлено")

    # Ждём появления НОВОГО ответа ассистента (с веб-поиском может быть дольше)
    timeout_s = 90
    start_wait = time.time()
    while time.time() - start_wait < timeout_s:
        n_now = page.locator(SEL_RESPONSE).count()
        if n_now > n_before:
            break
        time.sleep(2)
    else:
        print(f"  Ответ не появился за {timeout_s} сек — проверяю текущие ответы")

    # Ждём завершения генерации: ответ стабилен И нет stop-кнопки
    stop_sels = ', '.join([
        'button[aria-label="Stop streaming"]',
        'button[aria-label="Stop generating"]',
        'button[data-testid="stop-button"]',
        'button[data-testid="fruitjuice-stop-button"]',
    ])
    last_text  = ''
    stable_cnt = 0
    start_gen  = time.time()
    while time.time() - start_gen < 360:
        time.sleep(3)
        stop_visible  = page.locator(stop_sels).count() > 0
        responses_now = page.locator(SEL_RESPONSE).all()
        curr_text     = responses_now[-1].inner_text(timeout=8000) if responses_now else ''

        if not stop_visible and curr_text == last_text and len(curr_text) > 30:
            stable_cnt += 1
            if stable_cnt >= 3:  # 3 проверки по 3 сек = 9 сек стабильности
                break
        else:
            stable_cnt = 0
            last_text  = curr_text

    elapsed = time.time() - start_gen
    print(f"  Генерация: {elapsed:.0f} сек")

    # Забираем ответ
    responses = page.locator(SEL_RESPONSE).all()
    print(f"  Найдено ответов: {len(responses)}")
    if len(responses) <= n_before:
        print("  Новый ответ не появился!")
        return ''
    return responses[-1].inner_text(timeout=15000)


# ══════════════════════════════════════════════════════════════════════════════
#  ОСНОВНАЯ ЛОГИКА
# ══════════════════════════════════════════════════════════════════════════════

CDP_PORT = 9222

def wait_for_cdp():
    """Ждёт пока CDP порт станет доступен. Если нет — просит запустить bat-файл."""
    import urllib.request
    try:
        urllib.request.urlopen(f'http://127.0.0.1:{CDP_PORT}/json/version', timeout=2)
        print("CDP доступен — подключаюсь к Chrome.")
        return
    except Exception:
        pass

    print()
    print("=" * 60)
    print("  Chrome с отладочным портом не найден.")
    print()
    print("  Запусти один раз:")
    print(r"  scripts\launch_chatgpt.bat")
    print()
    print("  Это откроет Chrome с ChatGPT. Войди в аккаунт.")
    print("  Скрипт продолжит автоматически.")
    print("=" * 60)

    for attempt in range(300):  # ждём до 5 мин
        time.sleep(2)
        try:
            urllib.request.urlopen(f'http://127.0.0.1:{CDP_PORT}/json/version', timeout=2)
            print("Chrome готов — продолжаю.")
            return
        except Exception:
            if attempt % 15 == 14:
                print(f"  Жду Chrome... ({(attempt+1)*2} сек)")
    raise RuntimeError("Chrome с CDP портом не появился за 5 минут")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--rows', type=str, default=None, help='Диапазон строк Excel, напр. 2-20')
    parser.add_argument('--retry-dash', action='store_true',
                        help='Перепрогнать строки где OEM = "—"')
    parser.add_argument('--retry-alts', action='store_true',
                        help='Перепрогнать строки где Альтернативные = "—"')
    parser.add_argument('--file', type=str, default=None,
                        help='Путь к Excel файлу (по умолчанию EXCEL_FILE в коде)')
    args = parser.parse_args()

    if args.file:
        global EXCEL_FILE
        EXCEL_FILE = Path(args.file)

    wait_for_cdp()

    with sync_playwright() as pw:
        browser = pw.chromium.connect_over_cdp(f'http://127.0.0.1:{CDP_PORT}')
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        # Ищем вкладку ChatGPT, если она уже открыта
        chatgpt_page = None
        for p in context.pages:
            if 'chatgpt.com' in p.url:
                chatgpt_page = p
                break
        page = chatgpt_page or context.pages[0] if context.pages else context.new_page()

        # Открываем ChatGPT
        print("Открываю ChatGPT...")
        page.goto(CHATGPT_URL, wait_until='domcontentloaded', timeout=60000)
        time.sleep(3)

        # Проверяем логин
        if not is_logged_in(page):
            wait_for_login(page)

        if is_rate_limited(page):
            wait_for_rate_limit_reset(page)

        # Загружаем Excel
        print(f"Читаю: {EXCEL_FILE}")
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.worksheets[0]

        # Определяем строки для обработки
        row_range = None
        if args.rows:
            lo, hi = map(int, args.rows.split('-'))
            row_range = (lo, hi)

        todos = []
        for r in range(2, 502):
            if row_range and not (row_range[0] <= r <= row_range[1]):
                continue
            article = ws.cell(r, COL_ARTICLE).value
            if not article:
                continue
            oem_val  = str(ws.cell(r, COL_OEM).value  or '').strip()
            alts_val = str(ws.cell(r, COL_ALTS).value or '').strip()
            if args.retry_alts:
                # Перепрогон по аналогам: берём строки где alts = "—"
                if alts_val != '—':
                    continue
            elif args.retry_dash:
                # Перепрогон по OEM: берём строки где oem = "—"
                if oem_val != '—':
                    continue
            else:
                # Обычный режим: пропускаем уже заполненные
                if oem_val:
                    continue

            todos.append(r)

        total = len(todos)
        print(f"Строк для обработки: {total}\n")

        if not total:
            print("Нечего обрабатывать — все строки уже заполнены.")
            browser.close()
            return

        processed  = 0
        save_timer = 0

        for batch_start in range(0, total, BATCH_SIZE):
            batch_rows = todos[batch_start: batch_start + BATCH_SIZE]

            # Собираем items для промпта
            items = []
            for local_idx, r in enumerate(batch_rows, start=1):
                items.append({
                    'idx':     local_idx,
                    'article': str(ws.cell(r, COL_ARTICLE).value or '').strip(),
                    'brand':   str(ws.cell(r, COL_BRAND).value   or '').strip(),
                    'name':    str(ws.cell(r, COL_NAME).value     or '').strip()[:120],
                })

            prompt = build_prompt(items)

            print(f"Батч {batch_start//BATCH_SIZE + 1}: строки {batch_rows[0]-1}–{batch_rows[-1]-1}  ({len(batch_rows)} арт.)")

            # Проверка на rate limit перед отправкой
            if is_rate_limited(page):
                wait_for_rate_limit_reset(page)

            try:
                response_text = send_message(page, prompt)
            except Exception as e:
                print(f"  Ошибка отправки: {e}")
                time.sleep(10)
                continue

            if not response_text:
                print(f"  Пустой ответ — пропускаю батч")
                continue

            # Проверяем на server error
            body_lower = response_text.lower()
            if any(p in body_lower for p in SERVER_ERROR_PHRASES):
                print(f"  Ошибка сервера — повтор через 15 сек")
                time.sleep(15)
                continue

            # Парсим ответ
            parsed = parse_response(response_text, len(batch_rows))

            if not parsed:
                print(f"  Не удалось распарсить ответ:")
                print(f"  {response_text[:200]}")
                continue

            # Записываем в Excel
            for local_idx, r in enumerate(batch_rows, start=1):
                data    = parsed.get(local_idx, {})
                article = str(ws.cell(r, COL_ARTICLE).value or '').strip()
                oem     = data.get('oem',  '')
                alts    = data.get('alts', '')

                # Убираем исходный артикул из списка аналогов
                if alts and article:
                    parts = [p.strip() for p in alts.split(';') if p.strip().upper() != article.upper()]
                    alts  = '; '.join(parts)

                # Пишем OEM только если не было реальных данных (не затираем хорошие)
                cur_oem = str(ws.cell(r, COL_OEM).value or '').strip()
                if oem and cur_oem in ('', '—'):
                    set_cell(ws, r, COL_OEM, oem)
                elif not oem and cur_oem in ('', '—'):
                    set_cell(ws, r, COL_OEM, '—')

                # При --retry-dash не трогаем alts если там уже хорошие данные
                cur_alts = str(ws.cell(r, COL_ALTS).value or '').strip()
                if args.retry_dash and cur_alts not in ('', '—'):
                    pass  # Сохраняем существующие alts
                else:
                    set_cell(ws, r, COL_ALTS, alts or '—')

                print(f"  [{r-1:3d}] {article[:18]:18s}  OEM: {oem[:40]:40s}  Alts: {alts[:35]}")

            processed  += len(batch_rows)
            save_timer += 1

            if save_timer >= SAVE_EVERY:
                wb.save(EXCEL_FILE)
                print(f"  >> Сохранено ({processed}/{total})\n")
                save_timer = 0

            # Пауза между батчами (с запасом для веб-поиска)
            time.sleep(20)

        # Финальное сохранение
        wb.save(EXCEL_FILE)
        print(f"\nГотово! Обработано: {processed} строк")
        print(f"Файл сохранён: {EXCEL_FILE}")
        # Chrome НЕ закрываем — пользователь продолжает работу


if __name__ == '__main__':
    main()
