"""
ozon_order_sync.py
══════════════════════════════════════════════════════════════════════════════
Мониторит новые FBS-заказы напрямую из Ozon API и размещает заказ у Микадо.

Поток:
  Ozon API (awaiting_packaging) → [этот скрипт] → Микадо + Telegram

Логика (каждые POLL_INTERVAL_MINUTES минут):
  1. GET /v3/posting/fbs/list?status=awaiting_packaging — все неупакованные заказы
  2. Фильтруем уже обработанные по posting_number (хранится в data/order_sync_state.json)
  3. Для каждого нового заказа:
     a. Извлечь позиции: offer_id без суффикса "-con" → код Mikado
     b. Проверить наличие по актуальному прайсу Mikado
     c. Авторизоваться на mikado-parts.ru → для каждой позиции:
        - ввести каталожный номер в поиск → перейти на карточку
        - проверить наличие Волгоград → заполнить «Заказ:» → нажать «Заказать»
     d. Отправить уведомление в Telegram (заказ + результат)
     e. При ошибке Mikado → Telegram-алерт для ручной обработки
  4. Сохранить обработанные posting_number

Переменные .env:
    OZON_CLIENT_ID=...
    OZON_API_KEY=...
    MIKADO_CODE=35275
    MIKADO_PASSWORD=...
    TG_BOT_TOKEN=...
    TG_CHAT_ID=...

Запуск (непрерывный, раз в 15 мин):
  uv run --with requests,openpyxl scripts/ozon_order_sync.py

Разовый / тест:
  uv run --with requests,openpyxl scripts/ozon_order_sync.py --once
  uv run --with requests,openpyxl scripts/ozon_order_sync.py --once --dry-run

ВАЖНО — форма заказа Mikado:
  HTML-парсинг страниц mikado-parts.ru подобран эвристически.
  Если заказ не проходит: F12 → Network → выполните поиск и нажмите «Заказать» →
  найдите POST-запрос, скопируйте имена полей и action URL.
  Скорректируйте параметры в mikado_search_and_order().
"""

import sys
import io
import json
import time
import logging
import argparse
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Установи зависимости: uv run --with requests,openpyxl scripts/ozon_order_sync.py")
    sys.exit(1)

try:
    from telegram_notify import tg_order, tg_mikado_error, tg_alert
    _TG_OK = True
except ImportError:
    _TG_OK = False

try:
    from daemon_guard import single_instance
except ImportError:
    def single_instance(_): pass

from mikado_price_fetcher import download_mikado_price_bytes

# ─── Константы ────────────────────────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent.parent
ENV_FILE       = BASE_DIR / ".env"
STATE_FILE     = BASE_DIR / "data" / "order_sync_state.json"
LOG_FILE       = BASE_DIR / "logs" / "ozon_order_sync.log"
ORDERS_DIR     = BASE_DIR / "data" / "orders"

MIKADO_PRICE_URL  = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=YOUR_MIKADO_PRICE_KEY"
)
MIKADO_LOGIN_URL  = "https://mikado-parts.ru/office/SECURE.asp"
MIKADO_SEARCH_URL = "https://mikado-parts.ru/office/galleyp.asp"
MIKADO_ORDER_URL  = "https://mikado-parts.ru/office/pp0.asp"

OZON_API_BASE         = "https://api-seller.ozon.ru"
ARTICLE_SUFFIX        = "-con"
POLL_INTERVAL_MINUTES = 15
MAX_PROCESSED_RECORDS = 5000  # не хранить больше N posting_number в state

# ─── Логирование ──────────────────────────────────────────────────────────────
for d in (LOG_FILE.parent, ORDERS_DIR, STATE_FILE.parent):
    d.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ─── .env ─────────────────────────────────────────────────────────────────────
def load_env() -> dict:
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


# ─── State ────────────────────────────────────────────────────────────────────
def load_state() -> dict:
    if not STATE_FILE.exists():
        return {"processed": []}
    try:
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {"processed": []}


def save_state(state: dict) -> None:
    try:
        STATE_FILE.write_text(
            json.dumps(state, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as e:
        log.error(f"Ошибка сохранения state: {e}")


# ─── Ozon: получение заказов ──────────────────────────────────────────────────
def _ozon_headers(client_id: str, api_key: str) -> dict:
    return {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
    }


def get_ozon_new_orders(client_id: str, api_key: str) -> list[dict]:
    """
    Возвращает все FBS-отправления со статусом awaiting_packaging.
    Это заказы, которые нужно упаковать → именно их нужно закупить у Mikado.
    """
    if not client_id or not api_key:
        log.warning("OZON_CLIENT_ID / OZON_API_KEY не заданы — пропускаем опрос Ozon")
        return []

    from datetime import timezone

    all_postings: list[dict] = []
    offset = 0
    # Ozon API требует диапазон дат; берём последние 30 дней
    now_utc   = datetime.now(timezone.utc)
    since_str = (now_utc - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%SZ")
    to_str    = now_utc.strftime("%Y-%m-%dT%H:%M:%SZ")

    while True:
        try:
            resp = requests.post(
                f"{OZON_API_BASE}/v3/posting/fbs/list",
                headers=_ozon_headers(client_id, api_key),
                json={
                    "dir":    "asc",
                    "filter": {
                        "status": "awaiting_packaging",
                        "since":  since_str,
                        "to":     to_str,
                    },
                    "limit":  50,
                    "offset": offset,
                    "with":   {"analytics_data": False, "financial_data": False},
                },
                timeout=30,
            )
            resp.raise_for_status()
            result   = resp.json().get("result", {})
            postings = result.get("postings", [])
            all_postings.extend(postings)
            if not result.get("has_next"):
                break
            offset += 50
        except Exception as e:
            log.error(f"Ozon: ошибка получения заказов: {e}")
            break

    log.info(f"Ozon: получено {len(all_postings)} заказов (awaiting_packaging)")
    return all_postings


def parse_ozon_items(posting: dict) -> list[dict]:
    """Извлекает позиции из Ozon-отправления."""
    items = []
    for prod in posting.get("products", []):
        offer_id = str(prod.get("offer_id", "")).strip()
        if not offer_id:
            continue
        mikado_code = offer_id.removesuffix(ARTICLE_SUFFIX)
        items.append({
            "offer_id":    offer_id,
            "mikado_code": mikado_code,
            "name":        prod.get("name", ""),
            "quantity":    int(prod.get("quantity", 1)),
            "price_rub":   float(prod.get("price", 0)),
        })
    return items


# ─── Прайс Mikado ─────────────────────────────────────────────────────────────
def load_mikado_price() -> dict[str, dict]:
    """Загружает прайс. Возвращает {code: {qty, price, name}}."""
    content = download_mikado_price_bytes(MIKADO_PRICE_URL, log)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    ws     = wb.active
    rows   = ws.iter_rows(values_only=True)
    header = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]

    code_idx = qty_idx = price_idx = name_idx = None
    for i, h in enumerate(header):
        if h == "code":       code_idx  = i
        elif h == "qty":      qty_idx   = i
        elif h == "priceout": price_idx = i
        elif h == "prodname": name_idx  = i

    if code_idx is None:
        wb.close()
        log.error("Mikado: колонка Code не найдена")
        return {}

    db: dict[str, dict] = {}
    for row in rows:
        raw = row[code_idx] if len(row) > code_idx else None
        if not raw:
            continue
        code  = str(raw).strip()
        qty   = 0
        price = 0.0
        name  = ""
        if qty_idx   is not None and len(row) > qty_idx:
            try:    qty   = max(0, int(float(str(row[qty_idx] or 0))))
            except: pass
        if price_idx is not None and len(row) > price_idx:
            try:    price = float(str(row[price_idx] or 0))
            except: pass
        if name_idx  is not None and len(row) > name_idx:
            name = str(row[name_idx] or "").strip()
        db[code] = {"qty": qty, "price": price, "name": name}

    wb.close()
    in_stock = sum(1 for v in db.values() if v["qty"] > 0)
    log.info(f"Mikado: прайс — {len(db)} позиций, в наличии: {in_stock}")
    return db


# ─── Mikado: авторизация + корзина ────────────────────────────────────────────
def mikado_login(code: str, password: str) -> "requests.Session | None":
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"
    })
    try:
        r = session.post(
            MIKADO_LOGIN_URL,
            data={"CODE": code, "PASSWORD": password, "INSERT": "OK"},
            timeout=20,
        )
        r.raise_for_status()
        html = r.content.decode("windows-1251", errors="replace")
        if any(kw in html for kw in ("Обслуживание клиентов", "Продолжить", "выход")):
            log.info("Mikado: авторизация успешна")
            return session
        log.error("Mikado: авторизация не прошла")
        return None
    except Exception as e:
        log.error(f"Mikado: ошибка авторизации: {e}")
        return None


def mikado_search_and_order(
    session: "requests.Session",
    code: str,
    qty: int,
    dry_run: bool = False,
) -> dict:
    """
    Полный цикл заказа одной позиции у Микадо:
      1. POST galleyp.asp CODE=<код> → редирект на SearchCodeG.asp
      2. Берём первую ссылку на карточку товара
      3. Проверяем наличие Волгоград
      4. POST pp0.asp?MODE=AddOrd — оформление заказа

    Возвращает:
      {"ok": bool, "volgograd_qty": int, "ordered": int, "message": str}
    """
    import re, time
    from urllib.parse import urljoin

    office_base = "https://mikado-parts.ru/office/"

    try:
        r = session.post(
            MIKADO_SEARCH_URL,
            data={"CODE": code, "INSERT": ""},
            timeout=20,
        )
        r.raise_for_status()
        html = r.content.decode("windows-1251", errors="replace")
    except Exception as e:
        return {"ok": False, "volgograd_qty": 0, "ordered": 0,
                "message": f"Ошибка поиска: {e}"}

    m = re.search(r"href='(galleyp\.asp\?code=[^']+)'", html, re.IGNORECASE)
    if not m:
        return {"ok": False, "volgograd_qty": 0, "ordered": 0,
                "message": f"Товар не найден в поиске: {code}"}

    product_url = urljoin(office_base, m.group(1))

    try:
        r2 = session.get(product_url, timeout=20)
        r2.raise_for_status()
        html2 = r2.content.decode("windows-1251", errors="replace")
    except Exception as e:
        return {"ok": False, "volgograd_qty": 0, "ordered": 0,
                "message": f"Ошибка страницы товара: {e}"}

    if "Волгоград" not in html2:
        return {"ok": False, "volgograd_qty": 0, "ordered": 0,
                "message": f"Нет на складе Волгоград: {code}"}

    mq = re.search(r'name=MaxQTY[^>]+value=(\d+)', html2, re.IGNORECASE)
    max_qty = int(mq.group(1)) if mq else 0
    if max_qty == 0:
        return {"ok": False, "volgograd_qty": 0, "ordered": 0,
                "message": f"Волгоград: MaxQTY=0 для {code}"}

    order_qty = min(qty, max_qty)

    if dry_run:
        return {"ok": True, "volgograd_qty": max_qty, "ordered": order_qty,
                "message": f"[DRY-RUN] Волгоград {max_qty} шт — заказали бы {order_qty}"}

    form_data: dict = {}
    for hm in re.finditer(r'<input[^>]+>', html2, re.IGNORECASE):
        tag = hm.group(0)
        nm = re.search(r'\bname=(["\']?)(\w+)\1', tag, re.IGNORECASE)
        vl = re.search(r'\bvalue=(["\']?)([^"\'> ]*)\1', tag, re.IGNORECASE)
        if nm:
            form_data[nm.group(2)] = vl.group(2) if vl else ""

    form_data["VOLUME"] = str(order_qty)
    form_data["INSERT"] = "Заказать"
    form_data.pop("searchcode", None)

    try:
        r3 = session.post(
            MIKADO_ORDER_URL,
            params={"MODE": "AddOrd", "R": int(time.time() * 1000)},
            data=form_data,
            timeout=20,
        )
        resp = r3.content.decode("windows-1251", errors="replace")
        success = any(kw in resp.lower() for kw in
                      ("добавлен", "принят", "оформл", "подтвержд", "ok", "успеш"))
        return {
            "ok": success,
            "volgograd_qty": max_qty,
            "ordered": order_qty if success else 0,
            "message": "Заказ принят" if success else f"Ответ Mikado: {resp[:300]}",
        }
    except Exception as e:
        return {"ok": False, "volgograd_qty": max_qty, "ordered": 0,
                "message": f"Ошибка отправки заказа: {e}"}


# ─── Excel-отчёт ──────────────────────────────────────────────────────────────
def save_order_report(order_id: str, items: list[dict], price_db: dict) -> Path:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Заказ Микадо"

    H_FILL  = PatternFill("solid", fgColor="1A1A2E")
    H_FONT  = Font(bold=True, color="FFFFFF")
    OK_FILL = PatternFill("solid", fgColor="C6EFCE")
    WN_FILL = PatternFill("solid", fgColor="FFEB9C")
    NO_FILL = PatternFill("solid", fgColor="FFC7CE")

    for col, h in enumerate(
        ["Код Mikado", "Название (Ozon)", "Название (Mikado)",
         "Нужно, шт.", "Наличие Mikado", "Цена закупки", "Статус"], 1
    ):
        c = ws.cell(1, col, h)
        c.font      = H_FONT
        c.fill      = H_FILL
        c.alignment = Alignment(horizontal="center")

    for ri, item in enumerate(items, 2):
        code  = item["mikado_code"]
        info  = price_db.get(code, {})
        avail = info.get("qty", 0)
        need  = item["quantity"]
        if avail >= need:
            status, fill = "✓ В наличии", OK_FILL
        elif avail > 0:
            status, fill = f"⚠ Мало ({avail})", WN_FILL
        else:
            status, fill = "✗ Нет", NO_FILL

        vals = [code, item["name"], info.get("name", ""),
                need, avail, info.get("price", 0), status]
        for col, val in enumerate(vals, 1):
            c = ws.cell(ri, col, val)
            if col == 7:
                c.fill = fill

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["G"].width = 16
    ws.freeze_panes = "A2"

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = ORDERS_DIR / f"order_{order_id}_{ts}.xlsx"
    wb.save(path)
    return path


# ─── Обработка одного заказа ─────────────────────────────────────────────────
def process_order(
    posting:        dict,
    price_db:       dict,
    mikado_session: "requests.Session | None",
    env:            dict,
    dry_run:        bool,
) -> bool:
    order_id = posting.get("posting_number", "—")
    items    = parse_ozon_items(posting)

    if not items:
        log.warning(f"[{order_id}] Заказ пустой — пропускаем")
        return True

    for it in items:
        info = price_db.get(it["mikado_code"], {})
        it["mikado_qty"]   = info.get("qty", 0)
        it["mikado_price"] = info.get("price", 0)

    in_stock  = [it for it in items if it["mikado_qty"] >= it["quantity"]]
    low_stock = [it for it in items if 0 < it["mikado_qty"] < it["quantity"]]
    no_stock  = [it for it in items if it["mikado_qty"] == 0]

    log.info(
        f"[{order_id}] позиций={len(items)}  "
        f"✓{len(in_stock)}  ⚠{len(low_stock)}  ✗{len(no_stock)}"
    )
    for it in items:
        sign = "✓" if it["mikado_qty"] >= it["quantity"] else ("⚠" if it["mikado_qty"] > 0 else "✗")
        log.info(
            f"  {sign} {it['mikado_code']:<16} ×{it['quantity']}  "
            f"(склад: {it['mikado_qty']} шт, {it['mikado_price']} руб)  "
            f"{it['name'][:45]}"
        )

    tg_tok = env.get("TG_BOT_TOKEN", "")
    tg_cid = env.get("TG_CHAT_ID", "")
    if _TG_OK and tg_tok:
        try:
            tg_order(tg_tok, tg_cid, order_id, items)
        except Exception as e:
            log.warning(f"[{order_id}] Telegram уведомление не отправлено: {e}")

    failed_items: list[dict] = []

    if not mikado_session:
        log.warning(f"[{order_id}] Сессия Mikado недоступна — только отчёт")
        failed_items = list(items)
    else:
        for it in items:
            result = mikado_search_and_order(
                mikado_session, it["mikado_code"], it["quantity"], dry_run
            )
            vol = result["volgograd_qty"]
            msg = result["message"]
            if result["ok"]:
                log.info(
                    f"  ✓ {it['mikado_code']:<16} ×{result['ordered']}  "
                    f"(Волгоград: {vol} шт)  {msg}"
                )
                it["ordered_qty"] = result["ordered"]
                if result["ordered"] < it["quantity"]:
                    failed_items.append({**it, "_reason": f"Частично: заказано {result['ordered']} из {it['quantity']}"})
            else:
                log.warning(f"  ✗ {it['mikado_code']:<16}  {msg}")
                failed_items.append({**it, "_reason": msg})

    problem_items = failed_items + low_stock + no_stock
    if problem_items and _TG_OK and tg_tok:
        try:
            tg_mikado_error(tg_tok, tg_cid, order_id, problem_items)
        except Exception as e:
            log.warning(f"[{order_id}] Telegram алерт не отправлен: {e}")
    for it in no_stock + low_stock:
        log.warning(
            f"[{order_id}] ! {it['mikado_code']} — нужно ×{it['quantity']}, "
            f"есть {it['mikado_qty']} шт."
        )

    try:
        report = save_order_report(order_id, items, price_db)
        log.info(f"[{order_id}] Отчёт: {report.name}")
    except Exception as e:
        log.warning(f"[{order_id}] Отчёт не сохранён: {e}")

    return True


# ─── Один цикл ────────────────────────────────────────────────────────────────
def sync_once(env: dict, dry_run: bool = False) -> None:
    log.info("─" * 55)
    log.info(f"Опрос заказов Ozon {'[DRY-RUN] ' if dry_run else ''}")

    client_id = env.get("OZON_CLIENT_ID", "")
    api_key   = env.get("OZON_API_KEY", "")
    if not client_id or not api_key:
        log.error("OZON_CLIENT_ID / OZON_API_KEY не заданы — выход")
        return

    state     = load_state()
    processed = set(state.get("processed", []))

    # 1. Заказы из Ozon
    postings = get_ozon_new_orders(client_id, api_key)
    if not postings:
        log.info("Новых заказов нет")
        return

    # 2. Фильтр уже обработанных
    new_postings = [p for p in postings if p.get("posting_number") not in processed]
    log.info(f"Новых необработанных: {len(new_postings)} из {len(postings)}")
    if not new_postings:
        return

    # 3. Прайс Mikado
    price_db = load_mikado_price()

    # 4. Сессия Mikado (один раз для всех заказов)
    mikado_session = None
    if not dry_run and env.get("MIKADO_CODE") and env.get("MIKADO_PASSWORD"):
        mikado_session = mikado_login(env["MIKADO_CODE"], env["MIKADO_PASSWORD"])

    # 5. Обработка
    for posting in new_postings:
        pnum = posting.get("posting_number", "")
        try:
            done = process_order(posting, price_db, mikado_session, env, dry_run)
            if done:
                processed.add(pnum)
        except Exception:
            log.exception(f"[{pnum}] Необработанная ошибка")

    # 6. Сохранить state (обрезаем до MAX_PROCESSED_RECORDS)
    if not dry_run:
        state["processed"] = sorted(processed)[-MAX_PROCESSED_RECORDS:]
        save_state(state)

    log.info("Цикл завершён")


# ─── Точка входа ──────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(description="Синхронизация заказов Ozon → Mikado")
    parser.add_argument("--once",    action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    single_instance(__file__)
    env = load_env()

    if args.once or args.dry_run:
        sync_once(env, dry_run=args.dry_run)
        return

    log.info(f"Планировщик: опрос каждые {POLL_INTERVAL_MINUTES} мин")
    while True:
        try:
            sync_once(env)
        except Exception:
            log.exception("Необработанная ошибка в цикле")
        log.info(f"Следующий опрос через {POLL_INTERVAL_MINUTES} мин")
        time.sleep(POLL_INTERVAL_MINUTES * 60)


if __name__ == "__main__":
    main()
