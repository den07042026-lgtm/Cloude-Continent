"""
wb_order_sync.py
════════════════════════════════════════════════════════════════════════════
Мониторит новые FBS-заказы WB → заказывает у Mikado → Telegram.

Поток:
  WB FBS → [этот скрипт] → Mikado + Telegram

Логика (каждые POLL_INTERVAL_MINUTES минут):
  1. GET /api/v3/orders — новые заказы (cursor-пагинация)
  2. Для каждого нового заказа:
     a. order.article = vendorCode = Mikado Code (соглашение при загрузке WB)
     b. Проверить наличие по актуальному прайсу Mikado
     c. Авторизоваться на mikado-parts.ru → заказать позиции
     d. Telegram: заказ + результат
     e. Создать поставку WB и добавить заказы (FBS-флоу)
  3. Сохранить обработанные ID и timestamp

Важно — создание поставки WB:
  После заказа у Mikado нужно вручную передать заказы в работу в ЛК WB
  (Поставки → Новая поставка → добавить заказы → получить стикеры → привезти).
  Скрипт создаёт поставку автоматически и добавляет заказы — но наклеить
  стикеры и привезти нужно вручную.

Переменные .env:
  WB_API_KEY=...
  MIKADO_CODE=...
  MIKADO_PASSWORD=...
  TG_BOT_TOKEN=...
  TG_CHAT_ID=...

Запуск:
  uv run --with requests,openpyxl scripts/wb_order_sync.py
  uv run --with requests,openpyxl scripts/wb_order_sync.py --once [--dry-run]
"""

import sys
import io
import re
import json
import time
import logging
import argparse
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Установи зависимости: uv run --with requests,openpyxl scripts/wb_order_sync.py")
    sys.exit(1)

try:
    from telegram_notify import tg_alert
    _TG_OK = True
except ImportError:
    _TG_OK = False

# ─── Константы ────────────────────────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent.parent
ENV_FILE       = BASE_DIR / ".env"
STATE_FILE     = BASE_DIR / "data" / "wb_order_state.json"
LOG_FILE       = BASE_DIR / "logs" / "wb_order_sync.log"
ORDERS_DIR     = BASE_DIR / "data" / "orders_wb"
PRICE_FALLBACK = Path("C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx")

MIKADO_PRICE_URL  = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=BBE2E029-54CF-4D9E-9FAC-9FE25E85B300"
)
MIKADO_LOGIN_URL  = "https://mikado-parts.ru/office/SECURE.asp"
MIKADO_SEARCH_URL = "https://mikado-parts.ru/office/galleyp.asp"
MIKADO_ORDER_URL  = "https://mikado-parts.ru/office/pp0.asp"

WB_BASE            = "https://marketplace-api.wildberries.ru"
POLL_INTERVAL_MIN  = 15

# ─── Логирование ──────────────────────────────────────────────────────────────
for d in (LOG_FILE.parent, ORDERS_DIR, STATE_FILE.parent):
    d.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ─── .env ─────────────────────────────────────────────────────────────────────
def load_env() -> dict:
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


def _wb_headers(token: str) -> dict:
    return {"Authorization": token, "Content-Type": "application/json"}


# ─── State ────────────────────────────────────────────────────────────────────
def load_state() -> dict:
    if not STATE_FILE.exists():
        return {"processed": [], "last_ts": 0}
    try:
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {"processed": [], "last_ts": 0}


def save_state(state: dict) -> None:
    try:
        STATE_FILE.write_text(
            json.dumps(state, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as e:
        log.error(f"Ошибка сохранения state: {e}")


# ─── WB: получение заказов ────────────────────────────────────────────────────
def get_wb_orders(token: str, since_ts: int = 0) -> list[dict]:
    """
    Возвращает все новые WB FBS-заказы начиная с since_ts (Unix timestamp).
    Использует cursor-пагинацию (/api/v3/orders?limit=1000&next=N).
    """
    orders: list[dict] = []
    next_cursor = 0
    first_page  = True

    while True:
        params: dict = {"limit": 1000, "next": next_cursor}
        if first_page and since_ts:
            params["dateFrom"] = since_ts

        try:
            r = requests.get(
                f"{WB_BASE}/api/v3/orders",
                headers=_wb_headers(token),
                params=params,
                timeout=20,
            )
            r.raise_for_status()
            data   = r.json()
            batch  = data.get("orders", [])
            orders.extend(batch)
            next_cursor = data.get("next", 0)
            first_page  = False

            if not batch or not next_cursor:
                break
        except Exception as e:
            log.error(f"WB: ошибка получения заказов: {e}")
            break

    # Только новые (waiting = ещё не собраны)
    new = [o for o in orders if o.get("wbStatus") in ("waiting", "")]
    log.info(f"WB: заказов получено {len(orders)}, новых: {len(new)}")
    return new


# ─── WB: создать поставку и добавить заказы ──────────────────────────────────
def create_wb_supply(token: str, order_ids: list[int], dry_run: bool = False) -> str | None:
    """
    Создаёт поставку WB и добавляет к ней заказы.
    Возвращает supply_id или None при ошибке.
    """
    if dry_run:
        log.info(f"[DRY-RUN] WB: создали бы поставку для {len(order_ids)} заказов")
        return "DRY-RUN-SUPPLY"

    supply_name = f"Поставка WB {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    try:
        r = requests.post(
            f"{WB_BASE}/api/v3/supplies",
            headers=_wb_headers(token),
            json={"name": supply_name},
            timeout=15,
        )
        r.raise_for_status()
        supply_id = r.json().get("id", "")
        log.info(f"WB: создана поставка {supply_id} — '{supply_name}'")
    except Exception as e:
        log.error(f"WB: ошибка создания поставки: {e}")
        return None

    # Добавляем заказы по одному
    added = 0
    for oid in order_ids:
        try:
            r = requests.patch(
                f"{WB_BASE}/api/v3/supplies/{supply_id}/orders/{oid}",
                headers=_wb_headers(token),
                timeout=10,
            )
            r.raise_for_status()
            added += 1
        except Exception as e:
            log.warning(f"WB: не удалось добавить заказ {oid} в поставку: {e}")

    log.info(f"WB: в поставку {supply_id} добавлено {added}/{len(order_ids)} заказов")
    return supply_id


# ─── Mikado: прайс ────────────────────────────────────────────────────────────
def load_mikado_price() -> dict[str, dict]:
    """Возвращает {code: {qty, price, name}}."""
    content = None
    try:
        resp = requests.get(MIKADO_PRICE_URL, timeout=60)
        resp.raise_for_status()
        if resp.content[:2] == b"PK":
            content = resp.content
    except Exception as e:
        log.warning(f"Mikado прайс онлайн: {e}, берём локальный")

    if content:
        wb_file = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    elif PRICE_FALLBACK.exists():
        wb_file = openpyxl.load_workbook(PRICE_FALLBACK, read_only=True, data_only=True)
    else:
        log.error("Mikado: прайс недоступен")
        return {}

    ws     = wb_file.active
    rows   = ws.iter_rows(values_only=True)
    header = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]

    code_idx = qty_idx = price_idx = name_idx = None
    for i, h in enumerate(header):
        if h == "code":       code_idx  = i
        elif h == "qty":      qty_idx   = i
        elif h == "priceout": price_idx = i
        elif h == "prodname": name_idx  = i

    if code_idx is None:
        wb_file.close()
        return {}

    db: dict[str, dict] = {}
    for row in rows:
        raw = row[code_idx] if len(row) > code_idx else None
        if not raw:
            continue
        code = str(raw).strip()
        qty  = 0; price = 0.0; name = ""
        if qty_idx   is not None and len(row) > qty_idx:
            try: qty   = max(0, int(float(str(row[qty_idx] or 0))))
            except: pass
        if price_idx is not None and len(row) > price_idx:
            try: price = float(str(row[price_idx] or 0))
            except: pass
        if name_idx  is not None and len(row) > name_idx:
            name = str(row[name_idx] or "").strip()
        db[code] = {"qty": qty, "price": price, "name": name}

    wb_file.close()
    log.info(f"Mikado: {len(db):,} позиций, в наличии: {sum(1 for v in db.values() if v['qty'] > 0):,}")
    return db


# ─── Mikado: авторизация + заказ ─────────────────────────────────────────────
def mikado_login(code: str, password: str) -> "requests.Session | None":
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"})
    try:
        r = session.post(
            MIKADO_LOGIN_URL,
            data={"CODE": code, "PASSWORD": password, "INSERT": "OK"},
            timeout=20,
        )
        r.raise_for_status()
        html = r.content.decode("windows-1251", errors="replace")
        if any(kw in html for kw in ("Обслуживание клиентов", "Продолжить", "выход")):
            log.info("Mikado: авторизация успешна")
            return session
        log.error("Mikado: авторизация не прошла")
        return None
    except Exception as e:
        log.error(f"Mikado: ошибка авторизации: {e}")
        return None


def mikado_search_and_order(
    session: "requests.Session",
    code:    str,
    qty:     int,
    dry_run: bool = False,
) -> dict:
    """Поиск + заказ одной позиции у Mikado. Аналог из ozon_order_sync."""
    from urllib.parse import urljoin
    office_base = "https://mikado-parts.ru/office/"

    try:
        r = session.post(MIKADO_SEARCH_URL, data={"CODE": code, "INSERT": ""}, timeout=20)
        r.raise_for_status()
        html = r.content.decode("windows-1251", errors="replace")
    except Exception as e:
        return {"ok": False, "volgograd_qty": 0, "ordered": 0, "message": f"Ошибка поиска: {e}"}

    m = re.search(r"href='(galleyp\.asp\?code=[^']+)'", html, re.IGNORECASE)
    if not m:
        return {"ok": False, "volgograd_qty": 0, "ordered": 0, "message": f"Не найден: {code}"}

    try:
        r2 = session.get(urljoin(office_base, m.group(1)), timeout=20)
        r2.raise_for_status()
        html2 = r2.content.decode("windows-1251", errors="replace")
    except Exception as e:
        return {"ok": False, "volgograd_qty": 0, "ordered": 0, "message": f"Ошибка карточки: {e}"}

    if "Волгоград" not in html2:
        return {"ok": False, "volgograd_qty": 0, "ordered": 0, "message": f"Нет на складе Волгоград: {code}"}

    mq = re.search(r'name=MaxQTY[^>]+value=(\d+)', html2, re.IGNORECASE)
    max_qty = int(mq.group(1)) if mq else 0
    if max_qty == 0:
        return {"ok": False, "volgograd_qty": 0, "ordered": 0, "message": f"MaxQTY=0: {code}"}

    order_qty = min(qty, max_qty)

    if dry_run:
        return {"ok": True, "volgograd_qty": max_qty, "ordered": order_qty,
                "message": f"[DRY-RUN] Волгоград {max_qty} шт"}

    form_data: dict = {}
    for hm in re.finditer(r'<input[^>]+>', html2, re.IGNORECASE):
        tag = hm.group(0)
        nm  = re.search(r'\bname=(["\']?)(\w+)\1', tag, re.IGNORECASE)
        vl  = re.search(r'\bvalue=(["\']?)([^"\'> ]*)\1', tag, re.IGNORECASE)
        if nm:
            form_data[nm.group(2)] = vl.group(2) if vl else ""

    form_data["VOLUME"] = str(order_qty)
    form_data["INSERT"] = "Заказать"
    form_data.pop("searchcode", None)

    try:
        r3   = session.post(MIKADO_ORDER_URL, params={"MODE": "AddOrd", "R": int(time.time() * 1000)},
                            data=form_data, timeout=20)
        resp = r3.content.decode("windows-1251", errors="replace")
        ok   = any(kw in resp.lower() for kw in ("добавлен", "принят", "оформл", "подтвержд", "ok", "успеш"))
        return {"ok": ok, "volgograd_qty": max_qty, "ordered": order_qty if ok else 0,
                "message": "Заказ принят" if ok else f"Ответ: {resp[:300]}"}
    except Exception as e:
        return {"ok": False, "volgograd_qty": max_qty, "ordered": 0, "message": f"Ошибка POST: {e}"}


# ─── Excel-отчёт ──────────────────────────────────────────────────────────────
def save_order_report(wb_order_id: int, items: list[dict], price_db: dict) -> Path:
    wb_out = openpyxl.Workbook()
    ws     = wb_out.active
    ws.title = "Заказ WB"

    H_FILL  = PatternFill("solid", fgColor="1A1A2E")
    H_FONT  = Font(bold=True, color="FFFFFF")
    OK_FILL = PatternFill("solid", fgColor="C6EFCE")
    WN_FILL = PatternFill("solid", fgColor="FFEB9C")
    NO_FILL = PatternFill("solid", fgColor="FFC7CE")

    for col, h in enumerate(
        ["WB Order ID", "Артикул (Mikado)", "Название", "Нужно, шт.", "Наличие Mikado", "Цена закупки", "Статус"], 1
    ):
        c       = ws.cell(1, col, h)
        c.font  = H_FONT
        c.fill  = H_FILL
        c.alignment = Alignment(horizontal="center")

    for ri, item in enumerate(items, 2):
        code  = item["mikado_code"]
        info  = price_db.get(code, {})
        avail = info.get("qty", 0)
        need  = item["quantity"]
        if avail >= need:   status, fill = "✓ В наличии", OK_FILL
        elif avail > 0:     status, fill = f"⚠ Мало ({avail})", WN_FILL
        else:               status, fill = "✗ Нет", NO_FILL

        vals = [wb_order_id, code, info.get("name", item.get("name", "")),
                need, avail, info.get("price", 0), status]
        for col, val in enumerate(vals, 1):
            c = ws.cell(ri, col, val)
            if col == 7:
                c.fill = fill

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["G"].width = 18
    ws.freeze_panes = "A2"

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = ORDERS_DIR / f"wb_order_{wb_order_id}_{ts}.xlsx"
    wb_out.save(path)
    return path


# ─── Обработка одного заказа ─────────────────────────────────────────────────
def process_order(
    order:          dict,
    price_db:       dict,
    mikado_session: "requests.Session | None",
    env:            dict,
    dry_run:        bool,
) -> bool:
    order_id  = order.get("id")
    article   = order.get("article", "").strip()   # vendorCode = Mikado code
    nm_id     = order.get("nmId", "")
    qty       = order.get("quantity", 1) or 1
    price_rub = order.get("convertedPrice", 0) / 100 if order.get("convertedPrice") else 0

    if not article:
        log.warning(f"[WB #{order_id}] Нет артикула — пропускаем")
        return True

    items = [{"mikado_code": article, "name": article, "quantity": qty, "price_rub": price_rub, "nm_id": nm_id}]

    # Обогащаем данными из прайса
    for it in items:
        info = price_db.get(it["mikado_code"], {})
        it["mikado_qty"]   = info.get("qty", 0)
        it["mikado_price"] = info.get("price", 0)
        it["name"]         = info.get("name", it["name"])

    log.info(
        f"[WB #{order_id}] article={article} nmId={nm_id} "
        f"qty={qty} mikado_qty={items[0]['mikado_qty']}"
    )

    tg_tok = env.get("TG_BOT_TOKEN", "")
    tg_cid = env.get("TG_CHAT_ID", "")

    # Mikado: заказ
    order_result = None
    if mikado_session:
        order_result = mikado_search_and_order(mikado_session, article, qty, dry_run)
        ok  = order_result["ok"]
        msg = order_result["message"]
        sign = "✓" if ok else "✗"
        log.info(f"  {sign} Mikado {article} ×{qty}: {msg}")
    else:
        log.warning(f"[WB #{order_id}] Mikado недоступен — только отчёт")

    # Excel отчёт
    try:
        report = save_order_report(order_id, items, price_db)
        log.info(f"[WB #{order_id}] Отчёт: {report.name}")
    except Exception as e:
        log.warning(f"[WB #{order_id}] Отчёт не сохранён: {e}")

    # Telegram
    if _TG_OK and tg_tok:
        mikado_status = ""
        if order_result:
            mikado_status = "✓ заказан у Mikado" if order_result["ok"] else f"⚠ {order_result['message']}"
        else:
            mikado_status = "⚠ Mikado недоступен"

        avail = items[0]["mikado_qty"]
        stock_icon = "✓" if avail >= qty else ("⚠" if avail > 0 else "✗")

        text = (
            f"📦 Новый заказ WB #{order_id}\n"
            f"  Артикул: {article}\n"
            f"  Название: {items[0]['name'][:60]}\n"
            f"  Количество: {qty} шт\n"
            f"  Цена продажи: {price_rub:.0f} ₽\n"
            f"  Наличие Mikado: {stock_icon} {avail} шт\n"
            f"  Закупка: {items[0]['mikado_price']:.0f} ₽\n"
            f"  Mikado: {mikado_status}\n"
            f"  → Создай поставку в ЛК WB и добавь заказ #{order_id}"
        )
        try:
            tg_alert(tg_tok, tg_cid, text)
        except Exception as e:
            log.warning(f"[WB #{order_id}] Telegram не отправлен: {e}")

    return True


# ─── Один цикл ────────────────────────────────────────────────────────────────
def sync_once(env: dict, dry_run: bool = False) -> None:
    log.info("─" * 55)
    log.info(f"Опрос заказов WB {'[DRY-RUN] ' if dry_run else ''}")

    token = env.get("WB_API_KEY", "")
    if not token:
        log.error("WB_API_KEY не задан в .env")
        return

    state     = load_state()
    processed = set(state.get("processed", []))
    since_ts  = state.get("last_ts", 0)

    # 1. Получаем заказы
    orders = get_wb_orders(token, since_ts)
    if not orders:
        log.info("Новых заказов нет")
        return

    # 2. Фильтруем обработанные
    new_orders = [o for o in orders if o.get("id") not in processed]
    log.info(f"Необработанных: {len(new_orders)} из {len(orders)}")
    if not new_orders:
        return

    # 3. Прайс Mikado
    price_db = load_mikado_price()

    # 4. Сессия Mikado
    mikado_session = None
    if not dry_run and env.get("MIKADO_CODE") and env.get("MIKADO_PASSWORD"):
        mikado_session = mikado_login(env["MIKADO_CODE"], env["MIKADO_PASSWORD"])

    # 5. Обработка заказов
    processed_ids: list[int] = []
    for order in new_orders:
        oid = order.get("id")
        try:
            done = process_order(order, price_db, mikado_session, env, dry_run)
            if done:
                processed_ids.append(oid)
                processed.add(oid)
        except Exception:
            log.exception(f"[WB #{oid}] Необработанная ошибка")

    # 6. Создать поставку WB для всех новых заказов
    if processed_ids and not dry_run:
        supply_id = create_wb_supply(token, processed_ids, dry_run)
        if supply_id:
            log.info(f"WB: поставка {supply_id} готова — наклей стикеры и привези в ПВЗ")
            tg_tok = env.get("TG_BOT_TOKEN", "")
            tg_cid = env.get("TG_CHAT_ID", "")
            if _TG_OK and tg_tok:
                tg_alert(tg_tok, tg_cid,
                         f"WB поставка {supply_id} создана ({len(processed_ids)} заказов)\n"
                         f"Наклей стикеры в ЛК WB → Поставки → {supply_id}")

    # 7. Сохранить state
    if not dry_run:
        state["processed"] = sorted(str(p) for p in processed)
        state["last_ts"]   = int(datetime.now(timezone.utc).timestamp())
        save_state(state)

    log.info("Цикл завершён")


# ─── Точка входа ──────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(description="Синхронизация заказов WB → Mikado")
    parser.add_argument("--once",    action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    env = load_env()

    if args.once or args.dry_run:
        sync_once(env, dry_run=args.dry_run)
        return

    log.info(f"Планировщик: опрос каждые {POLL_INTERVAL_MIN} мин")
    while True:
        try:
            sync_once(env)
        except Exception:
            log.exception("Необработанная ошибка в цикле")
        log.info(f"Следующий опрос через {POLL_INTERVAL_MIN} мин")
        time.sleep(POLL_INTERVAL_MIN * 60)


if __name__ == "__main__":
    main()
