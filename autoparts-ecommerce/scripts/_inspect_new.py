import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

EXCEL = r'C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ_new.xlsx'
wb = openpyxl.load_workbook(EXCEL)
ws = wb.worksheets[0]
print(f'Строк: {ws.max_row}, Колонок: {ws.max_column}')
print()
for c in range(1, ws.max_column + 1):
    v  = ws.cell(1, c).value
    ex = ws.cell(2, c).value
    if v:
        print(f'  col{c:2d}: {str(v)[:30]:30s}  ex: {str(ex or "")[:50]}')

print()
total = ws.max_row - 1
oem_good  = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r,7).value and str(ws.cell(r,7).value).strip() not in ('','—'))
oem_dash  = sum(1 for r in range(2, ws.max_row+1) if str(ws.cell(r,7).value or '').strip() == '—')
oem_empty = sum(1 for r in range(2, ws.max_row+1) if not str(ws.cell(r,7).value or '').strip())
alt_good  = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r,9).value and str(ws.cell(r,9).value).strip() not in ('','—'))
alt_dash  = sum(1 for r in range(2, ws.max_row+1) if str(ws.cell(r,9).value or '').strip() == '—')
alt_empty = sum(1 for r in range(2, ws.max_row+1) if not str(ws.cell(r,9).value or '').strip())
print(f'Всего строк данных: {total}')
print(f'OEM (col7):  заполнено={oem_good}, прочерк={oem_dash}, пусто={oem_empty}')
print(f'Alts (col9): заполнено={alt_good}, прочерк={alt_dash}, пусто={alt_empty}')
