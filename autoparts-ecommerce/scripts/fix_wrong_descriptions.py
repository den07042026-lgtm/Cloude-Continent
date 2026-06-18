"""
fix_wrong_descriptions.py
═════════════════════════
Очищает ячейки «Описание», в которые попал промпт вместо готового текста.
Признак плохой ячейки: содержимое начинается с «Ты копирайтер».

Запуск:
  uv run --with openpyxl scripts/fix_wrong_descriptions.py
"""

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

import openpyxl

DST_DIR      = Path(r"C:\Users\Admin\Desktop\На сортировку 26.04")
COL_DESC     = "Описание"
BAD_PREFIX   = "Ты копирайтер"

total_cleared = 0

for xlsx in sorted(DST_DIR.glob("*.xlsx")):
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.active

    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col = headers.get(COL_DESC)
    if col is None:
        continue

    cleared = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, col).value
        if val and str(val).strip().startswith(BAD_PREFIX):
            ws.cell(row, col).value = None
            cleared += 1

    if cleared:
        wb.save(xlsx)
        print(f"  {xlsx.name}: очищено {cleared} ячеек")
        total_cleared += cleared

print(f"\nИтого очищено: {total_cleared} ячеек")
print("Теперь запусти основной скрипт заново.")
