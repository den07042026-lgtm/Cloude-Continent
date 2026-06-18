"""
wb_niche_analyzer.py
════════════════════════════════════════════════════════════════════════════
Анализ ниш WB через MPStats — ОДИН запрос вместо 500+.

Алгоритм:
  1. GET /wb/get/category/items?path=Автотовары → 2600+ предметов с OOS/продажами
  2. Фильтруем авто-предметы по ключевым словам
  3. Для каждого предмета keyword-match по нашему каталогу (Mikado + Автолига)
  4. Считаем маржу WB FBS на наших ценах vs средняя цена WB
  5. Скор = OOS% × ln(продажи+1) × max(0, маржа/12%)
  6. Экспорт топ-300 в Excel

Запуск:
    uv run --with openpyxl --with requests --with python-dotenv --with xlrd \\
           scripts/wb_niche_analyzer.py
    python scripts/wb_niche_analyzer.py [--days 30] [--min-sales 100] [--top 300]
"""

import argparse
import json
import logging
import math
import os
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

try:
    import requests
except ImportError:
    print("pip install requests"); sys.exit(1)

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(*a, **kw): pass

sys.stdout.reconfigure(encoding="utf-8")

# ─── Пути ────────────────────────────────────────────────────────────────────

BASE_DIR  = Path(__file__).parent.parent
load_dotenv(BASE_DIR / ".env")

MPSTATS_TOKEN = os.getenv("MPSTATS_TOKEN", "")
MIKADO_PRICE  = Path("C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx")
CACHE_DIR     = BASE_DIR / "data" / "analytics" / "cache"
OUT_DIR       = BASE_DIR / "data" / "analytics"
LOG_FILE      = BASE_DIR / "logs" / "wb_niche_analyzer.log"

# ─── WB FBS тарифы (Волгоград, май 2026) ─────────────────────────────────────

WB_ACQ_PCT       = 0.015   # эквайринг
WB_TAX_PCT       = 0.06    # УСН 6%
WB_SPP_RESERVE   = 0.07    # резерв на СПП
WB_RET_RATE      = 0.03    # доля возвратов
WB_PACKAGING     = 30.0    # упаковка, ₽
WB_DELIVERY_BASE = 50.6    # базовая ставка доставки FBS, ₽
WB_DELIVERY_L    = 15.4    # доп. за литр, ₽/л
WB_RETURN_BASE   = 136.0   # базовая ставка возврата, ₽
WB_RETURN_L      = 14.0    # доп. за литр при возврате, ₽/л
DEFAULT_VOLUME_L = 3.0     # объём по умолчанию (автозапчасть ~3 л)
DEFAULT_COMMISSION = 0.30  # fallback комиссия, если в предмете не указана
TARGET_MARGIN    = 0.12    # целевая маржа 12%

# ─── Стоп-слова и авто-фильтры ────────────────────────────────────────────────

# Предметы с этими словами НЕ являются автозапчастями
STOP_SUBJECTS = (
    "мотоцикл", "мотозапчаст", "мотото", "велосипед", "самокат",
    "скутер", "квадроцикл", "снегоход",
    "садовая техника", "мойки высокого давления", "запчасти для садов",
    "пылесос", "стиральн", "холодильник", "чайник",
    "ручной инструмент", "шуруповерт", "перфоратор", "дрель",
    "электроинструмент", "строительн", "кровельн",
    "канцеляр", "офис", "школьн",
    "косметик", "парфюм", "бытов химия",
)

# Предметы с этими словами ТОЧНО относятся к нашему сегменту
AUTO_PARTS_KW = (
    "автозапчаст",
    "запчаст", "колодк", "амортизатор", "тормоз",
    "подвеск", "рулев", "сцеплен", "ремен",
    "шаровая", "шаровой", "стойк стабил",
    "свеч", "зажигани", "фильтр масл", "фильтр воздуш",
    "фильтр салон", "фильтр топлив",
    "подшипник ступ", "ступиц",
    "сальник", "прокладк", "крепеж для авт",
    "зеркала автомоб", "зеркало автомоб",
    "фары автомоб", "фонар автомоб",
    "датчик", "катушк зажиган",
    "насос топлив", "бензонасос",
    "цилиндр автомоб", "реле автомоб",
    "тросов", "привод шруса", "шрус",
    "масла моторн", "масла трансмисс", "антифриз", "тосол",
    "разъем автомоб",
    "ремкомплект",
    "уплотнитель автомоб",
    "предохранител автомоб",
    "аккумулятор",
)

# Более широкий список для keyword-matching (включает пограничные предметы)
AUTO_BROAD_KW = AUTO_PARTS_KW + (
    "масл", "жидкост", "охлаждени", "топлив",
    "автомоб", "автоэлектр",
    "мотор", "двигател", "коробк",
)

log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  Расчёт маржи WB FBS
# ══════════════════════════════════════════════════════════════════════════════

def calc_wb_margin(purchase: float, sell: float,
                   volume_l: float = DEFAULT_VOLUME_L,
                   commission: float = DEFAULT_COMMISSION) -> float:
    """Маржа WB FBS (0..1). commission — доля 0..1."""
    if sell <= 0 or purchase <= 0:
        return 0.0
    delivery    = WB_DELIVERY_BASE + WB_DELIVERY_L * volume_l
    return_cost = (WB_RETURN_BASE + WB_RETURN_L * volume_l) * WB_RET_RATE
    com         = sell * commission
    acq         = sell * WB_ACQ_PCT
    spp         = sell * WB_SPP_RESERVE
    proceeds    = sell - com - acq - spp - delivery
    tax         = max(0.0, proceeds) * WB_TAX_PCT
    total_cost  = purchase + com + acq + spp + delivery + return_cost + WB_PACKAGING + tax
    return (sell - total_cost) / sell


def score_niche(oos_pct: float, sales: float, margin: float) -> float:
    return (oos_pct / 100.0) * math.log1p(sales) * max(0.0, margin / TARGET_MARGIN)


# ══════════════════════════════════════════════════════════════════════════════
#  MPStats
# ══════════════════════════════════════════════════════════════════════════════

def fetch_subjects(d1: str, d2: str, force: bool = False) -> list[dict]:
    """
    GET /wb/get/category/items → все предметы "Автотовары".
    Кэш на 1 сутки.
    """
    cache = CACHE_DIR / f"wb_auto_subjects_{d1}.json"

    if not force and cache.exists():
        age_hours = (time.time() - cache.stat().st_mtime) / 3600
        if age_hours < 24:
            log.info(f"Загружаем из кэша: {cache.name}")
            return json.loads(cache.read_text(encoding="utf-8"))

    log.info("MPStats: загружаем предметы Автотовары...")
    url = "https://mpstats.io/api/wb/get/category/items"
    params = {"path": "Автотовары", "d1": d1, "d2": d2, "startRow": 0, "endRow": 5000}
    headers = {"X-Mpstats-TOKEN": MPSTATS_TOKEN}

    for attempt in range(3):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=60)
            if r.status_code == 429:
                wait = 60 * (attempt + 1)
                log.warning(f"Rate-limit MPStats, ждём {wait}с...")
                time.sleep(wait)
                continue
            if r.status_code == 401:
                log.error("MPStats: неверный токен (401)")
                sys.exit(1)
            if r.status_code != 200:
                log.error(f"MPStats: HTTP {r.status_code} — {r.text[:200]}")
                return []
            data = r.json()
            rows = data if isinstance(data, list) else (data.get("data") or data.get("items") or [])
            log.info(f"MPStats: получено {len(rows)} предметов")
            cache.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")
            return rows
        except Exception as e:
            log.warning(f"MPStats попытка {attempt+1}: {e}")
            time.sleep(5)

    return []


def is_auto_subject(name: str) -> bool:
    """True если предмет — автозапчасть/авто-товар (не мото, не инструмент)."""
    low = name.lower()
    if any(kw in low for kw in STOP_SUBJECTS):
        return False
    return any(kw in low for kw in AUTO_BROAD_KW)


# ══════════════════════════════════════════════════════════════════════════════
#  Загрузка каталогов
# ══════════════════════════════════════════════════════════════════════════════

def load_mikado_items() -> list[dict]:
    """Возвращает [{brand, name, price}] из Mikado."""
    if not MIKADO_PRICE.exists():
        log.warning(f"Mikado: файл не найден — {MIKADO_PRICE}")
        return []
    try:
        wb = openpyxl.load_workbook(str(MIKADO_PRICE), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        log.error(f"Mikado: {e}")
        return []

    items = []
    ci_brand = ci_name = ci_price = None
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            hdr = [str(c).strip().lower() if c else "" for c in row]
            ci_brand = next((i for i, h in enumerate(hdr) if h in ("brandname", "brand")), None)
            ci_name  = next((i for i, h in enumerate(hdr) if h in ("prodname",  "name")),  None)
            ci_price = next((i for i, h in enumerate(hdr) if h in ("priceout",  "price")), None)
            if ci_brand is not None and ci_price is not None:
                header_found = True
            continue
        try:
            brand = str(row[ci_brand] or "").strip()
            name  = str(row[ci_name]  or "").strip() if ci_name is not None else ""
            price = float(row[ci_price] or 0)
        except (TypeError, ValueError, IndexError):
            continue
        if brand and price > 0:
            items.append({"brand": brand, "name": name.lower(), "price": price})

    try: wb.close()
    except Exception: pass

    log.info(f"Mikado: {len(items):,} позиций")
    return items


def load_autoliga_items() -> list[dict]:
    """Возвращает [{brand, name, price}] из Автолиги."""
    scripts_dir = Path(__file__).parent
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    try:
        from autoliga_loader import load_autoliga
    except ImportError as e:
        log.error(f"autoliga_loader: {e}")
        return []
    catalog = load_autoliga()
    items = [
        {"brand": v.get("brand", ""), "name": v.get("name", "").lower(), "price": float(v.get("price") or 0)}
        for v in catalog.values()
        if v.get("brand") and float(v.get("price") or 0) > 0
    ]
    log.info(f"Автолига: {len(items):,} позиций")
    return items


# ══════════════════════════════════════════════════════════════════════════════
#  Keyword matching: предмет WB ↔ наш каталог
# ══════════════════════════════════════════════════════════════════════════════

# Стоп-слова при разборе названия предмета
NAME_STOPWORDS = frozenset({
    "для", "авто", "автомобильные", "автомобильное", "автомобильный",
    "автомобиля", "автомобилей", "и", "в", "на", "с", "из",
    "запчасти", "запчастей", "к", "по", "деталь", "детали",
    "ремонт", "комплект", "крепеж",
})

def _subject_stems(subject_name: str) -> list[str]:
    """Ключевые стемы из второй части названия предмета (после '/')."""
    parts = subject_name.split("/")
    search_part = parts[-1].strip().lower() if len(parts) > 1 else subject_name.lower()
    words = [w.strip(".,;:()-") for w in search_part.split()]
    stems = []
    for w in words:
        if w in NAME_STOPWORDS or len(w) < 4:
            continue
        stems.append(w[:6])   # первые 6 букв как стем
    return stems


def match_catalog_items(subject_name: str, catalog_items: list[dict]) -> list[dict]:
    """Возвращает товары каталога, чьё название совпадает со стемами предмета."""
    stems = _subject_stems(subject_name)
    if not stems:
        return []
    matched = []
    for item in catalog_items:
        n = item["name"]
        if any(s in n for s in stems):
            matched.append(item)
    return matched


# ══════════════════════════════════════════════════════════════════════════════
#  Excel
# ══════════════════════════════════════════════════════════════════════════════

COLUMNS = [
    ("rank",          "#",              4),
    ("subject",       "Предмет WB",    40),
    ("sales_month",   "Прод/мес",      10),
    ("oos_pct",       "OOS%",           7),
    ("avg_price",     "Ср.цена WB",    12),
    ("our_cost_med",  "Наша себест.",  13),
    ("margin_pct",    "Маржа%",         9),
    ("commission_pct","Ком.WB%",        9),
    ("sellers",       "Продавцов",     11),
    ("items_wb",      "Товаров на WB", 14),
    ("revenue",       "Выручка/мес",   13),
    ("our_sku",       "Наших SKU",     10),
    ("our_brands",    "Наши бренды",   30),
    ("score",         "Скор",          10),
]


def write_excel(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ниши WB"

    keys    = [c[0] for c in COLUMNS]
    headers = [c[1] for c in COLUMNS]
    widths  = [c[2] for c in COLUMNS]

    ws.append(headers)

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    for r in rows:
        ws.append([r.get(k) for k in keys])

    # Подсветка OOS% (столбец 4 = D)
    oos_col = 4
    for ri in range(2, len(rows) + 2):
        cell = ws.cell(row=ri, column=oos_col)
        v = float(cell.value or 0)
        if v >= 50:
            cell.fill = PatternFill("solid", fgColor="C00000")
            cell.font = Font(color="FFFFFF", bold=True)
        elif v >= 40:
            cell.fill = PatternFill("solid", fgColor="FF4444")
            cell.font = Font(color="FFFFFF", bold=True)
        elif v >= 25:
            cell.fill = PatternFill("solid", fgColor="FFA500")
        elif v >= 15:
            cell.fill = PatternFill("solid", fgColor="FFD700")

    # Подсветка маржи (столбец 7 = G)
    margin_col = 7
    for ri in range(2, len(rows) + 2):
        cell = ws.cell(row=ri, column=margin_col)
        v = float(cell.value or 0)
        if v >= 15:
            cell.fill = PatternFill("solid", fgColor="70AD47")
            cell.font = Font(color="FFFFFF")
        elif v >= 10:
            cell.fill = PatternFill("solid", fgColor="FFEB84")
        elif v < 0:
            cell.fill = PatternFill("solid", fgColor="FFCCCC")

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(str(path))


# ══════════════════════════════════════════════════════════════════════════════
#  main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="WB Niche Analyzer")
    parser.add_argument("--days",       type=int,   default=30)
    parser.add_argument("--min-sales",  type=float, default=100, help="Мин.продаж/мес для предмета")
    parser.add_argument("--min-oos",    type=float, default=0.0)
    parser.add_argument("--top",        type=int,   default=300)
    parser.add_argument("--skip-mikado",   action="store_true")
    parser.add_argument("--skip-autoliga", action="store_true")
    parser.add_argument("--no-catalog-match", action="store_true", help="Не делать keyword match")
    parser.add_argument("--refresh",    action="store_true", help="Обновить кэш MPStats")
    args = parser.parse_args()

    for d in (CACHE_DIR, OUT_DIR, LOG_FILE.parent):
        d.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
        datefmt="%H:%M:%S",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(str(LOG_FILE), encoding="utf-8", errors="replace"),
        ],
    )

    if not MPSTATS_TOKEN:
        log.error("MPSTATS_TOKEN не задан в .env"); sys.exit(1)

    d2 = datetime.now().strftime("%Y-%m-%d")
    d1 = (datetime.now() - timedelta(days=args.days)).strftime("%Y-%m-%d")
    log.info(f"Период: {d1}…{d2} | мин.продаж: {args.min_sales} | топ: {args.top}")

    # ─── 1. Предметы с WB ────────────────────────────────────────────────────
    all_subjects = fetch_subjects(d1, d2, force=args.refresh)
    if not all_subjects:
        log.error("Нет данных из MPStats"); sys.exit(1)

    auto_subjects = [s for s in all_subjects if is_auto_subject(s.get("name", ""))]
    log.info(f"Всего предметов: {len(all_subjects)}, авто-предметов: {len(auto_subjects)}")

    # ─── 2. Каталог ──────────────────────────────────────────────────────────
    catalog_items: list[dict] = []
    if not args.no_catalog_match:
        if not args.skip_mikado:
            catalog_items.extend(load_mikado_items())
        if not args.skip_autoliga:
            catalog_items.extend(load_autoliga_items())
        log.info(f"Каталог: {len(catalog_items):,} позиций суммарно")

    # ─── 3. Анализ ───────────────────────────────────────────────────────────
    results = []

    for s in auto_subjects:
        sales    = float(s.get("sales", 0) or 0)
        oos_pct  = float(s.get("lost_profit_percent", 0) or 0)
        avg_price= float(s.get("avg_price", 0) or 0)
        purchase = float(s.get("purchase", 0) or 0)    # MPStats: ср. закупочная на WB
        sellers  = int(s.get("sellers", 0) or 0)
        items_wb = int(s.get("items", 0) or 0)
        revenue  = float(s.get("revenue", 0) or 0)
        comm_pct = float(s.get("commision_fbs", DEFAULT_COMMISSION * 100) or (DEFAULT_COMMISSION * 100)) / 100.0
        name     = s.get("name", "")

        if sales < args.min_sales:
            continue
        if oos_pct < args.min_oos:
            continue
        if avg_price <= 0:
            continue

        # Keyword-match нашего каталога
        matched_items: list[dict] = []
        if catalog_items:
            matched_items = match_catalog_items(name, catalog_items)

        # Себестоимость: медиана нашего каталога → fallback MPStats purchase → 40% avg_price
        if matched_items:
            prices_sorted = sorted(i["price"] for i in matched_items)
            our_cost = prices_sorted[len(prices_sorted) // 2]
            brand_counts: dict[str, int] = {}
            for it in matched_items:
                b = it["brand"]
                brand_counts[b] = brand_counts.get(b, 0) + 1
            top_brands_str = ", ".join(
                b for b, _ in sorted(brand_counts.items(), key=lambda x: -x[1])[:5]
            )
        else:
            our_cost = purchase if purchase > 0 else avg_price * 0.40
            top_brands_str = ""

        margin = calc_wb_margin(our_cost, avg_price, commission=comm_pct)
        sc = score_niche(oos_pct, sales, margin)

        results.append({
            "subject":       name,
            "sales_month":   int(sales),
            "oos_pct":       round(oos_pct, 1),
            "avg_price":     round(avg_price),
            "our_cost_med":  round(our_cost),
            "margin_pct":    round(margin * 100, 1),
            "commission_pct":round(comm_pct * 100, 1),
            "sellers":       sellers,
            "items_wb":      items_wb,
            "revenue":       round(revenue),
            "our_sku":       len(matched_items),
            "our_brands":    top_brands_str,
            "score":         round(sc, 4),
        })

    if not results:
        log.warning("Нет результатов — попробуй уменьшить --min-sales")
        return

    results.sort(key=lambda x: x["score"], reverse=True)

    # Нумерация
    for i, r in enumerate(results, 1):
        r["rank"] = i

    top = results[:args.top]

    # ─── 4. Экспорт ──────────────────────────────────────────────────────────
    ts    = datetime.now().strftime("%Y%m%d_%H%M")
    fpath = OUT_DIR / f"wb_niches_{ts}.xlsx"
    write_excel(top, fpath)
    log.info(f"Готово! {len(top)} строк → {fpath}")

    # Топ-30 в консоль
    w = 100
    print(f"\n{'─'*w}")
    print(f"{'#':>3} {'Предмет':<38} {'OOS%':>5} {'Прод':>7} {'Цена':>7} {'Маржа':>6} {'SKU':>5} {'Скор':>8}")
    print(f"{'─'*w}")
    for r in top[:30]:
        subj_s = r["subject"][:37]
        print(
            f"{r['rank']:>3} {subj_s:<38} {r['oos_pct']:>4.0f}% "
            f"{r['sales_month']:>7,} {r['avg_price']:>6.0f}₽ "
            f"{r['margin_pct']:>5.1f}% {r['our_sku']:>5} {r['score']:>8.4f}"
        )
    print(f"{'─'*w}")
    print(f"\nВсего авто-предметов: {len(results)}  |  Топ сохранён: {fpath}\n")


if __name__ == "__main__":
    main()
