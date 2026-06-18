"""
wb_item_analyzer.py  —  Уровень 2
════════════════════════════════════════════════════════════════════════════
Для топ-ниш из wb_niche_analyzer.py ищет конкретные товары на WB
и получает per-item аналитику через MPStats.

Поток:
  Наш каталог (Mikado + Автолига) × топ-ниши Level 1
  → приоритетные OEM-артикулы
  → WB поиск по артикулу (curl_cffi, Chrome TLS — обходит TLS-фингерпринт)
  → nmID → MPStats /wb/get/item/{nmId}/by_category
  → Скор = OOS% × продажи × max(0, маржа/12%)
  → Excel топ-1000 SKU для листинга

Запуск:
    uv run --with openpyxl --with requests --with python-dotenv --with xlrd \\
           --with curl-cffi python scripts/wb_item_analyzer.py
    python scripts/wb_item_analyzer.py [--top-niches 25] [--articles 1000]
"""

import argparse
import json
import logging
import math
import os
import random
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

try:
    import requests as _requests
except ImportError:
    print("pip install requests"); sys.exit(1)

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(*a, **kw): pass

# curl_cffi — Chrome TLS fingerprint, обязателен для WB search
try:
    from curl_cffi import requests as cffi_requests
    CURL_AVAILABLE = True
except ImportError:
    CURL_AVAILABLE = False
    print("ВНИМАНИЕ: curl_cffi не установлен. WB-поиск будет использовать requests (риск блокировки).")
    print("  Установи: pip install curl-cffi")

sys.stdout.reconfigure(encoding="utf-8")

# ─── Пути ────────────────────────────────────────────────────────────────────

BASE_DIR  = Path(__file__).parent.parent
load_dotenv(BASE_DIR / ".env")

MPSTATS_TOKEN = os.getenv("MPSTATS_TOKEN", "")
MIKADO_PRICE  = Path("C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx")
CACHE_DIR     = BASE_DIR / "data" / "analytics" / "cache"
OUT_DIR       = BASE_DIR / "data" / "analytics"
LOG_FILE      = BASE_DIR / "logs" / "wb_item_analyzer.log"

# ─── Тарифы WB FBS (Волгоград, май 2026) ────────────────────────────────────

WB_ACQ_PCT       = 0.015
WB_TAX_PCT       = 0.06
WB_SPP_RESERVE   = 0.07
WB_RET_RATE      = 0.03
WB_PACKAGING     = 30.0
WB_DELIVERY_BASE = 50.6
WB_DELIVERY_L    = 15.4
WB_RETURN_BASE   = 136.0
WB_RETURN_L      = 14.0
DEFAULT_VOLUME_L = 1.5     # небольшой товар (фильтр, колодки) ~1.5 л
DEFAULT_COMMISSION = 0.30  # fallback; берётся из MPStats по предмету
TARGET_MARGIN    = 0.12

# ─── Параметры запросов ──────────────────────────────────────────────────────

WB_DELAY        = (3.0, 6.0)   # случайная пауза между WB-запросами, с
WB_RETRY_WAITS  = [60, 180, 360]  # паузы при 429, с
MPSTATS_DELAY   = 1.0
MPSTATS_RETRIES = 3

log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  WB FBS маржа
# ══════════════════════════════════════════════════════════════════════════════

def calc_wb_margin(purchase: float, sell: float,
                   volume_l: float = DEFAULT_VOLUME_L,
                   commission: float = DEFAULT_COMMISSION) -> float:
    if sell <= 0 or purchase <= 0:
        return 0.0
    delivery    = WB_DELIVERY_BASE + WB_DELIVERY_L * volume_l
    return_cost = (WB_RETURN_BASE + WB_RETURN_L * volume_l) * WB_RET_RATE
    com         = sell * commission
    acq         = sell * WB_ACQ_PCT
    spp         = sell * WB_SPP_RESERVE
    proceeds    = sell - com - acq - spp - delivery
    tax         = max(0.0, proceeds) * WB_TAX_PCT
    total       = purchase + com + acq + spp + delivery + return_cost + WB_PACKAGING + tax
    return (sell - total) / sell


def score_item(oos_pct: float, monthly_sales: float, margin: float) -> float:
    return (oos_pct / 100.0) * math.log1p(monthly_sales) * max(0.0, margin / TARGET_MARGIN)


# ══════════════════════════════════════════════════════════════════════════════
#  WB поиск (curl_cffi — Chrome TLS)
# ══════════════════════════════════════════════════════════════════════════════

_WB_SESSION = None

def _get_wb_session():
    global _WB_SESSION
    if _WB_SESSION is not None:
        return _WB_SESSION

    if CURL_AVAILABLE:
        s = cffi_requests.Session(impersonate="chrome124")
    else:
        s = _requests.Session()
        s.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/124.0.0.0 Safari/537.36",
            "Accept-Language": "ru-RU,ru;q=0.9",
        })

    # Прогрев — получаем cookies WB
    try:
        s.get("https://www.wildberries.ru/", timeout=15)
        time.sleep(random.uniform(1.5, 2.5))
    except Exception as e:
        log.warning(f"WB прогрев: {e}")

    _WB_SESSION = s
    return s


def _wb_search_cache(article: str) -> Path:
    safe = article.replace(" ", "_").replace("/", "-").replace("\\", "-")[:60]
    return CACHE_DIR / f"wbs_{safe}.json"


def search_wb(article: str, limit: int = 6) -> list[dict]:
    """
    Поиск WB по артикулу. Возвращает [{nm_id, name, brand, subject_id, subject}].
    Кэшируется навсегда (данные о наличии товара на WB меняются редко).
    """
    cache = _wb_search_cache(article)
    if cache.exists():
        try:
            data = json.loads(cache.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return data
        except Exception:
            cache.unlink(missing_ok=True)

    s = _get_wb_session()

    url = "https://search.wb.ru/exactmatch/ru/common/v5/search"
    params = {
        "query":     article,
        "resultset": "catalog",
        "limit":     min(limit * 4, 100),
        "sort":      "popular",
        "page":      1,
        "dest":      -1257786,
        "appType":   1,
        "curr":      "rub",
        "spp":       27,
    }

    for attempt, wait in enumerate([0] + WB_RETRY_WAITS):
        if wait:
            log.warning(f"  WB rate-limit (попытка {attempt}), ждём {wait}с...")
            time.sleep(wait)
            _WB_SESSION = None  # пересоздаём сессию

        try:
            r = s.get(url, params=params, timeout=20)

            if r.status_code == 429:
                continue

            if r.status_code != 200:
                log.debug(f"  WB [{article}]: HTTP {r.status_code}")
                return []

            raw = r.json()
            products_raw = (
                raw.get("products")
                or (raw.get("data") or {}).get("products")
                or []
            )

            products = []
            for p in products_raw:
                nm_id = p.get("id")
                if not nm_id:
                    continue
                products.append({
                    "nm_id":      nm_id,
                    "name":       p.get("name", ""),
                    "brand":      p.get("brand", ""),
                    "subject_id": p.get("subjectId", 0),
                    "subject":    p.get("subjectName", ""),
                    "rating":     p.get("rating", 0),
                    "feedbacks":  p.get("feedbacks", 0),
                })
                if len(products) >= limit:
                    break

            cache.write_text(json.dumps(products, ensure_ascii=False), encoding="utf-8")
            return products

        except Exception as e:
            log.debug(f"  WB [{article}]: {e}")
            if attempt < len(WB_RETRY_WAITS):
                continue
            return []

    return []


# ══════════════════════════════════════════════════════════════════════════════
#  MPStats per-item
# ══════════════════════════════════════════════════════════════════════════════

_MP_SESSION = None

def _mp_headers():
    return {"X-Mpstats-TOKEN": MPSTATS_TOKEN, "Content-Type": "application/json"}


def _mpstats_item_cache(nm_id: int, d1: str) -> Path:
    return CACHE_DIR / f"mpitem_{nm_id}_{d1}.json"


def get_item_stats(nm_id: int, d1: str, d2: str) -> dict | None:
    """
    GET /api/wb/get/item/{nm_id}/by_category → ежедневная аналитика.
    Возвращает агрегат: oos_pct, monthly_sales, avg_price, commission.
    """
    cache = _mpstats_item_cache(nm_id, d1)
    if cache.exists():
        try:
            return json.loads(cache.read_text(encoding="utf-8"))
        except Exception:
            cache.unlink(missing_ok=True)

    url = f"https://mpstats.io/api/wb/get/item/{nm_id}/by_category"
    params = {"d1": d1, "d2": d2}

    for attempt in range(MPSTATS_RETRIES):
        try:
            r = _requests.get(url, headers=_mp_headers(), params=params, timeout=30)

            if r.status_code == 429:
                wait = 30 * (attempt + 1)
                log.warning(f"  MPStats rate-limit, ждём {wait}с...")
                time.sleep(wait)
                continue

            if r.status_code == 401:
                log.error("MPStats: неверный токен"); sys.exit(1)

            if r.status_code == 404:
                return None  # товар не найден в MPStats

            if r.status_code != 200:
                log.debug(f"  MPStats item {nm_id}: HTTP {r.status_code}")
                return None

            raw = r.json()
            result = _parse_item_stats(raw)
            if result:
                cache.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")
            return result

        except Exception as e:
            log.debug(f"  MPStats item {nm_id}: {e}")
            time.sleep(3)

    return None


def _parse_item_stats(raw: dict) -> dict | None:
    """Агрегирует ежедневные данные MPStats в суммарные метрики за период."""
    if not isinstance(raw, dict):
        return None

    # Может быть как прямой dict, так и {"data": {...}}
    data = raw.get("data", raw)

    balance     = data.get("balance",     data.get("stocks", []))
    sales       = data.get("sales",       [])
    final_price = data.get("final_price", data.get("price", []))

    # Фолбэк: скалярные поля
    if not isinstance(sales, list):
        return {
            "oos_pct":       float(data.get("lost_profit_percent", 0) or 0),
            "monthly_sales": float(data.get("sales", 0) or 0),
            "avg_price":     float(data.get("avg_price", 0) or data.get("final_price", 0) or 0),
            "commission":    float(data.get("commission_fbs", DEFAULT_COMMISSION * 100) or (DEFAULT_COMMISSION * 100)) / 100.0,
            "days":          30,
        }

    period      = len(balance) or len(sales) or 1
    oos_days    = sum(1 for b in balance if (b or 0) == 0)
    total_sales = sum(s or 0 for s in sales)
    prices      = [p for p in final_price if p and p > 0]
    avg_price   = sum(prices) / len(prices) if prices else 0

    if total_sales < 1:
        return None

    return {
        "oos_pct":       round(oos_days / period * 100, 1),
        "monthly_sales": round(total_sales * 30 / period, 1),  # нормализуем к 30 дням
        "avg_price":     round(avg_price, 0),
        "commission":    float(data.get("commission_fbs", DEFAULT_COMMISSION * 100) or (DEFAULT_COMMISSION * 100)) / 100.0,
        "days":          period,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  Загрузка каталогов + топ-ниш
# ══════════════════════════════════════════════════════════════════════════════

def load_mikado_catalog() -> list[dict]:
    """[{article, brand, name, price}]"""
    if not MIKADO_PRICE.exists():
        log.warning(f"Mikado: {MIKADO_PRICE} не найден")
        return []
    try:
        wb = openpyxl.load_workbook(str(MIKADO_PRICE), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        log.error(f"Mikado: {e}"); return []

    items = []
    ci = {}
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            hdr = [str(c).strip().lower() if c else "" for c in row]
            ci = {
                "code":  next((i for i, h in enumerate(hdr) if h == "code"),    None),
                "brand": next((i for i, h in enumerate(hdr) if h in ("brandname", "brand")), None),
                "name":  next((i for i, h in enumerate(hdr) if h in ("prodname", "name")),   None),
                "price": next((i for i, h in enumerate(hdr) if h in ("priceout", "price")),  None),
            }
            if ci["code"] is not None and ci["price"] is not None:
                header_found = True
            continue
        try:
            article = str(row[ci["code"]]  or "").strip()
            brand   = str(row[ci["brand"]] or "").strip() if ci["brand"] is not None else ""
            name    = str(row[ci["name"]]  or "").strip() if ci["name"]  is not None else ""
            price   = float(row[ci["price"]] or 0)
        except (TypeError, ValueError, IndexError):
            continue
        if article and price > 0:
            items.append({"article": article, "brand": brand,
                          "name": name.lower(), "price": price, "source": "mikado"})

    try: wb.close()
    except Exception: pass
    log.info(f"Mikado: {len(items):,} позиций")
    return items


def load_autoliga_catalog() -> list[dict]:
    """[{article, brand, name, price}]"""
    scripts_dir = Path(__file__).parent
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    try:
        from autoliga_loader import load_autoliga
    except ImportError as e:
        log.error(f"autoliga_loader: {e}"); return []
    catalog = load_autoliga()
    items = []
    for v in catalog.values():
        article = v.get("article", "")
        price   = float(v.get("price") or 0)
        if article and price > 0:
            items.append({
                "article": article,
                "brand":   v.get("brand", ""),
                "name":    v.get("name", "").lower(),
                "price":   price,
                "source":  "autoliga",
            })
    log.info(f"Автолига: {len(items):,} позиций")
    return items


def load_top_niches(top_n: int = 25) -> list[dict]:
    """
    Загружает топ-N ниш из последнего wb_niches_*.xlsx.
    Возвращает [{subject, score, oos_pct, avg_price, sales_month, commission_pct}]
    """
    files = sorted(OUT_DIR.glob("wb_niches_*.xlsx"),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        log.warning("wb_niches_*.xlsx не найден — запусти сначала wb_niche_analyzer.py")
        return []

    path = files[0]
    log.info(f"Ниши Level 1: {path.name}")
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        log.error(f"Ниши: {e}"); return []

    niches = []
    hdr_map = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri == 0:
            hdr_map = {str(v).strip(): i for i, v in enumerate(row) if v}
            continue
        if len(niches) >= top_n:
            break
        def _v(name, default=0):
            i = hdr_map.get(name)
            return row[i] if i is not None and i < len(row) else default
        subject = str(_v("Предмет WB", "")).strip()
        if not subject:
            continue
        niches.append({
            "subject":        subject,
            "score":          float(_v("Скор") or 0),
            "oos_pct":        float(_v("OOS%") or 0),
            "avg_price":      float(_v("Ср.цена WB") or 0),
            "sales_month":    float(_v("Прод/мес") or 0),
            "commission_pct": float(_v("Ком.WB%") or 0),
        })

    try: wb.close()
    except Exception: pass
    log.info(f"Загружено {len(niches)} ниш (топ-{top_n})")
    return niches


# ══════════════════════════════════════════════════════════════════════════════
#  Keyword matching ниши → каталог (упрощённый)
# ══════════════════════════════════════════════════════════════════════════════

NAME_STOPWORDS = frozenset({
    "для", "авто", "автомобильные", "автомобильное", "автомобильный",
    "автомобиля", "и", "в", "на", "с", "из", "к", "по",
    "запчасти", "деталь", "детали",
})


def _stems(text: str) -> list[str]:
    words = [w.strip(".,;:()-") for w in text.lower().split()]
    return [w[:6] for w in words if w not in NAME_STOPWORDS and len(w) >= 4]


def match_items_to_niche(subject: str, catalog: list[dict]) -> list[dict]:
    """Товары каталога, чьё название пересекается со стемами названия ниши."""
    parts = subject.split("/")
    search_part = parts[-1].strip() if len(parts) > 1 else subject
    stems = _stems(search_part)
    if not stems:
        return []
    return [item for item in catalog if any(s in item["name"] for s in stems)]


# ══════════════════════════════════════════════════════════════════════════════
#  Приоритизация артикулов
# ══════════════════════════════════════════════════════════════════════════════

def build_priority_list(niches: list[dict], catalog: list[dict],
                        max_articles: int = 1000) -> list[dict]:
    """
    Для каждой ниши находим совпадающие артикулы, назначаем им niche_score,
    дедупликация по артикулу (берём ниш с максимальным score).
    Сортируем: нишевый скор × (1/log(цена)) — дешёвые с высоким OOS первые.
    """
    by_article: dict[str, dict] = {}

    for niche in niches:
        matched = match_items_to_niche(niche["subject"], catalog)
        for item in matched:
            art = item["article"]
            priority = niche["score"] / math.log1p(max(item["price"], 1))
            if art not in by_article or priority > by_article[art]["priority"]:
                by_article[art] = {
                    **item,
                    "niche":          niche["subject"],
                    "niche_score":    niche["score"],
                    "niche_oos":      niche["oos_pct"],
                    "niche_avg_price":niche["avg_price"],
                    "niche_comm":     niche["commission_pct"] / 100.0,
                    "priority":       priority,
                }

    result = sorted(by_article.values(), key=lambda x: -x["priority"])
    log.info(f"Приоритетных артикулов: {len(result)}, берём топ-{max_articles}")
    return result[:max_articles]


# ══════════════════════════════════════════════════════════════════════════════
#  Excel
# ══════════════════════════════════════════════════════════════════════════════

COLUMNS = [
    ("rank",          "#",              4),
    ("article",       "Артикул",       16),
    ("brand",         "Бренд",         14),
    ("our_price",     "Наша цена",     11),
    ("wb_avg_price",  "Цена WB",       10),
    ("margin_pct",    "Маржа%",         9),
    ("oos_pct",       "OOS%",           7),
    ("monthly_sales", "Прод/мес",      10),
    ("score",         "Скор",           9),
    ("niche",         "Ниша",          42),
    ("wb_name",       "Название WB",   35),
    ("wb_brand",      "Бренд WB",      14),
    ("nm_id",         "nmID",          12),
    ("feedbacks",     "Отзывов",        9),
    ("source",        "Источник",      10),
    ("wb_link",       "Ссылка WB",     35),
]


def write_excel(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WB Товары"

    keys    = [c[0] for c in COLUMNS]
    headers = [c[1] for c in COLUMNS]
    widths  = [c[2] for c in COLUMNS]

    ws.append(headers)

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r.get(k) for k in keys])

    oos_col    = next(i+1 for i, c in enumerate(COLUMNS) if c[0] == "oos_pct")
    margin_col = next(i+1 for i, c in enumerate(COLUMNS) if c[0] == "margin_pct")

    for ri in range(2, len(rows) + 2):
        oos = float(ws.cell(row=ri, column=oos_col).value or 0)
        if oos >= 50:
            ws.cell(row=ri, column=oos_col).fill = PatternFill("solid", fgColor="C00000")
            ws.cell(row=ri, column=oos_col).font = Font(color="FFFFFF", bold=True)
        elif oos >= 35:
            ws.cell(row=ri, column=oos_col).fill = PatternFill("solid", fgColor="FF4444")
            ws.cell(row=ri, column=oos_col).font = Font(color="FFFFFF", bold=True)
        elif oos >= 20:
            ws.cell(row=ri, column=oos_col).fill = PatternFill("solid", fgColor="FFA500")

        mgn = float(ws.cell(row=ri, column=margin_col).value or 0)
        if mgn >= 15:
            ws.cell(row=ri, column=margin_col).fill = PatternFill("solid", fgColor="70AD47")
            ws.cell(row=ri, column=margin_col).font = Font(color="FFFFFF")
        elif mgn >= 10:
            ws.cell(row=ri, column=margin_col).fill = PatternFill("solid", fgColor="FFEB84")
        elif mgn < 0:
            ws.cell(row=ri, column=margin_col).fill = PatternFill("solid", fgColor="FFCCCC")

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(str(path))


# ══════════════════════════════════════════════════════════════════════════════
#  main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="WB Item Analyzer — Level 2")
    parser.add_argument("--days",        type=int,   default=30)
    parser.add_argument("--top-niches",  type=int,   default=25,  help="Сколько ниш из Level 1")
    parser.add_argument("--articles",    type=int,   default=1000, help="Макс. артикулов к поиску")
    parser.add_argument("--wb-results",  type=int,   default=5,   help="Товаров WB на артикул")
    parser.add_argument("--min-sales",   type=float, default=3.0, help="Мин.продаж/мес на WB")
    parser.add_argument("--min-oos",     type=float, default=10.0,help="Мин.OOS%%")
    parser.add_argument("--min-margin",  type=float, default=5.0, help="Мин.маржа%%")
    parser.add_argument("--top-results", type=int,   default=1000,help="Строк в итоговом отчёте")
    parser.add_argument("--min-price",   type=float, default=100.0,help="Мин.наша цена (отсекает болты/гайки)")
    parser.add_argument("--skip-mikado",   action="store_true")
    parser.add_argument("--skip-autoliga", action="store_true")
    parser.add_argument("--dry-run",       action="store_true",  help="Только каталог, без WB/MPStats")
    parser.add_argument("--refresh-wb",    action="store_true",  help="Очистить кэш WB поиска")
    args = parser.parse_args()

    for d in (CACHE_DIR, OUT_DIR, LOG_FILE.parent):
        d.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
        datefmt="%H:%M:%S",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(str(LOG_FILE), encoding="utf-8", errors="replace"),
        ],
    )

    if not MPSTATS_TOKEN:
        log.error("MPSTATS_TOKEN не задан в .env"); sys.exit(1)

    if args.refresh_wb:
        deleted = sum(1 for f in CACHE_DIR.glob("wbs_*.json") if f.unlink() or True)
        log.info(f"Кэш WB очищен: {deleted} файлов")

    d2 = datetime.now().strftime("%Y-%m-%d")
    d1 = (datetime.now() - timedelta(days=args.days)).strftime("%Y-%m-%d")
    log.info(f"Период: {d1}…{d2} | ниш: {args.top_niches} | артикулов: {args.articles}")

    if CURL_AVAILABLE:
        log.info("curl_cffi доступен — WB поиск через Chrome TLS fingerprint")
    else:
        log.warning("curl_cffi НЕ доступен — WB может блокировать запросы")

    # ─── 1. Загружаем ниши и каталог ─────────────────────────────────────────
    niches = load_top_niches(args.top_niches)
    if not niches and not args.dry_run:
        log.error("Нет ниш — сначала запусти wb_niche_analyzer.py"); sys.exit(1)

    catalog: list[dict] = []
    if not args.skip_mikado:
        catalog.extend(load_mikado_catalog())
    if not args.skip_autoliga:
        catalog.extend(load_autoliga_catalog())
    log.info(f"Каталог: {len(catalog):,} позиций")

    # ─── 2. Приоритетный список артикулов ────────────────────────────────────
    # Отсекаем слишком дешёвые позиции (болты, гайки — неправильный keyword match)
    catalog = [item for item in catalog if item["price"] >= args.min_price]
    log.info(f"Каталог после фильтра цены (>={args.min_price:.0f}₽): {len(catalog):,}")

    if niches:
        priority = build_priority_list(niches, catalog, args.articles)
    else:
        # dry-run без ниш: берём первые N из каталога
        priority = [
            {**item, "niche": "N/A", "niche_score": 0, "niche_oos": 0,
             "niche_avg_price": 0, "niche_comm": DEFAULT_COMMISSION, "priority": 0}
            for item in catalog[:args.articles]
        ]

    if not priority:
        log.error("Нет артикулов для анализа — проверь каталоги и ниши"); sys.exit(1)

    log.info(f"К обработке: {len(priority)} артикулов")

    if args.dry_run:
        log.info("=== DRY RUN — WB и MPStats не вызываются ===")
        for i, item in enumerate(priority[:20], 1):
            log.info(f"  {i:>3}. {item['article']:<20} {item['brand']:<15} "
                     f"{item['price']:>7.0f}₽  ниша: {item['niche'][:40]}")
        return

    # ─── 3. WB поиск + MPStats ────────────────────────────────────────────────
    results: list[dict] = []
    total = len(priority)

    for idx, item in enumerate(priority, 1):
        article    = item["article"]
        our_price  = item["price"]
        niche_comm = item.get("niche_comm", DEFAULT_COMMISSION)

        cached = _wb_search_cache(article).exists()
        log.info(f"[{idx}/{total}] {article} ({item['brand']}) "
                 f"{'[кэш]' if cached else '[WB]'}")

        wb_products = search_wb(article, limit=args.wb_results)
        if not cached:
            time.sleep(random.uniform(*WB_DELAY))

        if not wb_products:
            continue

        # Для каждого nmID — MPStats аналитика
        for prod in wb_products:
            nm_id = prod["nm_id"]
            stats = get_item_stats(nm_id, d1, d2)
            time.sleep(MPSTATS_DELAY)

            if not stats:
                continue

            oos_pct       = stats["oos_pct"]
            monthly_sales = stats["monthly_sales"]
            wb_avg_price  = stats["avg_price"] or item.get("niche_avg_price", 0)
            commission    = stats.get("commission") or niche_comm

            if monthly_sales < args.min_sales:
                continue
            if oos_pct < args.min_oos:
                continue

            margin = calc_wb_margin(our_price, wb_avg_price, commission=commission)
            if margin * 100 < args.min_margin:
                continue

            sc = score_item(oos_pct, monthly_sales, margin)

            results.append({
                "article":       article,
                "brand":         item["brand"],
                "our_price":     round(our_price),
                "wb_avg_price":  round(wb_avg_price),
                "margin_pct":    round(margin * 100, 1),
                "oos_pct":       round(oos_pct, 1),
                "monthly_sales": round(monthly_sales, 1),
                "score":         round(sc, 4),
                "niche":         item["niche"],
                "wb_name":       prod["name"][:60],
                "wb_brand":      prod["brand"],
                "nm_id":         nm_id,
                "feedbacks":     prod["feedbacks"],
                "source":        item["source"],
                "wb_link":       f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
            })

    if not results:
        log.warning("Нет результатов — возможно WB блокирует или кэш пуст")
        log.warning("Проверь: IP не заблокирован, curl_cffi установлен")
        return

    # ─── 4. Дедупликация и сортировка ────────────────────────────────────────
    # Если один артикул нашли несколько раз (разные nmID) — берём лучший результат
    by_article: dict[str, dict] = {}
    for r in results:
        art = r["article"]
        if art not in by_article or r["score"] > by_article[art]["score"]:
            by_article[art] = r

    final = sorted(by_article.values(), key=lambda x: -x["score"])[:args.top_results]

    for i, r in enumerate(final, 1):
        r["rank"] = i

    log.info(f"Найдено уникальных артикулов: {len(final)}")

    # ─── 5. Excel ─────────────────────────────────────────────────────────────
    ts    = datetime.now().strftime("%Y%m%d_%H%M")
    fpath = OUT_DIR / f"wb_items_{ts}.xlsx"
    write_excel(final, fpath)
    log.info(f"Готово! {len(final)} строк → {fpath}")

    # Топ-25 в консоль
    w = 110
    print(f"\n{'─'*w}")
    print(f"{'#':>3} {'Артикул':<18} {'Бренд':<12} {'Цена':>7} {'WB цена':>8} "
          f"{'Маржа':>6} {'OOS%':>5} {'Прод':>6} {'Скор':>8}")
    print(f"{'─'*w}")
    for r in final[:25]:
        print(
            f"{r['rank']:>3} {r['article']:<18} {r['brand'][:11]:<12} "
            f"{r['our_price']:>6.0f}₽ {r['wb_avg_price']:>7.0f}₽ "
            f"{r['margin_pct']:>5.1f}% {r['oos_pct']:>4.0f}% "
            f"{r['monthly_sales']:>6.0f} {r['score']:>8.4f}"
        )
    print(f"{'─'*w}")
    print(f"\n{fpath}\n")


if __name__ == "__main__":
    main()
