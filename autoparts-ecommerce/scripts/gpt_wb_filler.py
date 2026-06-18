# /// script
# requires-python = ">=3.10"
# dependencies = ["playwright", "openpyxl"]
# ///
"""
gpt_wb_filler.py
════════════════
Заполняет столбцы Параметры / Применяемость / Описание / Вес / Длина / Ширина / Высота
в файле «Топ-500 ВБ_new.xlsx» через ChatGPT (браузер, Playwright).
Описание — 1200–1500 символов для маркетплейса WB.
Уже заполненные ячейки пропускаются. Сохранение каждые 5 обработанных строк.

Установка браузера (один раз):
  uv run --with playwright python -m playwright install chromium

Первый запуск — войти в ChatGPT:
  cd C:\\Users\\Admin\\Documents\\Autoparts_Ecommerce
  uv run --with playwright,openpyxl scripts/gpt_wb_filler.py --login

Обычный запуск:
  uv run --with playwright,openpyxl scripts/gpt_wb_filler.py

Тест (строки 2–4, видимый браузер):
  uv run --with playwright,openpyxl scripts/gpt_wb_filler.py --rows 2-4 --debug

Конкретная строка:
  uv run --with playwright,openpyxl scripts/gpt_wb_filler.py --rows 5
"""

import re
import sys
import time
import argparse
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("Playwright не установлен.")
    print("  uv run --with playwright python -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment
except ImportError:
    print("openpyxl не установлен")
    sys.exit(1)


# ──────────────────────────────────────────────────────────
XLSX_PATH   = Path(r"C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ_new.xlsx")
PROFILE_DIR = Path.home() / ".chatgpt_playwright"
CHATGPT_URL = "https://chatgpt.com/"
SAVE_EVERY  = 5   # сохранять каждые N обработанных строк
# ──────────────────────────────────────────────────────────


# ══════════════════════════════════════════════════════════
# ПРОМПТ
# ══════════════════════════════════════════════════════════

def build_prompt(row: dict) -> str:
    missing = row["missing"]

    format_parts = []
    if "Параметры" in missing:
        format_parts.append(
            "ПАРАМЕТРЫ:\n"
            "[технические характеристики, каждая на новой строке в формате «Ключ: Значение»]"
        )
    if "Применяемость" in missing:
        format_parts.append(
            "ПРИМЕНЯЕМОСТЬ:\n"
            "[список автомобилей через точку с запятой: Марка Модель ГодНачала-ГодКонца]"
        )
    if "Описание" in missing:
        format_parts.append(
            "ОПИСАНИЕ:\n"
            "[текст 1200–1500 символов для маркетплейса WB;\n"
            " информативный тон, без восклицательных знаков, без HTML, абзацами]"
        )
    if "Вес" in missing:
        format_parts.append("ВЕС_Г: [целое число или нет_данных]")
    if "Длина" in missing:
        format_parts.append("ДЛИНА_ММ: [целое число или нет_данных]")
    if "Ширина" in missing:
        format_parts.append("ШИРИНА_ММ: [целое число или нет_данных]")
    if "Высота" in missing:
        format_parts.append("ВЫСОТА_ММ: [целое число или нет_данных]")

    format_str = "\n\n".join(format_parts)

    # Строим контекст для поиска
    search_hints = []
    if row["oem"]:
        search_hints.append(f"OEM: {row['oem']}")
    if row["alt"]:
        search_hints.append(f"аналоги: {row['alt'][:120]}")
    search_str = "; ".join(search_hints)

    return (
        f"Найди в интернете технические данные об автозапчасти и заполни поля ниже.\n"
        f"Используй поиск по OEM-номерам на сайтах: tecdoc.net, exist.ru, autodoc.ru, "
        f"emex.ru, сайтах производителей. Если одни источники не дают размеры — ищи в других.\n\n"
        f"ТОВАР:\n"
        f"  Наименование: {row['name']}\n"
        f"  Бренд: {row['brand']}\n"
        f"  Артикул: {row['code']}\n"
        f"  {search_str}\n\n"
        f"Ответ ТОЛЬКО в этом формате (секции разделяй пустой строкой, никаких пояснений до или после):\n\n"
        f"{format_str}\n\n"
        f"Числа — только цифры без единиц измерения. "
        f"Если данные не найдены нигде — пиши «нет_данных»."
    )


# ══════════════════════════════════════════════════════════
# ПАРСИНГ ОТВЕТА
# ══════════════════════════════════════════════════════════

# Маркеры секций
_SECTION_MARKERS = (
    "ПАРАМЕТРЫ", "ПРИМЕНЯЕМОСТЬ", "ОПИСАНИЕ",
    "ВЕС_Г", "ДЛИНА_ММ", "ШИРИНА_ММ", "ВЫСОТА_ММ",
)
_SPLIT_RE = re.compile(
    r'\n(' + '|'.join(_SECTION_MARKERS) + r')\s*:',
    re.IGNORECASE,
)
_NUMERIC_CLEANUP_RE = re.compile(r'\d+')


def parse_response(text: str, missing: list) -> dict:
    # Убираем markdown-жирность (**) и лишние пробелы
    text = re.sub(r'\*+', '', text)
    # Гарантируем перенос строки перед первой секцией
    text = "\n" + text

    # Разбиваем на секции
    parts = _SPLIT_RE.split(text)
    # parts: [до_первой_секции, key1, val1, key2, val2, ...]
    sections: dict[str, str] = {}
    for i in range(1, len(parts), 2):
        if i + 1 >= len(parts):
            break
        key = parts[i].upper().strip()
        val = parts[i + 1].strip()
        # Обрезаем значение на случай вложенных маркеров
        val = _SPLIT_RE.split(val)[0].strip()
        sections[key] = val

    result: dict = {}

    def _is_empty(val: str) -> bool:
        return not val or val.lower() in ("нет_данных", "нет данных", "-", "")

    # Текстовые секции
    for field, key in [("Параметры", "ПАРАМЕТРЫ"), ("Применяемость", "ПРИМЕНЯЕМОСТЬ"), ("Описание", "ОПИСАНИЕ")]:
        if field not in missing:
            continue
        val = sections.get(key, "")
        if not _is_empty(val):
            result[field] = val

    # Числовые поля
    for field, key in [("Вес", "ВЕС_Г"), ("Длина", "ДЛИНА_ММ"), ("Ширина", "ШИРИНА_ММ"), ("Высота", "ВЫСОТА_ММ")]:
        if field not in missing:
            continue
        raw = sections.get(key, "")
        # Также ищем «KEY: число» прямо в тексте (резерв)
        if _is_empty(raw):
            m = re.search(rf'{key}\s*:?\s*(\d+)', text, re.IGNORECASE)
            raw = m.group(1) if m else ""
        if raw:
            m = _NUMERIC_CLEANUP_RE.search(raw)
            if m:
                result[field] = int(m.group(0))

    return result


# ══════════════════════════════════════════════════════════
# PLAYWRIGHT — ВЗАИМОДЕЙСТВИЕ С CHATGPT
# ══════════════════════════════════════════════════════════

_INPUT_SELECTORS = [
    '#prompt-textarea',
    'div[contenteditable="true"][id="prompt-textarea"]',
    '[data-testid="composer-input"]',
    'div.ProseMirror[contenteditable="true"]',
    'p[data-placeholder*="Message"]',
    'p[data-placeholder*="Сообщение"]',
    'div[contenteditable="true"]',
    'textarea',
]

_SEND_SELECTORS = [
    'button[data-testid="send-button"]',
    'button[aria-label*="Send"]',
    'button[aria-label*="send"]',
]

# Кнопка веб-поиска в ChatGPT (селекторы меняются с обновлениями UI)
_SEARCH_SELECTORS = [
    'button[aria-label="Search the web"]',
    'button[aria-label*="Search"]',
    'button[data-testid="composer-search-button"]',
    'button[aria-label*="Поиск"]',
    '[aria-label*="web search"]',
    'button[aria-label*="browse"]',
]

_STOP_SELECTORS = [
    'button[data-testid="stop-button"]',
    'button[aria-label*="Stop"]',
    'button[aria-label*="stop"]',
]


def _find_visible(page, selectors: list, timeout: int = 2000):
    """Ждёт появления первого видимого элемента из списка селекторов."""
    for sel in selectors:
        try:
            page.wait_for_selector(sel, state="visible", timeout=timeout)
            return page.locator(sel).last
        except Exception:
            continue
    return None


def dismiss_modals(page):
    """Закрывает попапы и баннеры ChatGPT (upgrade, cookies, welcome)."""
    close_selectors = [
        'button[aria-label="Close"]',
        'button[data-testid="close-button"]',
        'button:has-text("Dismiss")',
        'button:has-text("Maybe later")',
        'button:has-text("No thanks")',
        'button:has-text("OK")',
        'button:has-text("Ok")',
        'button:has-text("Got it")',
        'button:has-text("Stay on free plan")',
        'button:has-text("Keep current plan")',
        'button:has-text("Skip for now")',
        'button:has-text("Continue")',
        'button:has-text("Понятно")',
        'button:has-text("Понял")',
        'button:has-text("Хорошо")',
        'button:has-text("Закрыть")',
        'button:has-text("ОК")',
        'button:has-text("Ок")',
        'button:has-text("Остаться")',
        'button:has-text("Продолжить")',
        'button:has-text("Позже")',
        '[data-testid="modal-close"]',
    ]
    for sel in close_selectors:
        try:
            btn = page.locator(sel).first
            if btn.is_visible():
                btn.click()
                time.sleep(0.4)
        except Exception:
            continue
    try:
        page.keyboard.press("Escape")
        time.sleep(0.3)
    except Exception:
        pass


def type_message(page, text: str):
    dismiss_modals(page)
    # Ждём поле ввода (по 8 сек на каждый селектор)
    el = _find_visible(page, _INPUT_SELECTORS, timeout=8_000)
    if el is None:
        # Повторная попытка после дополнительного ожидания
        time.sleep(4)
        dismiss_modals(page)
        el = _find_visible(page, _INPUT_SELECTORS, timeout=8_000)
    if el is None:
        raise RuntimeError("Поле ввода ChatGPT не найдено")

    el.click()
    time.sleep(0.3)

    page.evaluate(
        """([selectors, text]) => {
            let el = null;
            for (const sel of selectors) {
                el = document.querySelector(sel);
                if (el) break;
            }
            if (!el) return;
            el.focus();
            if (el.isContentEditable) {
                document.execCommand('selectAll', false, null);
                document.execCommand('insertText', false, text);
            } else {
                const setter = Object.getOwnPropertyDescriptor(
                    HTMLTextAreaElement.prototype, 'value').set;
                setter.call(el, text);
                el.dispatchEvent(new Event('input',  {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
            }
        }""",
        [_INPUT_SELECTORS, text],
    )
    time.sleep(0.5)


def enable_web_search(page) -> bool:
    """Включает веб-поиск в ChatGPT (если кнопка есть в UI). Возвращает True при успехе."""
    btn = _find_visible(page, _SEARCH_SELECTORS, timeout=2000)
    if btn:
        try:
            btn.click()
            time.sleep(0.5)
            return True
        except Exception:
            pass
    return False


def click_send(page):
    btn = _find_visible(page, _SEND_SELECTORS, timeout=3000)
    if btn:
        btn.click()
    else:
        page.keyboard.press("Enter")


def is_generating(page) -> bool:
    return _find_visible(page, _STOP_SELECTORS, timeout=500) is not None


def get_last_response(page) -> str:
    try:
        msgs = page.locator('[data-message-author-role="assistant"]').all()
        if msgs:
            return msgs[-1].inner_text(timeout=2000).strip()
    except Exception:
        pass
    try:
        # Fallback: article / generic message containers
        for sel in ['article', '[class*="message"]:last-child']:
            items = page.locator(sel).all()
            for item in reversed(items):
                txt = item.inner_text(timeout=1000).strip()
                if txt and len(txt) > 30:
                    return txt
    except Exception:
        pass
    return ""


_RATE_LIMIT_PHRASES = (
    "too many requests", "sending messages too quickly",
    "you've reached", "reached our limit",
    "слишком много запросов", "слишком быстро",
    "come back later", "try again later",
)

def _is_rate_limited(text: str) -> bool:
    low = text.lower()
    return any(p in low for p in _RATE_LIMIT_PHRASES)


def wait_for_response(page, timeout_sec: int = 180) -> str:
    print("    жду", end="", flush=True)

    # Ждём появления текста (до 60 сек), попутно закрываем попапы
    for _ in range(30):
        dismiss_modals(page)
        text = get_last_response(page)
        if text and len(text) > 20:
            break
        time.sleep(2)
        print(".", end="", flush=True)
    else:
        print(" [нет ответа]", flush=True)
        return ""

    # Ждём стабилизации, закрываем попапы каждые 10 сек
    prev, stable = "", 0
    for tick in range(timeout_sec):
        if tick % 10 == 0:
            dismiss_modals(page)
        cur = get_last_response(page)
        if cur == prev and cur:
            stable += 1
            if stable >= 3 and not is_generating(page):
                print(f" ({len(cur)} симв.)", flush=True)
                return cur
        else:
            stable = 0
            if tick % 5 == 0:
                print(".", end="", flush=True)
        prev = cur
        time.sleep(1)

    print(f" [таймаут, {len(prev)} симв.]", flush=True)
    return prev


def process_row(page, row: dict, delay: int, debug: bool, retries: int = 2) -> dict | None:
    for attempt in range(1, retries + 1):
        try:
            # Новый чат = переходим на главную
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30_000)
            try:
                page.wait_for_load_state("networkidle", timeout=10_000)
            except Exception:
                pass
            time.sleep(4)
            dismiss_modals(page)

            if debug and attempt == 1:
                page.screenshot(path=str(XLSX_PATH.parent / f"_debug_{row['code']}_start.png"))

            prompt = build_prompt(row)
            type_message(page, prompt)

            # Включаем веб-поиск (работает если есть в UI — Plus/Pro аккаунт)
            web_search_on = enable_web_search(page)
            if web_search_on:
                print("    [web search ON]", end=" ", flush=True)

            if debug and attempt == 1:
                page.screenshot(path=str(XLSX_PATH.parent / f"_debug_{row['code']}_typed.png"))
                print("    [debug] промпт введён, нажми Enter для отправки...")
                input()

            click_send(page)
            time.sleep(2)

            # Веб-поиск занимает дольше — увеличиваем таймаут
            response = wait_for_response(page, timeout_sec=300 if web_search_on else 180)

            if debug:
                debug_file = XLSX_PATH.parent / f"_debug_{row['code']}_response.txt"
                debug_file.write_text(response, encoding="utf-8")
                print(f"    [debug] ответ сохранён: {debug_file.name}")

            if not response:
                if attempt < retries:
                    print(f"    [!] Пустой ответ, попытка {attempt + 1}/{retries}...")
                    time.sleep(15)
                continue

            if _is_rate_limited(response):
                wait_min = 10
                print(f"    [!] ChatGPT: рейт-лимит — жду {wait_min} мин...")
                time.sleep(wait_min * 60)
                continue

            result = parse_response(response, row["missing"])
            if result:
                return result

            print(f"    [!] Ответ не распознан (попытка {attempt})")
            print(f"        Начало: {response[:150]!r}")
            if attempt < retries:
                time.sleep(10)

        except RuntimeError as e:
            print(f"    [!] {e}")
            if debug:
                page.screenshot(path=str(XLSX_PATH.parent / f"_debug_{row['code']}_error.png"))
            break
        except Exception as e:
            print(f"    [!] Ошибка (попытка {attempt}): {e}")
            if attempt < retries:
                time.sleep(15)

    return None


# ══════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════

def safe_save(wb, path: Path):
    for attempt in range(1, 6):
        try:
            wb.save(str(path))
            print(f"    [сохранено: {path.name}]")
            return
        except PermissionError:
            print(f"\n    [!] Файл занят — закройте Excel и нажмите Enter (попытка {attempt}/5)...")
            input()
    raise PermissionError(f"Не удалось сохранить {path} после 5 попыток")


# ══════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Заполняет столбцы Топ-500 ВБ через ChatGPT (браузер)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--login",  action="store_true", help="Открыть браузер для входа в ChatGPT")
    ap.add_argument("--debug",  action="store_true", help="Показывать браузер, сохранять скриншоты и ответы")
    ap.add_argument("--rows",   default=None,        help="Диапазон строк Excel: 2-10 или одна строка: 5")
    ap.add_argument("--delay",  default=45, type=int, help="Пауза между строками, сек (по умолч. 45)")
    args = ap.parse_args()

    PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    # Загружаем книгу
    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb.active
    headers = {
        ws.cell(1, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }

    col_code  = headers.get("Код / Артикул",               1)
    col_name  = headers.get("Наименование",                 2)
    col_brand = headers.get("Бренд",                        3)
    col_oem   = headers.get("OEM номера")
    col_alt   = headers.get("Альтернативные артикулы товара")

    # Целевые столбцы (GPT-поле → номер столбца Excel)
    target_map: dict[str, int | None] = {
        "Параметры":     headers.get("Параметры"),
        "Применяемость": headers.get("Применяемость"),
        "Описание":      headers.get("Описание"),
        "Вес":           headers.get("Вес, г"),
        "Длина":         headers.get("Длина, мм"),
        "Ширина":        headers.get("Ширина, мм"),
        "Высота":        headers.get("Высота, мм"),
    }

    # Диапазон строк
    max_row = ws.max_row
    row_start, row_end = 2, max_row
    if args.rows:
        if "-" in args.rows:
            a, b = args.rows.split("-")
            row_start, row_end = int(a), int(b)
        else:
            row_start = row_end = int(args.rows)
    row_end = min(row_end, max_row)

    print()
    print("═" * 62)
    print("  GPT WB Filler  [ChatGPT Browser + Playwright]")
    print(f"  Файл:    {XLSX_PATH.name}")
    print(f"  Строки:  {row_start}–{row_end} ({row_end - row_start + 1} шт.)")
    print(f"  Сохран.: каждые {SAVE_EVERY} обработанных строк")
    print(f"  Пауза:   {args.delay} сек между строками")
    print("═" * 62)

    headless = False  # headless блокируется ChatGPT — всегда показываем браузер

    with sync_playwright() as pw:
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR),
                channel="chrome",
                headless=headless,
                viewport={"width": 1280, "height": 900},
                locale="ru-RU",
                args=["--disable-blink-features=AutomationControlled"],
            )
        except Exception:
            # Chrome не установлен — используем встроенный Chromium
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR),
                headless=headless,
                viewport={"width": 1280, "height": 900},
                locale="ru-RU",
                args=["--disable-blink-features=AutomationControlled"],
            )

        page = context.new_page()

        # ── Режим входа ───────────────────────────────────────────
        if args.login:
            print("\n  Открываю браузер ChatGPT...")
            print("  Залогинься в свой аккаунт (chat.openai.com или chatgpt.com).")
            print("  Когда окажешься в чате — нажми Enter в этом терминале.")
            try:
                page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=20_000)
            except Exception:
                pass
            input("  > ")
            print("  ✓ Сессия сохранена в профиле браузера.")
            context.close()
            return

        # ── Проверяем авторизацию ─────────────────────────────────
        print("\n  Открываю ChatGPT...")
        try:
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30_000)
        except Exception as e:
            print(f"  ✗ Не удалось открыть ChatGPT: {e}")
            context.close()
            sys.exit(1)

        time.sleep(3)
        if any(k in page.url.lower() for k in ("login", "auth", "signin", "sign-in")):
            print("  ✗ Не авторизован! Запусти сначала:")
            print("    uv run --with playwright,openpyxl scripts/gpt_wb_filler.py --login")
            context.close()
            sys.exit(1)

        print("  ✓ Авторизован\n")

        stats:     dict[str, int] = {"done": 0, "skipped": 0, "failed": 0}
        since_save: int           = 0

        for row_num in range(row_start, row_end + 1):
            code = ws.cell(row_num, col_code).value
            if not code:
                continue

            code  = str(code).strip()
            name  = str(ws.cell(row_num, col_name).value  or "")
            brand = str(ws.cell(row_num, col_brand).value or "")
            oem   = str(ws.cell(row_num, col_oem).value   or "") if col_oem  else ""
            alt   = str(ws.cell(row_num, col_alt).value   or "") if col_alt  else ""

            # Какие целевые столбцы пусты?
            missing = [
                field for field, col in target_map.items()
                if col is not None and not ws.cell(row_num, col).value
            ]

            if not missing:
                print(f"  [{row_num:3}] {code}: пропуск (всё заполнено)")
                stats["skipped"] += 1
                continue

            print(f"\n  [{row_num:3}/{row_end}] {brand} {code} — {name[:45]}")
            print(f"         нужно:  {', '.join(missing)}")

            row_data = {
                "code": code, "name": name, "brand": brand,
                "oem": oem, "alt": alt, "missing": missing,
            }
            result = process_row(page, row_data, delay=args.delay, debug=args.debug)

            if result:
                filled = []
                for field, col in target_map.items():
                    if col is None or field not in missing:
                        continue
                    if ws.cell(row_num, col).value:  # уже заполнено — не трогаем
                        continue
                    val = result.get(field)
                    if val is None:
                        continue
                    cell = ws.cell(row_num, col, val)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    filled.append(field)

                print(f"         ✓ заполнено: {', '.join(filled) if filled else '(пусто)'}")
                stats["done"] += 1
                since_save += 1
            else:
                print("         ✗ не удалось")
                stats["failed"] += 1

            # Периодическое сохранение
            if since_save >= SAVE_EVERY:
                safe_save(wb, XLSX_PATH)
                since_save = 0

            if row_num < row_end:
                time.sleep(args.delay)

        context.close()

    # Финальное сохранение
    safe_save(wb, XLSX_PATH)

    print()
    print("═" * 62)
    print("  Готово!")
    print(f"  Заполнено:  {stats['done']} строк")
    print(f"  Пропущено:  {stats['skipped']} (уже заполнены)")
    print(f"  Ошибок:     {stats['failed']}")
    print(f"  Файл:       {XLSX_PATH}")
    print("═" * 62)
    print()


if __name__ == "__main__":
    main()
