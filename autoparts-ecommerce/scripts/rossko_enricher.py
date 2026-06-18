"""
rossko_enricher.py — дополнение Excel данными с volzsky.rossko.ru
══════════════════════════════════════════════════════════════════

Читает Excel файл (например, mikado_data.xlsx), находит строки с пустыми ячейками
в колонках: Параметры, OEM номера, Применяемость, Альтернативные артикулы товара, Изображения.
Для каждой такой строки ищет товар на volzsky.rossko.ru и заполняет пустые ячейки.

Запуск:
  uv run --with requests,openpyxl scripts/rossko_enricher.py --file mikado_data.xlsx
  uv run --with requests,openpyxl scripts/rossko_enricher.py --file mikado_data.xlsx --rows 2-50
  uv run --with requests,openpyxl scripts/rossko_enricher.py --file data.xlsx --out result.xlsx

Аргументы:
  --file      Путь к Excel файлу для обогащения  [обязательный]
  --out       Путь для сохранения результата  [по умолчанию: file_rsk.xlsx рядом с исходным]
  --rows      Диапазон строк, например 2-50  [по умолчанию: все строки]
  --login     Логин на volzsky.rossko.ru  [или ROSSKO_LOGIN в .env]
  --password  Пароль  [или ROSSKO_PASSWORD в .env]
  --delay     Пауза между запросами в секундах  [по умолчанию: 1.5]
  --env       Путь к .env файлу  [по умолчанию: .env рядом со скриптом]
  --images    Папка для сохранения изображений  [по умолчанию: images/ рядом с файлом]
  --no-images Не скачивать изображения
"""

import re
import sys
import time
import argparse
import urllib.parse
from datetime import datetime
from hashlib import md5
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
    from openpyxl.styles import Alignment
except ImportError:
    print("Установи зависимости: uv run --with requests,openpyxl ...")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

SITE_BASE   = "https://volzsky.rossko.ru"
UTILS_URL   = f"{SITE_BASE}/utils/"
SEARCH_URL  = f"{UTILS_URL}?action=search"

PRODUCT_CARD_BASE = "https://productcard.rossko.ru"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)

# Колонки которые обогащаем (если пустые)
ENRICH_COLS = [
    "Параметры",
    "OEM номера",
    "Применяемость",
    "Альтернативные артикулы товара",
    "Изображения",
]


# ══════════════════════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ══════════════════════════════════════════════════════════════════════════════

def load_env(env_path: Path) -> dict:
    env = {}
    if env_path and env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


def normalize_code(mikado_code: str) -> str:
    """
    Нормализует артикул Mikado для поиска на Rossko.
    f-a22025   → a22025   (однобуквенный префикс)
    ab-1234    → 1234     (двухбуквенный префикс)
    xzk-sg-31  → xzk-sg-31 (длинный префикс — оставляем как есть)
    """
    code = str(mikado_code).strip().lower()
    m = re.match(r"^[a-z]{1,2}-(.+)$", code)
    if m:
        return m.group(1)
    return code


def _get_jhash(code: int) -> int:
    """
    Воспроизводит алгоритм get_jhash() из JS сайта rossko.ru
    (используется для обхода бот-защиты).
    """
    x = 123456789
    k = 0
    for i in range(1677696):
        x = ((x + code) ^ (x + (x % 3) + (x % 17) + code) ^ i) % 16776960
        if x % 117 == 0:
            k = (k + 1) % 1111
    return k


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 1 — АВТОРИЗАЦИЯ (с обходом JS-бот-защиты)
# ══════════════════════════════════════════════════════════════════════════════

def login(username: str, password: str) -> requests.Session:
    """
    Авторизуется на volzsky.rossko.ru.

    Сайт защищён JS-бот-фильтром. Алгоритм:
      1. GET / → сервер выставляет cookie __js_p_=CODE,1800,0,0,0
      2. Вычисляем jhash(CODE) и выставляем __jhash_ и __jua_ cookies
      3. GET / → 302 с настоящим __hash_ cookie (сессия открыта)
      4. POST /utils/ action=auth → PHPSESSID + auth cookie
    """
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    # ── Шаг 1: получаем __js_p_ cookie ────────────────────────────────────────
    session.get(SITE_BASE + "/", timeout=20)

    js_p = session.cookies.get("__js_p_", "")
    if js_p:
        code_str = js_p.split(",")[0]
        try:
            code = int(code_str)
        except ValueError:
            code = 0

        if code:
            # ── Шаг 2: вычисляем jhash ────────────────────────────────────────
            print("  Вычисляю jhash (займёт ~3 сек)...", end=" ", flush=True)
            jhash = _get_jhash(code)
            print(f"hash={jhash}")

            ua_enc = urllib.parse.quote(USER_AGENT)

            # ── Шаг 3: отправляем jhash, получаем __hash_ cookie ─────────────
            session.cookies.set("__jhash_", str(jhash))
            session.cookies.set("__jua_", ua_enc)
            session.get(SITE_BASE + "/", timeout=20, allow_redirects=True)

    # ── Шаг 4: авторизация ─────────────────────────────────────────────────────
    resp = session.post(
        UTILS_URL,
        data={
            "auth[email]":    username,
            "auth[password]": password,
            "action":         "auth",
            "type":           "header",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
        timeout=20,
    )
    resp.raise_for_status()

    try:
        data = resp.json()
        if data.get("err") is False:
            uid = data.get("user_id", "?")
            print(f"  ✓ Авторизация успешна (user_id={uid})")
        else:
            print(f"  ✗ Ошибка авторизации: {data}")
    except Exception:
        print(f"  ✗ Неожиданный ответ авторизации: {resp.text[:200]}")

    _bypass_productcard(session)
    return session


def _bypass_productcard(session: requests.Session):
    """
    Обходит JS-защиту productcard.rossko.ru в рамках той же сессии.
    productcard.rossko.ru — отдельный домен с собственным __js_p_ challenge.
    """
    try:
        session.get(PRODUCT_CARD_BASE + "/", timeout=20)
    except Exception:
        return

    # Ищем __js_p_ именно для домена productcard.rossko.ru
    js_p = ""
    for c in session.cookies:
        if c.name == "__js_p_" and "productcard" in (c.domain or ""):
            js_p = c.value
            break

    if not js_p:
        return

    code_str = js_p.split(",")[0]
    try:
        code = int(code_str)
    except ValueError:
        return

    if not code:
        return

    print("  JS защита productcard.rossko.ru, вычисляю jhash...", end=" ", flush=True)
    jhash = _get_jhash(code)
    print(f"hash={jhash}")

    ua_enc = urllib.parse.quote(USER_AGENT)
    session.cookies.set("__jhash_", str(jhash), domain="productcard.rossko.ru", path="/")
    session.cookies.set("__jua_",   ua_enc,      domain="productcard.rossko.ru", path="/")
    session.get(PRODUCT_CARD_BASE + "/", timeout=20, allow_redirects=True)


def _make_api_headers(session: requests.Session) -> dict:
    """Заголовки для запросов к productcard.rossko.ru."""
    auth_val = session.cookies.get("auth", "")
    return {
        "Authorization-Session": auth_val,
        "Authorization-Domain":  SITE_BASE,
        "source":                "frontend",
        "Accept":                "application/json",
        "Content-Type":          "application/json",
        "User-Agent":            USER_AGENT,
    }


def _api_cookies(session: requests.Session) -> dict:
    """Cookies для передачи в productcard.rossko.ru."""
    return {
        "PHPSESSID": session.cookies.get("PHPSESSID", ""),
        "auth":      session.cookies.get("auth", ""),
    }


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 2 — ПОИСК ТОВАРА
# ══════════════════════════════════════════════════════════════════════════════

def search_product(session: requests.Session, article: str) -> list[dict]:
    """
    Ищет товар по артикулу через автодополнение поиска Rossko.
    Возвращает: [{"guid": "...", "label": "...", "value": "...", "count": N}]

    GET /utils/?action=search&term=ARTICLE
    Ответ: [{"label":"A22025 Fenox Амортизатор","value":"a22025","guid":"NSII...","count":1}]
    """
    resp = session.get(
        UTILS_URL,
        params={"action": "search", "term": article},
        headers={"Referer": SITE_BASE + "/"},
        timeout=20,
    )
    resp.raise_for_status()

    try:
        return resp.json()
    except Exception:
        return []


def select_best_match(results: list[dict], article: str, brand_hint: str = "") -> dict | None:
    """
    Выбирает наиболее подходящий товар из результатов поиска.
    Приоритет:
      1. Метка начинается с точного артикула + бренд-подсказка
      2. Метка начинается с точного артикула (первый)
      3. Первый результат
    """
    if not results:
        return None

    art_up    = article.upper().replace("-", "").replace(" ", "")
    brand_up  = brand_hint.upper() if brand_hint else ""

    def label_art(r: dict) -> str:
        label = r.get("label", "")
        return label.split(" ")[0].upper().replace("-", "").replace(" ", "")

    # Совпадение артикула + бренд
    if brand_up:
        for r in results:
            if label_art(r) == art_up and brand_up in r.get("label", "").upper():
                return r

    # Совпадение только артикула
    for r in results:
        if label_art(r) == art_up:
            return r

    return results[0]


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 3 — ЗАГРУЗКА ДАННЫХ КАРТОЧКИ ТОВАРА
# ══════════════════════════════════════════════════════════════════════════════

def fetch_product_card(session: requests.Session, guid: str) -> dict:
    """
    GET https://productcard.rossko.ru/api/Product/Card/{guid}
    Возвращает основные данные: характеристики, OEM, изображения.
    """
    url = f"{PRODUCT_CARD_BASE}/api/Product/Card/{guid}"
    resp = session.get(
        url,
        params={"CurrencyCode": "643", "newCart": "true"},
        headers=_make_api_headers(session),
        timeout=20,
    )
    resp.raise_for_status()
    try:
        return resp.json()
    except Exception:
        snippet = resp.text[:150] if resp.text else "(пусто)"
        print(f"    ✗ productcard API: пустой/некорректный JSON (статус {resp.status_code}): {snippet}")
        return {}


def fetch_applicability(session: requests.Session, guid: str) -> list:
    """
    GET https://productcard.rossko.ru/api/Product/CarTypes/{guid}
    Возвращает список марок с моделями и датами выпуска.
    """
    url = f"{PRODUCT_CARD_BASE}/api/Product/CarTypes/{guid}"
    resp = session.get(
        url,
        headers=_make_api_headers(session),
        timeout=20,
    )
    resp.raise_for_status()
    try:
        return resp.json()
    except Exception:
        return []


def fetch_crosses(session: requests.Session, guid: str) -> list[dict]:
    """
    GET https://productcard.rossko.ru/api/Product/Crosses/{guid}
    Возвращает аналоги (cross-references): [{brandName, partNumber, ...}]
    """
    url = f"{PRODUCT_CARD_BASE}/api/Product/Crosses/{guid}"
    resp = session.get(
        url,
        params={"CurrencyCode": "643", "newCart": "true"},
        headers=_make_api_headers(session),
        timeout=20,
    )
    resp.raise_for_status()
    try:
        data = resp.json()
        return data.get("crosses", [])
    except Exception:
        return []


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 4 — ПАРСИНГ И ФОРМАТИРОВАНИЕ ДАННЫХ
# ══════════════════════════════════════════════════════════════════════════════

def parse_characteristics(card_data: dict) -> tuple[str, str]:
    """
    Из ответа /api/Product/Card извлекает:
      - Параметры (все характеристики, кроме OEM)
      - OEM номера

    Возвращает (params_text, oem_text).

    Данные:
      characteristics: [
        {"name": "Вид амортизатора", "value": "давление газа"},
        {"name": "OEM", "value": "Porsche 1K0513029FA\nVAG 1K0513029FS\n..."}
      ]
    """
    main_part = card_data.get("mainPart", {})
    chars = main_part.get("characteristics", [])

    params: dict = {}
    oem_lines: list[str] = []

    for c in chars:
        name  = str(c.get("name", "")).strip()
        value = str(c.get("value", "")).strip()
        if not name or not value:
            continue

        if name.upper() == "OEM":
            # value содержит строки вида "Brand OEM_CODE\nBrand OEM_CODE"
            for line in value.splitlines():
                line = line.strip()
                if line:
                    oem_lines.append(line)
        else:
            if name in params:
                params[name] = f"{params[name]} / {value}"
            else:
                params[name] = value

    params_text = "\n".join(f"{k}: {v}" for k, v in params.items()) if params else ""

    # Форматируем OEM: "Brand: CODE; Brand: CODE"
    oem_formatted: list[str] = []
    for line in oem_lines[:20]:  # максимум 20
        parts = line.split(None, 1)  # разбиваем по первому пробелу
        if len(parts) == 2:
            oem_formatted.append(f"{parts[0]}: {parts[1]}")
        else:
            oem_formatted.append(line)
    oem_text = "; ".join(oem_formatted)

    return params_text, oem_text


def parse_applicability(apl_data: list) -> str:
    """
    Преобразует ответ /api/Product/CarTypes в строку:
    "AUDI A3 (2003–2012); SEAT ALTEA (2004–2009); ..."

    Данные:
      [{"name": "AUDI", "carModels": [{"name": "A3", "beginDate": "2003-05-01T00:00:00Z", "endDate": "2012-08-01T00:00:00Z"}]}]
    """
    if not apl_data:
        return ""

    parts: list[str] = []
    for brand_entry in apl_data:
        brand = brand_entry.get("name", "")
        for model in brand_entry.get("carModels", []):
            model_name = model.get("name", "")
            begin = model.get("beginDate", "")
            end   = model.get("endDate", "")

            year_from = begin[:4] if begin else ""
            year_to   = end[:4]   if end   else ""

            entry = f"{brand} {model_name}"
            if year_from or year_to:
                entry += f" ({year_from}–{year_to})"
            parts.append(entry)

    return "; ".join(parts)


def format_crosses(crosses: list[dict]) -> str:
    """
    Форматирует кросс-артикулы в строку:
    "LYNXauto G121007LR; FENOX A22025; ..."
    """
    if not crosses:
        return ""
    codes = []
    for c in crosses[:20]:
        brand = c.get("brandName", "")
        part  = c.get("partNumber", "")
        if brand and part:
            codes.append(f"{brand} {part}")
        elif part:
            codes.append(part)
    return "; ".join(codes)


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 5 — СКАЧИВАНИЕ ИЗОБРАЖЕНИЙ
# ══════════════════════════════════════════════════════════════════════════════

def download_images(session: requests.Session, card_data: dict,
                    article: str, images_dir: Path) -> list[str]:
    """
    Скачивает изображения товара из URL'ов в карточке.

    card_data["mainPart"]["images"] содержит прямые URL:
    ["https://imgs.rossko.ru/2B/C9/GUID/1.jpeg", ...]

    imgs.rossko.ru требует куки авторизации (без них отдаёт HTML).

    Сохраняет файлы как {article}_rsk_1.jpg, _rsk_2.jpg, ...
    Дубли по MD5 пропускает.
    Возвращает список имён сохранённых файлов.
    """
    if not images_dir:
        return []

    main_part = card_data.get("mainPart", {})
    image_urls = main_part.get("images", [])
    if not image_urls:
        return []

    saved: list[str] = []
    seen:  set        = set()
    idx = 1

    # imgs.rossko.ru требует куки авторизации с основного сайта
    img_cookies = _api_cookies(session)
    img_headers = {
        "User-Agent": USER_AGENT,
        "Referer":    SITE_BASE + "/",
    }

    for url in image_urls:
        try:
            ri = requests.get(url, timeout=20,
                              headers=img_headers, cookies=img_cookies)
            ri.raise_for_status()
            content = ri.content

            if content[:2] == b"\xff\xd8":
                ext = ".jpg"
            elif content[:8] == b"\x89PNG\r\n\x1a\n":
                ext = ".png"
            else:
                continue

            if len(content) < 1024:
                continue

            digest = md5(content).hexdigest()
            if digest in seen:
                continue
            seen.add(digest)

            fname = f"{article}_rsk_{idx}{ext}"
            (images_dir / fname).write_bytes(content)
            saved.append(fname)
            idx += 1

        except Exception:
            continue

    return saved


# ══════════════════════════════════════════════════════════════════════════════
#  ШАГ 6 — ОБОГАЩЕНИЕ EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def enrich_excel(
    file_path: Path,
    out_path: Path,
    session: requests.Session,
    images_dir: Path | None,
    delay: float = 1.5,
    row_start: int = 2,
    row_end: int | None = None,
) -> tuple[int, int]:
    """
    Читает Excel, находит строки с пустыми целевыми ячейками,
    заполняет данными с Rossko.
    Возвращает (обработано, обогащено).
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_idx: dict[str, int] = {}
    for i, h in enumerate(header_row, 1):
        if h:
            col_idx[str(h).strip()] = i

    print(f"  Столбцов в файле: {len(col_idx)}")

    code_col  = col_idx.get("Код (Mikado)")
    brand_col = col_idx.get("Бренд")

    if not code_col:
        print("  ✗ Колонка 'Код (Mikado)' не найдена в файле")
        return 0, 0

    target_cols: dict[str, int | None] = {name: col_idx.get(name) for name in ENRICH_COLS}
    available = [name for name, ci in target_cols.items() if ci is not None]
    print(f"  Целевые колонки найдены: {available}")
    if not available:
        print("  ✗ Ни одна из целевых колонок не найдена")
        return 0, 0

    effective_end = row_end if row_end else ws.max_row
    processed = 0
    enriched  = 0

    for row_num in range(row_start, effective_end + 1):
        mikado_code = ws.cell(row_num, code_col).value
        if not mikado_code:
            continue
        mikado_code = str(mikado_code).strip()
        brand_hint  = str(ws.cell(row_num, brand_col).value or "").strip() if brand_col else ""

        # Определяем какие ячейки пустые
        empty: dict[str, int] = {}
        for col_name, ci in target_cols.items():
            if ci and not ws.cell(row_num, ci).value:
                empty[col_name] = ci

        if not empty:
            continue

        processed += 1
        print(f"\n  [{row_num}] {mikado_code}  (пусто: {', '.join(empty)})")

        search_code = normalize_code(mikado_code)

        try:
            # ── Поиск ─────────────────────────────────────────────────────────
            results = search_product(session, search_code)
            if not results:
                print(f"    ✗ Не найдено (поиск: «{search_code}»)")
                time.sleep(delay)
                continue

            best = select_best_match(results, search_code, brand_hint)
            if not best:
                print(f"    ✗ Нет подходящего совпадения")
                time.sleep(delay)
                continue

            guid  = best.get("guid", "")
            label = best.get("label", "")
            print(f"    → {label} (guid={guid})")

            # ── Карточка товара ────────────────────────────────────────────────
            time.sleep(0.3)
            card_data = fetch_product_card(session, guid)

            params_text, oem_text = parse_characteristics(card_data)

            row_enriched = False

            # Параметры
            if "Параметры" in empty and params_text:
                _write_cell(ws, row_num, empty["Параметры"], params_text)
                param_count = len([l for l in params_text.splitlines() if l])
                print(f"    ✓ Параметры: {param_count} значений")
                row_enriched = True

            # OEM номера
            if "OEM номера" in empty and oem_text:
                _write_cell(ws, row_num, empty["OEM номера"], oem_text)
                oem_count = len(oem_text.split(";"))
                print(f"    ✓ OEM: {oem_count} номеров")
                row_enriched = True

            # ── Применяемость ──────────────────────────────────────────────────
            if "Применяемость" in empty:
                time.sleep(0.3)
                apl_data = fetch_applicability(session, guid)
                compat_text = parse_applicability(apl_data)
                if compat_text:
                    _write_cell(ws, row_num, empty["Применяемость"], compat_text)
                    model_count = sum(
                        len(b.get("carModels", [])) for b in apl_data
                    )
                    print(f"    ✓ Применяемость: {model_count} моделей")
                    row_enriched = True

            # ── Альтернативные артикулы ────────────────────────────────────────
            if "Альтернативные артикулы товара" in empty:
                time.sleep(0.3)
                crosses = fetch_crosses(session, guid)
                crosses_text = format_crosses(crosses)
                if crosses_text:
                    _write_cell(ws, row_num, empty["Альтернативные артикулы товара"],
                                crosses_text)
                    print(f"    ✓ Аналоги: {len(crosses)} артикулов")
                    row_enriched = True

            # ── Изображения ────────────────────────────────────────────────────
            if "Изображения" in empty and images_dir:
                imgs = download_images(session, card_data, mikado_code, images_dir)
                if imgs:
                    _write_cell(ws, row_num, empty["Изображения"], ", ".join(imgs))
                    print(f"    ✓ Фото: {', '.join(imgs)}")
                    row_enriched = True
                else:
                    print(f"    ○ Фото: не найдено")

            if row_enriched:
                enriched += 1

        except Exception as e:
            print(f"    ✗ Ошибка: {e}")

        time.sleep(delay)

    wb.save(out_path)
    return processed, enriched


def _write_cell(ws, row: int, col: int, value: str):
    ws.cell(row, col).value = value
    ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Дополнение Excel данными с volzsky.rossko.ru",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--file",      required=True,           help="Путь к Excel файлу")
    ap.add_argument("--out",       default=None,            help="Путь для сохранения результата")
    ap.add_argument("--rows",      default=None,            help="Диапазон строк, например 2-50")
    ap.add_argument("--login",     default=None,            help="Логин на volzsky.rossko.ru")
    ap.add_argument("--password",  default=None,            help="Пароль")
    ap.add_argument("--delay",     default=1.5, type=float, help="Пауза между запросами (сек)")
    ap.add_argument("--env",       default=None,            help="Путь к .env файлу")
    ap.add_argument("--images",    default=None,            help="Папка для сохранения изображений")
    ap.add_argument("--no-images", action="store_true",     help="Не скачивать изображения")
    args = ap.parse_args()

    # ── Пути ──────────────────────────────────────────────────────────────────
    file_path = Path(args.file)
    if not file_path.exists():
        print(f"Файл не найден: {file_path}")
        sys.exit(1)

    if args.out:
        out_path = Path(args.out)
    else:
        out_path = file_path.parent / (file_path.stem + "_rsk" + file_path.suffix)

    if args.no_images:
        images_dir = None
    elif args.images:
        images_dir = Path(args.images)
        images_dir.mkdir(parents=True, exist_ok=True)
    else:
        images_dir = file_path.parent / "images"
        images_dir.mkdir(exist_ok=True)

    # ── Логин и пароль ────────────────────────────────────────────────────────
    env_path = Path(args.env) if args.env else Path(__file__).parent.parent / ".env"
    env      = load_env(env_path)
    username = args.login    or env.get("ROSSKO_LOGIN",    "")
    password = args.password or env.get("ROSSKO_PASSWORD", "")

    if not username or not password:
        print("Укажите логин и пароль: --login XXXXX --password XXXXX")
        print("  или добавьте ROSSKO_LOGIN и ROSSKO_PASSWORD в .env файл")
        sys.exit(1)

    # ── Диапазон строк ────────────────────────────────────────────────────────
    if args.rows:
        row_start, row_end = map(int, args.rows.split("-"))
    else:
        wb      = openpyxl.load_workbook(file_path, read_only=True)
        row_end = wb.active.max_row or 1000
        wb.close()
        row_start = 2

    # ── Шапка ─────────────────────────────────────────────────────────────────
    print()
    print("═" * 60)
    print("  Rossko Enricher")
    print(f"  Файл:      {file_path.name}  (строки {row_start}–{row_end})")
    print(f"  Результат: {out_path}")
    print("═" * 60)

    # ── Авторизация ───────────────────────────────────────────────────────────
    print("\n[1/3] Авторизация...")
    session = login(username, password)

    # ── Обогащение ───────────────────────────────────────────────────────────
    print(f"\n[2/3] Обогащаю данные...")
    print(f"  Изображения: {'отключены (--no-images)' if images_dir is None else str(images_dir)}")

    processed, enriched = enrich_excel(
        file_path=file_path,
        out_path=out_path,
        session=session,
        images_dir=images_dir,
        delay=args.delay,
        row_start=row_start,
        row_end=row_end,
    )

    # ── Итог ──────────────────────────────────────────────────────────────────
    print()
    print("═" * 60)
    print("  Готово!")
    print(f"  Строк с пустыми ячейками: {processed}")
    print(f"  Строк обогащено:          {enriched}")
    print(f"  Результат: {out_path}")
    if images_dir:
        print(f"  Изображения: {images_dir}")
    print("═" * 60)
    print()


if __name__ == "__main__":
    main()
