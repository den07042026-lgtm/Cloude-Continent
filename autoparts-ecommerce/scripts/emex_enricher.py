"""
emex_enricher.py
Заполняет Топ-500 ВБ.xlsx:
  col 7  OEM номера            — пропускаем (нет бесплатного источника)
  col 8  Применяемость         — улучшаем из названия товара
  col 9  Альтернативные артикулы — emex.ru makes.list (бренды с тем же артикулом)
"""
import os, sys, io, json, re, time, random
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

EXCEL_FILE  = r'C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ.xlsx'
CACHE_DIR   = r'C:\Users\Admin\Documents\Autoparts_Ecommerce\data\analytics\emex_cache'
BATCH_SAVE  = 50   # сохранять Excel каждые N строк

os.makedirs(CACHE_DIR, exist_ok=True)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,*/*;q=0.9',
    'Accept-Language': 'ru-RU,ru;q=0.9',
}
session = requests.Session()
session.headers.update(HEADERS)

THIN = Border(
    left  =Side(style='thin', color='CCCCCC'),
    right =Side(style='thin', color='CCCCCC'),
    top   =Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# ── Применяемость из названия ──────────────────────────────────────────────────
# Типы продуктов на которые начинаются названия (рус+англ)
PRODUCT_PREFIXES = [
    r'ФИЛЬТР\s+(?:ВОЗДУШНЫЙ|МАСЛЯНЫЙ|ТОПЛИВНЫЙ|САЛОННЫЙ)',
    r'ФИЛЬТР\s+ТОПЛИВА', r'ФИЛЬТР\s+МАСЛА', r'ФИЛЬТР\s+(?:УГ\.?\s*)?САЛОНА',
    r'Фильтр\s+(?:воздушный|масляный|топливный|салонный|угольный)',
    r'СВЕЧА\s+(?:ЗАЖИГАНИЯ|НАКАЛИВАНИЯ)',
    r'Свеча\s+зажигания',
    r'КАТУШКА\s+(?:ЗАЖИГАНИЯ)?',
    r'САЙЛЕНТБЛОК', r'ШАРОВАЯ\s+ОПОРА', r'ОПОРА\s+СТОЙКИ',
    r'АМОРТИЗАТОР', r'ПОДШИПНИК', r'СТУПИЧНЫЙ\s+ПОДШИПНИК',
    r'ТОРМОЗН[ЫЫЙ\s]+(?:ДИСК|БАРАБАН|КОЛОДК)',
    r'РЕМЕНЬ', r'РОЛИК', r'ПОМПА', r'ТЕРМОСТАТ',
    r'КРЫШКА\s+(?:ГОЛОВКИ|КЛАПАНА)', r'ПРОКЛАДКА',
    r'КЛЮЧ\s+СВЕЧНОЙ',
]

PRODUCT_PREFIX_RE = re.compile(
    r'^(?:' + '|'.join(PRODUCT_PREFIXES) + r')\s+',
    re.IGNORECASE
)

# Суффиксы для удаления в конце (несколько паттернов)
SUFFIX_PATTERNS = [
    # (BRAND) ARTICLE — с круглыми скобками
    re.compile(r'\s*\([A-Z0-9а-яА-Я\s_/.-]{2,30}\)\s*[A-Z0-9/._-]{3,30}\s*$', re.IGNORECASE),
    # BRAND /ARTICLE/ — с косыми скобками
    re.compile(r'\s+[A-Z]{2,10}\s+/[A-Z][A-Z0-9/._-]{3,25}/\s*$', re.IGNORECASE),
    # /ARTICLE/ — просто артикул в слэшах
    re.compile(r'\s*/[A-Z][A-Z0-9/._-]{3,25}/\s*$', re.IGNORECASE),
    # BRAND ARTICLE в конце (без скобок, если BRAND = одно слово из букв)
    re.compile(r'\s+[A-Z]{2,8}\s+[A-Z][A-Z0-9]{4,20}\s*$'),
]


def remove_suffixes(s: str) -> str:
    for pat in SUFFIX_PATTERNS:
        s = pat.sub('', s).strip()
    return s


def extract_applicability_from_name(name: str) -> str:
    """Извлекает применяемость из названия товара."""
    if not name:
        return ''
    s = name.strip()
    # Убираем префикс-тип товара
    s = PRODUCT_PREFIX_RE.sub('', s).strip()
    # Убираем суффиксы (несколько проходов)
    s = remove_suffixes(s)
    # Если слишком коротко или только цифры
    if len(s) < 4 or re.fullmatch(r'[\d\s]+', s):
        return ''
    # Если похоже на просто артикул без пробелов
    if re.fullmatch(r'[A-Z0-9/._-]+', s, re.IGNORECASE) and len(s) < 15:
        return ''
    return s


# ── Emex HTML парсинг ──────────────────────────────────────────────────────────
def load_cache(article: str, brand: str) -> dict | None:
    key = re.sub(r'[^a-zA-Z0-9]', '_', f'{brand}_{article}')[:50]
    path = os.path.join(CACHE_DIR, f'{key}.json')
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    return None


def save_cache(article: str, brand: str, data: dict):
    key = re.sub(r'[^a-zA-Z0-9]', '_', f'{brand}_{article}')[:50]
    path = os.path.join(CACHE_DIR, f'{key}.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)


def fetch_emex_data(article: str, brand: str) -> dict:
    """Получает данные с emex.ru: makes list, имя. Возвращает dict."""
    cached = load_cache(article, brand)
    if cached is not None:
        return cached

    url = f'https://emex.ru/products/{article}/{brand}'
    try:
        time.sleep(random.uniform(0.8, 1.5))
        r = session.get(url, timeout=20)
        if r.status_code != 200:
            result = {'makes': [], 'name': '', 'error': r.status_code}
            save_cache(article, brand, result)
            return result

        html = r.text
        # Разэкранируем
        decoded = html.replace('\\"', '"').replace('\\/', '/').replace('\\n', ' ').replace('\\t', ' ')

        # Находим начало Redux state
        idx = decoded.find('"applePay":{"isAvailable"')
        if idx < 0:
            result = {'makes': [], 'name': '', 'error': 'state not found'}
            save_cache(article, brand, result)
            return result

        # Идём назад до {
        start = idx
        for i in range(idx, max(0, idx-100), -1):
            if decoded[i] == '{':
                start = i
                break

        # Находим конец объекта
        depth = 0
        in_str = False
        pos = start
        while pos < len(decoded):
            c = decoded[pos]
            if c == '"' and (pos == 0 or decoded[pos-1] != '\\'):
                in_str = not in_str
            if not in_str:
                if c == '{': depth += 1
                elif c == '}':
                    depth -= 1
                    if depth == 0:
                        break
            pos += 1

        json_str = decoded[start:pos+1]
        state = json.loads(json_str)
        details = state.get('details', {})

        makes_list = details.get('makes', {}).get('list', [])
        makes = [
            {'make': m.get('make', ''), 'num': m.get('num', ''), 'name': m.get('name', '')}
            for m in makes_list
        ]

        result = {
            'makes': makes,
            'name':  details.get('name', ''),
            'num':   details.get('num', ''),
            'make':  details.get('make', ''),
        }
        save_cache(article, brand, result)
        return result

    except Exception as e:
        result = {'makes': [], 'name': '', 'error': str(e)[:100]}
        save_cache(article, brand, result)
        return result


# Бренды-автопроизводители (OEM). Сравниваем без учёта регистра/пробелов.
OEM_BRANDS = {
    'hyundai', 'kia', 'hyundai/kia', 'hyundai / kia',
    'honda', 'toyota', 'lexus',
    'volkswagen', 'vw', 'audi', 'skoda', 'seat', 'porsche',
    'bmw', 'mini',
    'mercedes', 'mercedes-benz', 'smart',
    'ford', 'lincoln',
    'opel', 'vauxhall', 'chevrolet', 'gm', 'general motors',
    'renault', 'dacia',
    'peugeot', 'citroen', 'ds',
    'nissan', 'infiniti',
    'mazda', 'mitsubishi', 'subaru', 'suzuki', 'daihatsu', 'isuzu',
    'volvo',
    'fiat', 'alfa romeo', 'lancia', 'jeep', 'chrysler', 'dodge', 'ram',
    'land rover', 'jaguar', 'range rover',
    'lada', 'avtovaz', 'avtovaz (lada)', 'gaz', 'uaz', 'zaz',
    'geely', 'chery', 'haval', 'jac', 'byd', 'great wall', 'changan', 'lifan',
    'ssangyong', 'daewoo',
    'tesla', 'volvo truck', 'man', 'daf', 'scania', 'iveco',
    'caterpillar', 'john deere', 'kubota',
}


def _is_oem_brand(brand_name: str) -> bool:
    s = brand_name.strip().lower()
    if s in OEM_BRANDS:
        return True
    # Составные: "Renault, Nissan" / "General Motors, ACDelco" / "Renault Trucks"
    for part in re.split(r'[,/]', s):
        part = part.strip()
        if part in OEM_BRANDS:
            return True
        # Первое слово составного ("renault trucks" → "renault")
        first = part.split()[0] if part else ''
        if first in OEM_BRANDS:
            return True
    return False


def format_oem_numbers(makes: list, own_brand: str, article: str = '') -> str:
    """Формирует строку OEM-номеров из брендов-автопроизводителей."""
    brand_norm = own_brand.strip().upper()
    article_up = article.strip().upper()
    # Если артикул начинается с букв бренда (LECAR000031501 → LECAR) —
    # записи с тем же номером — это совместимость, не OEM-номер.
    brand_short = re.sub(r'[^A-Z]', '', brand_norm)[:5]
    article_is_branded = (len(brand_short) >= 3
                          and article_up.startswith(brand_short))

    lines = []
    seen = set()
    for m in makes:
        mb = m.get('make', '').strip()
        mn = m.get('num', '').strip()
        if mb.upper() == brand_norm:
            continue
        if not (mb and mn and _is_oem_brand(mb)):
            continue
        # Пропускаем совпадения "тот же артикул" для брендованных номеров
        if article_is_branded and mn.upper() == article_up:
            continue
        key = f'{mb.upper()} {mn.upper()}'
        if key not in seen:
            seen.add(key)
            lines.append(f'{mb} {mn}')
    return ' / '.join(lines[:10])


def format_alternatives(makes: list, own_brand: str, article: str) -> str:
    """Форматирует список альтернативных артикулов (только aftermarket)."""
    brand_norm = own_brand.strip().upper()
    lines = []
    for m in makes:
        mb = m.get('make', '').strip()
        mn = m.get('num', '').strip()
        if mb.upper() == brand_norm:
            continue
        if _is_oem_brand(mb):
            continue  # OEM идут в отдельную колонку
        if mb and mn:
            lines.append(f'{mb} {mn}')
    return ' / '.join(lines[:15])  # WB ограничение ~15 значений


def set_cell(ws, row, col, value, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=wrap)
    cell.border    = THIN


# ── Основная логика ────────────────────────────────────────────────────────────
def main():
    print(f'Читаю: {EXCEL_FILE}')
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.worksheets[0]

    total = sum(1 for r in range(2, 502) if ws.cell(r, 1).value)
    print(f'Строк с данными: {total}')
    print()

    updated_applic = 0
    updated_oem    = 0
    updated_alts   = 0
    errors         = 0

    for row_idx in range(2, 502):
        article = ws.cell(row_idx, 1).value
        name    = ws.cell(row_idx, 2).value
        brand   = ws.cell(row_idx, 3).value

        if not article:
            continue

        data_num  = row_idx - 1
        article_s = str(article).strip()
        brand_s   = str(brand).strip() if brand else ''

        # ── Применяемость ──────────────────────────────────────────────────────
        current_applic = ws.cell(row_idx, 8).value
        name_s = str(name).strip() if name else ''

        extracted = extract_applicability_from_name(name_s)
        if extracted and (not current_applic or len(str(current_applic)) < len(extracted)):
            set_cell(ws, row_idx, 8, extracted, wrap=True)
            updated_applic += 1

        # ── OEM + Альтернативные артикулы ──────────────────────────────────────
        current_oem  = ws.cell(row_idx, 7).value
        current_alts = ws.cell(row_idx, 9).value

        need_fetch = brand_s and (not current_oem or not current_alts)
        if need_fetch:
            data  = fetch_emex_data(article_s, brand_s)
            makes = data.get('makes', [])
            if 'error' in data:
                errors += 1

            if makes:
                if not current_oem:
                    oem = format_oem_numbers(makes, brand_s, article_s)
                    if oem:
                        set_cell(ws, row_idx, 7, oem, wrap=True)
                        updated_oem += 1

                if not current_alts:
                    alts = format_alternatives(makes, brand_s, article_s)
                    if alts:
                        set_cell(ws, row_idx, 9, alts, wrap=False)
                        updated_alts += 1

            pct = data_num / total * 100
            if data_num % 10 == 0 or data_num <= 5:
                err_note = f' err={data.get("error","")[:30]}' if 'error' in data else ''
                makes_n = len(makes) if need_fetch else 0
                print(f'  #{data_num:3d}/{total}  {article_s[:20]:20s}  {brand_s[:12]:12s}  '
                      f'makes={makes_n:2d}{err_note}  [{pct:.0f}%]')
        else:
            if data_num % 50 == 0:
                print(f'  #{data_num:3d}/{total}  applic={updated_applic} oem={updated_oem} alts={updated_alts}')

        # Сохраняем батч
        if data_num % BATCH_SAVE == 0 and data_num > 0:
            wb.save(EXCEL_FILE)
            print(f'  >> Сохранено в {EXCEL_FILE}  (строка {data_num})')

    # Финальное сохранение
    wb.save(EXCEL_FILE)
    print()
    print(f'Готово!')
    print(f'  Применяемость улучшена: {updated_applic} строк')
    print(f'  OEM номера заполнены:   {updated_oem} строк')
    print(f'  Альтернативные артикулы заполнены: {updated_alts} строк')
    print(f'  Ошибок emex.ru: {errors}')


if __name__ == '__main__':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    main()
