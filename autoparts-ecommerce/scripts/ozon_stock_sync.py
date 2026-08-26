"""
ozon_stock_sync.py
══════════════════
Синхронизирует остатки с прайсом Микадо по расписанию.

Расписание: 00:01, 04:01, 08:01, 12:01, 16:01, 20:01 (каждые 4 часа).
При старте — немедленная синхронизация, затем ожидание следующего слота.

Логика одного цикла:
  1. Скачать актуальный прайс с mikado-parts.ru
  2. Распарсить Code + QTY → {артикул: кол-во}
  3. Обновить остатки в МойСклад через инвентаризацию
     (МойСклад автоматически синхронизирует с Ozon)
  4. Fallback: если MOYSKLAD_TOKEN не задан — обновить Ozon напрямую

Переменные .env:
    MOYSKLAD_TOKEN=...
    MOYSKLAD_STORE_NAME=...   # имя FBS-склада в МойСклад (если пусто — первый склад)
    TG_BOT_TOKEN=...
    TG_CHAT_ID=...
    OZON_CLIENT_ID=...        # только для fallback-режима
    OZON_API_KEY=...
    OZON_WAREHOUSE_ID=...

Запуск (непрерывный, по расписанию):
  uv run --with requests,openpyxl scripts/ozon_stock_sync.py

Разовый запуск:
  uv run --with requests,openpyxl scripts/ozon_stock_sync.py --once

Сухой прогон:
  uv run --with requests,openpyxl scripts/ozon_stock_sync.py --once --dry-run
"""

import sys
import io
import time
import logging
import argparse
import json
import re
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, str(Path(__file__).parent))
try:
    from telegram_notify import tg_stock_done, tg_alert
    _TG_OK = True
except ImportError:
    _TG_OK = False

sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
except ImportError:
    print("Установи зависимости: uv run --with requests,openpyxl scripts/ozon_stock_sync.py")
    sys.exit(1)

# ─── Константы ────────────────────────────────────────────────────────────────
BASE_DIR        = Path(__file__).parent.parent
ENV_FILE        = BASE_DIR / ".env"
LOG_FILE        = BASE_DIR / "logs" / "ozon_stock_sync.log"
PRICE_FALLBACK  = Path("C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx")

MIKADO_PRICE_URL = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=BBE2E029-54CF-4D9E-9FAC-9FE25E85B300"
)

OZON_API_BASE       = "https://api-seller.ozon.ru"
MS_BASE             = "https://api.moysklad.ru/api/remap/1.2"
MOYSKLAD_API_BASE   = MS_BASE  # обратная совместимость

SYNC_SLOTS      = [0, 4, 8, 12, 16, 20]  # часы запуска (всегда HH:01)
OZON_BATCH_SIZE = 100   # API принимает до 100 позиций за раз
MS_BATCH_SIZE   = 100   # МойСклад: позиций в одной инвентаризации
AUTOLIGA_DEFAULT_MAX_AGE_HOURS = 36
BRAND_ATTRIBUTE_ID = 85

# ─── Логирование ──────────────────────────────────────────────────────────────
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ─── .env ─────────────────────────────────────────────────────────────────────
def load_env() -> dict:
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    # Локальный файл секретов используется только как fallback и никогда не логируется.
    secret_file = Path(env.get("OZON_SECRETS_FILE", "C:/Users/Admin/Desktop/1.txt"))
    if secret_file.exists() and (not env.get("OZON_CLIENT_ID") or not env.get("OZON_API_KEY")):
        text = secret_file.read_text(encoding="utf-8")
        for key, label in (("OZON_CLIENT_ID", "Client-Id"), ("OZON_API_KEY", "Api-Key"), ("OZON_WAREHOUSE_ID", "Warehouse-Id")):
            match = re.search(rf"(?im)^\s*{re.escape(label)}\s*[:=]\s*([^\r\n]+)", text)
            if match and not env.get(key):
                env[key] = match.group(1).strip()
    return env


def _norm_article(value: str) -> str:
    """Нормализация артикула только для сопоставления, не для отправки в API."""
    return re.sub(r"[^0-9A-ZА-Я]", "", str(value or "").upper().replace("Ё", "Е"))


def _norm_brand(value: str) -> str:
    return re.sub(r"[^0-9A-ZА-Я]", "", str(value or "").upper().replace("Ё", "Е"))


def _article_from_offer(offer_id: str) -> str:
    value = str(offer_id or "").strip()
    return re.sub(r"-con$", "", value, flags=re.IGNORECASE)


# ─── Mikado: скачать прайс ────────────────────────────────────────────────────
def download_mikado_price(url: str) -> bytes | None:
    """
    Скачивает Excel-прайс «Региональный склад Волгоград» с Mikado.
    URL содержит Key — авторизация сессией не требуется.
    Возвращает bytes файла или None при ошибке.
    """
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        if resp.content[:2] == b"PK":   # magic bytes ZIP/xlsx
            log.info(f"Mikado: прайс скачан ({len(resp.content):,} байт)")
            return resp.content
        log.warning(
            f"Mikado: ответ не похож на Excel "
            f"(первые байты: {resp.content[:4].hex()})"
        )
        return None
    except Exception as e:
        log.error(f"Mikado: ошибка скачивания прайса: {e}")
        return None


# ─── Mikado: парсинг прайса ───────────────────────────────────────────────────
def parse_price(content: bytes | None, fallback: Path) -> dict[str, int]:
    """
    Читает Excel прайса (скачанный или локальный).
    Возвращает {артикул (Code): наличие (QTY)}.

    Колонки прайса: Prodnum | Code | BrandName | Prodname | PriceOut | QTY | ...
    """
    if content:
        source = "скачанный прайс"
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    elif fallback.exists():
        source = str(fallback)
        log.warning(f"Используем локальный прайс: {fallback}")
        wb = openpyxl.load_workbook(fallback, read_only=True, data_only=True)
    else:
        log.error("Нет прайса: ни скачанного, ни локального файла")
        return {}

    ws = wb.active
    all_rows = ws.iter_rows(values_only=True)
    header_row = next(all_rows, None)
    if not header_row:
        log.error("Прайс пустой")
        wb.close()
        return {}

    headers_lower = [str(v).strip().lower() if v else "" for v in header_row]

    code_idx = qty_idx = None
    for i, h in enumerate(headers_lower):
        if h == "code":
            code_idx = i
        elif h == "qty":
            qty_idx = i

    if code_idx is None or qty_idx is None:
        log.error(
            f"Прайс: не найдены нужные колонки. "
            f"Ожидалось 'Code' и 'QTY', найдено: {[str(v) for v in header_row]}"
        )
        wb.close()
        return {}

    log.info(f"Прайс ({source}): колонки Code[{code_idx}] QTY[{qty_idx}]")

    result: dict[str, int] = {}
    for row in all_rows:
        raw_code = row[code_idx] if len(row) > code_idx else None
        raw_qty  = row[qty_idx]  if len(row) > qty_idx  else None
        if not raw_code:
            continue
        article = str(raw_code).strip()
        try:
            qty = max(0, int(float(str(raw_qty)))) if raw_qty is not None else 0
        except (ValueError, TypeError):
            qty = 0
        result[article] = qty

    wb.close()
    log.info(f"Прайс: {len(result)} позиций, "
             f"в наличии: {sum(1 for q in result.values() if q > 0)}")
    return result


def parse_mikado_catalog(content: bytes | None, fallback: Path) -> dict[str, list[dict]]:
    """Возвращает индекс Mikado: normalized_article -> варианты с брендом и остатком."""
    if content:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    elif fallback.exists():
        wb = openpyxl.load_workbook(fallback, read_only=True, data_only=True)
    else:
        return {}

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None) or []
    headers = [str(v).strip().lower() if v else "" for v in header]
    try:
        code_idx = headers.index("code")
        qty_idx = headers.index("qty")
    except ValueError:
        wb.close()
        return {}
    brand_idx = next((i for i, h in enumerate(headers) if h in {"brandname", "brand"}), None)

    index: dict[str, list[dict]] = {}
    for row in rows:
        code = str(row[code_idx] or "").strip() if len(row) > code_idx else ""
        key = _norm_article(code)
        if not key:
            continue
        brand = str(row[brand_idx] or "").strip() if brand_idx is not None and len(row) > brand_idx else ""
        try:
            qty = max(0, int(float(row[qty_idx] or 0)))
        except (ValueError, TypeError):
            qty = 0
        index.setdefault(key, []).append({"article": code, "brand": brand, "qty": qty, "source": "mikado"})
    wb.close()
    return index


def load_autoliga_catalog(max_age_hours: float) -> tuple[dict[str, list[dict]], Path | None]:
    """Загружает свежий прайс Автолиги и индексирует и артикул, и заводской код."""
    try:
        from autoliga_loader import load_autoliga, find_autoliga_file
        file = find_autoliga_file()
        if file is None:
            log.error("Автолига: файл прайса не найден")
            return {}, None
        age_hours = (time.time() - file.stat().st_mtime) / 3600
        if age_hours > max_age_hours:
            log.error(
                f"Автолига: прайс устарел ({age_hours:.1f} ч; допустимо {max_age_hours:.1f} ч): {file}"
            )
            return {}, file
        raw = load_autoliga(file)
        index: dict[str, list[dict]] = {}
        seen: set[tuple[str, str, str]] = set()
        for item in raw.values():
            entry = {
                "article": str(item.get("article", "")).strip(),
                "brand": str(item.get("brand", "")).strip(),
                "qty": max(0, int(float(item.get("stock", 0) or 0))),
                "source": "autoliga",
            }
            for raw_key in (item.get("article"), item.get("oem")):
                key = _norm_article(raw_key)
                marker = (key, _norm_brand(entry["brand"]), entry["article"])
                if key and marker not in seen:
                    index.setdefault(key, []).append(entry)
                    seen.add(marker)
        log.info(f"Автолига: свежий прайс {file.name}, индекс {len(index):,} артикулов")
        return index, file
    except Exception as e:
        log.error(f"Автолига: ошибка загрузки: {e}")
        return {}, None


def _select_supplier_qty(candidates: list[dict], brand: str) -> tuple[int, list[str], str]:
    """Выбирает безопасный остаток: бренд обязателен при неоднозначном артикуле."""
    if not candidates:
        return 0, [], "not_found"
    brand_key = _norm_brand(brand)
    branded = [x for x in candidates if brand_key and _norm_brand(x.get("brand", "")) == brand_key]
    selected = branded
    if not selected:
        candidate_brands = {_norm_brand(x.get("brand", "")) for x in candidates if x.get("brand")}
        if len(candidate_brands) <= 1:
            selected = candidates
        else:
            return 0, [], "brand_conflict"
    # Если товар доступен у обоих поставщиков, складываем, но ограничиваем витринный
    # остаток четырьмя единицами, чтобы снизить риск одновременной продажи.
    by_source: dict[str, int] = {}
    for item in selected:
        source = item["source"]
        by_source[source] = max(by_source.get(source, 0), int(item.get("qty", 0)))
    total = min(4, sum(by_source.values()))
    return total, sorted(k for k, v in by_source.items() if v > 0), "matched"


def build_target_stocks(products: list[dict], mikado: dict, autoliga: dict) -> tuple[dict[str, int], dict]:
    """Строит полный набор остатков, включая обязательные нули."""
    targets: dict[str, int] = {}
    stats = {"mikado": 0, "autoliga": 0, "both": 0, "zero": 0, "brand_conflict": 0}
    for product in products:
        offer_id = str(product.get("offer_id", "")).strip()
        if not offer_id:
            continue
        key = _norm_article(_article_from_offer(offer_id))
        candidates = [*mikado.get(key, []), *autoliga.get(key, [])]
        qty, sources, reason = _select_supplier_qty(candidates, product.get("brand", ""))
        targets[offer_id] = qty
        if reason == "brand_conflict":
            stats["brand_conflict"] += 1
        if not sources:
            stats["zero"] += 1
        elif len(sources) == 2:
            stats["both"] += 1
        else:
            stats[sources[0]] += 1
    return targets, stats


# ─── МойСклад: список артикулов ───────────────────────────────────────────────
def get_moysklad_articles(token: str) -> list[str]:
    """
    Возвращает список артикулов товаров из МойСклад.
    Если токен не задан — возвращает [] (фильтрация по МойСклад пропускается).
    """
    if not token:
        log.info("MOYSKLAD_TOKEN не задан — синхронизируем весь прайс без фильтрации")
        return []

    articles: list[str] = []
    offset = 0
    limit  = 1000
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept-Encoding": "gzip",
    }

    while True:
        try:
            resp = requests.get(
                f"{MOYSKLAD_API_BASE}/entity/product",
                headers=headers,
                params={"limit": limit, "offset": offset},
                timeout=30,
            )
            resp.raise_for_status()
            rows = resp.json().get("rows", [])
            for row in rows:
                art = row.get("article")
                if art:
                    articles.append(str(art).strip())
            if len(rows) < limit:
                break
            offset += limit
        except Exception as e:
            log.error(f"МойСклад: ошибка получения товаров: {e}")
            break

    log.info(f"МойСклад: получено {len(articles)} артикулов")
    return articles


# ─── МойСклад: обновить остатки через оприходование/списание ─────────────────
def _ms_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def _ms_get(token: str, path: str, params: dict | None = None) -> dict:
    r = requests.get(f"{MS_BASE}{path}", headers=_ms_headers(token),
                     params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def _get_ms_org(token: str) -> str | None:
    try:
        rows = _ms_get(token, "/entity/organization", {"limit": 1}).get("rows", [])
        return rows[0]["meta"]["href"] if rows else None
    except Exception as e:
        log.error(f"МойСклад: ошибка получения организации: {e}")
        return None


def _get_ms_store(token: str, store_name: str = "") -> str | None:
    try:
        rows = _ms_get(token, "/entity/store", {"limit": 100}).get("rows", [])
        if not rows:
            return None
        if store_name:
            match = next((r for r in rows if r.get("name", "") == store_name), None)
            if match:
                return match["meta"]["href"]
            log.warning(f"МойСклад: склад '{store_name}' не найден, берём первый")
        return rows[0]["meta"]["href"]
    except Exception as e:
        log.error(f"МойСклад: ошибка получения склада: {e}")
        return None


def _build_ms_product_map(token: str) -> dict[str, str]:
    """Возвращает {article: product_meta_href} для всех товаров МойСклад."""
    result: dict[str, str] = {}
    offset = 0
    while True:
        try:
            data = _ms_get(token, "/entity/product", {"limit": 1000, "offset": offset})
            rows = data.get("rows", [])
            for row in rows:
                article = (row.get("article") or row.get("code") or "").strip()
                if article:
                    result[article] = row["meta"]["href"]
            if len(rows) < 1000:
                break
            offset += 1000
        except Exception as e:
            log.error(f"МойСклад: ошибка загрузки товаров (offset={offset}): {e}")
            break
    log.info(f"МойСклад: загружено {len(result)} товаров")
    return result


def _get_ms_current_stock(token: str, store_href: str) -> dict[str, int]:
    """Возвращает {article: qty} текущих остатков по складу."""
    result: dict[str, int] = {}
    offset = 0
    while True:
        try:
            r = requests.get(
                f"{MS_BASE}/report/stock/all",
                headers=_ms_headers(token),
                params={"limit": 1000, "offset": offset,
                        "filter": f"store={store_href}"},
                timeout=30,
            )
            r.raise_for_status()
            data = r.json()
            rows = data.get("rows", [])
            for row in rows:
                art = (row.get("article") or "").strip()
                if art:
                    result[art] = int(row.get("quantity", 0))
            if len(rows) < 1000:
                break
            offset += 1000
        except Exception as e:
            log.error(f"МойСклад: ошибка получения остатков: {e}")
            break
    log.info(f"МойСклад: текущие остатки — {len(result)} позиций")
    return result


def _ms_post_doc(token: str, endpoint: str, org_href: str, store_href: str,
                 positions: list[dict]) -> int:
    """Создаёт проведённый документ (оприходование или списание) батчами."""
    updated = 0
    for start in range(0, len(positions), MS_BATCH_SIZE):
        batch = positions[start: start + MS_BATCH_SIZE]
        try:
            r = requests.post(
                f"{MS_BASE}/entity/{endpoint}",
                headers=_ms_headers(token),
                json={
                    "applicable":   True,
                    "organization": {"meta": {"href": org_href, "type": "organization",
                                               "mediaType": "application/json"}},
                    "store":        {"meta": {"href": store_href, "type": "store",
                                               "mediaType": "application/json"}},
                    "positions": batch,
                },
                timeout=60,
            )
            r.raise_for_status()
            updated += len(batch)
        except Exception as e:
            log.error(f"МойСклад: ошибка {endpoint} batch [{start}:{start+MS_BATCH_SIZE}]: {e}")
    return updated


def update_moysklad_stocks(
    token:      str,
    stocks:     dict[str, int],
    store_name: str = "",
    dry_run:    bool = False,
) -> int:
    """
    Синхронизирует остатки в МойСклад через дельта-документы:
      +дельта → Оприходование (/entity/enter)
      -дельта → Списание (/entity/loss)
    МойСклад автоматически передаёт остатки в Ozon (автосинк каждые 15 мин).
    Возвращает количество обновлённых позиций.
    """
    if not token:
        log.warning("MOYSKLAD_TOKEN не задан — обновление МойСклад пропущено")
        return 0

    org_href   = _get_ms_org(token)
    store_href = _get_ms_store(token, store_name)
    if not org_href or not store_href:
        log.error("МойСклад: не удалось получить организацию/склад")
        return 0

    product_map  = _build_ms_product_map(token)
    current_stock = _get_ms_current_stock(token, store_href)

    to_enter:    list[dict] = []
    to_writeoff: list[dict] = []

    for code, target_qty in stocks.items():
        href = product_map.get(code) or product_map.get(f"{code}-con")
        if not href:
            continue
        art = code if code in product_map else f"{code}-con"
        current_qty = current_stock.get(art, current_stock.get(code, 0))
        delta = target_qty - current_qty
        if delta == 0:
            continue
        pos = {"assortment": {"meta": {"href": href, "type": "product",
                                        "mediaType": "application/json"}},
               "quantity": abs(delta)}
        if delta > 0:
            pos["price"] = 0
            to_enter.append(pos)
        else:
            to_writeoff.append(pos)

    log.info(f"МойСклад: дельта — оприходовать={len(to_enter)}, списать={len(to_writeoff)}")

    if dry_run:
        log.info("[DRY-RUN] без записи в МойСклад")
        return len(to_enter) + len(to_writeoff)

    updated = 0
    if to_enter:
        n = _ms_post_doc(token, "enter", org_href, store_href, to_enter)
        updated += n
        log.info(f"МойСклад: оприходовано {n} позиций")
    if to_writeoff:
        n = _ms_post_doc(token, "loss", org_href, store_href, to_writeoff)
        updated += n
        log.info(f"МойСклад: списано {n} позиций")
    if not to_enter and not to_writeoff:
        log.info("МойСклад: остатки актуальны, изменений нет")

    return updated


# ─── Ozon: каталог и остатки ─────────────────────────────────────────────────
def _ozon_headers(client_id: str, api_key: str) -> dict:
    return {"Client-Id": client_id, "Api-Key": api_key, "Content-Type": "application/json"}


def _ozon_post(client_id: str, api_key: str, path: str, payload: dict) -> dict:
    response = requests.post(
        f"{OZON_API_BASE}{path}",
        headers=_ozon_headers(client_id, api_key),
        json=payload,
        timeout=60,
    )
    response.raise_for_status()
    return response.json()


def get_ozon_products(client_id: str, api_key: str) -> list[dict]:
    """Получает полный каталог и бренд каждого SKU; частичный ответ считается ошибкой."""
    offers: list[str] = []
    last_id = ""
    for _ in range(100):
        payload = {"filter": {"visibility": "ALL"}, "limit": 1000}
        if last_id:
            payload["last_id"] = last_id
        page = _ozon_post(client_id, api_key, "/v3/product/list", payload)
        items = page.get("result", {}).get("items", [])
        offers.extend(str(x.get("offer_id", "")).strip() for x in items if x.get("offer_id"))
        next_id = page.get("result", {}).get("last_id", "")
        if len(items) < 1000:
            break
        if not next_id or next_id == last_id:
            raise RuntimeError("Ozon: каталог оборвался на пагинации")
        last_id = next_id
    if not offers:
        raise RuntimeError("Ozon: получен пустой каталог — обновление запрещено")

    products: list[dict] = []
    for start in range(0, len(offers), 100):
        part = offers[start:start + 100]
        page = _ozon_post(
            client_id, api_key, "/v4/product/info/attributes",
            {"filter": {"offer_id": part}, "limit": 1000, "sort_dir": "ASC"},
        )
        rows = page.get("result", [])
        by_offer = {str(x.get("offer_id", "")): x for x in rows}
        for offer_id in part:
            row = by_offer.get(offer_id, {})
            brand = ""
            for attr in row.get("attributes", []):
                if int(attr.get("id", 0) or 0) == BRAND_ATTRIBUTE_ID:
                    values = attr.get("values", [])
                    brand = str(values[0].get("value", "")).strip() if values else ""
                    break
            products.append({"offer_id": offer_id, "brand": brand})
    if len(products) != len(offers):
        raise RuntimeError(f"Ozon: неполный каталог {len(products)}/{len(offers)}")
    log.info(f"Ozon: получено {len(products):,} SKU, с брендом {sum(bool(x['brand']) for x in products):,}")
    return products


def update_ozon_stocks(
    client_id: str,
    api_key: str,
    warehouse_id: int,
    stocks: dict[str, int],
    dry_run: bool = False,
) -> tuple[int, int]:
    """
    Обновляет FBS-остатки на Ozon через POST /v2/products/stocks.
    stocks: {offer_id: qty}  — offer_id уже с суффиксом -con.
    warehouse_id обязателен.
    """
    if not client_id or not api_key:
        log.warning("OZON_CLIENT_ID / OZON_API_KEY не заданы — обновление Ozon пропущено")
        return 0, len(stocks)
    if not warehouse_id:
        log.warning("OZON_WAREHOUSE_ID не задан — обновление Ozon пропущено")
        return 0, len(stocks)

    if dry_run:
        log.info(f"[DRY-RUN] Ozon: обновилось бы {len(stocks)} позиций")
        return len(stocks), 0

    headers = {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
    }

    items     = list(stocks.items())
    total_ok  = 0
    total_err = 0

    for start in range(0, len(items), OZON_BATCH_SIZE):
        batch = items[start : start + OZON_BATCH_SIZE]
        payload_stocks = [
            {"offer_id": offer_id, "warehouse_id": warehouse_id, "stock": qty}
            for offer_id, qty in batch
        ]
        try:
            resp = requests.post(
                f"{OZON_API_BASE}/v2/products/stocks",
                headers=headers,
                json={"stocks": payload_stocks},
                timeout=30,
            )
            resp.raise_for_status()
            results = resp.json().get("result", [])
            errors  = [r for r in results if r.get("errors")]
            total_ok  += len(payload_stocks) - len(errors)
            total_err += len(errors)
            for err in errors[:3]:
                log.warning(f"  Ozon [{err.get('offer_id')}]: {err.get('errors')}")
        except Exception as e:
            log.error(f"Ozon: ошибка отправки батча [{start}:{start+OZON_BATCH_SIZE}]: {e}")
            total_err += len(batch)

    log.info(f"Ozon: обновлено {total_ok}/{len(items)}, ошибок: {total_err}")
    return total_ok, total_err


# ─── Один цикл синхронизации ──────────────────────────────────────────────────
def sync_once(env: dict, dry_run: bool = False) -> None:
    log.info("─" * 55)
    log.info(f"Синхронизация {'[DRY-RUN] ' if dry_run else ''}Микадо + Автолига → Ozon")

    # 1. Скачиваем прайс (Key-авторизация в URL, сессия не нужна)
    price_url     = env.get("MIKADO_PRICE_URL", MIKADO_PRICE_URL)
    price_content = download_mikado_price(price_url)

    # 2. Загружаем оба источника. При сбое одного ничего не записываем: иначе
    # корректные товары поставщика были бы массово обнулены.
    mikado = parse_mikado_catalog(price_content, PRICE_FALLBACK)
    try:
        max_age = float(env.get("AUTOLIGA_MAX_AGE_HOURS", AUTOLIGA_DEFAULT_MAX_AGE_HOURS))
    except ValueError:
        max_age = AUTOLIGA_DEFAULT_MAX_AGE_HOURS
    autoliga, autoliga_file = load_autoliga_catalog(max_age)
    if not mikado or not autoliga:
        log.error(
            "ЗАЩИТА: один из обязательных прайсов пуст или устарел; "
            "Ozon и МойСклад не изменены"
        )
        return

    # 3. Получаем полный список SKU Ozon. Обновляем именно весь каталог, чтобы
    # отсутствующие у обоих поставщиков позиции получили честный ноль.
    client_id = env.get("OZON_CLIENT_ID", "")
    api_key = env.get("OZON_API_KEY", "")
    if not client_id or not api_key:
        log.error("OZON_CLIENT_ID / OZON_API_KEY не заданы — цикл остановлен")
        return
    try:
        products = get_ozon_products(client_id, api_key)
    except Exception as e:
        log.error(f"ЗАЩИТА: не удалось получить полный каталог Ozon: {e}")
        return

    stocks, match_stats = build_target_stocks(products, mikado, autoliga)
    if len(stocks) != len(products):
        log.error(f"ЗАЩИТА: неполный расчёт остатков {len(stocks)}/{len(products)}")
        return
    in_stock = sum(qty > 0 for qty in stocks.values())
    log.info(
        f"Сопоставление: Mikado={match_stats['mikado']}, "
        f"Автолига={match_stats['autoliga']}, оба={match_stats['both']}, "
        f"ноль={match_stats['zero']}, конфликт бренда={match_stats['brand_conflict']}"
    )

    # 4. Сохраняем аудит до отправки. В нём нет токенов и закупочных цен.
    report_dir = BASE_DIR / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    report_path = report_dir / f"ozon_stock_plan_{datetime.now():%Y-%m-%d_%H-%M-%S}.json"
    report_path.write_text(json.dumps({
        "created_at": datetime.now().isoformat(),
        "dry_run": dry_run,
        "autoliga_file": str(autoliga_file) if autoliga_file else None,
        "products": len(products),
        "positive": in_stock,
        "zero": len(stocks) - in_stock,
        "matching": match_stats,
        "stocks": stocks,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info(f"Аудит плана: {report_path}")

    # 5. Прямая запись в Ozon — единый источник истины для кабинета.
    try:
        warehouse_id = int(env.get("OZON_WAREHOUSE_ID", "0"))
    except ValueError:
        warehouse_id = 0
    updated, errors = update_ozon_stocks(
        client_id=client_id,
        api_key=api_key,
        warehouse_id=warehouse_id,
        stocks=stocks,
        dry_run=dry_run,
    )

    # 6. Telegram — итог синхронизации
    if _TG_OK and env.get("TG_BOT_TOKEN") and not dry_run:
        tg_stock_done(
            env["TG_BOT_TOKEN"], env.get("TG_CHAT_ID", ""),
            total=len(stocks), in_stock=in_stock, ms_updated=updated,
        )
    log.info(f"Цикл завершён: обновлено {updated}, ошибок {errors}")


# ─── Расписание ───────────────────────────────────────────────────────────────
def _seconds_until_next_slot() -> float:
    """Возвращает секунды до ближайшего слота из SYNC_SLOTS (всегда HH:01)."""
    now = datetime.now()
    cur_min = now.hour * 60 + now.minute

    for h in SYNC_SLOTS:
        slot_min = h * 60 + 1
        if slot_min > cur_min:
            return (slot_min - cur_min) * 60 - now.second

    # Следующий слот — завтра в 00:01
    return (24 * 60 + 1 - cur_min) * 60 - now.second


# ─── Точка входа ──────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Синхронизация остатков Микадо → Озон"
    )
    parser.add_argument("--once",    action="store_true", help="Запустить один раз и выйти")
    parser.add_argument("--dry-run", action="store_true", help="Без отправки в МойСклад/Ozon")
    args = parser.parse_args()

    env = load_env()

    if args.once or args.dry_run:
        sync_once(env, dry_run=args.dry_run)
        return

    slots_str = ", ".join(f"{h:02d}:01" for h in SYNC_SLOTS)
    log.info(f"Планировщик запущен: синхронизация в {slots_str}")

    # Немедленный запуск при старте
    try:
        sync_once(env)
    except Exception:
        log.exception("Ошибка при первоначальной синхронизации")

    while True:
        wait = _seconds_until_next_slot()
        next_dt = datetime.now() + timedelta(seconds=wait)
        log.info(f"Следующий запуск: {next_dt.strftime('%d.%m %H:%M')}")
        time.sleep(wait)
        try:
            sync_once(env)
        except Exception:
            log.exception("Необработанная ошибка в цикле синхронизации")


if __name__ == "__main__":
    main()
