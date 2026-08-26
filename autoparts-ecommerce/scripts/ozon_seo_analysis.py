"""
ozon_seo_analysis.py
Полный SEO-аудит магазина на Ozon через Seller API.

Что делает:
  1. Загружает все товары  (v3/products/list)
  2. Детали каждого        (v2/product/info/list): фото, описание, статус, fbs_sku
  3. SEO-баллы             (v1/analytics/search/score/article)
  4. Аналитика 30 дней     (v1/analytics/data): показы, корзины, сессии
  5. Атрибуты              (v3/products/info/attributes): кол-во заполненных
  6. Excel-отчёт           data/ozon_seo_ДАТА.xlsx  (3 листа)
  7. Топ-проблемы в консоль

Запуск:
  cd C:\\Users\\Admin\\Documents\\Autoparts_Ecommerce
  uv run --with requests,openpyxl,python-dotenv scripts/ozon_seo_analysis.py
  uv run --with requests,openpyxl,python-dotenv scripts/ozon_seo_analysis.py --limit 500
  uv run --with requests,openpyxl,python-dotenv scripts/ozon_seo_analysis.py --days 60
"""

import sys
import os
import time
import logging
import argparse
from pathlib import Path
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
except ImportError:
    sys.exit("Установи: pip install requests")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError:
    sys.exit("Установи: pip install openpyxl")
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
except ImportError:
    pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-7s %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ── Конфигурация ─────────────────────────────────────────────────────────────
OZON_CLIENT_ID = os.getenv("OZON_CLIENT_ID", "")
OZON_API_KEY   = os.getenv("OZON_API_KEY", "")
BASE_URL       = "https://api-seller.ozon.ru"

HEADERS = {
    "Client-Id":    OZON_CLIENT_ID,
    "Api-Key":      OZON_API_KEY,
    "Content-Type": "application/json",
    "Accept":       "application/json",
}

# ── Цвета Excel ───────────────────────────────────────────────────────────────
CLR_HEADER  = "1F4E79"   # тёмно-синий
CLR_WARN    = "FFF2CC"   # жёлтый
CLR_CRIT    = "FFE0E0"   # красный
CLR_OK      = "E2EFDA"   # зелёный
CLR_SUBHDR  = "BDD7EE"   # голубой

# ═════════════════════════════ API-обёртки ═══════════════════════════════════

def _post(path: str, body: dict, retry: int = 3) -> dict:
    url = BASE_URL + path
    for attempt in range(retry):
        try:
            r = requests.post(url, json=body, headers=HEADERS, timeout=45)
            if r.status_code == 429:
                wait = int(r.headers.get("Retry-After", 15))
                log.warning("Rate-limit %s → ждём %ds", path, wait)
                time.sleep(wait)
                continue
            if r.status_code >= 500:
                log.warning("HTTP %d %s (попытка %d)", r.status_code, path, attempt + 1)
                time.sleep(3 ** attempt)
                continue
            r.raise_for_status()
            return r.json()
        except requests.RequestException as e:
            if attempt == retry - 1:
                log.error("Ошибка %s: %s", path, e)
                return {}
            time.sleep(2 ** attempt)
    return {}


def fetch_products(limit_total: int = 0) -> list[dict]:
    """Все товары: [{product_id, offer_id}]"""
    out, last_id = [], ""
    while True:
        batch_size = min(1000, limit_total - len(out)) if limit_total else 1000
        resp = _post("/v3/product/list", {
            "filter": {"visibility": "ALL"},
            "last_id": last_id,
            "limit": batch_size,
        })
        items = resp.get("result", {}).get("items", [])
        if not items:
            break
        out.extend(items)
        last_id = resp.get("result", {}).get("last_id", "")
        if not last_id or len(items) < batch_size or (limit_total and len(out) >= limit_total):
            break
        time.sleep(0.3)
    log.info("Товаров всего: %d", len(out))
    return out


def fetch_info(product_ids: list[int]) -> dict[int, dict]:
    """{product_id: info} — название, фото, описание, sku"""
    result: dict[int, dict] = {}
    # v3/product/info/list принимает до 1000, возвращает {"items": [...]}
    for i in range(0, len(product_ids), 1000):
        batch = product_ids[i:i+1000]
        resp = _post("/v3/product/info/list", {"product_id": batch})
        for item in resp.get("items", []):
            result[item.get("id", 0)] = item
        time.sleep(0.25)
    log.info("Детали получены: %d товаров", len(result))
    return result


def fetch_seo_scores() -> dict[str, dict]:
    """SEO-баллы {offer_id: {score, score_items}}"""
    result, offset = {}, 0
    while True:
        resp = _post("/v1/analytics/search/score/article", {
            "limit": 1000,
            "offset": offset,
        })
        r = resp.get("result", {})
        items = r.get("items") or []
        if not items:
            if not result:
                log.warning("SEO-баллы: пустой ответ (endpoint может быть недоступен на вашем тарифе)")
            break
        for item in items:
            oid = item.get("offer_id", "")
            if oid:
                result[oid] = item
        total = r.get("total", 0)
        offset += len(items)
        if offset >= total or len(items) < 1000:
            break
        time.sleep(0.3)
    log.info("SEO-баллы: %d товаров", len(result))
    return result


def fetch_analytics(date_from: str, date_to: str) -> dict[str, dict]:
    """Аналитика {sku_str: {hits_view, hits_view_search, hits_tocart, ...}}"""
    metrics = [
        "hits_view",          # все показы товара
        "hits_view_search",   # показы в поиске
        "hits_view_pdp",      # просмотры карточки
        "hits_tocart",        # добавлено в корзину
        "session_view_search",# уникальные сессии в поиске
    ]
    result, offset = {}, 0
    while True:
        resp = _post("/v1/analytics/data", {
            "date_from": date_from,
            "date_to":   date_to,
            "dimension": [{"name": "sku"}],
            "filters":   [],
            "limit":     1000,
            "offset":    offset,
            "metrics":   metrics,
        })
        data = resp.get("result", {}).get("data", [])
        if not data:
            break
        for row in data:
            dims = row.get("dimensions", [{}])
            sku_id = dims[0].get("id", "") if dims else ""
            vals = row.get("metrics", [0] * len(metrics))
            result[sku_id] = dict(zip(metrics, vals))
        offset += len(data)
        if len(data) < 1000:
            break
        time.sleep(0.3)
    log.info("Аналитика: %d SKU за %s–%s", len(result), date_from, date_to)
    return result


def fetch_attributes(product_ids: list[int]) -> dict[int, int]:
    """{product_id: кол-во_заполненных_атрибутов}"""
    result: dict[int, int] = {}
    str_ids = [str(p) for p in product_ids]
    last_id = ""
    while True:
        resp = _post("/v3/products/info/attributes", {
            "filter": {"product_id": str_ids, "visibility": "ALL"},
            "limit":    1000,
            "last_id":  last_id,
            "sort_dir": "ASC",
        })
        items = resp.get("result", [])
        if not items:
            break
        for item in items:
            pid  = item.get("id", 0)
            attrs = item.get("attributes", [])
            filled = sum(1 for a in attrs if a.get("values"))
            result[pid] = filled
        last_id = resp.get("last_id", "")
        if not last_id or len(items) < 1000:
            break
        time.sleep(0.3)
    log.info("Атрибуты: %d товаров", len(result))
    return result

# ═════════════════════════════ Анализ ════════════════════════════════════════

def score_grade(score) -> str:
    if score is None:
        return "—"
    s = float(score)
    if s >= 80:
        return "✅ Отлично"
    if s >= 60:
        return "⚡ Средне"
    return "❌ Плохо"


def build_rows(
    products:   list[dict],
    info_map:   dict[int, dict],
    seo_map:    dict[str, dict],
    analytics:  dict[str, dict],
    attr_map:   dict[int, int],
) -> list[dict]:
    rows = []
    for p in products:
        pid      = p.get("product_id", 0)
        offer_id = p.get("offer_id", "")
        info     = info_map.get(pid, {})

        name        = info.get("name", "")
        images      = info.get("images", []) or []
        img_count   = len(images)
        has_desc    = bool(info.get("description_category_id"))
        desc_text   = info.get("description", "")  # может быть пустым
        status      = info.get("status", {})
        state       = status.get("state_name", info.get("state", "")) if isinstance(status, dict) else ""
        fbs_sku     = str(info.get("fbs_sku", ""))
        fbo_sku     = str(info.get("fbo_sku", ""))
        cat_id      = info.get("description_category_id", "")
        type_id     = info.get("type_id", "")
        price       = info.get("price", "") or info.get("marketing_price", "")

        seo = seo_map.get(offer_id, {})
        seo_score = seo.get("score")
        seo_items: list[dict] = seo.get("score_items", [])
        seo_by_key = {it.get("key"): it for it in seo_items}

        def si(key: str):
            item = seo_by_key.get(key, {})
            v  = item.get("value")
            mx = item.get("max_value")
            if v is None:
                return None
            return round(v / mx * 100) if mx else 0

        # Аналитика — ищем по обоим SKU
        an = analytics.get(fbs_sku) or analytics.get(fbo_sku) or {}
        hits_total   = an.get("hits_view", 0)
        hits_search  = an.get("hits_view_search", 0)
        hits_pdp     = an.get("hits_view_pdp", 0)
        hits_cart    = an.get("hits_tocart", 0)
        sessions     = an.get("session_view_search", 0)
        ctr_search   = round(hits_cart / hits_search * 100, 2) if hits_search else 0

        filled_attrs = attr_map.get(pid, 0)

        row = {
            "Артикул":             offer_id,
            "Название":            name,
            "Категория ID":        cat_id,
            "Тип ID":              type_id,
            "Цена":                price,
            "Статус":              state,
            "Фото (кол-во)":       img_count,
            "Описание (есть)":     "Да" if desc_text else "Нет",
            "Атрибутов заполнено": filled_attrs,
            "SEO-балл (%)":        seo_score,
            "SEO-оценка":          score_grade(seo_score),
            # SEO детали
            "SEO: Название (%)":   si("title"),
            "SEO: Описание (%)":   si("description"),
            "SEO: Фото (%)":       si("image"),
            "SEO: Атрибуты (%)":   si("attributes"),
            "SEO: Ключ.фразы (%)": si("search_keywords"),
            "SEO: Rich-контент (%)": si("rich_content"),
            # Аналитика 30 дней
            "Показы (всего)":      int(hits_total),
            "Показы (поиск)":      int(hits_search),
            "Просмотры карточки":  int(hits_pdp),
            "В корзину":           int(hits_cart),
            "Сессии (поиск)":      int(sessions),
            "CTR в корзину (%)":   ctr_search,
        }
        rows.append(row)

    # Сортировка: сначала с SEO-баллом, потом без
    rows.sort(key=lambda r: (
        r["SEO-балл (%)"] is None,
        -(r["SEO-балл (%)"] or 0),
    ))
    return rows


# ═════════════════════════════ Excel-отчёт ═══════════════════════════════════

def _hdr_cell(ws, row, col, text, bg=CLR_HEADER, fg="FFFFFF", bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, color=fg, name="Calibri", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def write_audit_sheet(ws, rows: list[dict]):
    cols = list(rows[0].keys()) if rows else []
    widths = {
        "Артикул": 22,
        "Название": 40,
        "Категория ID": 14,
        "Тип ID": 10,
        "Цена": 10,
        "Статус": 14,
        "Фото (кол-во)": 12,
        "Описание (есть)": 14,
        "Атрибутов заполнено": 18,
        "SEO-балл (%)": 14,
        "SEO-оценка": 14,
        "SEO: Название (%)": 16,
        "SEO: Описание (%)": 16,
        "SEO: Фото (%)": 12,
        "SEO: Атрибуты (%)": 16,
        "SEO: Ключ.фразы (%)": 18,
        "SEO: Rich-контент (%)": 18,
        "Показы (всего)": 14,
        "Показы (поиск)": 14,
        "Просмотры карточки": 18,
        "В корзину": 12,
        "Сессии (поиск)": 14,
        "CTR в корзину (%)": 16,
    }

    # Заголовки
    for ci, col in enumerate(cols, 1):
        _hdr_cell(ws, 1, ci, col)
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 14)

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    red   = PatternFill("solid", fgColor=CLR_CRIT)
    warn  = PatternFill("solid", fgColor=CLR_WARN)
    green = PatternFill("solid", fgColor=CLR_OK)
    bdr   = _thin_border()

    for ri, row in enumerate(rows, 2):
        seo  = row.get("SEO-балл (%)")
        imgs = row.get("Фото (кол-во)", 0)
        desc = row.get("Описание (есть)", "Нет")
        cart = row.get("В корзину", 0)

        for ci, col in enumerate(cols, 1):
            val = row[col]
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = bdr
            c.font = Font(name="Calibri", size=9)
            c.alignment = Alignment(vertical="center", wrap_text=(col == "Название"))

        # Подсветка строки по SEO-баллу
        seo_col = cols.index("SEO-балл (%)") + 1
        seo_cell = ws.cell(row=ri, column=seo_col)
        if seo is None:
            seo_cell.fill = warn
        elif seo < 50:
            seo_cell.fill = red
            seo_cell.font = Font(name="Calibri", size=9, bold=True, color="C00000")
        elif seo < 75:
            seo_cell.fill = warn
        else:
            seo_cell.fill = green

        # Фото
        img_col = cols.index("Фото (кол-во)") + 1
        img_cell = ws.cell(row=ri, column=img_col)
        if imgs == 0:
            img_cell.fill = red
        elif imgs < 3:
            img_cell.fill = warn

        # Описание
        desc_col = cols.index("Описание (есть)") + 1
        desc_cell = ws.cell(row=ri, column=desc_col)
        if desc == "Нет":
            desc_cell.fill = red

    # Автофильтр
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def write_problems_sheet(ws, rows: list[dict]):
    _hdr_cell(ws, 1, 1, "Категория проблемы", bg=CLR_HEADER)
    _hdr_cell(ws, 1, 2, "Артикул",           bg=CLR_HEADER)
    _hdr_cell(ws, 1, 3, "Название",           bg=CLR_HEADER)
    _hdr_cell(ws, 1, 4, "SEO-балл (%)",       bg=CLR_HEADER)
    _hdr_cell(ws, 1, 5, "Деталь / Значение",  bg=CLR_HEADER)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 40
    ws.freeze_panes = "A2"

    ri = 2
    red  = PatternFill("solid", fgColor=CLR_CRIT)
    warn = PatternFill("solid", fgColor=CLR_WARN)
    bdr  = _thin_border()

    def add(category, row, detail, critical=False):
        nonlocal ri
        fill = red if critical else warn
        for ci, val in enumerate([
            category,
            row["Артикул"],
            row["Название"],
            row["SEO-балл (%)"],
            detail,
        ], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.border = bdr
            c.font = Font(name="Calibri", size=9)
        ri += 1

    for row in rows:
        seo  = row.get("SEO-балл (%)")
        imgs = row.get("Фото (кол-во)", 0)
        desc = row.get("Описание (есть)", "Нет")
        cart = row.get("В корзину", 0)
        shows = row.get("Показы (поиск)", 0)

        if seo is not None and seo < 50:
            add("🔴 Критически низкий SEO", row, f"Балл {seo}% — нужна срочная доработка", critical=True)
        elif seo is not None and seo < 70:
            add("🟡 Низкий SEO-балл", row, f"Балл {seo}%", critical=False)

        if imgs == 0:
            add("🔴 Нет фото", row, "0 фотографий — товар не отображается", critical=True)
        elif imgs < 3:
            add("🟡 Мало фото", row, f"Только {imgs} фото (рекомендуется ≥5)", critical=False)

        if desc == "Нет":
            add("🟡 Нет описания", row, "Описание отсутствует — влияет на поиск", critical=False)

        kw_score = row.get("SEO: Ключ.фразы (%)")
        if kw_score is not None and kw_score < 50:
            add("🟡 Мало ключевых фраз", row, f"SEO-балл ключевых фраз: {kw_score}%", critical=False)

        attrs = row.get("Атрибутов заполнено", 0)
        if attrs < 5:
            add("🟡 Мало атрибутов", row, f"Заполнено только {attrs} атрибут(ов)", critical=False)

        if shows > 100 and cart == 0:
            add("🟡 Показы без корзины", row, f"{shows} показов в поиске, 0 в корзину", critical=False)


def write_summary_sheet(ws, rows: list[dict], date_from: str, date_to: str):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    def add_row(label, value, ri, bg=None):
        a = ws.cell(row=ri, column=1, value=label)
        b = ws.cell(row=ri, column=2, value=value)
        a.font = b.font = Font(name="Calibri", size=10)
        if bg:
            a.fill = b.fill = PatternFill("solid", fgColor=bg)
        a.alignment = Alignment(horizontal="left", vertical="center")
        b.alignment = Alignment(horizontal="center", vertical="center")

    _hdr_cell(ws, 1, 1, "Показатель",       bg=CLR_HEADER)
    _hdr_cell(ws, 1, 2, "Значение",          bg=CLR_HEADER)
    _hdr_cell(ws, 3, 1, f"Период аналитики: {date_from} – {date_to}", bg=CLR_SUBHDR, fg="000000", bold=False)
    ws.merge_cells("A3:B3")

    total = len(rows)
    with_seo = [r for r in rows if r.get("SEO-балл (%)") is not None]
    no_seo   = total - len(with_seo)
    avg_seo  = round(sum(r["SEO-балл (%)"] for r in with_seo) / len(with_seo), 1) if with_seo else 0
    crit_seo = sum(1 for r in with_seo if r["SEO-балл (%)"] < 50)
    low_seo  = sum(1 for r in with_seo if 50 <= r["SEO-балл (%)"] < 70)
    ok_seo   = sum(1 for r in with_seo if r["SEO-балл (%)"] >= 70)
    no_photo = sum(1 for r in rows if r["Фото (кол-во)"] == 0)
    few_photo = sum(1 for r in rows if 1 <= r["Фото (кол-во)"] < 3)
    no_desc  = sum(1 for r in rows if r["Описание (есть)"] == "Нет")
    total_shows   = sum(r["Показы (поиск)"] for r in rows)
    total_cart    = sum(r["В корзину"] for r in rows)
    avg_ctr       = round(total_cart / total_shows * 100, 2) if total_shows else 0
    zero_shows    = sum(1 for r in rows if r["Показы (поиск)"] == 0)
    shows_no_cart = sum(1 for r in rows if r["Показы (поиск)"] > 50 and r["В корзину"] == 0)

    data = [
        ("── Общее ──────────────────────────────", "", CLR_SUBHDR),
        ("Всего товаров",                  total,       None),
        ("С SEO-баллом",                   len(with_seo), None),
        ("Без SEO-балла",                  no_seo,      CLR_WARN if no_seo else None),
        ("── SEO-баллы ──────────────────────────", "", CLR_SUBHDR),
        ("Средний SEO-балл (%)",            avg_seo,     CLR_OK if avg_seo >= 70 else CLR_WARN),
        (f"Отлично (≥80%)",                ok_seo,      CLR_OK),
        (f"Средне (50–79%)",               low_seo,     CLR_WARN),
        (f"🔴 Критично (<50%)",             crit_seo,    CLR_CRIT),
        ("── Контент ────────────────────────────", "", CLR_SUBHDR),
        ("Нет фотографий (0 шт.)",         no_photo,    CLR_CRIT if no_photo else None),
        ("Мало фото (1-2 шт.)",            few_photo,   CLR_WARN if few_photo else None),
        ("Нет описания",                   no_desc,     CLR_WARN if no_desc else None),
        ("── Аналитика (поиск) ──────────────────", "", CLR_SUBHDR),
        ("Суммарно показов в поиске",      total_shows, None),
        ("Суммарно добавлено в корзину",   total_cart,  None),
        ("Средний CTR в корзину (%)",       avg_ctr,     CLR_OK if avg_ctr >= 2 else CLR_WARN),
        ("Товаров без единого показа",      zero_shows,  CLR_WARN if zero_shows else None),
        ("Товаров: >50 показов, 0 корзин", shows_no_cart, CLR_WARN if shows_no_cart else None),
    ]

    ri = 4
    for label, value, bg in data:
        add_row(label, value, ri, bg)
        ws.row_dimensions[ri].height = 18
        ri += 1


def save_excel(rows: list[dict], date_from: str, date_to: str, out_path: Path):
    wb = openpyxl.Workbook()

    ws_audit = wb.active
    ws_audit.title = "SEO-аудит"
    write_audit_sheet(ws_audit, rows)

    ws_prob = wb.create_sheet("Проблемы")
    write_problems_sheet(ws_prob, rows)

    ws_sum = wb.create_sheet("Сводка")
    write_summary_sheet(ws_sum, rows, date_from, date_to)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("Excel сохранён: %s", out_path)


# ═════════════════════════════ Консоль ═══════════════════════════════════════

def print_summary(rows: list[dict], date_from: str, date_to: str):
    total = len(rows)
    with_seo = [r for r in rows if r.get("SEO-балл (%)") is not None]
    avg_seo  = round(sum(r["SEO-балл (%)"] for r in with_seo) / len(with_seo), 1) if with_seo else 0
    crit_seo = [r for r in with_seo if r["SEO-балл (%)"] < 50]
    no_photo = [r for r in rows if r["Фото (кол-во)"] == 0]
    no_desc  = [r for r in rows if r["Описание (есть)"] == "Нет"]
    total_shows = sum(r["Показы (поиск)"] for r in rows)
    total_cart  = sum(r["В корзину"] for r in rows)
    ctr = round(total_cart / total_shows * 100, 2) if total_shows else 0
    zero_shows = sum(1 for r in rows if r["Показы (поиск)"] == 0)

    print("\n" + "═" * 60)
    print("   OZON SEO-АУДИТ — СВОДКА")
    print("═" * 60)
    print(f"  Товаров:             {total}")
    print(f"  С SEO-баллом:        {len(with_seo)} из {total}")
    print(f"  Средний SEO-балл:    {avg_seo}%")
    print(f"  Критично (<50%):     {len(crit_seo)}")
    print(f"  Без фото:            {len(no_photo)}")
    print(f"  Без описания:        {len(no_desc)}")
    print(f"\n  Аналитика {date_from} – {date_to}:")
    print(f"  Показов в поиске:    {total_shows:,}")
    print(f"  Добавлено в корзину: {total_cart:,}")
    print(f"  Средний CTR:         {ctr}%")
    print(f"  Нет ни одного показа: {zero_shows}")
    print("═" * 60)

    if crit_seo:
        print(f"\n  🔴 ТОП-10 с критично низким SEO (<50%):")
        for r in crit_seo[:10]:
            print(f"    [{r['SEO-балл (%)']:5.1f}%] {r['Артикул'][:30]:<30} {r['Название'][:40]}")

    if no_photo:
        print(f"\n  📷 Без фото ({len(no_photo)} шт.) — первые 5:")
        for r in no_photo[:5]:
            print(f"    {r['Артикул']}")

    print()


# ═════════════════════════════ Main ══════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Ozon SEO-аудит через API")
    parser.add_argument("--limit", type=int, default=0,
                        help="Кол-во товаров (0 = все)")
    parser.add_argument("--days",  type=int, default=30,
                        help="Период аналитики в днях (default=30)")
    parser.add_argument("--output", type=str, default="",
                        help="Путь к Excel-отчёту")
    args = parser.parse_args()

    if not OZON_CLIENT_ID or not OZON_API_KEY:
        sys.exit("Укажи OZON_CLIENT_ID и OZON_API_KEY в .env")

    date_to   = datetime.now().strftime("%Y-%m-%d")
    date_from = (datetime.now() - timedelta(days=args.days)).strftime("%Y-%m-%d")

    out_name  = args.output or f"data/ozon_seo_{datetime.now().strftime('%Y%m%d')}.xlsx"
    out_path  = Path(__file__).parent.parent / out_name

    log.info("── Шаг 1/5: Загрузка списка товаров")
    products = fetch_products(args.limit)
    if not products:
        sys.exit("Не удалось получить товары. Проверь API ключи.")

    product_ids = [p["product_id"] for p in products]

    log.info("── Шаг 2/5: Детали товаров")
    info_map = fetch_info(product_ids)

    log.info("── Шаг 3/5: SEO-баллы")
    seo_map = fetch_seo_scores()

    log.info("── Шаг 4/5: Аналитика %s – %s", date_from, date_to)
    analytics = fetch_analytics(date_from, date_to)

    log.info("── Шаг 5/5: Атрибуты товаров")
    attr_map = fetch_attributes(product_ids)

    log.info("── Формирование отчёта")
    rows = build_rows(products, info_map, seo_map, analytics, attr_map)

    print_summary(rows, date_from, date_to)
    save_excel(rows, date_from, date_to, out_path)

    print(f"  Готово! Открой файл:\n  {out_path}\n")


if __name__ == "__main__":
    main()
