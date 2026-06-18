"""
deepseek_description_generator.py
══════════════════════════════════
Генерирует описания товаров для Ozon через веб-интерфейс chat.deepseek.com.
Автоматизирует браузер через Playwright — без API ключа.

Установка браузера (один раз):
  uv run --with playwright python -m playwright install chromium

Первый запуск — войти в аккаунт DeepSeek:
  uv run --with playwright scripts/deepseek_description_generator.py --login

Обычный запуск (все артикулы):
  uv run --with playwright scripts/deepseek_description_generator.py

Обработать конкретные строки (по порядку в папке):
  uv run --with playwright scripts/deepseek_description_generator.py --rows 1-5

Если селекторы перестали работать — сделай скриншот для отладки:
  uv run --with playwright scripts/deepseek_description_generator.py --debug
"""

import re
import sys
import json
import time
import argparse
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Playwright не установлен.")
    print("Шаг 1: uv run --with playwright python -m playwright install chromium")
    print("Шаг 2: uv run --with playwright scripts/deepseek_description_generator.py --login")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

BASE_DIR       = Path(__file__).parent.parent
SCRAPER_OUTPUT = BASE_DIR / "data/suppliers/mikado/scraper_output"
DESC_DIR       = SCRAPER_OUTPUT / "descriptions"
XLSX_PATH      = SCRAPER_OUTPUT / "mikado_data.xlsx"
PROFILE_DIR    = Path.home() / ".deepseek_playwright"   # сохранённая сессия

DEEPSEEK_URL   = "https://chat.deepseek.com"

# Селекторы DeepSeek (актуальны на апрель 2025)
# Если скрипт сломается после обновления сайта — запусти --debug и проверь
SEL_INPUT      = 'textarea'                          # поле ввода сообщения
SEL_SEND_BTN   = '[aria-label="send message"]'       # кнопка отправки
SEL_STOP_BTN   = '[aria-label="stop"]'               # кнопка остановки генерации
SEL_RESPONSE   = 'div.ds-markdown'                  # блок с ответом ассистента
SEL_RESPONSE_FB= '[class*="markdown"]'              # fallback если основной не работает


# ══════════════════════════════════════════════════════════════════════════════
#  ФОРМИРОВАНИЕ ПРОМПТА
# ══════════════════════════════════════════════════════════════════════════════

def build_prompt(data: dict) -> str:
    """
    Строит промпт для DeepSeek из данных артикула.
    Оптимизирован для товаров автозапчастей (амортизаторы, подвеска).
    """
    name   = data.get("name", "")
    brand  = data.get("brand", "")
    code   = data.get("code", "").upper().replace("F-", "")
    params = data.get("params", {})
    compat = data.get("compatibility", "")

    # Параметры строкой
    params_lines = "\n".join(f"  {k}: {v}" for k, v in params.items()) if params else "  (нет данных)"

    # Применяемость — ограничиваем длину, оставляем самые читаемые модели
    if compat:
        models = [m.strip() for m in compat.split(";") if m.strip()]
        # Берём до 10 моделей, остальные считаем
        if len(models) > 10:
            shown  = "; ".join(models[:10])
            others = len(models) - 10
            compat_short = f"{shown} и ещё {others} моделей"
        else:
            compat_short = "; ".join(models)
    else:
        compat_short = "уточняется по артикулу"

    prompt = f"""Ты копирайтер интернет-магазина автозапчастей. Напиши описание товара для маркетплейса Ozon.

ТРЕБОВАНИЯ:
- Длина: от 600 до 900 символов (считай строго)
- Без HTML, без заголовков, без списков с тире или цифрами
- Сплошной текст из 3 абзацев
- Только русский язык
- Тон: спокойный, информативный, располагающий — без агрессивных призывов и восклицательных знаков
- Артикул {brand} {code} обязательно упомяни в тексте

ДАННЫЕ О ТОВАРЕ:
Бренд: {brand}
Артикул: {code}
Наименование: {name}

Технические характеристики:
{params_lines}

Применяется на автомобилях:
  {compat_short}

СТРУКТУРА (строго три абзаца):
1. Что это за деталь, её роль в подвеске, тип и конструкция. Естественно упомяни артикул {brand} {code}.
2. Для каких автомобилей подходит — назови 2–3 марки из списка применяемости.
3. Коротко о бренде {brand} и надёжности детали. Мягкий, ненавязчивый призыв.

Напиши только текст описания, без комментариев и пояснений."""

    return prompt


# ══════════════════════════════════════════════════════════════════════════════
#  ВЗАИМОДЕЙСТВИЕ С БРАУЗЕРОМ
# ══════════════════════════════════════════════════════════════════════════════

def find_input(page) -> object:
    """
    Ищет поле ввода на странице DeepSeek.
    Пробует несколько селекторов по порядку.
    """
    candidates = [
        'textarea',
        'div[contenteditable="true"]',
        '#chat-input',
        '[data-testid="chat-input"]',
        '[placeholder*="Message"]',
        '[placeholder*="сообщение"]',
    ]
    for sel in candidates:
        try:
            el = page.locator(sel).last
            if el.is_visible(timeout=1500):
                return el
        except Exception:
            continue
    return None


def type_prompt(page, text: str):
    """
    Вводит текст промпта в поле ввода.
    Использует JavaScript для надёжной вставки многострочного текста.
    """
    inp = find_input(page)
    if inp is None:
        raise RuntimeError("Поле ввода не найдено — попробуй --debug чтобы увидеть страницу")

    inp.click()
    time.sleep(0.3)

    # Очищаем поле
    page.keyboard.press("Control+a")
    page.keyboard.press("Delete")
    time.sleep(0.2)

    # Вставляем через буфер обмена — надёжнее для длинного текста с переносами
    page.evaluate("""(text) => {
        const el = document.querySelector('textarea') ||
                   document.querySelector('div[contenteditable="true"]');
        if (!el) return;
        if (el.tagName === 'TEXTAREA') {
            const nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                window.HTMLTextAreaElement.prototype, 'value').set;
            nativeInputValueSetter.call(el, text);
            el.dispatchEvent(new Event('input', { bubbles: true }));
        } else {
            el.focus();
            document.execCommand('selectAll');
            document.execCommand('insertText', false, text);
        }
    }""", text)

    time.sleep(0.5)


def click_send(page):
    """Отправляет сообщение — сначала кнопкой, потом Enter."""
    # Пробуем кнопку отправки
    btn_candidates = [
        SEL_SEND_BTN,
        'button[type="submit"]',
        '[aria-label*="Send"]',
        '[aria-label*="send"]',
        '[aria-label*="Отправить"]',
        'button:has(svg):last-of-type',
    ]
    for sel in btn_candidates:
        try:
            btn = page.locator(sel).last
            if btn.is_visible(timeout=800):
                btn.click()
                return
        except Exception:
            continue

    # Fallback: просто Enter
    page.keyboard.press("Enter")


def get_last_response_text(page) -> str:
    """
    Извлекает текст последнего ответа ассистента.
    Пробует несколько селекторов.
    """
    candidates = [
        SEL_RESPONSE,
        SEL_RESPONSE_FB,
        '[class*="assistant"] [class*="markdown"]',
        '[class*="assistant"] [class*="content"]',
        '[data-role="assistant"] [class*="markdown"]',
        '[class*="chat-message"]:last-child [class*="content"]',
    ]
    for sel in candidates:
        try:
            els = page.locator(sel).all()
            if els:
                # Берём последний элемент (последний ответ)
                text = els[-1].inner_text(timeout=2000).strip()
                if len(text) > 50:
                    return text
        except Exception:
            continue
    return ""


def is_generating(page) -> bool:
    """Проверяет идёт ли генерация ответа (есть ли кнопка Stop)."""
    candidates = [
        SEL_STOP_BTN,
        '[aria-label*="stop"]',
        '[aria-label*="Stop"]',
        '[aria-label*="остановить"]',
        'button:has([class*="stop"])',
    ]
    for sel in candidates:
        try:
            if page.locator(sel).first.is_visible(timeout=500):
                return True
        except Exception:
            continue
    return False


def wait_for_response(page, timeout_sec: int = 120) -> str:
    """
    Ждёт пока DeepSeek сгенерирует ответ.
    Стратегия: polling — текст стабилен 3 проверки по 2 сек подряд.
    """
    print("    Жду ответ", end="", flush=True)

    # Сначала ждём появления любого текста (генерация началась)
    started = False
    for _ in range(15):  # до 30 сек на начало
        text = get_last_response_text(page)
        if text and len(text) > 30:
            started = True
            break
        time.sleep(2)
        print(".", end="", flush=True)

    if not started:
        print(" (текст не появился)")
        return ""

    # Ждём стабилизации — текст перестал меняться
    prev_text  = ""
    stable_cnt = 0

    for tick in range(timeout_sec // 2):
        current = get_last_response_text(page)

        if current and current == prev_text:
            stable_cnt += 1
            if stable_cnt >= 3:
                print(f" готово ({len(current)} симв.)")
                return current
        else:
            stable_cnt = 0
            if tick % 3 == 0:
                print(".", end="", flush=True)

        prev_text = current
        time.sleep(2)

    print(f" таймаут ({len(prev_text)} симв.)")
    return prev_text


# ══════════════════════════════════════════════════════════════════════════════
#  ОБРАБОТКА ОДНОГО АРТИКУЛА
# ══════════════════════════════════════════════════════════════════════════════

def process_article(page, json_path: Path, out_dir: Path, debug: bool = False) -> dict | None:
    """Генерирует описание для одного артикула."""

    data = json.load(open(json_path, encoding="utf-8"))
    code = data.get("code", json_path.stem)
    name = data.get("name", "")[:35]

    out_file = out_dir / f"{code}_desc.txt"
    if out_file.exists():
        print(f"  [{code}] пропуск — описание уже есть")
        return {"code": code, "description": out_file.read_text(encoding="utf-8"), "skipped": True}

    print(f"  [{code}]  {data.get('brand','')}  {name}")

    # Открываем новый чат (навигация на главную = чистый чат)
    try:
        page.goto(DEEPSEEK_URL, wait_until="domcontentloaded", timeout=30000)
        time.sleep(2)
    except PWTimeout:
        print("    ✗ Таймаут загрузки страницы")
        return None

    if debug:
        page.screenshot(path=str(out_dir / f"{code}_debug_before.png"))
        print(f"    Скриншот: {code}_debug_before.png")

    prompt = build_prompt(data)

    try:
        type_prompt(page, prompt)

        if debug:
            page.screenshot(path=str(out_dir / f"{code}_debug_typed.png"))
            print(f"    Скриншот: {code}_debug_typed.png  (промпт введён)")
            input("    Нажми Enter для отправки... ")

        click_send(page)
        time.sleep(1)

        response = wait_for_response(page)

        if not response:
            if debug:
                page.screenshot(path=str(out_dir / f"{code}_debug_noresponse.png"))
            print("    ✗ Пустой ответ — попробуй --debug чтобы увидеть страницу")
            return None

        out_file.write_text(response, encoding="utf-8")
        print(f"    ✓ {out_file.name}")
        return {"code": code, "description": response, "skipped": False}

    except RuntimeError as e:
        print(f"    ✗ {e}")
        if debug:
            page.screenshot(path=str(out_dir / f"{code}_debug_error.png"))
        return None
    except Exception as e:
        print(f"    ✗ Неожиданная ошибка: {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  СОХРАНЕНИЕ В EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def update_excel(descriptions: list[dict], xlsx_path: Path):
    """Добавляет/обновляет колонку «Описание Ozon» в mikado_data.xlsx."""
    if not HAS_OPENPYXL:
        print("  openpyxl не установлен — Excel не обновлён")
        return
    if not xlsx_path.exists():
        print(f"  {xlsx_path.name} не найден — Excel не обновлён")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Находим или создаём колонку
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    if "Описание Ozon" not in headers:
        desc_col = ws.max_column + 1
        hcell = ws.cell(1, desc_col, "Описание Ozon")
        hcell.font      = Font(bold=True, color="FFFFFF", size=10)
        hcell.fill      = PatternFill("solid", fgColor="1A1A2E")
        hcell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[hcell.column_letter].width = 65
    else:
        desc_col = headers["Описание Ozon"]

    # Строим индекс: код → строка Excel
    code_col   = headers.get("Код (Mikado)", 1)
    code_to_row = {
        str(ws.cell(r, code_col).value).strip(): r
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, code_col).value
    }

    updated = 0
    for d in descriptions:
        if not d or d.get("skipped"):
            continue
        row = code_to_row.get(d["code"])
        if row:
            cell = ws.cell(row, desc_col, d["description"])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            updated += 1

    wb.save(xlsx_path)
    print(f"\n  Excel: +{updated} описаний → {xlsx_path.name}")


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Генерация описаний для Ozon через DeepSeek",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--data",     default=str(SCRAPER_OUTPUT), help="Папка с JSON файлами скрапера")
    ap.add_argument("--rows",     default=None,  help="Диапазон артикулов: 1-5 или просто 3")
    ap.add_argument("--login",    action="store_true", help="Открыть браузер для входа в аккаунт DeepSeek")
    ap.add_argument("--debug",    action="store_true", help="Показывать браузер, делать скриншоты")
    ap.add_argument("--delay",    default=8, type=int, help="Пауза между артикулами, сек (по умолч. 8)")
    args = ap.parse_args()

    data_dir = Path(args.data)
    out_dir  = data_dir / "descriptions"
    out_dir.mkdir(parents=True, exist_ok=True)
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    # ── Список JSON файлов ────────────────────────────────────────────────────
    json_files = sorted(data_dir.glob("*.json"))
    if not json_files:
        print(f"JSON файлы не найдены в: {data_dir}")
        sys.exit(1)

    if args.rows:
        if "-" in args.rows:
            a, b = map(int, args.rows.split("-"))
            json_files = json_files[a - 1 : b]
        else:
            idx = int(args.rows) - 1
            json_files = [json_files[idx]] if 0 <= idx < len(json_files) else []

    if not json_files:
        print("Нет файлов для обработки (проверь --rows)")
        sys.exit(1)

    # ── Шапка ────────────────────────────────────────────────────────────────
    print()
    print("═" * 60)
    print("  DeepSeek Description Generator")
    print(f"  Артикулов:  {len(json_files)}")
    print(f"  Источник:   {data_dir}")
    print(f"  Результат:  {out_dir}")
    print(f"  Профиль:    {PROFILE_DIR}")
    print("═" * 60)

    headless = not (args.login or args.debug)

    with sync_playwright() as pw:
        context = pw.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=headless,
            viewport={"width": 1280, "height": 900},
            locale="ru-RU",
            args=["--disable-blink-features=AutomationControlled"],  # скрываем автоматизацию
        )
        page = context.new_page()

        # ── Режим входа ───────────────────────────────────────────────────────
        if args.login:
            print("\n  Открываю браузер DeepSeek...")
            print("  Залогинься в свой аккаунт DeepSeek.")
            page.goto(DEEPSEEK_URL, wait_until="domcontentloaded")
            print("  Когда окажешься в чате — нажми Enter здесь.")
            input("  > ")
            context.storage_state(path=str(PROFILE_DIR / "state.json"))
            print("  ✓ Сессия сохранена. Теперь можно запускать без --login.")
            context.close()
            return

        # ── Проверяем авторизацию ─────────────────────────────────────────────
        print("\n  Открываю DeepSeek...")
        page.goto(DEEPSEEK_URL, wait_until="domcontentloaded", timeout=30000)
        time.sleep(3)

        # Если редиректит на страницу входа — нужно залогиниться
        if any(kw in page.url.lower() for kw in ("login", "signin", "auth", "sign-in")):
            print("\n  ✗ Не авторизован!")
            print("  Запусти сначала:")
            print("    uv run --with playwright scripts/deepseek_description_generator.py --login")
            context.close()
            sys.exit(1)

        if args.debug:
            page.screenshot(path=str(out_dir / "_debug_start.png"))
            print(f"  Скриншот: {out_dir}/_debug_start.png")
            print(f"  URL: {page.url}")

        print("  ✓ Авторизован\n")

        # ── Основной цикл ────────────────────────────────────────────────────
        results = []
        for i, json_file in enumerate(json_files, 1):
            print(f"[{i:2}/{len(json_files)}]", end=" ")
            result = process_article(page, json_file, out_dir, debug=args.debug)
            if result:
                results.append(result)
            if i < len(json_files):
                time.sleep(args.delay)

        context.close()

    # ── Обновляем Excel ───────────────────────────────────────────────────────
    if results:
        update_excel(results, XLSX_PATH)

    done    = sum(1 for r in results if r and not r.get("skipped"))
    skipped = sum(1 for r in results if r and r.get("skipped"))
    failed  = len(json_files) - len(results)

    print()
    print("═" * 60)
    print(f"  Готово!")
    print(f"  Сгенерировано: {done}")
    print(f"  Пропущено:     {skipped}  (уже были)")
    print(f"  Ошибок:        {failed}")
    print(f"  Описания:      {out_dir}")
    print("═" * 60)
    print()


if __name__ == "__main__":
    main()
