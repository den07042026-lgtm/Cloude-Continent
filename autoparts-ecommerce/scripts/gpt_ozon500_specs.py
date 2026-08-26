# /// script
# requires-python = ">=3.10"
# dependencies = ["playwright", "openpyxl"]
# ///
"""
gpt_ozon500_specs.py
═════════════════════
Заполняет столбцы «OEM-номера» и «Альтернативные артикулы» в
"500 обогатитель.xlsx" через ChatGPT (браузер, Playwright, тот же CDP-механизм,
что и gpt_ozon500_descriptions.py — переиспользует его функции).

Приоритет — ПОЛНОТА: собрать максимум подтверждённых номеров по нескольким
каталогам (TecDoc, Exist, Emex, Autodoc, каталоги производителя), но не
выдумывать номера, которых не удалось подтвердить.

Уже заполненные строки (хотя бы одно из двух полей непусто) пропускаются —
можно продолжать после прерывания. Сохранение происходит часто и в finally,
чтобы сбой одной строки/сети не терял накопленный прогресс (см. историю гонки
записи в fill_applicability_mikado.py / gpt_ozon500_applicability.py).

Требует уже запущенный "тестовый" Chrome с CDP-портом 9222, авторизованный в
ChatGPT (scripts/launch_chatgpt.bat).

Тест:
  uv run --with playwright,openpyxl scripts/gpt_ozon500_specs.py --rows 2-4 --debug

Обычный запуск (продолжает с непройденных строк):
  uv run --with playwright,openpyxl scripts/gpt_ozon500_specs.py --delay 60
"""

import re
import sys
import time
import argparse
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).parent))

from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Alignment

from gpt_ozon500_descriptions import (
    CHATGPT_URL, CDP_PORT, SAVE_EVERY,
    load_category_map, dismiss_modals, type_message, enable_web_search,
    click_send, wait_for_response, safe_save, _is_rate_limited,
)

XLSX_PATH = Path(r"C:\Users\Admin\Desktop\Озон-500\500 обогатитель.xlsx")

_EMPTY_MARKERS = {"-", "—", "нет", "нет данных", "не указано", "n/a", ""}

_SPECS_PROMPT_TEMPLATE = """Ты — технический специалист по автомобильным запчастям с доступом к профессиональным каталогам подбора (TecDoc, Exist, Emex, Autodoc, каталоги производителей, агрегаторы кросс-номеров).

Исходные данные:

Наименование: {name}

Бренд: {brand}

Артикул: {code}

Категория детали: {category}

Собери МАКСИМАЛЬНО ПОЛНЫЙ список:

1. OEM-номера — все оригинальные номера автопроизводителей, под которыми эта САМАЯ деталь (именно тот компонент, что указан в наименовании как основной) устанавливалась на конвейере. Если деталь подходит под несколько марок/моделей и у каждой свой OEM-номер — перечисли все варианты, не ограничивайся одним.

2. Альтернативные артикулы — артикулы ЭТОЙ ЖЕ детали у других производителей запчастей (кросс-номера, взаимозаменяемые аналоги основного компонента). Не включай сам артикул {code} бренда {brand} в этот список.

КРИТИЧЕСКИ ВАЖНО — не путай компоненты составных позиций:
- Если наименование включает основную деталь ПЛЮС сопутствующий элемент (например "фильтр с прокладкой", "к-т", "комплект", "с датчиком", "с крепежом" и т.п.) — все номера в ответе должны относиться ТОЛЬКО к основной детали (например к самому фильтру), а не к прокладке/датчику/крепежу.
- Номер прокладки, уплотнителя, датчика или другого сопутствующего элемента — это номер ДРУГОЙ физической детали, он НЕ является ни OEM, ни аналогом основной детали. Не включай такие номера в список ни при каких обстоятельствах, даже если они встречаются в одном источнике рядом с основным артикулом.
- Перед тем как включить номер в ответ, убедись, что он относится именно к тому же типу детали, что и основная (например, номер фильтра — только с другим номером фильтра, а не с номером прокладки к нему).

Правила:
- Полнота важнее краткости: ищи по нескольким источникам и объединяй результаты, а не ограничивайся первым найденным номером.
- Каждый номер должен быть реально подтверждён источником. Ничего не выдумывай, не генерируй "правдоподобные" номера и не додумывай по аналогии с похожими артикулами.
- Если по какому-то из двух пунктов подтверждённых данных нет вообще — оставь после двоеточия пустое место (ничего не пиши). КАТЕГОРИЧЕСКИ ЗАПРЕЩЕНО писать фразы "не найдено", "данные отсутствуют", "информация не найдена" и подобные.

Ответ дай СТРОГО двумя строками в формате, без заголовков, пояснений, markdown и списков:

OEM-НОМЕРА: номер1; номер2; номер3
АЛЬТЕРНАТИВНЫЕ АРТИКУЛЫ: номер1; номер2; номер3"""


def build_prompt(row: dict) -> str:
    return _SPECS_PROMPT_TEMPLATE.format(
        name=row["name"],
        brand=row["brand"],
        code=row["code"],
        category=row.get("category") or "не указана",
    )


_FIELD_PATTERNS = [
    ("oem", re.compile(r'^\s*oem[-\s]?номера?\s*:\s*(.*)$', re.I)),
    ("alt", re.compile(r'^\s*альтернативн\w*\s+артикул\w*\s*:\s*(.*)$', re.I)),
]


def parse_specs(text: str) -> dict:
    text = re.sub(r'\*+', '', text)
    fields = {"oem": "", "alt": ""}
    found_labels = 0
    for line in text.splitlines():
        line = line.strip()
        for key, pat in _FIELD_PATTERNS:
            m = pat.match(line)
            if m:
                found_labels += 1
                val = m.group(1).strip(" .;")
                if val and val.lower() not in _EMPTY_MARKERS:
                    fields[key] = val
    fields["_found_labels"] = found_labels
    return fields


def process_row(page, row: dict, debug: bool, retries: int = 2) -> dict | None:
    for attempt in range(1, retries + 1):
        try:
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30_000)
            try:
                page.wait_for_load_state("networkidle", timeout=10_000)
            except Exception:
                pass
            time.sleep(4)
            dismiss_modals(page)

            try:
                page_text = page.locator("body").inner_text(timeout=3000)
            except Exception:
                page_text = ""
            if _is_rate_limited(page_text):
                print("    [!] ChatGPT: рейт-лимит (баннер страницы) — жду 20 мин...")
                time.sleep(1200)
                continue

            prompt = build_prompt(row)
            type_message(page, prompt)

            web_search_on = enable_web_search(page)
            if web_search_on:
                print("    [web search ON]", end=" ", flush=True)

            if debug and attempt == 1:
                print("    [debug] промпт введён, нажми Enter для отправки...")
                input()

            click_send(page)
            time.sleep(2)

            response = wait_for_response(page, timeout_sec=300 if web_search_on else 180)

            if not response:
                if attempt < retries:
                    print(f"    [!] Пустой ответ, попытка {attempt + 1}/{retries}...")
                    time.sleep(15)
                continue

            if _is_rate_limited(response):
                print("    [!] ChatGPT: рейт-лимит — жду 20 мин...")
                time.sleep(1200)
                continue

            fields = parse_specs(response)
            if fields["_found_labels"] < 2:
                print(f"    [!] Не распознан формат ответа (меток: {fields['_found_labels']}/2), попытка {attempt}")
                if attempt < retries:
                    time.sleep(10)
                continue
            return fields

        except RuntimeError as e:
            print(f"    [!] {e}")
            break
        except Exception as e:
            print(f"    [!] Ошибка (попытка {attempt}): {e}")
            if attempt < retries:
                time.sleep(15)
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--debug", action="store_true", help="Пауза перед отправкой (проверка промпта)")
    ap.add_argument("--rows", default=None, help="Диапазон строк Excel: 2-10 или одна строка: 5")
    ap.add_argument("--delay", default=60, type=int, help="Пауза между строками, сек")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb.active
    category_map = load_category_map()

    col_code, col_name, col_brand = 1, 2, 3
    col_oem, col_alt = 8, 9

    max_row = ws.max_row
    row_start, row_end = 2, max_row
    if args.rows:
        if "-" in args.rows:
            a, b = args.rows.split("-")
            row_start, row_end = int(a), int(b)
        else:
            row_start = row_end = int(args.rows)
    row_end = min(row_end, max_row)

    print()
    print("=" * 62)
    print("  GPT Ozon-500 Specs [OEM-номера / Альтернативные артикулы]")
    print(f"  Файл:   {XLSX_PATH.name}")
    print(f"  Строки: {row_start}-{row_end} ({row_end - row_start + 1} шт.)")
    print("=" * 62)

    with sync_playwright() as pw:
        browser = pw.chromium.connect_over_cdp(f"http://127.0.0.1:{CDP_PORT}")
        context = browser.contexts[0] if browser.contexts else browser.new_context()

        page = None
        for p in context.pages:
            if "chatgpt.com" in p.url:
                page = p
                break
        if page is None:
            page = context.pages[0] if context.pages else context.new_page()

        print("\n  Открываю ChatGPT...")
        try:
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30_000)
        except Exception as e:
            print(f"  Не удалось открыть ChatGPT: {e}")
            sys.exit(1)

        time.sleep(3)
        if any(k in page.url.lower() for k in ("login", "auth", "signin", "sign-in")):
            print("  Не авторизован! Войди в ChatGPT в открытом окне тестового Chrome и перезапусти скрипт.")
            sys.exit(1)

        print("  Авторизован\n")

        stats = {"done": 0, "skipped": 0, "failed": 0}
        since_save = 0

        try:
            for row_num in range(row_start, row_end + 1):
                code = ws.cell(row_num, col_code).value
                if not code:
                    continue
                code = str(code).strip()
                name = str(ws.cell(row_num, col_name).value or "")
                brand = str(ws.cell(row_num, col_brand).value or "")

                if ws.cell(row_num, col_oem).value or ws.cell(row_num, col_alt).value:
                    print(f"  [{row_num:3}] {code}: пропуск (уже заполнено)")
                    stats["skipped"] += 1
                    continue

                print(f"\n  [{row_num:3}/{row_end}] {brand} {code} — {name[:45]}")

                category = category_map.get(code.lower(), "")
                row_data = {"code": code, "name": name, "brand": brand, "category": category}
                result = process_row(page, row_data, debug=args.debug)

                if result:
                    for col, key in ((col_oem, "oem"), (col_alt, "alt")):
                        cell = ws.cell(row_num, col, result[key])
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                    print(f"         + OEM: {result['oem'][:70] or '(пусто)'}")
                    print(f"         + аналоги: {result['alt'][:70] or '(пусто)'}")
                    stats["done"] += 1
                    since_save += 1
                else:
                    print("         не удалось")
                    stats["failed"] += 1

                if since_save >= SAVE_EVERY:
                    safe_save(wb, XLSX_PATH)
                    since_save = 0

                if row_num < row_end:
                    time.sleep(args.delay)
        finally:
            safe_save(wb, XLSX_PATH)

        # НЕ закрываем context/browser - это подключение по CDP к уже открытому
        # "тестовому" Chrome пользователя, а не браузер, запущенный этим скриптом.

    print()
    print("=" * 62)
    print("  Готово!")
    print(f"  Заполнено: {stats['done']}")
    print(f"  Пропущено: {stats['skipped']}")
    print(f"  Ошибок:    {stats['failed']}")
    print("=" * 62)
    print()


if __name__ == "__main__":
    main()
