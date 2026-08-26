# /// script
# requires-python = ">=3.10"
# dependencies = ["playwright", "openpyxl"]
# ///
"""
gpt_ozon500_descriptions.py
════════════════════════════
Заполняет столбец «Описание» в "500 обогатитель.xlsx" через ChatGPT (браузер,
Playwright) — адаптировано из scripts/gpt_wb_filler.py, тот же механизм
(персистентный профиль браузера, без API-ключа).

Описание: 1200-1500 символов, технически грамотное, для Ozon и WB - то, что
нужно знать покупателю, чтобы не ошибиться с выбором (совместимость,
сторона установки, на что обратить внимание).

Уже заполненные строки пропускаются - можно продолжать после прерывания.
Пути захардкожены (кириллица в argv ломается - см. GUIDE_OZON_500.md).

Требует уже запущенный "тестовый" Chrome с CDP-портом 9222 и авторизацией в
ChatGPT (scripts/launch_chatgpt.bat открывает его, если ещё не запущен).

Тест (несколько строк, с паузой перед отправкой для проверки промпта):
  uv run --with playwright,openpyxl scripts/gpt_ozon500_descriptions.py --rows 2-4 --debug

Обычный запуск (продолжает с непройденных строк):
  uv run --with playwright,openpyxl scripts/gpt_ozon500_descriptions.py
"""

import re
import sys
import time
import argparse
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Alignment

XLSX_PATH = Path(r"C:\Users\Admin\Desktop\Озон-500\500 обогатитель.xlsx")
CATEGORY_SRC_PATH = Path(r"C:\Users\Admin\Desktop\Озон-500\Топ-500_Ozon_2026-07-06.xlsx")
CHATGPT_URL = "https://chatgpt.com/"
SAVE_EVERY = 5
CDP_PORT = 9222  # "тестовый" Chrome, запущен через scripts/launch_chatgpt.bat, уже залогинен


def load_category_map() -> dict[str, str]:
    """Код -> Категория из исходного файла отбора (столбец E листа Топ-500)."""
    wb = openpyxl.load_workbook(str(CATEGORY_SRC_PATH), data_only=True)
    ws = wb["Топ-500"]
    return {str(r[1]).strip().lower(): (r[4] or "") for r in ws.iter_rows(min_row=2, values_only=True)}


# Промпт составлен самим ChatGPT после анализа реальных карточек Ozon/WB
# (см. scripts/gpt_design_prompt.py и data/analytics/top500_ozon/designed_prompt.txt)
_DESIGNED_PROMPT_TEMPLATE = """Ты — технический специалист по автомобильным запчастям и профессиональный автор карточек товаров для маркетплейсов Ozon и Wildberries.

Твоя задача — написать технически грамотное описание товара объёмом строго 1200-1500 символов.

Исходные данные:

Наименование: {name}

Бренд: {brand}

Артикул: {code}

Категория детали: {category}

Если в наименовании отсутствуют важные технические сведения (тип конструкции, сторона установки, комплектность, размеры, материал, особенности исполнения, применяемость, OEM-номера и т.п.), обязательно выполни поиск по артикулу и бренду в открытых каталогах производителей и автомобильных каталогах (TecDoc, Exist, Emex, Autodoc, каталоги производителя и аналогичные источники) и используй только подтверждённые данные. Не выдумывай характеристики.

Если конкретную характеристику подтвердить не удалось - просто не упоминай её вообще, как будто вопрос о ней не стоял. КАТЕГОРИЧЕСКИ ЗАПРЕЩЕНО писать фразы о том, что данные не найдены, не подтверждены или отсутствуют в каталогах (например "сведения о материале не приводятся", "данные по размерам отсутствуют", "информация не подтверждена" и т.п.) - такие фразы недопустимы в готовом тексте ни в каком виде. Текст должен содержать только то, что удалось узнать, и ничего не должно напоминать о том, чего не удалось найти.

Описание должно быть написано как техническое описание детали, а не рекламный текст.

Структура текста:

Первый абзац — что это за деталь, её назначение и принцип работы именно для данной категории запчастей.

Второй абзац — технические особенности изделия. Включай только подтверждённые характеристики: тип детали, конструкцию, материал, особенности исполнения, сторону установки, наличие прокладок, крепежа, датчиков, уплотнений, керамический состав, газовое или масляное исполнение амортизатора, размеры, фильтрующий материал, особенности фрикционного слоя и другие параметры, если они известны.

Далее укажи основные автомобили, для которых предназначена деталь. Используй модели и годы выпуска из наименования, при наличии уточняй модификации или платформы. Не перечисляй чрезмерно длинные списки — оставляй только наиболее важную применяемость.

Последний абзац обязательно должен помочь покупателю избежать ошибки при выборе. Укажи, что перед покупкой необходимо сверить артикул, OEM-номер (если известен), модификацию автомобиля, год выпуска, тип двигателя, коробки передач, сторону установки (если применимо), размеры детали и комплектность поставки.

Требования к стилю:

только фактическая информация;

никаких рекламных обещаний;

никаких фраз вроде «высокое качество», «отличный выбор», «надёжное решение», «идеально подходит», «обеспечивает максимальную эффективность», «лучший вариант», «премиальное качество» и аналогичных маркетинговых клише;

не обращаться к читателю;

не использовать списки, таблицы и маркированные пункты;

не использовать эмодзи;

не использовать HTML и Markdown;

не повторять одно и то же разными словами;

писать естественным техническим языком, как в качественных карточках автозапчастей;

учитывать особенности конкретной категории детали. Для фильтров описывать очистку рабочей среды, конструкцию фильтрующего элемента и комплектность; для тормозных колодок — тип, материал фрикционного слоя, наличие противоскрипных пластин, датчиков износа, фасок и пазов (если подтверждено); для амортизаторов — тип (газовый, масляный, газомасляный), сторону установки, конструкцию; для сайлентблоков — материал, место установки и назначение; для катушек зажигания — тип, назначение и особенности конструкции; для других деталей делать акцент на их реальной функции и технических особенностях.

Текст должен выглядеть как описание, написанное техническим специалистом для карточки товара на Ozon и Wildberries, помогать подобрать правильную запчасть и не содержать недостоверной информации.

Ответ - ТОЛЬКО текст описания, без заголовков, без пояснений до или после."""


def build_prompt(row: dict) -> str:
    return _DESIGNED_PROMPT_TEMPLATE.format(
        name=row["name"],
        brand=row["brand"],
        code=row["code"],
        category=row.get("category") or "не указана",
    )


# ══════════════════════════════════════════════════════════
# PLAYWRIGHT — те же селекторы/функции, что в gpt_wb_filler.py
# ══════════════════════════════════════════════════════════

_INPUT_SELECTORS = [
    '#prompt-textarea',
    'div[contenteditable="true"][id="prompt-textarea"]',
    '[data-testid="composer-input"]',
    'div.ProseMirror[contenteditable="true"]',
    'p[data-placeholder*="Message"]',
    'p[data-placeholder*="Сообщение"]',
    'div[contenteditable="true"]',
    'textarea',
]

_SEND_SELECTORS = [
    'button[data-testid="send-button"]',
    'button[aria-label*="Send"]',
    'button[aria-label*="send"]',
]

_STOP_SELECTORS = [
    'button[data-testid="stop-button"]',
    'button[aria-label*="Stop"]',
    'button[aria-label*="stop"]',
]

# Кнопка веб-поиска в ChatGPT (селекторы меняются с обновлениями UI)
_SEARCH_SELECTORS = [
    'button[aria-label="Search the web"]',
    'button[aria-label*="Search"]',
    'button[data-testid="composer-search-button"]',
    'button[aria-label*="Поиск"]',
    '[aria-label*="web search"]',
    'button[aria-label*="browse"]',
]


def enable_web_search(page) -> bool:
    """Включает веб-поиск в ChatGPT (если кнопка есть в UI). Возвращает True при успехе."""
    btn = _find_visible(page, _SEARCH_SELECTORS, timeout=2000)
    if btn:
        try:
            btn.click()
            time.sleep(0.5)
            return True
        except Exception:
            pass
    return False


def _find_visible(page, selectors: list, timeout: int = 2000):
    for sel in selectors:
        try:
            page.wait_for_selector(sel, state="visible", timeout=timeout)
            return page.locator(sel).last
        except Exception:
            continue
    return None


def dismiss_modals(page):
    close_selectors = [
        'button[aria-label="Close"]', 'button[data-testid="close-button"]',
        'button:has-text("Dismiss")', 'button:has-text("Maybe later")',
        'button:has-text("No thanks")', 'button:has-text("OK")', 'button:has-text("Ok")',
        'button:has-text("Got it")', 'button:has-text("Stay on free plan")',
        'button:has-text("Keep current plan")', 'button:has-text("Skip for now")',
        'button:has-text("Continue")', 'button:has-text("Понятно")',
        'button:has-text("Понял")', 'button:has-text("Хорошо")',
        'button:has-text("Закрыть")', 'button:has-text("ОК")', 'button:has-text("Ок")',
        'button:has-text("Остаться")', 'button:has-text("Продолжить")',
        'button:has-text("Позже")', '[data-testid="modal-close"]',
    ]
    for sel in close_selectors:
        try:
            btn = page.locator(sel).first
            if btn.is_visible():
                btn.click()
                time.sleep(0.4)
        except Exception:
            continue
    try:
        page.keyboard.press("Escape")
        time.sleep(0.3)
    except Exception:
        pass


def type_message(page, text: str):
    dismiss_modals(page)
    el = _find_visible(page, _INPUT_SELECTORS, timeout=8_000)
    if el is None:
        time.sleep(4)
        dismiss_modals(page)
        el = _find_visible(page, _INPUT_SELECTORS, timeout=8_000)
    if el is None:
        raise RuntimeError("Поле ввода ChatGPT не найдено")

    el.click()
    time.sleep(0.3)

    page.evaluate(
        """([selectors, text]) => {
            let el = null;
            for (const sel of selectors) {
                el = document.querySelector(sel);
                if (el) break;
            }
            if (!el) return;
            el.focus();
            if (el.isContentEditable) {
                document.execCommand('selectAll', false, null);
                document.execCommand('insertText', false, text);
            } else {
                const setter = Object.getOwnPropertyDescriptor(
                    HTMLTextAreaElement.prototype, 'value').set;
                setter.call(el, text);
                el.dispatchEvent(new Event('input',  {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
            }
        }""",
        [_INPUT_SELECTORS, text],
    )
    time.sleep(0.5)


def click_send(page):
    btn = _find_visible(page, _SEND_SELECTORS, timeout=3000)
    if btn:
        btn.click()
    else:
        page.keyboard.press("Enter")


def is_generating(page) -> bool:
    return _find_visible(page, _STOP_SELECTORS, timeout=500) is not None


def get_last_response(page) -> str:
    try:
        msgs = page.locator('[data-message-author-role="assistant"]').all()
        if msgs:
            return msgs[-1].inner_text(timeout=2000).strip()
    except Exception:
        pass
    try:
        for sel in ['article', '[class*="message"]:last-child']:
            items = page.locator(sel).all()
            for item in reversed(items):
                txt = item.inner_text(timeout=1000).strip()
                if txt and len(txt) > 30:
                    return txt
    except Exception:
        pass
    return ""


_RATE_LIMIT_PHRASES = (
    "too many requests", "sending messages too quickly",
    "you've reached", "reached our limit",
    "слишком много запросов", "слишком быстро",
    "come back later", "try again later",
    "запросы слишком часто", "доступ к вашим диалогам",
    "временно ограничен", "подождите несколько минут",
)


def _is_rate_limited(text: str) -> bool:
    low = text.lower()
    return any(p in low for p in _RATE_LIMIT_PHRASES)


def wait_for_response(page, timeout_sec: int = 180) -> str:
    print("    жду", end="", flush=True)
    for _ in range(30):
        dismiss_modals(page)
        text = get_last_response(page)
        if text and len(text) > 20:
            break
        time.sleep(2)
        print(".", end="", flush=True)
    else:
        print(" [нет ответа]", flush=True)
        return ""

    prev, stable = "", 0
    for tick in range(timeout_sec):
        if tick % 10 == 0:
            dismiss_modals(page)
        cur = get_last_response(page)
        if cur == prev and cur:
            stable += 1
            if stable >= 3 and not is_generating(page):
                print(f" ({len(cur)} симв.)", flush=True)
                return cur
        else:
            stable = 0
            if tick % 5 == 0:
                print(".", end="", flush=True)
        prev = cur
        time.sleep(1)

    print(f" [таймаут, {len(prev)} симв.]", flush=True)
    return prev


_UI_ARTIFACT_LINES = {
    "редактировать", "edit", "копировать", "copy", "поделиться", "share",
    "регенерировать", "regenerate", "прочитать вслух", "read aloud",
}


_CITATION_LINE_RE = re.compile(
    r'^(\+\d+|[A-ZА-ЯЁ][\wа-яё-]*\.(?:ру|ru|com|рф)|[\wа-яёА-ЯЁ][\wа-яё-]{1,20}\.(?:ру|ru|com|рф))\s*$',
    re.IGNORECASE,
)


def clean_description(text: str) -> str:
    text = re.sub(r'\*+', '', text)
    text = re.sub(r'^(ОПИСАНИЕ|Описание)\s*:?\s*\n?', '', text.strip())
    lines = text.split("\n")
    # убираем случайно захваченные кнопки интерфейса ChatGPT (Edit/Copy/...) в начале
    while lines and lines[0].strip().lower() in _UI_ARTIFACT_LINES:
        lines.pop(0)
    # убираем "плашки" цитируемых источников веб-поиска (напр. "Тачка.Ру", "+1"),
    # захваченные как отдельные короткие строки внутри текста
    lines = [ln for ln in lines if not _CITATION_LINE_RE.match(ln.strip())]
    text = "\n".join(lines)
    # схлопываем более двух подряд пустых строк (могли остаться после удаления цитат)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def process_row(page, row: dict, debug: bool, retries: int = 2) -> str | None:
    for attempt in range(1, retries + 1):
        try:
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30_000)
            try:
                page.wait_for_load_state("networkidle", timeout=10_000)
            except Exception:
                pass
            time.sleep(4)
            dismiss_modals(page)

            # лимит ChatGPT может показываться баннером на странице, а не как
            # обычное сообщение ассистента - проверяем весь текст страницы
            try:
                page_text = page.locator("body").inner_text(timeout=3000)
            except Exception:
                page_text = ""
            if _is_rate_limited(page_text):
                print("    [!] ChatGPT: рейт-лимит (баннер страницы) — жду 20 мин...")
                time.sleep(1200)
                continue

            prompt = build_prompt(row)
            type_message(page, prompt)

            web_search_on = enable_web_search(page)
            if web_search_on:
                print("    [web search ON]", end=" ", flush=True)

            if debug and attempt == 1:
                print("    [debug] промпт введён, нажми Enter для отправки...")
                input()

            click_send(page)
            time.sleep(2)

            response = wait_for_response(page, timeout_sec=300 if web_search_on else 180)

            if not response:
                if attempt < retries:
                    print(f"    [!] Пустой ответ, попытка {attempt + 1}/{retries}...")
                    time.sleep(15)
                continue

            if _is_rate_limited(response):
                print("    [!] ChatGPT: рейт-лимит — жду 20 мин...")
                time.sleep(1200)
                continue

            desc = clean_description(response)
            if len(desc) < 300:
                print(f"    [!] Слишком короткий ответ ({len(desc)} симв.), попытка {attempt}")
                if attempt < retries:
                    time.sleep(10)
                continue
            return desc

        except RuntimeError as e:
            print(f"    [!] {e}")
            break
        except Exception as e:
            print(f"    [!] Ошибка (попытка {attempt}): {e}")
            if attempt < retries:
                time.sleep(15)
    return None


def safe_save(wb, path: Path):
    for attempt in range(1, 6):
        try:
            wb.save(str(path))
            print(f"    [сохранено: {path.name}]")
            return
        except PermissionError:
            print(f"\n    [!] Файл занят — закройте Excel и нажмите Enter (попытка {attempt}/5)...")
            input()
    raise PermissionError(f"Не удалось сохранить {path} после 5 попыток")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--debug", action="store_true", help="Пауза перед отправкой (проверка промпта)")
    ap.add_argument("--rows", default=None, help="Диапазон строк Excel: 2-10 или одна строка: 5")
    ap.add_argument("--delay", default=20, type=int, help="Пауза между строками, сек")
    ap.add_argument("--redo-all", action="store_true",
                     help="Очистить существующие описания и сгенерировать заново все")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb.active
    category_map = load_category_map()

    col_code, col_name, col_brand, col_desc = 1, 2, 3, 6

    if args.redo_all:
        cleared = 0
        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row_num, col_desc).value:
                ws.cell(row_num, col_desc).value = None
                cleared += 1
        print(f"  --redo-all: очищено {cleared} существующих описаний")

    max_row = ws.max_row
    row_start, row_end = 2, max_row
    if args.rows:
        if "-" in args.rows:
            a, b = args.rows.split("-")
            row_start, row_end = int(a), int(b)
        else:
            row_start = row_end = int(args.rows)
    row_end = min(row_end, max_row)

    print()
    print("=" * 62)
    print("  GPT Ozon-500 Descriptions [ChatGPT Browser + Playwright]")
    print(f"  Файл:   {XLSX_PATH.name}")
    print(f"  Строки: {row_start}-{row_end} ({row_end - row_start + 1} шт.)")
    print("=" * 62)

    with sync_playwright() as pw:
        browser = pw.chromium.connect_over_cdp(f"http://127.0.0.1:{CDP_PORT}")
        context = browser.contexts[0] if browser.contexts else browser.new_context()

        # Ищем уже открытую вкладку ChatGPT, иначе берём первую или создаём новую
        page = None
        for p in context.pages:
            if "chatgpt.com" in p.url:
                page = p
                break
        if page is None:
            page = context.pages[0] if context.pages else context.new_page()

        print("\n  Открываю ChatGPT...")
        try:
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30_000)
        except Exception as e:
            print(f"  Не удалось открыть ChatGPT: {e}")
            sys.exit(1)

        time.sleep(3)
        if any(k in page.url.lower() for k in ("login", "auth", "signin", "sign-in")):
            print("  Не авторизован! Войди в ChatGPT в открытом окне тестового Chrome и перезапусти скрипт.")
            sys.exit(1)

        print("  Авторизован\n")

        stats = {"done": 0, "skipped": 0, "failed": 0}
        since_save = 0

        for row_num in range(row_start, row_end + 1):
            code = ws.cell(row_num, col_code).value
            if not code:
                continue
            code = str(code).strip()
            name = str(ws.cell(row_num, col_name).value or "")
            brand = str(ws.cell(row_num, col_brand).value or "")

            if ws.cell(row_num, col_desc).value:
                print(f"  [{row_num:3}] {code}: пропуск (уже заполнено)")
                stats["skipped"] += 1
                continue

            print(f"\n  [{row_num:3}/{row_end}] {brand} {code} — {name[:45]}")

            category = category_map.get(code.lower(), "")
            row_data = {"code": code, "name": name, "brand": brand, "category": category}
            result = process_row(page, row_data, debug=args.debug)

            if result:
                cell = ws.cell(row_num, col_desc, result)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                print(f"         + описание {len(result)} симв.")
                stats["done"] += 1
                since_save += 1
            else:
                print("         не удалось")
                stats["failed"] += 1

            if since_save >= SAVE_EVERY:
                safe_save(wb, XLSX_PATH)
                since_save = 0

            if row_num < row_end:
                time.sleep(args.delay)

        # НЕ закрываем context/browser - это подключение по CDP к уже открытому
        # "тестовому" Chrome пользователя, а не браузер, запущенный этим скриптом.

    safe_save(wb, XLSX_PATH)

    print()
    print("=" * 62)
    print("  Готово!")
    print(f"  Заполнено: {stats['done']}")
    print(f"  Пропущено: {stats['skipped']}")
    print(f"  Ошибок:    {stats['failed']}")
    print("=" * 62)
    print()


if __name__ == "__main__":
    main()
