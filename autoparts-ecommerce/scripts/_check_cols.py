import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
wb = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ.xlsx')
ws = wb.worksheets[0]
for c in range(1, ws.max_column+1):
    v  = ws.cell(1, c).value
    ex = ws.cell(2, c).value
    if v:
        print(f'  col{c:2d}: {str(v)[:30]:30s}  ex: {str(ex or "")[:60]}')
