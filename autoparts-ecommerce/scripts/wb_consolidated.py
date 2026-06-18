"""
wb_consolidated.py
════════════════════════════════════════════════════════════════════════════
Сводный файл из трёх источников матчей:
  - wb_art_match_*.xlsx        (Автолига, vendor_codes метод)
  - wb_matches_OEM_*.xlsx      (Автолига, basket OEM метод)
  - wb_matches_MIKADO_*.xlsx   (Микадо, basket OEM метод)

Шаги:
  1. Парсим все три файла → уникальные позиции по nm_id (дубли: берём где закупочная цена ниже)
  2. Подтягиваем свежие данные MPStats (sales_30d, oos_pct, цена) для каждого nm_id
  3. Считаем скор дефицита: oos_pct × log(1 + sales_30d)
  4. Делим на два листа:
       «Совпадение бренда»   — бренд поставщика == бренд WB
       «Разные бренды»       — не совпадают, но хорошие показатели
  5. Сортируем по скору сверху вниз

Запуск:
  cd C:\\Users\\Admin\\Documents\\Autoparts_Ecommerce
  uv run --with openpyxl,requests,python-dotenv scripts/wb_consolidated.py
"""

import sys, os, re, math, logging, time
from pathlib import Path
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl не установлен"); sys.exit(1)

try:
    import requests
except ImportError:
    print("requests не установлен"); sys.exit(1)

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

BASE_DIR       = Path(__file__).parent.parent
ANALYTICS_DIR  = BASE_DIR / "data" / "analytics"
LOG_FILE       = BASE_DIR / "logs" / "wb_consolidated.log"
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)

MPSTATS_TOKEN  = os.getenv("MPSTATS_TOKEN", "")
MPSTATS_WORKERS = 8   # параллельных запросов к MPStats

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ─── WB тарифы (FBS, Автозапчасти) ──────────────────────────────────────────
_WB_COMMISSION  = 0.25
_WB_TAX         = 0.06
_WB_RET_RATE    = 0.03
_WB_PACKAGING   = 30
_WB_DEL_BASE    = 50.6
_WB_DEL_LITER   = 15.4
_WB_RET_BASE    = 136.0
_WB_RET_LITER   = 14.0
_WB_DEFAULT_VOL = 3.0

# Порог «хороших показателей» для листа «Разные бренды»
_GOOD_SCORE_THRESHOLD = 20.0   # oos_pct × log(1+sales)


def _norm_art(s: str) -> str:
    return re.sub(r"[^A-ZА-ЯЁ0-9]", "", s.upper())


def _is_valid_article(norm: str) -> bool:
    """Отсеивает слишком короткие и чисто числовые артикулы-«мусор»."""
    if len(norm) < 5:
        return False
    if norm.isdigit() and len(norm) < 7:
        return False
    return True


def _norm_brand(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def _brands_match(a: str, b: str) -> bool:
    a, b = _norm_brand(a), _norm_brand(b)
    if not a or not b:
        return False
    return a == b or a.startswith(b) or b.startswith(a)


def _wb_margin(purchase: float, sell: float, liters: float = _WB_DEFAULT_VOL) -> float:
    if sell <= 0:
        return 0.0
    commission = sell * _WB_COMMISSION
    delivery   = _WB_DEL_BASE + liters * _WB_DEL_LITER
    ret_cost   = _WB_RET_RATE * (_WB_RET_BASE + liters * _WB_RET_LITER)
    tax        = max(0.0, sell - commission - delivery) * _WB_TAX
    profit     = sell - purchase - commission - delivery - ret_cost - tax - _WB_PACKAGING
    return profit / sell


def _wb_find_price(purchase: float, target: float) -> int | None:
    if purchase <= 0:
        return None
    if _wb_margin(purchase, 500_000) < target:
        return None
    lo, hi = 50, 500_000
    while lo < hi:
        mid = (lo + hi) // 2
        if _wb_margin(purchase, mid) >= target - 1e-6:
            hi = mid
        else:
            lo = mid + 1
    return lo if _wb_margin(purchase, lo) >= target - 1e-6 else None


# ─── Парсинг исходных файлов ──────────────────────────────────────────────────

def _latest(mask: str) -> Path | None:
    files = sorted(ANALYTICS_DIR.glob(mask), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def _parse_wb_art_match(path: Path) -> list[dict]:
    """wb_art_match_*.xlsx — листы 'Топ дефицит' и 'В наличии' (+ 'Нет в наличии')."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    out = []
    for sheet_name in ("Топ дефицит", "В наличии", "Нет в наличии"):
        if sheet_name not in wb.sheetnames:
            continue
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Заголовки: №, Арт.произв.(Автолига), Бренд(Автолига), Название, Остаток,
        #            Цена закупки, nm_id WB, Арт.произв.(WB), Название WB, Бренд WB,
        #            Предмет WB, Цена WB, OOS%, Продаж/мес, Скор, Маржа, Цена мин, Цена опт, Ссылка
        for r in rows[1:]:
            if not r[6]:
                continue
            try:
                nm_id = int(r[6])
            except (TypeError, ValueError):
                continue
            sup_art = str(r[1] or "").strip()
            if not _is_valid_article(_norm_art(sup_art)):
                continue
            out.append({
                "nm_id":         nm_id,
                "sup_article":   sup_art,
                "sup_brand":     str(r[2] or "").strip(),
                "sup_price":     float(r[5] or 0),
                "wb_art":        str(r[7] or "").strip(),
                "wb_name":       str(r[8] or "").strip(),
                "wb_brand":      str(r[9] or "").strip(),
                "wb_subject":    str(r[10] or "").strip(),
                "source":        "Автолига",
            })
    wb.close()
    log.info(f"  wb_art_match: {len(out)} позиций")
    return out


def _parse_simple(path: Path, source: str) -> list[dict]:
    """wb_matches_OEM_*.xlsx и wb_matches_MIKADO_*.xlsx — лист 'Матчи'."""
    wb   = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws   = wb["Матчи"]
    rows = list(ws.iter_rows(values_only=True))
    # Заголовки: Артикул, Бренд, Цена (руб), nm_id, Бренд WB, Цена WB, Продажи/30д,
    #            Категория WB, Артикул WB, Ссылка WB
    out = []
    for r in rows[1:]:
        if not r[3]:
            continue
        try:
            nm_id = int(r[3])
        except (TypeError, ValueError):
            continue
        sup_art = str(r[0] or "").strip()
        if not _is_valid_article(_norm_art(sup_art)):
            continue
        out.append({
            "nm_id":       nm_id,
            "sup_article": sup_art,
            "sup_brand":   str(r[1] or "").strip(),
            "sup_price":   float(r[2] or 0),
            "wb_art":      str(r[8] or "").strip(),
            "wb_name":     "",
            "wb_brand":    str(r[4] or "").strip(),
            "wb_subject":  str(r[7] or "").strip(),
            "source":      source,
        })
    wb.close()
    log.info(f"  {source}: {len(out)} позиций")
    return out


def load_all_sources() -> list[dict]:
    """Загружает все три источника, дедуплицирует по nm_id (оставляем min цену закупки)."""
    art_match = _latest("wb_art_match_*.xlsx")
    oem_match = _latest("wb_matches_OEM_*.xlsx")
    mik_match = _latest("wb_matches_MIKADO_*.xlsx")

    if not art_match: log.warning("wb_art_match_*.xlsx не найден")
    if not oem_match: log.warning("wb_matches_OEM_*.xlsx не найден")
    if not mik_match: log.warning("wb_matches_MIKADO_*.xlsx не найден")

    all_rows: list[dict] = []
    if art_match: all_rows += _parse_wb_art_match(art_match)
    if oem_match: all_rows += _parse_simple(oem_match, "Автолига")
    if mik_match: all_rows += _parse_simple(mik_match, "Микадо")

    # Дедупликация по nm_id — берём строку с меньшей ненулевой ценой закупки
    by_nm: dict[int, dict] = {}
    for row in all_rows:
        nm = row["nm_id"]
        if nm not in by_nm:
            by_nm[nm] = row
        else:
            ep = by_nm[nm]["sup_price"]
            np_ = row["sup_price"]
            if np_ > 0 and (ep <= 0 or np_ < ep):
                by_nm[nm] = row

    result = list(by_nm.values())
    log.info(f"Итого уникальных nm_id: {len(result)}")
    return result


# ─── MPStats: свежие данные ───────────────────────────────────────────────────

def _fetch_mpstats(nm_id: int, d1: str, d2: str) -> dict | None:
    """Возвращает {sales_30d, oos_pct, wb_price, wb_name} или None при ошибке."""
    url     = f"https://mpstats.io/api/wb/get/item/{nm_id}/sales"
    headers = {"X-Mpstats-TOKEN": MPSTATS_TOKEN}
    for attempt in range(3):
        try:
            r = requests.get(url, headers=headers,
                             params={"d1": d1, "d2": d2}, timeout=20)
            if r.status_code == 200:
                data = r.json()
                if not data:
                    return {"sales_30d": 0, "oos_pct": 0.0,
                            "wb_price_final": 0, "wb_price_full": 0}
                sales_30d = sum(int(d.get("sales") or 0) for d in data)
                oos_days  = sum(1 for d in data if int(d.get("balance") or 0) == 0)
                oos_pct   = round(oos_days / len(data) * 100, 1)
                last      = data[-1]
                # final_price = цена для покупателя (после скидок WB)
                # price       = «зачёркнутая» полная цена (без скидок)
                wb_price_final = float(last.get("final_price") or last.get("price") or 0)
                wb_price_full  = float(last.get("price") or 0)
                return {"sales_30d": sales_30d, "oos_pct": oos_pct,
                        "wb_price_final": wb_price_final,
                        "wb_price_full":  wb_price_full}
            if r.status_code == 429:
                time.sleep(2 * (attempt + 1))
                continue
            return None
        except Exception:
            if attempt < 2:
                time.sleep(1)
    return None


def enrich_from_db(positions: list[dict]) -> list[dict]:
    """Фоллбэк: берём sales_30d, oos_pct, price_rub из wb_products (локальная БД)."""
    import sqlite3
    db_path = BASE_DIR / "data" / "analytics" / "wb_index.db"
    conn = sqlite3.connect(str(db_path), timeout=10)
    rows = conn.execute(
        "SELECT nm_id, price_rub, sales_30d, oos_pct, name FROM wb_products"
    ).fetchall()
    conn.close()
    db: dict[int, tuple] = {r[0]: r for r in rows}
    for p in positions:
        r = db.get(p["nm_id"])
        if r:
            p["wb_price_fresh"] = float(r[1] or 0)
            p["wb_price_full"]  = float(r[1] or 0)
            p["sales_30d"]      = int(r[2]   or 0)
            p["oos_pct"]        = float(r[3] or 0)
            if not p.get("wb_name") and r[4]:
                p["wb_name"] = str(r[4])
        else:
            p.setdefault("wb_price_fresh", 0)
            p.setdefault("wb_price_full",  0)
            p.setdefault("sales_30d",      0)
            p.setdefault("oos_pct",        0.0)
    log.info(f"БД: обогащено {sum(1 for p in positions if p.get('sales_30d',0)>0)} позиций")
    return positions


def enrich_mpstats(positions: list[dict], use_db: bool = False) -> list[dict]:
    if use_db:
        log.info("Режим --use-db: берём данные из wb_products (локальная БД)...")
        return enrich_from_db(positions)

    if not MPSTATS_TOKEN:
        log.warning("MPSTATS_TOKEN не задан — используем БД")
        return enrich_from_db(positions)

    d2 = datetime.now().strftime("%Y-%m-%d")
    d1 = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    log.info(f"Запрашиваем MPStats для {len(positions)} nm_id  ({d1} → {d2})...")

    results: dict[int, dict] = {}
    ok = err = 0
    t0 = time.time()

    with ThreadPoolExecutor(max_workers=MPSTATS_WORKERS) as exe:
        futures = {exe.submit(_fetch_mpstats, p["nm_id"], d1, d2): p["nm_id"]
                   for p in positions}
        for i, fut in enumerate(as_completed(futures), 1):
            nm_id = futures[fut]
            data  = fut.result()
            if data:
                results[nm_id] = data
                ok += 1
            else:
                err += 1
            if i % 100 == 0 or i == len(positions):
                elapsed = time.time() - t0
                rate    = i / elapsed if elapsed > 0 else 0
                log.info(f"  [{i}/{len(positions)}]  ok={ok}  err={err}  {rate:.1f} req/с")
            # Если первые 30 все упали — сеть недоступна, переключаемся на БД
            if i == 30 and ok == 0:
                log.warning("MPStats недоступен — переключаемся на локальную БД")
                return enrich_from_db(positions)

    log.info(f"MPStats: получено {ok}, ошибок {err}")

    for p in positions:
        mp = results.get(p["nm_id"])
        if mp:
            p["sales_30d"]      = mp["sales_30d"]
            p["oos_pct"]        = mp["oos_pct"]
            p["wb_price_fresh"] = mp["wb_price_final"]
            p["wb_price_full"]  = mp["wb_price_full"]
        else:
            p.setdefault("sales_30d",      0)
            p.setdefault("oos_pct",        0.0)
            p.setdefault("wb_price_fresh", 0)
            p.setdefault("wb_price_full",  0)

    return positions


# ─── Скоры и цены ────────────────────────────────────────────────────────────

def calc_scores(positions: list[dict]) -> None:
    for p in positions:
        sales = p.get("sales_30d") or 0
        oos   = p.get("oos_pct")   or 0.0
        price = p.get("wb_price_fresh") or 0.0
        purch = p.get("sup_price")  or 0.0
        p["score"]        = round(oos * math.log1p(sales), 2)
        p["sell_min"]     = _wb_find_price(purch, 0.0)  if purch > 0 else None
        p["sell_opt"]     = _wb_find_price(purch, 0.15) if purch > 0 else None
        p["margin_at_wb"] = round(_wb_margin(purch, price) * 100, 1) if purch > 0 and price > 0 else None


# ─── Excel ───────────────────────────────────────────────────────────────────

_BORDER = Border(
    left=Side(style="thin",  color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="EEEEEE"),
)
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YLW_FILL = PatternFill("solid", fgColor="FFEB9C")
GRN_FILL = PatternFill("solid", fgColor="C6EFCE")
BLU_FILL = PatternFill("solid", fgColor="BDD7EE")

HEADERS = [
    ("№",                        4),
    ("Артикул\n(поставщик)",    20),
    ("Бренд\n(поставщик)",      16),
    ("Источник",                10),
    ("Цена\nзакупки, ₽",       13),
    ("nm_id WB",                12),
    ("Арт. произв.\n(WB)",     22),
    ("Название WB",             45),
    ("Бренд WB",                16),
    ("Предмет WB",              30),
    ("Цена WB\n(со скидкой), ₽", 14),
    ("Цена WB\n(полная), ₽",   13),
    ("OOS %",                    8),
    ("Продаж/\nмес",            10),
    ("Скор\nдефицит",            9),
    ("Маржа при\nцене WB,%",   14),
    ("Цена мин.\n(0%), ₽",     13),
    ("Цена опт.\n(15%), ₽",    14),
    ("Ссылка WB",               12),
]

OOS_COL    = 13
MARGIN_COL = 16
LINK_COL   = 19
SRC_COL    = 4


def _write_header(ws, color: str) -> None:
    fill  = PatternFill("solid", fgColor=color)
    font  = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, (title, width) in enumerate(HEADERS, 1):
        c = ws.cell(1, ci, title)
        c.font = font; c.fill = fill
        c.alignment = align; c.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def _write_rows(ws, items: list[dict]) -> None:
    src_colors = {"Автолига": "1F4E79", "Микадо": "1A5C1A"}
    for ri, p in enumerate(items, 2):
        price_final = p.get("wb_price_fresh") or 0
        price_full  = p.get("wb_price_full")  or 0
        row = [
            ri - 1,
            p["sup_article"],
            p["sup_brand"],
            p["source"],
            p["sup_price"] or "—",
            p["nm_id"],
            p["wb_art"],
            p["wb_name"],
            p["wb_brand"],
            p["wb_subject"],
            price_final or "—",
            price_full  or "—",
            p.get("oos_pct", 0),
            p.get("sales_30d", 0),
            p.get("score", 0),
            p["margin_at_wb"] if p.get("margin_at_wb") is not None else "—",
            p["sell_min"]     if p.get("sell_min")     is not None else "—",
            p["sell_opt"]     if p.get("sell_opt")     is not None else "—",
            "→ WB",
        ]
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.border = _BORDER
            c.alignment = Alignment(vertical="center")

        # Ссылка
        lc = ws.cell(ri, LINK_COL)
        lc.hyperlink = f"https://www.wildberries.ru/catalog/{p['nm_id']}/detail.aspx"
        lc.font = Font(color="0563C1", underline="single")

        # Источник — цветная метка
        sc = ws.cell(ri, SRC_COL)
        color = src_colors.get(p["source"], "555555")
        sc.fill = PatternFill("solid", fgColor=color)
        sc.font = Font(color="FFFFFF", bold=True, size=9)
        sc.alignment = Alignment(horizontal="center", vertical="center")

        # OOS цвет
        oos = p.get("oos_pct", 0)
        if oos >= 30:
            ws.cell(ri, OOS_COL).fill = RED_FILL
        elif oos >= 15:
            ws.cell(ri, OOS_COL).fill = YLW_FILL

        # Маржа цвет
        mrg = p.get("margin_at_wb")
        if mrg is not None:
            ws.cell(ri, MARGIN_COL).fill = (
                GRN_FILL if mrg >= 15 else
                YLW_FILL if mrg >= 0  else
                RED_FILL
            )

    if not items:
        ws.cell(2, 1, "Нет данных")


def export_excel(matched: list[dict], diff_brand: list[dict],
                 suspicious: list[dict]) -> Path:
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    sheets = [
        ("Топ дефицит",      "8B0000", matched),
        ("Разные бренды",    "555555", diff_brand),
        ("Подозрительные",   "7B4800", suspicious),
    ]
    for title, color, items in sheets:
        ws = wb_out.create_sheet(title)
        _write_header(ws, color)
        _write_rows(ws, items)
        log.info(f"  Лист '{title}': {len(items)} строк")

    ws_info = wb_out.create_sheet("Инфо")
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_info.append(["Сформировано", ts])
    ws_info.append(["Топ дефицит (бренды совпадают)", str(len(matched))])
    ws_info.append(["Разные бренды (хорошие показатели)", str(len(diff_brand))])
    ws_info.append(["Подозрительные (опт.цена < 40% WB)", str(len(suspicious))])
    ws_info.append(["Примечание", "Подозрительные — вероятные ложные матчи, требуют ручной проверки"])
    ws_info.column_dimensions["A"].width = 40
    ws_info.column_dimensions["B"].width = 18

    ts_fn = datetime.now().strftime("%Y%m%d_%H%M")
    path  = ANALYTICS_DIR / f"wb_consolidated_{ts_fn}.xlsx"
    wb_out.save(str(path))
    return path


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--use-db", action="store_true",
                        help="Использовать кэш из wb_products вместо запросов к MPStats")
    args = parser.parse_args()

    log.info("=" * 62)
    log.info("WB Consolidated  |  Автолига + Микадо → сводный дефицит")
    log.info("=" * 62)

    log.info("Загружаем источники...")
    positions = load_all_sources()

    log.info("Обогащаем данными MPStats...")
    positions = enrich_mpstats(positions, use_db=args.use_db)

    log.info("Считаем скоры и цены...")
    calc_scores(positions)

    # Убираем убыточные (закупка >= цена WB)
    before    = len(positions)
    positions = [
        p for p in positions
        if not (p["sup_price"] > 0 and p.get("wb_price_fresh", 0) > 0
                and p["sup_price"] >= p["wb_price_fresh"])
    ]
    if before != len(positions):
        log.info(f"Убыточных отфильтровано: {before - len(positions)}")

    # Подозрительные: наша оптимальная цена < 40% от цены WB → вероятный ложный матч
    suspicious = []
    clean      = []
    for p in positions:
        wb_price  = p.get("wb_price_fresh") or 0
        opt_price = p.get("sell_opt") or 0
        if wb_price > 0 and opt_price > 0 and opt_price < wb_price * 0.40:
            suspicious.append(p)
        else:
            clean.append(p)
    if suspicious:
        log.info(f"Подозрительных (опт.цена < 40% от WB): {len(suspicious)} → вкладка 'Подозрительные'")
    positions = clean

    # Разделяем по совпадению бренда
    matched        = [p for p in positions if _brands_match(p["sup_brand"], p["wb_brand"])]
    diff_brand_all = [p for p in positions if not _brands_match(p["sup_brand"], p["wb_brand"])]
    diff_brand     = [p for p in diff_brand_all if p.get("score", 0) >= _GOOD_SCORE_THRESHOLD]

    # Сортировка по скору
    matched.sort(key=lambda x: x["score"], reverse=True)
    diff_brand.sort(key=lambda x: x["score"], reverse=True)
    suspicious.sort(key=lambda x: x["score"], reverse=True)

    log.info(f"Совпадение бренда: {len(matched)}  |  Разные бренды (хорошие): {len(diff_brand)} из {len(diff_brand_all)}")

    log.info("Экспортируем Excel...")
    path = export_excel(matched, diff_brand, suspicious)

    log.info("")
    log.info(f"ИТОГО: {len(matched)} позиций в «Топ дефицит»")
    if matched:
        log.info("\nТОП-10:")
        for i, p in enumerate(matched[:10], 1):
            log.info(
                f"  {i:2}. [{p['source']:8s}] OOS={p['oos_pct']:5.1f}%  "
                f"прод={p['sales_30d']:4d}  скор={p['score']:7.2f}  "
                f"WB={p.get('wb_price_fresh',0):>7.0f}₽  "
                f"закуп={p['sup_price']:>7.0f}₽  "
                f"| {p['wb_brand']} {p['wb_name'][:30] or p['sup_article']}"
            )
    log.info(f"\nФайл: {path}")


if __name__ == "__main__":
    main()
