"""
wb_unified_catalog.py
════════════════════════════════════════════════════════════════════════════
Объединяет каталоги Mikado и Автолиги в единый словарь.

Источники:
  - Mikado: полный прайс (код, бренд, название, цена, остаток)
             + OEM-номера из mikado_data.xlsx (где есть, пока 3 позиции)
  - Автолига: полный каталог по OEM-номерам

Ключи словаря:
  - Для позиций с известным OEM → ключ = нормализованный OEM
  - Для позиций Mikado без OEM  → ключ = "MK:<prodnum>"
  - Позиции Автолиги всегда keyed by OEM

Использование:
    from wb_unified_catalog import load_unified_catalog
    catalog = load_unified_catalog()

Запуск напрямую:
    uv run --with openpyxl,xlrd,requests scripts/wb_unified_catalog.py
"""

import sys
import io
import re
import logging
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    import requests
except ImportError:
    print("Установи зависимости: uv run --with openpyxl,requests scripts/wb_unified_catalog.py")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent))
from autoliga_loader import load_autoliga

BASE_DIR       = Path(__file__).parent.parent
SCRAPER_DATA   = BASE_DIR / "data" / "suppliers" / "mikado" / "scraper_output" / "mikado_data.xlsx"
PRICE_FALLBACK = Path("C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx")
MIKADO_PRICE_URL = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=BBE2E029-54CF-4D9E-9FAC-9FE25E85B300"
)

log = logging.getLogger(__name__)


# ─── Нормализация OEM ─────────────────────────────────────────────────────────

def normalize_oem(s: str) -> str:
    return re.sub(r"[\s\-\.]", "", s).upper().strip()


def parse_oem_column(raw: str) -> list[str]:
    """
    Парсит колонку OEM из mikado_data.xlsx.
    Формат: "AUDI: 1K0513029JA; AUDI: 1K0 513 029 HM; ..."
    """
    if not raw:
        return []
    oems = set()
    for part in raw.split(";"):
        part = part.strip()
        oem_raw = part.split(":", 1)[1].strip() if ":" in part else part
        norm = normalize_oem(oem_raw)
        if norm and len(norm) >= 4:
            oems.add(norm)
    return list(oems)


# ─── Загрузка полного прайса Mikado ──────────────────────────────────────────

def _load_mikado_full() -> dict[str, dict]:
    """
    Загружает полный прайс Mikado.
    Возвращает {prodnum: {brand, name, price, stock, code}}
    """
    content = None
    try:
        resp = requests.get(MIKADO_PRICE_URL, timeout=60)
        resp.raise_for_status()
        if resp.content[:2] == b"PK":
            content = resp.content
            log.info(f"Mikado: прайс скачан ({len(content):,} байт)")
    except Exception as e:
        log.warning(f"Mikado: онлайн недоступен ({e}), берём локальный")

    try:
        src = io.BytesIO(content) if content else (
            PRICE_FALLBACK if PRICE_FALLBACK.exists() else None
        )
        if src is None:
            log.error("Mikado: прайс недоступен")
            return {}

        wb  = openpyxl.load_workbook(src, read_only=True, data_only=True)
        ws  = wb.active
        rows = ws.iter_rows(values_only=True)
        hdr  = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]

        def find(names: list[str]) -> int | None:
            for i, h in enumerate(hdr):
                if any(n == h for n in names):
                    return i
            return None

        c_prodnum = find(["prodnum"])
        c_code    = find(["code"])
        c_brand   = find(["brandname"])
        c_name    = find(["prodname"])
        c_price   = find(["priceout"])
        c_qty     = find(["qty"])

        db: dict[str, dict] = {}
        for row in rows:
            def get_str(idx):
                return str(row[idx]).strip() if idx is not None and len(row) > idx and row[idx] else ""
            def get_float(idx):
                try:
                    return float(row[idx]) if idx is not None and len(row) > idx and row[idx] else 0.0
                except Exception:
                    return 0.0

            prodnum = get_str(c_prodnum)
            if not prodnum:
                continue
            price = get_float(c_price)
            if price <= 0:
                continue

            db[prodnum] = {
                "prodnum": prodnum,
                "code":    get_str(c_code),
                "brand":   get_str(c_brand),
                "name":    get_str(c_name),
                "price":   price,
                "stock":   get_float(c_qty),
            }

        wb.close()
        log.info(f"Mikado: загружено {len(db):,} позиций")
        return db

    except Exception as e:
        log.error(f"Mikado: ошибка загрузки прайса — {e}")
        return {}


# ─── Загрузка OEM-данных из scraper_output ───────────────────────────────────

def _load_mikado_oems() -> dict[str, list[str]]:
    """
    Загружает OEM-номера из mikado_data.xlsx (scraper output).
    Возвращает {prodnum: [normalized_oem, ...]}
    Доступно только для позиций, которые были проскрапены.
    """
    if not SCRAPER_DATA.exists():
        return {}

    try:
        wb   = openpyxl.load_workbook(SCRAPER_DATA, read_only=True, data_only=True)
        ws   = wb.active
        rows = ws.iter_rows(values_only=True)
        hdr  = [str(v).strip() if v else "" for v in (next(rows, []) or [])]

        def col(parts: list[str]) -> int | None:
            for i, h in enumerate(hdr):
                if any(p.lower() in h.lower() for p in parts):
                    return i
            return None

        c_code = col(["Код"])
        c_oem  = col(["OEM"])

        result: dict[str, list[str]] = {}
        for row in rows:
            code = str(row[c_code]).strip() if c_code is not None and len(row) > c_code and row[c_code] else ""
            if not code or code.lower() == "none":
                continue
            oems = parse_oem_column(
                str(row[c_oem]) if c_oem is not None and len(row) > c_oem and row[c_oem] else ""
            )
            if oems:
                result[code] = oems

        wb.close()
        log.info(f"Mikado OEM: загружено для {len(result)} позиций из scraper_output")
        return result

    except Exception as e:
        log.warning(f"Mikado OEM: ошибка чтения scraper_output — {e}")
        return {}


# ─── Основная функция ─────────────────────────────────────────────────────────

def load_unified_catalog() -> dict[str, dict]:
    """
    Строит единый каталог Mikado + Автолига.

    Структура каждой записи:
    {
        "oems":        list[str],   # нормализованные OEM-номера (пусто если не известны)
        "mikado_code": str | None,  # prodnum Mikado
        "brand":       str,
        "name":        str,
        "best_price":  float,       # мин. закупочная цена
        "best_source": str,         # "mikado" | "autoliga"
        "in_stock":    bool,
        "suppliers": {
            "mikado":   {"price": float, "stock": float},  # если есть
            "autoliga": {"price": float, "stock": float},  # если есть
        }
    }
    """
    mikado_items = _load_mikado_full()
    mikado_oems  = _load_mikado_oems()   # {prodnum: [oem, ...]} — только скрапнутые
    autoliga     = load_autoliga()

    catalog: dict[str, dict] = {}

    # ── Шаг 1: все позиции Mikado ─────────────────────────────────────────────
    # Строим обратный индекс OEM → prodnum
    oem_to_prodnum: dict[str, str] = {}

    for prodnum, item in mikado_items.items():
        oems = mikado_oems.get(prodnum, [])

        entry = {
            "oems":        oems,
            "mikado_code": prodnum,
            "brand":       item["brand"],
            "name":        item["name"],
            "best_price":  item["price"],
            "best_source": "mikado",
            "in_stock":    item["stock"] > 0,
            "suppliers":   {"mikado": {"price": item["price"], "stock": item["stock"]}},
        }

        if oems:
            # Регистрируем по OEM-номерам
            for oem in oems:
                if oem not in catalog:
                    catalog[oem] = entry
                    oem_to_prodnum[oem] = prodnum
        else:
            # Нет OEM → ключ по внутреннему коду
            key = f"MK:{prodnum}"
            catalog[key] = entry

    # ── Шаг 2: Автолига — матчинг с Mikado или добавление новых позиций ───────
    autoliga_new      = 0
    autoliga_enriched = 0

    for oem, al in autoliga.items():
        al_price    = al["price"]
        al_stock    = al["stock"]
        al_in_stock = al_stock > 0

        if oem in catalog:
            # Mikado уже есть по этому OEM — обогащаем
            entry = catalog[oem]
            entry["suppliers"]["autoliga"] = {"price": al_price, "stock": al_stock}
            entry["in_stock"] = entry["in_stock"] or al_in_stock
            if al_price < entry["best_price"]:
                entry["best_price"]  = al_price
                entry["best_source"] = "autoliga"
            autoliga_enriched += 1
        else:
            # Только у Автолиги
            catalog[oem] = {
                "oems":        [oem],
                "mikado_code": None,
                "brand":       al["brand"],
                "name":        al["name"],
                "best_price":  al_price,
                "best_source": "autoliga",
                "in_stock":    al_in_stock,
                "suppliers":   {"autoliga": {"price": al_price, "stock": al_stock}},
            }
            autoliga_new += 1

    # ── Статистика ────────────────────────────────────────────────────────────
    total       = len(catalog)
    both        = sum(1 for e in catalog.values() if len(e["suppliers"]) == 2)
    in_stock    = sum(1 for e in catalog.values() if e["in_stock"])
    with_oem    = sum(1 for e in catalog.values() if e["oems"])
    mikado_only = sum(1 for e in catalog.values()
                      if "mikado" in e["suppliers"] and "autoliga" not in e["suppliers"])

    log.info(
        f"Единый каталог: {total:,} позиций | "
        f"оба поставщика: {both:,} | "
        f"только Mikado: {mikado_only:,} | "
        f"только Автолига: {autoliga_new:,} | "
        f"в наличии: {in_stock:,} | "
        f"с OEM: {with_oem:,}"
    )
    return catalog


# ─── Быстрая проверка ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    catalog = load_unified_catalog()

    print("\nПримеры с двумя поставщиками:")
    count = 0
    for oem, e in catalog.items():
        if len(e["suppliers"]) == 2:
            mk = e["suppliers"]["mikado"]
            al = e["suppliers"]["autoliga"]
            print(
                f"  {oem}: {e['brand']} | {e['name'][:35]} | "
                f"Mikado {mk['price']:.0f} ₽ / Автолига {al['price']:.0f} ₽ → лучшая {e['best_price']:.0f} ₽"
            )
            count += 1
            if count >= 5:
                break

    print("\nПримеры Mikado без OEM (ключ MK:...):")
    count = 0
    for key, e in catalog.items():
        if key.startswith("MK:") and not e["oems"]:
            print(f"  {key}: {e['brand']} | {e['name'][:40]} | {e['best_price']:.0f} ₽")
            count += 1
            if count >= 5:
                break
