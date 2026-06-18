"""
mikado_match.py
════════════════════════════════════════════════════════════════════════════
Матчит артикулы Микадо (колонка Code) с wb_card_oem (basket OEM).
Формат вывода и метод матчинга идентичны wb_matches_OEM_*.xlsx (Автолига).

Запуск:
  cd C:\\Users\\Admin\\Documents\\Autoparts_Ecommerce
  uv run --with openpyxl scripts/mikado_match.py
  uv run ... mikado_match.py --file "C:\\Users\\Admin\\Downloads\\mikado_price_34 (2).xlsx"
"""

import sys, re, json, sqlite3, argparse
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Установи: uv run --with openpyxl scripts/mikado_match.py"); sys.exit(1)

BASE_DIR   = Path(__file__).parent.parent
DB_PATH    = BASE_DIR / "data" / "analytics" / "wb_index.db"
OUTPUT_DIR = BASE_DIR / "data" / "analytics"
DEFAULT_F  = Path(r"C:\Users\Admin\Downloads\mikado_price_34 (2).xlsx")

_AUTO_KW = (
    "авт", "запч", "тормоз", "амортизатор", "подвеск", "рулев",
    "сцепл", "ремен", "фильтр", "подшипник", "сальник", "датчик",
    "насос", "привод", "генератор", "стартер", "трос", "рычаг",
    "масла", "аккумулятор", "колодк", "прокладк", "шрус", "пружин",
)

# Артикулы короче этого порога требуют совпадения бренда (слишком неуникальны)
_MIN_ART_WITHOUT_BRAND = 10


def _norm(s: str) -> str:
    return re.sub(r"[^A-ZА-ЯЁ0-9]", "", s.upper())


def _norm_brand(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def _brands_match(a: str, b: str) -> bool:
    """Совпадение брендов: точное или один является префиксом другого."""
    a = _norm_brand(a)
    b = _norm_brand(b)
    if not a or not b:
        return False
    return a == b or a.startswith(b) or b.startswith(a)


def load_mikado(path: Path) -> dict:
    """norm_code → {article, brand, price}"""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = [str(h).strip() if h else "" for h in rows[0]]
    try:
        ci = headers.index("Code")
        bi = headers.index("BrandName")
        pi = headers.index("PriceOut")
    except ValueError as e:
        print(f"Колонка не найдена: {e}"); sys.exit(1)

    out: dict = {}
    for r in rows[1:]:
        code  = str(r[ci]).strip() if r[ci] else ""
        brand = str(r[bi]).strip() if r[bi] else ""
        try:
            price = float(r[pi]) if r[pi] else 0.0
        except (ValueError, TypeError):
            price = 0.0
        if not code:
            continue
        n = _norm(code)
        if len(n) < 4:
            continue
        if n not in out:
            out[n] = {"article": code, "brand": brand, "price": price}

    print(f"Микадо: {len(out):,} уникальных артикулов")
    return out


def run_match(conn: sqlite3.Connection, mikado: dict) -> list[tuple]:
    kw_conds = " OR ".join(f"LOWER(p.subject) LIKE '%{kw}%'" for kw in _AUTO_KW)
    rows = conn.execute(
        "SELECT p.nm_id, p.brand, p.price_rub, p.sales_30d, p.subject, c.oem_list "
        "FROM wb_products p JOIN wb_card_oem c ON p.nm_id = c.nm_id "
        f"WHERE c.oem_list != '[]' "
        f"AND (p.subject IS NULL OR p.subject = '' OR {kw_conds})"
    ).fetchall()

    print(f"WB карточек с OEM данными (авто): {len(rows):,}")

    matches: list[tuple] = []
    seen: set[tuple] = set()

    for nm_id, wb_brand, wb_price, sales, subject, oem_json in rows:
        try:
            vc_list = json.loads(oem_json)
        except Exception:
            continue
        for vc in vc_list:
            if not vc:
                continue
            n = _norm(str(vc))
            if len(n) < 4:
                continue
            if n not in mikado:
                continue
            it = mikado[n]
            # Короткий артикул без совпадения бренда — ложное срабатывание
            if len(n) < _MIN_ART_WITHOUT_BRAND and not _brands_match(it["brand"], wb_brand or ""):
                continue
            key = (n, nm_id)
            if key in seen:
                continue
            seen.add(key)
            matches.append((
                it["article"], it["brand"], it["price"],
                nm_id, wb_brand or "", wb_price or 0, sales or 0,
                subject or "", str(vc),
                f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
            ))

    matches.sort(key=lambda x: x[6], reverse=True)
    print(f"Совпадений: {len(matches):,}")
    return matches


def save_excel(matches: list[tuple], total_mikado: int) -> Path:
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="D6E4F0")

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "Матчи"

    headers = [
        "Артикул Mikado", "Бренд Mikado", "Цена Mikado (руб)",
        "nm_id WB", "Бренд WB", "Цена WB (руб)",
        "Продажи/30д", "Категория WB", "Артикул WB", "Ссылка WB",
    ]
    col_widths = [20, 16, 17, 12, 16, 14, 13, 50, 20, 18]

    ws.append(headers)
    for ci, cell in enumerate(ws[1], 1):
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, m in enumerate(matches, 1):
        (mk_art, mk_brand, mk_price,
         nm_id, wb_brand, wb_price, sales,
         subject, wb_art, url) = m

        ws.append([mk_art, mk_brand, mk_price,
                   nm_id, wb_brand, wb_price, sales,
                   subject, wb_art, url])

        row_idx = i + 1
        link_cell = ws.cell(row=row_idx, column=10)
        link_cell.hyperlink = url
        link_cell.font = Font(color="0563C1", underline="single", size=10)

        if i % 2 == 0:
            for ci in range(1, 10):
                ws.cell(row=row_idx, column=ci).fill = ALT_FILL

    # Лист Инфо
    ws_info = wb_out.create_sheet("Инфо")
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_info.append(["Сформировано", ts])
    ws_info.append(["Артикулов Микадо", f"{total_mikado:,}"])
    ws_info.append(["Совпадений", f"{len(matches):,}"])
    ws_info.column_dimensions["A"].width = 22
    ws_info.column_dimensions["B"].width = 18

    ts_fn  = datetime.now().strftime("%Y%m%d_%H%M")
    path   = OUTPUT_DIR / f"wb_matches_MIKADO_{ts_fn}.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb_out.save(str(path))
    print(f"Сохранено: {path}")
    return path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default=str(DEFAULT_F))
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(f"Файл не найден: {path}"); sys.exit(1)

    print("Загрузка прайса Микадо...")
    mikado = load_mikado(path)

    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")

    print("Матчинг с wb_card_oem...")
    matches = run_match(conn, mikado)
    conn.close()

    save_excel(matches, len(mikado))


if __name__ == "__main__":
    main()
