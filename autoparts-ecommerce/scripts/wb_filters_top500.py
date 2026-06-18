"""
wb_filters_top500.py — ТОП-500 ФИЛЬТРОВ для WB: Автолига + Микадо
==================================================================
Цель: выбрать 500 фильтров с максимальной ожидаемой прибылью.
Ранжирование: маржа_руб × log(продажи бренда) — больше зарабатываем первыми.
Источник: 7 512 уникальных фильтров из двух прайсов (после дедупликации).
"""

import xlrd, openpyxl, json, os, re, math
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AUTOLIGA_FILE = "data/suppliers/autoliga/PriceALVLG0411.xls"
MIKADO_FILE   = "data/suppliers/mikado/mikado_price_live.xlsx"
CACHE_DIR     = "data/analytics/cache"
OUTPUT_FILE   = "data/analytics/wb_filters_top500.xlsx"

# WB FBS: Автозапчасти, май 2026
WB_COMMISSION  = 0.25
WB_ACQUIRING   = 0.015
WB_RETURN_RATE = 0.03
TAX_RATE       = 0.06
WB_LOGISTICS   = 75
MIN_MARGIN_PCT = 15
TARGET_MARGIN  = 25
MAX_PER_BRAND  = 20   # лимит на бренд из 500 фильтров

AUTO_KEYS = [
    "авто", "запчаст", "шин", "диск", "тормоз", "подвеск", "двигател",
    "кузов", "электрик", "трансмисс", "масл", "фильтр", "сцеплени",
    "рулев", "охлажден", "технич", "мотоцикл", "выхлоп", "топлив",
    "аккумулят", "зажиган", "амортизатор",
]

# Популярные авто в названиях фильтров → подсказка по применяемости
CAR_HINTS = {
    "lada": "LADA", "ваз": "LADA", "granta": "Granta", "гранта": "Granta",
    "vesta": "Vesta", "веста": "Vesta", "niva": "Niva", "нива": "Niva",
    "priora": "Priora", "приора": "Priora", "kalina": "Kalina",
    "rio": "Kia Rio", "solaris": "Hyundai Solaris", "солярис": "Hyundai Solaris",
    "creta": "Hyundai Creta", "крета": "Hyundai Creta",
    "logan": "Renault Logan", "логан": "Renault Logan",
    "focus": "Ford Focus", "фокус": "Ford Focus",
    "polo": "VW Polo", "поло": "VW Polo",
    "golf": "VW Golf", "passat": "VW Passat",
    "octavia": "Skoda Octavia", "rapid": "Skoda Rapid",
    "e46": "BMW E46", "e39": "BMW E39", "e60": "BMW E60",
    "w211": "Merc W211", "w203": "Merc W203",
    "camry": "Toyota Camry", "corolla": "Toyota Corolla",
    "almera": "Nissan Almera", "x-trail": "Nissan X-Trail",
    "lacetti": "Chevrolet Lacetti", "cruze": "Chevrolet Cruze",
    "sportage": "Kia Sportage", "tucson": "Hyundai Tucson",
    "haval": "Haval", "geely": "Geely", "chery": "Chery",
    "tiggo": "Chery Tiggo", "jolion": "Haval Jolion",
    "largus": "Lada Largus", "ларгус": "Lada Largus",
    "duster": "Renault Duster", "дастер": "Renault Duster",
    "outlander": "Mitsubishi Outlander",
    "peugeot": "Peugeot", "citroen": "Citroen",
}


# ── Утилиты ───────────────────────────────────────────────────────────────────

def norm(s):
    return re.sub(r'[\s\-\.\/\\]', '', str(s)).upper()

def is_auto(subj):
    return any(k in subj.lower() for k in AUTO_KEYS)

def is_filter(name):
    n = name.lower()
    return 'фильтр' in n or 'filter' in n

def filter_type(name):
    n = name.lower()
    if 'масл' in n or ('oil' in n and 'filter' in n):  return 'масляный'
    if 'воздуш' in n or 'air' in n:                    return 'воздушный'
    if 'салон' in n or 'cabin' in n or 'cabin' in n:   return 'салонный'
    if 'топлив' in n or 'fuel' in n:                   return 'топливный'
    if 'акпп' in n or 'трансм' in n or 'gearbox' in n: return 'АКПП'
    return 'прочий'

def calc_margin(buy, sell):
    fees = sell * (WB_COMMISSION + WB_ACQUIRING + WB_RETURN_RATE)
    tax  = sell * TAX_RATE
    net  = sell - fees - WB_LOGISTICS - tax - buy
    pct  = (net / sell * 100) if sell > 0 else 0
    return round(net, 2), round(pct, 1)

def min_sell(buy, target_pct):
    coeff = (1 - WB_COMMISSION - WB_ACQUIRING - WB_RETURN_RATE - TAX_RATE) - target_pct / 100
    return (WB_LOGISTICS + buy) / coeff if coeff > 0 else 1e9

def get_cars(name):
    n = name.lower()
    found = []
    for kw, label in CAR_HINTS.items():
        if kw in n and label not in found:
            found.append(label)
    return ', '.join(found[:4]) if found else ''


# ── Загрузка прайсов ──────────────────────────────────────────────────────────

def load_autoliga():
    book = xlrd.open_workbook(AUTOLIGA_FILE, encoding_override='cp1251')
    ws   = book.sheet_by_index(0)
    items = []
    for r in range(9, ws.nrows):
        try:
            brand   = str(ws.cell_value(r, 1)).strip()
            article = str(ws.cell_value(r, 2)).strip()
            name    = str(ws.cell_value(r, 4)).strip()
            stock   = float(str(ws.cell_value(r, 6)).replace(',', '.') or 0)
            price   = float(str(ws.cell_value(r, 7)).replace(',', '.') or 0)
        except:
            continue
        if not brand or not article or stock <= 0 or price <= 0 or not is_filter(name):
            continue
        items.append({'brand': brand, 'article': article, 'name': name,
                      'stock': int(stock), 'buy': price, 'source': 'Автолига'})
    return items


def load_mikado():
    wb = openpyxl.load_workbook(MIKADO_FILE, read_only=True, data_only=True)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            article = str(row[1]).strip() if row[1] else ''
            brand   = str(row[2]).strip() if row[2] else ''
            name    = str(row[3]).strip() if row[3] else ''
            price   = float(str(row[4]).replace(',', '.')) if row[4] else 0.0
            qty_raw = str(row[5]).strip() if row[5] else '0'
            stock   = 10 if qty_raw.startswith('>') else int(float(qty_raw))
        except:
            continue
        if not brand or not article or stock <= 0 or price <= 0 or not is_filter(name):
            continue
        items.append({'brand': brand, 'article': article, 'name': name,
                      'stock': stock, 'buy': price, 'source': 'Микадо'})
    wb.close()
    return items


def merge(al_items, mk_items):
    index = {}
    for it in al_items:
        key = (norm(it['brand']), norm(it['article']))
        if key not in index or it['buy'] < index[key]['buy']:
            index[key] = it.copy()
    for it in mk_items:
        key = (norm(it['brand']), norm(it['article']))
        if key in index:
            if it['buy'] < index[key]['buy']:
                src = 'Авт+Мик'
                merged = it.copy()
                merged['source'] = src
                index[key] = merged
            else:
                index[key]['source'] = 'Авт+Мик'
        else:
            index[key] = it.copy()
    return list(index.values())


# ── MPStats: статистика по брендам ───────────────────────────────────────────

def build_brand_stats():
    brand_data = defaultdict(list)
    for fn in os.listdir(CACHE_DIR):
        if not fn.startswith('mpsb_'):
            continue
        path = os.path.join(CACHE_DIR, fn)
        with open(path, encoding='utf-8') as f:
            try:   data = json.load(f)
            except: continue
        for p in data:
            if not is_auto(p.get('subject', '')):
                continue
            b = p.get('brand', '').upper().strip()
            if b:
                brand_data[b].append({
                    'sales': float(p.get('sales_30d', 0) or 0),
                    'oos':   float(p.get('oos_pct', 0) or 0),
                    'price': float(p.get('price', 0) or 0),
                })
    stats = {}
    for b, prods in brand_data.items():
        sales = [p['sales'] for p in prods]
        oos   = [p['oos']   for p in prods]
        prices= [p['price'] for p in prods if p['price'] > 0]
        stats[b] = {
            'avg_sales':    sum(sales) / len(sales) if sales else 0,
            'avg_oos':      sum(oos) / len(oos) if oos else 0,
            'wb_price_avg': sum(prices) / len(prices) if prices else 0,
            'product_count': len(prods),
        }
    return stats


# ── Скоринг: максимизация прибыли ────────────────────────────────────────────

def score_item(item, bstats):
    buy      = item['buy']
    brand_up = item['brand'].upper()
    bs       = bstats.get(brand_up, {})

    wb_avg  = bs.get('wb_price_avg', 0)
    p_min15 = min_sell(buy, MIN_MARGIN_PCT)
    p_rec25 = min_sell(buy, TARGET_MARGIN)

    # Цена продажи: если есть данные WB — ставим чуть ниже рынка, но не ниже 15%
    if wb_avg >= p_min15 and wb_avg > 0:
        sell = min(wb_avg * 0.95, p_rec25)
        sell = max(sell, p_min15)
    else:
        sell = p_rec25

    mg_rub, mg_pct = calc_margin(buy, sell)
    if mg_pct < MIN_MARGIN_PCT:
        return {}

    avg_sales = bs.get('avg_sales', 0)
    avg_oos   = bs.get('avg_oos', 0)
    wb_cnt    = bs.get('product_count', 0)

    # Ожидаемая прибыль в месяц: маржа × спрос
    # Нет данных о конкретных продажах фильтра, но знаем бренд в среднем
    # demand_factor = log(продажи бренда+1) → чем больше продаётся бренд, тем лучше
    demand_factor = math.log1p(avg_sales) if avg_sales > 0 else 0.5

    # Дефицит: чем выше OOS у конкурентов, тем больше наш потенциал
    oos_boost = 1.0 + (avg_oos / 100) * 1.5   # 1.0–2.5

    # Голубой океан: меньше конкурентов = легче продать
    if   wb_cnt == 0:  ocean = 1.8
    elif wb_cnt < 10:  ocean = 1.4
    elif wb_cnt < 30:  ocean = 1.2
    else:              ocean = max(1.0 / math.sqrt(wb_cnt / 10), 0.5)

    # Итоговый скор: ожидаемая прибыль с учётом спроса и конкуренции
    profit_score = mg_rub * demand_factor * oos_boost * ocean

    # Ожидаемая прибыль/мес (оценка): маржа × ожидаемые продажи
    # Осторожная оценка: берём 5% от продаж бренда на 1 артикул
    est_sales_per_sku = max(avg_sales * 0.05, 1.0)
    est_monthly_profit = mg_rub * est_sales_per_sku

    # Обоснование
    parts = []
    if avg_sales > 0: parts.append(f"Продажи бренда WB: {avg_sales:.0f} шт/мес")
    if avg_oos > 0:   parts.append(f"OOS конкурентов: {avg_oos:.0f}%")
    if wb_avg > 0:    parts.append(f"Ср.цена WB: {wb_avg:.0f} руб")
    parts.append(f"Маржа: {mg_pct:.1f}% = {mg_rub:.0f} руб/шт")
    parts.append(f"Оцен.прибыль/мес: ~{est_monthly_profit:.0f} руб")
    if wb_cnt > 0:    parts.append(f"Конкурентов WB: {wb_cnt}")

    return {
        'score':         round(profit_score, 4),
        'sell':          round(sell),
        'mg_pct':        mg_pct,
        'mg_rub':        mg_rub,
        'avg_sales':     round(avg_sales, 1),
        'oos':           round(avg_oos, 1),
        'wb_cnt':        wb_cnt,
        'est_profit_mo': round(est_monthly_profit),
        'reason':        '; '.join(parts),
    }


# ── Excel экспорт ─────────────────────────────────────────────────────────────

C_HDR  = "1F4E79"
C_GOLD = "FFF2CC"
C_SILV = "EBF3FB"
C_BRNZ = "FCE4D6"
C_AL   = "E3F2FD"   # голубой — только Автолига
C_MK   = "E8F5E9"   # зелёный — только Микадо
C_BOTH = "FFF9C4"   # жёлтый — оба поставщика

THIN = Border(
    left  =Side(style='thin', color='D0D0D0'),
    right =Side(style='thin', color='D0D0D0'),
    top   =Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

COLS = [
    ('#',                    5),
    ('Артикул',             25),
    ('Бренд',               16),
    ('Тип фильтра',         14),
    ('Название',            50),
    ('Источник',            12),
    ('Применяемость',       28),
    ('Закупка, руб',        13),
    ('Продажа WB, руб',     15),
    ('Маржа, %',            10),
    ('Маржа, руб/шт',       13),
    ('Оцен.прибыль/мес',    16),
    ('Остаток',              9),
    ('Продажи бренда/мес',  18),
    ('OOS конкур., %',      14),
    ('Конкурентов WB',      14),
    ('Подробное обоснование', 80),
]


def _cell(ws, r, c, v, bold=False, bg=None, wrap=False, align='left', fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    fc = 'FFFFFF' if bg == C_HDR else '000000'
    cell.font      = Font(bold=bold, size=10, color=fc)
    cell.border    = THIN
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    if fmt:
        cell.number_format = fmt
    return cell


def export(results, pool_size, both_count):
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb = Workbook()

    # ── Лист 1: Топ-500 фильтров ──────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Топ-500 фильтров'
    ws.freeze_panes = 'B2'

    for col, (h, w) in enumerate(COLS, 1):
        _cell(ws, 1, col, h, bold=True, bg=C_HDR, align='center', wrap=True)
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w
    ws.row_dimensions[1].height = 40

    FMTS = [None, None, None, None, None, None, None,
            '#,##0', '#,##0', '0.0', '#,##0', '#,##0',
            '#,##0', '#,##0.0', '0.0', '#,##0', None]

    for i, it in enumerate(results[:500], 2):
        rank = i - 1
        src  = it.get('source', '')
        if rank <= 50:
            bg = C_GOLD
        elif rank <= 150:
            bg = C_SILV
        elif rank <= 300:
            bg = C_BRNZ
        else:
            bg = C_BOTH if src == 'Авт+Мик' else (C_MK if src == 'Микадо' else C_AL)

        row = [
            rank,
            it['article'],
            it['brand'],
            filter_type(it['name']),
            it['name'],
            src,
            get_cars(it['name']),
            it['buy'],
            it.get('sell', ''),
            it.get('mg_pct', ''),
            it.get('mg_rub', ''),
            it.get('est_profit_mo', ''),
            it['stock'],
            it.get('avg_sales', ''),
            it.get('oos', ''),
            it.get('wb_cnt', ''),
            it.get('reason', ''),
        ]
        for col, (val, fmt) in enumerate(zip(row, FMTS), 1):
            _cell(ws, i, col, val, bg=bg, wrap=(col == len(COLS)), fmt=fmt)
        ws.row_dimensions[i].height = 30

    # ── Лист 2: По типам фильтров ─────────────────────────────────────────────
    ws2 = wb.create_sheet('По типам')
    type_data = defaultdict(lambda: {'cnt': 0, 'mg_rub': [], 'profit': [], 'src': defaultdict(int)})
    for it in results[:500]:
        ft = filter_type(it['name'])
        type_data[ft]['cnt']     += 1
        type_data[ft]['mg_rub'].append(it.get('mg_rub', 0))
        type_data[ft]['profit'].append(it.get('est_profit_mo', 0))
        type_data[ft]['src'][it.get('source', '')] += 1

    _cell(ws2, 1, 1, 'Тип фильтра', bold=True, bg=C_HDR)
    _cell(ws2, 1, 2, 'Кол-во в топ-500', bold=True, bg=C_HDR, align='center')
    _cell(ws2, 1, 3, 'Ср.маржа руб/шт', bold=True, bg=C_HDR, align='center')
    _cell(ws2, 1, 4, 'Оцен.прибыль/мес итого', bold=True, bg=C_HDR, align='center')
    _cell(ws2, 1, 5, 'Автолига', bold=True, bg=C_HDR, align='center')
    _cell(ws2, 1, 6, 'Микадо', bold=True, bg=C_HDR, align='center')
    _cell(ws2, 1, 7, 'Оба', bold=True, bg=C_HDR, align='center')
    ws2.column_dimensions['A'].width = 18
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws2.column_dimensions[col].width = 20

    for r, (ft, d) in enumerate(sorted(type_data.items(), key=lambda x: -x[1]['cnt']), 2):
        avg_mg = sum(d['mg_rub']) / len(d['mg_rub']) if d['mg_rub'] else 0
        tot_pr = sum(d['profit'])
        ws2.cell(r, 1).value = ft
        ws2.cell(r, 2).value = d['cnt']
        ws2.cell(r, 3).value = round(avg_mg)
        ws2.cell(r, 4).value = round(tot_pr)
        ws2.cell(r, 5).value = d['src'].get('Автолига', 0)
        ws2.cell(r, 6).value = d['src'].get('Микадо', 0)
        ws2.cell(r, 7).value = d['src'].get('Авт+Мик', 0)
    # Итого
    total_row = ws2.max_row + 2
    ws2.cell(total_row, 1).value = 'ИТОГО'
    ws2.cell(total_row, 1).font = Font(bold=True)
    ws2.cell(total_row, 2).value = 500
    all_pr = sum(it.get('est_profit_mo', 0) for it in results[:500])
    ws2.cell(total_row, 4).value = round(all_pr)
    ws2.cell(total_row, 4).font = Font(bold=True)
    ws2.cell(total_row, 4).number_format = '#,##0'

    # ── Лист 3: По брендам ────────────────────────────────────────────────────
    ws3 = wb.create_sheet('По брендам')
    bd = defaultdict(lambda: {'cnt': 0, 'mg': [], 'profit': [], 'src': set(), 'types': defaultdict(int)})
    for it in results[:500]:
        b = it['brand']
        bd[b]['cnt']    += 1
        bd[b]['mg'].append(it.get('mg_rub', 0))
        bd[b]['profit'].append(it.get('est_profit_mo', 0))
        bd[b]['src'].add(it.get('source', ''))
        bd[b]['types'][filter_type(it['name'])] += 1

    ws3.append(['Бренд', 'Кол-во', 'Источник', 'Ср.маржа руб/шт',
                'Оцен.прибыль/мес', 'Типы фильтров'])
    ws3.row_dimensions[1].height = 20
    for b, d in sorted(bd.items(), key=lambda x: -sum(x[1]['profit'])):
        avg_mg  = sum(d['mg']) / len(d['mg']) if d['mg'] else 0
        tot_pr  = sum(d['profit'])
        types_s = ', '.join(f"{t}:{n}" for t, n in sorted(d['types'].items(), key=lambda x: -x[1]))
        ws3.append([b, d['cnt'], '+'.join(sorted(d['src'])),
                    round(avg_mg), round(tot_pr), types_s])
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws3.column_dimensions[col].width = 22

    # ── Лист 4: Источники ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Источники')
    src_cnt = defaultdict(int)
    for it in results[:500]:
        src_cnt[it.get('source', '?')] += 1
    ws4.append(['Источник', 'Позиций в топ-500', '% от топ-500'])
    for src, n in sorted(src_cnt.items(), key=lambda x: -x[1]):
        ws4.append([src, n, round(n / 5, 1)])
    ws4.append([])
    ws4.append(['Всего в пуле Автолига:', 2913, ''])
    ws4.append(['Всего в пуле Микадо:', 4688, ''])
    ws4.append(['Совпадают (brand+art):', both_count, ''])
    ws4.append(['Объединённый пул:', pool_size, ''])
    for col in ['A', 'B', 'C']:
        ws4.column_dimensions[col].width = 28

    # ── Лист 5: Методология ───────────────────────────────────────────────────
    ws5 = wb.create_sheet('Методология')
    notes = [
        ('ЦЕЛЬ', 'Выбрать 500 фильтров с максимальной ожидаемой прибылью'),
        ('', ''),
        ('ФОРМУЛА СКОРА', ''),
        ('profit_score', 'маржа_руб x log(продажи_бренда+1) x oos_boost x ocean'),
        ('oos_boost',    '1 + avg_oos/100 x 1.5  (чем выше дефицит, тем лучше наши шансы)'),
        ('ocean',        '0 WB->1.8 / <10->1.4 / <30->1.2 / >30->1/sqrt(cnt/10)'),
        ('', ''),
        ('ОЦЕНКА ПРИБЫЛИ/МЕС', ''),
        ('Формула',      'маржа_руб x (продажи_бренда x 0.05)'),
        ('Коэфф. 0.05',  'осторожная оценка: 1 артикул = 5% от суммарных продаж бренда'),
        ('', ''),
        ('МАРЖА WB FBS', ''),
        ('Комиссия',     '25%'),
        ('Эквайринг',    '1.5%'),
        ('Возвраты',     '3%'),
        ('УСН',          '6%'),
        ('Логистика',    '75 руб.'),
        ('Мин. маржа',   '>= 15%'),
        ('', ''),
        ('ЦВЕТА', ''),
        ('Золото (1-50)',    'Абсолютный топ — заходить в первую очередь'),
        ('Серебро (51-150)', 'Высокий приоритет'),
        ('Бронза (151-300)', 'Средний приоритет'),
        ('Оба поставщика',  'Есть у обоих — можно выбирать лучшую поставку'),
        ('Только Микадо',   'Зелёный — только из Микадо'),
        ('Только Автолига', 'Голубой — только из Автолиги'),
        ('', ''),
        ('ДАННЫЕ MPSTATS', 'Кэш 683 брендов, апр-май 2026. Продажи — среднее по бренду, не по артикулу.'),
    ]
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 75
    for r_data in notes:
        row_idx = ws5.max_row + 1
        ws5.append(list(r_data))
        if r_data[1] == '':
            ws5.cell(row_idx, 1).font = Font(bold=True, size=11)

    wb.save(OUTPUT_FILE)
    return os.path.abspath(OUTPUT_FILE)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print('=' * 65)
    print('  WB Filters Top-500  |  Avtoliga + Mikado  |  Max Profit')
    print('=' * 65)

    print('\n[1/4] Loading filters from Avtoliga...')
    al = load_autoliga()
    print(f'  {len(al):,} filters')

    print('[2/4] Loading filters from Mikado...')
    mk = load_mikado()
    print(f'  {len(mk):,} filters')

    print('[3/4] Merging...')
    pool = merge(al, mk)
    both = sum(1 for it in pool if it['source'] == 'Авт+Мик')
    print(f'  Pool: {len(pool):,} | Both suppliers: {both} | AL only: {sum(1 for it in pool if it["source"]=="Автолига")} | MK only: {sum(1 for it in pool if it["source"]=="Микадо")}')

    print('[4/4] Scoring and ranking...')
    bstats = build_brand_stats()

    scored, skipped = [], 0
    for it in pool:
        s = score_item(it, bstats)
        if not s:
            skipped += 1
            continue
        it.update(s)
        scored.append(it)

    scored.sort(key=lambda x: x['score'], reverse=True)

    brand_cnt = defaultdict(int)
    final = []
    for it in scored:
        b = it['brand']
        if brand_cnt[b] < MAX_PER_BRAND:
            brand_cnt[b] += 1
            final.append(it)

    print(f'  Scored: {len(scored):,} | Skipped (margin<15%): {skipped} | After brand cap: {len(final):,}')

    path = export(final, len(pool), both)
    sz   = os.path.getsize(OUTPUT_FILE) // 1024
    print(f'\n  Saved: {path}  ({sz} KB)')

    # Итоги
    total_profit = sum(it.get('est_profit_mo', 0) for it in final[:500])
    avg_margin   = sum(it.get('mg_pct', 0) for it in final[:500]) / min(len(final), 500)

    from collections import Counter
    type_cnt = Counter(filter_type(it['name']) for it in final[:500])
    src_cnt  = Counter(it.get('source', '') for it in final[:500])
    brand_top = Counter(it['brand'] for it in final[:500]).most_common(10)

    print('\n' + '=' * 65)
    print(f'  Top-500 фильтров:')
    print(f'  Оцен.прибыль/мес (все 500): ~{total_profit:,.0f} руб')
    print(f'  Средняя маржа:               {avg_margin:.1f}%')
    print()
    print('  По типам:')
    for ft, n in type_cnt.most_common():
        try:
            print(f'    {ft:<18} {n:>3}')
        except UnicodeEncodeError:
            print(f'    [?]                {n:>3}')
    print()
    print('  По источникам:')
    for src, n in src_cnt.most_common():
        try:
            print(f'    {src:<16} {n:>3}')
        except UnicodeEncodeError:
            print(f'    [?]              {n:>3}')
    print()
    print('  Топ-10 брендов:')
    for b, n in brand_top:
        try:
            print(f'    {b:<20} {n:>3} шт')
        except UnicodeEncodeError:
            print(f'    [?]                  {n:>3} шт')


if __name__ == '__main__':
    main()
