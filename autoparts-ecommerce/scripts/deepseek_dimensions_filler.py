"""
deepseek_dimensions_filler.py
══════════════════════════════
Добавляет столбцы «Вес, г», «Длина, мм», «Ширина, мм», «Высота, мм»
после столбца «Описание» — через браузер chat.deepseek.com, без API ключа.

Исходная папка:    C:\\Users\\Admin\\Desktop\\На сортировку 26.04\\
Папка результатов: C:\\Users\\Admin\\Desktop\\На сортировку 08.05\\

Установка браузера (один раз):
  uv run --with playwright python -m playwright install chromium

Первый запуск — войти в DeepSeek:
  uv run --with playwright,openpyxl scripts/deepseek_dimensions_filler.py --login

Обработать все файлы:
  uv run --with playwright,openpyxl scripts/deepseek_dimensions_filler.py

Один файл для теста:
  uv run --with playwright,openpyxl scripts/deepseek_dimensions_filler.py --file "Колодки тормозные.xlsx" --rows 2-30

Продолжение после прерывания:
  Просто запустить снова — пропустит строки, которые уже заполнены.
"""

import re
import sys
import json
import time
import shutil
import argparse
import traceback
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Playwright не установлен.")
    print("Шаг 1: uv run --with playwright python -m playwright install chromium")
    print("Шаг 2: uv run --with playwright,openpyxl scripts/deepseek_dimensions_filler.py --login")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("openpyxl не установлен.")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

SRC_DIR      = Path(r"C:\Users\Admin\Desktop\На сортировку 26.04")
OUT_DIR      = Path(r"C:\Users\Admin\Desktop\На сортировку 08.05")
PROFILE_DIR  = Path.home() / ".deepseek_playwright"
DEEPSEEK_URL = "https://chat.deepseek.com"

BATCH_SIZE = 30   # строк за один запрос
SAVE_EVERY = 5    # сохранять Excel каждые N батчей (~150 строк)

NEW_COLS  = ["Вес, г", "Длина, мм", "Ширина, мм", "Высота, мм"]
COL_AFTER = "Описание"

# Признаки ошибки/лимита в тексте ответа
ERROR_PHRASES = [
    "something went wrong",
    "try again",
    "server error",
    "error occurred",
    "service unavailable",
]


# ══════════════════════════════════════════════════════════════════════════════
#  ПРОМПТ
# ══════════════════════════════════════════════════════════════════════════════

INSTRUCTION = (
    "Ты эксперт по автозапчастям. Для каждой позиции ниже укажи "
    "вес и габариты ТРАНСПОРТНОЙ УПАКОВКИ — то есть внешние размеры коробки или пакета "
    "вместе с самой деталью внутри, именно так, как товар будет отправлен покупателю.\n"
    "ВАЖНО: это НЕ размеры самой детали, а размеры упаковки (коробки/пакета) снаружи.\n"
    "Правила:\n"
    "- Если деталь идёт комплектом (2 шт. и т.п.) — вес/размер всего комплекта в упаковке\n"
    "- Опирайся на типичные размеры упаковки для данного вида запчасти\n"
    "- Верни ТОЛЬКО JSON объект, без пояснений, без markdown блоков:\n"
    '{"items": [{"idx": 0, "вес": 500, "длина": 200, "ширина": 150, "высота": 80}, ...]}\n'
    "вес — граммы, длина/ширина/высота — мм, всё целые числа.\n\n"
)


def build_prompt(items: list[dict]) -> str:
    lines = [INSTRUCTION]
    for item in items:
        lines.append(f"[{item['idx']}]")
        lines.append(f"Наименование: {item['name']}")
        if item.get("params"):
            lines.append(f"Параметры: {str(item['params'])[:200]}")
        if item.get("desc"):
            lines.append(f"Описание: {str(item['desc'])[:250]}")
        lines.append("")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  ПАРСИНГ ОТВЕТА
# ══════════════════════════════════════════════════════════════════════════════

def extract_json_items(text: str) -> list[dict]:
    clean = re.sub(r"```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    clean = clean.replace("```", "").strip()

    ITEM_KEYS = {"вес", "длина", "ширина", "высота", "idx"}

    def is_valid(lst):
        return bool(lst) and isinstance(lst[0], dict) and bool(ITEM_KEYS & lst[0].keys())

    def try_parse(s):
        try:
            parsed = json.loads(s.strip())
            if isinstance(parsed, dict) and "items" in parsed:
                items = parsed["items"]
                if isinstance(items, list) and is_valid(items):
                    return items
            if isinstance(parsed, list) and is_valid(parsed):
                return parsed
        except Exception:
            pass
        return None

    result = try_parse(clean)
    if result:
        return result

    for start in [i for i, c in enumerate(clean) if c == '{']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == '{':
                depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result:
                        return result
                    break

    for start in [i for i, c in enumerate(clean) if c == '[']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == '[':
                depth += 1
            elif c == ']':
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result:
                        return result
                    break

    return []


# ══════════════════════════════════════════════════════════════════════════════
#  БРАУЗЕР — ХЕЛПЕРЫ
# ══════════════════════════════════════════════════════════════════════════════

def find_textarea(page):
    """Ищет поле ввода DeepSeek."""
    selectors = [
        'textarea#chat-input',
        'textarea[placeholder]',
        'textarea',
        '[contenteditable="true"]',
    ]
    for sel in selectors:
        try:
            el = page.wait_for_selector(sel, timeout=3000, state="visible")
            if el:
                return page.locator(sel).first
        except Exception:
            continue
    return None


def type_into_textarea(page, text: str):
    """Вставляет текст в React-textarea DeepSeek через нативный сеттер."""
    el = find_textarea(page)
    if el is None:
        # Диагностика
        info = page.evaluate("""() => ({
            url: location.href,
            textareas: document.querySelectorAll('textarea').length,
            contenteditable: document.querySelectorAll('[contenteditable]').length,
        })""")
        raise RuntimeError(f"Поле ввода не найдено. Диагностика: {info}")

    el.click()
    time.sleep(0.3)

    # React-совместимый способ: нативный HTMLTextAreaElement setter
    ok = page.evaluate("""(text) => {
        const el = document.querySelector('textarea#chat-input')
                || document.querySelector('textarea');
        if (!el) return false;
        const setter = Object.getOwnPropertyDescriptor(
            window.HTMLTextAreaElement.prototype, 'value'
        ).set;
        setter.call(el, text);
        el.dispatchEvent(new Event('input',  { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        return true;
    }""", text)

    if not ok:
        # Запасной вариант: посимвольный ввод (медленно, но надёжно)
        el.fill("")
        el.type(text[:3000], delay=5)  # DeepSeek режет слишком длинные промпты

    time.sleep(0.5)


def click_send(page) -> bool:
    """Нажимает кнопку отправки."""
    btn_selectors = [
        'button[aria-label*="Send" i]',
        'button[aria-label*="Отправить" i]',
        '#send-button',
        'button[type="submit"]',
        # DeepSeek: иконка-стрелка в правом нижнем углу textarea
        'div.input-area button:last-child',
        'form button:last-child',
        'textarea ~ button',
        'textarea + button',
    ]
    for _ in range(8):
        for sel in btn_selectors:
            try:
                btn = page.locator(sel).last
                if btn.is_visible(timeout=400) and btn.is_enabled(timeout=400):
                    btn.click()
                    return True
            except Exception:
                continue
        time.sleep(0.5)

    # Fallback: Enter из поля ввода (работает в DeepSeek)
    try:
        el = find_textarea(page)
        if el:
            el.press("Enter")
            return True
    except Exception:
        pass

    page.keyboard.press("Enter")
    return False


def is_generating(page) -> bool:
    """Проверяет, идёт ли ещё генерация ответа."""
    stop_selectors = [
        'button[aria-label*="Stop" i]',
        'button[aria-label*="Стоп" i]',
        'button[aria-label*="stop" i]',
        '[data-testid="stop-button"]',
    ]
    for sel in stop_selectors:
        try:
            if page.locator(sel).first.is_visible(timeout=300):
                return True
        except Exception:
            continue

    # Проверяем через JS наличие анимации/спиннера
    try:
        generating = page.evaluate("""() =>
            !!document.querySelector('.loading, .generating, .spinner, [class*="loading"], [class*="generat"]')
        """)
        return bool(generating)
    except Exception:
        return False


def get_last_response(page) -> str:
    """Возвращает текст последнего ответа DeepSeek."""
    selectors = [
        # DeepSeek-специфичные
        '.ds-markdown',
        '[class*="markdown"]',
        # Общие для чат-интерфейсов
        '[data-message-author-role="assistant"]',
        '[data-role="assistant"]',
        '.assistant-message',
        '.message.assistant',
        # Широкий fallback: последний блок с контентом
        'main .prose',
        'main p:last-of-type',
    ]
    for sel in selectors:
        try:
            els = page.locator(sel).all()
            if not els:
                continue
            text = els[-1].inner_text(timeout=2000).strip()
            if len(text) > 10 and not text.startswith(INSTRUCTION[:30]):
                return text
        except Exception:
            continue
    return ""


def wait_for_response(page, timeout_sec: int = 180) -> str:
    """Ждёт стабилизации ответа DeepSeek."""
    print(" ожидание", end="", flush=True)

    # Ждём появления хоть какого-то текста (до 30 сек)
    for _ in range(20):
        if get_last_response(page):
            break
        time.sleep(1.5)
        print(".", end="", flush=True)

    prev, stable = "", 0
    for tick in range(timeout_sec // 2):
        cur = get_last_response(page)
        if cur and cur == prev:
            stable += 1
            if stable >= 3 and not is_generating(page):
                print(f" готово ({len(cur)} симв.)")
                return cur
        else:
            stable = 0
            if tick % 4 == 0:
                print(".", end="", flush=True)
        prev = cur
        time.sleep(2)

    print(f" таймаут ({len(prev)} симв.)")
    return prev


def open_new_chat(page):
    """Открывает новый пустой чат DeepSeek."""
    try:
        page.goto(DEEPSEEK_URL, wait_until="domcontentloaded", timeout=30000)
    except PWTimeout:
        pass
    except Exception as e:
        raise RuntimeError(f"Ошибка загрузки DeepSeek: {e}")

    # Проверяем авторизацию
    if any(kw in page.url.lower() for kw in ("login", "signin", "auth")):
        raise RuntimeError(
            "Не авторизован! Запусти сначала:\n"
            "  uv run --with playwright,openpyxl scripts/deepseek_dimensions_filler.py --login"
        )

    # Ждём поле ввода
    try:
        page.wait_for_selector("textarea, [contenteditable='true']", timeout=15000, state="visible")
    except Exception:
        time.sleep(3)

    # Если открылся старый чат — нажимаем «New chat»
    if "/chat/" in page.url or "/c/" in page.url:
        for btn_sel in [
            'button[aria-label*="New" i]',
            'button[aria-label*="Новый" i]',
            'a[href="/"]',
            '[data-testid="new-chat"]',
        ]:
            try:
                btn = page.locator(btn_sel).first
                if btn.is_visible(timeout=800):
                    btn.click()
                    time.sleep(1)
                    break
            except Exception:
                continue

    time.sleep(1)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ХЕЛПЕРЫ
# ══════════════════════════════════════════════════════════════════════════════

def read_headers(ws) -> dict:
    return {
        ws.cell(1, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }


def ensure_new_columns(ws, headers: dict) -> dict:
    existing = set(headers.keys())
    if all(col in existing for col in NEW_COLS):
        return headers

    insert_at = headers.get(COL_AFTER, ws.max_column) + 1
    to_add = [col for col in NEW_COLS if col not in existing]

    ws.insert_cols(insert_at, len(to_add))
    for i, col_name in enumerate(to_add):
        cell = ws.cell(1, insert_at + i, col_name)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1A3A5C")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 14

    return read_headers(ws)


# ══════════════════════════════════════════════════════════════════════════════
#  ОБРАБОТКА ФАЙЛА
# ══════════════════════════════════════════════════════════════════════════════

def process_file(page, src: Path, dst: Path, row_range: tuple | None, debug: bool):
    if not dst.exists():
        shutil.copy2(src, dst)
        print(f"  → Скопирован: {dst.name}")

    wb = openpyxl.load_workbook(dst)
    ws = wb.active

    headers = read_headers(ws)
    headers = ensure_new_columns(ws, headers)

    desc_col  = headers.get(COL_AFTER)
    name_col  = headers.get("Наименование")
    param_col = headers.get("Параметры")
    w_col     = headers.get("Вес, г")
    l_col     = headers.get("Длина, мм")
    wd_col    = headers.get("Ширина, мм")
    h_col     = headers.get("Высота, мм")

    if not desc_col:
        print(f"  ⚠ Колонка «{COL_AFTER}» не найдена — пропуск")
        wb.close()
        return

    todos = []
    for r in range(2, ws.max_row + 1):
        if row_range and not (row_range[0] <= r <= row_range[1]):
            continue
        if not ws.cell(r, desc_col).value:
            continue
        if w_col and ws.cell(r, w_col).value is not None:
            continue
        todos.append(r)

    total = len(todos)
    if not todos:
        print(f"  Нечего обрабатывать")
        wb.close()
        return

    print(f"  Строк для обработки: {total}")

    processed  = 0
    save_timer = 0

    for batch_no, batch_start in enumerate(range(0, total, BATCH_SIZE), start=1):
        batch_rows = todos[batch_start : batch_start + BATCH_SIZE]

        items = []
        for local_idx, r in enumerate(batch_rows):
            items.append({
                "idx":    local_idx,
                "name":   ws.cell(r, name_col).value  if name_col  else "",
                "params": ws.cell(r, param_col).value if param_col else "",
                "desc":   ws.cell(r, desc_col).value  or "",
            })

        print(f"  Батч {batch_no} (стр. {batch_rows[0]}–{batch_rows[-1]})", end="", flush=True)

        # Новый чат перед каждым батчем — чистый контекст
        try:
            open_new_chat(page)
        except RuntimeError as e:
            print(f"\n  ✗ {e}")
            wb.save(dst)
            wb.close()
            raise

        prompt = build_prompt(items)

        # Вставляем промпт
        try:
            type_into_textarea(page, prompt)
        except RuntimeError as e:
            print(f"\n  ✗ {e}")
            if debug:
                shot = OUT_DIR / f"_debug_b{batch_no}_noinput.png"
                try:
                    page.screenshot(path=str(shot))
                    print(f"    Скриншот: {shot}")
                except Exception:
                    pass
            continue

        if debug:
            shot = OUT_DIR / f"_debug_b{batch_no}_typed.png"
            try:
                page.screenshot(path=str(shot))
                print(f"\n    Скриншот перед отправкой: {shot}")
                input("    Нажми Enter для отправки > ")
            except Exception:
                pass

        click_send(page)
        time.sleep(1.5)

        # 3 попытки получить валидный JSON
        results = []
        for attempt in range(1, 4):
            raw = wait_for_response(page)

            if not raw:
                print(f"  ✗ Пустой ответ (попытка {attempt}/3)")
            elif any(p in raw.lower() for p in ERROR_PHRASES):
                print(f"  ⚠ Ошибка сервера (попытка {attempt}/3) — жду 20 сек...")
                time.sleep(20)
                if attempt < 3:
                    try:
                        open_new_chat(page)
                        type_into_textarea(page, prompt)
                        click_send(page)
                        time.sleep(1.5)
                    except Exception:
                        pass
                continue
            else:
                results = extract_json_items(raw)
                if results:
                    break
                print(f"\n  ⚠ JSON не распознан (попытка {attempt}/3)")
                print(f"    Начало ответа: {raw[:300]}")

            if attempt < 3:
                time.sleep(8)

        if debug and raw:
            try:
                (OUT_DIR / f"_debug_b{batch_no}_response.txt").write_text(raw, encoding="utf-8")
            except Exception:
                pass

        if not results:
            print(f"  ✗ Батч пропущен")
            continue

        result_map = {r["idx"]: r for r in results}
        filled = 0
        for local_idx, r in enumerate(batch_rows):
            data = result_map.get(local_idx)
            if not data:
                continue
            if w_col:  ws.cell(r, w_col).value  = data.get("вес")
            if l_col:  ws.cell(r, l_col).value  = data.get("длина")
            if wd_col: ws.cell(r, wd_col).value = data.get("ширина")
            if h_col:  ws.cell(r, h_col).value  = data.get("высота")
            filled += 1

        processed  += filled
        save_timer += 1
        print(f" → {filled}/{len(batch_rows)}")

        if save_timer >= SAVE_EVERY:
            wb.save(dst)
            print(f"    💾 Сохранено ({processed} строк)")
            save_timer = 0

        time.sleep(2)

    wb.save(dst)
    print(f"  ✓ Итого: {processed}/{total} → {dst.name}")
    wb.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Заполнение веса и габаритов через DeepSeek (браузер)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--file",  default=None, help="Один файл (имя xlsx)")
    ap.add_argument("--rows",  default=None, help="Диапазон строк: 2-100")
    ap.add_argument("--login", action="store_true", help="Открыть браузер для входа в DeepSeek")
    ap.add_argument("--debug", action="store_true", help="Скриншоты + пауза перед отправкой")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    src_files = [SRC_DIR / args.file] if args.file else sorted(SRC_DIR.glob("*.xlsx"))

    row_range = None
    if args.rows:
        parts = args.rows.split("-")
        row_range = (int(parts[0]), int(parts[1]) if len(parts) > 1 else int(parts[0]))

    print()
    print("═" * 64)
    print("  DeepSeek Dimensions Filler (браузер)")
    print(f"  Батч:    {BATCH_SIZE} строк/запрос")
    print(f"  Файлов:  {len(src_files)}")
    print(f"  Из:      {SRC_DIR}")
    print(f"  В:       {OUT_DIR}")
    print("═" * 64)

    with sync_playwright() as pw:
        context = pw.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=False,
            viewport={"width": 1280, "height": 900},
            locale="ru-RU",
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = context.new_page()

        # ── Режим входа ──────────────────────────────────────────────────────
        if args.login:
            print("\n  Открываю DeepSeek...")
            print("  Войди в свой аккаунт в открывшемся браузере.")
            page.goto(DEEPSEEK_URL, wait_until="domcontentloaded")
            print("  Когда окажешься в чате — нажми Enter здесь.")
            input("  > ")
            print("  ✓ Сессия сохранена. Теперь можно запускать без --login.")
            context.close()
            return

        # ── Проверка авторизации ─────────────────────────────────────────────
        print("\n  Открываю DeepSeek...")
        try:
            page.goto(DEEPSEEK_URL, wait_until="domcontentloaded", timeout=30000)
        except PWTimeout:
            print("  ✗ Таймаут. Проверь интернет-соединение.")
            context.close()
            sys.exit(1)

        time.sleep(3)
        if any(kw in page.url.lower() for kw in ("login", "signin", "auth")):
            print("\n  ✗ Не авторизован!")
            print("  Запусти сначала:")
            print("    uv run --with playwright,openpyxl scripts/deepseek_dimensions_filler.py --login")
            context.close()
            sys.exit(1)

        print("  ✓ Авторизован\n")

        # ── Основной цикл ────────────────────────────────────────────────────
        total_files = len(src_files)
        for i, src in enumerate(src_files, 1):
            if not src.exists():
                print(f"[{i:3}/{total_files}] ✗ Не найден: {src.name}")
                continue

            dst = OUT_DIR / src.name
            print(f"[{i:3}/{total_files}] {src.name}")

            try:
                process_file(page, src, dst, row_range, debug=args.debug)
            except KeyboardInterrupt:
                print("\n\n  Прерывание — прогресс сохранён в последнем 💾")
                context.close()
                sys.exit(0)
            except RuntimeError as e:
                print(f"  ✗ {e}")
                context.close()
                sys.exit(1)
            except Exception as e:
                print(f"  ✗ Неожиданная ошибка: {e}")
                traceback.print_exc()

            print()

        context.close()

    print("═" * 64)
    print("  Всё готово!")
    print(f"  Результаты: {OUT_DIR}")
    print("═" * 64)
    print()


if __name__ == "__main__":
    main()
