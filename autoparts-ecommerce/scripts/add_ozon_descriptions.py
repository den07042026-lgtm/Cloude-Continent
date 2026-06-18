"""
add_ozon_descriptions.py
════════════════════════
Добавляет столбец «Описание» между «Альтернативные артикулы товара» и «Изображения»
в каждый Excel-файл из папки источника. Описание генерируется через GigaChat (браузер).

Источник:  C:/Users/Admin/Desktop/На сортировку 24.04(2)/
Результат: C:/Users/Admin/Desktop/На сортировку 26.04/

Установка браузера (один раз):
  uv run --with playwright python -m playwright install chromium

Первый запуск — войти в GigaChat:
  uv run --with playwright,openpyxl scripts/add_ozon_descriptions.py --login

Обычный запуск — все файлы:
  uv run --with playwright,openpyxl scripts/add_ozon_descriptions.py

Один конкретный файл:
  uv run --with playwright,openpyxl scripts/add_ozon_descriptions.py --file "Амортизатор подвески.xlsx"

Ограничить строки (для теста):
  uv run --with playwright,openpyxl scripts/add_ozon_descriptions.py --file "Амортизатор подвески.xlsx" --rows 2-5
"""

import sys
import time
import shutil
import argparse
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Playwright не установлен.")
    print("  uv run --with playwright python -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("openpyxl не установлен: uv run --with playwright,openpyxl ...")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  ПУТИ
# ══════════════════════════════════════════════════════════════════════════════

SRC_DIR      = Path(r"C:\Users\Admin\Desktop\На сортировку 24.04(2)")
DST_DIR      = Path(r"C:\Users\Admin\Desktop\На сортировку 26.04")
PROFILE_DIR  = Path.home() / ".gigachat_playwright"
GIGACHAT_URL = "https://giga.chat/"

COL_DESC_HEADER = "Описание"
COL_AFTER       = "Альтернативные артикулы товара"
COL_BEFORE      = "Изображения"

SAVE_EVERY = 5


# ══════════════════════════════════════════════════════════════════════════════
#  ПРОМПТ
# ══════════════════════════════════════════════════════════════════════════════

def build_prompt(code: str, name: str, brand: str, params: str, compat: str) -> str:
    params_block = params.strip() if params else "(нет данных)"

    if compat:
        models = [m.strip() for m in compat.split(";") if m.strip()]
        if len(models) > 10:
            compat_short = "; ".join(models[:10]) + f" и ещё {len(models) - 10} моделей"
        else:
            compat_short = "; ".join(models)
    else:
        compat_short = "уточняется по артикулу"

    return f"""Ты копирайтер интернет-магазина автозапчастей. Напиши описание товара для маркетплейса Ozon.

ТРЕБОВАНИЯ:
- Длина: от 600 до 900 символов (считай строго)
- Без HTML, без заголовков, без маркированных списков
- Сплошной текст из 3 абзацев
- Только русский язык
- Тон: спокойный, информативный, без агрессивных призывов и восклицательных знаков
- Артикул {brand} {code.upper()} обязательно упомяни в тексте

ДАННЫЕ О ТОВАРЕ:
Бренд: {brand}
Артикул: {code.upper()}
Наименование: {name}

Технические характеристики:
{params_block}

Применяется на автомобилях:
  {compat_short}

СТРУКТУРА (строго три абзаца):
1. Что это за деталь, её роль, тип и конструкция. Естественно упомяни артикул {brand} {code.upper()}.
2. Для каких автомобилей подходит — назови 2–3 марки из списка применяемости.
3. Коротко о бренде {brand} и надёжности детали. Мягкий, ненавязчивый призыв.

Напиши только текст описания, без комментариев и пояснений."""


# ══════════════════════════════════════════════════════════════════════════════
#  PLAYWRIGHT — ВЗАИМОДЕЙСТВИЕ С GIGACHAT
# ══════════════════════════════════════════════════════════════════════════════

def dismiss_modal(page):
    """Закрывает модальные окна нажатием Escape или кликом по backdrop."""
    try:
        page.keyboard.press("Escape")
        time.sleep(0.5)
    except Exception:
        pass
    try:
        # backdrop из лога: div[data-open="true"][data-backdrop]
        backdrop = page.locator('[data-backdrop]').first
        if backdrop.is_visible(timeout=500):
            page.mouse.click(5, 5)
            time.sleep(0.5)
    except Exception:
        pass


def find_input(page):
    time.sleep(0.5)
    # Точный id из лога GigaChat
    for sel in [
        '#chat-input-textarea',
        'textarea[placeholder*="Спрос"]',
        'textarea[placeholder*="Напиш"]',
        'textarea',
    ]:
        try:
            el = page.locator(sel).last
            if el.is_visible(timeout=2000):
                return el
        except Exception:
            continue
    return None


def type_prompt(page, text: str):
    dismiss_modal(page)
    inp = find_input(page)
    if inp is None:
        raise RuntimeError("Поле ввода не найдено")
    # Снимаем возможный оверлей через JS-фокус
    page.evaluate("""() => {
        const el = document.querySelector('#chat-input-textarea') ||
                   document.querySelector('textarea');
        if (el) el.focus();
    }""")
    time.sleep(0.3)
    # Вставляем текст через React-совместимый setter
    page.evaluate("""(text) => {
        const el = document.querySelector('#chat-input-textarea') ||
                   document.querySelector('textarea');
        if (!el) return;
        const setter = Object.getOwnPropertyDescriptor(
            window.HTMLTextAreaElement.prototype, 'value').set;
        setter.call(el, text);
        el.dispatchEvent(new Event('input',  { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
    }""", text)
    time.sleep(0.5)


def click_send(page):
    # Сначала пробуем кнопку отправки
    for sel in [
        'button[aria-label*="Отправить"]',
        'button[aria-label*="отправить"]',
        'button[type="submit"]',
        '[data-testid="send-button"]',
    ]:
        try:
            btn = page.locator(sel).last
            if btn.is_visible(timeout=800):
                btn.click()
                return
        except Exception:
            continue
    # Fallback: Enter в поле ввода
    page.locator('#chat-input-textarea').press("Enter")


def get_last_response(page) -> str:
    # Начало промпта — по нему исключаем блоки с сообщением пользователя
    PROMPT_MARKER = "Ты копирайтер"

    try:
        text = page.evaluate("""(marker) => {
            const paras = [...document.querySelectorAll('p')]
                .filter(p => (p.innerText || '').trim().length > 40);
            if (!paras.length) return '';

            // Идём с конца, пропускаем параграфы внутри пользовательского сообщения
            for (let i = paras.length - 1; i >= 0; i--) {
                const p = paras[i];
                if ((p.innerText || '').includes(marker)) continue;

                // Собираем родительский контейнер с 2+ параграфами
                let container = p;
                for (let j = 0; j < 8; j++) {
                    if (!container.parentElement) break;
                    const parent = container.parentElement;
                    const pCount = parent.querySelectorAll('p').length;
                    if (pCount >= 2 && !parent.innerText.includes(marker)) {
                        container = parent;
                        break;
                    }
                    container = parent;
                }

                const t = container.innerText.trim();
                if (t.length > 50 && !t.startsWith(marker)) return t;
            }
            return '';
        }""", PROMPT_MARKER)
        if text and len(text) > 50:
            return text
    except Exception:
        pass

    # Fallback: последний большой листовой элемент, не содержащий промпт
    try:
        text = page.evaluate("""(marker) => {
            const all = [...document.querySelectorAll('div, p, span')];
            const big = all.filter(el =>
                el.children.length === 0 &&
                (el.innerText || '').length > 100 &&
                !(el.innerText || '').includes(marker)
            );
            return big.length ? big[big.length - 1].innerText.trim() : '';
        }""", PROMPT_MARKER)
        if text and len(text) > 50:
            return text
    except Exception:
        pass
    return ""


PROMPT_MARKER = "Ты копирайтер"


def is_valid_response(text: str) -> bool:
    """Возвращает False если текст — это промпт, а не ответ ГигаЧата."""
    return bool(text) and PROMPT_MARKER not in text


def wait_for_response(page, timeout_sec: int = 120) -> str:
    print("жду", end="", flush=True)

    # Ждём появления валидного текста (до 60 сек)
    for _ in range(30):
        candidate = get_last_response(page)
        if is_valid_response(candidate) and len(candidate) > 30:
            break
        time.sleep(2)
        print(".", end="", flush=True)
    else:
        try:
            html = page.content()
            debug_html = DST_DIR / "_debug_page.html"
            debug_html.write_text(html, encoding="utf-8")
            page.screenshot(path=str(DST_DIR / "_debug_page.png"))
            print(f" таймаут (0 симв.) → дамп: {debug_html}", flush=True)
        except Exception:
            print(" таймаут (0 симв.)", flush=True)
        return ""

    # Ждём стабилизации — только валидные ответы считаем
    prev, stable = "", 0
    for tick in range(timeout_sec):
        cur = get_last_response(page)
        if not is_valid_response(cur):
            # Пришёл промпт вместо ответа — сбрасываем счётчик, ждём дальше
            stable = 0
            prev = ""
            time.sleep(1.5)
            continue
        if cur == prev:
            stable += 1
            if stable >= 2:
                print(f" ({len(cur)} симв.)", flush=True)
                return cur
        else:
            stable = 0
            if tick % 4 == 0:
                print(".", end="", flush=True)
        prev = cur
        time.sleep(1.5)

    print(f" таймаут ({len(prev)} симв.)", flush=True)
    return prev if is_valid_response(prev) else ""


def new_chat(page):
    """Открывает новый чат и закрывает модальные окна."""
    dismiss_modal(page)
    for sel in [
        'a[href="/"][aria-label*="чат"]',
        'button[aria-label*="новый чат"]',
        'button[aria-label*="Новый чат"]',
        '[data-testid="new-chat"]',
        'a[href="/"]',
    ]:
        try:
            btn = page.locator(sel).first
            if btn.is_visible(timeout=800):
                btn.click()
                time.sleep(1.5)
                dismiss_modal(page)
                return
        except Exception:
            continue
    # Fallback: переходим на главную страницу чата
    try:
        page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=20000)
    except Exception:
        pass
    time.sleep(2)
    dismiss_modal(page)


def get_description(page, code, name, brand, params, compat, retries: int = 3) -> str:
    for attempt in range(1, retries + 1):
        try:
            new_chat(page)
            prompt = build_prompt(code, name, brand, params, compat)
            type_prompt(page, prompt)
            click_send(page)
            time.sleep(1)
            return wait_for_response(page)
        except RuntimeError as e:
            screenshot = DST_DIR / f"_debug_{code}.png"
            try:
                page.screenshot(path=str(screenshot))
                print(f"  ✗ {e} → скриншот: {screenshot.name}")
            except Exception:
                print(f"  ✗ {e}")
            return ""
        except Exception as e:
            if attempt < retries:
                print(f"  ↺ ошибка, попытка {attempt + 1}/{retries}: {e}", flush=True)
                time.sleep(10)
            else:
                print(f"  ✗ {e}", flush=True)
                return ""
    return ""


# ══════════════════════════════════════════════════════════════════════════════
#  РАБОТА С EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def prepare_workbook(src_path: Path, dst_path: Path):
    if not dst_path.exists():
        shutil.copy2(src_path, dst_path)

    wb = openpyxl.load_workbook(dst_path)
    ws = wb.active

    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    col_after  = headers.get(COL_AFTER)
    col_before = headers.get(COL_BEFORE)

    if not col_after or not col_before:
        print(f"    ! Колонки не найдены: {COL_AFTER!r}={col_after}, {COL_BEFORE!r}={col_before}")
        print(f"    Заголовки файла: {list(headers.keys())}")
        return None

    if COL_DESC_HEADER not in headers:
        insert_at = col_after + 1
        ws.insert_cols(insert_at)

        hcell = ws.cell(1, insert_at, COL_DESC_HEADER)
        hcell.font      = Font(bold=True, color="FFFFFF", size=10)
        hcell.fill      = PatternFill("solid", fgColor="1A1A2E")
        hcell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[hcell.column_letter].width = 65

        wb.save(dst_path)
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    col_desc   = headers[COL_DESC_HEADER]
    col_code   = headers.get("Код (Mikado)", 1)
    col_name   = headers.get("Наименование", 2)
    col_brand  = headers.get("Бренд", 3)
    col_params = headers.get("Параметры", 7)
    col_compat = headers.get("Применяемость", 9)
    col_oem    = headers.get("OEM номера", 8)
    col_images = headers.get("Изображения")

    return wb, ws, col_code, col_desc, col_name, col_brand, col_params, col_compat, col_oem, col_images


def process_file(page, src_path: Path, dst_path: Path,
                 row_start: int, row_end: int | None, delay: int):
    print(f"\n  Файл: {src_path.name}")

    result = prepare_workbook(src_path, dst_path)
    if result is None:
        return

    wb, ws, col_code, col_desc, col_name, col_brand, col_params, col_compat, col_oem, col_images = result

    max_row = ws.max_row
    end     = min(row_end, max_row) if row_end else max_row

    done = skipped = failed = incomplete = 0

    for row in range(row_start, end + 1):
        code = ws.cell(row, col_code).value
        if not code:
            continue

        code   = str(code).strip()
        name   = str(ws.cell(row, col_name).value or "")
        brand  = str(ws.cell(row, col_brand).value or "")
        params = str(ws.cell(row, col_params).value or "")
        compat = str(ws.cell(row, col_compat).value or "")
        oem    = str(ws.cell(row, col_oem).value or "") if col_oem else ""
        images = str(ws.cell(row, col_images).value or "") if col_images else ""

        if not all([params.strip(), oem.strip(), compat.strip(), images.strip()]):
            incomplete += 1
            continue

        existing = ws.cell(row, col_desc).value
        if existing and str(existing).strip():
            skipped += 1
            continue

        print(f"    [{row}/{end}] {brand} {code} — ", end="", flush=True)

        desc = get_description(page, code, name, brand, params, compat)

        if desc:
            cell = ws.cell(row, col_desc, desc)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            done += 1
        else:
            failed += 1

        if (done + failed) % SAVE_EVERY == 0:
            wb.save(dst_path)

        if row < end:
            time.sleep(delay)

    wb.save(dst_path)
    print(f"    → сгенерировано: {done}, пропущено (уже есть): {skipped}, неполные данные: {incomplete}, ошибок: {failed}")
    print(f"    → сохранено: {dst_path.name}")


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="Добавляет «Описание» в Excel через GigaChat (браузер)")
    ap.add_argument("--file",    default=None, help="Конкретный файл из папки источника")
    ap.add_argument("--rows",    default=None, help="Диапазон строк: 2-10 (по умолч. все)")
    ap.add_argument("--login",   action="store_true", help="Открыть браузер для входа в GigaChat")
    ap.add_argument("--delay",   default=3, type=int, help="Пауза между артикулами, сек")
    ap.add_argument("--profile", default=None, help="Папка профиля браузера")
    args = ap.parse_args()

    DST_DIR.mkdir(parents=True, exist_ok=True)
    profile_dir = Path(args.profile) if args.profile else PROFILE_DIR
    profile_dir.mkdir(parents=True, exist_ok=True)

    row_start, row_end = 2, None
    if args.rows:
        if "-" in args.rows:
            a, b = args.rows.split("-")
            row_start, row_end = int(a), int(b)
        else:
            row_start = row_end = int(args.rows)

    if args.file:
        files = [SRC_DIR / args.file]
    else:
        files = sorted(SRC_DIR.glob("*.xlsx"))

    if not files and not args.login:
        print("Файлы не найдены")
        sys.exit(1)

    print()
    print("═" * 60)
    print("  Ozon Description Generator  [GigaChat]")
    print(f"  Файлов:    {len(files)}")
    print(f"  Строки:    {row_start}–{row_end or 'все'}")
    print(f"  Источник:  {SRC_DIR}")
    print(f"  Результат: {DST_DIR}")
    print("═" * 60)

    headless = not args.login

    with sync_playwright() as pw:
        # Пробуем использовать установленный Chrome (у него правильные сетевые настройки)
        # Если Chrome не установлен — падаем на встроенный Chromium
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(profile_dir),
                channel="chrome",
                headless=headless,
                viewport={"width": 1280, "height": 900},
                locale="ru-RU",
                args=["--disable-blink-features=AutomationControlled"],
            )
        except Exception:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(profile_dir),
                headless=headless,
                viewport={"width": 1280, "height": 900},
                locale="ru-RU",
                args=["--disable-blink-features=AutomationControlled"],
            )
        page = context.new_page()

        if args.login:
            print("\n  Открываю браузер.")
            print(f"  Перейди вручную на {GIGACHAT_URL} и залогинься.")
            print("  После входа нажми Enter в этом окне терминала.")
            try:
                page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=15000)
            except Exception:
                pass  # браузер открыт, пользователь сам перейдёт
            input("  Нажми Enter когда окажешься в чате... ")
            context.storage_state(path=str(profile_dir / "state.json"))
            print("  ✓ Сессия сохранена.")
            context.close()
            return

        # Проверяем авторизацию
        try:
            page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=30000)
        except Exception as e:
            print(f"\n  ✗ Не удалось открыть {GIGACHAT_URL}: {e}")
            print("  Проверь подключение к интернету и доступность giga.chat")
            context.close()
            sys.exit(1)
        time.sleep(3)
        if any(k in page.url.lower() for k in ("login", "signin", "auth", "sso", "id.sber")):
            print("\n  ✗ Не авторизован! Запусти сначала:")
            print("    uv run --with playwright,openpyxl scripts/add_ozon_descriptions.py --login")
            context.close()
            sys.exit(1)
        print("  ✓ Авторизован в GigaChat\n")

        for i, src_path in enumerate(files, 1):
            dst_path = DST_DIR / src_path.name
            print(f"[{i}/{len(files)}]", end=" ")
            process_file(page, src_path, dst_path, row_start, row_end, args.delay)

        context.close()

    print()
    print("═" * 60)
    print("  Всё готово!")
    print(f"  Файлы: {DST_DIR}")
    print("═" * 60)


if __name__ == "__main__":
    main()
