import openpyxl, sys, io, importlib
sys.path.insert(0, 'scripts')
import emex_enricher as enr
importlib.reload(enr)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ.xlsx')
ws = wb.worksheets[0]

count = 0
for r in range(2, 502):
    alts = str(ws.cell(r, 9).value or '')
    if not alts:
        continue
    parts = [p.strip() for p in alts.split('/')]
    oem_parts = [p for p in parts if p and enr._is_oem_brand(p.split(' ')[0])]
    if not oem_parts:
        continue
    art   = str(ws.cell(r, 1).value or '').strip()
    brand = str(ws.cell(r, 3).value or '').strip()
    d     = enr.load_cache(art, brand)
    makes = d.get('makes', []) if d else []
    for m in makes:
        mb = m.get('make', '')
        mn = m.get('num', '')
        if enr._is_oem_brand(mb):
            print(f'Row {r-1}  art={art}  brand={brand}  OEM_make={mb!r}  num={mn!r}')
    count += 1
    if count >= 10:
        break
