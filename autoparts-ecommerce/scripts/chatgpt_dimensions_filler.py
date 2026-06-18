"""
chatgpt_dimensions_filler.py
════════════════════════════
Добавляет столбцы «Вес, г», «Длина, мм», «Ширина, мм», «Высота, мм»
после столбца «Описание» — через браузер chat.openai.com, без API ключа.

Исходная папка:    C:\\Users\\Admin\\Desktop\\На сортировку 26.04\\
Папка результатов: C:\\Users\\Admin\\Desktop\\На сортировку 08.05\\

Установка браузера (один раз):
  uv run --with playwright python -m playwright install chromium

Первый запуск — войти в ChatGPT:
  uv run --with playwright,openpyxl scripts/chatgpt_dimensions_filler.py --login

Обработать все файлы:
  uv run --with playwright,openpyxl scripts/chatgpt_dimensions_filler.py

Один файл для теста:
  uv run --with playwright,openpyxl scripts/chatgpt_dimensions_filler.py --file "Амортизатор подвески.xlsx" --rows 2-30

Продолжение после прерывания:
  Просто запустить снова — пропустит строки, которые уже заполнены.

Если селекторы сломались после обновления сайта:
  uv run --with playwright,openpyxl scripts/chatgpt_dimensions_filler.py --file "Амортизатор подвески.xlsx" --rows 2-5 --debug
"""

import re
import sys
import json
import time
import shutil
import argparse
import traceback
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Playwright не установлен.")
    print("Шаг 1: uv run --with playwright python -m playwright install chromium")
    print("Шаг 2: uv run --with playwright,openpyxl scripts/chatgpt_dimensions_filler.py --login")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("openpyxl не установлен.")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

SRC_DIR     = Path(r"C:\Users\Admin\Desktop\На сортировку 26.04")
OUT_DIR     = Path(r"C:\Users\Admin\Desktop\На сортировку 08.05")
PROFILE_DIR = Path.home() / ".chatgpt_playwright"
CHATGPT_URL = "https://chatgpt.com"

BATCH_SIZE = 50   # строк за один запрос к ChatGPT
SAVE_EVERY = 3    # сохранять Excel каждые N батчей (~45 строк)

NEW_COLS  = ["Вес, г", "Длина, мм", "Ширина, мм", "Высота, мм"]
COL_AFTER = "Описание"

# Селекторы ChatGPT (актуальны на май 2025)
# Если сломались — запусти с --debug и проверь скриншоты
SEL_INPUT    = '#prompt-textarea'
SEL_SEND_BTN = 'button[data-testid="send-button"]'
SEL_STOP_BTN = 'button[aria-label="Stop streaming"]'
SEL_RESPONSE = '[data-message-author-role="assistant"]'


# ══════════════════════════════════════════════════════════════════════════════
#  ПРОМПТ
# ══════════════════════════════════════════════════════════════════════════════

INSTRUCTION = (
    "Ты эксперт по автозапчастям. Для каждой позиции ниже укажи "
    "вес и габариты ТРАНСПОРТНОЙ УПАКОВКИ — то есть внешние размеры коробки или пакета "
    "вместе с самой деталью внутри, именно так, как товар будет отправлен покупателю.\n"
    "ВАЖНО: это НЕ размеры самой детали, а размеры упаковки (коробки/пакета) снаружи.\n"
    "Правила:\n"
    "- Если деталь идёт комплектом (2 шт. и т.п.) — вес/размер всего комплекта в упаковке\n"
    "- Опирайся на типичные размеры упаковки для данного вида запчасти\n"
    "- Верни ТОЛЬКО JSON объект, без пояснений, без markdown блоков:\n"
    '{"items": [{"idx": 0, "вес": 500, "длина": 200, "ширина": 150, "высота": 80}, ...]}\n'
    "вес — граммы (деталь + упаковка), длина/ширина/высота — мм (внешние стороны упаковки), "
    "всё целые числа.\n\n"
)


def build_prompt(items: list[dict]) -> str:
    lines = [INSTRUCTION]
    for item in items:
        lines.append(f"[{item['idx']}]")
        lines.append(f"Наименование: {item['name']}")
        if item.get("params"):
            lines.append(f"Параметры: {str(item['params'])[:200]}")
        if item.get("desc"):
            lines.append(f"Описание: {str(item['desc'])[:250]}")
        lines.append("")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  ПАРСИНГ ОТВЕТА ChatGPT
# ══════════════════════════════════════════════════════════════════════════════

def extract_json_items(text: str) -> list[dict]:
    """Вытаскивает список items из ответа ChatGPT.
    Обрабатывает markdown-блоки, произвольный текст вокруг JSON, вложенные объекты.
    """
    # Убираем markdown-блоки
    clean = re.sub(r"```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    clean = clean.replace("```", "").strip()

    ITEM_KEYS = {"вес", "длина", "ширина", "высота", "idx"}

    def is_valid_items(lst: list) -> bool:
        """Список словарей с хотя бы одним ключом из ожидаемых."""
        return (
            bool(lst)
            and isinstance(lst[0], dict)
            and bool(ITEM_KEYS & lst[0].keys())
        )

    def try_parse(s: str):
        s = s.strip()
        try:
            parsed = json.loads(s)
            if isinstance(parsed, dict) and "items" in parsed:
                items = parsed["items"]
                if isinstance(items, list) and is_valid_items(items):
                    return items
            if isinstance(parsed, list) and is_valid_items(parsed):
                return parsed
        except Exception:
            pass
        return None

    # 1. Весь очищенный текст как JSON
    result = try_parse(clean)
    if result is not None:
        return result

    # 2. Сканируем: находим каждый '{' и пытаемся вырезать валидный объект
    for start, opener, closer in [(i, '{', '}') for i, c in enumerate(clean) if c == '{']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == opener:
                depth += 1
            elif c == closer:
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result is not None:
                        return result
                    break

    # 3. Сканируем: находим каждый '[' и пытаемся вырезать валидный массив
    for start in [i for i, c in enumerate(clean) if c == '[']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == '[':
                depth += 1
            elif c == ']':
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result is not None:
                        return result
                    break

    return []


# ══════════════════════════════════════════════════════════════════════════════
#  ВЗАИМОДЕЙСТВИЕ С БРАУЗЕРОМ
# ══════════════════════════════════════════════════════════════════════════════

RATE_LIMIT_PHRASES = [
    "достигнут лимит",
    "message limit",
    "you've reached",
    "rate limit",
    "попробуйте позже",
    "try again",
    "upgrade",
]

SERVER_ERROR_PHRASES = [
    "something went wrong",
    "if this issue persists",
    "help.openai.com",
    "unexpected error",
    "server error",
    "try again later",
]
RATE_LIMIT_WAIT_MIN = 65  # ждать 65 минут перед повтором


def is_rate_limited(page) -> bool:
    """Проверяет есть ли на странице сообщение о лимите."""
    try:
        body = page.locator("body").inner_text(timeout=2000).lower()
        return any(phrase in body for phrase in RATE_LIMIT_PHRASES)
    except Exception:
        return False


def wait_for_rate_limit_reset(page):
    """Ждёт сброса лимита сообщений ChatGPT с обратным отсчётом."""
    print(f"\n  ⏳ Лимит сообщений ChatGPT — жду {RATE_LIMIT_WAIT_MIN} мин. до сброса...")
    for remaining in range(RATE_LIMIT_WAIT_MIN, 0, -1):
        print(f"\r  ⏳ Осталось: {remaining} мин.  ", end="", flush=True)
        time.sleep(60)
    print(f"\r  ✓ Продолжаю работу...                    ")
    # Перезагружаем страницу после ожидания
    try:
        page.goto(CHATGPT_URL, wait_until="networkidle", timeout=40000)
    except Exception:
        pass


def dismiss_modals(page):
    """Закрывает модальные окна ChatGPT. При лимите сообщений — ждёт сброса."""
    # Проверяем лимит сообщений
    if is_rate_limited(page):
        wait_for_rate_limit_reset(page)
        return

    modal_ids = [
        '#modal-conversation-history-rate-limit',
        '[data-testid="modal-conversation-history-rate-limit"]',
    ]
    for modal_sel in modal_ids:
        try:
            if page.locator(modal_sel).is_visible(timeout=800):
                for btn_sel in [
                    f'{modal_sel} button',
                    'button:has-text("OK")',
                    'button:has-text("Понятно")',
                    'button:has-text("Got it")',
                    'button:has-text("Close")',
                    'button:has-text("Закрыть")',
                ]:
                    try:
                        btn = page.locator(btn_sel).first
                        if btn.is_visible(timeout=500):
                            btn.click()
                            time.sleep(0.8)
                            return
                    except Exception:
                        continue
        except Exception:
            continue
    try:
        page.keyboard.press("Escape")
        time.sleep(0.4)
    except Exception:
        pass


def _page_diagnostics(page) -> str:
    """Возвращает строку с диагностикой текущего состояния страницы."""
    try:
        info = page.evaluate("""() => ({
            url:            location.href,
            title:          document.title,
            promptTextarea: !!document.querySelector('#prompt-textarea'),
            contentEditable:document.querySelectorAll('[contenteditable="true"]').length,
            roleTextbox:    document.querySelectorAll('[role="textbox"]').length,
            textareas:      document.querySelectorAll('textarea').length,
            buttons:        document.querySelectorAll('button').length,
        })""")
        return str(info)
    except Exception as e:
        return f"(JS ошибка: {e})"


def find_input(page):
    """Ищет поле ввода ChatGPT (Lexical-редактор)."""
    dismiss_modals(page)

    selectors = [
        '#prompt-textarea',
        'div[id="prompt-textarea"]',
        'div[contenteditable="true"][data-lexical-editor="true"]',
        '[role="textbox"]',
        'div[contenteditable="true"]',
        '[data-testid="prompt-textarea"]',
        'textarea',
        '[placeholder*="Message"]',
        '[placeholder*="Сообщение"]',
    ]
    for sel in selectors:
        try:
            el = page.wait_for_selector(sel, timeout=2500, state="visible")
            if el:
                return page.locator(sel).first
        except Exception:
            continue

    # Ничего не нашли — печатаем диагностику
    print(f"\n    Диагностика: {_page_diagnostics(page)}")
    return None


def type_prompt(page, text: str):
    """Вставляет текст в ChatGPT.
    ProseMirror игнорирует fill() и navigator.clipboard — он обновляет модель
    только через реальный ClipboardEvent с DataTransfer. Именно это и делаем.
    """
    inp = find_input(page)
    if inp is None:
        raise RuntimeError("Поле ввода не найдено — попробуй --debug")

    inp.click()
    time.sleep(0.5)

    # Способ 1: прямой ClipboardEvent с DataTransfer
    # ProseMirror перехватывает именно это событие для вставки текста
    try:
        ok = page.evaluate("""(text) => {
            const el = document.querySelector('#prompt-textarea')
                    || document.querySelector('[role="textbox"]')
                    || document.querySelector('div[contenteditable="true"]');
            if (!el) return false;
            el.focus();
            document.execCommand('selectAll');
            document.execCommand('delete');
            const dt = new DataTransfer();
            dt.setData('text/plain', text);
            el.dispatchEvent(new ClipboardEvent('paste', {
                bubbles: true, cancelable: true, clipboardData: dt
            }));
            return true;
        }""", text)
        time.sleep(1.0)
        if ok:
            return
    except Exception:
        pass

    # Способ 2: системный буфер через navigator.clipboard + Ctrl+V
    try:
        page.evaluate("async (t) => await navigator.clipboard.writeText(t)", text)
        time.sleep(0.4)
        inp.click()
        time.sleep(0.2)
        page.keyboard.press("Control+v")
        time.sleep(0.8)
        return
    except Exception:
        pass

    raise RuntimeError("Не удалось вставить текст — попробуй --debug")


def click_send(page) -> bool:
    """Нажимает кнопку отправки. Ждёт пока она станет активной после вставки текста."""
    btn_selectors = [
        'button[data-testid="send-button"]',
        'button[aria-label="Send prompt"]',
        'button[aria-label="Отправить"]',
        'button[aria-label*="send" i]',
        'button[type="submit"]',
    ]

    # Ждём до 5 сек пока хоть одна кнопка станет видимой и enabled
    for _ in range(10):
        for sel in btn_selectors:
            try:
                btn = page.locator(sel).last
                if btn.is_visible(timeout=300) and btn.is_enabled(timeout=300):
                    btn.click()
                    return True
            except Exception:
                continue
        time.sleep(0.5)

    # Если кнопка не найдена — Enter (ChatGPT: Enter = отправить)
    page.keyboard.press("Enter")
    return False


def is_generating(page) -> bool:
    candidates = [
        SEL_STOP_BTN,
        'button[aria-label*="Stop"]',
        'button[aria-label*="stop"]',
        'button[aria-label*="Остановить"]',
        '[data-testid="stop-button"]',
    ]
    for sel in candidates:
        try:
            if page.locator(sel).first.is_visible(timeout=400):
                return True
        except Exception:
            continue
    return False


def get_last_response(page) -> str:
    """Возвращает текст последнего ответа АССИСТЕНТА.
    Фильтрует сообщения пользователя (они содержат наш промпт).
    """
    # Только специфичные селекторы ассистента — без широких fallback
    assistant_selectors = [
        '[data-message-author-role="assistant"] .markdown',
        '[data-message-author-role="assistant"]',
        '[data-message-role="assistant"] .markdown',
        '[data-message-role="assistant"]',
        # Структура через article/conversation-turn
        'article[data-testid*="conversation-turn"] [data-message-author-role="assistant"]',
    ]
    for sel in assistant_selectors:
        try:
            els = page.locator(sel).all()
            if not els:
                continue
            text = els[-1].inner_text(timeout=2000).strip()
            # Пропускаем — это наш собственный промпт, а не ответ ChatGPT
            if text.startswith(INSTRUCTION[:30]):
                continue
            if len(text) > 10:
                return text
        except Exception:
            continue

    # Fallback: берём ВСЕ сообщения, отбрасываем пользовательские
    try:
        all_msgs = page.locator('[data-message-author-role]').all()
        for el in reversed(all_msgs):
            try:
                role = el.get_attribute("data-message-author-role", timeout=500)
                if role != "assistant":
                    continue
                text = el.inner_text(timeout=1500).strip()
                if len(text) > 10 and not text.startswith(INSTRUCTION[:30]):
                    return text
            except Exception:
                continue
    except Exception:
        pass

    return ""


def wait_for_response(page, timeout_sec: int = 150) -> str:
    """Ждёт стабилизации ответа (текст не меняется 3 проверки подряд)."""
    print("    ожидание", end="", flush=True)

    # Ждём появления хоть какого-то текста (до 40 сек)
    started = False
    for _ in range(27):
        if get_last_response(page):
            started = True
            break
        time.sleep(1.5)
        print(".", end="", flush=True)

    if not started:
        # Диагностика: что вообще на странице
        try:
            info = page.evaluate("""() => ({
                url:       location.href,
                assistant: document.querySelectorAll('[data-message-author-role="assistant"]').length,
                articles:  document.querySelectorAll('article').length,
                markdown:  document.querySelectorAll('.markdown').length,
                generating: !!document.querySelector('button[aria-label*="Stop"]'),
            })""")
            print(f"\n    Диагностика ответа: {info}")
        except Exception:
            pass
        print(f" таймаут (ответ не появился)")
        return ""

    prev, stable = "", 0
    for tick in range(timeout_sec // 2):
        cur = get_last_response(page)
        if cur and cur == prev:
            stable += 1
            if stable >= 3 and not is_generating(page):
                print(f" готово ({len(cur)} симв.)")
                return cur
        else:
            stable = 0
            if tick % 4 == 0:
                print(".", end="", flush=True)
        prev = cur
        time.sleep(2)

    print(f" таймаут ({len(prev)} симв.)")
    return prev


def open_new_chat(page, debug: bool = False, out_dir: Path = None, tag: str = ""):
    """Открывает новый пустой чат на chatgpt.com."""
    try:
        page.goto(CHATGPT_URL, wait_until="networkidle", timeout=40000)
    except PWTimeout:
        # networkidle иногда не наступает — пробуем продолжить
        pass
    except Exception as e:
        raise RuntimeError(f"Ошибка загрузки ChatGPT: {e}")

    if any(kw in page.url.lower() for kw in ("login", "signin", "auth")):
        raise RuntimeError(
            "Не авторизован! Запусти сначала:\n"
            "  uv run --with playwright,openpyxl scripts/chatgpt_dimensions_filler.py --login"
        )

    # Ждём появления поля ввода
    try:
        page.wait_for_selector(
            '#prompt-textarea, div[contenteditable="true"], [role="textbox"]',
            timeout=15000, state="visible",
        )
    except Exception:
        time.sleep(4)

    dismiss_modals(page)

    # Скриншот сохраняется всегда при debug, или в out_dir при любом запуске
    if (debug or out_dir) and out_dir:
        shot = out_dir / f"_debug_{tag}_newchat.png"
        try:
            page.screenshot(path=str(shot))
            if debug:
                print(f"\n    Скриншот: {shot.name}")
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ХЕЛПЕРЫ
# ══════════════════════════════════════════════════════════════════════════════

def read_headers(ws) -> dict:
    return {
        ws.cell(1, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }


def ensure_new_columns(ws, headers: dict) -> dict:
    existing = set(headers.keys())
    if all(col in existing for col in NEW_COLS):
        return headers

    insert_at = headers.get(COL_AFTER, ws.max_column) + 1
    to_add = [col for col in NEW_COLS if col not in existing]

    ws.insert_cols(insert_at, len(to_add))
    for i, col_name in enumerate(to_add):
        cell = ws.cell(1, insert_at + i, col_name)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1A3A5C")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 14

    return read_headers(ws)


# ══════════════════════════════════════════════════════════════════════════════
#  ОБРАБОТКА ФАЙЛА
# ══════════════════════════════════════════════════════════════════════════════

def process_file(page, src: Path, dst: Path,
                 row_range: tuple | None, debug: bool):
    if not dst.exists():
        shutil.copy2(src, dst)
        print(f"  → Скопирован: {dst.name}")

    wb = openpyxl.load_workbook(dst)
    ws = wb.active

    headers = read_headers(ws)
    headers = ensure_new_columns(ws, headers)

    desc_col  = headers.get(COL_AFTER)
    name_col  = headers.get("Наименование")
    param_col = headers.get("Параметры")
    w_col     = headers.get("Вес, г")
    l_col     = headers.get("Длина, мм")
    wd_col    = headers.get("Ширина, мм")
    h_col     = headers.get("Высота, мм")

    if not desc_col:
        print(f"  ⚠ Колонка «{COL_AFTER}» не найдена — пропуск")
        wb.close()
        return

    # Строки для обработки (есть Описание + новые колонки пустые)
    todos = []
    for r in range(2, ws.max_row + 1):
        if row_range and not (row_range[0] <= r <= row_range[1]):
            continue
        if not ws.cell(r, desc_col).value:
            continue
        if w_col and ws.cell(r, w_col).value is not None:
            continue
        todos.append(r)

    total = len(todos)
    if not todos:
        print(f"  Нечего обрабатывать")
        wb.close()
        return

    print(f"  Строк для обработки: {total}")

    processed  = 0
    save_timer = 0
    out_dir_dbg = dst.parent

    for batch_no, batch_start in enumerate(range(0, total, BATCH_SIZE), start=1):
        batch_rows = todos[batch_start : batch_start + BATCH_SIZE]

        items = []
        for local_idx, r in enumerate(batch_rows):
            items.append({
                "idx":    local_idx,
                "name":   ws.cell(r, name_col).value  if name_col  else "",
                "params": ws.cell(r, param_col).value if param_col else "",
                "desc":   ws.cell(r, desc_col).value  or "",
            })

        print(f"  Батч {batch_no} (стр. {batch_rows[0]}–{batch_rows[-1]})", end="", flush=True)

        # Новый чат перед каждым батчем — гарантия чистого контекста
        try:
            open_new_chat(page, debug=debug, out_dir=out_dir_dbg,
                          tag=f"b{batch_no}")
        except RuntimeError as e:
            print(f"\n  ✗ {e}")
            wb.save(dst)
            wb.close()
            raise

        prompt = build_prompt(items)

        try:
            type_prompt(page, prompt)
        except RuntimeError as e:
            print(f"\n  ✗ {e}")
            # Скриншот всегда — помогает понять что на странице
            try:
                shot = out_dir_dbg / f"_debug_b{batch_no}_noinput.png"
                page.screenshot(path=str(shot))
                print(f"    Скриншот сохранён: {shot}")
            except Exception:
                pass
            continue

        if debug:
            page.screenshot(path=str(out_dir_dbg / f"_debug_b{batch_no}_typed.png"))
            print(f"\n    Скриншот сохранён. Нажми Enter для отправки...")
            input("    > ")

        sent = click_send(page)

        # Убеждаемся что сообщение ушло — URL должен смениться на /c/xxxx
        try:
            page.wait_for_url("**/c/**", timeout=8000)
        except Exception:
            # Проверяем: может это лимит сообщений?
            if is_rate_limited(page):
                wait_for_rate_limit_reset(page)
                # Повторяем батч: открываем новый чат и пробуем снова
                try:
                    open_new_chat(page, debug=debug, out_dir=out_dir_dbg, tag=f"b{batch_no}r")
                    type_prompt(page, prompt)
                    click_send(page)
                    page.wait_for_url("**/c/**", timeout=8000)
                except Exception:
                    pass
            elif not sent:
                print(f"\n  ⚠ Сообщение, возможно, не отправлено (URL не изменился)")

        time.sleep(1)

        # До 3 попыток получить валидный ответ
        raw_response = ""
        results = []
        for attempt in range(1, 4):
            raw_response = wait_for_response(page)

            if not raw_response:
                print(f"  ✗ Пустой ответ (попытка {attempt}/3)")
            elif any(p in raw_response.lower() for p in SERVER_ERROR_PHRASES):
                print(f"  ⚠ Ошибка сервера ChatGPT (попытка {attempt}/3) — жду 30 сек...")
                time.sleep(30)
                # Повторяем запрос в новом чате
                if attempt < 3:
                    try:
                        open_new_chat(page, debug=debug, out_dir=out_dir_dbg, tag=f"b{batch_no}r{attempt}")
                        type_prompt(page, prompt)
                        click_send(page)
                        page.wait_for_url("**/c/**", timeout=8000)
                    except Exception:
                        pass
                    time.sleep(1)
                continue
            else:
                results = extract_json_items(raw_response)
                if results:
                    break
                print(f"  ⚠ JSON не распознан (попытка {attempt}/3)")
                print(f"    Начало ответа: {raw_response[:300]}")

            if attempt < 3:
                time.sleep(10)

        if debug and raw_response:
            (out_dir_dbg / f"_debug_b{batch_no}_response.txt").write_text(
                raw_response, encoding="utf-8"
            )

        if not results:
            print(f"  ✗ Батч пропущен после 3 попыток")
            continue

        result_map = {r["idx"]: r for r in results}
        filled = 0
        for local_idx, r in enumerate(batch_rows):
            data = result_map.get(local_idx)
            if not data:
                continue
            if w_col:  ws.cell(r, w_col).value  = data.get("вес")
            if l_col:  ws.cell(r, l_col).value  = data.get("длина")
            if wd_col: ws.cell(r, wd_col).value = data.get("ширина")
            if h_col:  ws.cell(r, h_col).value  = data.get("высота")
            filled += 1

        processed  += filled
        save_timer += 1
        print(f" → заполнено {filled}/{len(batch_rows)}")

        if save_timer >= SAVE_EVERY:
            wb.save(dst)
            print(f"    💾 Сохранено (итого: {processed})")
            save_timer = 0

        # Пауза между батчами (ChatGPT иногда ограничивает частоту запросов)
        time.sleep(3)

    wb.save(dst)
    print(f"  ✓ Итого: {processed}/{total} → {dst.name}")
    wb.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Заполнение веса и габаритов через ChatGPT (браузер)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--file",  default=None, help="Один файл (имя xlsx)")
    ap.add_argument("--rows",  default=None, help="Диапазон строк: 2-100")
    ap.add_argument("--login", action="store_true", help="Открыть браузер для входа в ChatGPT")
    ap.add_argument("--debug", action="store_true", help="Показывать браузер + скриншоты")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    src_files = [SRC_DIR / args.file] if args.file else sorted(SRC_DIR.glob("*.xlsx"))

    row_range = None
    if args.rows:
        parts = args.rows.split("-")
        row_range = (int(parts[0]), int(parts[1]) if len(parts) > 1 else int(parts[0]))

    print()
    print("═" * 64)
    print("  ChatGPT Dimensions Filler")
    print(f"  Батч:    {BATCH_SIZE} строк/запрос")
    print(f"  Файлов:  {len(src_files)}")
    print(f"  Из:      {SRC_DIR}")
    print(f"  В:       {OUT_DIR}")
    print("═" * 64)

    # Headless режим ChatGPT не поддерживает — Cloudflare блокирует автоматизацию.
    # Браузер всегда открывается видимым; можно свернуть окно и не трогать.
    with sync_playwright() as pw:
        context = pw.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=False,
            viewport={"width": 1280, "height": 900},
            locale="ru-RU",
            permissions=["clipboard-read", "clipboard-write"],
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = context.new_page()

        # ── Режим входа ──────────────────────────────────────────────────────
        if args.login:
            print("\n  Открываю ChatGPT...")
            print("  Войди в свой аккаунт.")
            page.goto(CHATGPT_URL, wait_until="domcontentloaded")
            print("  Когда окажешься в чате — нажми Enter здесь.")
            input("  > ")
            print("  ✓ Сессия сохранена. Теперь можно запускать без --login.")
            context.close()
            return

        # ── Проверка авторизации ─────────────────────────────────────────────
        print("\n  Открываю ChatGPT...")
        try:
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30000)
        except PWTimeout:
            print("  ✗ Таймаут. Проверь интернет-соединение.")
            context.close()
            sys.exit(1)

        time.sleep(3)
        if any(kw in page.url.lower() for kw in ("login", "signin", "auth")):
            print("\n  ✗ Не авторизован!")
            print("  Запусти сначала:")
            print("    uv run --with playwright,openpyxl scripts/chatgpt_dimensions_filler.py --login")
            context.close()
            sys.exit(1)

        print("  ✓ Авторизован\n")

        # ── Основной цикл ────────────────────────────────────────────────────
        total_files = len(src_files)
        for i, src in enumerate(src_files, 1):
            if not src.exists():
                print(f"[{i:3}/{total_files}] ✗ Не найден: {src.name}")
                continue

            dst = OUT_DIR / src.name
            print(f"[{i:3}/{total_files}] {src.name}")

            try:
                process_file(page, src, dst, row_range, debug=args.debug)
            except KeyboardInterrupt:
                print("\n\n  Прерывание — прогресс сохранён в последнем 💾")
                context.close()
                sys.exit(0)
            except RuntimeError as e:
                print(f"  ✗ {e}")
                context.close()
                sys.exit(1)
            except Exception as e:
                print(f"  ✗ Неожиданная ошибка: {e}")
                traceback.print_exc()

            print()

        context.close()

    print("═" * 64)
    print("  Всё готово!")
    print(f"  Результаты: {OUT_DIR}")
    print("═" * 64)
    print()


if __name__ == "__main__":
    main()
