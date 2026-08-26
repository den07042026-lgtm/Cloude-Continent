# /// script
# requires-python = ">=3.10"
# dependencies = ["playwright", "openpyxl"]
# ///
"""
gpt_ozon500_applicability.py
══════════════════════════════
Дозаполняет столбец «Применяемость» в "500 обогатитель.xlsx" через ChatGPT
(браузер, Playwright, тот же CDP-механизм, что и gpt_ozon500_descriptions.py) —
только для строк, где применяемость не удалось получить с сайта Mikado
(fill_applicability_mikado.py).

Точность приоритетнее полноты: промпт требует включать только автомобили,
подтверждённые источником, и оставлять пусто, если ничего подтвердить не
удалось — никаких выдуманных моделей.

Требует уже запущенный "тестовый" Chrome с CDP-портом 9222, авторизованный в
ChatGPT (scripts/launch_chatgpt.bat).

Тест:
  uv run --with playwright,openpyxl scripts/gpt_ozon500_applicability.py --rows 12-15 --debug

Обычный запуск (только пустые строки, продолжает после прерывания):
  uv run --with playwright,openpyxl scripts/gpt_ozon500_applicability.py --delay 60
"""

import re
import sys
import time
import argparse
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).parent))

from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Alignment

from gpt_ozon500_descriptions import (
    CHATGPT_URL, CDP_PORT, SAVE_EVERY,
    load_category_map, dismiss_modals, type_message, enable_web_search,
    click_send, wait_for_response, safe_save, _is_rate_limited,
)

XLSX_PATH = Path(r"C:\Users\Admin\Desktop\Озон-500\500 обогатитель.xlsx")

_APPLICABILITY_PROMPT_TEMPLATE = """Ты — специалист по подбору автозапчастей. Нужно определить точный список автомобилей (марка, модель, поколение/кузов, годы выпуска), для которых подходит указанная деталь. Точность здесь важнее полноты — покупатель будет ориентироваться на этот список при выборе, ошибка приведёт к покупке неподходящей детали и возврату.

Исходные данные:

Наименование: {name}

Бренд: {brand}

Артикул: {code}

Категория детали: {category}

Выполни поиск по открытым каталогам подбора запчастей (TecDoc, Exist, Emex, Autodoc, каталог производителя {brand} и аналогичные источники) и найди применяемость именно для артикула {code} бренда {brand}.

Правила:
- Включай в ответ ТОЛЬКО автомобили, совместимость которых подтверждена хотя бы одним источником.
- Если данные из разных источников расходятся — включай только те позиции, которые подтверждаются согласованно.
- Ничего не выдумывай и не обобщай "по аналогии" с похожими артикулами.
- Если не удалось найти ни одного подтверждённого варианта - оставь список пустым (ничего не пиши после двоеточия). КАТЕГОРИЧЕСКИ ЗАПРЕЩЕНО писать фразы "не найдено", "данные отсутствуют", "информация не найдена" и подобные - в ответе не должно быть ни слова о том, чего не удалось найти.

Ответ дай СТРОГО одной строкой в формате:

ПРИМЕНЯЕМОСТЬ: МАРКА МОДЕЛЬ (кузов/поколение, если известно) (годы выпуска); МАРКА2 МОДЕЛЬ2 (годы) - и так далее через "; "

Без заголовков, пояснений, markdown, до или после строки ответа."""


def build_prompt(row: dict) -> str:
    return _APPLICABILITY_PROMPT_TEMPLATE.format(
        name=row["name"],
        brand=row["brand"],
        code=row["code"],
        category=row.get("category") or "не указана",
    )


_APP_LINE_RE = re.compile(r'^\s*применяемость\s*:\s*(.*)$', re.I | re.M)
_EMPTY_MARKERS = {"-", "—", "нет", "нет данных", "не указано", "n/a", ""}


def parse_applicability(text: str) -> str | None:
    """Возвращает найденный список или '' (пусто, но распознано), либо None (формат не распознан)."""
    text = re.sub(r'\*+', '', text)
    m = _APP_LINE_RE.search(text)
    if not m:
        return None
    val = m.group(1).strip(" .;\n")
    if val.lower() in _EMPTY_MARKERS:
        return ""
    return val


def process_row(page, row: dict, debug: bool, retries: int = 2) -> tuple[bool, str]:
    """Возвращает (успех, текст). Текст может быть пустой строкой (подтверждённое отсутствие данных)."""
    for attempt in range(1, retries + 1):
        try:
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30_000)
            try:
                page.wait_for_load_state("networkidle", timeout=10_000)
            except Exception:
                pass
            time.sleep(4)
            dismiss_modals(page)

            try:
                page_text = page.locator("body").inner_text(timeout=3000)
            except Exception:
                page_text = ""
            if _is_rate_limited(page_text):
                print("    [!] ChatGPT: рейт-лимит (баннер страницы) — жду 20 мин...")
                time.sleep(1200)
                continue

            prompt = build_prompt(row)
            type_message(page, prompt)

            web_search_on = enable_web_search(page)
            if web_search_on:
                print("    [web search ON]", end=" ", flush=True)

            if debug and attempt == 1:
                print("    [debug] промпт введён, нажми Enter для отправки...")
                input()

            click_send(page)
            time.sleep(2)

            response = wait_for_response(page, timeout_sec=300 if web_search_on else 180)

            if not response:
                if attempt < retries:
                    print(f"    [!] Пустой ответ, попытка {attempt + 1}/{retries}...")
                    time.sleep(15)
                continue

            if _is_rate_limited(response):
                print("    [!] ChatGPT: рейт-лимит — жду 20 мин...")
                time.sleep(1200)
                continue

            parsed = parse_applicability(response)
            if parsed is None:
                print(f"    [!] Не распознан формат ответа, попытка {attempt}")
                if attempt < retries:
                    time.sleep(10)
                continue
            return True, parsed

        except RuntimeError as e:
            print(f"    [!] {e}")
            break
        except Exception as e:
            print(f"    [!] Ошибка (попытка {attempt}): {e}")
            if attempt < retries:
                time.sleep(15)
    return False, ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--debug", action="store_true", help="Пауза перед отправкой (проверка промпта)")
    ap.add_argument("--rows", default=None, help="Диапазон строк Excel: 2-10 или одна строка: 5")
    ap.add_argument("--delay", default=60, type=int, help="Пауза между строками, сек")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb.active
    category_map = load_category_map()

    col_code, col_name, col_brand, col_app = 1, 2, 3, 7

    max_row = ws.max_row
    row_start, row_end = 2, max_row
    if args.rows:
        if "-" in args.rows:
            a, b = args.rows.split("-")
            row_start, row_end = int(a), int(b)
        else:
            row_start = row_end = int(args.rows)
    row_end = min(row_end, max_row)

    print()
    print("=" * 62)
    print("  GPT Ozon-500 Applicability (добор для Mikado-пропусков)")
    print(f"  Файл:   {XLSX_PATH.name}")
    print(f"  Строки: {row_start}-{row_end} ({row_end - row_start + 1} шт.)")
    print("=" * 62)

    with sync_playwright() as pw:
        browser = pw.chromium.connect_over_cdp(f"http://127.0.0.1:{CDP_PORT}")
        context = browser.contexts[0] if browser.contexts else browser.new_context()

        page = None
        for p in context.pages:
            if "chatgpt.com" in p.url:
                page = p
                break
        if page is None:
            page = context.pages[0] if context.pages else context.new_page()

        print("\n  Открываю ChatGPT...")
        try:
            page.goto(CHATGPT_URL, wait_until="domcontentloaded", timeout=30_000)
        except Exception as e:
            print(f"  Не удалось открыть ChatGPT: {e}")
            sys.exit(1)

        time.sleep(3)
        if any(k in page.url.lower() for k in ("login", "auth", "signin", "sign-in")):
            print("  Не авторизован! Войди в ChatGPT в открытом окне тестового Chrome и перезапусти скрипт.")
            sys.exit(1)

        print("  Авторизован\n")

        stats = {"done": 0, "confirmed_empty": 0, "skipped": 0, "failed": 0}
        since_save = 0

        try:
            for row_num in range(row_start, row_end + 1):
                code = ws.cell(row_num, col_code).value
                if not code:
                    continue
                code = str(code).strip()
                name = str(ws.cell(row_num, col_name).value or "")
                brand = str(ws.cell(row_num, col_brand).value or "")

                if ws.cell(row_num, col_app).value:
                    print(f"  [{row_num:3}] {code}: пропуск (уже заполнено)")
                    stats["skipped"] += 1
                    continue

                print(f"\n  [{row_num:3}/{row_end}] {brand} {code} — {name[:45]}")

                category = category_map.get(code.lower(), "")
                row_data = {"code": code, "name": name, "brand": brand, "category": category}
                ok, text = process_row(page, row_data, debug=args.debug)

                if ok:
                    if text:
                        cell = ws.cell(row_num, col_app, text)
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                        print(f"         + {text[:70]}")
                        stats["done"] += 1
                    else:
                        print("         подтверждённых данных не нашлось (оставлено пустым)")
                        stats["confirmed_empty"] += 1
                    since_save += 1
                else:
                    print("         не удалось получить ответ")
                    stats["failed"] += 1

                if since_save >= SAVE_EVERY:
                    safe_save(wb, XLSX_PATH)
                    since_save = 0

                if row_num < row_end:
                    time.sleep(args.delay)
        finally:
            safe_save(wb, XLSX_PATH)

        # НЕ закрываем context/browser - это подключение по CDP к уже открытому
        # "тестовому" Chrome пользователя, а не браузер, запущенный этим скриптом.

    print()
    print("=" * 62)
    print("  Готово!")
    print(f"  Заполнено:          {stats['done']}")
    print(f"  Подтверждённо пусто: {stats['confirmed_empty']}")
    print(f"  Пропущено:          {stats['skipped']}")
    print(f"  Не удалось:         {stats['failed']}")
    print("=" * 62)
    print()


if __name__ == "__main__":
    main()
