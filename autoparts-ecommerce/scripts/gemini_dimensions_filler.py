"""
gemini_dimensions_filler.py
════════════════════════════
Добавляет столбцы «Вес, г», «Длина, мм», «Ширина, мм», «Высота, мм»
после столбца «Описание» — через Google Gemini API.

Бесплатный лимит: 1500 запросов/день, 15 запросов/мин.
При батче 30 строк = 45 000 строк/день — хватит на все файлы за один запуск.

Получить бесплатный API ключ: https://aistudio.google.com → Get API key

Исходная папка:    C:\\Users\\Admin\\Desktop\\На сортировку 26.04\\
Папка результатов: C:\\Users\\Admin\\Desktop\\На сортировку 08.05\\

Запуск:
  set GEMINI_API_KEY=AIza...
  uv run --with google-genai,openpyxl scripts/gemini_dimensions_filler.py

Один файл для теста:
  uv run --with google-genai,openpyxl scripts/gemini_dimensions_filler.py --file "Автобаферы.xlsx" --rows 2-30

Продолжение после прерывания:
  Запустить снова — пропустит уже заполненные строки.
"""

import os
import re
import sys
import json
import time
import shutil
import argparse
import traceback
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from google import genai
    from google.genai import types as genai_types
except ImportError:
    print("google-genai не установлен. Запусти:")
    print("  uv run --with google-genai,openpyxl scripts/gemini_dimensions_filler.py")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("openpyxl не установлен.")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

SRC_DIR    = Path(r"C:\Users\Admin\Desktop\На сортировку 26.04")
OUT_DIR    = Path(r"C:\Users\Admin\Desktop\На сортировку 08.05")
MODEL_NAME = "gemini-2.0-flash"

BATCH_SIZE  = 30   # строк за один запрос
SAVE_EVERY  = 5    # сохранять Excel каждые N батчей (~150 строк)
RPM_DELAY   = 4.5  # пауза между запросами (сек) — 15 RPM = 1 запрос/4 сек

NEW_COLS  = ["Вес, г", "Длина, мм", "Ширина, мм", "Высота, мм"]
COL_AFTER = "Описание"


# ══════════════════════════════════════════════════════════════════════════════
#  ПРОМПТ
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = (
    "Ты эксперт по автозапчастям. Для каждой позиции укажи "
    "вес и габариты ТРАНСПОРТНОЙ УПАКОВКИ — внешние размеры коробки или пакета "
    "вместе с самой деталью внутри, именно так, как товар будет отправлен покупателю. "
    "ВАЖНО: это НЕ размеры самой детали, а размеры упаковки снаружи. "
    "Если деталь идёт комплектом — вес и размер всего комплекта в упаковке. "
    "Отвечай ТОЛЬКО валидным JSON объектом без пояснений и markdown:\n"
    '{"items": [{"idx": 0, "вес": 500, "длина": 200, "ширина": 150, "высота": 80}, ...]}\n'
    "вес — граммы, длина/ширина/высота — мм, всё целые числа."
)


def build_prompt(items: list[dict]) -> str:
    lines = []
    for item in items:
        lines.append(f"[{item['idx']}]")
        lines.append(f"Наименование: {item['name']}")
        if item.get("params"):
            lines.append(f"Параметры: {str(item['params'])[:200]}")
        if item.get("desc"):
            lines.append(f"Описание: {str(item['desc'])[:250]}")
        lines.append("")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  ПАРСИНГ ОТВЕТА
# ══════════════════════════════════════════════════════════════════════════════

def extract_json_items(text: str) -> list[dict]:
    clean = re.sub(r"```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    clean = clean.replace("```", "").strip()

    ITEM_KEYS = {"вес", "длина", "ширина", "высота", "idx"}

    def is_valid(lst):
        return bool(lst) and isinstance(lst[0], dict) and bool(ITEM_KEYS & lst[0].keys())

    def try_parse(s):
        try:
            parsed = json.loads(s.strip())
            if isinstance(parsed, dict) and "items" in parsed:
                items = parsed["items"]
                if isinstance(items, list) and is_valid(items):
                    return items
            if isinstance(parsed, list) and is_valid(parsed):
                return parsed
        except Exception:
            pass
        return None

    result = try_parse(clean)
    if result:
        return result

    for start in [i for i, c in enumerate(clean) if c == '{']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == '{':
                depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result:
                        return result
                    break

    for start in [i for i, c in enumerate(clean) if c == '[']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == '[':
                depth += 1
            elif c == ']':
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result:
                        return result
                    break

    return []


# ══════════════════════════════════════════════════════════════════════════════
#  GEMINI API
# ══════════════════════════════════════════════════════════════════════════════

def ask_gemini(client, items: list[dict]) -> list[dict]:
    """Отправляет батч в Gemini, возвращает список items с габаритами."""
    user_text = build_prompt(items)

    for attempt in range(1, 4):
        try:
            response = client.models.generate_content(
                model=MODEL_NAME,
                contents=user_text,
                config=genai_types.GenerateContentConfig(
                    system_instruction=SYSTEM_PROMPT,
                    temperature=0.1,
                ),
            )
            raw = response.text.strip()
            results = extract_json_items(raw)
            if results:
                return results
            print(f"\n  ⚠ JSON не распознан (попытка {attempt}/3)")
            print(f"    Начало ответа: {raw[:300]}")
        except Exception as e:
            err = str(e).lower()
            if "quota" in err or "rate" in err or "429" in err:
                wait = 60 * attempt
                print(f"\n  ⏳ Лимит запросов — жду {wait} сек...")
                time.sleep(wait)
            else:
                print(f"\n  ⚠ Ошибка API (попытка {attempt}/3): {e}")
                time.sleep(10 * attempt)

    return []


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ХЕЛПЕРЫ
# ══════════════════════════════════════════════════════════════════════════════

def read_headers(ws) -> dict:
    return {
        ws.cell(1, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }


def ensure_new_columns(ws, headers: dict) -> dict:
    existing = set(headers.keys())
    if all(col in existing for col in NEW_COLS):
        return headers

    insert_at = headers.get(COL_AFTER, ws.max_column) + 1
    to_add = [col for col in NEW_COLS if col not in existing]

    ws.insert_cols(insert_at, len(to_add))
    for i, col_name in enumerate(to_add):
        cell = ws.cell(1, insert_at + i, col_name)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1A3A5C")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 14

    return read_headers(ws)


# ══════════════════════════════════════════════════════════════════════════════
#  ОБРАБОТКА ФАЙЛА
# ══════════════════════════════════════════════════════════════════════════════

def process_file(client, src: Path, dst: Path, row_range: tuple | None):
    if not dst.exists():
        shutil.copy2(src, dst)
        print(f"  → Скопирован: {dst.name}")

    wb = openpyxl.load_workbook(dst)
    ws = wb.active

    headers = read_headers(ws)
    headers = ensure_new_columns(ws, headers)

    desc_col  = headers.get(COL_AFTER)
    name_col  = headers.get("Наименование")
    param_col = headers.get("Параметры")
    w_col     = headers.get("Вес, г")
    l_col     = headers.get("Длина, мм")
    wd_col    = headers.get("Ширина, мм")
    h_col     = headers.get("Высота, мм")

    if not desc_col:
        print(f"  ⚠ Колонка «{COL_AFTER}» не найдена — пропуск")
        wb.close()
        return

    todos = []
    for r in range(2, ws.max_row + 1):
        if row_range and not (row_range[0] <= r <= row_range[1]):
            continue
        if not ws.cell(r, desc_col).value:
            continue
        if w_col and ws.cell(r, w_col).value is not None:
            continue
        todos.append(r)

    total = len(todos)
    if not todos:
        print(f"  Нечего обрабатывать")
        wb.close()
        return

    print(f"  Строк для обработки: {total}")

    processed  = 0
    save_timer = 0

    for batch_no, batch_start in enumerate(range(0, total, BATCH_SIZE), start=1):
        batch_rows = todos[batch_start : batch_start + BATCH_SIZE]

        items = []
        for local_idx, r in enumerate(batch_rows):
            items.append({
                "idx":    local_idx,
                "name":   ws.cell(r, name_col).value  if name_col  else "",
                "params": ws.cell(r, param_col).value if param_col else "",
                "desc":   ws.cell(r, desc_col).value  or "",
            })

        print(f"  Батч {batch_no} (стр. {batch_rows[0]}–{batch_rows[-1]})", end="", flush=True)

        results = ask_gemini(client, items)

        if not results:
            print(f" ✗ пропущен")
            continue

        result_map = {r["idx"]: r for r in results}
        filled = 0
        for local_idx, r in enumerate(batch_rows):
            data = result_map.get(local_idx)
            if not data:
                continue
            if w_col:  ws.cell(r, w_col).value  = data.get("вес")
            if l_col:  ws.cell(r, l_col).value  = data.get("длина")
            if wd_col: ws.cell(r, wd_col).value = data.get("ширина")
            if h_col:  ws.cell(r, h_col).value  = data.get("высота")
            filled += 1

        processed  += filled
        save_timer += 1
        print(f" → {filled}/{len(batch_rows)}")

        if save_timer >= SAVE_EVERY:
            wb.save(dst)
            print(f"    💾 Сохранено ({processed} строк)")
            save_timer = 0

        time.sleep(RPM_DELAY)

    wb.save(dst)
    print(f"  ✓ Итого: {processed}/{total} → {dst.name}")
    wb.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Заполнение веса и габаритов через Google Gemini API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--file", default=None, help="Один файл (имя xlsx)")
    ap.add_argument("--rows", default=None, help="Диапазон строк: 2-100")
    args = ap.parse_args()

    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        # Пробуем .env в папке проекта
        env_path = Path(__file__).parent.parent / ".env"
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8").splitlines():
                if line.startswith("GEMINI_API_KEY="):
                    api_key = line.split("=", 1)[1].strip()
                    break

    if not api_key:
        print("\n  ✗ API ключ не найден!")
        print("  Получи бесплатный ключ на https://aistudio.google.com")
        print("  Затем задай его:")
        print("    set GEMINI_API_KEY=AIza...")
        print("  Или добавь строку GEMINI_API_KEY=AIza... в файл .env проекта")
        sys.exit(1)

    client = genai.Client(api_key=api_key)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    src_files = [SRC_DIR / args.file] if args.file else sorted(SRC_DIR.glob("*.xlsx"))

    row_range = None
    if args.rows:
        parts = args.rows.split("-")
        row_range = (int(parts[0]), int(parts[1]) if len(parts) > 1 else int(parts[0]))

    print()
    print("═" * 64)
    print("  Gemini Dimensions Filler")
    print(f"  Модель:  {MODEL_NAME}")
    print(f"  Батч:    {BATCH_SIZE} строк/запрос")
    print(f"  Файлов:  {len(src_files)}")
    print(f"  Из:      {SRC_DIR}")
    print(f"  В:       {OUT_DIR}")
    print("═" * 64)
    print()

    for i, src in enumerate(src_files, 1):
        if not src.exists():
            print(f"[{i:3}/{len(src_files)}] ✗ Не найден: {src.name}")
            continue

        dst = OUT_DIR / src.name
        print(f"[{i:3}/{len(src_files)}] {src.name}")

        try:
            process_file(client, src, dst, row_range)
        except KeyboardInterrupt:
            print("\n\n  Прерывание — прогресс сохранён в последнем 💾")
            sys.exit(0)
        except Exception as e:
            print(f"  ✗ Ошибка: {e}")
            traceback.print_exc()

        print()

    print("═" * 64)
    print("  Всё готово!")
    print(f"  Результаты: {OUT_DIR}")
    print("═" * 64)
    print()


if __name__ == "__main__":
    main()
