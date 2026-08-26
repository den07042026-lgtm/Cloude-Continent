"""
mikado_fetch_photos.py
════════════════════════════════════════════════════════════════════════════
Точечно докачивает фото с mikado-parts.ru по списку конкретных артикулов
(не по всему прайсу), для тех позиций топ-500 Ozon, для которых фото не
нашлось нигде на компьютере (см. без_фото.txt).

Переиспользует login()/fetch_product()/fetch_images() из mikado_scraper.py.

Запуск:
  cd C:\\Users\\Admin\\Documents\\Autoparts_Ecommerce
  uv run --with requests scripts/mikado_fetch_photos.py --codes "C:\\Users\\Admin\\Desktop\\Озон-500\\без_фото.txt" --out "C:\\Users\\Admin\\Desktop\\Озон-500\\Изображения"
"""

import sys
import time
import argparse
import logging
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from mikado_scraper import login, fetch_product, fetch_images, load_env

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-7s %(message)s")
log = logging.getLogger(__name__)

ROOT = Path(__file__).parent.parent
ENV_FILE = ROOT / ".env"
MIKADO_PRICE_FRESH = ROOT / "data" / "analytics" / "top500_ozon" / "mikado_price_fresh.xlsx"


def read_codes(path: Path) -> list[str]:
    lines = path.read_text(encoding="utf-8").splitlines()
    codes = []
    for line in lines[1:]:  # пропускаем заголовок "Код\tБренд\tНаименование"
        if not line.strip():
            continue
        code = line.split("\t")[0].strip()
        if code:
            codes.append(code)
    return codes


def read_all_codes_from_excel(path: Path) -> list[str]:
    """Все 500 кодов из топ-500 Excel (столбец B листа Топ-500), не только без_фото.txt."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Топ-500"]
    codes = [str(r[1]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[1]]
    wb.close()
    return codes


def load_code_to_prodnum(price_path: Path) -> dict[str, str]:
    """Сайт Mikado принимает в URL Prodnum (с брендовым префиксом типа xzk-of-4434g),
    а НЕ Code (of-4434g) - это разные колонки прайса."""
    wb = openpyxl.load_workbook(price_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    mapping = {}
    for r in rows:
        code = r[idx["Code"]]
        prodnum = r[idx["Prodnum"]]
        if code and prodnum:
            mapping[str(code).strip().lower()] = str(prodnum).strip()
    wb.close()
    return mapping


# Захардкожено вместо required=True CLI-аргументов: кириллица в путях ломается
# при передаче через argv (и в Bash, и в PowerShell) - падает FileNotFoundError
# с мусорными байтами вместо кириллицы. Пути внутри .py-файла читаются нормально.
DEFAULT_CODES = Path(r"C:\Users\Admin\Desktop\Озон-500\без_фото.txt")
DEFAULT_OUT = Path(r"C:\Users\Admin\Desktop\Озон-500\Изображения")
EXCEL_PATH = Path(r"C:\Users\Admin\Desktop\Озон-500\Топ-500_Ozon_2026-07-06.xlsx")
ALL_OUT = Path(r"C:\Users\Admin\Desktop\Озон-500\Изображения\mikado_all")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--codes", default=None, help="Путь к файлу без_фото.txt (по умолчанию: DEFAULT_CODES)")
    ap.add_argument("--out", default=None, help="Папка для сохранения фото (по умолчанию: DEFAULT_OUT)")
    ap.add_argument("--delay", type=float, default=1.5, help="Пауза между запросами, сек")
    ap.add_argument("--all", action="store_true",
                    help="Прогнать ВСЕ 500 позиций из Excel (не только без_фото.txt), "
                         "сохранить в отдельную папку Изображения/mikado_all для сравнения источников")
    args = ap.parse_args()

    if args.all:
        codes_path = None
        out_dir = ALL_OUT
        out_dir.mkdir(parents=True, exist_ok=True)
        codes = read_all_codes_from_excel(EXCEL_PATH)
        log.info("Режим --all: все позиции из Excel = %d", len(codes))
    else:
        codes_path = Path(args.codes) if args.codes else DEFAULT_CODES
        out_dir = Path(args.out) if args.out else DEFAULT_OUT
        out_dir.mkdir(parents=True, exist_ok=True)
        codes = read_codes(codes_path)
        log.info("Артикулов к докачке: %d", len(codes))

    code_to_prodnum = load_code_to_prodnum(MIKADO_PRICE_FRESH)
    log.info("Загружена карта Code->Prodnum: %d записей", len(code_to_prodnum))

    env = load_env(ENV_FILE)
    mikado_code = env.get("MIKADO_CODE", "")
    mikado_password = env.get("MIKADO_PASSWORD", "")
    if not mikado_code or not mikado_password:
        sys.exit("Нет MIKADO_CODE/MIKADO_PASSWORD в .env")

    if args.all:
        already = {p.stem.rsplit("_", 1)[0] for p in out_dir.iterdir()} if out_dir.exists() else set()
        before = len(codes)
        codes = [c for c in codes if c not in already]
        log.info("Уже скачано ранее (--all, пропускаем): %d, осталось: %d", before - len(codes), len(codes))

    session = login(mikado_code, mikado_password)

    found, not_found, errors = [], [], []
    for i, code in enumerate(codes, 1):
        prodnum = code_to_prodnum.get(code.lower())
        if not prodnum:
            log.warning("[%d/%d] %s - нет Prodnum в прайсе, пропуск", i, len(codes), code)
            not_found.append(code)
            continue
        try:
            html, data = fetch_product(session, prodnum)
            artid = data.get("_artid")
            if not artid:
                log.warning("[%d/%d] %s (%s) - нет ARTID (страница не найдена?)", i, len(codes), code, prodnum)
                not_found.append(code)
                time.sleep(args.delay)
                continue
            # fetch_images сохраняет как {code}_N.ext - передаём НАШ code, чтобы имена
            # файлов совпадали с колонкой "Код Mikado" в итоговом Excel top-500.
            saved = fetch_images(session, artid, code, out_dir)
            if saved:
                log.info("[%d/%d] %s (%s) - сохранено %d фото", i, len(codes), code, prodnum, len(saved))
                found.append(code)
            else:
                log.info("[%d/%d] %s (%s) - фото на сайте нет", i, len(codes), code, prodnum)
                not_found.append(code)
        except Exception as e:
            log.warning("[%d/%d] %s (%s) - ошибка: %s", i, len(codes), code, prodnum, e)
            errors.append(code)
        time.sleep(args.delay)

    log.info("=" * 60)
    log.info("Готово. Найдено фото: %d / %d", len(found), len(codes))
    log.info("Без фото на сайте Mikado: %d", len(not_found))
    log.info("Ошибок: %d", len(errors))

    if not args.all:
        still_missing = out_dir.parent / "без_фото_после_mikado.txt"
        with open(still_missing, "w", encoding="utf-8") as f:
            f.write("Код\n")
            for c in not_found + errors:
                f.write(c + "\n")
        log.info("Список всё ещё без фото: %s", still_missing)


if __name__ == "__main__":
    main()
