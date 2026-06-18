"""
wb_catalog_mapper.py
════════════════════════════════════════════════════════════════════════════
Таблица соответствия: наш каталог (Mikado + Автолига) → ниши WB.

Этап 1.5 — без nmID (WB search заблокирован), но с полными данными ниш:
  • артикул/OEM, бренд, название, остаток, цена
  • сопоставленный предмет WB, OOS%, продажи/мес, средняя цена, маржа
  • приоритет = OOS% × ln(продажи+1) × max(0, маржа/12%)

Выход (data/analytics/):
  wb_catalog_map_<ts>.xlsx  — 3 листа:
    "Маппинг"     — каждый артикул + лучшая WB-ниша  (сортировка по приоритету)
    "Топ_нишам"   — агрегат по нишам: сколько наших SKU, оборот, OOS%
    "Для_поиска"  — список OEM-артикулов для wb_item_analyzer.py (когда WB разблокируется)
  wb_search_queue_<ts>.txt  — OEM-артикулы построчно (очередь для WB-поиска)

Запуск:
    uv run --with openpyxl --with requests --with python-dotenv --with xlrd \\
           scripts/wb_catalog_mapper.py [--min-price 100] [--min-oos 15] [--top 5000]
"""

import argparse
import json
import logging
import math
import os
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

try:
    import requests
except ImportError:
    print("pip install requests"); sys.exit(1)

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(*a, **kw): pass

sys.stdout.reconfigure(encoding="utf-8")

# ─── Пути ─────────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent.parent
load_dotenv(BASE_DIR / ".env")
MPSTATS_TOKEN = os.getenv("MPSTATS_TOKEN", "")
MIKADO_FILE   = Path("C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx")
CACHE_DIR     = BASE_DIR / "data" / "analytics" / "cache"
OUT_DIR       = BASE_DIR / "data" / "analytics"
LOG_FILE      = BASE_DIR / "logs" / "wb_catalog_mapper.log"

# ─── WB FBS тарифы ────────────────────────────────────────────────────────────
WB_ACQ_PCT       = 0.015
WB_TAX_PCT       = 0.06
WB_SPP_RESERVE   = 0.07
WB_RET_RATE      = 0.03
WB_PACKAGING     = 30.0
WB_DELIVERY_BASE = 50.6
WB_DELIVERY_L    = 15.4
WB_RETURN_BASE   = 136.0
WB_RETURN_L      = 14.0
DEFAULT_VOLUME_L = 3.0
DEFAULT_COMM     = 0.30
TARGET_MARGIN    = 0.12

# ─── Авто-фильтры (из wb_niche_analyzer.py) ───────────────────────────────────
STOP_SUBJECTS = (
    "мотоцикл", "мотозапчаст", "мотото", "велосипед", "самокат",
    "скутер", "квадроцикл", "снегоход",
    "садовая техника", "мойки высокого давления", "запчасти для садов",
    "пылесос", "стиральн", "холодильник", "чайник",
    "ручной инструмент", "шуруповерт", "перфоратор", "дрель",
    "электроинструмент", "строительн", "кровельн",
    "канцеляр", "офис", "школьн",
    "косметик", "парфюм", "бытов химия",
)
AUTO_BROAD_KW = (
    "автозапчаст", "запчаст", "колодк", "амортизатор", "тормоз",
    "подвеск", "рулев", "сцеплен", "ремен",
    "шаровая", "шаровой", "стойк стабил",
    "свеч", "зажигани", "фильтр масл", "фильтр воздуш",
    "фильтр салон", "фильтр топлив",
    "подшипник ступ", "ступиц",
    "сальник", "прокладк", "крепеж для авт",
    "зеркала автомоб", "зеркало автомоб",
    "фары автомоб", "фонар автомоб",
    "датчик", "катушк зажиган",
    "насос топлив", "бензонасос",
    "цилиндр автомоб", "реле автомоб",
    "тросов", "привод шруса", "шрус",
    "масла моторн", "масла трансмисс", "антифриз", "тосол",
    "разъем автомоб", "ремкомплект",
    "уплотнитель автомоб", "предохранител автомоб", "аккумулятор",
    "масл", "жидкост", "охлаждени", "топлив",
    "автомоб", "автоэлектр", "мотор", "двигател", "коробк",
)
NAME_STOPWORDS = frozenset({
    "для", "авто", "автомобильные", "автомобильное", "автомобильный",
    "автомобиля", "автомобилей", "и", "в", "на", "с", "из",
    "запчасти", "запчастей", "к", "по", "деталь", "детали",
    "ремонт", "комплект", "крепеж",
})

log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  Утилиты
# ══════════════════════════════════════════════════════════════════════════════

def calc_wb_margin(purchase: float, sell: float,
                   volume_l: float = DEFAULT_VOLUME_L,
                   commission: float = DEFAULT_COMM) -> float:
    if sell <= 0 or purchase <= 0:
        return 0.0
    delivery    = WB_DELIVERY_BASE + WB_DELIVERY_L * volume_l
    return_cost = (WB_RETURN_BASE + WB_RETURN_L * volume_l) * WB_RET_RATE
    com  = sell * commission
    acq  = sell * WB_ACQ_PCT
    spp  = sell * WB_SPP_RESERVE
    proc = sell - com - acq - spp - delivery
    tax  = max(0.0, proc) * WB_TAX_PCT
    cost = purchase + com + acq + spp + delivery + return_cost + WB_PACKAGING + tax
    return (sell - cost) / sell


def score_niche(oos_pct: float, sales: float, margin: float) -> float:
    return (oos_pct / 100.0) * math.log1p(sales) * max(0.0, margin / TARGET_MARGIN)


def _normalize_oem(s: str) -> str:
    return s.replace(" ", "").replace("-", "").replace(".", "").upper().strip()


def is_auto_subject(name: str) -> bool:
    low = name.lower()
    if any(kw in low for kw in STOP_SUBJECTS):
        return False
    return any(kw in low for kw in AUTO_BROAD_KW)


def _subject_stems(subject_name: str) -> list[str]:
    parts = subject_name.split("/")
    search_part = parts[-1].strip().lower() if len(parts) > 1 else subject_name.lower()
    words = [w.strip(".,;:()-") for w in search_part.split()]
    return [w[:6] for w in words if w not in NAME_STOPWORDS and len(w) >= 4]


def item_name_lower(item: dict) -> str:
    return item.get("name_lower") or item.get("name", "").lower()


def match_item_to_subject(item_name_lc: str, subject_stems: list[str]) -> bool:
    return any(s in item_name_lc for s in subject_stems)


# ══════════════════════════════════════════════════════════════════════════════
#  MPStats
# ══════════════════════════════════════════════════════════════════════════════

def fetch_subjects(d1: str, d2: str, force: bool = False) -> list[dict]:
    cache = CACHE_DIR / f"wb_auto_subjects_{d1}.json"
    if not force and cache.exists():
        age = (time.time() - cache.stat().st_mtime) / 3600
        if age < 24:
            log.info(f"Кэш предметов: {cache.name}")
            return json.loads(cache.read_text(encoding="utf-8"))

    log.info("MPStats: загружаем предметы Автотовары...")
    url    = "https://mpstats.io/api/wb/get/category/items"
    params = {"path": "Автотовары", "d1": d1, "d2": d2, "startRow": 0, "endRow": 5000}
    hdrs   = {"X-Mpstats-TOKEN": MPSTATS_TOKEN}

    for attempt in range(3):
        try:
            r = requests.get(url, headers=hdrs, params=params, timeout=60)
            if r.status_code == 429:
                time.sleep(60 * (attempt + 1)); continue
            if r.status_code != 200:
                log.error(f"MPStats {r.status_code}: {r.text[:200]}"); return []
            data = r.json()
            rows = data if isinstance(data, list) else (data.get("data") or data.get("items") or [])
            log.info(f"MPStats: {len(rows)} предметов")
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            cache.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")
            return rows
        except Exception as e:
            log.warning(f"MPStats попытка {attempt+1}: {e}"); time.sleep(5)
    return []


# ══════════════════════════════════════════════════════════════════════════════
#  Загрузка каталогов
# ══════════════════════════════════════════════════════════════════════════════

def load_mikado() -> list[dict]:
    """Возвращает [{oem, article, brand, name, name_lower, price, stock, source}]."""
    if not MIKADO_FILE.exists():
        log.warning(f"Mikado: файл не найден {MIKADO_FILE}"); return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(MIKADO_FILE), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        log.error(f"Mikado: {e}"); return []

    items = []
    hdr_map: dict[str, int] = {}
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            hdr = [str(c).strip().lower() if c else "" for c in row]
            for i, h in enumerate(hdr):
                if h in ("prodnum", "code", "brandname", "prodname", "priceout", "qty"):
                    hdr_map[h] = i
            if "brandname" in hdr_map and "priceout" in hdr_map:
                header_found = True
            continue

        def _get(key, default=""):
            idx = hdr_map.get(key)
            if idx is None: return default
            v = row[idx]
            return v if v is not None else default

        try:
            price = float(_get("priceout", 0))
        except (TypeError, ValueError):
            continue
        if price <= 0:
            continue

        article_raw = str(_get("prodnum", "")).strip()   # f-a22025
        code_raw    = str(_get("code",    "")).strip()   # a22025
        brand       = str(_get("brandname", "")).strip()
        name        = str(_get("prodname",  "")).strip()
        try:
            stock = float(_get("qty", 0))
        except (TypeError, ValueError):
            stock = 0.0

        # OEM = нормализованный code
        oem = _normalize_oem(code_raw) if code_raw else _normalize_oem(article_raw)
        if not oem or not brand:
            continue

        items.append({
            "oem":       oem,
            "article":   article_raw or code_raw,
            "brand":     brand,
            "name":      name,
            "name_lower": name.lower(),
            "price":     price,
            "stock":     stock,
            "source":    "mikado",
        })

    try: wb.close()
    except Exception: pass
    log.info(f"Mikado: {len(items):,} позиций")
    return items


def load_autoliga_items() -> list[dict]:
    """Возвращает [{oem, article, brand, name, name_lower, price, stock, source}]."""
    scripts_dir = Path(__file__).parent
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    try:
        from autoliga_loader import load_autoliga
    except ImportError as e:
        log.error(f"autoliga_loader: {e}"); return []

    raw = load_autoliga()
    items = []
    for v in raw.values():
        price = float(v.get("price") or 0)
        brand = v.get("brand", "")
        if price <= 0 or not brand:
            continue
        name = v.get("name", "")
        items.append({
            "oem":       v.get("oem", ""),
            "article":   v.get("article", ""),
            "brand":     brand,
            "name":      name,
            "name_lower": name.lower(),
            "price":     price,
            "stock":     float(v.get("stock") or 0),
            "source":    "autoliga",
        })
    log.info(f"Автолига: {len(items):,} позиций")
    return items


# ══════════════════════════════════════════════════════════════════════════════
#  Маппинг
# ══════════════════════════════════════════════════════════════════════════════

def build_subject_index(subjects: list[dict], min_sales: float, min_oos: float) -> list[dict]:
    """
    Строит список предметов с рассчитанными стемами и скором.
    """
    result = []
    for s in subjects:
        name     = s.get("name", "")
        sales    = float(s.get("sales", 0) or 0)
        oos_pct  = float(s.get("lost_profit_percent", 0) or 0)
        avg_price= float(s.get("avg_price", 0) or 0)
        if not is_auto_subject(name): continue
        if sales < min_sales: continue
        if oos_pct < min_oos: continue
        if avg_price <= 0: continue

        comm_pct = float(s.get("commision_fbs", DEFAULT_COMM * 100) or (DEFAULT_COMM * 100)) / 100.0
        stems    = _subject_stems(name)
        if not stems: continue

        result.append({
            "subject":    name,
            "subject_id": s.get("id", ""),
            "sales":      int(sales),
            "oos_pct":    round(oos_pct, 1),
            "avg_price":  round(avg_price),
            "comm_pct":   comm_pct,
            "sellers":    int(s.get("sellers", 0) or 0),
            "items_wb":   int(s.get("items", 0) or 0),
            "revenue":    round(float(s.get("revenue", 0) or 0)),
            "stems":      stems,
        })

    result.sort(key=lambda x: x["oos_pct"] * math.log1p(x["sales"]), reverse=True)
    log.info(f"Предметов для маппинга: {len(result)}")
    return result


def map_items(catalog: list[dict], subject_index: list[dict],
              price_min_ratio: float = 0.05,
              price_max_ratio: float = 8.0) -> list[dict]:
    """
    Для каждого товара каталога находит лучшую WB-нишу.
    Возвращает список строк для Excel.
    """
    rows = []
    for item in catalog:
        name_lc  = item["name_lower"]
        our_price = item["price"]

        best = None
        best_sc = -1.0

        for subj in subject_index:
            if not match_item_to_subject(name_lc, subj["stems"]):
                continue
            avg_p = subj["avg_price"]
            if avg_p > 0:
                ratio = our_price / avg_p
                if ratio < price_min_ratio or ratio > price_max_ratio:
                    continue
            else:
                ratio = 0.0
            margin = calc_wb_margin(our_price, avg_p, commission=subj["comm_pct"])
            sc = score_niche(subj["oos_pct"], subj["sales"], margin)
            if sc > best_sc:
                best_sc = sc
                best = {**subj, "margin": margin, "score": sc, "price_ratio": round(ratio, 2)}

        row = {
            "oem":       item["oem"],
            "article":   item["article"],
            "brand":     item["brand"],
            "name":      item["name"],
            "stock":     item["stock"],
            "our_price": round(our_price),
            "source":    item["source"],
        }

        if best:
            row.update({
                "subject":     best["subject"],
                "oos_pct":     best["oos_pct"],
                "sales":       best["sales"],
                "avg_price":   best["avg_price"],
                "price_ratio": best.get("price_ratio", 0.0),
                "margin_pct":  round(best["margin"] * 100, 1),
                "comm_pct":    round(best["comm_pct"] * 100, 1),
                "sellers":     best["sellers"],
                "items_wb":    best["items_wb"],
                "score":       round(best_sc, 4),
            })
        else:
            row.update({
                "subject":     "",
                "oos_pct":     0.0,
                "sales":       0,
                "avg_price":   0,
                "price_ratio": 0.0,
                "margin_pct":  0.0,
                "comm_pct":    0.0,
                "sellers":     0,
                "items_wb":    0,
                "score":       0.0,
            })

        rows.append(row)

    rows.sort(key=lambda x: x["score"], reverse=True)
    log.info(f"Строк в маппинге: {len(rows):,}")
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  Excel-экспорт
# ══════════════════════════════════════════════════════════════════════════════

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(color="FFFFFF", bold=True)
RED_FILL   = PatternFill("solid", fgColor="C00000")
ORANGE_FILL= PatternFill("solid", fgColor="FFA500")
YELLOW_FILL= PatternFill("solid", fgColor="FFD700")
GREEN_FILL = PatternFill("solid", fgColor="70AD47")
LIGHTRED   = PatternFill("solid", fgColor="FFCCCC")
WHITE_FONT = Font(color="FFFFFF")
BOLD_WHITE = Font(color="FFFFFF", bold=True)


def _write_header(ws, cols):
    ws.append([c[1] for c in cols])
    for i, c in enumerate(cols, 1):
        cell = ws.cell(1, i)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28


def write_mapping_sheet(ws, rows: list[dict], top: int) -> None:
    cols = [
        ("score",      "Приоритет",    10),
        ("brand",      "Бренд",        14),
        ("article",    "Артикул",      16),
        ("oem",        "OEM",          16),
        ("name",       "Название",     40),
        ("stock",      "Остаток",       8),
        ("our_price",  "Наша цена",    10),
        ("subject",    "Ниша WB",      38),
        ("oos_pct",    "OOS%",          7),
        ("sales",      "Прод/мес",      9),
        ("avg_price",  "Ср.цена WB",   11),
        ("margin_pct", "Маржа%",        8),
        ("comm_pct",   "Ком.%",         7),
        ("price_ratio", "Цена/WBср",     9),
        ("sellers",    "Продавцов",    10),
        ("items_wb",   "SKU на WB",    10),
        ("source",     "Источник",      9),
    ]
    _write_header(ws, cols)
    keys = [c[0] for c in cols]

    for r in rows[:top]:
        ws.append([r.get(k) for k in keys])

    # Цветовая подсветка OOS% (столбец 9)
    oos_c = 9
    for ri in range(2, min(top, len(rows)) + 2):
        v = float(ws.cell(ri, oos_c).value or 0)
        if v >= 50:
            ws.cell(ri, oos_c).fill = RED_FILL;    ws.cell(ri, oos_c).font = BOLD_WHITE
        elif v >= 40:
            ws.cell(ri, oos_c).fill = PatternFill("solid", fgColor="FF4444"); ws.cell(ri, oos_c).font = BOLD_WHITE
        elif v >= 25:
            ws.cell(ri, oos_c).fill = ORANGE_FILL
        elif v >= 15:
            ws.cell(ri, oos_c).fill = YELLOW_FILL

    # Маржа (столбец 12)
    mrg_c = 12
    for ri in range(2, min(top, len(rows)) + 2):
        v = float(ws.cell(ri, mrg_c).value or 0)
        if v >= 15:
            ws.cell(ri, mrg_c).fill = GREEN_FILL; ws.cell(ri, mrg_c).font = WHITE_FONT
        elif v >= 10:
            ws.cell(ri, mrg_c).fill = PatternFill("solid", fgColor="FFEB84")
        elif v < 0:
            ws.cell(ri, mrg_c).fill = LIGHTRED

    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = c[2]
    ws.freeze_panes = "A2"


def write_niches_sheet(ws, rows: list[dict]) -> None:
    """Агрегат по нишам: сколько наших SKU, суммарный остаток, OOS%, продажи."""
    by_niche: dict[str, dict] = {}
    for r in rows:
        subj = r.get("subject") or "— нет ниши —"
        if subj not in by_niche:
            by_niche[subj] = {
                "subject":   subj,
                "oos_pct":   r.get("oos_pct", 0),
                "sales":     r.get("sales", 0),
                "avg_price": r.get("avg_price", 0),
                "sku_count": 0,
                "in_stock":  0,
                "brands":    set(),
                "score_max": 0.0,
            }
        n = by_niche[subj]
        n["sku_count"] += 1
        if r.get("stock", 0) > 0:
            n["in_stock"] += 1
        if r.get("brand"):
            n["brands"].add(r["brand"])
        if r.get("score", 0) > n["score_max"]:
            n["score_max"] = r["score"]

    cols = [
        ("subject",   "Ниша WB",          40),
        ("oos_pct",   "OOS%",               7),
        ("sales",     "Прод/мес",           9),
        ("avg_price", "Ср.цена WB",        11),
        ("sku_count", "Наших SKU",         10),
        ("in_stock",  "В наличии",         10),
        ("brands_str","Наши бренды",       35),
        ("score_max", "Макс.скор",         10),
    ]
    _write_header(ws, cols)
    keys = [c[0] for c in cols]

    agg = sorted(by_niche.values(), key=lambda x: x["sku_count"], reverse=True)
    for n in agg:
        n["brands_str"] = ", ".join(sorted(n["brands"])[:8])
        ws.append([n.get(k) for k in keys])

    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = c[2]
    ws.freeze_panes = "A2"


def write_search_sheet(ws, rows: list[dict]) -> None:
    """OEM-артикулы с ненулевым приоритетом для wb_item_analyzer.py."""
    search_rows = [r for r in rows if r.get("score", 0) > 0 and r.get("oem")]
    # Дедуп по OEM, берём с макс. скором
    by_oem: dict[str, dict] = {}
    for r in search_rows:
        oem = r["oem"]
        if oem not in by_oem or r["score"] > by_oem[oem]["score"]:
            by_oem[oem] = r
    deduped = sorted(by_oem.values(), key=lambda x: x["score"], reverse=True)

    cols = [
        ("oem",       "OEM (для поиска)",  20),
        ("article",   "Артикул",           16),
        ("brand",     "Бренд",             14),
        ("name",      "Название",          40),
        ("subject",   "Ниша WB",           35),
        ("score",     "Приоритет",         10),
        ("oos_pct",   "OOS%",               7),
        ("sales",     "Прод/мес",           9),
        ("avg_price", "Ср.цена WB",        11),
        ("our_price", "Наша цена",         10),
        ("margin_pct","Маржа%",             8),
    ]
    _write_header(ws, cols)
    keys = [c[0] for c in cols]

    for r in deduped:
        ws.append([r.get(k) for k in keys])

    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = c[2]
    ws.freeze_panes = "A2"

    log.info(f"Очередь поиска WB: {len(deduped):,} уникальных OEM")
    return deduped


# ══════════════════════════════════════════════════════════════════════════════
#  main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="WB Catalog Mapper")
    parser.add_argument("--days",       type=int,   default=30)
    parser.add_argument("--min-sales",  type=float, default=200,
                        help="Мин.продаж/мес для предмета WB")
    parser.add_argument("--min-oos",    type=float, default=5.0,
                        help="Мин.OOS%% для предмета WB")
    parser.add_argument("--min-price",  type=float, default=100,
                        help="Мин.наша цена закупки (фильтр шурупов)")
    parser.add_argument("--top",        type=int,   default=5000,
                        help="Макс.строк в листе Маппинг")
    parser.add_argument("--price-min-ratio", type=float, default=0.05,
                        help="Мин. наша_цена/avg_wb (отсев дешевле 5%% ср.цены ниши)")
    parser.add_argument("--price-max-ratio", type=float, default=8.0,
                        help="Макс. наша_цена/avg_wb (отсев дороже 800%% ср.цены ниши)")
    parser.add_argument("--skip-mikado",   action="store_true")
    parser.add_argument("--skip-autoliga", action="store_true")
    parser.add_argument("--refresh",    action="store_true",
                        help="Обновить кэш MPStats")
    args = parser.parse_args()

    for d in (CACHE_DIR, OUT_DIR, LOG_FILE.parent):
        d.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
        datefmt="%H:%M:%S",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(str(LOG_FILE), encoding="utf-8", errors="replace"),
        ],
    )

    if not MPSTATS_TOKEN:
        log.error("MPSTATS_TOKEN не задан в .env"); sys.exit(1)

    d2 = datetime.now().strftime("%Y-%m-%d")
    d1 = (datetime.now() - timedelta(days=args.days)).strftime("%Y-%m-%d")
    log.info(f"Период: {d1}…{d2} | мин.продаж ниши: {args.min_sales} | мин.OOS: {args.min_oos}%% | мин.цена: {args.min_price}₽")

    # ─── 1. MPStats ───────────────────────────────────────────────────────────
    subjects_raw = fetch_subjects(d1, d2, force=args.refresh)
    if not subjects_raw:
        log.error("Нет данных MPStats"); sys.exit(1)

    subject_index = build_subject_index(subjects_raw, args.min_sales, args.min_oos)
    log.info(f"Предметов в индексе: {len(subject_index)}")

    # ─── 2. Каталог ───────────────────────────────────────────────────────────
    catalog: list[dict] = []
    if not args.skip_mikado:
        catalog.extend(load_mikado())
    if not args.skip_autoliga:
        catalog.extend(load_autoliga_items())

    # Фильтр по минимальной цене
    before = len(catalog)
    catalog = [it for it in catalog if it["price"] >= args.min_price]
    log.info(f"Каталог после фильтра цены ≥{args.min_price}₽: {len(catalog):,} (было {before:,})")

    if not catalog:
        log.error("Каталог пустой"); sys.exit(1)

    # ─── 3. Маппинг ───────────────────────────────────────────────────────────
    mapped_rows = map_items(catalog, subject_index,
                            price_min_ratio=args.price_min_ratio,
                            price_max_ratio=args.price_max_ratio)

    matched     = sum(1 for r in mapped_rows if r["score"] > 0)
    not_matched = len(mapped_rows) - matched
    log.info(f"Сопоставлено: {matched:,}  |  Без ниши: {not_matched:,}")

    # ─── 4. Excel ─────────────────────────────────────────────────────────────
    ts    = datetime.now().strftime("%Y%m%d_%H%M")
    fpath = OUT_DIR / f"wb_catalog_map_{ts}.xlsx"

    wb_out = openpyxl.Workbook()

    ws_map = wb_out.active
    ws_map.title = "Маппинг"
    write_mapping_sheet(ws_map, mapped_rows, args.top)

    ws_niches = wb_out.create_sheet("Топ_нишам")
    write_niches_sheet(ws_niches, mapped_rows)

    ws_search = wb_out.create_sheet("Для_поиска")
    search_queue = write_search_sheet(ws_search, mapped_rows)

    wb_out.save(str(fpath))
    log.info(f"Excel сохранён: {fpath}")

    # ─── 5. Текстовый файл очереди поиска ────────────────────────────────────
    txt_path = OUT_DIR / f"wb_search_queue_{ts}.txt"
    with open(txt_path, "w", encoding="utf-8") as f:
        for r in search_queue:
            f.write(f"{r['oem']}\n")
    log.info(f"Очередь поиска WB: {txt_path} ({len(search_queue)} OEM-артикулов)")

    # ─── Консольный отчёт ────────────────────────────────────────────────────
    w = 115
    print(f"\n{'─'*w}")
    print(f"{'#':>4} {'Бренд':12} {'Артикул':14} {'Название':35} {'Ниша WB':30} {'OOS':>4} {'Прод':>6} {'Маржа':>6} {'Скор':>7}")
    print(f"{'─'*w}")
    for i, r in enumerate(mapped_rows[:40], 1):
        if r["score"] <= 0:
            break
        print(
            f"{i:>4} {r['brand'][:11]:12} {r['article'][:13]:14} {r['name'][:34]:35} "
            f"{r['subject'][:29]:30} {r['oos_pct']:>3.0f}%% {r['sales']:>6,} "
            f"{r['margin_pct']:>5.1f}%% {r['score']:>7.3f}"
        )
    print(f"{'─'*w}")
    print(f"\nИтого в каталоге:  {len(catalog):,} позиций")
    print(f"С WB-нишей:        {matched:,}")
    print(f"Без ниши:          {not_matched:,}")
    print(f"Уникальных OEM:    {len(search_queue):,} (готовы к поиску на WB)")
    print(f"\nФайлы:")
    print(f"  {fpath}")
    print(f"  {txt_path}\n")


if __name__ == "__main__":
    main()
