"""
ozon_direct_update.py
═══════════════════════════════════════════════════════════════════════════════
Разовый / плановый пересчёт цен и остатков Ozon напрямую через Ozon Seller API.

Отличия от price_recalc.py:
  - Цены обновляются напрямую в Ozon (не через МойСклад)
  - Остатки тоже напрямую в Ozon
  - Формула: наценка (Оптимум) от (закупка + упаковка), а не маржа 12%
  - Габариты берутся из Ozon карточек (не из mikado_data.xlsx)
  - Прайс Mikado используется для закупочных цен И остатков (QTY)

Тиры Оптимум (наценка на закупка + упаковка):
  < 500 ₽  → 30%
  500–1199 → 25%
  1200–2499 → 22%
  2500–3499 → 20%
  ≥ 3500   → 17%

Запуск:
  uv run --with requests,openpyxl scripts/ozon_direct_update.py --once
  uv run --with requests,openpyxl scripts/ozon_direct_update.py --once --dry-run
  uv run --with requests,openpyxl scripts/ozon_direct_update.py --once --skip-prices
  uv run --with requests,openpyxl scripts/ozon_direct_update.py --once --skip-stocks

Переменные .env:
    OZON_CLIENT_ID=...
    OZON_API_KEY=...
    OZON_WAREHOUSE_ID=...
    TG_BOT_TOKEN=...   (опционально)
    TG_CHAT_ID=...     (опционально)
"""

import sys
import io
import math
import json
import time
import logging
import argparse
from collections import Counter
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
except ImportError:
    print("Установи: uv run --with requests,openpyxl scripts/ozon_direct_update.py")
    sys.exit(1)

try:
    from telegram_notify import tg_alert
    _TG_OK = True
except ImportError:
    _TG_OK = False

# ─── Пути ─────────────────────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).parent.parent
ENV_FILE  = BASE_DIR / ".env"
LOG_FILE  = BASE_DIR / "logs" / "ozon_direct_update.log"
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)

MIKADO_PRICE_URL = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=YOUR_MIKADO_PRICE_KEY"
)
PRICE_FALLBACK = Path("C:/Users/Admin/Desktop/mikado_price_34 (3).xlsx")

OZON_BASE = "https://api-seller.ozon.ru"

# ─── Логирование ──────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ─── Ценовые константы (FBS Ozon) ─────────────────────────────────────────────
FBS_TIERS = [100, 300, 1500, 5000, 10000]
FBS_RATES = [0.14, 0.20, 0.44, 0.44, 0.44, 0.44]

ACQ_PCT  = 0.015   # эквайринг (зашит, не меняется)
TAX_PCT  = 0.06    # УСН 6%
RET_RATE = 0.03    # % возвратов
REVERSE  = 80      # обратная логистика, ₽
OTHER    = 20      # упаковка, ₽

LOG_FBS = [
    (0.5, 75), (1, 90), (2, 115), (5, 155), (10, 210),
    (15, 265), (20, 315), (25, 365), (30, 420), (50, 620),
]
DEFAULT_LOGISTICS = 115  # ₽ — если габариты не найдены в Ozon карточке

# ─── Защитные ограничения ─────────────────────────────────────────────────────
PRICE_DROP_LIMIT = 0.60   # не снижать цену более чем на 40% за один пересчёт
MIN_MARKUP_FLOOR = 0.05   # не ставить цену если наценка ниже 5%

# Коды (строчными), исключённые из пересчёта.
# Причина: в прайсе Mikado есть позиции с таким же кодом, но это другой товар.
SKIP_CODES: frozenset[str] = frozenset({
    "gf-1904",  # Mikado: газовая пружина багажника Citroen C4 (не наш товар)
})


# ─── .env ─────────────────────────────────────────────────────────────────────
def load_env() -> dict:
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


# ─── Ценовые формулы ──────────────────────────────────────────────────────────
def _fbs_rate(sell: float) -> float:
    for thresh, rate in zip(FBS_TIERS, FBS_RATES):
        if sell < thresh:
            return rate
    return FBS_RATES[-1]


def _log_cost(weight_kg: float) -> float:
    for lim, cost in LOG_FBS:
        if weight_kg <= lim:
            return cost
    return LOG_FBS[-1][1] + math.ceil(weight_kg - LOG_FBS[-1][0]) * 15


def calc_logistics(volume_weight_kg: float) -> float:
    return _log_cost(volume_weight_kg)


def calc_profit(purchase: float, sell: float, logistics: float) -> float:
    comm_rate   = _fbs_rate(sell)
    commission  = sell * comm_rate
    acquiring   = sell * ACQ_PCT
    return_loss = RET_RATE * (logistics + REVERSE)
    proceeds    = sell - commission - acquiring - logistics
    tax         = max(0.0, proceeds) * TAX_PCT
    total_cost  = purchase + commission + acquiring + logistics + return_loss + OTHER + tax
    return sell - total_cost


def get_optimal_markup(purchase: float) -> float:
    """Тиры Оптимум: наценка от закупочной цены."""
    if purchase < 500:  return 0.30
    if purchase < 1200: return 0.25
    if purchase < 2500: return 0.22
    if purchase < 3500: return 0.20
    return 0.17


def find_markup_price(purchase: float, logistics: float) -> int | None:
    """Минимальная цена при которой наценка (прибыль / (закупка + упаковка)) ≥ Оптимум."""
    target = get_optimal_markup(purchase)
    base   = purchase + OTHER
    for s in range(50, 500_001):
        profit = calc_profit(purchase, s, logistics)
        if profit / base >= target - 1e-6:
            return s
    return None


# ─── Mikado: прайс ────────────────────────────────────────────────────────────
def load_mikado(url: str, fallback: Path) -> tuple[dict[str, float], dict[str, int]]:
    """
    Скачивает (или читает локальный) прайс Mikado.
    Возвращает ({code: purchase_price}, {code: qty}).
    """
    content = None
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        if resp.content[:2] == b"PK":
            content = resp.content
            log.info(f"Mikado: прайс скачан ({len(content):,} байт)")
    except Exception as e:
        log.warning(f"Mikado: онлайн недоступен ({e}), берём локальный")

    if content:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    elif fallback.exists():
        log.warning(f"Mikado: используем локальный файл {fallback}")
        wb = openpyxl.load_workbook(fallback, read_only=True, data_only=True)
    else:
        log.error("Mikado: прайс недоступен")
        return {}, {}

    ws = wb.active
    all_rows = ws.iter_rows(values_only=True)
    header = [str(v).strip().lower() if v else "" for v in (next(all_rows, []) or [])]

    code_idx = price_idx = qty_idx = None
    for i, h in enumerate(header):
        if h == "code":     code_idx  = i
        elif h == "priceout": price_idx = i
        elif h == "qty":    qty_idx   = i

    if code_idx is None:
        log.error(f"Mikado: колонка 'Code' не найдена. Заголовок: {header}")
        wb.close()
        return {}, {}

    raw_prices: dict[str, list[float]] = {}
    raw_qtys:   dict[str, int]         = {}

    for row in all_rows:
        raw = row[code_idx] if len(row) > code_idx else None
        if not raw:
            continue
        code  = str(raw).strip().lower()
        price = 0.0
        if price_idx is not None and len(row) > price_idx:
            try: price = float(str(row[price_idx] or 0))
            except: pass
        qty = 0
        if qty_idx is not None and len(row) > qty_idx:
            try: qty = max(0, int(float(str(row[qty_idx] or 0))))
            except: pass

        if price > 0:
            raw_prices.setdefault(code, []).append(price)
        raw_qtys[code] = raw_qtys.get(code, 0) + qty

    wb.close()

    # Исключаем артикулы с конфликтующими ценами
    prices: dict[str, float] = {}
    for code, pl in raw_prices.items():
        if len({round(p, 2) for p in pl}) == 1:
            prices[code] = pl[0]
        else:
            log.warning(f"  Mikado дубль: '{code}' — конфликт цен {pl} → исключён")

    log.info(f"Mikado: {len(prices)} позиций с ценой, "
             f"в наличии: {sum(1 for q in raw_qtys.values() if q > 0)}")
    return prices, raw_qtys


# ─── Ozon API helpers ─────────────────────────────────────────────────────────
def _ozon_headers(client_id: str, api_key: str) -> dict:
    return {
        "Client-Id": client_id,
        "Api-Key":   api_key,
        "Content-Type": "application/json",
    }


def get_ozon_products(client_id: str, api_key: str) -> list[dict]:
    """
    Возвращает все товары Ozon: [{product_id, offer_id}].
    Пагинация через last_id.
    """
    headers  = _ozon_headers(client_id, api_key)
    products = []
    last_id  = ""
    page     = 0

    while True:
        page += 1
        try:
            resp = requests.post(
                f"{OZON_BASE}/v3/product/list",
                headers=headers,
                json={"filter": {"visibility": "ALL"}, "last_id": last_id, "limit": 1000},
                timeout=30,
            )
            resp.raise_for_status()
            result  = resp.json().get("result", {})
            items   = result.get("items", [])
            last_id = result.get("last_id", "")
            products.extend(items)
            log.info(f"Ozon товары: страница {page}, получено {len(items)}, всего {len(products)}")
            if not items or not last_id:
                break
        except Exception as e:
            log.error(f"Ozon: ошибка получения товаров (страница {page}): {e}")
            break

    log.info(f"Ozon: итого {len(products)} товаров")
    return products


def get_ozon_volume_weights(client_id: str, api_key: str, product_ids: list[int]) -> dict[str, float]:
    """
    Возвращает объёмный вес товаров (кг) из Ozon: {offer_id: volume_weight_kg}.
    Использует POST /v3/product/info/list (батчи по 1000).
    volume_weight — готовое тарифное значение Ozon (max(факт, объём)).
    """
    headers = _ozon_headers(client_id, api_key)
    weights: dict[str, float] = {}
    no_weight = 0

    for start in range(0, len(product_ids), 1000):
        batch = product_ids[start : start + 1000]
        try:
            resp = requests.post(
                f"{OZON_BASE}/v3/product/info/list",
                headers=headers,
                json={"product_id": batch},
                timeout=60,
            )
            resp.raise_for_status()
            items = resp.json().get("items", [])
            for item in items:
                offer_id = item.get("offer_id", "")
                vw = item.get("volume_weight") or 0
                if vw and float(vw) > 0:
                    weights[offer_id] = float(vw)
                else:
                    no_weight += 1
        except Exception as e:
            log.error(f"Ozon: ошибка получения объёмного веса (batch {start}): {e}")

    log.info(f"Ozon: объёмный вес получен для {len(weights)} товаров, без веса: {no_weight}")
    return weights


def update_ozon_prices(
    client_id: str, api_key: str,
    prices: list[dict],   # [{offer_id, price}]
    dry_run: bool = False,
) -> tuple[int, int]:
    """
    Обновляет цены через POST /v1/product/import/prices.
    Возвращает (ok, fail).
    """
    if dry_run:
        log.info(f"[DRY-RUN] Ozon цены: обновилось бы {len(prices)} позиций")
        return len(prices), 0

    headers = _ozon_headers(client_id, api_key)
    ok = fail = 0

    for start in range(0, len(prices), 1000):
        batch = prices[start : start + 1000]
        payload = [
            {
                "offer_id":      p["offer_id"],
                "price":         str(p["price"]),
                "old_price":     "0",
                "min_price":     "0",
                "currency_code": "RUB",
            }
            for p in batch
        ]
        try:
            resp = requests.post(
                f"{OZON_BASE}/v1/product/import/prices",
                headers=headers,
                json={"prices": payload},
                timeout=60,
            )
            resp.raise_for_status()
            result = resp.json().get("result", [])
            errors = [r for r in result if not r.get("updated")]
            ok   += len(batch) - len(errors)
            fail += len(errors)
            for e in errors:
                log.warning(f"  Ozon цена: {e.get('offer_id')} — {e.get('errors')}")
        except Exception as e:
            log.error(f"Ozon: ошибка обновления цен (batch {start}): {e}")
            fail += len(batch)

    return ok, fail


def update_ozon_stocks(
    client_id: str, api_key: str,
    warehouse_id: int,
    stocks: list[dict],   # [{offer_id, stock}]
    dry_run: bool = False,
) -> tuple[int, int]:
    """
    Обновляет FBS-остатки через POST /v2/products/stocks.
    Возвращает (ok, fail).
    """
    if dry_run:
        log.info(f"[DRY-RUN] Ozon остатки: обновилось бы {len(stocks)} позиций")
        return len(stocks), 0

    headers = _ozon_headers(client_id, api_key)
    ok = fail = 0

    for start in range(0, len(stocks), 100):
        batch = stocks[start : start + 100]
        payload = [
            {"offer_id": s["offer_id"], "stock": s["stock"], "warehouse_id": warehouse_id}
            for s in batch
        ]
        try:
            resp = requests.post(
                f"{OZON_BASE}/v2/products/stocks",
                headers=headers,
                json={"stocks": payload},
                timeout=60,
            )
            resp.raise_for_status()
            result = resp.json().get("result", [])
            errors = [r for r in result if r.get("errors")]
            ok   += len(batch) - len(errors)
            fail += len(errors)
            for e in errors:
                log.warning(f"  Ozon остаток: {e.get('offer_id')} — {e.get('errors')}")
        except Exception as e:
            log.error(f"Ozon: ошибка обновления остатков (batch {start}): {e}")
            fail += len(batch)

    return ok, fail


# ─── Основная логика ──────────────────────────────────────────────────────────
def run_once(env: dict, dry_run: bool, skip_prices: bool, skip_stocks: bool) -> None:
    log.info("═" * 60)
    log.info(f"ozon_direct_update — старт {'[DRY-RUN] ' if dry_run else ''}"
             f"{'[SKIP-PRICES] ' if skip_prices else ''}"
             f"{'[SKIP-STOCKS] ' if skip_stocks else ''}")

    client_id    = env.get("OZON_CLIENT_ID", "")
    api_key      = env.get("OZON_API_KEY", "")
    warehouse_id_s = env.get("OZON_WAREHOUSE_ID", "")

    if not client_id or not api_key:
        log.error("OZON_CLIENT_ID / OZON_API_KEY не заданы — выход")
        return
    if not skip_stocks and not warehouse_id_s:
        log.error("OZON_WAREHOUSE_ID не задан — выход")
        return

    try:
        warehouse_id = int(warehouse_id_s)
    except (ValueError, TypeError):
        log.error(f"OZON_WAREHOUSE_ID не число: {warehouse_id_s!r}")
        return

    # 1. Прайс Mikado
    price_db, qty_db = load_mikado(MIKADO_PRICE_URL, PRICE_FALLBACK)
    if not price_db:
        log.error("Прайс Mikado пуст — выход")
        return

    # 2. Список товаров Ozon
    ozon_products = get_ozon_products(client_id, api_key)
    if not ozon_products:
        log.error("Ozon: товары не получены — выход")
        return

    # Защита от дублей артикулов
    art_counts = Counter(p["offer_id"].removesuffix("-con").lower() for p in ozon_products)
    dup_articles = {a for a, c in art_counts.items() if c > 1}
    if dup_articles:
        log.warning(f"⚠ Дублирующиеся артикулы в Ozon ({len(dup_articles)} шт.) → пропущены")

    # 3. Объёмный вес из Ozon карточек
    product_ids = [p["product_id"] for p in ozon_products]
    vol_weights = get_ozon_volume_weights(client_id, api_key, product_ids)

    # 4. Расчёт цен и формирование батчей
    price_updates: list[dict] = []
    stock_updates: list[dict] = []

    stats = {"skipped": 0, "no_price": 0, "no_weight": 0,
             "price_ok": 0, "stock_ok": 0}

    for prod in ozon_products:
        offer_id = prod["offer_id"]
        mk_code  = offer_id.removesuffix("-con")
        mk_key   = mk_code.lower()

        # SKIP_CODES
        if mk_key in SKIP_CODES:
            stats["skipped"] += 1
            continue

        # Дубли
        if mk_key in dup_articles:
            stats["skipped"] += 1
            continue

        purchase = price_db.get(mk_key)
        if not purchase:
            stats["no_price"] += 1
            continue

        # Цена
        if not skip_prices:
            vw = vol_weights.get(offer_id)
            if vw:
                logistics = calc_logistics(vw)
                stats["price_ok"] += 1
            else:
                logistics = DEFAULT_LOGISTICS
                stats["no_weight"] += 1

            new_price = find_markup_price(purchase, logistics)
            if new_price is None:
                log.warning(f"  {mk_code}: не подобрать цену (закупка={purchase:.0f} ₽)")
                stats["skipped"] += 1
                continue

            markup = calc_profit(purchase, new_price, logistics) / (purchase + OTHER)
            if markup < MIN_MARKUP_FLOOR:
                log.warning(f"  {mk_code}: наценка {markup*100:.1f}% < {MIN_MARKUP_FLOOR*100:.0f}% → пропущен")
                stats["skipped"] += 1
                continue

            log.info(
                f"  {mk_code:<18} закупка={purchase:.0f} ₽  "
                f"лог={logistics:.0f} ₽  "
                f"наценка={markup*100:.0f}%  → {new_price} ₽"
            )
            price_updates.append({"offer_id": offer_id, "price": new_price})

        # Остатки
        if not skip_stocks:
            qty = qty_db.get(mk_key, 0)
            stock_updates.append({"offer_id": offer_id, "stock": qty})
            stats["stock_ok"] += 1

    # 5. Применяем
    prices_ok = prices_fail = stocks_ok = stocks_fail = 0

    if not skip_prices and price_updates:
        log.info(f"Ozon: отправляем {len(price_updates)} обновлений цен...")
        prices_ok, prices_fail = update_ozon_prices(client_id, api_key, price_updates, dry_run)
        log.info(f"Ozon цены: обновлено {prices_ok}, ошибок {prices_fail}")

    if not skip_stocks and stock_updates:
        log.info(f"Ozon: отправляем {len(stock_updates)} обновлений остатков...")
        stocks_ok, stocks_fail = update_ozon_stocks(
            client_id, api_key, warehouse_id, stock_updates, dry_run
        )
        log.info(f"Ozon остатки: обновлено {stocks_ok}, ошибок {stocks_fail}")

    in_stock = sum(1 for s in stock_updates if s["stock"] > 0)

    log.info("─" * 60)
    log.info(
        f"Итог: цены обновлено={prices_ok} ошибок={prices_fail} | "
        f"остатки обновлено={stocks_ok} ошибок={stocks_fail} | "
        f"в наличии={in_stock} | без цены Mikado={stats['no_price']} | "
        f"дефолт лог={stats['no_weight']} | пропущено={stats['skipped']}"
    )

    # Telegram
    if _TG_OK:
        try:
            tg_alert(
                f"{'[DRY-RUN] ' if dry_run else ''}ozon_direct_update\n"
                f"Цены: +{prices_ok} ❌{prices_fail}\n"
                f"Остатки: +{stocks_ok} ❌{stocks_fail}  в наличии: {in_stock}\n"
                f"Без цены Mikado: {stats['no_price']}  дефолт логистика: {stats['no_weight']}"
            )
        except Exception:
            pass


# ─── CLI ──────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(description="Ozon direct price+stock update")
    parser.add_argument("--once",         action="store_true", help="Разовый запуск")
    parser.add_argument("--dry-run",      action="store_true", help="Не отправлять в Ozon")
    parser.add_argument("--skip-prices",  action="store_true", help="Не обновлять цены")
    parser.add_argument("--skip-stocks",  action="store_true", help="Не обновлять остатки")
    args = parser.parse_args()

    env = load_env()
    run_once(env, dry_run=args.dry_run, skip_prices=args.skip_prices, skip_stocks=args.skip_stocks)


if __name__ == "__main__":
    main()
