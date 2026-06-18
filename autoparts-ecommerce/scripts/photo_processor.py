"""
photo_processor.py — обработка фотографий товаров для Ozon
══════════════════════════════════════════════════════════

Что делает:
  На каждый товар создаёт 2 фотографии:
    Фото 1 (_main)  — чистый товар на белом фоне, без текста (главное фото Ozon)
    Фото 2 (_info)  — товар + параметры из таблицы + логотип магазина

Входные данные:
  - Папка с исходными фото (images/)
  - Excel файл mikado_data.xlsx (берёт колонку "Параметры")
  - Логотип магазина (PNG с прозрачным фоном, опционально)

Запуск:
  uv run --with rembg,pillow,openpyxl scripts/photo_processor.py --data путь/к/mikado_data.xlsx
  uv run --with rembg,pillow,openpyxl scripts/photo_processor.py --data путь/к/mikado_data.xlsx --rows 2-5
  uv run --with rembg,pillow,openpyxl scripts/photo_processor.py --data путь/к/mikado_data.xlsx --rows 2-5 --logo путь/к/логотип.png

Аргументы:
  --data      Путь к mikado_data.xlsx  [обязательный]
  --rows      Строки Excel для обработки, например 2-5 или 3 [по умолчанию: все строки]
  --images    Папка с исходными фото   [по умолчанию: images/ рядом с data]
  --out       Папка для результатов    [по умолчанию: photos_ozon/ рядом с data]
  --logo      Путь к PNG логотипу      [опционально]
  --size      Размер итогового фото    [по умолчанию: 1000]
  --size      Размер итогового фото [по умолчанию: 1000]

Результаты сохраняются в подпапку с датой и временем:
  photos_ozon/2026-04-14_10-30/
      f-a22025_1_main.jpg
      f-a22025_1_info.jpg
      ...
"""

import re
import sys
import argparse
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    print("Установи зависимости: uv run --with rembg,pillow scripts/photo_processor.py")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Установи зависимости: uv run --with rembg,pillow,openpyxl scripts/photo_processor.py")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНСТАНТЫ ДИЗАЙНА
# ══════════════════════════════════════════════════════════════════════════════

IMG_W         = 900           # ширина итогового холста
IMG_H         = 1200          # высота итогового холста (3:4)
BG_COLOR      = (255, 255, 255)   # белый фон
TEXT_COLOR    = (40,  40,  40)    # тёмно-серый текст
ACCENT_COLOR  = (20,  20,  20)    # почти чёрный для заголовков
LINE_COLOR    = (210, 210, 210)   # светло-серая разделительная линия
PARAM_COLOR   = (80,  80,  80)    # серый для значений параметров

LOGO_MARGIN   = 24            # отступ от краёв
STRIPE_H      = 20            # высота тёмно-оранжевой полосы сверху и снизу
STRIPE_COLOR  = (255, 204, 51)  # цвет полосы с bg_inpainted.jpg

HEADER_H      = 200           # высота верхней зоны: лого + название + артикул
FOOTER_H      = 220           # высота нижней зоны с параметрами
PADDING       = 10            # минимальный отступ по краям
IMG_SHIFT_DOWN = 41           # нижний край фото уходит на 41px за пределы холста


# ══════════════════════════════════════════════════════════════════════════════
#  ЧТЕНИЕ ДАННЫХ ИЗ EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def read_data(xlsx_path: Path) -> dict[str, dict]:
    """
    Читает mikado_data.xlsx.
    Возвращает словарь {артикул: {name, brand, params, row, ...}}.
    Поле row — номер строки в Excel (начиная с 2).
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Заголовки из строки 1
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    def col(name: str) -> int | None:
        for i, h in enumerate(headers):
            if name.lower() in h.lower():
                return i
        return None

    idx_code   = col("код")
    idx_name   = col("наименование")
    idx_brand  = col("бренд")
    idx_params = col("параметры")
    idx_compat = col("применяемость")
    idx_images = col("изображения")
    idx_profit = col("прибыль")

    data = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue

        # Пропускаем строки без прибыли
        if idx_profit is not None:
            profit_val = row[idx_profit]
            if profit_val is None or str(profit_val).strip() in ("", "None"):
                continue

        # Ключ: артикул если есть, иначе номер строки
        code = str(row[idx_code] or "").strip() if idx_code is not None else ""
        if not code:
            code = f"row_{row_num}"

        data[code] = {
            "code":       code,
            "name":       str(row[idx_name]   or "") if idx_name   is not None else "",
            "brand":      str(row[idx_brand]  or "") if idx_brand  is not None else "",
            "params":     str(row[idx_params] or "") if idx_params is not None else "",
            "compat":     str(row[idx_compat] or "") if idx_compat is not None else "",
            "image_file": str(row[idx_images] or "") if idx_images is not None else "",
            "row":        row_num,
        }

    wb.close()
    return data


def filter_by_rows(items: dict[str, dict], rows_arg: str) -> dict[str, dict]:
    """
    Фильтрует товары по номерам строк Excel.
    rows_arg примеры: "2-5", "3", "2-4,7,10-12"
    """
    allowed = set()
    for part in rows_arg.split(","):
        part = part.strip()
        if "-" in part:
            a, b = part.split("-", 1)
            allowed.update(range(int(a), int(b) + 1))
        else:
            allowed.add(int(part))
    return {code: item for code, item in items.items() if item["row"] in allowed}


# ══════════════════════════════════════════════════════════════════════════════
#  УДАЛЕНИЕ ФОНА
# ══════════════════════════════════════════════════════════════════════════════

def remove_background(img: Image.Image) -> Image.Image:
    """
    Удаляет фон с помощью rembg. Возвращает RGBA изображение.
    alpha_matting=True — чёткие края, убирает остатки текста/вотермарок
    на границе товара.
    """
    try:
        from rembg import remove
        rgba = remove(
            img,
            alpha_matting=True,
            alpha_matting_foreground_threshold=240,
            alpha_matting_background_threshold=10,
            alpha_matting_erode_size=10,
        )
        return rgba
    except ImportError:
        print("  ⚠  rembg не установлен — фон не удаляется")
        return img.convert("RGBA")


def place_on_white(fg: Image.Image, canvas_size: int, padding: int) -> Image.Image:
    """
    Помещает изображение (RGBA) на белый квадратный холст.
    Товар масштабируется чтобы занять ~75% площади, выравнивается по центру.
    """
    canvas = Image.new("RGB", (canvas_size, canvas_size), BG_COLOR)

    # Максимальный размер товара с учётом отступов
    max_dim = canvas_size - padding * 2
    fg.thumbnail((max_dim, max_dim), Image.LANCZOS)

    # Центрирование
    x = (canvas_size - fg.width)  // 2
    y = (canvas_size - fg.height) // 2

    if fg.mode == "RGBA":
        canvas.paste(fg, (x, y), mask=fg.split()[3])
    else:
        canvas.paste(fg, (x, y))

    return canvas


# ══════════════════════════════════════════════════════════════════════════════
#  ОБРЕЗКА ПОЛОСЫ АРТИКУЛА
# ══════════════════════════════════════════════════════════════════════════════

def _find_strip_y(img: Image.Image) -> int:
    """
    Возвращает Y-координату начала светлой полосы с артикулом поставщика.

    Структура исходного фото:
      [тонкая тёмная рамка сверху, несколько px]
      [основное фото товара]
      [тонкая тёмная полоска-разделитель — такая же, как рамка сверху]
      [светлая полоса с артикулом]

    Сканирование снизу вверх: идём от низа изображения вверх и находим
    первую тонкую тёмную полосу — это и есть разделитель. Всё ниже неё
    является полосой артикула и скрывается под нижней оранжевой границей.
    """
    gray   = img.convert("L")
    w, h   = gray.size
    pixels = gray.load()

    def row_vals(y):
        return [pixels[x, y] for x in range(0, w, max(1, w // 100))]

    means = [sum(v := row_vals(y)) / len(v) for y in range(h)]

    # Сглаживание окном 3
    sm = means.copy()
    for i in range(1, h - 1):
        sm[i] = (means[i-1] + means[i] + means[i+1]) / 3

    # ── Шаг 1: характеристика верхней рамки ──────────────────────────────────
    border_end = 1
    for i in range(1, min(25, h)):
        if sm[i] > sm[0] + 12:
            border_end = i
            break
        border_end = i + 1

    border_bright  = sum(sm[:border_end]) / border_end
    border_thick   = border_end
    content_bright = sum(sm[border_end:border_end + 30]) / min(30, h - border_end)

    # Адаптивный порог: 55% пути от яркости рамки к яркости контента
    dark_thresh = border_bright + (content_bright - border_bright) * 0.55
    max_band_h  = max(8, border_thick + 5)

    def coverage(y_start, y_end):
        """Доля пикселей темнее порога в средней строке полосы (по сырым пикселям)."""
        vals = row_vals((y_start + y_end) // 2)
        return sum(1 for v in vals if v <= dark_thresh) / len(vals)

    # ── Шаг 2: сканирование снизу вверх по СЫРЫМ значениям ───────────────────
    # Сглаживание убивает тонкие полосы (1–2px): сырые значения точнее.
    # Контекст (before/after) берём из сглаженных — они стабильнее.
    scan_end = int(h * 0.65)
    i = h - 2
    while i > scan_end:
        if means[i] <= dark_thresh:          # сырое значение строки
            band_bottom = i
            while i > scan_end and means[i] <= dark_thresh:
                i -= 1
            band_top = i + 1
            band_h   = band_bottom - band_top + 1

            if 1 <= band_h <= max_band_h:
                cov    = coverage(band_top, band_bottom)
                after  = sum(sm[band_bottom + 1:min(h, band_bottom + 21)]) / min(20, h - band_bottom - 1)
                before = sum(sm[max(0, band_top - 20):band_top]) / max(1, min(20, band_top))
                band   = sum(means[band_top:band_bottom + 1]) / band_h  # сырое среднее полосы

                if (cov  >= 0.55            # разделитель на ≥55% ширины
                        and band < before - 2   # темнее продукта сверху
                        and band < after  - 2): # темнее артикула снизу
                    return band_bottom + 1
        else:
            i -= 1

    return h


def crop_article_strip(img: Image.Image) -> Image.Image:
    """Обрезает нижнюю полосу с артикулом (используется для главного фото)."""
    return img.crop((0, 0, img.width, _find_strip_y(img)))


def _fit_to_width(img: Image.Image, w: int, h: int) -> Image.Image:
    """
    Масштабирует изображение по ширине w, прижимает к низу холста w×h.
    """
    scale   = w / img.width
    new_h   = int(img.height * scale)
    resized = img.resize((w, new_h), Image.LANCZOS)
    canvas  = Image.new("RGB", (w, h), BG_COLOR)
    y_paste = h + IMG_SHIFT_DOWN - new_h
    canvas.paste(resized, (0, y_paste))
    return canvas


# ══════════════════════════════════════════════════════════════════════════════
#  ФОТО 1 — ГЛАВНОЕ
# ══════════════════════════════════════════════════════════════════════════════

def make_main_photo(img: Image.Image, w: int, h: int) -> Image.Image:
    """
    Фото 1: исходное фото без полосы артикула, растянутое по ширине холста w×h.
    """
    return _fit_to_width(crop_article_strip(img), w, h)


# ══════════════════════════════════════════════════════════════════════════════
#  ФОТО 2 — ИНФОРМАЦИОННОЕ (товар + логотип)
# ══════════════════════════════════════════════════════════════════════════════

# Словарь сокращений → полные слова (применяется к названию товара)
_ABBREVS = [
    ("Аморт.",  "Амортизатор "),
    ("зад.",    "задний "),
    ("пер.",    "передний "),
    ("л/пр",    "левый/правый "),
    ("лев.",    "левый "),
    ("прав.",   "правый "),
    ("пр.",     "правый "),
    ("ст.",     "стойка "),
    ("оп.",     "опора "),
    ("к-т",     "комплект "),
    ("п/у",     "с пневмоупором "),
]


def expand_abbreviations(text: str) -> str:
    """Раскрывает сокращения в названии товара."""
    for abbr, full in _ABBREVS:
        text = text.replace(abbr, full)
    # Убираем двойные пробелы после замен
    while "  " in text:
        text = text.replace("  ", " ")
    return text.strip()


def get_font(size: int, bold: bool = False):
    """Загружает системный шрифт. Если не найден — использует встроенный."""
    font_candidates = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
    ]
    bold_candidates = [
        "C:/Windows/Fonts/arialbd.ttf",
        "C:/Windows/Fonts/segoeuib.ttf",
        "C:/Windows/Fonts/tahomabd.ttf",
    ]
    candidates = bold_candidates if bold else font_candidates
    for path in candidates:
        if Path(path).exists():
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                continue
    return ImageFont.load_default()


def parse_params(params_text: str) -> list[tuple[str, str]]:
    """
    Разбирает строку параметров на пары (ключ, значение).
    Формат строки: "Ключ: Значение\nКлюч2: Значение2\n..."
    """
    result = []
    for line in params_text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        if ":" in line:
            key, _, val = line.partition(":")
            result.append((key.strip(), val.strip()))
        else:
            result.append((line, ""))
    return result


def _fit_bg(bg: Image.Image, w: int, h: int) -> Image.Image:
    """
    Подгоняет фон под прямоугольник w×h с кадрированием по центру
    (без искажений пропорций).
    """
    bg   = bg.convert("RGB")
    bw, bh = bg.size
    scale = max(w / bw, h / bh)
    new_w = int(bw * scale)
    new_h = int(bh * scale)
    bg    = bg.resize((new_w, new_h), Image.LANCZOS)
    x0    = (new_w - w) // 2
    y0    = (new_h - h) // 2
    return bg.crop((x0, y0, x0 + w, y0 + h))


def make_info_photo(img: Image.Image, item: dict, logo: Image.Image | None,
                    w: int, h: int, bg: Image.Image | None = None) -> Image.Image:
    """
    Компоновка (снизу вверх):
      ┌────────────────────────────────────────────┐  y=0
      │████████ полоса ████████████████████████████│  STRIPE_H
      │  [ЛОГО 3/5h]     Название (54px)           │  ← header
      │                  Артикул  (72px)           │
      │      ИСХОДНОЕ ФОТО (полная ширина)         │
      │████████ полоса ████████████████████████████│  h-STRIPE_H … h
      └────────────────────────────────────────────┘  y=h
    """
    # ── Слой 1: фон или белый ─────────────────────────────────────────────────
    canvas = _fit_bg(bg, w, h) if bg is not None else Image.new("RGB", (w, h), BG_COLOR)

    # ── Слой 2: исходное фото ────────────────────────────────────────────────
    # Концепция: верхняя граница полосы артикула = верхняя граница нижней
    # оранжевой полосы. Лого и текст — в оставшейся зоне сверху.
    scale   = w / img.width
    new_h   = int(img.height * scale)
    resized = img.resize((w, new_h), Image.LANCZOS)

    # Нижний край фото касается верхней границы нижней оранжевой полосы
    y_img = (h - STRIPE_H) - new_h
    canvas.paste(resized, (0, y_img))

    # Свободная зона для лого и текста — между верхней полосой и фото
    header_top = STRIPE_H
    header_h   = max(y_img - STRIPE_H, 40)

    # ── Слой 3а: лого — верхний левый угол, 1/2 от прежнего размера ──────────
    draw = ImageDraw.Draw(canvas)
    logo_resized = None
    if logo and header_h > 40:
        logo_copy  = logo.copy()
        target_h   = int(HEADER_H * 3 / 5 * 2 * 0.8)      # фиксированный, одинаковый на обоих фото
        max_logo_w = w // 2 - LOGO_MARGIN * 2
        ratio      = min(max_logo_w / logo_copy.width, target_h / logo_copy.height)
        new_lw     = max(1, int(logo_copy.width  * ratio))
        new_lh     = max(1, int(logo_copy.height * ratio))
        logo_copy  = logo_copy.resize((new_lw, new_lh), Image.LANCZOS)
        logo_resized = logo_copy                          # сохраняем для compat-фото
        lx = LOGO_MARGIN
        ly = header_top + LOGO_MARGIN                     # верхний левый угол зоны
        if logo_copy.mode == "RGBA":
            canvas.paste(logo_copy, (lx, ly), mask=logo_copy.split()[3])
        else:
            canvas.paste(logo_copy, (lx, ly))
        logo_right = lx + new_lw + LOGO_MARGIN
    else:
        logo_right = LOGO_MARGIN

    # ── Слой 3б: текст — только в зоне справа от логотипа ───────────────────
    name       = expand_abbreviations(item.get("name", "")).upper()
    code       = item.get("code", "")
    zone_x1    = logo_right
    zone_x2    = w - LOGO_MARGIN
    zone_cx    = (zone_x1 + zone_x2) // 2
    max_text_w = zone_x2 - zone_x1

    font_code  = ImageFont.truetype("C:/Windows/Fonts/ariblk.ttf", 72)

    def wrap_text(text, font, max_w):
        words, lines, buf = text.split(), [], ""
        for word in words:
            test = f"{buf} {word}".strip()
            if draw.textbbox((0, 0), test, font=font)[2] > max_w and buf:
                lines.append(buf)
                buf = word
            else:
                buf = test
        if buf:
            lines.append(buf)
        return lines

    # Автоподбор размера шрифта: уменьшаем пока весь блок не уместится в header_h
    code_lh = draw.textbbox((0, 0), "АЙ", font=font_code)[3] + 4
    font_name = None
    for size in [52, 46, 40, 34, 28, 22]:
        f = ImageFont.truetype("C:/Windows/Fonts/arialbd.ttf", size)
        lines = wrap_text(name, f, max_text_w)
        lh    = draw.textbbox((0, 0), "АЙ", font=f)[3] + 4
        gap   = lh * 2
        bh    = lh * len(lines) + (gap + code_lh if code else 0)
        if bh <= header_h:
            font_name  = f
            name_lines = lines
            name_lh    = lh
            break
    if font_name is None:   # крайний случай: берём минимальный размер
        font_name  = ImageFont.truetype("C:/Windows/Fonts/arialbd.ttf", 22)
        name_lines = wrap_text(name, font_name, max_text_w)
        name_lh    = draw.textbbox((0, 0), "АЙ", font=font_name)[3] + 4

    gap     = name_lh * 2
    block_h = name_lh * len(name_lines) + (gap + code_lh if code else 0)
    y_text  = header_top + (header_h - block_h) // 2

    for line in name_lines:
        line_w = draw.textbbox((0, 0), line, font=font_name)[2]
        draw.text((zone_cx - line_w // 2, y_text), line, font=font_name, fill=ACCENT_COLOR)
        y_text += name_lh

    if code:
        y_text += gap
        code_upper = code.upper()
        bbox   = draw.textbbox((0, 0), code_upper, font=font_code)
        # bbox может иметь ненулевые bbox[0]/bbox[1] — учитываем смещение
        b_left, b_top, b_right, b_bottom = bbox
        code_w = b_right  - b_left
        code_h = b_bottom - b_top

        # Позиция рисования текста: чтобы визуальный центр текста = zone_cx
        tx = zone_cx - code_w // 2 - b_left
        ty = y_text            - b_top

        pad    = 12
        radius = 10

        # Подложка строго вокруг визуальных границ текста
        rx0 = tx + b_left  - pad
        ry0 = ty + b_top   - pad
        rx1 = tx + b_right  + pad
        ry1 = ty + b_bottom + pad

        # Градиентная подложка: от краёв (разбавлен белым) к центру (чистый цвет)
        sr, sg, sb = STRIPE_COLOR
        for i in range(pad):
            t = i / pad
            r = int(sr + (255 - sr) * (1 - t) * 0.55)
            g = int(sg + (255 - sg) * (1 - t) * 0.55)
            b = int(sb + (255 - sb) * (1 - t) * 0.55)
            draw.rounded_rectangle(
                [rx0 + i, ry0 + i, rx1 - i, ry1 - i],
                radius=max(1, radius - i),
                fill=(r, g, b),
            )
        draw.rounded_rectangle(
            [rx0 + pad, ry0 + pad, rx1 - pad, ry1 - pad],
            radius=1,
            fill=STRIPE_COLOR,
        )

        # Текст поверх подложки
        draw.text((tx, ty), code_upper, font=font_code, fill=(0, 0, 0))

    # ── Слой 4: полосы поверх всего ──────────────────────────────────────────
    draw.rectangle([(0, 0),          (w, STRIPE_H)],      fill=STRIPE_COLOR)
    draw.rectangle([(0, h-STRIPE_H), (w, h)],             fill=STRIPE_COLOR)

    return canvas, logo_resized


# ══════════════════════════════════════════════════════════════════════════════
#  ФОТО 3 — ПРИМЕНЯЕМОСТЬ
# ══════════════════════════════════════════════════════════════════════════════

def make_compat_photo(item: dict, logo_resized: Image.Image | None,
                      w: int, h: int, bg: Image.Image | None = None) -> Image.Image:
    """
    Компоновка:
      ┌────────────────────────────────────────────┐  y=0
      │████████ полоса ████████████████████████████│  STRIPE_H
      │  [ЛОГО]          ПРИМЕНЯЕМОСТЬ             │  ← header
      │                                            │
      │  текст применяемости по всей ширине        │
      │  ...                                       │
      │████████ полоса ████████████████████████████│  h-STRIPE_H … h
      └────────────────────────────────────────────┘  y=h

    logo_resized — уже отмасштабированный логотип из make_info_photo (те же пиксели).
    """
    # ── Слой 1: фон или белый ─────────────────────────────────────────────────
    canvas = _fit_bg(bg, w, h) if bg is not None else Image.new("RGB", (w, h), BG_COLOR)
    draw   = ImageDraw.Draw(canvas)

    header_top = STRIPE_H

    # ── Лого — вставляем готовый, без пересчёта ───────────────────────────────
    logo_right  = LOGO_MARGIN
    logo_bottom = header_top + LOGO_MARGIN

    if logo_resized:
        lx = LOGO_MARGIN
        ly = header_top + LOGO_MARGIN
        if logo_resized.mode == "RGBA":
            canvas.paste(logo_resized, (lx, ly), mask=logo_resized.split()[3])
        else:
            canvas.paste(logo_resized, (lx, ly))
        logo_right  = lx + logo_resized.width  + LOGO_MARGIN
        logo_bottom = ly + logo_resized.height + LOGO_MARGIN

    # высота шапки = от верхней полосы до нижнего края логотипа
    header_bottom = logo_bottom
    header_h      = header_bottom - header_top

    # ── Заголовок "ПРИМЕНЯЕМОСТЬ" — зона справа от лого ──────────────────────
    font_title = ImageFont.truetype("C:/Windows/Fonts/arialbd.ttf", 52)
    font_text  = ImageFont.truetype("C:/Windows/Fonts/arial.ttf",   30)

    zone_x1 = logo_right
    zone_x2 = w - LOGO_MARGIN
    zone_cx = (zone_x1 + zone_x2) // 2

    title     = "ПРИМЕНЯЕМОСТЬ"
    tb        = draw.textbbox((0, 0), title, font=font_title)
    title_h   = tb[3] - tb[1]
    title_y   = header_top + (header_h - title_h) // 2 - tb[1]
    title_x   = zone_cx - (tb[2] - tb[0]) // 2 - tb[0]
    draw.text((title_x, title_y), title, font=font_title, fill=ACCENT_COLOR)

    # ── Текст применяемости — ниже шапки ──────────────────────────────────────
    compat_raw  = item.get("compat", "").strip()
    text_top    = header_bottom + LOGO_MARGIN
    text_bottom = h - STRIPE_H - LOGO_MARGIN
    max_text_w  = w - LOGO_MARGIN * 2
    available_h = text_bottom - text_top

    def wrap_line(text, font, max_w):
        words, lines, buf = text.split(), [], ""
        for word in words:
            test = f"{buf} {word}".strip()
            if draw.textbbox((0, 0), test, font=font)[2] > max_w and buf:
                lines.append(buf)
                buf = word
            else:
                buf = test
        if buf:
            lines.append(buf)
        return lines or [""]

    def build_wrapped(font, max_w):
        result = []
        for paragraph in compat_raw.splitlines():
            paragraph = paragraph.strip()
            if paragraph:
                result.extend(wrap_line(paragraph, font, max_w))
            else:
                result.append("")
        return result

    # Автоподбор размера шрифта чтобы весь текст поместился
    font_size = 30
    while font_size >= 12:
        font_text = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", font_size)
        line_h    = draw.textbbox((0, 0), "АЙ", font=font_text)[3] + 4
        wrapped   = build_wrapped(font_text, max_text_w)
        if line_h * len(wrapped) <= available_h:
            break
        font_size -= 1

    def draw_justified(line, is_last, x0, y, max_w, font, color):
        """Рисует строку с выравниванием по ширине. Последняя строка — по левому краю."""
        words = line.split()
        if is_last or len(words) <= 1:
            draw.text((x0, y), line, font=font, fill=color)
            return
        words_w = sum(draw.textbbox((0, 0), word, font=font)[2]
                      - draw.textbbox((0, 0), word, font=font)[0]
                      for word in words)
        gap = (max_w - words_w) / (len(words) - 1)
        x = x0
        for word in words:
            draw.text((int(x), y), word, font=font, fill=color)
            wb = draw.textbbox((0, 0), word, font=font)
            x += (wb[2] - wb[0]) + gap

    y = text_top
    for i, line in enumerate(wrapped):
        if line:
            is_last = (i == len(wrapped) - 1) or (i + 1 < len(wrapped) and not wrapped[i + 1])
            draw_justified(line, is_last, LOGO_MARGIN, y, max_text_w, font_text, ACCENT_COLOR)
        y += line_h

    # ── Полосы поверх всего ───────────────────────────────────────────────────
    draw.rectangle([(0, 0),          (w, STRIPE_H)], fill=STRIPE_COLOR)
    draw.rectangle([(0, h-STRIPE_H), (w, h)],        fill=STRIPE_COLOR)

    return canvas


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Обработка фото товаров для Ozon",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--data",      required=True,             help="Путь к mikado_data.xlsx")
    ap.add_argument("--rows",      default=None,              help="Строки Excel: 2-5 или 3 или 2-4,7")
    ap.add_argument("--images",    default=None,              help="Папка с исходными фото")
    ap.add_argument("--out",       default=None,              help="Папка для результатов")
    ap.add_argument("--logo",      default=None,              help="Путь к PNG логотипу")
    ap.add_argument("--bg",        default=None,              help="Путь к фону инфо-фото (JPG/PNG)")
    args = ap.parse_args()

    data_file  = Path(args.data)
    if not data_file.exists():
        print(f"Файл не найден: {data_file}")
        sys.exit(1)

    images_dir = Path(args.images) if args.images else data_file.parent / "images"

    # Результаты — в подпапку с датой и временем запуска
    base_out  = Path(args.out) if args.out else data_file.parent / "photos_ozon"
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out_dir   = base_out / timestamp
    out_dir.mkdir(parents=True, exist_ok=True)

    # Фон инфо-фото
    bg = None
    if args.bg:
        bg_path = Path(args.bg)
        if bg_path.exists():
            bg = Image.open(bg_path).convert("RGB")
            print(f"  Фон: {bg_path.name}  ({bg.width}×{bg.height}px)")
        else:
            print(f"  ⚠  Фон не найден: {bg_path}")

    # Логотип
    logo = None
    if args.logo:
        logo_path = Path(args.logo)
        if logo_path.exists():
            logo = Image.open(logo_path).convert("RGBA")
            print(f"  Логотип: {logo_path.name}  ({logo.width}×{logo.height}px)")
        else:
            print(f"  ⚠  Логотип не найден: {logo_path}")

    # Данные из Excel
    print(f"\n  Читаю данные из {data_file.name}...")
    items = read_data(data_file)

    # Фильтрация по строкам
    if args.rows:
        items = filter_by_rows(items, args.rows)
        print(f"  Строки {args.rows} → выбрано товаров: {len(items)}")
        for code, item in items.items():
            print(f"    строка {item['row']}: {code}  {item['name'][:40]}")
    else:
        print(f"  Найдено товаров: {len(items)}")

    if not items:
        print("  ⚠  Нет товаров для обработки")
        sys.exit(0)

    print(f"\n  Товаров к обработке: {len(items)}")
    print(f"  Размер:              {IMG_W}×{IMG_H}px")
    print(f"  Результаты:          {out_dir}")
    print()
    print("─" * 55)

    done = 0
    skip = 0

    for code, item in items.items():
        image_file = item.get("image_file", "").strip()
        if not image_file:
            print(f"  ⚠  Строка {item['row']}: не указан файл изображения — пропуск")
            skip += 1
            continue

        # Ячейка может содержать несколько файлов через запятую, точку с запятой или перенос
        import re as _re
        raw_names = [n.strip() for n in _re.split(r"[,;\n]+", image_file) if n.strip()]

        # Разрешаем имена в файловую систему (регистронезависимо)
        def resolve(name: str) -> Path | None:
            p = images_dir / name
            if p.exists():
                return p
            # Поиск без учёта регистра
            low = name.lower()
            for f in images_dir.iterdir():
                if f.name.lower() == low:
                    return f
            return None

        print(f"  Строка {item['row']}:  {item['name'][:50]}")

        # Флаг: compat-фото создаётся один раз на товар (при первом найденном фото)
        logo_resized_shared = None
        compat_saved = False

        for raw_name in raw_names:
            photo_path = resolve(raw_name)
            if photo_path is None:
                print(f"  ⚠  Файл не найден: {raw_name}")
                skip += 1
                continue

            stem = photo_path.stem
            try:
                img = Image.open(photo_path).convert("RGB")

                # Фото 1 — название и артикул (для каждого файла)
                print(f"    → инфо ({photo_path.name})...", end=" ", flush=True)
                info_photo, logo_resized = make_info_photo(img, item, logo, IMG_W, IMG_H, bg)
                info_path = out_dir / f"{stem}_info.jpg"
                info_photo.save(info_path, "JPEG", quality=92)
                print(f"✓  {info_path.name}")

                if logo_resized_shared is None:
                    logo_resized_shared = logo_resized

                done += 1

            except Exception as e:
                print(f"\n    ✗ Ошибка ({raw_name}): {e}")

        # Фото 2 — применяемость (один раз на товар)
        if not compat_saved and logo_resized_shared is not None:
            try:
                first_stem = Path(raw_names[0]).stem
                print(f"    → применяемость...", end=" ", flush=True)
                compat_photo = make_compat_photo(item, logo_resized_shared, IMG_W, IMG_H, bg)
                compat_path  = out_dir / f"{first_stem}_compat.jpg"
                compat_photo.save(compat_path, "JPEG", quality=92)
                print(f"✓  {compat_path.name}")
                compat_saved = True
            except Exception as e:
                print(f"\n    ✗ Ошибка (compat): {e}")

        print()

    print("─" * 55)
    print(f"  Готово!")
    print(f"  Обработано: {done} фото  |  Пропущено: {skip}")
    print(f"  Результаты: {out_dir}")
    print()


if __name__ == "__main__":
    main()
