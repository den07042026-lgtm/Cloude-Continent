"""
autoliga_top500.py
════════════════════════════════════════════════════════════════════════════
Анализ полного прайса Автолиги → топ-500 товаров для выхода на WB.

Методы матчинга (применяются последовательно):
  М1. OEM-матч: Автолига артикул → wb_card_oem (basket CDN данные)
  М2. VendorCode-матч: артикул → vendor_codes таблица
  М3. Бренд+нормализованный артикул → wb_products прямо
  М4. WB поиск по артикулу (API) для топ-брендов без совпадений

Скоринг (0-100 баллов):
  - Маржа при цене WB          (30 баллов макс)
  - Спрос (продажи/мес)        (25 баллов макс)
  - Дефицит (OOS%)             (25 баллов макс)
  - Ценовое преимущество       (15 баллов макс)
  - Наличие у поставщика       (5  баллов макс)

Запуск:
  cd C:\\Users\\Admin\\Documents\\Autoparts_Ecommerce
  uv run --with openpyxl,xlrd,requests,python-dotenv scripts/autoliga_top500.py
"""

import sys, os, re, json, math, time, logging, sqlite3
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import xlrd
except ImportError:
    print("Установи: uv run --with openpyxl,xlrd scripts/autoliga_top500.py"); sys.exit(1)

try:
    import requests
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

BASE_DIR      = Path(__file__).parent.parent
AUTOLIGA_DIR  = BASE_DIR / "data" / "suppliers" / "autoliga"
DB_PATH       = BASE_DIR / "data" / "analytics" / "wb_index.db"
OUTPUT_DIR    = BASE_DIR / "data" / "analytics"
LOG_FILE      = BASE_DIR / "logs" / "autoliga_top500.log"
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)

MPSTATS_TOKEN = os.getenv("MPSTATS_TOKEN", "")
TOP_N         = 500
WB_SEARCH_WORKERS = 6
WB_SEARCH_LIMIT   = 300   # макс артикулов для WB-поиска

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# WB тарифы FBS Автозапчасти
_COMM = 0.25; _TAX = 0.06; _RET = 0.03
_PKG = 30; _DEL_B = 50.6; _DEL_L = 15.4
_RET_B = 136.0; _RET_L = 14.0; _VOL = 3.0

# Ценовой sweet spot: слишком дёшево — убивают комиссии
PRICE_MIN = 250
PRICE_MAX = 15000


# ─── Утилиты ──────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    return re.sub(r"[^A-ZА-ЯЁ0-9]", "", s.upper())

def _norm_brand(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", s.upper())

def _brands_eq(a: str, b: str) -> bool:
    a, b = _norm_brand(a), _norm_brand(b)
    if not a or not b: return False
    return a == b or a.startswith(b) or b.startswith(a)

def _margin(purchase: float, sell: float) -> float:
    if sell <= 0: return -99.0
    comm = sell * _COMM
    deliv = _DEL_B + _VOL * _DEL_L
    ret  = _RET * (_RET_B + _VOL * _RET_L)
    tax  = max(0.0, sell - comm - deliv) * _TAX
    return (sell - purchase - comm - deliv - ret - tax - _PKG) / sell * 100

def _find_price(purchase: float, target_margin: float) -> int | None:
    if purchase <= 0: return None
    if _margin(purchase, 500_000) < target_margin: return None
    lo, hi = 50, 500_000
    while lo < hi:
        mid = (lo + hi) // 2
        if _margin(purchase, mid) >= target_margin - 1e-6: hi = mid
        else: lo = mid + 1
    return lo if _margin(purchase, lo) >= target_margin - 1e-6 else None

def _is_valid_art(norm: str) -> bool:
    if len(norm) < 5: return False
    if norm.isdigit() and len(norm) < 7: return False
    return True


# ─── М0: Загрузка Автолиги ────────────────────────────────────────────────────

def load_autoliga() -> list[dict]:
    files = sorted(AUTOLIGA_DIR.glob("*.xls*"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        log.error("Прайс Автолиги не найден!"); sys.exit(1)
    path = files[0]
    log.info(f"Читаем прайс: {path.name}")

    wb  = xlrd.open_workbook(str(path), encoding_override="cp1251")
    ws  = wb.sheet_by_index(0)
    out = []
    for r in range(9, ws.nrows):
        row = ws.row_values(r)
        if len(row) < 8: continue
        brand   = str(row[1]).strip()
        art_raw = str(row[2]).strip()
        if art_raw.endswith(".0"): art_raw = art_raw[:-2]
        name    = str(row[4]).strip() if len(row) > 4 else ""
        try: stock = int(float(row[6]))
        except: stock = 0
        try: price = float(row[7])
        except: price = 0.0
        if not art_raw or price <= 0: continue
        n = _norm(art_raw)
        if not _is_valid_art(n): continue
        if price < PRICE_MIN or price > PRICE_MAX: continue
        out.append({
            "article": art_raw, "norm": n,
            "brand": brand, "name": name,
            "stock": stock, "price": price,
            "wb_matches": []  # заполним позже
        })
    log.info(f"Автолига: {len(out):,} позиций после фильтрации (цена {PRICE_MIN}–{PRICE_MAX}₽)")
    return out


# ─── М1: OEM-матч через wb_card_oem ──────────────────────────────────────────

def match_oem(al_items: list[dict], conn: sqlite3.Connection) -> int:
    log.info("М1: OEM-матч через wb_card_oem...")
    rows = conn.execute(
        "SELECT p.nm_id, p.brand, p.name, p.price_rub, p.sales_30d, "
        "       p.oos_pct, p.subject, c.oem_list "
        "FROM wb_products p JOIN wb_card_oem c ON p.nm_id=c.nm_id "
        "WHERE c.oem_list != '[]'"
    ).fetchall()

    # Индекс: norm_art → [(nm_id, wb_brand, wb_name, price, sales, oos, subject, orig_art)]
    idx: dict[str, list] = defaultdict(list)
    for nm_id, brand, name, price, sales, oos, subj, oem_json in rows:
        try: arts = json.loads(oem_json)
        except: continue
        for a in arts:
            n = _norm(str(a))
            if _is_valid_art(n):
                idx[n].append((nm_id, brand or "", name or "", float(price or 0),
                               int(sales or 0), float(oos or 0), subj or "", str(a)))

    new_matches = 0
    for item in al_items:
        hits = idx.get(item["norm"], [])
        for nm_id, wb_brand, wb_name, price, sales, oos, subj, orig_art in hits:
            if not _brands_eq(item["brand"], wb_brand): continue
            item["wb_matches"].append({
                "nm_id": nm_id, "wb_brand": wb_brand, "wb_name": wb_name,
                "wb_price": price, "sales_30d": sales, "oos_pct": oos,
                "subject": subj, "wb_art": orig_art, "method": "OEM"
            })
            new_matches += 1
    log.info(f"  М1 совпадений: {new_matches}")
    return new_matches


# ─── М2: VendorCode-матч ──────────────────────────────────────────────────────

def match_vendor_codes(al_items: list[dict], conn: sqlite3.Connection) -> int:
    log.info("М2: VendorCode-матч через vendor_codes...")
    try:
        rows = conn.execute(
            "SELECT v.nm_id, v.vc_raw, p.brand, p.name, p.price_rub, "
            "       p.sales_30d, p.oos_pct, p.subject "
            "FROM vendor_codes v JOIN wb_products p ON v.nm_id=p.nm_id "
            "WHERE v.vc_raw IS NOT NULL AND v.vc_raw != ''"
        ).fetchall()
    except Exception as e:
        log.warning(f"  vendor_codes недоступен: {e}"); return 0

    idx: dict[str, list] = defaultdict(list)
    for nm_id, vc_raw, brand, name, price, sales, oos, subj in rows:
        for part in re.split(r"[/;,]+", vc_raw or ""):
            n = _norm(part.strip())
            if _is_valid_art(n):
                idx[n].append((nm_id, brand or "", name or "", float(price or 0),
                               int(sales or 0), float(oos or 0), subj or "", part.strip()))

    already = {(item["norm"], m["nm_id"])
               for item in al_items for m in item["wb_matches"]}
    new_matches = 0
    for item in al_items:
        hits = idx.get(item["norm"], [])
        for nm_id, wb_brand, wb_name, price, sales, oos, subj, orig_art in hits:
            if (item["norm"], nm_id) in already: continue
            if not _brands_eq(item["brand"], wb_brand): continue
            item["wb_matches"].append({
                "nm_id": nm_id, "wb_brand": wb_brand, "wb_name": wb_name,
                "wb_price": price, "sales_30d": sales, "oos_pct": oos,
                "subject": subj, "wb_art": orig_art, "method": "VendorCode"
            })
            already.add((item["norm"], nm_id))
            new_matches += 1
    log.info(f"  М2 совпадений: {new_matches}")
    return new_matches


# ─── М3: Бренд + нормализованный артикул в wb_products.name ──────────────────

_M3_MIN_LEN = 7  # минимальная длина артикула для Brand+Name матча

def match_brand_art(al_items: list[dict], conn: sqlite3.Connection) -> int:
    log.info("М3: Бренд+артикул в wb_products (токенный матч)...")
    rows = conn.execute(
        "SELECT nm_id, brand, name, price_rub, sales_30d, oos_pct, subject "
        "FROM wb_products WHERE brand IS NOT NULL AND name IS NOT NULL"
    ).fetchall()

    # Индекс: norm_brand → [(nm_id, name_tokens_set, wb_name, price, sales, oos, subj)]
    # name_tokens_set — множество нормализованных токенов из wb_products.name
    brand_idx: dict[str, list] = defaultdict(list)
    for nm_id, brand, name, price, sales, oos, subj in rows:
        nb = _norm_brand(brand or "")
        if not nb: continue
        # Токенизация: разбиваем название на отдельные слова/коды
        tokens = set(re.findall(r"[A-ZА-ЯЁ0-9]{3,}", (name or "").upper()))
        if tokens:
            brand_idx[nb].append((nm_id, tokens, name or "",
                                  float(price or 0), int(sales or 0),
                                  float(oos or 0), subj or ""))

    already = {(item["norm"], m["nm_id"])
               for item in al_items for m in item["wb_matches"]}
    new_matches = 0
    for item in al_items:
        # Только длинные артикулы — короткие дают слишком много ложных совпадений
        if len(item["norm"]) < _M3_MIN_LEN:
            continue
        nb = _norm_brand(item["brand"])
        hits = brand_idx.get(nb, [])
        for nm_id, name_tokens, wb_name, price, sales, oos, subj in hits:
            if (item["norm"], nm_id) in already: continue
            # Артикул должен быть точным токеном в названии WB (не подстрокой!)
            if item["norm"] not in name_tokens: continue
            item["wb_matches"].append({
                "nm_id": nm_id, "wb_brand": item["brand"], "wb_name": wb_name,
                "wb_price": price, "sales_30d": sales, "oos_pct": oos,
                "subject": subj, "wb_art": item["article"], "method": "Brand+Name"
            })
            already.add((item["norm"], nm_id))
            new_matches += 1
    log.info(f"  М3 совпадений: {new_matches}")
    return new_matches


# ─── М4: WB search API для топ-брендов ───────────────────────────────────────

_WB_DEST = "-1257786"

def _wb_search_article(article: str, brand: str) -> list[dict]:
    """Ищет артикул на WB через search API. Возвращает список найденных товаров."""
    url = "https://search.wb.ru/exactmatch/ru/common/v9/search"
    params = {"appType": 1, "curr": "rub", "dest": _WB_DEST,
              "query": article, "resultset": "catalog", "sort": "popular",
              "spp": 30, "suppressSpellcheck": "false"}
    try:
        r = requests.get(url, params=params, timeout=8,
                         headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code != 200: return []
        data = r.json()
        products = (data.get("data") or {}).get("products") or []
        results = []
        for p in products[:5]:
            wb_brand = p.get("brand") or ""
            if not _brands_eq(brand, wb_brand): continue
            nm_id = int(p.get("id") or 0)
            if not nm_id: continue
            results.append({
                "nm_id":    nm_id,
                "wb_brand": wb_brand,
                "wb_name":  p.get("name") or "",
                "wb_price": float((p.get("sizes") or [{}])[0].get("price", {}).get("product", 0)) / 100
                            if (p.get("sizes") or [{}])[0].get("price") else
                            float(p.get("priceU", 0)) / 100,
                "sales_30d": 0,
                "oos_pct":   0.0,
                "subject":   "",
                "wb_art":    article,
                "method":    "WB_Search",
            })
        return results
    except Exception:
        return []


def match_wb_search(al_items: list[dict], conn: sqlite3.Connection) -> int:
    log.info("М4: WB search API для брендов без совпадений...")

    # Считаем бренды по числу неподобранных позиций
    unmatched_by_brand: dict[str, list] = defaultdict(list)
    for item in al_items:
        if not item["wb_matches"]:
            unmatched_by_brand[item["brand"]].append(item)

    # Сортируем бренды: у которых уже есть хоть несколько матчей — ищем ещё
    # Топ-бренды по числу артикулов
    top_brands = sorted(unmatched_by_brand.keys(),
                        key=lambda b: len(unmatched_by_brand[b]), reverse=True)

    # Выбираем артикулы для поиска (макс WB_SEARCH_LIMIT)
    to_search: list[dict] = []
    for brand in top_brands:
        items = unmatched_by_brand[brand]
        # Для каждого бренда берём первые 20 артикулов
        to_search.extend(items[:20])
        if len(to_search) >= WB_SEARCH_LIMIT:
            break
    to_search = to_search[:WB_SEARCH_LIMIT]
    log.info(f"  Поиск {len(to_search)} артикулов по WB API...")

    already = {(item["norm"], m["nm_id"])
               for item in al_items for m in item["wb_matches"]}
    # nm_id → WB данные из нашей БД
    wb_db: dict[int, tuple] = {
        r[0]: r for r in conn.execute(
            "SELECT nm_id, brand, name, price_rub, sales_30d, oos_pct, subject "
            "FROM wb_products"
        ).fetchall()
    }

    new_matches = 0
    ok = err = 0
    t0 = time.time()

    with ThreadPoolExecutor(max_workers=WB_SEARCH_WORKERS) as exe:
        futures = {exe.submit(_wb_search_article, it["article"], it["brand"]): it
                   for it in to_search}
        for i, fut in enumerate(as_completed(futures), 1):
            item = futures[fut]
            results = fut.result()
            for hit in results:
                nm_id = hit["nm_id"]
                if (item["norm"], nm_id) in already: continue
                # Обогащаем данными из БД если есть
                if nm_id in wb_db:
                    r = wb_db[nm_id]
                    hit["wb_brand"]  = r[1] or hit["wb_brand"]
                    hit["wb_name"]   = r[2] or hit["wb_name"]
                    hit["wb_price"]  = float(r[3] or hit["wb_price"])
                    hit["sales_30d"] = int(r[4] or 0)
                    hit["oos_pct"]   = float(r[5] or 0)
                    hit["subject"]   = r[6] or ""
                item["wb_matches"].append(hit)
                already.add((item["norm"], nm_id))
                new_matches += 1
                ok += 1
            if not results: err += 1
            if i % 50 == 0:
                log.info(f"  [{i}/{len(to_search)}]  найдено={ok}  без рез.={err}  "
                         f"{i/(time.time()-t0):.1f} req/с")

    log.info(f"  М4 совпадений: {new_matches}")
    return new_matches


# ─── Скоринг ──────────────────────────────────────────────────────────────────

def score_match(item: dict, match: dict) -> float:
    """Возвращает скор 0-100 для пары (Автолига артикул, WB карточка)."""
    purchase  = item["price"]
    wb_price  = match["wb_price"]
    sales     = match["sales_30d"]
    oos       = match["oos_pct"]
    stock     = item["stock"]

    if wb_price <= 0: return 0.0

    margin = _margin(purchase, wb_price)
    opt_15 = _find_price(purchase, 15.0)

    # 1. Маржа (30 баллов)
    if margin >= 25:    margin_score = 30
    elif margin >= 15:  margin_score = 25
    elif margin >= 5:   margin_score = 15
    elif margin >= 0:   margin_score = 8
    elif margin >= -10: margin_score = 2
    else:               margin_score = 0

    # 2. Спрос (25 баллов)
    demand_score = min(25, math.log1p(sales) / math.log1p(200) * 25)

    # 3. Дефицит OOS (25 баллов)
    deficit_score = min(25, oos / 100 * 25)

    # 4. Ценовое преимущество (15 баллов)
    if opt_15 and wb_price > 0:
        price_gap = (wb_price - opt_15) / wb_price
        price_score = min(15, max(0, price_gap * 50))
    else:
        price_score = 0

    # 5. Наличие (5 баллов)
    stock_score = min(5, stock / 10)

    total = margin_score + demand_score + deficit_score + price_score + stock_score
    return round(total, 2)


def classify_strategy(item: dict, match: dict, score: float) -> str:
    """Определяет стратегию входа."""
    margin   = _margin(item["price"], match["wb_price"])
    oos      = match["oos_pct"]
    sales    = match["sales_30d"]
    opt_15   = _find_price(item["price"], 15.0) or 0
    gap_pct  = (match["wb_price"] - opt_15) / match["wb_price"] * 100 if match["wb_price"] > 0 and opt_15 > 0 else 0

    if oos >= 50 and sales >= 15:
        return "Острый дефицит"
    elif margin >= 20 and gap_pct >= 15:
        return "Ценовое преимущество"
    elif oos >= 25 and sales >= 8:
        return "Дефицит"
    elif margin >= 15 and sales >= 20:
        return "Стабильный спрос"
    elif margin >= 10 and sales >= 10:
        return "Хорошая маржа"
    elif gap_pct >= 25:
        return "Ценовой арбитраж"
    else:
        return "Потенциал"


def build_justification(item: dict, match: dict, score: float, strategy: str) -> str:
    margin  = _margin(item["price"], match["wb_price"])
    opt_0   = _find_price(item["price"], 0.0)
    opt_15  = _find_price(item["price"], 15.0)
    gap     = ((match["wb_price"] - (opt_15 or match["wb_price"])) / match["wb_price"] * 100
               if match["wb_price"] > 0 and opt_15 else 0)

    parts = [f"[{strategy}]"]
    parts.append(f"WB: {match['sales_30d']} прод/мес, OOS {match['oos_pct']:.0f}%.")
    parts.append(f"Цена WB {match['wb_price']:.0f}₽, закупка {item['price']:.0f}₽.")

    if margin >= 0:
        parts.append(f"Маржа при текущей цене WB: {margin:.1f}%.")
    else:
        parts.append(f"При цене WB убыток {abs(margin):.1f}%.")

    if opt_0:  parts.append(f"Безубыточность от {opt_0}₽.")
    if opt_15: parts.append(f"15% маржа от {opt_15}₽.")

    if gap > 0:
        parts.append(f"Ценовое окно: можно зайти на {gap:.0f}% ниже рынка с 15% маржой.")

    if match["oos_pct"] >= 50:
        parts.append(f"Конкурент в OOS {match['oos_pct']:.0f}% времени — постоянный дефицит.")
    elif match["oos_pct"] >= 25:
        parts.append(f"Периодический дефицит у конкурентов ({match['oos_pct']:.0f}% OOS).")

    if match["sales_30d"] >= 50:
        parts.append("Высокий спрос в нише.")
    elif match["sales_30d"] >= 20:
        parts.append("Стабильный спрос.")

    if item["stock"] >= 20:
        parts.append(f"Хороший запас у поставщика ({item['stock']} шт).")

    if match["method"] != "OEM":
        parts.append(f"Матч: {match['method']}.")

    return " ".join(parts)


# ─── Финальный отбор ──────────────────────────────────────────────────────────

def select_top(al_items: list[dict]) -> list[dict]:
    """Выбирает лучший WB-матч для каждого артикула и скорит."""
    candidates = []
    for item in al_items:
        if not item["wb_matches"]: continue
        best = max(item["wb_matches"], key=lambda m: score_match(item, m))
        s = score_match(item, best)
        if s < 5: continue
        wb_p = best["wb_price"]
        pur  = item["price"]
        # Фильтр: WB цена должна быть хотя бы 80% от закупки (иначе ложный матч)
        if wb_p > 0 and pur > 0 and wb_p < pur * 0.80:
            continue  # ложный матч: WB дешевле нашей закупки
        # Фильтр: наша минимальная цена не должна быть > 3x WB
        opt_15 = _find_price(pur, 15.0) or 0
        if opt_15 > 0 and wb_p > 0 and opt_15 > wb_p * 3.0:
            continue  # подозрительный матч: цены на разных уровнях
        strategy = classify_strategy(item, best, s)
        justif   = build_justification(item, best, s, strategy)
        candidates.append({
            "article":      item["article"],
            "brand":        item["brand"],
            "name":         item["name"],
            "purchase":     item["price"],
            "stock":        item["stock"],
            "nm_id":        best["nm_id"],
            "wb_brand":     best["wb_brand"],
            "wb_name":      best["wb_name"],
            "wb_price":     best["wb_price"],
            "wb_art":       best["wb_art"],
            "wb_subject":   best["subject"],
            "sales_30d":    best["sales_30d"],
            "oos_pct":      best["oos_pct"],
            "margin_at_wb": round(_margin(item["price"], best["wb_price"]), 1),
            "sell_min":     _find_price(item["price"], 0.0),
            "sell_opt":     _find_price(item["price"], 15.0),
            "method":       best["method"],
            "score":        s,
            "strategy":     strategy,
            "justification": justif,
        })

    candidates.sort(key=lambda x: x["score"], reverse=True)
    # Ограничиваем 1 карточкой WB на nm_id
    seen_nm: set = set()
    result = []
    for c in candidates:
        if c["nm_id"] in seen_nm: continue
        seen_nm.add(c["nm_id"])
        result.append(c)
        if len(result) >= TOP_N: break

    log.info(f"Кандидатов всего: {len(candidates)}, в топ-{TOP_N}: {len(result)}")
    return result


# ─── Excel ────────────────────────────────────────────────────────────────────

STRAT_COLORS = {
    "Острый дефицит":      "8B0000",
    "Дефицит":             "C0392B",
    "Ценовое преимущество":"1A5C1A",
    "Стабильный спрос":    "1F4E79",
    "Хорошая маржа":       "2E86AB",
    "Ценовой арбитраж":    "7B4800",
    "Потенциал":           "555555",
}

_BORDER = Border(
    left=Side(style="thin",   color="CCCCCC"),
    right=Side(style="thin",  color="CCCCCC"),
    bottom=Side(style="thin", color="EEEEEE"),
)
RED_F = PatternFill("solid", fgColor="FFC7CE")
YLW_F = PatternFill("solid", fgColor="FFEB9C")
GRN_F = PatternFill("solid", fgColor="C6EFCE")
BLU_F = PatternFill("solid", fgColor="D6E4F0")

HEADERS = [
    ("№",                    4),
    ("Стратегия",           18),
    ("Артикул\nпроизводителя", 20),
    ("Бренд",               16),
    ("Название\n(Автолига)", 35),
    ("Цена\nзакупки, ₽",   12),
    ("Остаток\nшт.",         8),
    ("nm_id WB",            12),
    ("Название WB",         40),
    ("Бренд WB",            16),
    ("Предмет WB",          28),
    ("Цена WB, ₽",         12),
    ("OOS %",                8),
    ("Продаж/\nмес",        10),
    ("Скор",                 8),
    ("Маржа\nпри WB, %",   12),
    ("Цена мин.\n(0%), ₽", 13),
    ("Цена опт.\n(15%), ₽",14),
    ("Метод\nматча",        12),
    ("Ссылка WB",           12),
    ("Обоснование",         70),
]

MARGIN_COL = 16
OOS_COL    = 13
LINK_COL   = 20
STRAT_COL  = 2


def save_excel(items: list[dict]) -> Path:
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = f"Топ-{TOP_N}"

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append([h[0] for h in HEADERS])
    for ci, (_, w) in enumerate(HEADERS, 1):
        c = ws.cell(1, ci)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    for i, it in enumerate(items, 1):
        url = f"https://www.wildberries.ru/catalog/{it['nm_id']}/detail.aspx"
        row = [
            i,
            it["strategy"],
            it["article"],
            it["brand"],
            it["name"],
            it["purchase"],
            it["stock"],
            it["nm_id"],
            it["wb_name"],
            it["wb_brand"],
            it["wb_subject"],
            it["wb_price"] or "—",
            it["oos_pct"],
            it["sales_30d"],
            it["score"],
            it["margin_at_wb"] if it["margin_at_wb"] is not None else "—",
            it["sell_min"] or "—",
            it["sell_opt"] or "—",
            it["method"],
            url,
            it["justification"],
        ]
        ri = i + 1
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.border = _BORDER
            c.alignment = Alignment(vertical="center",
                                    wrap_text=(ci == len(HEADERS)))

        # Ссылка
        lc = ws.cell(ri, LINK_COL)
        lc.hyperlink = url
        lc.font = Font(color="0563C1", underline="single", size=9)

        # Стратегия — цветная метка
        sc = ws.cell(ri, STRAT_COL)
        color = STRAT_COLORS.get(it["strategy"], "555555")
        sc.fill = PatternFill("solid", fgColor=color)
        sc.font = Font(color="FFFFFF", bold=True, size=8)
        sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # OOS цвет
        oos = it["oos_pct"]
        if oos >= 50:   ws.cell(ri, OOS_COL).fill = RED_F
        elif oos >= 25: ws.cell(ri, OOS_COL).fill = YLW_F

        # Маржа цвет
        mrg = it["margin_at_wb"]
        if mrg is not None:
            ws.cell(ri, MARGIN_COL).fill = (
                GRN_F if mrg >= 15 else
                YLW_F if mrg >= 0  else
                RED_F
            )

        # Чередование строк
        if i % 2 == 0:
            for ci in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 17, 18, 19]:
                ws.cell(ri, ci).fill = BLU_F

    # Лист со статистикой
    ws2 = wb_out.create_sheet("Статистика")
    from collections import Counter
    strat_cnt = Counter(it["strategy"] for it in items)
    brand_cnt = Counter(it["brand"] for it in items)
    method_cnt = Counter(it["method"] for it in items)

    ws2.append(["Стратегия", "Кол-во"]); ws2.append([])
    for k, v in strat_cnt.most_common(): ws2.append([k, v])
    ws2.append([]); ws2.append(["Бренд (топ-20)", "Кол-во"])
    for k, v in brand_cnt.most_common(20): ws2.append([k, v])
    ws2.append([]); ws2.append(["Метод матча", "Кол-во"])
    for k, v in method_cnt.most_common(): ws2.append([k, v])

    avg_score  = sum(it["score"] for it in items) / len(items) if items else 0
    avg_margin = sum(it["margin_at_wb"] for it in items if it["margin_at_wb"]) / max(1, sum(1 for it in items if it["margin_at_wb"]))
    ws2.append([]); ws2.append(["Средний скор", round(avg_score, 1)])
    ws2.append(["Средняя маржа при цене WB", round(avg_margin, 1)])
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 12

    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / f"autoliga_top500_{ts}.xlsx"
    wb_out.save(str(path))
    return path


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 65)
    log.info(f"Autoliga Top-{TOP_N}  |  Полный анализ прайса → WB")
    log.info("=" * 65)
    t0 = time.time()

    # Загружаем Автолигу
    al_items = load_autoliga()

    # Подключаем БД
    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")

    # Применяем матчинг методами
    m1 = match_oem(al_items, conn)
    m2 = match_vendor_codes(al_items, conn)
    m3 = match_brand_art(al_items, conn)
    m4 = match_wb_search(al_items, conn)

    matched_items = sum(1 for it in al_items if it["wb_matches"])
    log.info(f"Итого артикулов с матчем: {matched_items:,} "
             f"(M1={m1} M2={m2} M3={m3} M4={m4})")

    conn.close()

    # Скоринг и отбор топ-N
    log.info("Скоринг и отбор...")
    top = select_top(al_items)

    # Статистика стратегий
    from collections import Counter
    strat = Counter(it["strategy"] for it in top)
    log.info("Стратегии в топе:")
    for k, v in strat.most_common():
        log.info(f"  {k:25s}: {v}")

    # Экспорт
    log.info("Экспорт Excel...")
    path = save_excel(top)

    elapsed = time.time() - t0
    log.info("=" * 65)
    log.info(f"Готово за {elapsed:.0f}с  |  Файл: {path}")
    log.info("=" * 65)

    # Топ-15
    log.info("\nТОП-15:")
    for i, it in enumerate(top[:15], 1):
        log.info(f"  {i:2}. [{it['strategy']:20s}] скор={it['score']:5.1f}  "
                 f"OOS={it['oos_pct']:4.0f}%  прод={it['sales_30d']:3d}  "
                 f"марж={it['margin_at_wb']:+5.1f}%  "
                 f"{it['brand']:12s} {it['article']}")


if __name__ == "__main__":
    main()
