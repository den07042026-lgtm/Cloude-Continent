"""
ozon_top500_selector.py
========================================================================
Отбор топ-500 товаров из свежего прайса Микадо для загрузки на Ozon
под новые 500 слотов. Закупка >= 300 руб. Цель: максимальный оборот и прибыль.

Источники:
  1. Ozon Seller API — наш текущий каталог (/v3/product/list) + аналитика
     продаж за 90 дней (/v1/analytics/data: revenue, ordered_units).
     Даёт РЕАЛЬНЫЙ сигнал, какие бренды/категории уже продаются у НАС на Ozon.
  2. Свежий прайс Микадо (MIKADO_PRICE_URL, без логина) — 39 125 позиций,
     закупочная цена, остаток, бренд.
  3. pricing_engine.OzonPricer — расчёт рекомендованной цены продажи и маржи
     по фактической ценовой политике магазина (get_optimal_markup, та же
     формула, что использует price_recalc.py в проде).
  4. Универсальные сигналы спроса (перенесены из wb_top500_combined.py):
     CAR_FLEET (популярность моделей авто в РФ), PART_FREQ (частота замены
     детали / "вечный спрос").

Результат: data/analytics/top500_ozon/Топ-500_Ozon_<дата>.xlsx с обоснованием
по каждой позиции.

Запуск:
  cd C:\\Users\\Admin\\Documents\\Autoparts_Ecommerce
  uv run --with requests,openpyxl scripts/ozon_top500_selector.py
"""

import sys
import os
import re
import json
import math
import time
import logging
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-7s %(message)s")
log = logging.getLogger(__name__)

ROOT = Path(__file__).parent.parent
ENV_FILE = ROOT / ".env"
CACHE_DIR = ROOT / "data" / "analytics" / "top500_ozon"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

MIKADO_PRICE_URL = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=YOUR_MIKADO_PRICE_KEY"
)

MIN_PURCHASE = 300.0
TOP_N = 500
MAX_PER_BRAND = 25
MAX_PER_CATEGORY = 30
MIN_MARGIN_PCT = 8.0
ANALYTICS_DAYS = 90
DEFAULT_LOGISTICS = 115  # руб, дефолт без габаритов (как в price_recalc.py)


def load_env() -> dict:
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


ENV = load_env()
OZON_CLIENT_ID = ENV.get("OZON_CLIENT_ID", "")
OZON_API_KEY = ENV.get("OZON_API_KEY", "")
OZON_BASE = "https://api-seller.ozon.ru"
HEADERS = {
    "Client-Id": OZON_CLIENT_ID,
    "Api-Key": OZON_API_KEY,
    "Content-Type": "application/json",
}


# ══════════════════════════════════════════════════════════════════════════
#  OZON PRICER (идентично scripts/pricing_engine.py OzonPricer + get_optimal_markup)
# ══════════════════════════════════════════════════════════════════════════

FBS_TIERS = [100, 300, 1500, 5000, 10000]
FBS_RATES = [0.14, 0.20, 0.44, 0.44, 0.44, 0.44]
ACQ_PCT = 0.015
TAX_PCT = 0.06
RET_RATE = 0.03
REVERSE = 80
OTHER = 30
OPTIMUM_OTHER = 20

LOG_TABLE = [
    (0.5, 75), (1, 90), (2, 115), (5, 155), (10, 210),
    (15, 265), (20, 315), (25, 365), (30, 420), (50, 620),
]


def _fbs_rate(sell: float) -> float:
    for thresh, rate in zip(FBS_TIERS, FBS_RATES):
        if sell < thresh:
            return rate
    return FBS_RATES[-1]


def calc_profit(purchase: float, sell: float, logistics: float) -> float:
    commission = sell * _fbs_rate(sell)
    acquiring = sell * ACQ_PCT
    return_loss = RET_RATE * (logistics + REVERSE)
    proceeds = sell - commission - acquiring - logistics
    tax = max(0.0, proceeds) * TAX_PCT
    total_cost = purchase + commission + acquiring + logistics + return_loss + OTHER + tax
    return sell - total_cost


def get_optimal_markup(purchase: float) -> float:
    if purchase < 500: return 0.30
    if purchase < 1200: return 0.25
    if purchase < 2500: return 0.22
    if purchase < 3500: return 0.20
    return 0.17


def find_price_optimum(purchase: float, logistics: float) -> int | None:
    target = get_optimal_markup(purchase)
    cost = purchase + OPTIMUM_OTHER
    for s in range(50, 500_001, 1):
        profit = calc_profit(purchase, s, logistics)
        if profit / cost >= target - 1e-6:
            return s
    return None


def calc_margin_pct(purchase: float, sell: float, logistics: float) -> float:
    if sell <= 0:
        return 0.0
    return calc_profit(purchase, sell, logistics) / sell * 100


# ══════════════════════════════════════════════════════════════════════════
#  УНИВЕРСАЛЬНЫЕ СИГНАЛЫ СПРОСА (перенесены из wb_top500_combined.py)
# ══════════════════════════════════════════════════════════════════════════

CAR_FLEET = {
    "2107": 1177, "2106": 850, "2109": 900, "2110": 600, "2114": 400,
    "2115": 350, "priora": 800, "приора": 800,
    "granta": 1070, "гранта": 1070,
    "vesta": 750, "веста": 750,
    "niva": 1004, "нива": 1004,
    "kalina": 650, "калина": 650,
    "largus": 350, "ларгус": 350,
    "rio": 1020, "рио": 1020,
    "sportage": 380, "спортейдж": 380,
    "cerato": 280, "серато": 280,
    "solaris": 926, "солярис": 926,
    "creta": 420, "крета": 420,
    "tucson": 310, "туксон": 310,
    "elantra": 250, "элантра": 250,
    "logan": 774, "логан": 774,
    "duster": 320, "дастер": 320,
    "sandero": 280, "сандеро": 280,
    "focus": 742, "фокус": 742,
    "mondeo": 220, "мондео": 220,
    "polo": 500, "поло": 500,
    "passat": 280, "пассат": 280,
    "camry": 350, "камри": 350,
    "rav4": 280, "rav 4": 280, "corolla": 260, "королла": 260,
    "almera": 380, "альмера": 380,
    "x-trail": 270, "qashqai": 200,
    "rapid": 350, "рапид": 350,
    "octavia": 280, "октавия": 280,
    "outlander": 280, "аутлендер": 280,
    "lancer": 250, "лансер": 250,
    "lacetti": 380, "лачетти": 380,
    "cruze": 250, "круз": 250,
    "aveo": 200, "авео": 200,
    "haval": 350, "хавал": 350, "jolion": 200, "h6": 150,
    "geely": 280, "джили": 280, "atlas": 180, "coolray": 120,
    "chery": 220, "чери": 220, "tiggo": 180,
}

PART_FREQ = {
    "фильтр масл": 6.0, "масляный фильтр": 6.0,
    "фильтр воздуш": 3.5, "воздушный фильтр": 3.5,
    "фильтр салон": 3.5, "салонный фильтр": 3.5,
    "фильтр топлив": 3.0, "топливный фильтр": 3.0,
    "свеч": 2.5,
    "масло мотор": 6.0, "моторное масло": 6.0,
    "антифриз": 1.5, "охлаждающ жидк": 1.5,
    "жидкость тормозн": 1.5,
    "тормозн колодк": 2.5, "колодки тормозн": 2.5,
    "тормозной диск": 1.5, "тормозн диск": 1.5,
    "амортизатор": 1.5, "аморт": 1.5,
    "стойк стабил": 2.0, "стабилизатор": 1.5,
    "сайлентблок": 2.0,
    "шаровая": 1.5,
    "ступичный подшипн": 1.5, "подшипник ступиц": 1.5,
    "рычаг подвески": 1.2,
    "шрус": 1.2,
    "ремень грм": 1.0, "ремень привода": 1.5, "ремень генератор": 1.5,
    "натяжитель": 1.0,
    "цепь грм": 0.8,
    "рулевая рейка": 0.8, "рулевой наконечн": 1.2,
    "помпа": 1.0, "водяной насос": 1.0,
    "радиатор": 0.8,
    "катушка зажиган": 1.2,
    "генератор": 0.6,
    "стартер": 0.5,
    "дворники": 2.0, "щетки стеклооч": 2.0,
    "прокладка гбц": 0.5, "болт гбц": 0.5,
    "форсунк": 1.0,
    "термостат": 0.8,
    "аккумулятор": 0.7,
    "лампа": 2.0, "лампочка": 2.0,
    "бачок расшир": 1.0,
    "трос": 0.9,
    "зеркало": 0.6,
    "кольцо поршнев": 0.8,
    "бампер": 0.5,
}

# ── Категории (человекочитаемые бакеты для сведения статистики) ────────────
CATEGORY_KW = [
    ("Фильтры масляные", ["фильтр масл", "масляный фильтр"]),
    ("Фильтры воздушные", ["фильтр воздуш", "воздушный фильтр"]),
    ("Фильтры салонные", ["фильтр салон", "салонный фильтр"]),
    ("Фильтры топливные", ["фильтр топлив", "топливный фильтр"]),
    ("Свечи зажигания", ["свеч"]),
    ("Колодки тормозные", ["колодк"]),
    ("Тормозные диски", ["тормозн диск", "тормозной диск", "диск.торм", "торм.диск", "диск торм"]),
    ("Подшипники ступицы", ["подшипник ступиц", "ступичный подшипн"]),
    ("Амортизаторы/стойки", ["амортизатор", "аморт.", "стойка амортиз"]),
    ("Опоры амортизатора", ["опора аморт", "опора.(верхн"]),
    ("Стойки стабилизатора", ["стойк стабил", "стабилизатор"]),
    ("Сайлентблоки", ["сайлентблок", "сайленблок"]),
    ("Шаровые опоры", ["шаровая"]),
    ("ШРУСы", ["шрус"]),
    ("Ремни ГРМ/привода", ["ремень грм", "ремень привода", "ремень генератор"]),
    ("Катушки зажигания", ["катушка зажиган"]),
    ("Генераторы", ["генератор"]),
    ("Стартеры", ["стартер"]),
    ("Насосы/помпы", ["помпа", "водяной насос", "насос топлив", "насос водян", "насос гур", "насос гидроусил"]),
    ("Форсунки", ["форсунк"]),
    ("Радиаторы", ["радиатор"]),
    ("Термостаты", ["термостат"]),
    ("Сальники", ["сальник"]),
    ("Тросы", ["трос"]),
    ("Бачки расширительные", ["бачок расшир", "бачёк расшир"]),
    ("Болты ГБЦ", ["болт гбц"]),
    ("Прокладки ГБЦ", ["прокладк"]),
    ("Кольца поршневые", ["кольцо поршнев", "кольца поршнев"]),
    ("Зеркала", ["зеркал"]),
    ("Бамперы", ["бампер"]),
    ("Дроссельные заслонки", ["дроссельн"]),
    ("Датчики", ["датчик"]),
    ("Фонари/лампы", ["фонар", "лампа", "лампочка", "фара"]),
    ("Аккумуляторы", ["аккумулятор", "акб"]),
    ("Рулевые рейки/наконечники", ["рулев"]),
    ("Рычаги подвески", ["рычаг подвес"]),
    # ── Категории, добавленные после аудита "Прочее" (42% пула не классифицировалось) ──
    ("Ролики натяжные/обводные", ["ролик"]),
    ("Диски/корзины сцепления", ["диск сцеплен", "корзина сцеплен", "ведомый диск", "нажимной диск"]),
    ("Пружины подвески", ["пружина"]),
    ("Ступицы в сборе", ["ступиц"]),
    ("Опоры/подушки двигателя", ["опора двигат", "подушка двигат", "опора кпп", "опора силового"]),
    ("Тормозные цилиндры", ["цилиндр тормозн", "рабочий цилиндр", "главный цилиндр", "цилиндр рабочий", "цилиндр главн"]),
    ("Поршни", ["поршень"]),
    ("Дворники/щётки стеклоочистителя", ["дворник", "щетка стеклооч", "щётка стеклооч", "щетки стеклооч"]),
    ("Клапаны", ["клапан"]),
    ("Выхлопная система", ["глушитель", "резонатор", "приемная труба", "приёмная труба"]),
    ("Рычаги/тяги подвески", ["рычаг", "тяга рулев", "тяга поперечн"]),
    ("Вкладыши двигателя", ["вкладыш"]),
    ("Крестовины карданные", ["крестовин"]),
    ("Тормозные барабаны", ["барабан"]),
    ("Патрубки/шланги/гофры", ["патрубок", "гофра", "шланг"]),
    ("Наконечники рулевые", ["наконечник"]),
    ("Пыльники ШРУС/аморт.", ["пыльник"]),
    ("Натяжители цепи/ремня", ["натяжитель"]),
    ("Втулки стабилизатора", ["втулк"]),
    ("Сцепление (комплект)", ["сцепл"]),
    ("Вентиляторы охлаждения", ["вентилятор"]),
    ("Подшипники (общие)", ["подшипник"]),
    ("Ремкомплекты тормозного суппорта", ["ремкомплект", "суппорта", "тормозного механизма"]),
]


def classify_category(name: str) -> str:
    low = name.lower()
    for cat, kws in CATEGORY_KW:
        if any(kw in low for kw in kws):
            return cat
    return "Прочее"


# Жидкости (моторное масло, антифриз/охлаждающая жидкость) исключаются из
# отбора по требованию пользователя - не сами фильтры/насосы, а именно
# готовая жидкость в таре.
LIQUID_EXCLUDE_KW = ["антифриз", "охлаждающая жидк", "охлаждающ. жидк", "тосол"]


def is_excluded_liquid(name: str) -> bool:
    low = name.lower()
    if any(kw in low for kw in LIQUID_EXCLUDE_KW):
        return True
    if "масло" in low and "фильтр" not in low and "насос" not in low:
        return True
    return False


def match_kw_value(name: str, table: dict) -> float:
    low = name.lower()
    best = 0.0
    for kw, val in table.items():
        if kw in low and val > best:
            best = val
    return best


# ══════════════════════════════════════════════════════════════════════════
#  ПУБЛИЧНЫЙ СПРОС OZON (ozon_public_demand_scan.py) — открытый источник:
#  среднее число отзывов по категории в реальной выдаче ozon.ru (без логина,
#  без платных API). Отзывы начисляются только подтверждённым покупателям,
#  поэтому это рабочий бесплатный прокси относительного спроса между
#  категориями именно на Ozon (в отличие от WB/MPStats).
# ══════════════════════════════════════════════════════════════════════════

PUBLIC_DEMAND_FILE = ROOT / "data" / "analytics" / "top500_ozon" / "ozon_public_demand.json"


def load_public_demand() -> dict[str, dict]:
    if not PUBLIC_DEMAND_FILE.exists():
        log.warning("Нет файла %s — сначала запустите ozon_public_demand_scan.py", PUBLIC_DEMAND_FILE)
        return {}
    return json.loads(PUBLIC_DEMAND_FILE.read_text(encoding="utf-8"))


# ══════════════════════════════════════════════════════════════════════════
#  OZON API
# ══════════════════════════════════════════════════════════════════════════

def _post(path: str, body: dict, retry: int = 4) -> dict:
    url = OZON_BASE + path
    for attempt in range(retry):
        try:
            r = requests.post(url, json=body, headers=HEADERS, timeout=45)
            if r.status_code == 429:
                wait = int(r.headers.get("Retry-After", 15))
                log.warning("Rate-limit %s -> ждём %ds", path, wait)
                time.sleep(wait)
                continue
            if r.status_code >= 500:
                time.sleep(3 ** attempt)
                continue
            r.raise_for_status()
            return r.json()
        except requests.RequestException as e:
            if attempt == retry - 1:
                log.error("Ошибка %s: %s", path, e)
                return {}
            time.sleep(2 ** attempt)
    return {}


def fetch_ozon_catalog() -> list[dict]:
    """Все товары каталога: [{product_id, offer_id, sku, archived}]"""
    out, last_id = [], ""
    while True:
        resp = _post("/v3/product/list", {
            "filter": {"visibility": "ALL"},
            "last_id": last_id,
            "limit": 1000,
        })
        items = resp.get("result", {}).get("items", [])
        if not items:
            break
        out.extend(items)
        last_id = resp.get("result", {}).get("last_id", "")
        if not last_id or len(items) < 1000:
            break
        time.sleep(0.3)
    log.info("Ozon: каталог получен, товаров = %d", len(out))
    return out


def fetch_ozon_analytics(days: int) -> dict[str, dict]:
    """{sku_str: {revenue, ordered_units, name}} за N дней."""
    date_to = datetime.now().strftime("%Y-%m-%d")
    date_from = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    metrics = ["revenue", "ordered_units"]
    result, offset = {}, 0
    while True:
        resp = _post("/v1/analytics/data", {
            "date_from": date_from,
            "date_to": date_to,
            "dimension": ["sku"],
            "filters": [],
            "limit": 1000,
            "offset": offset,
            "metrics": metrics,
        })
        data = resp.get("result", {}).get("data", [])
        if not data:
            break
        for row in data:
            dims = row.get("dimensions", [{}])
            sku_id = dims[0].get("id", "") if dims else ""
            name = dims[0].get("name", "") if dims else ""
            vals = row.get("metrics", [0, 0])
            result[sku_id] = {"revenue": vals[0], "ordered_units": vals[1], "name": name}
        offset += len(data)
        if len(data) < 1000:
            break
        time.sleep(0.3)
    log.info("Ozon: аналитика за %d дней получена, SKU = %d", days, len(result))
    return result


# ══════════════════════════════════════════════════════════════════════════
#  MIKADO — свежий прайс
# ══════════════════════════════════════════════════════════════════════════

def fetch_mikado_fresh(max_age_hours: float = 12.0) -> Path:
    out_path = CACHE_DIR / "mikado_price_fresh.xlsx"
    if out_path.exists():
        age_h = (time.time() - out_path.stat().st_mtime) / 3600
        if age_h <= max_age_hours:
            log.info("Микадо: использую кэш (%.1f ч. назад) -> %s", age_h, out_path)
            return out_path
    for attempt in range(3):
        try:
            r = requests.get(MIKADO_PRICE_URL, timeout=90)
            r.raise_for_status()
            if r.content[:2] != b"PK":
                raise RuntimeError("Микадо вернул не xlsx (возможно, требуется логин/сессия)")
            out_path.write_bytes(r.content)
            log.info("Микадо: свежий прайс скачан (%d байт) -> %s", len(r.content), out_path)
            return out_path
        except requests.RequestException as e:
            log.warning("Микадо: попытка %d не удалась (%s)", attempt + 1, e)
            time.sleep(5)
    if out_path.exists():
        log.warning("Микадо: сервер недоступен, использую последний имеющийся кэш (%s)", out_path)
        return out_path
    raise RuntimeError("Микадо недоступен и нет кэша")


def load_mikado_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    idx = {name: i for i, name in enumerate(header)}
    out = []
    for row in rows_iter:
        if not row or row[idx["Code"]] is None:
            continue
        try:
            price = float(str(row[idx["PriceOut"]]).replace(",", "."))
        except (TypeError, ValueError):
            continue
        try:
            qty = int(float(row[idx["QTY"]] or 0))
        except (TypeError, ValueError):
            qty = 0
        out.append({
            "code": str(row[idx["Code"]]).strip(),
            "brand": str(row[idx["BrandName"]] or "").strip(),
            "name": str(row[idx["Prodname"]] or "").strip(),
            "price": price,
            "qty": qty,
        })
    log.info("Микадо: строк прочитано = %d", len(out))
    return out


# ══════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════

def main():
    if not OZON_CLIENT_ID or not OZON_API_KEY:
        sys.exit("Нет OZON_CLIENT_ID/OZON_API_KEY в .env")

    # 1. Текущий каталог + аналитика Ozon
    catalog = fetch_ozon_catalog()
    (CACHE_DIR / "ozon_catalog.json").write_text(json.dumps(catalog, ensure_ascii=False), encoding="utf-8")

    analytics = fetch_ozon_analytics(ANALYTICS_DAYS)
    (CACHE_DIR / "ozon_analytics.json").write_text(json.dumps(analytics, ensure_ascii=False), encoding="utf-8")

    # активные (неархивные) offer_id -> код Mikado (без "-con"), sku
    active_codes = set()
    sku_to_offer = {}
    for item in catalog:
        if item.get("archived"):
            continue
        offer_id = item.get("offer_id", "")
        code = re.sub(r"-con$", "", offer_id, flags=re.IGNORECASE).lower()
        active_codes.add(code)
        sku_to_offer[str(item.get("sku", ""))] = code

    # 2. Свежий прайс Микадо
    mikado_path = fetch_mikado_fresh()
    mikado_rows = load_mikado_rows(mikado_path)

    code_to_mikado = {r["code"].lower(): r for r in mikado_rows}

    # 3. Агрегация оборота/продаж НАШИХ активных товаров по бренду и категории
    brand_agg = defaultdict(lambda: {"revenue": 0.0, "units": 0, "skus": 0})
    cat_agg = defaultdict(lambda: {"revenue": 0.0, "units": 0, "skus": 0})

    for sku, offer_code in sku_to_offer.items():
        stat = analytics.get(sku)
        mrow = code_to_mikado.get(offer_code)
        brand = mrow["brand"] if mrow else None
        name_for_cat = (stat["name"] if stat else "") or (mrow["name"] if mrow else "")
        category = classify_category(name_for_cat) if name_for_cat else "Прочее"

        cat_agg[category]["skus"] += 1
        if brand:
            brand_agg[brand]["skus"] += 1

        if stat:
            rev = stat.get("revenue", 0) or 0
            units = stat.get("ordered_units", 0) or 0
            cat_agg[category]["revenue"] += rev
            cat_agg[category]["units"] += units
            if brand:
                brand_agg[brand]["revenue"] += rev
                brand_agg[brand]["units"] += units

    log.info("Активных SKU на Ozon: %d, брендов с историей: %d, категорий: %d",
              len(active_codes), len(brand_agg), len(cat_agg))

    # 4. Кандидаты из Микадо: закупка >= 300, остаток > 0, ещё не на Ozon
    candidates = []
    seen_codes = set()
    for r in mikado_rows:
        code_l = r["code"].lower()
        if code_l in seen_codes:
            continue
        if code_l in active_codes:
            continue
        if r["price"] < MIN_PURCHASE:
            continue
        if r["qty"] <= 0:
            continue
        if is_excluded_liquid(r["name"]):
            continue
        seen_codes.add(code_l)
        candidates.append(r)

    log.info("Кандидатов после фильтра (закупка>=%.0f, в наличии, не на Ozon): %d",
              MIN_PURCHASE, len(candidates))

    # 5. Скоринг
    public_demand = load_public_demand()
    fallback_avg_reviews = (
        sorted(v["avg_reviews"] for v in public_demand.values())[len(public_demand) // 2]
        if public_demand else 0.0
    )
    log.info("Категорий с публичными данными спроса: %d, fallback (медиана)=%.0f",
              len(public_demand), fallback_avg_reviews)
    scored = []
    for r in candidates:
        purchase = r["price"]
        sell = find_price_optimum(purchase, DEFAULT_LOGISTICS)
        if sell is None:
            continue
        margin_pct = calc_margin_pct(purchase, sell, DEFAULT_LOGISTICS)
        if margin_pct < MIN_MARGIN_PCT:
            continue
        profit_rub = calc_profit(purchase, sell, DEFAULT_LOGISTICS)

        category = classify_category(r["name"])
        cat_stat = cat_agg.get(category, {"revenue": 0, "units": 0, "skus": 0})
        brand_stat = brand_agg.get(r["brand"], {"revenue": 0, "units": 0, "skus": 0})

        pd = public_demand.get(category)
        consumable_val = match_kw_value(r["name"], PART_FREQ)
        fleet_val = match_kw_value(r["name"], CAR_FLEET)

        if pd:
            avg_reviews = pd["avg_reviews"]
        elif category == "Прочее":
            avg_reviews = fallback_avg_reviews * 0.4  # неклассифицированное - консервативная оценка
        else:
            avg_reviews = fallback_avg_reviews  # категория есть, но скан не прошёл - нейтральная медиана
        ext_demand_component = math.log1p(avg_reviews)
        internal_bonus = 1 + min(math.log1p(cat_stat["units"]) * 0.15 + math.log1p(brand_stat["units"]) * 0.15, 0.5)
        consumable_mult = 1 + min(consumable_val / 6.0, 1.0) * 0.5
        fleet_mult = 1 + min(fleet_val / 1200, 1.0) * 0.35
        margin_score = min(margin_pct / 25.0, 1.2)
        profit_component = math.log1p(max(profit_rub, 0))

        score = (ext_demand_component * internal_bonus * consumable_mult
                 * fleet_mult * margin_score * profit_component)

        reasons = []
        if pd:
            reasons.append(
                f"категория «{category}»: на открытой выдаче Ozon в среднем "
                f"{avg_reviews:.0f} отзывов на карточку (медиана цены {pd['median_price']:,.0f} руб, "
                f"рейтинг {pd['avg_rating']:.1f}) - прокси реального спроса на площадке"
            )
        elif category == "Прочее":
            reasons.append("товар не удалось отнести к конкретной категории - спрос оценён консервативно (40% от медианы по всем категориям)")
        else:
            reasons.append(f"категория «{category}» - нет прямых данных публичной выдачи Ozon, использована медианная оценка спроса по категориям")
        if cat_stat["units"] > 0:
            reasons.append(
                f"у нас на Ozon категория уже продаётся: {cat_stat['revenue']:,.0f} руб "
                f"за {ANALYTICS_DAYS} дн. ({cat_stat['units']:.0f} шт.)"
            )
        if brand_stat["units"] > 0:
            reasons.append(f"бренд уже продавался у нас: {brand_stat['units']:.0f} шт. за {ANALYTICS_DAYS} дн.")
        if consumable_val >= 1.5:
            reasons.append("расходник с коротким циклом замены (стабильный повторный спрос)")
        if fleet_val >= 250:
            reasons.append(f"подходит к модели из топа автопарка РФ (популярность-индекс {fleet_val:.0f} тыс. шт.)")
        reasons.append(f"маржа {margin_pct:.1f} процентов при цене продажи {sell:,.0f} руб, прибыль {profit_rub:,.0f} руб/шт.")

        scored.append({
            **r,
            "category": category,
            "sell": sell,
            "margin_pct": margin_pct,
            "profit_rub": profit_rub,
            "score": score,
            "reason": "; ".join(reasons),
            "cat_units": cat_stat["units"],
            "cat_revenue": cat_stat["revenue"],
            "brand_units": brand_stat["units"],
            "brand_revenue": brand_stat["revenue"],
            "avg_reviews": avg_reviews,
        })

    scored.sort(key=lambda x: x["score"], reverse=True)

    # 6. Отбор топ-500 с ограничением на бренд И категорию (разнообразие).
    # Без лимита на категорию отбор смещается в сторону категорий, где у
    # поставщика просто больше SKU (артефакт ассортимента Микадо), а не
    # туда, где выше реальный расчётный score на позицию.
    final = []
    brand_count = defaultdict(int)
    category_count = defaultdict(int)
    for r in scored:
        if brand_count[r["brand"]] >= MAX_PER_BRAND:
            continue
        if category_count[r["category"]] >= MAX_PER_CATEGORY:
            continue
        final.append(r)
        brand_count[r["brand"]] += 1
        category_count[r["category"]] += 1
        if len(final) >= TOP_N:
            break

    log.info("Отобрано финально: %d позиций", len(final))

    write_excel(final, scored, cat_agg, brand_agg, public_demand)


def write_excel(final: list[dict], all_scored: list[dict], cat_agg: dict, brand_agg: dict, public_demand: dict):
    wb = openpyxl.Workbook()

    # ── Лист 1: Топ-500 ──
    ws = wb.active
    ws.title = "Топ-500"
    headers = [
        "№", "Код Mikado", "Бренд", "Наименование", "Категория",
        "Закупка, ₽", "Остаток Mikado, шт.", "Рек. цена продажи, ₽",
        "Маржа, %", "Прибыль/шт., ₽", "Score", "Обоснование",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for i, r in enumerate(final, 1):
        ws.append([
            i, r["code"], r["brand"], r["name"], r["category"],
            round(r["price"], 2), r["qty"], r["sell"],
            round(r["margin_pct"], 1), round(r["profit_rub"], 0), round(r["score"], 2),
            r["reason"],
        ])

    widths = [5, 12, 16, 45, 20, 11, 14, 16, 9, 12, 8, 90]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # ── Лист 2: Методология ──
    ws2 = wb.create_sheet("Методология")
    total_purchase = sum(r["price"] for r in final)
    total_profit_per_cycle = sum(r["profit_rub"] for r in final)
    lines = [
        ("Отбор топ-500 товаров для загрузки на Ozon", ""),
        ("Дата формирования", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Источники данных", ""),
        ("1. Ozon Seller API", "/v3/product/list - текущий каталог (наши активные SKU); "
                                f"/v1/analytics/data - оборот и продажи за {ANALYTICS_DAYS} дн."),
        ("2. Прайс Микадо", f"свежая выгрузка через API (без логина), {MIKADO_PRICE_URL[:60]}..."),
        ("3. Открытая выдача Ozon", "ozon_public_demand_scan.py - бесплатный источник без платных API: "
                                     "поиск по ozon.ru для каждой товарной категории, среднее число отзывов "
                                     "(=прокси накопленных продаж) и медианная цена по выборке 12-16 карточек"),
        ("4. Ценообразование", "pricing_engine / price_recalc.py - наценка 30/25/22/20/17% "
                                "по ценовым тиерам закупки (политика Оптимум), логистика FBS 115 руб по умолчанию"),
        ("5. Универсальные сигналы", "PART_FREQ (частота замены детали), CAR_FLEET (популярность модели "
                                      "в автопарке РФ, данные Автостата) - перенесены из методологии wb_top500_combined.py"),
        ("", ""),
        ("Фильтры отбора", ""),
        ("Мин. закупка", f"{MIN_PURCHASE:.0f} руб"),
        ("Мин. маржа", f"{MIN_MARGIN_PCT:.0f}%"),
        ("Остаток на складе Микадо", "> 0"),
        ("Жидкости (масло, антифриз)", "исключены по требованию пользователя (моторное масло, антифриз/тосол)"),
        ("Уже загружен на Ozon", "исключается (сверка по офферу код-con)"),
        ("Макс. позиций на бренд", f"{MAX_PER_BRAND} (для разнообразия ассортимента)"),
        ("Макс. позиций на категорию", f"{MAX_PER_CATEGORY} (иначе отбор смещается в категории, где у Микадо просто больше SKU)"),
        ("", ""),
        ("Формула скоринга", "score = ext_demand x internal_bonus x consumable_mult x fleet_mult x margin_score x profit_component"),
        ("ext_demand (главный сигнал)", "log(1 + среднее число отзывов по категории на открытой выдаче Ozon) - "
                                         "объективный внешний спрос, не зависит от нашей истории продаж"),
        ("internal_bonus (мягкий, некритичный)", "1 + min(log(1+продажи_категории)x0.15 + log(1+продажи_бренда)x0.15; 0.5) "
                                                   "- бонус до +50% по факт. продажам у нас на Ozon за 90 дн., не доминирует"),
        ("consumable_mult", "буст до x1.5 для расходников короткого цикла (масл./возд. фильтры, свечи, колодки...)"),
        ("fleet_mult", "буст до x1.35 для деталей к самым массовым моделям в автопарке РФ"),
        ("margin_score", "маржа/25%, потолок x1.2"),
        ("profit_component", "log(1+прибыль руб/шт.) - абсолютная прибыль важнее при равной марже"),
        ("", ""),
        ("Итоги по отобранным 500 позициям", ""),
        ("Суммарная закупка (единоразово, по 1 шт. каждой позиции)", f"{total_purchase:,.0f} руб"),
        ("Суммарная прибыль с одного оборота (1 шт. каждой позиции)", f"{total_profit_per_cycle:,.0f} руб"),
        ("Кандидатов рассмотрено всего", f"{len(all_scored):,}"),
        ("", ""),
        ("Важное ограничение", "Платные источники (MPStats и т.п.) не используются. Спрос оценивается по "
                                "двум открытым сигналам: (1) публичная выдача Ozon по категории - число "
                                "отзывов как прокси продаж, без привязки к конкретному SKU/OEM; (2) наша "
                                "фактическая статистика продаж за 90 дней (используется как мягкая поправка, "
                                "т.к. у нас пока всего ~7 активных категорий и мало данных). Точность оценки "
                                "спроса ниже, чем при наличии полноценной SKU-аналитики по конкретному товару."),
    ]
    for row in lines:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 90
    for row in ws2.iter_rows(min_row=1, max_row=len(lines)):
        row[0].font = Font(bold=True)

    # ── Лист 3: Статистика по категориям (наша текущая продажи) ──
    ws3 = wb.create_sheet("Наши категории (90 дн)")
    ws3.append(["Категория", "Оборот, ₽ (90 дн)", "Продано, шт.", "Активных SKU"])
    for c in range(1, 5):
        ws3.cell(1, c).font = Font(bold=True)
    for cat, stat in sorted(cat_agg.items(), key=lambda x: -x[1]["revenue"]):
        ws3.append([cat, round(stat["revenue"], 0), stat["units"], stat["skus"]])
    ws3.column_dimensions["A"].width = 30

    # ── Лист 4: Статистика по брендам ──
    ws4 = wb.create_sheet("Наши бренды (90 дн)")
    ws4.append(["Бренд", "Оборот, ₽ (90 дн)", "Продано, шт.", "Активных SKU"])
    for c in range(1, 5):
        ws4.cell(1, c).font = Font(bold=True)
    for brand, stat in sorted(brand_agg.items(), key=lambda x: -x[1]["revenue"])[:100]:
        ws4.append([brand, round(stat["revenue"], 0), stat["units"], stat["skus"]])
    ws4.column_dimensions["A"].width = 25

    # ── Лист 5: Публичный спрос Ozon по категориям (открытый источник) ──
    ws5 = wb.create_sheet("Спрос Ozon (публичный)")
    ws5.append(["Категория", "Среднее число отзывов на карточку", "Медианная цена, руб", "Ср. рейтинг", "Выборка карточек"])
    for c in range(1, 6):
        ws5.cell(1, c).font = Font(bold=True)
    for cat, pd in sorted(public_demand.items(), key=lambda x: -x[1]["avg_reviews"]):
        ws5.append([cat, round(pd["avg_reviews"], 0), round(pd["median_price"], 0),
                    round(pd["avg_rating"], 2), pd["sample_size"]])
    ws5.column_dimensions["A"].width = 30

    out_path = CACHE_DIR / f"Топ-500_Ozon_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(out_path)
    log.info("Excel сохранён: %s", out_path)
    print(f"\nГОТОВО: {out_path}")


if __name__ == "__main__":
    main()
