"""
wb_top500_combined.py — ТОП-500 WB: Автолига + Микадо, 8 сигналов
===================================================================
Источники данных:
  1. Автолига прайс (PriceALVLG0411.xls, 26.05.2026, 41 143 арт.)
  2. Микадо прайс   (mikado_price_live.xlsx, актуальный, 37 643 арт.)
  3. MPStats кэш    (683 бренда, 52 695 авто-товаров WB)
  4. Drive2 / Drom  (топ-15 авто РФ × частота замены деталей)
  5. Автостат 2025  (объём автопарка по моделям)
  6. WB-аналитика   (рост категорий: масла +313%, подвеска +155%)

8 сигналов: S1 Дефицит, S2 Голубой океан, S3 Автопарк×поломка,
            S4 Расходник, S5 Сезон, S6 Уникальность, S7 Ниша, S8 Широкий
Дедупликация: совпадение (бренд_норм + артикул_норм) → оставляем дешевле.
"""

import xlrd, openpyxl, json, os, re, math
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Пути ──────────────────────────────────────────────────────────────────────
AUTOLIGA_FILE = "data/suppliers/autoliga/PriceALVLG0411.xls"
MIKADO_FILE   = "data/suppliers/mikado/mikado_price_live.xlsx"
CACHE_DIR     = "data/analytics/cache"
OUTPUT_FILE   = "data/analytics/wb_top500_combined.xlsx"

# ── WB FBS: Автозапчасти, май 2026 ────────────────────────────────────────────
WB_COMMISSION  = 0.25
WB_ACQUIRING   = 0.015
WB_RETURN_RATE = 0.03
TAX_RATE       = 0.06
WB_LOGISTICS   = 75
MIN_MARGIN_PCT = 15
TARGET_MARGIN  = 25
MAX_PER_BRAND  = 15

# ══════════════════════════════════════════════════════════════════════════════
#  ДАННЫЕ ИЗ ИССЛЕДОВАНИЙ
# ══════════════════════════════════════════════════════════════════════════════

CAR_FLEET = {
    "2107": 1177, "2106": 850, "2109": 900, "2110": 600, "2114": 400,
    "2115": 350, "priora": 800, "приора": 800,
    "granta": 1070, "гранта": 1070,
    "vesta": 750, "веста": 750,
    "niva": 1004, "нива": 1004, "4x4": 1004,
    "kalina": 650, "калина": 650,
    "largus": 350, "ларгус": 350,
    "rio": 1020, "рио": 1020,
    "sportage": 380, "спортейдж": 380,
    "cerato": 280, "серато": 280,
    "ceed": 200, "soul": 150,
    "solaris": 926, "солярис": 926,
    "creta": 420, "крета": 420,
    "tucson": 310, "туксон": 310,
    "elantra": 250, "элантра": 250,
    "ix35": 290, "santa fe": 200,
    "logan": 774, "логан": 774,
    "duster": 320, "дастер": 320,
    "sandero": 280, "сандеро": 280,
    "megane": 180, "лагуна": 150,
    "focus": 742, "фокус": 742,
    "mondeo": 220, "мондео": 220,
    "polo": 500, "поло": 500,
    "passat": 280, "пассат": 280,
    "golf": 220, "tiguan": 180,
    "camry": 350, "камри": 350,
    "rav4": 280, "corolla": 260, "королла": 260,
    "land cruiser": 220,
    "almera": 380, "альмера": 380,
    "x-trail": 270, "qashqai": 200, "tiida": 150,
    "rapid": 350, "рапид": 350,
    "octavia": 280, "октавия": 280,
    "fabia": 180,
    "outlander": 280, "аутлендер": 280,
    "lancer": 250, "лансер": 250,
    "asx": 180,
    "lacetti": 380, "лачетти": 380,
    "cruze": 250, "круз": 250,
    "aveo": 200, "авео": 200,
    "e34": 150, "e36": 140, "e46": 250, "e39": 180, "e60": 220,
    "w210": 130, "w211": 180, "w203": 160, "w204": 200,
    "peugeot": 300, "citroen": 250, "308": 120, "407": 100,
    "haval": 350, "jolion": 200, "h6": 150,
    "geely": 280, "atlas": 180, "coolray": 120,
    "chery": 220, "tiggo": 180,
    "jac": 80, "changan": 100, "belgee": 50,
}

PART_FREQ = {
    "фильтр масл":      6.0,
    "масляный фильтр":  6.0,
    "фильтр воздуш":    3.5,
    "воздушный фильтр": 3.5,
    "фильтр салон":     3.5,
    "салонный фильтр":  3.5,
    "свеч":             2.5,
    "spark plug":       2.5,
    "масло мотор":      6.0,
    "моторное масло":   6.0,
    "антифриз":         1.5,
    "охлаждающ жидк":   1.5,
    "жидкость тормозн": 1.5,
    "тормозн колодк":   2.5,
    "колодки тормозн":  2.5,
    "brake pad":        2.5,
    "тормозной диск":   1.5,
    "амортизатор":      1.5,
    "стойк стабил":     2.0,
    "стабилизатор":     1.5,
    "сайлентблок":      2.0,
    "шаровая":          1.5,
    "ступичный подшипн":1.5,
    "подшипник ступиц": 1.5,
    "рычаг подвески":   1.2,
    "шрус":             1.2,
    "ремень грм":       1.0,
    "ремень привода":   1.5,
    "ремень генератор": 1.5,
    "натяжитель":       1.0,
    "цепь грм":         0.8,
    "рулевая рейка":    0.8,
    "рулевой наконечн": 1.2,
    "помпа":            1.0,
    "водяной насос":    1.0,
    "радиатор":         0.8,
    "катушка зажиган":  1.2,
    "генератор":        0.6,
    "стартер":          0.5,
    "дворники":         2.0,
    "щетки стеклооч":   2.0,
    "прокладка гбц":    0.5,
    "форсунк":          1.0,
    "термостат":        0.8,
    "аккумулятор":      0.7,
    "лампа":            2.0,
    "лампочка":         2.0,
}

SUMMER_KW = [
    "антифриз", "охлаждающ", "coolant", "радиатор охлажд",
    "масло мотор", "моторное масло", "motor oil",
    "кондиционер", "компрессор кондиц",
    "тормозн", "brake",
    "амортизатор", "подвеск", "shock",
    "шин", "диск колесн",
    "аккумулятор", "акб", "battery",
    "дворник", "щетк стеклооч", "wiper",
    "фильтр воздуш", "air filter",
    "омыватель",
]

WIDE_APPLY_KW = [
    "универсальн", "подходит для всех", "все авто",
    "масло", "антифриз", "тормозная жидк", "омыватель",
    "w", "5w", "10w", "15w",
]

AUTO_KEYS = [
    "авто", "запчаст", "шин", "диск", "тормоз", "подвеск", "двигател",
    "кузов", "электрик", "трансмисс", "масл", "фильтр", "сцеплени",
    "рулев", "охлажден", "технич", "мотоцикл", "выхлоп", "топлив",
    "аккумулят", "зажиган", "амортизатор",
]

WB_TOP_NICHES = {
    "масло моторное":       {"revenue_est": 800, "growth": 3.13, "sellers": 450},
    "тормозные колодки":    {"revenue_est": 650, "growth": 1.80, "sellers": 380},
    "фильтр масляный":      {"revenue_est": 500, "growth": 1.70, "sellers": 320},
    "амортизатор":          {"revenue_est": 480, "growth": 1.55, "sellers": 290},
    "фильтр воздушный":     {"revenue_est": 350, "growth": 1.65, "sellers": 280},
    "шины":                 {"revenue_est": 1200, "growth": 2.74, "sellers": 600},
    "аккумулятор":          {"revenue_est": 420, "growth": 1.40, "sellers": 200},
    "свечи зажигания":      {"revenue_est": 380, "growth": 1.50, "sellers": 350},
    "фильтр салонный":      {"revenue_est": 280, "growth": 1.55, "sellers": 310},
    "сайлентблок":          {"revenue_est": 320, "growth": 1.45, "sellers": 260},
    "ремень грм":           {"revenue_est": 250, "growth": 1.35, "sellers": 180},
    "дворники":             {"revenue_est": 230, "growth": 1.60, "sellers": 250},
    "антифриз":             {"revenue_est": 300, "growth": 1.80, "sellers": 200},
    "стойки стабилизатора": {"revenue_est": 260, "growth": 1.45, "sellers": 220},
    "шаровая опора":        {"revenue_est": 210, "growth": 1.40, "sellers": 200},
    "катушка зажигания":    {"revenue_est": 190, "growth": 1.50, "sellers": 170},
    "ступичный подшипник":  {"revenue_est": 220, "growth": 1.45, "sellers": 190},
    "рулевой наконечник":   {"revenue_est": 180, "growth": 1.35, "sellers": 160},
    "помпа":                {"revenue_est": 160, "growth": 1.30, "sellers": 140},
    "тормозной диск":       {"revenue_est": 290, "growth": 1.50, "sellers": 230},
}


# ══════════════════════════════════════════════════════════════════════════════
#  УТИЛИТЫ
# ══════════════════════════════════════════════════════════════════════════════

def norm(s):
    return re.sub(r'[\s\-\.\/\\]', '', str(s)).upper()

def is_auto(subj):
    return any(k in subj.lower() for k in AUTO_KEYS)

def calc_margin(buy, sell):
    fees = sell * (WB_COMMISSION + WB_ACQUIRING + WB_RETURN_RATE)
    tax  = sell * TAX_RATE
    net  = sell - fees - WB_LOGISTICS - tax - buy
    pct  = (net / sell * 100) if sell > 0 else 0
    return round(net, 2), round(pct, 1)

def min_price(buy, target):
    coeff = (1 - WB_COMMISSION - WB_ACQUIRING - WB_RETURN_RATE - TAX_RATE) - target / 100
    return (WB_LOGISTICS + buy) / coeff if coeff > 0 else 1e9

def get_fleet_score(name):
    t = name.lower()
    total, matched = 0, []
    for model, fleet in CAR_FLEET.items():
        if model in t:
            total += fleet
            matched.append(model)
    return min(total / 500, 3.0), matched

def get_repair_freq(name):
    t = name.lower()
    best, matched_part = 0.0, ""
    for kw, freq in PART_FREQ.items():
        if kw in t and freq > best:
            best = freq
            matched_part = kw
    return min(best / 3.0, 2.0), matched_part

def get_niche_score(name):
    t = name.lower()
    best_score, best_niche = 0.0, ""
    for niche, data in WB_TOP_NICHES.items():
        if any(kw in t for kw in niche.split()):
            score = math.log1p(data["revenue_est"]) * data["growth"] / math.sqrt(data["sellers"])
            if score > best_score:
                best_score = score
                best_niche = niche
    return min(best_score / 5, 2.0), best_niche

def is_summer(name):   return any(k in name.lower() for k in SUMMER_KW)
def is_wide_apply(name): return any(k in name.lower() for k in WIDE_APPLY_KW)

# Чёрный список конкретных артикулов (norm_brand, norm_article)
BLACKLIST = {
    ('MATRIX',     '74044'),           # круг лепестковый — не запчасть
    ('MITSUBISHI', 'MR571476'),        # OEM фильтр Pajero III — слишком дорого
    ('MITSUBISHI', '1500A608'),        # OEM фильтр Pajero Sport — слишком дорого
    ('VAG',        '4M0133843C'),      # OEM фильтр Touareg — слишком дорого
    ('VAG',        '4M0133843G'),      # OEM фильтр Touareg — слишком дорого
    ('BRAVE',      'BRRP423'),         # втулка стаб. 86 руб — слишком дёшево
    ('ТОЛЬЯТТИ',   '21050370311200'),  # планка АКБ 30 руб — слишком дёшево
}


def is_excluded(item: dict) -> bool:
    """Исключает: масло, лампочки, антифриз, закупка < 150₽, чёрный список."""
    n   = item['name'].lower()
    buy = item['buy']

    # Чёрный список конкретных позиций
    if (norm(item['brand']), norm(item['article'])) in BLACKLIST:
        return True
    # Цена закупки ниже порога
    if buy < 150:
        return True
    # Масло любого типа — масляный фильтр оставляем
    if 'масло' in n and 'фильтр' not in n:
        return True
    # Вязкость масла (5w-30, 0w-20 и т.д.)
    if re.search(r'\d+w[-\s]\d+', n):
        return True
    # Антифриз / охлаждающая жидкость
    if 'антифриз' in n or 'coolant' in n:
        return True
    # Лампочки и лампы
    if 'лампа' in n or 'лампочка' in n:
        return True
    return False


# ══════════════════════════════════════════════════════════════════════════════
#  ЗАГРУЗКА ПРАЙСОВ
# ══════════════════════════════════════════════════════════════════════════════

def load_autoliga() -> list:
    book = xlrd.open_workbook(AUTOLIGA_FILE, encoding_override='cp1251')
    ws   = book.sheet_by_index(0)
    items = []
    for r in range(9, ws.nrows):
        try:
            brand   = str(ws.cell_value(r, 1)).strip()
            article = str(ws.cell_value(r, 2)).strip()
            name    = str(ws.cell_value(r, 4)).strip()
            stock   = float(str(ws.cell_value(r, 6)).replace(',', '.') or 0)
            price   = float(str(ws.cell_value(r, 7)).replace(',', '.') or 0)
        except Exception:
            continue
        if not brand or not article or stock <= 0 or price <= 0:
            continue
        items.append({
            "brand":   brand,
            "article": article,
            "name":    name,
            "stock":   int(stock),
            "buy":     price,
            "source":  "Автолига",
        })
    return items


def _parse_mikado_qty(raw) -> int:
    if raw is None:
        return 0
    s = str(raw).strip()
    if s.startswith('>'):
        return 10
    try:
        return int(float(s))
    except Exception:
        return 0


def load_mikado() -> list:
    wb = openpyxl.load_workbook(MIKADO_FILE, read_only=True, data_only=True)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            article = str(row[1]).strip() if row[1] else ""
            brand   = str(row[2]).strip() if row[2] else ""
            name    = str(row[3]).strip() if row[3] else ""
            price   = float(str(row[4]).replace(',', '.')) if row[4] else 0.0
            stock   = _parse_mikado_qty(row[5])
        except Exception:
            continue
        if not brand or not article or stock <= 0 or price <= 0:
            continue
        items.append({
            "brand":   brand,
            "article": article,
            "name":    name,
            "stock":   stock,
            "buy":     price,
            "source":  "Микадо",
        })
    wb.close()
    return items


def merge_suppliers(al_items: list, mk_items: list) -> list:
    """
    Дедупликация: если brand_norm + article_norm совпадают у обоих поставщиков,
    оставляем позицию с наименьшей ценой закупки, источник = 'Авт+Мик'.
    """
    # Индекс по ключу (brand_norm, article_norm)
    index = {}
    for it in al_items:
        key = (norm(it['brand']), norm(it['article']))
        if key not in index or it['buy'] < index[key]['buy']:
            index[key] = it.copy()

    # Добавляем Микадо: если ключ уже есть → сравниваем цену
    for it in mk_items:
        key = (norm(it['brand']), norm(it['article']))
        if key in index:
            existing = index[key]
            if it['buy'] < existing['buy']:
                merged = it.copy()
                merged['source'] = 'Авт+Мик'
                index[key] = merged
            else:
                index[key]['source'] = 'Авт+Мик'
        else:
            index[key] = it.copy()

    return list(index.values())


# ══════════════════════════════════════════════════════════════════════════════
#  MPSTATS: агрегат по брендам
# ══════════════════════════════════════════════════════════════════════════════

def build_brand_stats() -> dict:
    brand_data = defaultdict(list)
    for fn in os.listdir(CACHE_DIR):
        if not fn.startswith('mpsb_'):
            continue
        with open(os.path.join(CACHE_DIR, fn), encoding='utf-8') as f:
            try: data = json.load(f)
            except: continue
        for p in data:
            if not is_auto(p.get('subject', '')):
                continue
            b = p.get('brand', '').upper().strip()
            if b:
                brand_data[b].append({
                    'sales': float(p.get('sales_30d', 0) or 0),
                    'oos':   float(p.get('oos_pct', 0) or 0),
                    'price': float(p.get('price', 0) or 0),
                })
    stats = {}
    for b, prods in brand_data.items():
        sales = [p['sales'] for p in prods]
        oos   = [p['oos']   for p in prods]
        price = [p['price'] for p in prods if p['price'] > 0]
        stats[b] = {
            'avg_sales':     sum(sales) / len(sales) if sales else 0,
            'max_sales':     max(sales) if sales else 0,
            'avg_oos':       sum(oos) / len(oos) if oos else 0,
            'max_oos':       max(oos) if oos else 0,
            'wb_price_avg':  sum(price) / len(price) if price else 0,
            'wb_price_min':  min(price) if price else 0,
            'product_count': len(prods),
        }
    return stats


# ══════════════════════════════════════════════════════════════════════════════
#  СКОРИНГ
# ══════════════════════════════════════════════════════════════════════════════

def score_item(item: dict, brand_stats: dict) -> dict:
    buy      = item['buy']
    name     = item['name']
    name_lo  = name.lower()
    brand_up = item['brand'].upper()
    bs       = brand_stats.get(brand_up, {})

    wb_avg  = bs.get('wb_price_avg', 0)
    p_min15 = min_price(buy, MIN_MARGIN_PCT)
    p_rec25 = min_price(buy, TARGET_MARGIN)

    if wb_avg >= p_min15 and wb_avg > 0:
        sell = min(wb_avg * 0.95, p_rec25)
        sell = max(sell, p_min15)
    else:
        sell = p_rec25

    mg_rub, mg_pct = calc_margin(buy, sell)
    if mg_pct < MIN_MARGIN_PCT:
        return {}

    cnt = bs.get('product_count', 0)
    oos = bs.get('avg_oos', 0)

    s1 = 1.0 + (oos / 100) * 2.0

    if   cnt == 0: s2 = 1.8
    elif cnt < 10: s2 = 1.4
    elif cnt < 30: s2 = 1.2
    else:          s2 = max(1.0 / math.sqrt(cnt / 10), 0.5)

    fleet_score, fleet_cars   = get_fleet_score(name_lo)
    repair_score, repair_part = get_repair_freq(name_lo)
    s3 = 1.0 + fleet_score * 0.3 + repair_score * 0.3

    freq_raw, _ = get_repair_freq(name_lo)
    s4 = 1.0 + min(freq_raw / 2.0, 0.5)

    s5 = 1.25 if is_summer(name_lo) else 1.0
    s6 = 1.5 if cnt == 0 else (1.2 if cnt < 5 else 1.0)

    niche_score, niche_name = get_niche_score(name_lo)
    s7 = 1.0 + niche_score * 0.4

    s8 = 1.15 if is_wide_apply(name_lo) else 1.0

    avg_sales = bs.get('avg_sales', 0)
    demand    = math.log1p(avg_sales) if avg_sales > 0 else 0.3
    m_score   = min(mg_pct / 30.0, 1.0)

    score = demand * s1 * s2 * m_score * s3 * s4 * s5 * s6 * s7 * s8

    strats = []
    if oos >= 25:          strats.append(f"Дефицит OOS {oos:.0f}%")
    if s2 > 1.1:           strats.append(f"Голубой океан ({cnt} WB)")
    if fleet_score > 0.5:  strats.append(f"Топ авто РФ: {', '.join(fleet_cars[:2])}")
    if repair_score > 0.5: strats.append(f"Частая замена: {repair_part}")
    if s4 > 1.1:           strats.append("Расходник")
    if s5 > 1.0:           strats.append("Сезон лето")
    if s6 > 1.0:           strats.append("Слабая конк. WB")
    if niche_score > 0.5:  strats.append(f"Топ-ниша WB: {niche_name}")
    if s8 > 1.0:           strats.append("Широкая применяемость")

    parts = []
    if avg_sales > 0: parts.append(f"WB продажи бренда: {avg_sales:.0f} шт/мес")
    if oos > 0:       parts.append(f"OOS бренда: {oos:.0f}%")
    if wb_avg > 0:    parts.append(f"Ср.цена WB: {wb_avg:.0f} руб")
    if fleet_cars:    parts.append(f"Авто в парке: {'+'.join(fleet_cars[:3])}")
    if repair_part:   parts.append(f"Деталь меняют: {repair_part} (x{freq_raw:.1f}/100к км)")
    if niche_name:    parts.append(f"WB-ниша: {niche_name}")
    parts.append(f"Маржа: {mg_pct:.1f}% / {mg_rub:.0f} руб")
    parts.append(f"Источник: {item['source']}")

    return {
        'score':       round(score, 5),
        'sell':        round(sell),
        'mg_pct':      mg_pct,
        'mg_rub':      mg_rub,
        'avg_sales':   round(avg_sales, 1),
        'oos':         round(oos, 1),
        'wb_cnt':      cnt,
        'fleet_cars':  ', '.join(fleet_cars[:3]),
        'repair_part': repair_part,
        'niche':       niche_name,
        'strats':      ', '.join(strats) if strats else 'Маржа',
        'reason':      '; '.join(parts),
        's1': round(s1,2), 's2': round(s2,2), 's3': round(s3,2),
        's4': round(s4,2), 's5': round(s5,2), 's6': round(s6,2),
        's7': round(s7,2), 's8': round(s8,2),
    }


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ЭКСПОРТ
# ══════════════════════════════════════════════════════════════════════════════

C_HEADER = "1F4E79"
C_GOLD   = "FFF2CC"
C_SILV   = "EBF3FB"
C_BRNZ   = "FCE4D6"
C_GREEN  = "E2EFDA"
C_MK     = "E8F5E9"  # светло-зелёный — Микадо

THIN = Border(
    left  =Side(style="thin", color="D0D0D0"),
    right =Side(style="thin", color="D0D0D0"),
    top   =Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

MAIN_COLS = [
    ("#",                       5),
    ("Артикул",                27),
    ("Бренд",                  18),
    ("Название",               45),
    ("Источник",               12),
    ("Закупка, руб",           13),
    ("Продажа WB, руб",        16),
    ("Маржа, %",               10),
    ("Маржа, руб",             11),
    ("WB продажи шт/мес",      18),
    ("OOS %",                  10),
    ("Остаток",                10),
    ("WB товаров бренда",      18),
    ("Авто в парке",           20),
    ("Деталь - частота",       22),
    ("WB-ниша",                22),
    ("Стратегии",              40),
    ("Подробное обоснование",  90),
]


def _cell(ws, r, c, v, bold=False, bg=None, wrap=False, align="left", fmt=None, color=None):
    cell = ws.cell(row=r, column=c, value=v)
    fc   = "FFFFFF" if bg == C_HEADER else (color or "000000")
    cell.font      = Font(bold=bold, size=10, color=fc)
    cell.border    = THIN
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt
    return cell


def export(results: list, al_cnt: int, mk_cnt: int, both_cnt: int) -> str:
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb = Workbook()

    # ── Лист 1: Топ-500 ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Top-500 (без масла и ламп)"
    ws.freeze_panes = "B2"

    for col, (h, w) in enumerate(MAIN_COLS, 1):
        _cell(ws, 1, col, h, bold=True, bg=C_HEADER, align="center", wrap=True)
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w
    ws.row_dimensions[1].height = 45

    FMTS = [None, None, None, None, None, '#,##0', '#,##0', '0.0', '#,##0',
            '#,##0.0', '0.0', '#,##0', '#,##0', None, None, None, None, None]

    for i, it in enumerate(results[:500], 2):
        rank = i - 1
        src  = it.get('source', '')
        if rank <= 50:
            bg = C_GOLD
        elif rank <= 150:
            bg = C_SILV
        elif rank <= 300:
            bg = C_BRNZ
        else:
            bg = C_MK if src == 'Микадо' else None

        row = [
            rank, it['article'], it['brand'], it['name'],
            src,
            it['buy'], it.get('sell', ''), it.get('mg_pct', ''), it.get('mg_rub', ''),
            it.get('avg_sales', ''), it.get('oos', ''), it['stock'], it.get('wb_cnt', ''),
            it.get('fleet_cars', ''), it.get('repair_part', ''), it.get('niche', ''),
            it.get('strats', ''), it.get('reason', ''),
        ]
        for col, (val, fmt) in enumerate(zip(row, FMTS), 1):
            _cell(ws, i, col, val, bg=bg, wrap=(col == len(MAIN_COLS)), fmt=fmt)
        ws.row_dimensions[i].height = 32

    # ── Лист 2: Сигналы (детали скора) ───────────────────────────────────────
    ws2 = wb.create_sheet("Сигналы топ-50")
    sig_cols = ["#", "Артикул", "Бренд", "Источник", "Score", "S1 Дефицит", "S2 Океан",
                "S3 Автопарк", "S4 Расходник", "S5 Лето", "S6 Уникальн",
                "S7 Ниша", "S8 Широкий", "Маржа %"]
    for c, h in enumerate(sig_cols, 1):
        _cell(ws2, 1, c, h, bold=True, bg=C_HEADER, align="center")
        ws2.column_dimensions[ws2.cell(1, c).column_letter].width = 13
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 18

    for i, it in enumerate(results[:50], 2):
        row2 = [i-1, it['article'], it['brand'], it.get('source', ''), it['score'],
                it.get('s1'), it.get('s2'), it.get('s3'), it.get('s4'),
                it.get('s5'), it.get('s6'), it.get('s7'), it.get('s8'),
                it.get('mg_pct')]
        for c, val in enumerate(row2, 1):
            _cell(ws2, i, c, val, align="center", fmt=('0.00' if c >= 5 else None))

    # ── Лист 3: По источникам ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("По источникам")
    src_cnt = defaultdict(int)
    for it in results[:500]:
        src_cnt[it.get('source', '?')] += 1
    ws3.append(["Источник", "Позиций в топ-500", "% от топ-500"])
    for src, n in sorted(src_cnt.items(), key=lambda x: -x[1]):
        ws3.append([src, n, round(n / 5, 1)])
    ws3.append([])
    ws3.append(["Всего Автолига:", al_cnt, ""])
    ws3.append(["Всего Микадо:", mk_cnt, ""])
    ws3.append(["Дедуплицировано (совпад. бренд+арт):", both_cnt, ""])
    for col in ['A', 'B', 'C']:
        ws3.column_dimensions[col].width = 35

    # ── Лист 4: Стратегии ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Стратегии")
    cnt4 = defaultdict(int)
    for it in results[:500]:
        for s in it.get('strats', '').split(', '):
            if s.strip(): cnt4[s.strip()] += 1
    ws4.append(["Стратегия", "Товаров в топ-500", "% от топ-500"])
    for s, n in sorted(cnt4.items(), key=lambda x: -x[1]):
        ws4.append([s, n, round(n / 5, 1)])
    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 15

    # ── Лист 5: WB Топ-ниши ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("WB Топ-ниши")
    ws5.append(["Ниша", "Выручка WB млн/мес*", "Рост 2024", "Продавцов", "В нашем топ-500"])
    niche_cnt = defaultdict(int)
    for it in results[:500]:
        if it.get('niche'): niche_cnt[it['niche']] += 1
    for niche, data in sorted(WB_TOP_NICHES.items(), key=lambda x: -x[1]['revenue_est']):
        ws5.append([niche, data['revenue_est'], f"x{data['growth']:.2f}",
                    data['sellers'], niche_cnt.get(niche, 0)])
    ws5.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E']: ws5.column_dimensions[col].width = 18
    ws5.append([])
    ws5.append(["* Оценочные данные на основе открытых отчётов MPStats/ShopStat 2025"])

    # ── Лист 6: Топ брендов ───────────────────────────────────────────────────
    ws6 = wb.create_sheet("Бренды")
    bd = defaultdict(lambda: {'cnt': 0, 'sales': 0, 'oos': 0, 'mg': [], 'src': set()})
    for it in results[:500]:
        b = it['brand']
        bd[b]['cnt']   += 1
        bd[b]['sales'] += it.get('avg_sales', 0)
        bd[b]['oos']   += it.get('oos', 0)
        bd[b]['mg'].append(it.get('mg_pct', 0))
        bd[b]['src'].add(it.get('source', ''))
    ws6.append(["Бренд", "В топ-500", "Источник", "WB продажи шт/мес", "OOS %", "Ср.маржа %"])
    for b, d in sorted(bd.items(), key=lambda x: -x[1]['cnt']):
        avg_mg  = sum(d['mg']) / len(d['mg']) if d['mg'] else 0
        avg_oos = d['oos'] / d['cnt'] if d['cnt'] else 0
        ws6.append([b, d['cnt'], '+'.join(sorted(d['src'])),
                    round(d['sales']), round(avg_oos, 1), round(avg_mg, 1)])
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws6.column_dimensions[col].width = 20

    # ── Лист 7: Методология ───────────────────────────────────────────────────
    ws7 = wb.create_sheet("Методология")
    notes = [
        ("ИСТОЧНИКИ ДАННЫХ", ""),
        ("Автолига прайс",   "PriceALVLG0411.xls, 26.05.2026, 41 143 арт. со стоком"),
        ("Микадо прайс",     "mikado_price_live.xlsx, актуальный (31.05.2026), 37 643 арт."),
        ("Дедупликация",     "brand_norm + article_norm: при совпадении берём дешевле"),
        ("Исключения",       "Масло (любое, кроме масляных фильтров) и лампочки/лампы"),
        ("MPStats кеш",      "683 бренда, 52 695 авто-товаров WB (апр-май 2026)"),
        ("Drive2 / Drom",    "Топ-15 авто РФ: частота замены 30+ видов деталей"),
        ("Автостат 2025",    "Объём автопарка 30 моделей (тыс. шт.)"),
        ("WB-аналитика",     "Рост ниш 2024: масла +313%, шины +274%, подвеска +155%"),
        ("", ""),
        ("ФОРМУЛА МАРЖИ WB FBS", ""),
        ("Комиссия WB",  "25% (Автозапчасти FBS, май 2026)"),
        ("Эквайринг",    "1.5%"),
        ("Возвраты",     "3%"),
        ("УСН",          "6% от выручки"),
        ("Логистика",    "75 руб. FBS Волгоград (ок. 0.3 кг)"),
        ("Мин. маржа",   ">= 15% (иначе товар отброшен)"),
        ("Цель. маржа",  "25% - рекомендованная цена продажи"),
        ("", ""),
        ("ФОРМУЛА СКОРА (8 СИГНАЛОВ)", ""),
        ("S1 Дефицит",       "1 + avg_oos/100 x 2  (1.0-3.0)"),
        ("S2 Голубой океан", "0 WB-прод.->1.8 / <10->1.4 / <30->1.2 / >30->1/sqrt(cnt/10)"),
        ("S3 Автопарк",      "1 + fleet_score x 0.3 + repair_score x 0.3  (1.0-2.8)"),
        ("S4 Расходник",     "1 + min(freq/2, 0.5)  (1.0-1.5)"),
        ("S5 Сезон лето",    "x1.25 для летних категорий"),
        ("S6 Уникальность",  "x1.5 нет на WB; x1.2 мало (<5 WB-прод.)"),
        ("S7 WB-ниша",       "1 + niche_score x 0.4  (log выручка x рост / sqrt(продавцов))"),
        ("S8 Широкий",       "x1.15 для универсальных товаров"),
        ("Score итог",       "demand x S1 x S2 x margin_score x S3 x S4 x S5 x S6 x S7 x S8"),
        ("", ""),
        ("ВАЖНО", "Матчинг по артикулу на уровне бренда (MPStats кеш)."),
        ("", "Дедупликация между поставщиками по brand_norm + article_norm."),
        ("", "Для точного OEM-матчинга: wb_product_indexer.py --method oem через VPN."),
    ]
    ws7.column_dimensions['A'].width = 30
    ws7.column_dimensions['B'].width = 70
    for r_data in notes:
        row_idx = ws7.max_row + 1
        ws7.append(list(r_data))
        if r_data[1] == "":
            ws7.cell(row_idx, 1).font = Font(bold=True, size=11)

    wb.save(OUTPUT_FILE)
    return os.path.abspath(OUTPUT_FILE)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print("  WB Top-500 COMBINED  |  Avtoliga + Mikado  |  8 signals")
    print("=" * 70)

    print("\n[1/5] Loading Avtoliga price list...")
    al_items = load_autoliga()
    print(f"  Avtoliga: {len(al_items):,} items, {len(set(i['brand'] for i in al_items))} brands")

    print("\n[2/5] Loading Mikado price list...")
    mk_items = load_mikado()
    print(f"  Mikado:   {len(mk_items):,} items, {len(set(i['brand'] for i in mk_items))} brands")

    print("\n[3/5] Merging and deduplicating...")
    all_items = merge_suppliers(al_items, mk_items)
    both_cnt = sum(1 for it in all_items if it['source'] == 'Авт+Мик')
    al_only  = sum(1 for it in all_items if it['source'] == 'Автолига')
    mk_only  = sum(1 for it in all_items if it['source'] == 'Микадо')
    print(f"  Total pool: {len(all_items):,}  |  Autoliga only: {al_only}  |  Mikado only: {mk_only}  |  Both: {both_cnt}")

    # Исключаем по всем правилам
    before = len(all_items)
    all_items = [it for it in all_items if not is_excluded(it)]
    print(f"  Excluded (maslo/antifriz/lampy/<150/blacklist): {before - len(all_items):,}  |  After: {len(all_items):,}")

    print("\n[4/5] Building brand stats from MPStats cache...")
    bstats  = build_brand_stats()
    matched = sum(1 for i in all_items if i['brand'].upper() in bstats)
    print(f"  Brands in cache: {len(bstats):,}  |  Matched items: {matched:,}/{len(all_items):,}")

    print("\n[5/5] Scoring (8 signals) + ranking...")
    scored, skipped = [], 0
    for it in all_items:
        s = score_item(it, bstats)
        if not s:
            skipped += 1
            continue
        it.update(s)
        scored.append(it)
    scored.sort(key=lambda x: x['score'], reverse=True)

    brand_cnt = defaultdict(int)
    final = []
    for it in scored:
        b = it['brand']
        if brand_cnt[b] < MAX_PER_BRAND:
            brand_cnt[b] += 1
            final.append(it)

    print(f"  Scored: {len(scored):,}  |  Skipped: {skipped}  |  After cap ({MAX_PER_BRAND}/brand): {len(final):,}")

    print("\n  Exporting Excel...")
    path = export(final, len(al_items), len(mk_items), both_cnt)
    sz   = os.path.getsize(OUTPUT_FILE) // 1024
    print(f"  Saved: {path}  ({sz} KB)")

    # Preview top-20
    print("\n" + "=" * 120)
    print(f"{'#':>3}  {'Brand':<18} {'Article':<22} {'Src':<8} {'Buy':>7} {'Sell':>7} "
          f"{'Mg':>5} {'Sls':>5} {'OOS':>4}  {'Score':>7}")
    print("-" * 120)
    for i, it in enumerate(final[:20], 1):
        try:
            print(f"{i:>3}  {it['brand']:<18} {it['article']:<22} "
                  f"{it.get('source',''):<8} "
                  f"{it['buy']:>7.0f} {it.get('sell',0):>7.0f} "
                  f"{it.get('mg_pct',0):>4.0f}% "
                  f"{it.get('avg_sales',0):>5.0f} {it.get('oos',0):>3.0f}%  "
                  f"{it['score']:>7.4f}")
        except UnicodeEncodeError:
            print(f"{i:>3}  [encoding error]")

    # Source breakdown
    src_top = defaultdict(int)
    for it in final[:500]:
        src_top[it.get('source', '?')] += 1
    print(f"\n  Top-500 by source: ", end="")
    for src, n in sorted(src_top.items(), key=lambda x: -x[1]):
        try:
            print(f"{src}={n}", end="  ")
        except UnicodeEncodeError:
            print(f"[?]={n}", end="  ")
    print(f"\n  Unique brands in top-500: {len(brand_cnt)}")
    print(f"  Excel: {path}")


if __name__ == '__main__':
    main()
