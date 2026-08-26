# /// script
# requires-python = ">=3.10"
# dependencies = ["requests", "openpyxl"]
# ///
"""
fill_applicability_mikado.py
══════════════════════════════
Заполняет столбец «Применяемость» в "500 обогатитель.xlsx" данными с
mikado-parts.ru — тем же запросом, что открывается по клику на "Применимость"
в карточке товара (pp0.asp?MODE=APL), см. mikado_scraper.py.

Наш "Артикул" в файле — bare-код без префикса поставщика (например "of-4434g"),
а карточка на сайте открывается по полному коду Mikado (Prodnum, с префиксом,
например "xzk-of-4434g"). Соответствие Code -> Prodnum берётся из свежего
прайса Mikado (тот же API, что использует ozon_stock_sync.py).

Перед записью применяемости сверяется бренд и общий контекст названия
(word-overlap) между нашей строкой и карточкой Mikado — если совпадение
слабое, строка НЕ заполняется и попадает в лог несовпадений для ручной
проверки (защита от ошибок сопоставления).

Уже заполненные строки пропускаются — можно продолжать после прерывания.

Тест:
  uv run --with requests,openpyxl scripts/fill_applicability_mikado.py --rows 2-6

Обычный запуск:
  uv run --with requests,openpyxl scripts/fill_applicability_mikado.py --delay 0.5
"""

import re
import sys
import time
import argparse
import io
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).parent))

import requests
import openpyxl
from openpyxl.styles import Alignment

from mikado_scraper import login, fetch_product, fetch_compatibility, compat_to_text, load_env

XLSX_PATH = Path(r"C:\Users\Admin\Desktop\Озон-500\500 обогатитель.xlsx")
MIKADO_PRICE_URL = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=YOUR_MIKADO_PRICE_KEY"
)
PRICE_FALLBACK = Path(r"C:\Users\Admin\Documents\Ecommerce\mikado_price_34.xlsx")
MISMATCH_LOG = Path(r"C:\Users\Admin\Desktop\Озон-500\применяемость_несовпадения.txt")
SAVE_EVERY = 5

_STOPWORDS = {"и", "др", "с", "для", "на", "не", "или", "шт"}


def _tokens(text: str) -> set[str]:
    text = re.sub(r"[^\w\s]", " ", text.lower())
    return {w for w in text.split() if len(w) > 2 and w not in _STOPWORDS}


def name_similarity(a: str, b: str) -> float:
    ta, tb = _tokens(a), _tokens(b)
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def build_code_map() -> dict[str, str]:
    """Код (bare, lower) -> Prodnum (полный код с префиксом для galleyp.asp)."""
    try:
        resp = requests.get(MIKADO_PRICE_URL, timeout=60)
        resp.raise_for_status()
        if resp.content[:2] != b"PK":
            raise ValueError("ответ не похож на Excel")
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
        print("  Прайс Mikado: скачан свежий")
    except Exception as e:
        print(f"  Прайс Mikado: скачивание не удалось ({e}), использую локальный файл")
        wb = openpyxl.load_workbook(PRICE_FALLBACK, read_only=True, data_only=True)
    ws = wb.active
    mapping = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        prodnum, code = r[0], r[1]
        if code and prodnum:
            mapping[str(code).strip().lower()] = str(prodnum).strip()
    wb.close()
    return mapping


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rows", default=None, help="Диапазон строк Excel: 2-10 или одна строка: 5")
    ap.add_argument("--delay", default=0.6, type=float, help="Пауза между позициями, сек")
    ap.add_argument("--min-sim", default=0.12, type=float,
                     help="Порог схожести названий (word-overlap), ниже которого — несовпадение")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb.active
    col_code, col_name, col_brand, col_app = 1, 2, 3, 7

    max_row = ws.max_row
    row_start, row_end = 2, max_row
    if args.rows:
        if "-" in args.rows:
            a, b = args.rows.split("-")
            row_start, row_end = int(a), int(b)
        else:
            row_start = row_end = int(args.rows)
    row_end = min(row_end, max_row)

    print()
    print("=" * 62)
    print("  Mikado — Применяемость")
    print(f"  Файл:   {XLSX_PATH.name}")
    print(f"  Строки: {row_start}-{row_end} ({row_end - row_start + 1} шт.)")
    print("=" * 62)

    print("\n  Строю карту Код -> Prodnum...")
    code_map = build_code_map()
    print(f"  В прайсе {len(code_map)} позиций")

    env = load_env(Path(__file__).parent.parent / ".env")
    print("\n  Авторизация в Mikado...")
    session = login(env["MIKADO_CODE"], env["MIKADO_PASSWORD"])

    stats = {"done": 0, "skipped": 0, "not_in_price": 0, "no_artid": 0,
              "no_compat": 0, "mismatch": 0, "error": 0}
    mismatches = []
    since_save = 0

    try:
        for row_num in range(row_start, row_end + 1):
            code_raw = ws.cell(row_num, col_code).value
            if not code_raw:
                continue
            code = str(code_raw).strip().lower()
            our_name = str(ws.cell(row_num, col_name).value or "")
            our_brand = str(ws.cell(row_num, col_brand).value or "")

            if ws.cell(row_num, col_app).value:
                stats["skipped"] += 1
                continue

            try:
                prodnum = code_map.get(code)
                if not prodnum:
                    print(f"  [{row_num:3}] {code}: нет в текущем прайсе Mikado")
                    stats["not_in_price"] += 1
                    continue

                try:
                    _, data = fetch_product(session, prodnum)
                except Exception as e:
                    print(f"  [{row_num:3}] {code}: ошибка запроса карточки — {e}")
                    stats["error"] += 1
                    continue

                artid = data.get("_artid")
                if not artid:
                    print(f"  [{row_num:3}] {code}: карточка не найдена (нет ARTID)")
                    stats["no_artid"] += 1
                    continue

                mikado_name = data.get("name", "")
                mikado_brand = data.get("brand", "")
                sim = name_similarity(our_name, mikado_name)
                brand_ok = our_brand.strip().lower() == mikado_brand.strip().lower()

                if sim < args.min_sim and not brand_ok:
                    print(f"  [{row_num:3}] {code}: НЕСОВПАДЕНИЕ контекста "
                          f"(наше: «{our_name[:40]}» / Mikado: «{mikado_name[:40]}», "
                          f"бренд {our_brand}/{mikado_brand}) — пропуск")
                    mismatches.append(f"{row_num}\t{code}\t{our_brand}\t{our_name}\t{mikado_brand}\t{mikado_name}")
                    stats["mismatch"] += 1
                    continue

                time.sleep(0.4)
                try:
                    compat_rows = fetch_compatibility(session, artid)
                except Exception as e:
                    print(f"  [{row_num:3}] {code}: ошибка запроса применяемости — {e}")
                    stats["error"] += 1
                    continue
                text = compat_to_text(compat_rows)

                if not text:
                    print(f"  [{row_num:3}] {code}: применяемость на сайте не указана")
                    stats["no_compat"] += 1
                else:
                    cell = ws.cell(row_num, col_app, text)
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    print(f"  [{row_num:3}] {code}: {text[:70]}")
                    stats["done"] += 1
                    since_save += 1

            except Exception as e:
                print(f"  [{row_num:3}] {code}: непредвиденная ошибка — {e}")
                stats["error"] += 1

            if since_save >= SAVE_EVERY:
                wb.save(str(XLSX_PATH))
                since_save = 0

            if row_num < row_end:
                time.sleep(args.delay)

    finally:
        wb.save(str(XLSX_PATH))
        if mismatches:
            MISMATCH_LOG.write_text("\n".join(mismatches), encoding="utf-8")

    print()
    print("=" * 62)
    print("  Готово!")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    if mismatches:
        print(f"  Несовпадения записаны в: {MISMATCH_LOG}")
    print("=" * 62)
    print()


if __name__ == "__main__":
    main()
