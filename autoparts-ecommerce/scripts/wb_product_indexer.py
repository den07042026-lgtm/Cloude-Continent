"""
wb_product_indexer.py
════════════════════════════════════════════════════════════════════════════
Строит маппинг: наш каталог (Mikado + Автолига) → nm_id WB (топ 5-10 товаров).

МЕТОДЫ (применяются последовательно, результаты кэшируются в SQLite):

  М1. MPStats Brand Scan — для каждого бренда нашего каталога запрашиваем
      топ-100 товаров через MPStats /wb/get/brand.
      ~683 запроса × 1с = ~12 минут. БЕЗ запросов к WB — нет блокировок.
      Возвращает сразу: nm_id, sales, OOS%, commission_fbs.

  М2. OEM Exactmatch — для Автолиги (артикул = OEM-номер).
      Прямой поиск по OEM через WB search. Самая точная привязка.
      Поддерживает --proxy для обхода блокировок WB.

  М3. Brand+Name Match — из SQLite без доп. запросов.
      brand match (нормализованный) + пересечение стемов названия.
      Запускается автоматически после М1/М2.

Выход:
  data/analytics/wb_index.db       — SQLite (постоянный кэш, растёт)
  data/analytics/wb_matches_*.xlsx — Excel-отчёт

Запуск:
  uv run --with requests,openpyxl,xlrd,curl-cffi,python-dotenv \\
         scripts/wb_product_indexer.py
  uv run ... wb_product_indexer.py --method brand           # только М1 (MPStats)
  uv run ... wb_product_indexer.py --method oem             # только М2 (WB OEM)
  uv run ... wb_product_indexer.py --method oem --proxy http://host:port
  uv run ... wb_product_indexer.py --match-only             # только М3 из БД
  uv run ... wb_product_indexer.py --top-brands 50          # сколько брендов (0=все)
  uv run ... wb_product_indexer.py --top-oem 500            # лимит OEM за сессию
  uv run ... wb_product_indexer.py --no-export              # без Excel
  uv run ... wb_product_indexer.py --verify-oem             # OEM-верификация brand_name матчей
  uv run ... wb_product_indexer.py --verify-oem --top-verify 2000  # верифицировать 2000 пар
"""

import argparse
import json
import logging
import os
import random
import re
import sqlite3
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests as _req
    _REQ_OK = True
except ImportError:
    _req = None
    _REQ_OK = False

try:
    from curl_cffi import requests as _cffi
    CURL_OK = True
except ImportError:
    _cffi = None
    CURL_OK = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    XL_OK = True
except ImportError:
    XL_OK = False

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(*a, **kw): pass

# ─── Пути ─────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent.parent
load_dotenv(BASE_DIR / ".env")
CACHE_DIR   = BASE_DIR / "data" / "analytics" / "cache"
OUT_DIR     = BASE_DIR / "data" / "analytics"
LOG_FILE    = BASE_DIR / "logs" / "wb_product_indexer.log"
DB_PATH     = OUT_DIR / "wb_index.db"
MIKADO_FILE = Path("C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx")

# ─── MPStats ──────────────────────────────────────────────────────────────────
MPSTATS_TOKEN    = os.getenv("MPSTATS_TOKEN", "")
MPSTATS_BRAND_URL = "https://mpstats.io/api/wb/get/brand"
MP_DELAY         = (0.8, 1.5)       # пауза между MPStats запросами, с
MP_RETRY_WAITS   = [30, 90]         # при rate-limit

_today = datetime.now()
MP_D2  = _today.strftime("%Y-%m-%d")
MP_D1  = (_today - timedelta(days=30)).strftime("%Y-%m-%d")

# ─── WB Search API (только М2) ────────────────────────────────────────────────
WB_SEARCH_URL = "https://search.wb.ru/exactmatch/ru/common/v5/search"
WB_DEST       = -1257786
TOP_K         = 10
OEM_DELAY     = (7.0, 12.0)   # длиннее — ночной режим, меньше детектов
RETRY_WAITS   = [30, 190, 190]  # 190с = чуть больше 3 мин ротации IP

# Авто-фильтр для WB-товаров (М2)
AUTO_SUBJ_KW = frozenset({
    "автозапчаст", "запчаст", "колодк", "амортизатор", "тормоз",
    "подвеск", "рулев", "сцеплен", "ремен", "шаровая", "стойк",
    "свеч", "зажиган", "фильтр", "подшипник", "сальник", "прокладк",
    "датчик", "насос", "шрус", "привод", "масла", "аккумулятор",
    "генератор", "стартер", "ступиц", "зеркал", "трос", "наконечник",
    "рычаг", "пружин", "ремкомплект", "уплотнитель", "реле",
})
STOP_SUBJ_KW = frozenset({"мотоцикл", "велосипед", "скутер", "квадроцикл"})

# Верхнеуровневые авто-категории WB (по префиксу subject)
AUTO_TOP_CATS = ("автозапчаст", "масла и техн", "автоаксессуар", "шины", "автохим")


def _is_auto(subject: str) -> bool:
    """True если WB-товар относится к авторынку по полю subject."""
    s = (subject or "").lower()
    if not s:
        return True   # нет данных — оставляем
    if any(kw in s for kw in STOP_SUBJ_KW):
        return False
    # Товар авто если subject начинается с авто-категории ИЛИ содержит ключевое слово
    return any(s.startswith(cat) for cat in AUTO_TOP_CATS) or \
           any(kw in s for kw in AUTO_SUBJ_KW)


# Глобальный proxy (устанавливается из --proxy)
_PROXY: str | None = None

# ─── Логирование ──────────────────────────────────────────────────────────────
for _d in (CACHE_DIR, OUT_DIR, LOG_FILE.parent):
    _d.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  SQLite
# ══════════════════════════════════════════════════════════════════════════════

def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS wb_products (
            nm_id          INTEGER PRIMARY KEY,
            brand          TEXT,
            brand_norm     TEXT,
            name           TEXT,
            name_low       TEXT,
            subject        TEXT,
            subject_id     INTEGER,
            price_rub      REAL,
            sales_30d      REAL,
            oos_pct        REAL,
            commission_fbs REAL,
            fetched_at     TEXT,
            source         TEXT DEFAULT 'mpstats'
        );
        CREATE TABLE IF NOT EXISTS searched_brands (
            brand_norm  TEXT PRIMARY KEY,
            brand       TEXT,
            results     INTEGER,
            searched_at TEXT
        );
        CREATE TABLE IF NOT EXISTS wb_matches (
            our_article  TEXT,
            our_source   TEXT,
            our_brand    TEXT,
            our_price    REAL,
            nm_id        INTEGER,
            method       TEXT,
            score        REAL,
            oem_verified TEXT,
            PRIMARY KEY (our_article, nm_id)
        );
        CREATE TABLE IF NOT EXISTS wb_card_oem (
            nm_id      INTEGER PRIMARY KEY,
            oem_list   TEXT,
            fetched_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_brand_norm ON wb_products(brand_norm);
    """)
    # Миграция: добавляем новые колонки если их нет (для старых БД)
    existing = {row[1] for row in conn.execute("PRAGMA table_info(wb_products)")}
    for col, defn in [
        ("subject_id",     "INTEGER"),
        ("sales_30d",      "REAL"),
        ("oos_pct",        "REAL"),
        ("commission_fbs", "REAL"),
        ("source",         "TEXT DEFAULT 'mpstats'"),
    ]:
        if col not in existing:
            conn.execute(f"ALTER TABLE wb_products ADD COLUMN {col} {defn}")
    existing_m = {row[1] for row in conn.execute("PRAGMA table_info(wb_matches)")}
    if "oem_verified" not in existing_m:
        conn.execute("ALTER TABLE wb_matches ADD COLUMN oem_verified TEXT")
    conn.commit()
    return conn


# ══════════════════════════════════════════════════════════════════════════════
#  М1: MPStats Brand Scan
# ══════════════════════════════════════════════════════════════════════════════

def _norm_brand(b: str) -> str:
    return re.sub(r"[\s\-_\.]", "", b.lower()).strip()


def _mpstats_brand(brand_name: str) -> list[dict]:
    """
    MPStats /wb/get/brand?path={brand_name} → топ-100 товаров бренда.
    Кэш: CACHE_DIR/mpsb_{safe}.json (бессрочный).
    """
    if not MPSTATS_TOKEN:
        log.error("MPSTATS_TOKEN не задан — М1 недоступен")
        return []

    safe  = re.sub(r"[^\w]", "_", brand_name.lower())[:55]
    cache = CACHE_DIR / f"mpsb_{safe}.json"
    if cache.exists():
        try:
            return json.loads(cache.read_text(encoding="utf-8"))
        except Exception:
            cache.unlink(missing_ok=True)

    params  = {"path": brand_name, "d1": MP_D1, "d2": MP_D2, "startRow": 0, "endRow": 100}
    headers = {"X-Mpstats-TOKEN": MPSTATS_TOKEN}

    for attempt, wait in enumerate([0] + MP_RETRY_WAITS):
        if wait:
            log.warning(f"  MPStats rate-limit (попытка {attempt}), ждём {wait}с...")
            time.sleep(wait)
        try:
            r = _req.get(MPSTATS_BRAND_URL, headers=headers, params=params, timeout=30)
            if r.status_code == 429:
                continue
            if r.status_code == 401:
                log.error("MPStats: неверный токен (401)")
                return []
            if r.status_code != 200:
                log.debug(f"  MPStats [{brand_name}]: HTTP {r.status_code}")
                return []
            data  = r.json()
            rows  = data if isinstance(data, list) else (data or {}).get("data", [])
            prods = []
            for p in (rows or []):
                nm_id = p.get("id")
                if not nm_id:
                    continue
                subj = (p.get("subject") or "").strip()
                prods.append({
                    "nm_id":          int(nm_id),
                    "brand":          (p.get("brand") or "").strip(),
                    "name":           (p.get("name") or "").strip(),
                    "subject":        subj,
                    "subject_id":     p.get("subject_id") or 0,
                    "price":          p.get("final_price") or p.get("final_price_average") or 0,
                    "sales_30d":      p.get("sales") or 0,
                    "oos_pct":        p.get("lost_profit_percent") or 0,
                    "commission_fbs": (p.get("commission_fbs") or 30.0) / 100.0,
                })
            # Кэшируем даже пустой список — бренда нет на WB (легитимный пустой ответ)
            cache.write_text(json.dumps(prods, ensure_ascii=False), encoding="utf-8")
            return prods
        except Exception as e:
            log.debug(f"  MPStats [{brand_name}]: {e}")
            time.sleep(3)

    return []


def collect_brands(mikado: list[dict], autoliga: list[dict], top_n: int = 0) -> list[tuple]:
    """Собирает уникальные бренды. Возвращает [(brand_display, brand_norm, count)]."""
    counts: dict[str, dict] = {}
    for items in (mikado, autoliga):
        for it in items:
            b = it.get("brand", "").strip()
            if not b:
                continue
            bn = _norm_brand(b)
            if bn not in counts:
                counts[bn] = {"display": b, "count": 0}
            counts[bn]["count"] += 1
    result = sorted(
        [(v["display"], k, v["count"]) for k, v in counts.items()],
        key=lambda x: x[2], reverse=True
    )
    log.info(f"Уникальных брендов в каталоге: {len(result)}")
    return result[:top_n] if top_n else result


def run_brand_search(brands: list[tuple], conn: sqlite3.Connection,
                     force: bool = False) -> int:
    """М1: для каждого бренда → MPStats → SQLite. Возвращает кол-во новых nm_id."""
    if not _REQ_OK:
        log.error("requests не установлен — М1 недоступен")
        return 0
    if not MPSTATS_TOKEN:
        log.error("MPSTATS_TOKEN не задан в .env — М1 недоступен")
        return 0

    already = set(
        r[0] for r in conn.execute("SELECT brand_norm FROM searched_brands").fetchall()
    ) if not force else set()

    todo  = [(d, bn, cnt) for (d, bn, cnt) in brands if bn not in already]
    total = len(todo)
    log.info(f"Брендов к запросу MPStats: {total}  (уже готово: {len(brands)-total})")

    if not todo:
        log.info("Все бренды уже проверены — пропускаем М1")
        return 0

    new_total = 0
    now_str   = datetime.now().isoformat()

    for idx, (display, bnorm, cnt) in enumerate(todo):
        safe   = re.sub(r"[^\w]", "_", display.lower())[:55]
        cached = (CACHE_DIR / f"mpsb_{safe}.json").exists()
        log.info(f"  [{idx+1}/{total}] {display} ({cnt} поз.)  {'[кэш]' if cached else '[MPStats]'}")

        prods = _mpstats_brand(display)

        rows = [
            (p["nm_id"], p["brand"], _norm_brand(p["brand"]),
             p["name"], p["name"].lower(),
             p["subject"], p["subject_id"], p["price"],
             p["sales_30d"], p["oos_pct"], p["commission_fbs"],
             now_str, "mpstats")
            for p in prods
        ]
        if rows:
            conn.executemany("""
                INSERT OR REPLACE INTO wb_products
                (nm_id, brand, brand_norm, name, name_low, subject, subject_id,
                 price_rub, sales_30d, oos_pct, commission_fbs, fetched_at, source)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, rows)

        conn.execute("""
            INSERT OR REPLACE INTO searched_brands (brand_norm, brand, results, searched_at)
            VALUES (?,?,?,?)
        """, (bnorm, display, len(prods), now_str))
        conn.commit()

        new_total += len(prods)
        log.info(f"    → {len(prods)} товаров WB")

        if not cached:
            time.sleep(random.uniform(*MP_DELAY))

    return new_total


# ══════════════════════════════════════════════════════════════════════════════
#  М2: OEM Exactmatch (WB search, Автолига)
# ══════════════════════════════════════════════════════════════════════════════

_session = None
_WB_COOKIE_FILE = OUT_DIR / "wb_cookies.json"


def _load_wb_cookies() -> dict:
    """Загружает cookies из файла, сохранённого wb_get_cookie.py."""
    if _WB_COOKIE_FILE.exists():
        try:
            import json as _json
            c = _json.loads(_WB_COOKIE_FILE.read_text(encoding="utf-8"))
            if c:
                log.info(f"WB cookies загружены: {list(c.keys())}")
                return c
        except Exception as e:
            log.warning(f"Не удалось загрузить wb_cookies.json: {e}")
    return {}


def get_session():
    global _session
    if _session:
        return _session
    if CURL_OK:
        s = _cffi.Session(impersonate="chrome124")
        log.info("WB-сессия: curl_cffi Chrome124 ✓")
    elif _REQ_OK:
        s = _req.Session()
        s.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "ru-RU,ru;q=0.9",
        })
        log.warning("WB-сессия: requests (без Chrome-fingerprint)")
    else:
        raise RuntimeError("Нет requests / curl_cffi")

    # Прогрев через wildberries.ru только если нет прокси (прямое соединение)
    # С прокси пропускаем — 498 antibot "метит" сессию и усугубляет блокировку
    if not _PROXY:
        try:
            s.get("https://www.wildberries.ru/", timeout=12)
            time.sleep(random.uniform(1.5, 2.5))
        except Exception as e:
            log.warning(f"WB прогрев: {e}")

    # Загружаем WBAAS cookies в сессию
    wb_cookies = _load_wb_cookies()
    if wb_cookies:
        try:
            s.cookies.update(wb_cookies)
        except Exception:
            pass
    _session = s
    return s


def _oem_cache(oem: str) -> Path:
    safe = re.sub(r"[^\w]", "_", oem)[:55]
    return CACHE_DIR / f"wbidx_{safe}.json"


def _wb_search(query: str, limit: int = 10) -> list[dict]:
    """
    WB exactmatch search. Возвращает [{nm_id, brand, name, subject}].
    Не кэширует пустые ответы при подозрении на антискрейп (raw_count=0).
    """
    safe  = re.sub(r"[^\w]", "_", query.lower())[:55]
    cache = CACHE_DIR / f"wbidx_{safe}.json"
    if cache.exists():
        try:
            return json.loads(cache.read_text(encoding="utf-8"))
        except Exception:
            cache.unlink(missing_ok=True)

    params = {
        "query": query, "resultset": "catalog",
        "limit": limit * 3, "sort": "popular",
        "page": 1, "dest": WB_DEST,
        "appType": 1, "curr": "rub", "spp": 27,
    }
    proxies   = {"https": _PROXY, "http": _PROXY} if _PROXY else None
    raw_count = 0

    # Передаём WBAAS-токен напрямую в заголовке Cookie для search.wb.ru
    _wb_cookies_dict = _load_wb_cookies()
    _cookie_header   = "; ".join(f"{k}={v}" for k, v in _wb_cookies_dict.items()) if _wb_cookies_dict else ""
    _req_headers = {
        "Referer": "https://www.wildberries.ru/",
        "Origin":  "https://www.wildberries.ru",
        "Accept":  "application/json, text/plain, */*",
    }
    if _cookie_header:
        _req_headers["Cookie"] = _cookie_header

    for attempt, wait in enumerate([0] + RETRY_WAITS):
        if wait:
            log.warning(f"  WB rate-limit (попытка {attempt}), ждём {wait}с...")
            time.sleep(wait)
            global _session
            _session = None
        try:
            r = get_session().get(WB_SEARCH_URL, params=params,
                                  headers=_req_headers,
                                  proxies=proxies, timeout=20)
            if r.status_code == 429:
                continue
            if r.status_code != 200:
                log.debug(f"  WB [{query}]: HTTP {r.status_code}")
                break
            raw       = r.json()
            items     = raw.get("products") or (raw.get("data") or {}).get("products") or []
            raw_count = len(items)
            prods     = []
            for p in items:
                nm = p.get("id")
                if not nm:
                    continue
                subj     = (p.get("subjectName") or p.get("subject") or "").lower()
                name_low = (p.get("name") or "").lower()
                is_auto  = (
                    any(k in subj for k in AUTO_SUBJ_KW)
                    or any(k in name_low for k in AUTO_SUBJ_KW)
                )
                if is_auto and not any(k in subj for k in STOP_SUBJ_KW):
                    prods.append({
                        "nm_id":   int(nm),
                        "brand":   (p.get("brand") or "").strip(),
                        "name":    (p.get("name") or "").strip(),
                        "subject": (p.get("subjectName") or "").strip(),
                        "price":   (p.get("priceU") or p.get("salePriceU") or 0) / 100,
                    })
                    if len(prods) >= limit:
                        break
            if prods or raw_count > 0:
                cache.write_text(json.dumps(prods, ensure_ascii=False), encoding="utf-8")
            else:
                log.debug(f"  WB [{query}]: пустой ответ (антискрейп?) — не кэшируем")
            return prods
        except Exception as e:
            log.debug(f"  WB [{query}] попытка {attempt+1}: {e}")
            time.sleep(5)

    log.warning(f"  WB [{query}]: все ретраи исчерпаны, результат не кэшируется")
    return []


def run_oem_search(al_items: list[dict], conn: sqlite3.Connection,
                   limit_n: int = 0) -> int:
    """М2: OEM exactmatch для Автолиги. Возвращает кол-во новых матчей."""
    done = set(
        r[0] for r in conn.execute(
            "SELECT our_article FROM wb_matches WHERE our_source='autoliga' AND method='oem'"
        ).fetchall()
    )
    todo = [it for it in al_items if it["article"] not in done]
    if limit_n:
        todo = todo[:limit_n]

    log.info(f"OEM к поиску: {len(todo):,}  (уже есть: {len(done):,})")
    if not todo:
        return 0

    found   = 0
    now_str = datetime.now().isoformat()

    for idx, item in enumerate(todo):
        oem     = item.get("oem") or item["article"]
        article = item["article"]
        cached  = _oem_cache(oem).exists()

        if (idx + 1) % 200 == 0 or idx == 0:
            log.info(f"  OEM {idx+1}/{len(todo)}  найдено: {found}")

        prods = _wb_search(oem, limit=TOP_K)

        if prods:
            conn.executemany("""
                INSERT OR IGNORE INTO wb_products
                (nm_id, brand, brand_norm, name, name_low, subject, price_rub, fetched_at, source)
                VALUES (?,?,?,?,?,?,?,?,'wb')
            """, [(p["nm_id"], p["brand"], _norm_brand(p["brand"]),
                   p["name"], p["name"].lower(), p["subject"], p["price"], now_str)
                  for p in prods])
            conn.executemany("""
                INSERT OR REPLACE INTO wb_matches
                (our_article, our_source, our_brand, our_price, nm_id, method, score)
                VALUES (?,?,?,?,?,'oem',1.0)
            """, [(article, "autoliga", item["brand"], item["price"], p["nm_id"])
                  for p in prods])
            conn.commit()
            found += 1

        if not cached:
            time.sleep(random.uniform(*OEM_DELAY))

    log.info(f"М2 готово: матчи для {found}/{len(todo)} OEM-артикулов")
    return found


# ══════════════════════════════════════════════════════════════════════════════
#  М3: Brand+Name Match из SQLite
# ══════════════════════════════════════════════════════════════════════════════

# ── OEM-верификация ───────────────────────────────────────────────────────────

_OEM_CHAR_KEYS = frozenset({
    "oem номера", "номер oem", "oem номер", "oemномера",
    "артикул производителя", "oem", "оем", "oem-номера",
})

WB_CARD_URL  = "https://card.wb.ru/cards/v2/detail"
OEM_DELAY_V  = (0.8, 1.5)


def _norm_oem(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def fetch_card_oem(nm_id: int, conn: sqlite3.Connection) -> list[str]:
    """Возвращает нормализованные OEM-номера из карточки WB. Кэш в wb_card_oem."""
    row = conn.execute("SELECT oem_list FROM wb_card_oem WHERE nm_id=?", (nm_id,)).fetchone()
    if row is not None:
        return json.loads(row[0])

    if not _REQ_OK:
        return []
    try:
        params = {"appType": 1, "curr": "rub", "dest": WB_DEST, "spp": 27, "nm": nm_id}
        r = _req.get(WB_CARD_URL, params=params, timeout=15)
        if r.status_code != 200:
            return []
        data     = r.json()
        products = (data.get("data") or {}).get("products") or []
        if not products:
            return []

        oem_list: list[str] = []
        for char in (products[0].get("characteristics") or []):
            if isinstance(char, list):
                items = char
            else:
                items = [char]
            for item in items:
                key = (item.get("name") or "").lower().replace("-", " ").strip()
                if key in _OEM_CHAR_KEYS:
                    val = str(item.get("value") or "")
                    for part in re.split(r"[;,\n]", val):
                        norm = _norm_oem(part.strip())
                        if norm and len(norm) >= 5:
                            oem_list.append(norm)

        conn.execute("""
            INSERT OR REPLACE INTO wb_card_oem (nm_id, oem_list, fetched_at)
            VALUES (?, ?, ?)
        """, (nm_id, json.dumps(oem_list, ensure_ascii=False), datetime.now().isoformat()))
        conn.commit()
        # Пауза только после реального сетевого запроса, не из кэша
        time.sleep(random.uniform(*OEM_DELAY_V))
        return oem_list
    except Exception as e:
        log.debug(f"fetch_card_oem {nm_id}: {e}")
        return []


def run_oem_verify(al_items: list[dict], conn: sqlite3.Connection, limit: int = 0) -> None:
    """
    Верифицирует brand_name матчи сравнением OEM из карточки WB с нашим OEM.

    oem_verified:
      'yes'     — OEM совпал  (или метод oem — уже точный)
      'no'      — WB-карточка содержит OEM, но наш не найден
      'unknown' — карточка без OEM-данных, или нет OEM на нашей стороне (Mikado)
    """
    # Метод oem — уже точный, сразу ставим yes
    conn.execute("""
        UPDATE wb_matches SET oem_verified='yes'
        WHERE method='oem' AND oem_verified IS NULL
    """)
    # Mikado — OEM на нашей стороне нет → unknown
    conn.execute("""
        UPDATE wb_matches SET oem_verified='unknown'
        WHERE method='brand_name' AND our_source='mikado' AND oem_verified IS NULL
    """)
    conn.commit()

    # Строим карту article → [norm_oem] для Автолиги
    al_oem_map: dict[str, list[str]] = {}
    for it in al_items:
        article = it.get("article", "")
        oems = set()
        for raw in [it.get("oem") or "", it.get("article") or ""]:
            n = _norm_oem(raw)
            if n and len(n) >= 5:
                oems.add(n)
        if oems:
            al_oem_map[article] = list(oems)

    rows = conn.execute("""
        SELECT our_article, nm_id, score
        FROM wb_matches
        WHERE method='brand_name' AND our_source='autoliga' AND oem_verified IS NULL
        ORDER BY score DESC
    """).fetchall()
    if limit:
        rows = rows[:limit]

    # Группируем по nm_id: один WB-товар → несколько наших артикулов.
    # Так каждый nm_id запрашивается ровно один раз, пауза — только при сетевом запросе.
    by_nm: dict[int, list[str]] = {}
    for article, nm_id, score in rows:
        by_nm.setdefault(nm_id, []).append(article)

    unique_nm = len(by_nm)
    cached_nm = conn.execute(
        f"SELECT COUNT(*) FROM wb_card_oem WHERE nm_id IN ({','.join('?' * unique_nm)})",
        list(by_nm.keys()),
    ).fetchone()[0] if unique_nm <= 999 else 0  # sqlite limit

    log.info(
        f"OEM-верификация: {len(rows):,} пар → {unique_nm:,} уникальных nm_id "
        f"({cached_nm} в кэше, ~{unique_nm - cached_nm} сетевых запросов)"
    )

    verified = no_match = unknown = 0
    pending: list[tuple] = []

    for i, (nm_id, articles) in enumerate(by_nm.items()):
        wb_oems = fetch_card_oem(nm_id, conn)  # sleep только при реальном запросе

        for article in articles:
            our_oems = al_oem_map.get(article, [])
            if not our_oems or not wb_oems:
                result = "unknown"
                unknown += 1
            elif any(o in wb_oems for o in our_oems):
                result = "yes"
                verified += 1
            else:
                result = "no"
                no_match += 1
            pending.append((result, article, nm_id))

        if len(pending) >= 500:
            conn.executemany(
                "UPDATE wb_matches SET oem_verified=? WHERE our_article=? AND nm_id=?",
                pending,
            )
            conn.commit()
            pending.clear()

        if (i + 1) % 200 == 0:
            log.info(f"  [{i+1}/{unique_nm} nm_id]  ✓={verified}  ✗={no_match}  ?={unknown}")

    if pending:
        conn.executemany(
            "UPDATE wb_matches SET oem_verified=? WHERE our_article=? AND nm_id=?",
            pending,
        )
        conn.commit()
    log.info(f"OEM-верификация завершена: ✓={verified}  ✗={no_match}  ?={unknown}")


# ── Brand+Name Match ──────────────────────────────────────────────────────────

_NAME_STOP = frozenset({
    "для", "авто", "автомобил", "автомобиль", "запчаст", "запчасти",
    "деталь", "детали", "оригинал", "аналог", "новый", "новая",
    "комплект", "и", "в", "на", "с", "из", "к", "по",
})


def _stems(text: str) -> set[str]:
    words = re.split(r"[\s\-_/,\.\(\)]+", text.lower())
    return {w[:7] for w in words if len(w) >= 4 and w.strip("()[]") not in _NAME_STOP}


def match_in_db(article: str, brand: str, name: str, price: float,
                source: str, conn: sqlite3.Connection) -> list[dict]:
    """Ищет WB-товары по brand+name в SQLite. Топ TOP_K."""
    bnorm = _norm_brand(brand)
    if not bnorm:
        return []

    rows = conn.execute("""
        SELECT nm_id, brand, name, price_rub, subject, sales_30d, oos_pct
        FROM wb_products
        WHERE brand_norm LIKE ?
        LIMIT 5000
    """, (f"%{bnorm[:7]}%",)).fetchall()

    if not rows:
        return []

    # Оставляем только автозапчасти — отсекаем Bosch-дрели, Samsung-телефоны и т.п.
    rows = [r for r in rows if _is_auto(r[4])]
    if not rows:
        return []

    our_stems = _stems(name)
    results   = []

    for nm_id, wb_brand, wb_name, wb_price, subject, sales, oos in rows:
        wb_bnorm = _norm_brand(wb_brand)
        if not (wb_bnorm == bnorm or bnorm in wb_bnorm or wb_bnorm in bnorm):
            continue
        wb_stems = _stems(wb_name or "")
        if our_stems and wb_stems:
            common = our_stems & wb_stems
            score  = len(common) / max(len(our_stems), len(wb_stems))
        else:
            score = 0.05
        results.append({
            "nm_id":   nm_id,
            "brand":   wb_brand,
            "name":    wb_name,
            "price":   wb_price,
            "subject": subject,
            "sales":   sales or 0,
            "oos_pct": oos or 0,
            "score":   round(score, 3),
        })

    results.sort(key=lambda x: x["score"], reverse=True)
    return results[:TOP_K]


def run_match(mikado: list[dict], al_items: list[dict], conn: sqlite3.Connection) -> None:
    """М3: brand+name matching для Mikado и Автолиги."""
    total_in_db = conn.execute("SELECT COUNT(*) FROM wb_products").fetchone()[0]
    log.info(f"WB-товаров в SQLite: {total_in_db:,}")
    if total_in_db == 0:
        log.warning("Индекс пуст — сначала запусти --method brand или --method oem")
        return

    now_str = datetime.now().isoformat()

    def _do_match(items: list[dict], source: str) -> None:
        done = set(
            r[0] for r in conn.execute(
                "SELECT our_article FROM wb_matches WHERE our_source=?", (source,)
            ).fetchall()
        )
        todo    = [it for it in items if it["article"] not in done]
        matched = 0
        for it in todo:
            hits = match_in_db(it["article"], it["brand"], it["name"],
                               it["price"], source, conn)
            if hits:
                conn.executemany("""
                    INSERT OR REPLACE INTO wb_matches
                    (our_article, our_source, our_brand, our_price, nm_id, method, score)
                    VALUES (?,?,?,?,?,'brand_name',?)
                """, [(it["article"], source, it["brand"], it["price"],
                       h["nm_id"], h["score"]) for h in hits])
                matched += 1
        conn.commit()
        log.info(f"  {source}: {matched}/{len(todo)} сопоставлено по brand+name")

    _do_match(mikado, "mikado")
    _do_match(al_items, "autoliga")


# ══════════════════════════════════════════════════════════════════════════════
#  Загрузка каталогов
# ══════════════════════════════════════════════════════════════════════════════

def load_mikado() -> list[dict]:
    if not XL_OK:
        log.error("openpyxl не установлен")
        return []
    if not MIKADO_FILE.exists():
        log.warning(f"Mikado: файл не найден — {MIKADO_FILE}")
        return []
    try:
        wb  = openpyxl.load_workbook(str(MIKADO_FILE), read_only=True, data_only=True)
        ws  = wb.active
        rows = ws.iter_rows(values_only=True)
        hdr  = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]
        ci   = {k: next((i for i, h in enumerate(hdr) if h == k), None)
                for k in ("code", "brandname", "prodname", "priceout")}
        items = []
        for row in rows:
            def g(k, d=""):
                i = ci.get(k)
                return str(row[i]).strip() if i is not None and len(row) > i and row[i] else d
            code  = g("code")
            price = 0.0
            try: price = float(g("priceout", "0") or 0)
            except Exception: pass
            if code and price > 0:
                items.append({"article": code, "brand": g("brandname"),
                              "name": g("prodname"), "price": price, "source": "mikado"})
        wb.close()
        log.info(f"Mikado: {len(items):,} позиций")
        return items
    except Exception as e:
        log.error(f"Mikado: {e}")
        return []


def load_autoliga() -> list[dict]:
    try:
        sys.path.insert(0, str(Path(__file__).parent))
        from autoliga_loader import load_autoliga as _load
        raw = _load()
        items = []
        for oem, v in raw.items():
            price = float(v.get("price") or 0)
            if price <= 0:
                continue
            items.append({
                "article": v.get("article", oem),
                "oem":     oem,
                "brand":   v.get("brand", ""),
                "name":    v.get("name", ""),
                "price":   price,
                "source":  "autoliga",
            })
        log.info(f"Автолига: {len(items):,} позиций")
        return items
    except Exception as e:
        log.error(f"Автолига: {e}")
        return []


# ══════════════════════════════════════════════════════════════════════════════
#  Excel
# ══════════════════════════════════════════════════════════════════════════════

def _reliability(method: str, oem_verified, score) -> float:
    """
    Надёжность матча — насколько уверенно наш артикул = WB-товар.

    Уровни:
      1.00        oem-метод + OEM подтверждён
      0.95        oem-метод, ещё не верифицировался
      0.50–1.00   brand_name + OEM подтверждён  (0.50 + score×0.50)
      0.00–0.60   brand_name + неизвестно        (score × 0.60)
      0.00–0.45   brand_name + не проверялось    (score × 0.45)
      0.00        OEM отклонён
    """
    s = float(score or 0)
    if method == "oem":
        if oem_verified == "no":  return 0.00
        if oem_verified == "yes": return 1.00
        return 0.95                             # unknown / NULL
    # brand_name
    if oem_verified == "no":      return 0.00
    if oem_verified == "yes":     return round(0.50 + s * 0.50, 3)
    if oem_verified == "unknown": return round(s * 0.60, 3)
    return round(s * 0.45, 3)                  # NULL — не проверялось


def _attract(row) -> float:
    """
    Индекс выгодности позиции с учётом дефицитности.

    Формула: OOS × (1 + Продажи/100) × Маржа × Надёжность_матча

    OOS — главный сигнал, задаёт порядок.
    Продажи — аддитивный бонус (не обнуляют результат при sales=0).
    Маржа — фильтр (<5% → 0) и мультипликатор.
    Надёжность — штраф за слабый матч (oem='no' → 0).
    """
    method    = row[5]
    our_price = float(row[3] or 0)
    wb_price  = float(row[9] or 0)
    sales     = float(row[11] or 0)
    oos       = float(row[12] or 0)          # 0..100
    comm      = float(row[13] or 0.25)
    oem_v     = row[14]

    if oem_v == "no":
        return 0.0

    # Маржа: если цен нет — предполагаем 15% (не обнуляем сортировку)
    if wb_price > 0 and our_price > 0:
        revenue = wb_price * (1.0 - comm)
        margin  = (revenue - our_price) / revenue
    else:
        margin = 0.15

    if margin < 0.05:
        return 0.0

    # Продажи — бонус от 1× до 11× (при 1000 шт/мес), не обнуляют OOS
    sales_boost = 1.0 + min(sales, 1000) / 100.0

    raw = oos * sales_boost * margin

    rel_w = {"yes": 1.00, "unknown": 0.60, None: 0.50, "no": 0.00}.get(oem_v, 0.50)
    if method == "oem":
        rel_w = max(rel_w, 0.95)

    return round(raw * rel_w, 3)


def export_excel(conn: sqlite3.Connection) -> Path | None:
    if not XL_OK:
        log.warning("openpyxl не установлен")
        return None

    rows = conn.execute("""
        SELECT m.our_article, m.our_source, m.our_brand, m.our_price,
               m.nm_id, m.method, m.score,
               p.name, p.brand, p.price_rub, p.subject,
               p.sales_30d, p.oos_pct, p.commission_fbs,
               m.oem_verified
        FROM wb_matches m
        LEFT JOIN wb_products p ON m.nm_id = p.nm_id
        ORDER BY m.our_brand, m.our_article, m.score DESC
    """).fetchall()

    if not rows:
        log.warning("Нет матчей для экспорта")
        return None

    # Фильтр 1: только авто-категории (r[10] = subject)
    before = len(rows)
    rows = [r for r in rows if _is_auto(r[10])]
    log.info(f"Фильтр категорий: оставлено {len(rows):,} из {before:,} строк")

    # Фильтр 2: brand_name матчи со score < 0.5 ненадёжны — убираем
    before = len(rows)
    rows = [r for r in rows if r[5] == "oem" or (r[6] or 0) >= 0.5]
    log.info(f"Фильтр score<0.5: оставлено {len(rows):,} из {before:,} строк")

    # Сортируем по убыванию индекса выгодности (OOS × Продажи × Маржа × Надёжность)
    rows = sorted(rows, key=_attract, reverse=True)

    wb_out = openpyxl.Workbook()
    ws     = wb_out.active
    ws.title = "Маппинг"

    HEADERS = [
        "Наш артикул", "Источник", "Наш бренд", "Наша цена ₽",
        "nm_id WB", "Метод", "Скор", "OEM ✓", "Надёжность", "Дефицит▼",
        "Название WB", "Бренд WB", "Цена WB ₽", "Ниша WB",
        "Продаж/мес", "OOS%", "Комиссия FBS%", "Маржа%", "Ссылка WB",
    ]
    ws.append(HEADERS)
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill  = hfill
        cell.font  = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    OEM_FILL   = PatternFill("solid", fgColor="C6EFCE")
    BRAND_FILL = PatternFill("solid", fgColor="FFEB9C")

    OEM_YES_FILL = PatternFill("solid", fgColor="92D050")
    OEM_NO_FILL  = PatternFill("solid", fgColor="FF6B6B")
    OEM_UNK_FILL = PatternFill("solid", fgColor="D9D9D9")

    _OEM_LABEL    = {"yes": "✓", "no": "✗", "unknown": "?", None: ""}
    _OEM_CELL_FILL = {"yes": OEM_YES_FILL, "no": OEM_NO_FILL, "unknown": OEM_UNK_FILL}

    # Цвет ячейки «Надёжность» по порогам
    REL_GREEN  = PatternFill("solid", fgColor="00B050")  # ≥ 0.80
    REL_LIME   = PatternFill("solid", fgColor="92D050")  # 0.60–0.79
    REL_YELLOW = PatternFill("solid", fgColor="FFEB9C")  # 0.35–0.59
    REL_ORANGE = PatternFill("solid", fgColor="FFC7CE")  # 0.10–0.34
    REL_RED    = PatternFill("solid", fgColor="FF6B6B")  # < 0.10

    def _rel_fill(r: float) -> PatternFill:
        if r >= 0.80: return REL_GREEN
        if r >= 0.60: return REL_LIME
        if r >= 0.35: return REL_YELLOW
        if r >= 0.10: return REL_ORANGE
        return REL_RED

    for ri, row in enumerate(rows, 2):
        (*cols, wb_name, wb_brand, wb_price, subject, sales, oos, comm, oem_v) = row
        nm_id    = cols[4]
        method   = cols[5]
        score_v  = cols[6]
        link     = f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx" if nm_id else ""
        comm_pct = round((comm or 0) * 100, 1)
        oem_label = _OEM_LABEL.get(oem_v, "")
        rel = _reliability(method, oem_v, score_v)
        attr = _attract(row)

        # Маржа% для отображения
        wp = float(wb_price or 0)
        op = float(cols[3] or 0)
        cm = float(comm or 0.25)
        if wp > 0 and op > 0:
            margin_pct = round((wp * (1 - cm) - op) / (wp * (1 - cm)) * 100, 1)
        else:
            margin_pct = ""

        # cols[0..6] + OEM ✓ + Надёжность + Дефицит▼ + WB-данные + Маржа%
        ws.append(list(cols) + [oem_label, f"{rel:.0%}", attr,
                                 wb_name, wb_brand, wb_price, subject,
                                 sales, oos, comm_pct, margin_pct, "→ WB"])

        fill = OEM_FILL if method == "oem" else BRAND_FILL
        for col in range(1, len(HEADERS) + 1):
            ws.cell(ri, col).fill = fill

        # OEM ✓ (столбец 8)
        oem_cell = ws.cell(ri, 8)
        if oem_v in _OEM_CELL_FILL:
            oem_cell.fill = _OEM_CELL_FILL[oem_v]
        oem_cell.alignment = Alignment(horizontal="center")
        oem_cell.font = Font(bold=True)

        # Надёжность (столбец 9)
        rel_cell = ws.cell(ri, 9)
        rel_cell.fill = _rel_fill(rel)
        rel_cell.alignment = Alignment(horizontal="center")
        rel_cell.font = Font(bold=(rel >= 0.80))

        # Дефицит▼ (столбец 10) — числовой индекс, жирный если высокий
        def_cell = ws.cell(ri, 10)
        def_cell.alignment = Alignment(horizontal="center")
        if attr >= 10:
            def_cell.font = Font(bold=True, color="375623")
        elif attr == 0:
            def_cell.font = Font(color="AAAAAA")

        lc = ws.cell(ri, len(HEADERS))
        if link:
            lc.hyperlink = link
            lc.font  = Font(color="0563C1", underline="single")
            lc.value = "→ WB"

    widths = [18, 10, 16, 11, 12, 11, 7, 6, 9, 9, 52, 16, 11, 32, 10, 7, 11, 9, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb_out.create_sheet("Инфо")
    total_arts  = conn.execute("SELECT COUNT(DISTINCT our_article) FROM wb_matches").fetchone()[0]
    oem_arts    = conn.execute("SELECT COUNT(DISTINCT our_article) FROM wb_matches WHERE method='oem'").fetchone()[0]
    brand_arts  = conn.execute("SELECT COUNT(DISTINCT our_article) FROM wb_matches WHERE method='brand_name'").fetchone()[0]
    wb_total    = conn.execute("SELECT COUNT(*) FROM wb_products").fetchone()[0]
    brands_done = conn.execute("SELECT COUNT(*) FROM searched_brands").fetchone()[0]
    mp_count    = conn.execute("SELECT COUNT(*) FROM wb_products WHERE source='mpstats'").fetchone()[0]
    oem_yes = conn.execute("SELECT COUNT(*) FROM wb_matches WHERE oem_verified='yes'").fetchone()[0]
    oem_no  = conn.execute("SELECT COUNT(*) FROM wb_matches WHERE oem_verified='no'").fetchone()[0]
    oem_unk = conn.execute("SELECT COUNT(*) FROM wb_matches WHERE oem_verified='unknown'").fetchone()[0]
    for r_data in [
        ("Дата",                  datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("WB-товаров в БД",       wb_total),
        ("  из них MPStats",      mp_count),
        ("Брендов проверено",     brands_done),
        ("Артикулов с матчем",    total_arts),
        ("  в т.ч. OEM",         oem_arts),
        ("  в т.ч. brand+name",  brand_arts),
        ("OEM верификация ✓",     oem_yes),
        ("OEM верификация ✗",     oem_no),
        ("OEM верификация ?",     oem_unk),
        ("MPStats период",        f"{MP_D1} — {MP_D2}"),
        ("Цвет OEM ✓",            "✓=зелёный ✗=красный ?=серый | строка: oem=светло-зелёный, brand_name=жёлтый"),
    ]:
        ws2.append(list(r_data))
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 30

    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUT_DIR / f"wb_matches_{ts}.xlsx"
    wb_out.save(str(path))
    log.info(f"Excel: {path}  ({len(rows)} строк, {total_arts} артикулов)")
    return path


# ══════════════════════════════════════════════════════════════════════════════
#  main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    global _PROXY

    parser = argparse.ArgumentParser(description="WB Product Indexer")
    parser.add_argument("--method",       choices=["all", "brand", "oem", "match"], default="all",
                        help="Метод: all=все, brand=М1(MPStats), oem=М2(WB), match=только М3")
    parser.add_argument("--top-brands",   type=int, default=0,
                        help="Сколько брендов (0=все, сортировка по кол-ву позиций)")
    parser.add_argument("--top-oem",      type=int, default=0,
                        help="Лимит OEM-артикулов за сессию (0=все непроверенные)")
    parser.add_argument("--match-only",   action="store_true",
                        help="Алиас --method match")
    parser.add_argument("--force-brands", action="store_true",
                        help="Перезапустить поиск даже для уже проверенных брендов")
    parser.add_argument("--proxy",        type=str, default=None,
                        help="HTTP/SOCKS5 прокси для М2 WB-запросов: http://host:port")
    parser.add_argument("--verify-oem",  action="store_true",
                        help="OEM-верификация brand_name матчей через карточки WB")
    parser.add_argument("--top-verify",  type=int, default=0,
                        help="Лимит пар для OEM-верификации (0=все непроверенные)")
    parser.add_argument("--no-export",   action="store_true")
    args = parser.parse_args()

    if args.match_only:
        args.method = "match"
    if args.proxy:
        _PROXY = args.proxy
        log.info(f"Прокси для WB (М2): {_PROXY}")

    conn     = get_db()
    mikado   = load_mikado()
    al_items = load_autoliga()

    # ── М1: MPStats Brand Scan ────────────────────────────────────────────────
    if args.method in ("all", "brand"):
        log.info("=" * 65)
        log.info("МЕТОД 1: MPStats Brand Scan")
        log.info(f"  Период: {MP_D1} — {MP_D2}")
        log.info("=" * 65)

        brands = collect_brands(mikado, al_items, args.top_brands)
        new_n  = run_brand_search(brands, conn, force=args.force_brands)
        total  = conn.execute("SELECT COUNT(*) FROM wb_products").fetchone()[0]
        log.info(f"М1 готово: +{new_n} записей → итого в БД: {total:,}")

    # ── М2: OEM Exactmatch ────────────────────────────────────────────────────
    if args.method in ("all", "oem"):
        log.info("=" * 65)
        log.info("МЕТОД 2: OEM Exactmatch (WB search, Автолига)")
        if _PROXY:
            log.info(f"  Прокси: {_PROXY}")
        log.info("=" * 65)
        if not CURL_OK:
            log.warning("curl_cffi не установлен — риск блокировки. pip install curl-cffi")
        if al_items:
            run_oem_search(al_items, conn, limit_n=args.top_oem)

    # ── М3: Brand+Name Match ──────────────────────────────────────────────────
    if args.method in ("all", "match"):
        log.info("=" * 65)
        log.info("МЕТОД 3: Brand+Name Matching (из SQLite)")
        log.info("=" * 65)
        run_match(mikado, al_items, conn)

    # ── OEM-верификация ───────────────────────────────────────────────────────
    if args.verify_oem:
        log.info("=" * 65)
        log.info("OEM-ВЕРИФИКАЦИЯ: карточки WB")
        log.info("=" * 65)
        run_oem_verify(al_items, conn, limit=args.top_verify)

    # ── Итоговая статистика ───────────────────────────────────────────────────
    total_arts  = conn.execute("SELECT COUNT(DISTINCT our_article) FROM wb_matches").fetchone()[0]
    total_pairs = conn.execute("SELECT COUNT(*) FROM wb_matches").fetchone()[0]
    oem_arts    = conn.execute("SELECT COUNT(*) FROM wb_matches WHERE method='oem'").fetchone()[0]
    brand_arts  = conn.execute("SELECT COUNT(*) FROM wb_matches WHERE method='brand_name'").fetchone()[0]
    wb_total    = conn.execute("SELECT COUNT(*) FROM wb_products").fetchone()[0]

    log.info("\n" + "=" * 65)
    log.info("ИТОГ:")
    log.info(f"  WB-товаров в индексе:        {wb_total:,}")
    log.info(f"  Артикулов с матчем:          {total_arts:,}")
    log.info(f"    OEM (точные):              {oem_arts:,}")
    log.info(f"    brand+name:                {brand_arts:,}")
    log.info(f"  Всего пар артикул→nm_id:     {total_pairs:,}")
    log.info("=" * 65)

    top = conn.execute("""
        SELECT m.our_brand, m.our_article, m.method, m.score,
               p.name, p.price_rub, p.subject, p.sales_30d, p.oos_pct
        FROM wb_matches m
        LEFT JOIN wb_products p ON m.nm_id = p.nm_id
        ORDER BY m.score DESC
        LIMIT 20
    """).fetchall()
    if top:
        w = 135
        print(f"\n{'─'*w}")
        print(f"{'Бренд':14} {'Артикул':16} {'Метод':10} {'Скор':5}  "
              f"{'WB Название':38} {'Ниша':20} {'Цена':7} {'Прод':6} {'OOS':5}")
        print(f"{'─'*w}")
        for brand, art, method, score, wb_name, wb_price, subject, sales, oos in top:
            print(
                f"{(brand or '')[:13]:14} {art[:15]:16} {(method or '')[:9]:10} "
                f"{(score or 0):4.2f}  {(wb_name or '')[:37]:38} "
                f"{(subject or '')[:19]:20} {(wb_price or 0):6.0f}₽ "
                f"{(sales or 0):5.0f} {(oos or 0):4.0f}%"
            )
        print(f"{'─'*w}\n")

    if not args.no_export:
        export_excel(conn)

    conn.close()


if __name__ == "__main__":
    main()
