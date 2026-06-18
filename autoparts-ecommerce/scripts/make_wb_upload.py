"""
make_wb_upload.py
Создаёт файл "Топ-500 ВБ.xlsx" в формате образца "Барабан тормозной.xlsx".
Источник данных: wb_top500_combined.xlsx
Колонки образца (Склады → Поставщик):
  1  Код / Артикул
  2  Наименование
  3  Бренд
  4  Цена закупки, руб.
  5  Поставщик          ← бывший "Склады"
  6  Всего, шт.
  7  Параметры          ← цена продажи WB + маржа
  8  OEM номера
  9  Применяемость      ← Авто в парке
  10 Альтернативные артикулы товара
  11 Описание           ← подробное обоснование (аналитика)
  12 Вес, г
  13 Длина, мм
  14 Ширина, мм
  15 Высота, мм
  16 Изображения
"""

import os, sys, io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SRC_COMBINED = r"C:\Users\Admin\Documents\Autoparts_Ecommerce\data\analytics\wb_top500_combined.xlsx"
OUT_FILE     = r"C:\Users\Admin\Desktop\Топ-500 ВБ\Топ-500 ВБ.xlsx"

# ── Структура колонок ──────────────────────────────────────────────────────────
HEADERS = [
    ("Код / Артикул",                    22),
    ("Наименование",                     55),
    ("Бренд",                            18),
    ("Цена закупки, руб.",               16),
    ("Поставщик",                        14),
    ("Параметры",                        30),
    ("OEM номера",                       20),
    ("Применяемость",                    35),
    ("Альтернативные артикулы товара",   22),
    ("Описание",                         70),
    ("Вес, г",                           10),
    ("Длина, мм",                        12),
    ("Ширина, мм",                       12),
    ("Высота, мм",                       12),
    ("Изображения",                      20),
]

# Цвета из образца
C_DARK   = "1A1A2E"   # тёмно-синий (Описание)
C_BLUE   = "1A3A5C"   # синий (размеры)
C_WHITE  = "FFFFFF"
C_HEADER = "2E4057"   # для обычных шапок
C_AL     = "E3F2FD"   # голубой — Автолига
C_MK     = "E8F5E9"   # зелёный — Микадо
C_BOTH   = "FFF9C4"   # жёлтый — оба

THIN = Border(
    left  =Side(style='thin', color='CCCCCC'),
    right =Side(style='thin', color='CCCCCC'),
    top   =Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def header_style(col_idx):
    """Возвращает (bold, bg, fg, size) для шапки по номеру колонки (1-based)."""
    if col_idx == 11:                 # Описание — тёмно-синий
        return True, C_DARK, C_WHITE, 10
    elif 12 <= col_idx <= 15:         # Размеры — синий
        return True, C_BLUE, C_WHITE, 10
    else:
        return True, C_HEADER, C_WHITE, 10


def set_header(ws, row, col, value, width):
    bold, bg, fg, size = header_style(col)
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, size=size, color=fg)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    cell.border    = THIN
    ws.column_dimensions[cell.column_letter].width = width
    return cell


def set_data(ws, row, col, value, row_bg=None, wrap=False, fmt=None, align='left'):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                                wrap_text=wrap)
    cell.border    = THIN
    if row_bg:
        cell.fill = PatternFill('solid', fgColor=row_bg)
    if fmt:
        cell.number_format = fmt
    return cell


def source_bg(src: str) -> str | None:
    if src == 'Авт+Мик': return C_BOTH
    if src == 'Микадо':  return C_MK
    if src == 'Автолига': return C_AL
    return None


# Расширенные наименования для коротких названий из Микадо
NAME_OVERRIDES = {
    "caf100493c": "Фильтр воздушный CHAMPION Ford Focus II/III/C-Max/Kuga, Mazda 3/5, Volvo C30/S40/V40/V50 2004–2019",
    "ccf0021c":   "Фильтр салонный угольный CHAMPION Audi Q7 (2006–2015), Porsche Cayenne (2002–2010), VW Touareg/Transporter T5/Amarok/Multivan 2002–2016",
    "caf100689p": "Фильтр воздушный CHAMPION Opel Astra G/H, Vectra B/C, Zafira A/B, Chevrolet Lacetti/Cruze",
    "caf100715p": "Фильтр воздушный CHAMPION VW Polo/Fox/Lupo, Skoda Fabia I/II, Seat Ibiza/Cordoba 1.0–1.4",
    "caf100724p": "Фильтр воздушный CHAMPION Renault Clio II/III/Modus, Dacia Logan/Sandero, Nissan Micra K12/Note",
    "217003705010": "Катушка зажигания Lada Priora ВАЗ-2170/2171/2172 1.6L 16-клапанный",
    "ccf0046":    "Фильтр салонный CHAMPION Renault Kangoo I (1998–2007), Renault Clio II (1998–2005)",
    "ccf0070":    "Фильтр салонный CHAMPION Opel Astra H (2004–2010), Zafira B (2005–2011), Corsa D",
    "ccf0093c":   "Фильтр салонный угольный CHAMPION VW Touareg I (2004–2017), Audi Q7 4L (2007–2015), Porsche Cayenne (2003–2018)",
    "ccf0152":    "Фильтр салонный CHAMPION Ford Fiesta VI (2008–2017) 1.25–1.6/1.4TDCi/1.6TDCi",
    "ccf0153":    "Фильтр салонный CHAMPION Renault Koleos I (2008–2016) 2.0/2.5/2.0dCi",
    "ccf0327":    "Фильтр салонный CHAMPION Ford Mondeo III/IV (2004–2007), Jaguar X-Type (2005–2009)",
    "ccf0417":    "Фильтр салонный CHAMPION к-т 2 шт. VW Touareg I (2004–2010), Audi Q7 4L, Porsche Cayenne 955/957",
    "oe059/t10":  "Свеча зажигания медная CHAMPION L92YC — мотоциклы, ATV, садовая техника (M14×1.25, аналог NGK BP5HS)",
}


def build_params(sell, mg_pct, mg_rub, oos, wb_cnt, niche) -> str:
    parts = []
    if sell:    parts.append(f"Цена продажи WB: {int(sell)} руб.")
    if mg_pct:  parts.append(f"Маржа: {mg_pct:.1f}% ({int(mg_rub)} руб./шт.)")
    if oos:     parts.append(f"OOS конкурентов: {oos:.0f}%")
    if wb_cnt:  parts.append(f"Конкурентов WB: {wb_cnt}")
    if niche:   parts.append(f"Ниша: {niche}")
    return "\n".join(parts)


def main():
    # ── Читаем combined ────────────────────────────────────────────────────────
    print(f"Читаю: {SRC_COMBINED}")
    src_wb = openpyxl.load_workbook(SRC_COMBINED)
    src_ws = src_wb.worksheets[0]

    # Заголовки combined (строка 1):
    # 1=#, 2=Артикул, 3=Бренд, 4=Название, 5=Источник,
    # 6=Закупка, 7=Продажа, 8=Маржа%, 9=Маржа руб, 10=WB продажи,
    # 11=OOS%, 12=Остаток, 13=WB товаров бренда, 14=Авто в парке,
    # 15=Деталь-частота, 16=WB-ниша, 17=Стратегии, 18=Подробное обоснование

    rows = list(src_ws.iter_rows(min_row=2, max_row=501, values_only=True))
    print(f"  Строк данных: {len(rows)}")

    # ── Создаём новый Excel ────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Топ-500 ВБ"
    ws.freeze_panes = "A2"

    # Шапка
    for col_idx, (header, width) in enumerate(HEADERS, 1):
        set_header(ws, 1, col_idx, header, width)
    ws.row_dimensions[1].height = 40

    # Данные
    for i, row in enumerate(rows, 2):
        rank      = row[0]
        article   = row[1]
        brand     = row[2]
        name      = row[3]
        source    = row[4]   # Автолига / Микадо / Авт+Мик
        buy       = row[5]
        sell      = row[6]
        mg_pct    = row[7]
        mg_rub    = row[8]
        avg_sales = row[9]
        oos_pct   = row[10]
        stock     = row[11]
        wb_cnt    = row[12]
        cars      = row[13]  # Авто в парке
        part_freq = row[14]  # Деталь - частота
        niche     = row[15]
        strats    = row[16]
        reason    = row[17]

        # Применяемость — авто из названия
        applicability = str(cars) if cars else ''

        # Расширенное наименование если есть в словаре
        display_name = NAME_OVERRIDES.get(str(article).strip().lower(), name)

        values = [
            article,         # 1  Код / Артикул
            display_name,    # 2  Наименование
            brand,           # 3  Бренд
            buy,             # 4  Цена закупки
            source,          # 5  Поставщик
            None,            # 6  Параметры — очищено
            None,            # 7  OEM номера — нет данных
            applicability,   # 8  Применяемость
            None,            # 9  Альтернативные артикулы — нет данных
            None,            # 10 Описание — очищено
            None,            # 11 Вес, г
            None,            # 12 Длина, мм
            None,            # 13 Ширина, мм
            None,            # 14 Высота, мм
            None,            # 15 Изображения
        ]

        for col_idx, val in enumerate(values, 1):
            wrap  = col_idx in (2, 8)
            align = 'center' if col_idx in (4, 5, 11, 12, 13, 14) else 'left'
            fmt   = '#,##0.00' if col_idx == 4 else None
            set_data(ws, i, col_idx, val, row_bg=None, wrap=wrap,
                     fmt=fmt, align=align)

        ws.row_dimensions[i].height = 18

    # Сохраняем
    os.makedirs(os.path.dirname(OUT_FILE), exist_ok=True)
    wb.save(OUT_FILE)
    size = os.path.getsize(OUT_FILE) // 1024
    print(f"Сохранено: {OUT_FILE}  ({size} KB)")
    print(f"Строк данных: {len(rows)}")

    # Статистика
    from collections import Counter
    src_stat = Counter(str(r[4]) for r in rows if r[4])
    print("По поставщикам:")
    for src, n in src_stat.most_common():
        print(f"  {src}: {n}")


if __name__ == '__main__':
    main()
