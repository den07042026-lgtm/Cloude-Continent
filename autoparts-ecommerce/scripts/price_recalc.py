"""
price_recalc.py
═══════════════════════════════════════════════════════════════════════════════
Ежедневный пересчёт цен в 00:01.

Алгоритм:
  1. Загружает актуальный прайс Mikado (закупочные цены)
  2. Загружает габариты товаров из scraper_output/mikado_data.xlsx
  3. Получает текущие цены из Ozon API (/v4/product/info/prices)
  4. Для каждого offer_id в Ozon (формат <code>-con):
     - Находит закупочную цену из прайса Mikado
     - Считает логистику FBS по габаритам (или дефолт 115 ₽)
     - Применяет ценовую политику «Оптимум»:
         Целевая наценка на себестоимость (тиер по закупке):
           < 500 ₽  →  25%
           < 1200 ₽ →  20%
           < 2500 ₽ →  17%
           < 3500 ₽ →  15%
           ≥ 3500 ₽ →  12%
         Минимальная цена при которой profit/(purchase+20) ≥ target
  5. Применяет защиты: не снижать > 30% за раз, не ставить маржу < 5%
  6. Обновляет цены в Ozon через /v1/product/import/prices
  7. Telegram-отчёт

Переменные .env:
    OZON_CLIENT_ID=...
    OZON_API_KEY=...
    TG_BOT_TOKEN=...
    TG_CHAT_ID=...

Запуск (непрерывный демон, срабатывает в 00:01 каждый день):
  uv run --with requests,openpyxl scripts/price_recalc.py

Разовый пересчёт:
  uv run --with requests,openpyxl scripts/price_recalc.py --once
  uv run --with requests,openpyxl scripts/price_recalc.py --once --dry-run
"""

import sys
import io
import math
import json
import time
import logging
import argparse
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
except ImportError:
    print("Установи зависимости: uv run --with requests,openpyxl scripts/price_recalc.py")
    sys.exit(1)

try:
    from telegram_notify import tg_price_done, tg_alert
    _TG_OK = True
except ImportError:
    _TG_OK = False

try:
    from daemon_guard import single_instance
except ImportError:
    def single_instance(_): pass

try:
    from code_aliases import CODE_ALIASES
except ImportError:
    CODE_ALIASES = {}

from mikado_price_fetcher import download_mikado_price_bytes
from autoliga_loader import load_autoliga

# ─── Константы ────────────────────────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent.parent
ENV_FILE       = BASE_DIR / ".env"
LOG_FILE       = BASE_DIR / "logs" / "price_recalc_v2.log"
SCRAPER_DATA   = BASE_DIR / "data" / "suppliers" / "mikado" / "scraper_output" / "mikado_data.xlsx"
PRICE_LOG_FILE = BASE_DIR / "data" / "price_recalc_last.json"

MIKADO_PRICE_URL = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=YOUR_MIKADO_PRICE_KEY"
)
OZON_API_BASE = "https://api-seller.ozon.ru"
BATCH_SIZE    = 100

# ─── Защитные ограничения ─────────────────────────────────────────────────────
MIN_MARGIN_FLOOR = 0.05   # не ставить цену если маржа (profit/sell) ниже 5%

SKIP_CODES: frozenset[str] = frozenset()

# ─── Ценовая политика «Оптимум» ───────────────────────────────────────────────
FBS_TIERS = [100, 300, 1500, 5000, 10000]
FBS_RATES = [0.14, 0.20, 0.44, 0.44, 0.44, 0.44]

ACQ_PCT   = 0.015   # эквайринг
TAX_PCT   = 0.06    # УСН 6%
RET_RATE  = 0.03    # % возвратов
REVERSE   = 80      # обратная логистика, ₽
OTHER     = 20      # упаковка/этикетки, ₽ (соответствует ozon_calculator.html)
DEFAULT_LOGISTICS = 115  # ₽ для товаров без габаритов

LOG_FBS = [
    (0.5, 75), (1, 90), (2, 115), (5, 155), (10, 210),
    (15, 265), (20, 315), (25, 365), (30, 420), (50, 620),
]

# ─── Логирование ──────────────────────────────────────────────────────────────
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ─── .env ─────────────────────────────────────────────────────────────────────
def load_env() -> dict:
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


# ─── Ценовые формулы (Оптимум) ────────────────────────────────────────────────
def get_optimal_markup(purchase: float) -> float:
    if purchase < 500:  return 0.25
    if purchase < 1200: return 0.20
    if purchase < 2500: return 0.17
    if purchase < 3500: return 0.15
    return 0.12


def _fbs_rate(sell: float) -> float:
    for thresh, rate in zip(FBS_TIERS, FBS_RATES):
        if sell < thresh:
            return rate
    return FBS_RATES[-1]


def _log_cost(weight_kg: float) -> float:
    for lim, cost in LOG_FBS:
        if weight_kg <= lim:
            return cost
    return LOG_FBS[-1][1] + math.ceil(weight_kg - LOG_FBS[-1][0]) * 15


def calc_logistics(weight_g: float, length_mm: float, width_mm: float, height_mm: float) -> float:
    actual_kg = weight_g / 1000
    vol_kg    = length_mm * width_mm * height_mm / 5_000_000
    return _log_cost(max(actual_kg, vol_kg))


def calc_profit(purchase: float, sell: float, logistics: float) -> float:
    commission  = sell * _fbs_rate(sell)
    acquiring   = sell * ACQ_PCT
    return_loss = RET_RATE * (logistics + REVERSE)
    proceeds    = sell - commission - acquiring - logistics
    tax         = max(0.0, proceeds) * TAX_PCT
    total_cost  = purchase + commission + acquiring + logistics + return_loss + OTHER + tax
    return sell - total_cost


def find_rec_price(purchase: float, logistics: float) -> int | None:
    """Оптимум: минимальная цена при которой profit/(purchase+OTHER) >= target."""
    target = get_optimal_markup(purchase)
    cost   = purchase + OTHER
    for s in range(50, 500_001):
        profit = calc_profit(purchase, s, logistics)
        if profit / cost >= target - 1e-6:
            return s
    return None


# ─── Загрузка прайса Mikado ───────────────────────────────────────────────────
def load_mikado_price() -> dict[str, float]:
    """Загружает прайс. Возвращает {code_lower: purchase_price}."""
    content = download_mikado_price_bytes(MIKADO_PRICE_URL, log)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    ws     = wb.active
    rows   = ws.iter_rows(values_only=True)
    header = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]

    code_idx = price_idx = prodnum_idx = None
    for i, h in enumerate(header):
        if h == "code":       code_idx    = i
        elif h == "priceout": price_idx   = i
        elif h == "prodnum":  prodnum_idx = i

    if code_idx is None:
        wb.close()
        return {}

    raw_db: dict[str, list[float]] = {}
    for row in rows:
        raw = row[code_idx] if len(row) > code_idx else None
        if not raw:
            continue
        code  = str(raw).strip().lower()
        price = 0.0
        if price_idx is not None and len(row) > price_idx:
            try:    price = float(str(row[price_idx] or 0))
            except: pass
        if price > 0:
            raw_db.setdefault(code, []).append(price)
            # Также индексируем по Prodnum (уникален — нужен для CODE_ALIASES)
            if prodnum_idx is not None and len(row) > prodnum_idx and row[prodnum_idx]:
                prodnum = str(row[prodnum_idx]).strip().lower()
                if prodnum and prodnum != code:
                    raw_db.setdefault(prodnum, []).append(price)

    db: dict[str, float] = {}
    unsafe: list[str] = []
    for code, prices in raw_db.items():
        if len({round(p, 2) for p in prices}) == 1:
            db[code] = prices[0]
        else:
            unsafe.append(code)
            log.warning(f"  Mikado дубль: '{code}' — конфликт цен {prices} → исключён")
    if unsafe:
        log.warning(f"Mikado: {len(unsafe)} кодов с конфликтующими дублями исключены")

    wb.close()
    log.info(f"Mikado: цены загружены — {len(db)} позиций с ценой")
    return db


def _price_key(value: str) -> str:
    """Нормализованный артикул для сопоставления прайсов с offer_id Ozon."""
    return "".join(ch for ch in str(value).lower().strip() if ch.isalnum())


def load_supplier_prices() -> dict[str, float]:
    """Объединяет Микадо и Автолигу; при совпадении берёт меньшую закупку."""
    combined: dict[str, float] = {}

    def add(key: str, price: float) -> None:
        if not key or price <= 0:
            return
        for candidate in {str(key).lower().strip(), _price_key(key)}:
            if candidate:
                combined[candidate] = min(price, combined.get(candidate, price))

    mikado = load_mikado_price()
    for article, price in mikado.items():
        add(article, price)

    autoliga = load_autoliga()
    for item in autoliga.values():
        add(item.get("article", ""), float(item.get("price", 0) or 0))

    log.info(
        f"Единый прайс: Микадо={len(mikado)}; Автолига={len(autoliga)}; "
        f"индекс={len(combined)} ключей"
    )
    return combined


# ─── Загрузка габаритов из scraper_output ─────────────────────────────────────
def load_product_dims() -> dict[str, dict]:
    dims: dict[str, dict] = {}
    if not SCRAPER_DATA.exists():
        log.warning(f"Габариты: файл не найден {SCRAPER_DATA}")
        return dims

    try:
        wb = openpyxl.load_workbook(SCRAPER_DATA, read_only=True, data_only=True)
        ws = wb.active
        rows   = ws.iter_rows(values_only=True)
        header = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]

        idx = {}
        for kw, cols in [
            ("code",   ["код", "code", "артикул"]),
            ("weight", ["вес"]),
            ("length", ["длина"]),
            ("width",  ["ширина"]),
            ("height", ["высота"]),
        ]:
            for i, h in enumerate(header):
                if any(c in h for c in cols):
                    idx[kw] = i
                    break

        for row in rows:
            code = str(row[idx["code"]]).strip() if "code" in idx and len(row) > idx["code"] else ""
            if not code or code.lower() == "none":
                continue
            try:
                w = float(row[idx["weight"]] or 0) if "weight" in idx else 0
                l = float(row[idx["length"]] or 0) if "length" in idx else 0
                s = float(row[idx["width"]]  or 0) if "width"  in idx else 0
                h = float(row[idx["height"]] or 0) if "height" in idx else 0
                if all(v > 0 for v in (w, l, s, h)):
                    dims[code] = {"weight": w, "length": l, "width": s, "height": h}
            except Exception:
                pass

        wb.close()
        log.info(f"Габариты: загружено {len(dims)} позиций")
    except Exception as e:
        log.error(f"Габариты: ошибка чтения {SCRAPER_DATA}: {e}")

    return dims


# ─── Ozon API ─────────────────────────────────────────────────────────────────
def _ozon_headers(client_id: str, api_key: str) -> dict:
    return {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
    }


def get_ozon_prices(client_id: str, api_key: str) -> dict[str, int]:
    """
    Возвращает {offer_id: current_price_rub} для всех товаров Ozon.
    Использует /v5/product/info/prices (пагинация через cursor).
    """
    if not client_id or not api_key:
        log.warning("OZON credentials не заданы — текущие цены не загружены")
        return {}

    result: dict[str, int] = {}
    cursor = ""

    while True:
        try:
            body: dict = {
                "filter": {"visibility": "ALL"},
                "limit": 1000,
            }
            if cursor:
                body["cursor"] = cursor

            resp = requests.post(
                f"{OZON_API_BASE}/v5/product/info/prices",
                headers=_ozon_headers(client_id, api_key),
                json=body,
                timeout=30,
            )
            resp.raise_for_status()
            data  = resp.json()
            items = data.get("items", [])
            for item in items:
                offer_id  = item.get("offer_id", "")
                price_str = item.get("price", {}).get("price", "0")
                try:
                    result[offer_id] = int(float(price_str))
                except (ValueError, TypeError):
                    pass
            cursor = data.get("cursor", "")
            if not cursor or len(items) < 1000:
                break
        except Exception as e:
            log.error(f"Ozon: ошибка загрузки цен: {e}")
            break

    log.info(f"Ozon: загружено {len(result)} текущих цен")
    return result


def update_ozon_prices(
    client_id: str,
    api_key:   str,
    prices:    list[dict],  # [{offer_id, new_price}]
    dry_run:   bool = False,
) -> tuple[int, int]:
    """Обновляет цены через /v1/product/import/prices. Возвращает (ok, fail)."""
    if not client_id or not api_key:
        log.warning("OZON_CLIENT_ID / OZON_API_KEY не заданы")
        return 0, 0
    if dry_run:
        log.info(f"[DRY-RUN] Ozon: обновилось бы {len(prices)} цен")
        return len(prices), 0

    ok = fail = 0
    for i in range(0, len(prices), BATCH_SIZE):
        chunk = prices[i : i + BATCH_SIZE]
        payload = {
            "prices": [
                {
                    "offer_id":  p["offer_id"],
                    "price":     str(p["new_price"]),
                    "old_price": "0",
                    "min_price": "0",
                }
                for p in chunk
            ]
        }
        try:
            resp = requests.post(
                f"{OZON_API_BASE}/v1/product/import/prices",
                headers=_ozon_headers(client_id, api_key),
                json=payload,
                timeout=60,
            )
            resp.raise_for_status()
            result = resp.json().get("result", [])
            errs   = [r for r in result if not r.get("updated")]
            ok    += len(chunk) - len(errs)
            fail  += len(errs)
            for e in errs[:3]:
                log.warning(f"  Ozon price [{e.get('offer_id')}]: {e.get('errors')}")
        except Exception as e:
            log.error(f"Ozon: ошибка батча цен [{i // BATCH_SIZE + 1}]: {e}")
            fail += len(chunk)

    return ok, fail


# ─── Один пересчёт ────────────────────────────────────────────────────────────
def recalc_once(env: dict, dry_run: bool = False, allow_decrease: bool = False) -> None:
    log.info("═" * 55)
    log.info(f"Пересчёт цен {'[DRY-RUN] ' if dry_run else ''}(Оптимум) — старт")

    client_id = env.get("OZON_CLIENT_ID", "")
    api_key   = env.get("OZON_API_KEY", "")
    if not client_id or not api_key:
        log.error("OZON_CLIENT_ID / OZON_API_KEY не заданы — выход")
        return

    # 1. Загружаем данные
    price_db = load_supplier_prices()
    dims_db  = load_product_dims()
    if not price_db:
        log.error("Прайс Mikado пуст — пересчёт отменён")
        return

    # 2. Текущие цены Ozon (нужны для защиты от резкого снижения)
    current_prices = get_ozon_prices(client_id, api_key)

    # 3. Рассчитываем новые цены
    pending:   list[dict] = []
    skipped = unchanged = 0

    for offer_id, cur_price in current_prices.items():
        mk_code = offer_id.removesuffix("-con")
        mk_key  = mk_code.lower()

        if mk_key in SKIP_CODES:
            skipped += 1
            continue

        # Жёсткая привязка: если для этого кода задан конкретный Prodnum — используем его
        lookup_key = CODE_ALIASES.get(mk_key, mk_key)

        purchase = price_db.get(lookup_key) or price_db.get(_price_key(lookup_key))
        if not purchase:
            skipped += 1
            continue

        dims = dims_db.get(mk_code)
        if dims:
            logistics = calc_logistics(dims["weight"], dims["length"], dims["width"], dims["height"])
        else:
            logistics = DEFAULT_LOGISTICS

        new_price = find_rec_price(purchase, logistics)
        if new_price is None:
            alias_note = f" [алиас: {lookup_key}]" if lookup_key != mk_key else ""
            log.warning(f"  {mk_code}{alias_note}: не удалось подобрать цену (закупка={purchase:.0f} ₽)")
            skipped += 1
            continue

        # Защита 1: маржа не ниже минимального порога
        margin = calc_profit(purchase, new_price, logistics) / new_price
        if margin < MIN_MARGIN_FLOOR:
            log.warning(f"  {mk_code}: маржа {margin*100:.1f}% < {MIN_MARGIN_FLOOR*100:.0f}% — пропущен")
            skipped += 1
            continue

        # Защита 2: никогда не снижать цену автоматически
        if not allow_decrease and cur_price > 0 and new_price < cur_price:
            log.info(
                f"  {mk_code}: цена не снижается ({cur_price} → {new_price} ₽) — пропущен"
            )
            skipped += 1
            continue

        if cur_price == new_price:
            unchanged += 1
            continue

        markup     = get_optimal_markup(purchase)
        profit_val = calc_profit(purchase, new_price, logistics)
        alias_note = f" [{lookup_key}]" if lookup_key != mk_key else ""
        log.info(
            f"  {mk_code:<16}{alias_note}  закупка={purchase:.0f} ₽  "
            f"лог={logistics:.0f} ₽  наценка={markup*100:.0f}%  "
            f"цена: {cur_price} → {new_price} ₽  прибыль={profit_val:.0f} ₽"
        )
        pending.append({"offer_id": offer_id, "new_price": new_price})

    # 4. Отправляем в Ozon
    updated = fail = 0
    if pending:
        updated, fail = update_ozon_prices(client_id, api_key, pending, dry_run=dry_run)
        skipped += fail
        log.info(f"Ozon: обновлено {updated}, ошибок {fail}")
    else:
        log.info("Нет изменений для отправки")

    if dry_run:
        updated = len(pending)

    log.info(
        f"Пересчёт завершён: обновлено={updated}  "
        f"без изменений={unchanged}  пропущено={skipped}"
    )

    # Сохраняем лог последнего пересчёта
    try:
        PRICE_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
        PRICE_LOG_FILE.write_text(
            json.dumps({
                "ts":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "updated":   updated,
                "skipped":   skipped,
                "unchanged": unchanged,
            }, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception:
        pass

    if _TG_OK and env.get("TG_BOT_TOKEN") and not dry_run:
        tg_price_done(env["TG_BOT_TOKEN"], env.get("TG_CHAT_ID", ""), updated, skipped)


# ─── Планировщик: 01:10, 07:10, 13:10, 19:10 ─────────────────────────────────
# Сдвинуто на 1 час раньше исходной сетки (02/08/14/20) — синхронизировано
# со сдвигом ozon_stock_sync.py, чтобы обойти окно ~08:00-08:15.
_RECALC_HOURS = [1, 7, 13, 19]
_RECALC_MIN   = 10


def _seconds_until_next_slot() -> float:
    now     = datetime.now()
    cur_min = now.hour * 60 + now.minute
    for h in _RECALC_HOURS:
        slot_min = h * 60 + _RECALC_MIN
        if slot_min > cur_min:
            return (slot_min - cur_min) * 60 - now.second
    first_tomorrow = _RECALC_HOURS[0] * 60 + _RECALC_MIN
    return (24 * 60 - cur_min + first_tomorrow) * 60 - now.second


# ─── Точка входа ──────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(description="Пересчёт цен Mikado → Ozon (Оптимум) каждые 6 часов")
    parser.add_argument("--once",    action="store_true", help="Запустить один раз сейчас")
    parser.add_argument("--dry-run", action="store_true", help="Без записи в Ozon")
    parser.add_argument(
        "--force-decrease", action="store_true",
        help="Разрешить снижение цен при разовом принудительном пересчёте",
    )
    args = parser.parse_args()

    single_instance(__file__)
    env = load_env()

    if args.once or args.dry_run:
        recalc_once(env, dry_run=args.dry_run, allow_decrease=args.force_decrease)
        return

    slots_str = ", ".join(f"{h:02d}:{_RECALC_MIN:02d}" for h in _RECALC_HOURS)
    log.info(f"Планировщик цен запущен: пересчёт в {slots_str}")
    while True:
        wait = _seconds_until_next_slot()
        next_run = (datetime.now() + timedelta(seconds=wait)).strftime("%d.%m %H:%M")
        log.info(f"Следующий пересчёт в {next_run} (через {wait / 3600:.1f} ч)")
        time.sleep(wait)
        try:
            recalc_once(env)
        except Exception:
            log.exception("Необработанная ошибка в пересчёте цен")
        time.sleep(60)


if __name__ == "__main__":
    main()
