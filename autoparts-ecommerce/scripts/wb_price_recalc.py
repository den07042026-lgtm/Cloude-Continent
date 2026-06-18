"""
wb_price_recalc.py
═══════════════════════════════════════════════════════════════════════════════
Ежедневный пересчёт цен WB в 00:01.

Алгоритм:
  1. Загружает актуальный прайс Mikado (закупочные цены)
  2. Загружает габариты из scraper_output/mikado_data.xlsx
  3. Получает все товары WB (nmID + vendorCode + текущая цена)
  4. Для каждого товара с известной закупочной ценой:
     - Считает объём в литрах по габаритам (или DEFAULT_VOLUME)
     - Считает логистику FBS по тарифу склада Волгоград
     - Находит минимальную цену при маржа ≥ TARGET_MARGIN (12%)
     - Обновляет цену на WB через POST /api/v2/upload/task
  5. Telegram-отчёт

Тарифы WB FBS (склад Волгоград, актуально на 2026-05-19):
  Доставка:  50.6₽ + 15.4₽/литр
  Возврат:   136₽  + 14₽/литр (через ПВЗ)
  Комиссия:  25% (Автозапчасти)
  УСН:       6%
  Возвратность: 3%

Переменные .env:
    WB_API_KEY=...
    TG_BOT_TOKEN=...
    TG_CHAT_ID=...

Запуск (демон, срабатывает в 00:01):
  uv run --with requests,openpyxl scripts/wb_price_recalc.py

Разовый пересчёт:
  uv run --with requests,openpyxl scripts/wb_price_recalc.py --once
  uv run --with requests,openpyxl scripts/wb_price_recalc.py --once --dry-run
"""

import sys
import io
import json
import math
import time
import logging
import argparse
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
    import openpyxl
except ImportError:
    print("Установи зависимости: uv run --with requests,openpyxl scripts/wb_price_recalc.py")
    sys.exit(1)

try:
    from telegram_notify import tg_price_done, tg_alert
    _TG_OK = True
except ImportError:
    _TG_OK = False

# ─── Константы ────────────────────────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent.parent
ENV_FILE       = BASE_DIR / ".env"
LOG_FILE       = BASE_DIR / "logs" / "wb_price_recalc.log"
SCRAPER_DATA   = BASE_DIR / "data" / "suppliers" / "mikado" / "scraper_output" / "mikado_data.xlsx"
PRICE_FALLBACK = Path("C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx")
PRICE_LOG_FILE = BASE_DIR / "data" / "wb_price_recalc_last.json"

MIKADO_PRICE_URL = (
    "https://mikado-parts.ru/api/Price/GetPriceExcel"
    "?StockId=34&Key=BBE2E029-54CF-4D9E-9FAC-9FE25E85B300"
)

WB_PRICES_BASE = "https://discounts-prices-api.wildberries.ru"
BATCH_SIZE     = 1000   # WB принимает до 1000 позиций за раз

# ─── Тарифы WB FBS (склад Волгоград, 2026-05-19) ─────────────────────────────
WB_COMMISSION    = 0.25    # Автозапчасти FBS
TAX_PCT          = 0.06    # УСН 6%
RET_RATE         = 0.03    # доля возвратов 3%
PACKAGING        = 30      # упаковка, ₽
TARGET_MARGIN    = 0.12    # целевая маржа 12%

# Доставка до покупателя: база + литры × ставка
DELIVERY_BASE    = 50.6
DELIVERY_LITER   = 15.4

# Обратная логистика (возврат через ПВЗ): база + литры × ставка
RETURN_BASE      = 136.0
RETURN_LITER     = 14.0

# Дефолтный объём если нет габаритов (~типичная автозапчасть 20×15×10 см = 3 л)
DEFAULT_VOLUME   = 3.0

# Не снижать цену более чем на 30% за один пересчёт
PRICE_DROP_LIMIT = 0.70

# Коды (строчными), исключённые из автопересчёта (ложные совпадения с Mikado)
SKIP_CODES: frozenset[str] = frozenset({
    "gf-1904",  # Mikado: газовая пружина багажника Citroen C4 (не наш товар)
})

# ─── Логирование ──────────────────────────────────────────────────────────────
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ─── .env ─────────────────────────────────────────────────────────────────────
def load_env() -> dict:
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


# ─── Ценовые формулы ──────────────────────────────────────────────────────────
def _volume_liters(length_mm: float, width_mm: float, height_mm: float) -> float:
    """Объём коробки в литрах. Минимум 1 литр."""
    return max(1.0, length_mm * width_mm * height_mm / 1_000_000)


def _delivery_cost(liters: float) -> float:
    return DELIVERY_BASE + liters * DELIVERY_LITER


def _return_cost(liters: float) -> float:
    return RET_RATE * (RETURN_BASE + liters * RETURN_LITER)


def calc_profit(purchase: float, sell: float, liters: float) -> float:
    commission  = sell * WB_COMMISSION
    delivery    = _delivery_cost(liters)
    ret_cost    = _return_cost(liters)
    proceeds    = sell - commission - delivery
    tax         = max(0.0, proceeds) * TAX_PCT
    return sell - purchase - commission - delivery - ret_cost - tax - PACKAGING


def find_rec_price(purchase: float, liters: float,
                   target: float = TARGET_MARGIN) -> int | None:
    """Минимальная цена продажи при которой маржа ≥ target."""
    for s in range(50, 500_001):
        profit = calc_profit(purchase, s, liters)
        if s > 0 and profit / s >= target - 1e-6:
            return s
    return None


# ─── Загрузка прайса Mikado ───────────────────────────────────────────────────
def load_mikado_price() -> dict[str, float]:
    content = None
    try:
        resp = requests.get(MIKADO_PRICE_URL, timeout=60)
        resp.raise_for_status()
        if resp.content[:2] == b"PK":
            content = resp.content
            log.info(f"Mikado: прайс скачан ({len(content):,} байт)")
    except Exception as e:
        log.warning(f"Mikado онлайн недоступен ({e}), берём локальный")

    if content:
        wb_file = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    elif PRICE_FALLBACK.exists():
        wb_file = openpyxl.load_workbook(PRICE_FALLBACK, read_only=True, data_only=True)
    else:
        log.error("Mikado: прайс недоступен")
        return {}

    ws     = wb_file.active
    rows   = ws.iter_rows(values_only=True)
    header = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]

    code_idx = price_idx = None
    for i, h in enumerate(header):
        if h == "code":       code_idx  = i
        elif h == "priceout": price_idx = i

    raw_db: dict[str, list[float]] = {}
    for row in rows:
        raw = row[code_idx] if code_idx is not None and len(row) > code_idx else None
        if not raw:
            continue
        code  = str(raw).strip().lower()
        price = 0.0
        if price_idx is not None and len(row) > price_idx:
            try: price = float(str(row[price_idx] or 0))
            except: pass
        if price > 0:
            raw_db.setdefault(code, []).append(price)

    db: dict[str, float] = {}
    unsafe: list[str] = []
    for code, prices in raw_db.items():
        if len({round(p, 2) for p in prices}) == 1:
            db[code] = prices[0]
        else:
            unsafe.append(code)
            log.warning(f"  Mikado дубль: '{code}' — конфликт цен {prices} → исключён")
    if unsafe:
        log.warning(f"Mikado: {len(unsafe)} кодов с конфликтующими дублями исключены")

    wb_file.close()
    log.info(f"Mikado: {len(db)} позиций с ценой")
    return db


# ─── Загрузка габаритов ───────────────────────────────────────────────────────
def load_product_dims() -> dict[str, dict]:
    dims: dict[str, dict] = {}
    if not SCRAPER_DATA.exists():
        log.warning(f"Габариты: файл не найден {SCRAPER_DATA}")
        return dims
    try:
        wb_file = openpyxl.load_workbook(SCRAPER_DATA, read_only=True, data_only=True)
        ws      = wb_file.active
        rows    = ws.iter_rows(values_only=True)
        header  = [str(v).strip().lower() if v else "" for v in (next(rows, []) or [])]

        idx = {}
        for kw, cols in [
            ("code",   ["код", "code", "артикул"]),
            ("length", ["длина"]),
            ("width",  ["ширина"]),
            ("height", ["высота"]),
        ]:
            for i, h in enumerate(header):
                if any(c in h for c in cols):
                    idx[kw] = i
                    break

        for row in rows:
            code = str(row[idx["code"]]).strip() if "code" in idx and len(row) > idx["code"] else ""
            if not code or code.lower() == "none":
                continue
            try:
                l = float(row[idx["length"]] or 0) if "length" in idx else 0
                w = float(row[idx["width"]]  or 0) if "width"  in idx else 0
                h = float(row[idx["height"]] or 0) if "height" in idx else 0
                if all(v > 0 for v in (l, w, h)):
                    dims[code] = {"length": l, "width": w, "height": h}
            except Exception:
                pass

        wb_file.close()
        log.info(f"Габариты: {len(dims)} позиций")
    except Exception as e:
        log.error(f"Габариты: ошибка чтения: {e}")
    return dims


# ─── WB API: получение товаров с ценами ──────────────────────────────────────
def get_wb_goods(token: str) -> list[dict]:
    """
    Возвращает все товары WB: [{nmID, vendorCode, current_price}].
    vendorCode = Mikado code (соглашение при загрузке).
    """
    headers = {"Authorization": token, "Content-Type": "application/json"}
    result  = []
    offset  = 0
    limit   = 1000

    while True:
        try:
            r = requests.get(
                f"{WB_PRICES_BASE}/api/v2/list/goods/filter",
                headers=headers,
                params={"limit": limit, "offset": offset},
                timeout=30,
            )
            r.raise_for_status()
            goods = r.json().get("data", {}).get("listGoods", [])
            for g in goods:
                vc    = (g.get("vendorCode") or "").strip()
                nm_id = g.get("nmID")
                sizes = g.get("sizes", [])
                price = sizes[0].get("price", 0) if sizes else 0
                if vc and nm_id:
                    result.append({
                        "nmID":         nm_id,
                        "vendorCode":   vc,
                        "current_price": float(price),
                    })
            if len(goods) < limit:
                break
            offset += limit
        except Exception as e:
            log.error(f"WB: ошибка получения товаров (offset={offset}): {e}")
            break

    log.info(f"WB: загружено {len(result)} товаров")
    return result


# ─── WB API: обновление цен ───────────────────────────────────────────────────
def update_wb_prices(token: str, updates: list[dict]) -> tuple[int, int]:
    """
    updates: [{nmID, new_price}]
    Возвращает (обновлено, ошибок).
    """
    headers = {"Authorization": token, "Content-Type": "application/json"}
    ok = fail = 0

    for i in range(0, len(updates), BATCH_SIZE):
        chunk = updates[i: i + BATCH_SIZE]
        payload = {"data": [{"nmID": u["nmID"], "price": u["new_price"]} for u in chunk]}
        try:
            r = requests.post(
                f"{WB_PRICES_BASE}/api/v2/upload/task",
                headers=headers,
                json=payload,
                timeout=30,
            )
            r.raise_for_status()
            ok += len(chunk)
            log.info(f"WB: батч [{i}:{i+len(chunk)}] принят (taskId={r.json().get('data',{}).get('taskId','?')})")
        except Exception as e:
            log.error(f"WB: ошибка батча [{i}:{i+len(chunk)}]: {e}")
            fail += len(chunk)

    return ok, fail


# ─── Один пересчёт ────────────────────────────────────────────────────────────
def recalc_once(env: dict, dry_run: bool = False) -> None:
    log.info("═" * 55)
    log.info(f"WB пересчёт цен {'[DRY-RUN] ' if dry_run else ''}— старт")

    wb_token = env.get("WB_API_KEY", "")
    if not wb_token:
        log.error("WB_API_KEY не задан — выход")
        return

    # 1. Загружаем данные
    price_db = load_mikado_price()
    dims_db  = load_product_dims()
    if not price_db:
        log.error("Прайс Mikado пуст — пересчёт отменён")
        return

    # 2. Товары WB
    wb_goods = get_wb_goods(wb_token)
    if not wb_goods:
        log.error("WB: товары не загружены")
        return

    # 3. Рассчитываем цены
    pending:   list[dict] = []
    skipped   = unchanged = 0

    for good in wb_goods:
        vc    = good["vendorCode"]
        vc_key = vc.lower()

        if vc_key in SKIP_CODES:
            skipped += 1
            continue

        purchase = price_db.get(vc_key)
        if not purchase:
            skipped += 1
            continue

        dims = dims_db.get(vc_key) or dims_db.get(vc)
        if dims:
            liters = _volume_liters(dims["length"], dims["width"], dims["height"])
        else:
            liters = DEFAULT_VOLUME

        new_price = find_rec_price(purchase, liters)
        if new_price is None:
            log.warning(f"  {vc}: не удалось подобрать цену (закупка={purchase:.0f}₽)")
            skipped += 1
            continue

        # Защита: не снижать цену более чем на 30% за один пересчёт
        cur = good["current_price"]
        if cur > 0 and new_price < cur * PRICE_DROP_LIMIT:
            log.warning(
                f"  {vc}: подозрительное снижение {cur:.0f} → {new_price} ₽ "
                f"(−{(1 - new_price/cur)*100:.0f}%, лимит 30%) — пропущен"
            )
            skipped += 1
            continue

        if int(cur) == new_price:
            unchanged += 1
            continue

        log.info(
            f"  {vc:<16}  закупка={purchase:.0f}₽  "
            f"объём={liters:.1f}л  лог={_delivery_cost(liters):.0f}₽  "
            f"цена: {good['current_price']:.0f} → {new_price}₽"
        )

        if not dry_run:
            pending.append({"nmID": good["nmID"], "new_price": new_price})

    # 4. Отправляем на WB
    updated = 0
    if not dry_run and pending:
        updated, fail = update_wb_prices(wb_token, pending)
        skipped += fail
        log.info(f"WB: обновлено {updated}, ошибок {fail}")
    elif dry_run:
        updated = len(pending)

    log.info(
        f"Пересчёт завершён: обновлено={updated}  "
        f"без изменений={unchanged}  пропущено={skipped}"
    )

    # Сохраняем лог
    try:
        PRICE_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
        PRICE_LOG_FILE.write_text(
            json.dumps({
                "ts":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "updated":   updated,
                "skipped":   skipped,
                "unchanged": unchanged,
            }, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception:
        pass

    # Telegram
    if _TG_OK and env.get("TG_BOT_TOKEN") and not dry_run:
        tg_price_done(env["TG_BOT_TOKEN"], env.get("TG_CHAT_ID", ""), updated, skipped)


# ─── Планировщик ──────────────────────────────────────────────────────────────
def _seconds_until_0001() -> float:
    now   = datetime.now()
    today = now.replace(hour=0, minute=1, second=0, microsecond=0)
    if now >= today:
        today += timedelta(days=1)
    return (today - now).total_seconds()


# ─── Точка входа ──────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(description="Ежедневный пересчёт цен WB")
    parser.add_argument("--once",    action="store_true", help="Запустить один раз сейчас")
    parser.add_argument("--dry-run", action="store_true", help="Без отправки на WB")
    args = parser.parse_args()

    env = load_env()

    if args.once or args.dry_run:
        recalc_once(env, dry_run=args.dry_run)
        return

    log.info("Планировщик WB цен: пересчёт ежедневно в 00:01")
    while True:
        wait    = _seconds_until_0001()
        next_dt = (datetime.now() + timedelta(seconds=wait)).strftime("%d.%m %H:%M")
        log.info(f"Следующий пересчёт в {next_dt} (через {wait/3600:.1f}ч)")
        time.sleep(wait)
        try:
            recalc_once(env)
        except Exception:
            log.exception("Необработанная ошибка в пересчёте цен WB")
        time.sleep(120)


if __name__ == "__main__":
    main()
