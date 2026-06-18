"""
gigachat_dimensions_filler.py
══════════════════════════════
Добавляет столбцы «Вес, г», «Длина, мм», «Ширина, мм», «Высота, мм»
после столбца «Описание» — через браузер giga.chat.
Без API ключа, без лимита сообщений, работает headless.

Исходная папка:    C:\\Users\\Admin\\Desktop\\На сортировку 26.04\\
Папка результатов: C:\\Users\\Admin\\Desktop\\На сортировку 08.05\\

Первый запуск — войти в GigaChat (сессия сохранится):
  uv run --with playwright,openpyxl scripts/gigachat_dimensions_filler.py --login

Обработать все файлы:
  uv run --with playwright,openpyxl scripts/gigachat_dimensions_filler.py

Один файл для теста:
  uv run --with playwright,openpyxl scripts/gigachat_dimensions_filler.py --file "Автобаферы.xlsx" --rows 2-30

Продолжение после прерывания:
  Запустить снова — пропустит строки, которые уже заполнены.
"""

import re
import sys
import json
import time
import shutil
import argparse
import traceback
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Playwright не установлен.")
    print("  uv run --with playwright python -m playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("openpyxl не установлен.")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

SRC_DIR      = Path(r"C:\Users\Admin\Desktop\На сортировку 26.04")
OUT_DIR      = Path(r"C:\Users\Admin\Desktop\На сортировку 08.05")
PROFILE_DIR  = Path.home() / ".gigachat_dimensions_playwright"
GIGACHAT_URL = "https://giga.chat/"

BATCH_SIZE = 30   # строк за один запрос
SAVE_EVERY = 3    # сохранять Excel каждые N батчей (~90 строк)

NEW_COLS  = ["Вес, г", "Длина, мм", "Ширина, мм", "Высота, мм"]
COL_AFTER = "Описание"

# Начало нашего промпта — по нему фильтруем сообщения пользователя
PROMPT_MARKER = "Ты эксперт по автозапчастям"


# ══════════════════════════════════════════════════════════════════════════════
#  ПРОМПТ
# ══════════════════════════════════════════════════════════════════════════════

INSTRUCTION = (
    "Ты эксперт по автозапчастям. Для каждой позиции ниже укажи "
    "вес и габариты ТРАНСПОРТНОЙ УПАКОВКИ — внешние размеры коробки или пакета "
    "вместе с самой деталью внутри, именно так, как товар будет отправлен покупателю.\n"
    "ВАЖНО: это НЕ размеры самой детали, а размеры упаковки снаружи.\n"
    "Правила:\n"
    "- Если деталь идёт комплектом (2 шт. и т.п.) — вес/размер всего комплекта в упаковке\n"
    "- Опирайся на типичные размеры упаковки для данного вида запчасти\n"
    "- Верни ТОЛЬКО JSON объект, без пояснений, без markdown блоков:\n"
    '{"items": [{"idx": 0, "вес": 500, "длина": 200, "ширина": 150, "высота": 80}, ...]}\n'
    "вес — граммы (деталь + упаковка), длина/ширина/высота — мм (внешние стороны упаковки), "
    "всё целые числа.\n\n"
)


def build_prompt(items: list[dict]) -> str:
    lines = [INSTRUCTION]
    for item in items:
        lines.append(f"[{item['idx']}]")
        lines.append(f"Наименование: {item['name']}")
        if item.get("params"):
            lines.append(f"Параметры: {str(item['params'])[:200]}")
        if item.get("desc"):
            lines.append(f"Описание: {str(item['desc'])[:250]}")
        lines.append("")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  ПАРСИНГ ОТВЕТА
# ══════════════════════════════════════════════════════════════════════════════

def extract_json_items(text: str) -> list[dict]:
    clean = re.sub(r"```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    clean = clean.replace("```", "").strip()

    ITEM_KEYS = {"вес", "длина", "ширина", "высота", "idx"}

    def is_valid(lst):
        return bool(lst) and isinstance(lst[0], dict) and bool(ITEM_KEYS & lst[0].keys())

    def try_parse(s):
        try:
            parsed = json.loads(s.strip())
            if isinstance(parsed, dict) and "items" in parsed:
                items = parsed["items"]
                if isinstance(items, list) and is_valid(items):
                    return items
            if isinstance(parsed, list) and is_valid(parsed):
                return parsed
        except Exception:
            pass
        return None

    result = try_parse(clean)
    if result:
        return result

    for start in [i for i, c in enumerate(clean) if c == '{']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == '{':
                depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result:
                        return result
                    break

    for start in [i for i, c in enumerate(clean) if c == '[']:
        depth = 0
        for i, c in enumerate(clean[start:], start):
            if c == '[':
                depth += 1
            elif c == ']':
                depth -= 1
                if depth == 0:
                    result = try_parse(clean[start:i + 1])
                    if result:
                        return result
                    break

    return []


# ══════════════════════════════════════════════════════════════════════════════
#  БРАУЗЕР — GIGACHAT
# ══════════════════════════════════════════════════════════════════════════════

def dismiss_modal(page):
    try:
        page.keyboard.press("Escape")
        time.sleep(0.5)
    except Exception:
        pass
    try:
        backdrop = page.locator('[data-backdrop]').first
        if backdrop.is_visible(timeout=500):
            page.mouse.click(5, 5)
            time.sleep(0.5)
    except Exception:
        pass


def find_input(page):
    time.sleep(0.5)
    for sel in [
        '#chat-input-textarea',
        'textarea[placeholder*="Спрос"]',
        'textarea[placeholder*="Напиш"]',
        'textarea',
    ]:
        try:
            el = page.locator(sel).last
            if el.is_visible(timeout=2000):
                return el
        except Exception:
            continue

    # Диагностика — что сейчас на странице
    try:
        info = page.evaluate("""() => ({
            url:       location.href,
            title:     document.title,
            textareas: document.querySelectorAll('textarea').length,
            inputs:    document.querySelectorAll('input').length,
            buttons:   document.querySelectorAll('button').length,
        })""")
        print(f"\n    Диагностика: {info}")
    except Exception:
        pass
    return None


def type_prompt(page, text: str):
    dismiss_modal(page)
    inp = find_input(page)
    if inp is None:
        raise RuntimeError("Поле ввода GigaChat не найдено")

    page.evaluate("""() => {
        const el = document.querySelector('#chat-input-textarea') ||
                   document.querySelector('textarea');
        if (el) el.focus();
    }""")
    time.sleep(0.3)

    # React-совместимый setter — обновляет внутреннее состояние компонента
    page.evaluate("""(text) => {
        const el = document.querySelector('#chat-input-textarea') ||
                   document.querySelector('textarea');
        if (!el) return;
        const setter = Object.getOwnPropertyDescriptor(
            window.HTMLTextAreaElement.prototype, 'value').set;
        setter.call(el, text);
        el.dispatchEvent(new Event('input',  { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
    }""", text)
    time.sleep(0.5)


def click_send(page):
    for sel in [
        'button[aria-label*="Отправить"]',
        'button[aria-label*="отправить"]',
        'button[type="submit"]',
        '[data-testid="send-button"]',
    ]:
        try:
            btn = page.locator(sel).last
            if btn.is_visible(timeout=800):
                btn.click()
                return
        except Exception:
            continue
    page.locator('#chat-input-textarea').press("Enter")


def get_last_response(page) -> str:
    """Извлекает последний ответ GigaChat, исключая текст пользовательского промпта."""
    try:
        text = page.evaluate("""(marker) => {
            // Собираем весь текст страницы кроме блоков с маркером промпта
            const containers = [...document.querySelectorAll('p, pre, code')];
            const parts = [];
            for (const el of containers) {
                const t = (el.innerText || el.textContent || '').trim();
                if (t.length > 0 && !t.includes(marker)) {
                    parts.push(t);
                }
            }
            return parts.join('\\n');
        }""", PROMPT_MARKER)
        if text and len(text) > 10 and PROMPT_MARKER not in text:
            return text
    except Exception:
        pass
    return ""


def is_generating(page) -> bool:
    try:
        # GigaChat показывает анимацию или кнопку Stop во время генерации
        for sel in [
            '[aria-label*="Стоп"]',
            '[aria-label*="стоп"]',
            '[aria-label*="Stop"]',
            'button[aria-label*="остановить"]',
            '[class*="loading"]',
            '[class*="typing"]',
        ]:
            if page.locator(sel).first.is_visible(timeout=300):
                return True
    except Exception:
        pass
    return False


def wait_for_response(page, timeout_sec: int = 120) -> str:
    print("    жду", end="", flush=True)

    # Ждём появления текста (до 60 сек)
    for _ in range(30):
        candidate = get_last_response(page)
        if candidate and len(candidate) > 10:
            break
        time.sleep(2)
        print(".", end="", flush=True)
    else:
        print(" таймаут (ответ не появился)")
        return ""

    # Ждём стабилизации
    prev, stable = "", 0
    for tick in range(timeout_sec):
        cur = get_last_response(page)
        if cur == prev and cur:
            stable += 1
            if stable >= 2 and not is_generating(page):
                print(f" готово ({len(cur)} симв.)")
                return cur
        else:
            stable = 0
            if tick % 4 == 0:
                print(".", end="", flush=True)
        prev = cur
        time.sleep(1.5)

    print(f" таймаут ({len(prev)} симв.)")
    return prev


def new_chat(page):
    """Открывает новый чат GigaChat."""
    dismiss_modal(page)
    for sel in [
        'a[href="/"][aria-label*="чат"]',
        'button[aria-label*="новый чат"]',
        'button[aria-label*="Новый чат"]',
        '[data-testid="new-chat"]',
        'a[href="/"]',
    ]:
        try:
            btn = page.locator(sel).first
            if btn.is_visible(timeout=800):
                btn.click()
                time.sleep(1.5)
                dismiss_modal(page)
                return
        except Exception:
            continue
    # Fallback: переходим на главную
    try:
        page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=20000)
    except Exception:
        pass
    time.sleep(2)
    dismiss_modal(page)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ХЕЛПЕРЫ
# ══════════════════════════════════════════════════════════════════════════════

def read_headers(ws) -> dict:
    return {
        ws.cell(1, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }


def ensure_new_columns(ws, headers: dict) -> dict:
    existing = set(headers.keys())
    if all(col in existing for col in NEW_COLS):
        return headers

    insert_at = headers.get(COL_AFTER, ws.max_column) + 1
    to_add = [col for col in NEW_COLS if col not in existing]

    ws.insert_cols(insert_at, len(to_add))
    for i, col_name in enumerate(to_add):
        cell = ws.cell(1, insert_at + i, col_name)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1A3A5C")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 14

    return read_headers(ws)


# ══════════════════════════════════════════════════════════════════════════════
#  ОБРАБОТКА ФАЙЛА
# ══════════════════════════════════════════════════════════════════════════════

def process_file(page, src: Path, dst: Path,
                 row_range: tuple | None, debug: bool):
    if not dst.exists():
        shutil.copy2(src, dst)
        print(f"  → Скопирован: {dst.name}")

    wb = openpyxl.load_workbook(dst)
    ws = wb.active

    headers = read_headers(ws)
    headers = ensure_new_columns(ws, headers)

    desc_col  = headers.get(COL_AFTER)
    name_col  = headers.get("Наименование")
    param_col = headers.get("Параметры")
    w_col     = headers.get("Вес, г")
    l_col     = headers.get("Длина, мм")
    wd_col    = headers.get("Ширина, мм")
    h_col     = headers.get("Высота, мм")

    if not desc_col:
        print(f"  ⚠ Колонка «{COL_AFTER}» не найдена — пропуск")
        wb.close()
        return

    todos = []
    for r in range(2, ws.max_row + 1):
        if row_range and not (row_range[0] <= r <= row_range[1]):
            continue
        if not ws.cell(r, desc_col).value:
            continue
        if w_col and ws.cell(r, w_col).value is not None:
            continue
        todos.append(r)

    total = len(todos)
    if not todos:
        print(f"  Нечего обрабатывать")
        wb.close()
        return

    print(f"  Строк для обработки: {total}")

    processed  = 0
    save_timer = 0

    for batch_no, batch_start in enumerate(range(0, total, BATCH_SIZE), start=1):
        batch_rows = todos[batch_start : batch_start + BATCH_SIZE]

        items = []
        for local_idx, r in enumerate(batch_rows):
            items.append({
                "idx":    local_idx,
                "name":   ws.cell(r, name_col).value  if name_col  else "",
                "params": ws.cell(r, param_col).value if param_col else "",
                "desc":   ws.cell(r, desc_col).value  or "",
            })

        print(f"  Батч {batch_no} (стр. {batch_rows[0]}–{batch_rows[-1]})", end="", flush=True)

        new_chat(page)
        prompt = build_prompt(items)

        try:
            type_prompt(page, prompt)
        except RuntimeError as e:
            print(f"\n  ✗ {e}")
            continue

        if debug:
            page.screenshot(path=str(OUT_DIR / f"_debug_b{batch_no}_typed.png"))
            print(f"\n    Нажми Enter для отправки...")
            input("    > ")

        click_send(page)
        time.sleep(1)

        # До 3 попыток получить валидный JSON
        results = []
        for attempt in range(1, 4):
            raw = wait_for_response(page)

            if not raw:
                print(f"  ✗ Пустой ответ (попытка {attempt}/3)")
            else:
                results = extract_json_items(raw)
                if results:
                    break
                print(f"  ⚠ JSON не распознан (попытка {attempt}/3)")
                print(f"    Начало ответа: {raw[:300]}")

            if attempt < 3:
                time.sleep(10)
                new_chat(page)
                type_prompt(page, prompt)
                click_send(page)
                time.sleep(1)

        if not results:
            print(f"  ✗ Батч пропущен после 3 попыток")
            continue

        result_map = {r["idx"]: r for r in results}
        filled = 0
        for local_idx, r in enumerate(batch_rows):
            data = result_map.get(local_idx)
            if not data:
                continue
            if w_col:  ws.cell(r, w_col).value  = data.get("вес")
            if l_col:  ws.cell(r, l_col).value  = data.get("длина")
            if wd_col: ws.cell(r, wd_col).value = data.get("ширина")
            if h_col:  ws.cell(r, h_col).value  = data.get("высота")
            filled += 1

        processed  += filled
        save_timer += 1
        print(f" → заполнено {filled}/{len(batch_rows)}")

        if save_timer >= SAVE_EVERY:
            wb.save(dst)
            print(f"    💾 Сохранено (итого: {processed})")
            save_timer = 0

        time.sleep(2)

    wb.save(dst)
    print(f"  ✓ Итого: {processed}/{total} → {dst.name}")
    wb.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Заполнение веса и габаритов через GigaChat (браузер)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--file",  default=None, help="Один файл (имя xlsx)")
    ap.add_argument("--rows",  default=None, help="Диапазон строк: 2-100")
    ap.add_argument("--login", action="store_true", help="Открыть браузер для входа в GigaChat")
    ap.add_argument("--debug", action="store_true", help="Показывать браузер + скриншоты")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    src_files = [SRC_DIR / args.file] if args.file else sorted(SRC_DIR.glob("*.xlsx"))

    row_range = None
    if args.rows:
        parts = args.rows.split("-")
        row_range = (int(parts[0]), int(parts[1]) if len(parts) > 1 else int(parts[0]))

    print()
    print("═" * 64)
    print("  GigaChat Dimensions Filler")
    print(f"  Батч:    {BATCH_SIZE} строк/запрос")
    print(f"  Файлов:  {len(src_files)}")
    print(f"  Из:      {SRC_DIR}")
    print(f"  В:       {OUT_DIR}")
    print("═" * 64)

    # Запускаем в видимом режиме — GigaChat может блокировать headless
    headless = False

    with sync_playwright() as pw:
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR),
                channel="chrome",
                headless=headless,
                viewport={"width": 1280, "height": 900},
                locale="ru-RU",
                args=["--disable-blink-features=AutomationControlled"],
            )
        except Exception:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=str(PROFILE_DIR),
                headless=headless,
                viewport={"width": 1280, "height": 900},
                locale="ru-RU",
                args=["--disable-blink-features=AutomationControlled"],
            )
        page = context.new_page()

        # ── Режим входа ──────────────────────────────────────────────────────
        if args.login:
            print(f"\n  Открываю браузер GigaChat...")
            try:
                page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=15000)
            except Exception:
                pass
            print(f"  Войди в аккаунт на {GIGACHAT_URL}")
            print("  Когда окажешься в чате — нажми Enter здесь.")
            input("  > ")
            context.storage_state(path=str(PROFILE_DIR / "state.json"))
            print("  ✓ Сессия сохранена.")
            context.close()
            return

        # ── Проверка авторизации ─────────────────────────────────────────────
        print("\n  Открываю GigaChat...")
        try:
            page.goto(GIGACHAT_URL, wait_until="domcontentloaded", timeout=30000)
        except Exception as e:
            print(f"  ✗ Не удалось открыть GigaChat: {e}")
            context.close()
            sys.exit(1)

        time.sleep(3)
        if any(k in page.url.lower() for k in ("login", "signin", "auth", "sso", "id.sber")):
            print("\n  ✗ Не авторизован!")
            print("  Запусти сначала:")
            print("    uv run --with playwright,openpyxl scripts/gigachat_dimensions_filler.py --login")
            context.close()
            sys.exit(1)

        print("  ✓ Авторизован\n")

        # ── Основной цикл ────────────────────────────────────────────────────
        total_files = len(src_files)
        for i, src in enumerate(src_files, 1):
            if not src.exists():
                print(f"[{i:3}/{total_files}] ✗ Не найден: {src.name}")
                continue

            dst = OUT_DIR / src.name
            print(f"[{i:3}/{total_files}] {src.name}")

            try:
                process_file(page, src, dst, row_range, debug=args.debug)
            except KeyboardInterrupt:
                print("\n\n  Прерывание — прогресс сохранён в последнем 💾")
                context.close()
                sys.exit(0)
            except Exception as e:
                print(f"  ✗ Ошибка: {e}")
                traceback.print_exc()

            print()

        context.close()

    print("═" * 64)
    print("  Всё готово!")
    print(f"  Результаты: {OUT_DIR}")
    print("═" * 64)
    print()


if __name__ == "__main__":
    main()
