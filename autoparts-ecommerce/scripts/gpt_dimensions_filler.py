"""
gpt_dimensions_filler.py
════════════════════════
Добавляет столбцы «Вес, г», «Длина, мм», «Ширина, мм», «Высота, мм»
после столбца «Описание» и заполняет их через OpenAI GPT.

Исходная папка:    C:\\Users\\Admin\\Desktop\\На сортировку 26.04\\
Папка результатов: C:\\Users\\Admin\\Desktop\\На сортировку 08.05\\

Настройка API-ключа (один раз в сессии):
  set OPENAI_API_KEY=sk-...

Запуск всех файлов:
  uv run --with openai,openpyxl scripts/gpt_dimensions_filler.py

Только конкретный файл:
  uv run --with openai,openpyxl scripts/gpt_dimensions_filler.py --file "Амортизатор подвески.xlsx"

Конкретный файл + диапазон строк (для теста):
  uv run --with openai,openpyxl scripts/gpt_dimensions_filler.py --file "Амортизатор подвески.xlsx" --rows 2-30

Продолжение прерванной обработки:
  Просто запустить заново — скрипт пропустит уже заполненные строки.
"""

import os
import sys
import json
import time
import shutil
import argparse
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from openai import OpenAI
except ImportError:
    print("openai не установлен. Запусти:")
    print("  uv run --with openai,openpyxl scripts/gpt_dimensions_filler.py")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("openpyxl не установлен.")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

SRC_DIR    = Path(r"C:\Users\Admin\Desktop\На сортировку 26.04")
OUT_DIR    = Path(r"C:\Users\Admin\Desktop\На сортировку 08.05")
MODEL      = "gpt-4o-mini"
BATCH_SIZE = 15   # строк за один запрос к GPT
SAVE_EVERY = 3    # сохранять файл каждые N батчей (= каждые ~45 строк)

# Новые столбцы (порядок важен — именно так они будут вставлены)
NEW_COLS  = ["Вес, г", "Длина, мм", "Ширина, мм", "Высота, мм"]
COL_AFTER = "Описание"   # вставить новые столбцы сразу после этого


# ══════════════════════════════════════════════════════════════════════════════
#  ПРОМПТ
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = """Ты эксперт по автозапчастям. Тебе присылают список автозапчастей.
Для каждой позиции нужно указать приблизительные габариты и вес УПАКОВКИ (с учётом коробки/пакета):
  вес — в граммах (целое число)
  длина — наибольший размер, в мм (целое число)
  ширина — средний размер, в мм (целое число)
  высота — наименьший размер, в мм (целое число)

Правила:
- Если деталь продаётся комплектом (например «2 шт.») — давай вес/габариты всего комплекта
- Опирайся на типичные значения для данного вида запчасти, бренда и применяемости
- Если данных совсем недостаточно — давай наиболее вероятные значения, не ставь null
- Отвечай ТОЛЬКО JSON объектом вида {"items": [...]}
- Каждый элемент массива: {"idx": <число из запроса>, "вес": <число>, "длина": <число>, "ширина": <число>, "высота": <число>}
- Никаких пояснений, никаких markdown-блоков — только чистый JSON
"""


def build_user_prompt(items: list[dict]) -> str:
    lines = []
    for item in items:
        lines.append(f"[{item['idx']}]")
        lines.append(f"Наименование: {item['name']}")
        if item.get("params"):
            lines.append(f"Параметры: {str(item['params'])[:200]}")
        if item.get("desc"):
            lines.append(f"Описание: {str(item['desc'])[:250]}")
        lines.append("")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ХЕЛПЕРЫ
# ══════════════════════════════════════════════════════════════════════════════

def read_headers(ws) -> dict:
    """Возвращает {имя_колонки: номер_колонки} для первой строки."""
    return {
        ws.cell(1, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }


def ensure_new_columns(ws, headers: dict) -> dict:
    """
    Вставляет новые столбцы после COL_AFTER, если их ещё нет.
    Возвращает обновлённый словарь заголовков.
    """
    existing_cols = set(headers.keys())
    already_have  = all(col in existing_cols for col in NEW_COLS)
    if already_have:
        return headers

    # Позиция вставки: сразу после «Описание»
    if COL_AFTER in headers:
        insert_at = headers[COL_AFTER] + 1
    else:
        insert_at = ws.max_column + 1

    # Вставляем сразу N пустых колонок — это сдвигает всё вправо корректно
    to_add = [col for col in NEW_COLS if col not in existing_cols]
    ws.insert_cols(insert_at, len(to_add))

    # Заголовки для новых колонок
    for i, col_name in enumerate(to_add):
        cell = ws.cell(1, insert_at + i, col_name)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1A3A5C")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 14

    return read_headers(ws)


# ══════════════════════════════════════════════════════════════════════════════
#  API
# ══════════════════════════════════════════════════════════════════════════════

def ask_gpt(client: OpenAI, items: list[dict]) -> list[dict]:
    """Отправляет батч в GPT, возвращает список {idx, вес, длина, ширина, высота}."""
    user_msg = build_user_prompt(items)

    for attempt in range(3):
        try:
            resp = client.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": user_msg},
                ],
                temperature=0.1,
                response_format={"type": "json_object"},
                timeout=60,
            )
            raw = resp.choices[0].message.content.strip()
            parsed = json.loads(raw)

            # Ожидаем {"items": [...]}
            if "items" in parsed and isinstance(parsed["items"], list):
                return parsed["items"]
            # Fallback: если вернул просто массив под другим ключом
            for v in parsed.values():
                if isinstance(v, list):
                    return v
            return []

        except json.JSONDecodeError as e:
            print(f"    ⚠ JSON ошибка (попытка {attempt+1}/3): {e}")
        except Exception as e:
            print(f"    ⚠ Ошибка API (попытка {attempt+1}/3): {e}")

        time.sleep(8 * (attempt + 1))

    return []


# ══════════════════════════════════════════════════════════════════════════════
#  ОБРАБОТКА ФАЙЛА
# ══════════════════════════════════════════════════════════════════════════════

def process_file(client: OpenAI, src: Path, dst: Path, row_range: tuple | None = None):
    # Копируем исходник в папку результатов, если ещё не скопирован
    if not dst.exists():
        shutil.copy2(src, dst)
        print(f"  → Скопирован: {dst.name}")

    wb = openpyxl.load_workbook(dst)
    ws = wb.active

    # Убедиться, что нужные колонки есть
    headers = read_headers(ws)
    headers = ensure_new_columns(ws, headers)

    # Индексы нужных колонок
    desc_col  = headers.get(COL_AFTER)
    name_col  = headers.get("Наименование")
    param_col = headers.get("Параметры")
    w_col     = headers.get("Вес, г")
    l_col     = headers.get("Длина, мм")
    wd_col    = headers.get("Ширина, мм")
    h_col     = headers.get("Высота, мм")

    if not desc_col:
        print(f"  ⚠ Колонка «{COL_AFTER}» не найдена — пропуск файла")
        wb.close()
        return

    # Собираем строки для обработки
    todos = []
    for r in range(2, ws.max_row + 1):
        # Фильтр по диапазону строк (если задан)
        if row_range:
            lo, hi = row_range
            if not (lo <= r <= hi):
                continue

        # Только строки с заполненным «Описание»
        if not ws.cell(r, desc_col).value:
            continue

        # Пропускаем уже заполненные (resume)
        if w_col and ws.cell(r, w_col).value is not None:
            continue

        todos.append(r)

    total = len(todos)
    if not todos:
        print(f"  Нечего обрабатывать (все строки уже заполнены или нет описания)")
        wb.close()
        return

    print(f"  Строк для обработки: {total}")

    processed  = 0
    save_timer = 0

    for batch_no, batch_start in enumerate(range(0, total, BATCH_SIZE), start=1):
        batch_rows = todos[batch_start : batch_start + BATCH_SIZE]

        # Формируем items для промпта
        items = []
        for local_idx, r in enumerate(batch_rows):
            items.append({
                "idx":    local_idx,
                "name":   ws.cell(r, name_col).value  if name_col  else "",
                "params": ws.cell(r, param_col).value if param_col else "",
                "desc":   ws.cell(r, desc_col).value  or "",
            })

        batch_label = f"строки {batch_rows[0]}–{batch_rows[-1]}"
        print(f"  Батч {batch_no} ({batch_label})", end="", flush=True)

        results = ask_gpt(client, items)
        result_map = {r["idx"]: r for r in results}

        filled = 0
        for local_idx, r in enumerate(batch_rows):
            data = result_map.get(local_idx)
            if not data:
                continue
            if w_col:  ws.cell(r, w_col).value  = data.get("вес")
            if l_col:  ws.cell(r, l_col).value  = data.get("длина")
            if wd_col: ws.cell(r, wd_col).value = data.get("ширина")
            if h_col:  ws.cell(r, h_col).value  = data.get("высота")
            filled += 1

        processed  += filled
        save_timer += 1
        print(f" → заполнено {filled}/{len(batch_rows)}")

        # Периодическое сохранение
        if save_timer >= SAVE_EVERY:
            wb.save(dst)
            print(f"    💾 Сохранено (итого обработано: {processed})")
            save_timer = 0

        time.sleep(1.2)  # не превышаем rate limit

    # Финальное сохранение
    wb.save(dst)
    print(f"  ✓ Готово: {processed}/{total} строк → {dst.name}")
    wb.close()


# ══════════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Заполнение веса и габаритов через GPT",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--file",  default=None, help="Один конкретный файл (имя xlsx)")
    ap.add_argument("--rows",  default=None, help="Диапазон строк: 2-100 (только для одного файла)")
    args = ap.parse_args()

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print("\n  ✗ Не задан OPENAI_API_KEY!")
        print("  Задай ключ перед запуском:")
        print("    set OPENAI_API_KEY=sk-...")
        sys.exit(1)

    client = OpenAI(api_key=api_key)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Собираем список файлов
    if args.file:
        src_files = [SRC_DIR / args.file]
    else:
        src_files = sorted(SRC_DIR.glob("*.xlsx"))

    # Диапазон строк (только для одного файла, иначе игнорируется)
    row_range = None
    if args.rows:
        if "-" in args.rows:
            a, b = map(int, args.rows.split("-"))
        else:
            a = b = int(args.rows)
        row_range = (a, b)

    print()
    print("═" * 64)
    print("  GPT Dimensions Filler")
    print(f"  Модель:  {MODEL}")
    print(f"  Батч:    {BATCH_SIZE} строк/запрос")
    print(f"  Файлов:  {len(src_files)}")
    print(f"  Из:      {SRC_DIR}")
    print(f"  В:       {OUT_DIR}")
    print("═" * 64)
    print()

    total_files = len(src_files)
    for i, src in enumerate(src_files, 1):
        if not src.exists():
            print(f"[{i:3}/{total_files}] ✗ Не найден: {src.name}")
            continue

        dst = OUT_DIR / src.name
        print(f"[{i:3}/{total_files}] {src.name}")

        try:
            process_file(client, src, dst, row_range)
        except KeyboardInterrupt:
            print("\n\n  Прерывание — прогресс сохранён в последнем 💾")
            sys.exit(0)
        except Exception as e:
            print(f"  ✗ Ошибка: {e}")

        print()

    print("═" * 64)
    print("  Всё готово!")
    print(f"  Результаты: {OUT_DIR}")
    print("═" * 64)
    print()


if __name__ == "__main__":
    main()
