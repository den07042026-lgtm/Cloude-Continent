"""
Скрипт автоматического сбора данных с mikado-parts.ru.

Читает коды из прайса (колонка Code), авторизуется на сайте,
скачивает страницы деталей и сводит результат в Excel таблицу.

Запуск:
    uv run --with requests,openpyxl scripts/fetch_mikado.py
    uv run --with requests,openpyxl scripts/fetch_mikado.py --rows 2-20
    uv run --with requests,openpyxl scripts/fetch_mikado.py --rows 2-20 --ozon Амортизаторы_Ozon_заполнен.xlsx
"""

import re
import sys
import time
import json
import argparse
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "prompts"))
sys.stdout.reconfigure(encoding="utf-8")

# ─── Зависимости ─────────────────────────────────────────────────────────────
try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Установи зависимости: uv run --with requests,openpyxl scripts/fetch_mikado.py")
    sys.exit(1)

from mikado_parser import parse_mikado_page

# ─── Конфигурация ─────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent.parent
ENV_FILE    = BASE_DIR / ".env"
PRICE_FILE  = Path("C:/Users/Admin/Documents/Ecommerce/mikado_price_34.xlsx")
HTML_DIR    = BASE_DIR / "data/suppliers/mikado/html"
JSON_DIR    = BASE_DIR / "data/suppliers/mikado/json"
IMG_DIR     = BASE_DIR / "data/suppliers/mikado/images"
OUT_DIR     = BASE_DIR / "data/suppliers/mikado"

MIKADO_LOGIN_URL   = "https://mikado-parts.ru/office/SECURE.asp"
MIKADO_PRODUCT_URL = "https://mikado-parts.ru/office/galleyp.asp"
MIKADO_AJAX_URL    = "https://mikado-parts.ru/office/pp0.asp"

DELAY_SECONDS = 1.5  # пауза между запросами


# ─── Загрузка .env ────────────────────────────────────────────────────────────
def load_env() -> dict:
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


# ─── Авторизация ──────────────────────────────────────────────────────────────
def login(code: str, password: str) -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"
    })
    resp = session.post(
        MIKADO_LOGIN_URL,
        data={"CODE": code, "PASSWORD": password, "INSERT": "OK"},
        timeout=20
    )
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")
    if "Обслуживание клиентов" in html or "Продолжить" in html:
        print(f"  ✓ Авторизация успешна (код клиента: {code})")
    else:
        print(f"  ✗ Проверьте логин/пароль")
    return session


# ─── Скачать HTML страницы детали ─────────────────────────────────────────────
def fetch_product_html(session: requests.Session, code: str) -> str:
    """Скачивает HTML страницы детали. Возвращает текст в windows-1251."""
    resp = session.get(
        MIKADO_PRODUCT_URL,
        params={"code": code},
        timeout=20
    )
    resp.raise_for_status()
    return resp.content.decode("windows-1251", errors="replace")


# ─── Извлечь ARTID из HTML страницы детали ───────────────────────────────────
def extract_artid(html: str) -> str | None:
    """Извлекает ARTID из JavaScript на странице детали."""
    m = re.search(r"ARTID=(\d+)", html)
    return m.group(1) if m else None


# ─── Загрузить OEM номера ─────────────────────────────────────────────────────
def fetch_oem(session: requests.Session, artid: str, code: str) -> list[dict]:
    """
    Загружает таблицу OEM номеров для детали.
    Возвращает список {"manufacturer": "VW", "oem_code": "1K0513029JA"}.
    """
    resp = session.get(
        MIKADO_AJAX_URL,
        params={"MODE": "OEM", "ARTID": artid, "code": code},
        timeout=20
    )
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")

    oem_list = []
    # Строки таблицы: <tr ...><td>Manufacturer</td><td>OEM_CODE</td></tr>
    for row_m in re.finditer(r"<tr[^>]*>(.*?)</tr>", html, re.S | re.I):
        cells = re.findall(r"<td[^>]*>(.*?)</td>", row_m.group(1), re.S | re.I)
        if len(cells) >= 2:
            manufacturer = re.sub(r"<[^>]+>", "", cells[0]).strip()
            oem_code = re.sub(r"<[^>]+>", "", cells[1]).strip()
            if manufacturer and oem_code and manufacturer.lower() not in ("производитель", "марка", "бренд"):
                oem_list.append({"manufacturer": manufacturer, "oem_code": oem_code})
    return oem_list


# ─── Загрузить применяемость ──────────────────────────────────────────────────
def fetch_compatibility(session: requests.Session, artid: str) -> list[dict]:
    """
    Загружает таблицу применяемости через AJAX.
    Структура HTML: <tbody id='blockNNN'> с rowspan-ячейками для марки и модели.
    Возвращает список {"brand": ..., "model": ..., "modification": ...,
                        "year_from": ..., "year_to": ..., "engine": ...}.
    """
    resp = session.get(
        MIKADO_AJAX_URL,
        params={"MODE": "APL", "ARTID": artid},
        timeout=20
    )
    resp.raise_for_status()
    html = resp.content.decode("windows-1251", errors="replace")

    rows = []

    def _td_text(td_html: str) -> str:
        return re.sub(r"<[^>]+>", "", td_html).replace("&nbsp;", " ").strip()

    # Каждый tbody — группа одной марки (или несколько моделей)
    for tbody_m in re.finditer(r"<tbody[^>]*>(.*?)</tbody>", html, re.S | re.I):
        tbody = tbody_m.group(1)
        current_brand = ""
        current_model = ""

        for tr_m in re.finditer(r"<tr[^>]*>(.*?)</tr>", tbody, re.S | re.I):
            td_tags = re.findall(r"<td([^>]*)>(.*?)</td>", tr_m.group(1), re.S | re.I)
            if not td_tags:
                continue

            # Ячейки с rowspan идут первыми: brand (rs большой), затем model (rs меньше)
            rowspan_cells = [(attrs, _td_text(v)) for attrs, v in td_tags
                             if "rowspan" in attrs.lower()]

            if len(rowspan_cells) >= 2:
                # Строка начинает новую марку: td[0]=марка, td[1]=модель
                current_brand = rowspan_cells[0][1]
                current_model = rowspan_cells[1][1]
            elif len(rowspan_cells) == 1:
                # Продолжение марки, новая модель
                current_model = rowspan_cells[0][1]

            if not current_brand:
                continue

            # Все текстовые значения ячеек
            vals = [_td_text(v) for _, v in td_tags]

            # Год: формат MM.YYYY или YYYY
            years = [v for v in vals if re.match(r"\d{2}\.\d{4}$", v)]
            # Модификация: не марка, не модель, не год, не число
            skip = {current_brand, current_model}
            mods = [v for v in vals if v and v not in years
                    and v not in skip
                    and not re.match(r"^\d+(\s+\d+)*\s*$", v)]

            year_from = years[0] if len(years) > 0 else ""
            year_to   = years[1] if len(years) > 1 else ""
            mod = mods[0] if mods else ""

            rows.append({
                "brand": current_brand,
                "model": current_model,
                "modification": mod,
                "year_from": year_from,
                "year_to": year_to,
            })

    return rows


# ─── Скачать изображения ──────────────────────────────────────────────────────
def fetch_and_save_images(session: requests.Session, artid: str, code: str,
                           img_dir: Path) -> list[Path]:
    """
    Загружает изображения из вкладки КАРТИНКИ (pp0.asp?MODE=PIC).
    Сохраняет как {code}_1.jpg, {code}_2.jpg, ... в img_dir.
    Возвращает список сохранённых путей.
    """
    resp = session.get(
        MIKADO_AJAX_URL,
        params={"MODE": "PIC", "CODE": code, "ARTID": artid},
        timeout=20
    )
    resp.raise_for_status()
    pic_html = resp.content.decode("windows-1251", errors="replace")

    # Извлекаем src из всех <img> тегов
    img_srcs = re.findall(r"<img[^>]+src=[\'\"]?(wi/img\.asp[^\'\">\s]+)", pic_html, re.I)
    if not img_srcs:
        return []

    # Убираем дубли (одинаковый размер → один и тот же файл)
    seen_content = set()
    saved = []
    idx = 1
    for src in img_srcs:
        img_url = "https://mikado-parts.ru/office/" + src
        try:
            r = session.get(img_url, timeout=20)
            r.raise_for_status()
            # Проверяем что это действительно изображение (JPEG/PNG magic bytes)
            content = r.content
            if not (content[:2] == b'\xff\xd8' or content[:8] == b'\x89PNG\r\n\x1a\n'):
                continue
            # Дедупликация по хешу содержимого
            content_hash = hash(content)
            if content_hash in seen_content:
                continue
            seen_content.add(content_hash)
            # Определяем расширение
            ext = ".jpg" if content[:2] == b'\xff\xd8' else ".png"
            out_path = img_dir / f"{code}_{idx}{ext}"
            out_path.write_bytes(content)
            saved.append(out_path)
            idx += 1
        except Exception:
            continue
    return saved


def _compat_to_text(compat_rows: list[dict]) -> str:
    """Форматирует применяемость: уникальные Марка+Модель с диапазоном лет."""
    if not compat_rows:
        return ""

    # Группируем по марке+модели, берём мин/макс год
    groups: dict[str, dict] = {}
    for row in compat_rows:
        key = f"{row.get('brand', '')}|{row.get('model', '')}"
        if key not in groups:
            groups[key] = {
                "brand": row.get("brand", ""),
                "model": row.get("model", ""),
                "year_from": row.get("year_from", ""),
                "year_to":   row.get("year_to", ""),
            }
        else:
            # расширяем диапазон лет
            g = groups[key]
            if row.get("year_from") and (not g["year_from"] or row["year_from"] < g["year_from"]):
                g["year_from"] = row["year_from"]
            if row.get("year_to") and (not g["year_to"] or row["year_to"] > g["year_to"]):
                g["year_to"] = row["year_to"]

    parts = []
    for g in groups.values():
        line = f"{g['brand']} {g['model']}".strip()
        yf = re.sub(r"^\d{2}\.", "", g["year_from"]) if g["year_from"] else ""
        yt = re.sub(r"^\d{2}\.", "", g["year_to"]) if g["year_to"] else ""
        if yf or yt:
            line += f" ({yf}–{yt})"
        parts.append(line)
    return "; ".join(parts)


# ─── Читать коды из прайса ────────────────────────────────────────────────────
def read_price_codes(price_file: Path, row_start: int, row_end: int) -> list[dict]:
    """Читает строки из прайса. Ищет колонку Code автоматически."""
    wb = openpyxl.load_workbook(price_file, read_only=True, data_only=True)
    ws = wb.active

    # Читаем заголовки из первой строки
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(v).strip() if v else f"col{i}" for i, v in enumerate(first_row, 1)]

    # Ищем колонку Prodnum (код для URL: f-a22025) и Code (артикул: a22025)
    prodnum_col = None
    code_col = None
    for i, h in enumerate(headers, 1):
        if h.lower() == "prodnum":
            prodnum_col = i
        if h.lower() == "code":
            code_col = i

    # Колонка для формирования URL — Prodnum если есть, иначе Code
    url_col = prodnum_col or code_col
    if not url_col:
        raise ValueError("Колонки 'Prodnum' или 'Code' не найдены в прайсе")

    print(f"  Колонка для URL (Prodnum): {prodnum_col}, Code: {code_col}")

    items = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=row_start, max_row=row_end, values_only=True),
        start=row_start
    ):
        code = row[url_col - 1] if len(row) >= url_col else None
        if code:
            item = {"_row": row_idx, "_code": str(code).strip()}
            for col_idx, val in enumerate(row, start=1):
                h = headers[col_idx - 1] if col_idx <= len(headers) else f"col{col_idx}"
                item[h] = val
            items.append(item)

    wb.close()
    return items


# ─── Сохранить сводную таблицу Excel ─────────────────────────────────────────
def save_summary_excel(results: list[dict], out_path: Path):
    """Сохраняет сводную таблицу со всеми данными."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mikado данные"

    # Стили
    H_FILL  = PatternFill("solid", fgColor="1A1A2E")
    H_FONT  = Font(bold=True, color="FFFFFF", size=10)
    S_FILL1 = PatternFill("solid", fgColor="FFFFFF")
    S_FILL2 = PatternFill("solid", fgColor="FFF3EE")
    A_FONT  = Font(color="CC0000", bold=True)  # для аналогов дешевле нашей цены

    # Заголовки
    headers = [
        "Код (Mikado)",
        "Наше название",
        "Бренд",
        "Цена закупки, руб.",
        "Склады (наличие)",
        "Всего на складах, шт.",
        # Параметры
        "Тип амортизатора",
        "Конструкция",
        "Система",
        "Сторона установки",
        "Способ крепления",
        "Диаметр, мм",
        "Диаметр штока, мм",
        "Ходовая часть",
        "Парные артикулы",
        # OEM и применяемость
        "OEM номера",
        "Применяемость",
        # Аналоги
        "Аналог 1 — бренд",
        "Аналог 1 — код",
        "Аналог 1 — цена",
        "Аналог 2 — бренд",
        "Аналог 2 — код",
        "Аналог 2 — цена",
        "Аналог 3 — бренд",
        "Аналог 3 — код",
        "Аналог 3 — цена",
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, h)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Данные
    for row_idx, r in enumerate(results, start=2):
        fill = S_FILL1 if row_idx % 2 == 0 else S_FILL2

        params = r.get("params", {})
        analogs = sorted(
            [a for a in r.get("analogs", []) if a.get("price", 0) > 0],
            key=lambda x: x["price"]
        )

        stock_text = ", ".join(
            f"{s['warehouse']} {s['qty']}шт."
            for s in r.get("stock_items", [])
        )
        total_stock = sum(s["qty"] for s in r.get("stock_items", []))

        # OEM номера: группируем по производителю
        oem_details = r.get("oem_details", [])
        if oem_details:
            oem_text = "; ".join(
                f"{d['manufacturer']}: {d['oem_code']}" for d in oem_details[:10]
            )
        else:
            oem_numbers = r.get("oem_numbers", [])
            oem_text = "; ".join(oem_numbers[:10])

        row_values = [
            r.get("code", ""),
            r.get("name", ""),
            r.get("brand", ""),
            r.get("price", 0),
            stock_text,
            total_stock,
            # Параметры
            params.get("Тип амортизатора", ""),
            params.get("Конструкция амортизатора", ""),
            params.get("Система амортизатора", ""),
            params.get("Сторона установки", ""),
            params.get("Способ крепления амортизатора", ""),
            params.get("Диаметр [мм]", "").replace(" мм", ""),
            params.get("Диаметр входного штока [мм]", "").replace(" мм", ""),
            params.get("Ходовая часть", ""),
            params.get("парные номера артикулов", ""),
            # OEM и применяемость
            oem_text,
            r.get("compatibility", ""),
        ]
        # Первые 3 аналога
        for i in range(3):
            if i < len(analogs):
                a = analogs[i]
                row_values += [a.get("brand", ""), a.get("code", ""), a.get("price", 0)]
            else:
                row_values += ["", "", ""]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Если аналог дешевле нашей цены — подсветить красным
            if col_idx in (20, 23, 26) and isinstance(val, (int, float)):
                our_price = r.get("price", 0)
                if val > 0 and val < our_price:
                    cell.font = A_FONT

    # Ширина колонок
    #              Код  Назв  Бренд  Цена  Склады  Шт  Тип   Констр  Сист  Сторон  Крепл  Диам  Шток  Ход  Парн  OEM   Прим  А1бр  А1код  А1ц  А2бр  А2код  А2ц  А3бр  А3код  А3ц
    col_widths = [15,  35,   12,    14,   35,    12, 20,   18,     18,   20,    22,    10,   10,   14,  20,   40,   50,   12,   20,   12,  12,   20,   12,  12,   20,   12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"\n  Сводная таблица сохранена: {out_path}")


# ─── Заполнить Ozon-шаблон ────────────────────────────────────────────────────
def fill_ozon_template(results: list[dict], ozon_file: Path):
    """Дополняет файл Ozon шаблона данными из Mikado (только пустые ячейки)."""
    wb = openpyxl.load_workbook(ozon_file)
    ws = wb["Шаблон"]

    # Заголовки строки 2
    headers = {c: ws.cell(2, c).value for c in range(1, ws.max_column + 1)}

    # Индекс: артикул → номер строки
    art_to_row = {}
    for row_idx in range(5, ws.max_row + 1):
        art = ws.cell(row_idx, 2).value
        if art:
            art_to_row[str(art).strip()] = row_idx

    updated = 0
    for r in results:
        code = r.get("code", "")
        if code not in art_to_row:
            continue

        row_idx = art_to_row[code]
        params = r.get("params", {})

        fill_map = {
            27: _build_hashtags(r),
            28: _build_annotation(r),
            36: _extract_int(params.get("Диаметр [мм]", "")),
            43: "Беларусь" if r.get("brand", "").upper() in ("FENOX",) else None,
            44: "1 шт.",
            48: "8708800009",
        }

        for col, value in fill_map.items():
            if value and ws.cell(row_idx, col).value is None:
                ws.cell(row_idx, col).value = value
                updated += 1

    wb.save(ozon_file)
    print(f"  Ozon шаблон обновлён: {updated} ячеек заполнено → {ozon_file}")


def _build_hashtags(r: dict) -> str:
    tags = ["#амортизатор", "#подвеска"]
    brand = r.get("brand", "").lower().replace(" ", "_")
    if brand:
        tags.append(f"#{brand}")
    params = r.get("params", {})
    side = params.get("Сторона установки", "")
    if "задн" in side.lower() or "задний" in side.lower() or "задние" in side.lower():
        tags.append("#задний_амортизатор")
    elif "передн" in side.lower():
        tags.append("#передний_амортизатор")
    gas = params.get("Тип амортизатора", "")
    if "газ" in gas.lower():
        tags.append("#газовый_амортизатор")
    elif "масл" in gas.lower():
        tags.append("#масляный_амортизатор")
    # Марки авто из применяемости/аналогов
    compat = " ".join(a.get("name", "") for a in r.get("analogs", []))
    for brand_car, tag in [("skoda", "#skoda"), ("volkswagen", "#volkswagen"),
                            ("vw", "#volkswagen"), ("audi", "#audi"),
                            ("toyota", "#toyota"), ("nissan", "#nissan"),
                            ("kia", "#kia"), ("hyundai", "#hyundai"),
                            ("bmw", "#bmw"), ("mercedes", "#mercedes")]:
        if brand_car in compat.lower() and tag not in tags:
            tags.append(tag)
    return " ".join(tags[:12])


def _build_annotation(r: dict) -> str:
    params = r.get("params", {})
    parts = [f"Амортизатор {r.get('brand', '')} {r.get('code', '').upper().replace('F-', '')}."]
    system = params.get("Система амортизатора", "")
    gas = params.get("Тип амортизатора", "")
    diam = params.get("Диаметр [мм]", "")
    side = params.get("Сторона установки", "")
    if system or gas:
        parts.append(f"{system.capitalize()}, {gas}.".strip(", .") + ".")
    if diam:
        parts.append(f"Диаметр {diam}.")
    if side:
        parts.append(f"Установка: {side.lower()}.")
    # Применяемость из аналогов
    compat_cars = set()
    for a in r.get("analogs", [])[:3]:
        name = a.get("name", "")
        for car in re.findall(r"(Skoda\s+\w+|VW\s+\w+|Audi\s+\w+|Toyota\s+\w+|Nissan\s+\w+|Kia\s+\w+|Hyundai\s+\w+)", name):
            compat_cars.add(car)
    if compat_cars:
        parts.append(f"Совместим с: {'; '.join(sorted(compat_cars))}.")
    return " ".join(parts)


def _extract_int(s: str):
    m = re.search(r"(\d+)", str(s))
    return int(m.group(1)) if m else None


# ─── Главная функция ──────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", default="2-20", help="Диапазон строк прайса, например 2-20")
    parser.add_argument("--ozon", default=None, help="Путь к Ozon шаблону для автозаполнения")
    args = parser.parse_args()

    # Диапазон строк
    row_start, row_end = map(int, args.rows.split("-"))

    # Загрузка credentials
    env = load_env()
    code_cred = env.get("MIKADO_CODE", "35275")
    password   = env.get("MIKADO_PASSWORD", "")
    if not password:
        print("Ошибка: MIKADO_PASSWORD не найден в .env")
        sys.exit(1)

    # Папки
    HTML_DIR.mkdir(parents=True, exist_ok=True)
    JSON_DIR.mkdir(parents=True, exist_ok=True)
    IMG_DIR.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*55}")
    print(f"  Mikado Data Fetcher")
    print(f"  Прайс: {PRICE_FILE.name}, строки {row_start}–{row_end}")
    print(f"{'='*55}\n")

    # Читаем прайс
    print("[1/4] Читаю прайс...")
    items = read_price_codes(PRICE_FILE, row_start, row_end)
    print(f"  Найдено кодов: {len(items)}")
    for it in items:
        print(f"  → {it['_code']}")

    # Авторизация
    print("\n[2/4] Авторизация на mikado-parts.ru...")
    session = login(code_cred, password)

    # Скачивание и парсинг
    print(f"\n[3/4] Скачиваю и парсю {len(items)} страниц...")
    results = []
    for i, item in enumerate(items, start=1):
        art_code = item["_code"]
        print(f"  [{i:2}/{len(items)}] {art_code} ... ", end="", flush=True)

        html_path = HTML_DIR / f"{art_code}.html"
        json_path = JSON_DIR / f"{art_code}.json"

        try:
            html = fetch_product_html(session, art_code)

            # Сохраняем HTML
            html_path.write_bytes(html.encode("windows-1251", errors="replace"))

            # Парсим из строки напрямую
            import tempfile, os
            tmp = tempfile.NamedTemporaryFile(mode="wb", suffix=".html", delete=False)
            tmp.write(html.encode("windows-1251", errors="replace"))
            tmp.close()

            data = parse_mikado_page(tmp.name)
            data["_price_row"] = item  # добавляем строку из прайса
            os.unlink(tmp.name)

            # Загружаем OEM и Применяемость через AJAX
            artid = extract_artid(html)
            if artid:
                time.sleep(0.5)
                oem_rows = fetch_oem(session, artid, art_code)
                data["oem_numbers"] = [r["oem_code"] for r in oem_rows]
                data["oem_details"] = oem_rows

                time.sleep(0.5)
                compat_rows = fetch_compatibility(session, artid)
                data["compatibility_rows"] = compat_rows
                data["compatibility"] = _compat_to_text(compat_rows)

                time.sleep(0.5)
                saved_imgs = fetch_and_save_images(session, artid, art_code, IMG_DIR)
                data["images"] = [str(p.name) for p in saved_imgs]
                print(f"    OEM: {len(oem_rows)} шт., "
                      f"применяемость: {len(compat_rows)} строк, "
                      f"фото: {len(saved_imgs)} шт.", end="")
            else:
                print(f"    ARTID не найден", end="")

            # Сохраняем JSON
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            results.append(data)
            price_str = f"{data['price']:.0f} руб." if data['price'] else "цена не найдена"
            stock_str = f"{sum(s['qty'] for s in data['stock_items'])} шт."
            print(f"\n  ✓  {data['name'][:30]:<30} | {data['brand']:<10} | {price_str} | {stock_str}")

        except Exception as e:
            print(f"✗  Ошибка: {e}")

        if i < len(items):
            time.sleep(DELAY_SECONDS)

    # Сводная таблица
    print(f"\n[4/4] Сохраняю сводную таблицу...")
    out_path = OUT_DIR / "mikado_summary.xlsx"
    save_summary_excel(results, out_path)

    # Заполнить Ozon шаблон (если указан)
    if args.ozon:
        ozon_path = Path(args.ozon) if Path(args.ozon).is_absolute() else Path("C:/Users/Admin/Documents/Ecommerce") / args.ozon
        if ozon_path.exists():
            print(f"\n  Обновляю Ozon шаблон: {ozon_path.name}...")
            fill_ozon_template(results, ozon_path)
        else:
            print(f"  Ozon файл не найден: {ozon_path}")

    print(f"\n{'='*55}")
    print(f"  Готово! Обработано: {len(results)}/{len(items)}")
    print(f"  HTML: {HTML_DIR}")
    print(f"  JSON: {JSON_DIR}")
    print(f"  Таблица: {out_path}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
