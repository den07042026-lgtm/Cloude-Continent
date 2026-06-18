"""
update_photo_links.py
Обновляет ссылки на фото в уже готовом файле «Озон - на модерацию».
Получает свежие download-URL с публичного Яндекс.Диска и заменяет их в Excel.
"""

import os
import requests
import openpyxl

# ── Настройки ─────────────────────────────────────────────────────────────────

FILE_PATH = r"C:\Users\Admin\Desktop\Озон - на модерацию\Кольцо поршневое_10.05.2026.xlsx"
YADISK_PUBLIC_URL = "https://disk.yandex.ru/d/h92xccNNdiKyCA"

SHEET_MAIN = "Шаблон"
DATA_START = 5   # первая строка данных


# ── Яндекс.Диск ──────────────────────────────────────────────────────────────

def fetch_yadisk_public(folder_url: str) -> dict[str, str]:
    """Возвращает {стем_нижний_регистр: download_url} для всех файлов в папке."""
    index: dict[str, str] = {}
    offset, limit, total = 0, 1000, None

    print(f"Получаю список фото с Яндекс.Диска...", flush=True)

    while True:
        try:
            r = requests.get(
                "https://cloud-api.yandex.net/v1/disk/public/resources",
                params={"public_key": folder_url, "limit": limit,
                        "sort": "name", "offset": offset},
                timeout=30,
            )
            r.raise_for_status()
        except requests.RequestException as e:
            print(f"  ОШИБКА: {e}")
            break

        data = r.json()
        emb = data.get("_embedded", {})
        if total is None:
            total = emb.get("total", 0)
            print(f"  Всего файлов: {total}")

        for item in emb.get("items", []):
            if item.get("type") == "file":
                name = item.get("name", "")
                url = item.get("file", "")
                if name and url:
                    index[os.path.splitext(name)[0].lower()] = url

        offset += limit
        if not emb.get("items") or (total and offset >= total):
            break

    print(f"  Проиндексировано: {len(index)} файлов")
    return index


# ── Поиск фото по артикулу ────────────────────────────────────────────────────

def get_photo_urls(article: str, index: dict[str, str]) -> tuple[str | None, list[str]]:
    """
    Главное фото:    стем содержит артикул И '1_info'
    Дополнительные: все остальные стемы, содержащие артикул
    """
    art = article.lower()
    main = None
    extras: list[tuple[str, str]] = []

    for stem, url in index.items():
        if art not in stem:
            continue
        if "1_info" in stem:
            main = url
        else:
            extras.append((stem, url))

    extras.sort(key=lambda x: x[0])
    return main, [u for _, u in extras]


# ── Определение столбцов фото динамически ─────────────────────────────────────

def detect_photo_cols(ws) -> tuple[int, int]:
    """Ищет столбцы 'Ссылка на главное фото' и 'Ссылки на дополнительные фото' в строке 2."""
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    col_main = col_add = None
    for i, h in enumerate(header_row, 1):
        if not h:
            continue
        h_str = str(h).strip().rstrip("*").strip()
        if "главное фото" in h_str.lower():
            col_main = i
        elif "дополнительные фото" in h_str.lower() or "дополн" in h_str.lower():
            col_add = i
    return col_main, col_add


# ── Основная логика ───────────────────────────────────────────────────────────

def main() -> None:
    # 1. Получить свежий индекс фото
    photo_index = fetch_yadisk_public(YADISK_PUBLIC_URL)
    if not photo_index:
        print("Индекс пуст — выход.")
        return

    # 2. Открыть файл
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb[SHEET_MAIN]

    col_main, col_add = detect_photo_cols(ws)
    print(f"\nСтолбец главного фото:         {col_main}")
    print(f"Столбец доп. фото:             {col_add}")

    if not col_main or not col_add:
        print("Не удалось определить столбцы фото!")
        return

    # 3. Обойти строки и заменить URL
    updated = skipped = no_main = 0

    for row in ws.iter_rows(min_row=DATA_START):
        article_cell = row[1]   # col 2 (0-indexed = 1) — артикул с суффиксом -con
        if not article_cell.value:
            continue

        article_con = str(article_cell.value).strip()
        article = article_con.removesuffix("-con")   # raw Mikado code

        main_url, extra_urls = get_photo_urls(article, photo_index)

        if not main_url:
            no_main += 1
            skipped += 1
            continue

        ws.cell(row[0].row, col_main).value = main_url
        ws.cell(row[0].row, col_add).value = "\n".join(extra_urls) if extra_urls else None
        updated += 1

    # 4. Сохранить
    wb.save(FILE_PATH)
    wb.close()

    print(f"\nОбновлено:          {updated} строк")
    print(f"Нет главного фото:  {no_main} артикулов")
    print(f"Файл сохранён:      {FILE_PATH}")


if __name__ == "__main__":
    main()
