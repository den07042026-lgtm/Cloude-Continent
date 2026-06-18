#!/usr/bin/env python3
"""
fill_ozon_templates.py

Заполняет шаблоны Ozon данными из файлов с ценами.
Поддерживает загрузку фото с публичного или приватного Яндекс.Диска.

Запуск:
    .venv\\Scripts\\python.exe fill_ozon_templates.py
"""

import os
import re
import time
import subprocess
import openpyxl
import requests
from playwright.sync_api import sync_playwright

# ══════════════════════════════════════════════════════════════════════════════
# КОНФИГУРАЦИЯ ЗАПУСКА — меняй здесь при каждом новом батче
# ══════════════════════════════════════════════════════════════════════════════

SOURCE_DIR   = r"C:\Users\Admin\Desktop\Итого"
UPLOAD_DIR   = r"C:\Users\Admin\Desktop\На загрузку"
OUTPUT_DIR   = r"C:\Users\Admin\Desktop\Озон - на модерацию"

# Ограничить обработку конкретными категориями (пустой список = все совпадения)
PROCESS_ONLY = [
    "Амортизатор подвески",
    "Болты ГБЦ",
    "Зеркало заднего вида",
    "Кольцо поршневое",
]

# ── Локальные фото ─────────────────────────────────────────────────────────
# Папка с подпапками по категориям; артикул без фото здесь → позиция пропускается
IMG_LOCAL_DIR = r"C:\Users\Admin\Desktop\Итого\img"

# ── Яндекс.Диск ───────────────────────────────────────────────────────────
# Публичная папка — вставь ссылку вида disk.yandex.ru/d/...
# Как получить: правая кнопка по папке в браузере → Поделиться → скопировать ссылку
YADISK_PUBLIC_URL = ""   # не используется — каждая категория имеет свой URL

# True — папка содержит подпапки по категориям; индекс строится отдельно для каждой
YADISK_HAS_SUBFOLDERS = False

# Отдельные публичные ссылки по категориям (переопределяют YADISK_PUBLIC_URL)
CATEGORY_YADISK_URLS: dict[str, str] = {
    "Амортизатор подвески":  "https://disk.yandex.ru/d/f9JqR0kq_jcF8Q",
    "Болты ГБЦ":             "https://disk.yandex.ru/d/z_dGHmAt4GcWWg",
    "Зеркало заднего вида":  "https://disk.yandex.ru/d/oJQjnsbhTSheXA",
    "Кольцо поршневое":      "https://disk.yandex.ru/d/h92xccNNdiKyCA",
}

# Приватная папка через браузер (не нужна если заполнен YADISK_PUBLIC_URL)
YADISK_DISK_PATH  = ""

# Chrome для приватного доступа
CDP_URL        = "http://localhost:9222"
CHROME_EXE     = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
CHROME_PROFILE = r"C:\Users\Admin\AppData\Local\Google\Chrome\User Data"

# ══════════════════════════════════════════════════════════════════════════════
# Столбцы источника (0-indexed)
# ══════════════════════════════════════════════════════════════════════════════
# 0  Код (Mikado)   1  Наименование    2  Бренд
# 3  Цена закупки   4  Цена продажи    8  Параметры
# 9  OEM номера    10  Применяемость  11  Альт. артикулы товара
# 12 Описание      13  Вес, г         14  Длина, мм
# 15 Ширина, мм    16  Высота, мм
S = {
    'article':       0,
    'name':          1,
    'brand':         2,
    'price_sell':    4,
    'oem':           9,
    'applicability': 10,
    'alt_articles':  11,
    'description':   12,
    'weight':        13,
    'length':        14,
    'width':         15,
    'height':        16,
}

# ══════════════════════════════════════════════════════════════════════════════
# Маппинг заголовков шаблона на внутренние ключи (для динамического определения)
# Столбцы читаются из строки 2 каждого шаблона — не хардкодим номера колонок,
# т.к. они различаются между категориями Ozon.
# ══════════════════════════════════════════════════════════════════════════════
HEADER_TO_KEY = {
    'Артикул':                        'article',
    'Название товара':                 'name',
    'Цена, руб.':                      'price',
    'НДС, %':                          'vat',
    'Ускоренный сбор отзывов':         'reviews',
    'Вес в упаковке, г':               'weight_pkg',
    'Ширина упаковки, мм':             'width_pkg',
    'Высота упаковки, мм':             'height_pkg',
    'Длина упаковки, мм':              'length_pkg',
    'Ссылка на главное фото':          'photo_main',
    'Ссылки на дополнительные фото':   'photo_add',
    'Бренд':                           'brand',
    'Партномер':                       'part_number',
    'Тип':                             'type_',
    'Аннотация':                       'description',
    'OEM-номер':                       'oem',
    'Альтернативные артикулы товара':  'alt_articles',
}

# Лист «Для чего подходит»
A = {'article': 1, 'brand': 2, 'model': 3, 'mod': 4}

DATA_START    = 5
VAT_VALUE     = "Не облагается"

# ══════════════════════════════════════════════════════════════════════════════
# Яндекс.Диск — публичный (requests, без браузера)
# ══════════════════════════════════════════════════════════════════════════════

def fetch_yadisk_public(folder_url: str, subfolder: str = "") -> dict[str, str]:
    """
    Файлы из публичной папки Я.Диска. Возвращает {стем: url}.
    subfolder — имя подпапки внутри публичной папки (напр. "Трос акселератора").
    """
    index: dict[str, str] = {}
    offset, limit, total = 0, 1000, None

    label = f"  Яндекс.Диск / {subfolder}..." if subfolder else "  Яндекс.Диск (публичный)..."
    print(label, end=' ', flush=True)

    params_base: dict = {'public_key': folder_url, 'limit': limit, 'sort': 'name'}
    if subfolder:
        params_base['path'] = f'/{subfolder}'

    while True:
        try:
            r = requests.get(
                "https://cloud-api.yandex.net/v1/disk/public/resources",
                params={**params_base, 'offset': offset},
                timeout=30,
            )
            r.raise_for_status()
        except requests.RequestException as e:
            print(f"\n  ОШИБКА: {e}")
            break

        data = r.json()
        emb  = data.get('_embedded', {})
        if total is None:
            total = emb.get('total', 0)
            print(f"файлов: {total}", flush=True)

        for item in emb.get('items', []):
            if item.get('type') == 'file':
                name = item.get('name', '')
                url  = item.get('file', '')
                if name and url:
                    index[os.path.splitext(name)[0].lower()] = url

        offset += limit
        if not emb.get('items') or (total and offset >= total):
            break

    return index


# ══════════════════════════════════════════════════════════════════════════════
# Яндекс.Диск — приватный (через залогиненный Chrome)
# ══════════════════════════════════════════════════════════════════════════════

def fetch_yadisk_private(disk_path: str) -> dict[str, str]:
    """
    Файлы из приватной папки через Chrome с сессией Яндекса.
    Подключается по CDP, либо запускает Chrome с профилем.
    Возвращает {стем: url}.
    """
    index: dict[str, str] = {}

    with sync_playwright() as pw:
        # — Подключение к Chrome —
        ctx = None
        try:
            browser = pw.chromium.connect_over_cdp(CDP_URL)
            ctx = browser.contexts[0] if browser.contexts else None
            if ctx:
                print(f"  Chrome подключён по CDP ({CDP_URL})")
        except Exception as e:
            print(f"  CDP недоступен ({e}), запускаю Chrome с профилем...")

        if ctx is None:
            print("  Закрываю Chrome и перезапускаю с отладочным портом...")
            subprocess.run(
                ["taskkill", "/F", "/IM", "chrome.exe", "/T"],
                capture_output=True,
            )
            time.sleep(1)
            subprocess.Popen([
                CHROME_EXE,
                "--remote-debugging-port=9222",
                "--profile-directory=Default",
                f"--user-data-dir={CHROME_PROFILE}",
                "--no-first-run",
                "--no-default-browser-check",
            ])
            # Ждём запуска Chrome (до 15 секунд)
            for i in range(15):
                time.sleep(1)
                try:
                    r = requests.get(
                        "http://localhost:9222/json/version", timeout=2
                    )
                    if r.ok:
                        print(f"  Chrome готов (через {i+1}с)")
                        break
                except Exception:
                    pass
            else:
                print("  ✗ Chrome не запустился за 15 секунд — фото пропущены")
                return index
            try:
                browser = pw.chromium.connect_over_cdp(CDP_URL)
                ctx = browser.contexts[0] if browser.contexts else None
            except Exception as e2:
                print(f"  ✗ CDP после перезапуска: {e2}")
                return index

        page = ctx.new_page()

        # — Открываем Яндекс.Диск —
        print("  Открываю disk.yandex.ru...", end=' ', flush=True)
        try:
            page.goto("https://disk.yandex.ru/",
                      wait_until="domcontentloaded", timeout=30_000)
        except Exception as e:
            print(f"\n  ОШИБКА загрузки страницы: {e}")
            page.close()
            return index

        cur = page.url
        print(f"URL после загрузки: {cur}")

        if any(k in cur for k in ("passport", "login", "auth")):
            print("  ⚠  Нужна авторизация Яндекс!")
            print("  Войдите в открытом браузере, затем нажмите Enter...")
            input()
            page.goto("https://disk.yandex.ru/",
                      wait_until="domcontentloaded", timeout=30_000)

        # — Получаем список файлов —
        print(f"  Запрашиваю список файлов: {disk_path}")
        offset, limit, total = 0, 1000, None

        while True:
            result = page.evaluate(
                """([path, limit, offset]) =>
                    fetch('/api/v1/resources'
                        + '?path='   + encodeURIComponent(path)
                        + '&limit='  + limit
                        + '&offset=' + offset)
                    .then(r => r.ok ? r.json() : r.status)
                    .catch(e => 'ERR:' + String(e))
                """,
                [disk_path, limit, offset],
            )

            # Диагностика ответа
            if isinstance(result, (int, str)):
                print(f"  ⚠  API вернул: {result}")
                # Если 404 — возможно путь неверный, пробуем без /disk/
                if result == 404 and disk_path.startswith('/disk/'):
                    alt = disk_path[5:]   # убираем /disk/
                    print(f"  Пробую альтернативный путь: {alt}")
                    result = page.evaluate(
                        """([path, limit, offset]) =>
                            fetch('/api/v1/resources'
                                + '?path='   + encodeURIComponent(path)
                                + '&limit='  + limit
                                + '&offset=' + offset)
                            .then(r => r.ok ? r.json() : r.status)
                            .catch(e => 'ERR:' + String(e))
                        """,
                        [alt, limit, offset],
                    )
                    if isinstance(result, (int, str)):
                        print(f"  ⚠  Альт. путь тоже не работает: {result}")
                        break
                else:
                    break

            emb = result.get('_embedded', {})
            if total is None:
                total = emb.get('total', 0)
                print(f"  Файлов найдено: {total}", flush=True)

            items     = emb.get('items', [])
            no_url:   list[tuple[str, str]] = []

            for item in items:
                if item.get('type') != 'file':
                    continue
                name = item.get('name', '')
                url  = item.get('file', '')
                stem = os.path.splitext(name)[0].lower() if name else ''
                if not stem:
                    continue
                if url:
                    index[stem] = url
                else:
                    no_url.append((stem, item.get('path', '')))

            # Если поле file отсутствует — получаем URL батчем
            if no_url:
                print(f"  Запрашиваю {len(no_url)} download-URL батчем...",
                      end=' ', flush=True)
                urls = page.evaluate(
                    """paths => Promise.all(paths.map(p =>
                        fetch('/api/v1/resources/download?path='
                              + encodeURIComponent(p))
                        .then(r => r.json()).then(d => d.href || '')
                        .catch(() => '')
                    ))""",
                    [p for _, p in no_url],
                )
                for (stem, _), url in zip(no_url, urls):
                    if url:
                        index[stem] = url
                print("OK", flush=True)

            offset += limit
            if not items or (total is not None and offset >= total):
                break

        print(f"  Итого проиндексировано: {len(index)} файлов")
        if index:
            sample = list(index.keys())[:3]
            print(f"  Примеры имён: {sample}")

        page.close()

    return index


# ══════════════════════════════════════════════════════════════════════════════
# Локальный индекс фото (фильтр)
# ══════════════════════════════════════════════════════════════════════════════

def build_local_photo_set(category: str) -> set[str]:
    """
    Сканирует IMG_LOCAL_DIR/<category> и возвращает множество стемов (нижний регистр).
    Артикул, не встречающийся ни в одном стеме, пропускается без обращения к Яндекс.Диску.
    """
    folder = os.path.join(IMG_LOCAL_DIR, category)
    if not os.path.isdir(folder):
        print(f"  ⚠  Локальная папка фото не найдена: {folder}")
        return set()
    stems = {
        os.path.splitext(f)[0].lower()
        for f in os.listdir(folder)
        if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))
    }
    print(f"  Локальных фото: {len(stems)} стемов в {folder}")
    return stems


# ══════════════════════════════════════════════════════════════════════════════
# Динамическое определение столбцов шаблона
# ══════════════════════════════════════════════════════════════════════════════

def detect_template_columns(ws) -> dict[str, int]:
    """
    Читает строку 2 шаблона и возвращает {key: col_1indexed}.
    Использует HEADER_TO_KEY: сначала точное совпадение заголовка (без *),
    затем startswith — для полей вида «Партномер (артикул производителя)*».
    """
    result: dict[str, int] = {}
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    for col_idx, cell_val in enumerate(header_row, 1):
        if not cell_val:
            continue
        cell_str = str(cell_val).strip().rstrip('*').strip()
        for header, key in HEADER_TO_KEY.items():
            if cell_str == header or cell_str.startswith(header):
                if key not in result:   # первое совпадение побеждает
                    result[key] = col_idx
                break
    return result


# ══════════════════════════════════════════════════════════════════════════════
# Нормализация текстовых полей
# ══════════════════════════════════════════════════════════════════════════════

def normalize_name(name: str) -> str:
    """
    Кириллические слова, написанные целиком капсом → нижний регистр
    (первое слово — с заглавной). Латиница, цифры, бренды — без изменений.
    """
    if not name:
        return name
    words = name.split()
    result = []
    for i, word in enumerate(words):
        letters = [c for c in word if c.isalpha()]
        is_cyrillic_caps = (
            letters
            and all(c.isupper() for c in letters)
            and all('Ѐ' <= c <= 'ӿ' for c in letters)
        )
        if is_cyrillic_caps:
            result.append(word.capitalize() if i == 0 else word.lower())
        else:
            result.append(word)
    return ' '.join(result)


def clean_description(text: str) -> str:
    """
    Удаляет символы вне допустимых диапазонов: ASCII, кириллица,
    типографские символы «»–—''„"". Убирает иероглифы и прочие экзотические символы.
    """
    if not text:
        return text
    ALLOWED_EXTRA = {
        0x00AB, 0x00BB,                    # « »
        0x2013, 0x2014,                    # – —
        0x2018, 0x2019, 0x201C, 0x201D,    # ' ' " "
        0x201E,                            # „
    }
    out = []
    for c in text:
        cp = ord(c)
        if cp <= 0x007F or 0x0400 <= cp <= 0x04FF or cp in ALLOWED_EXTRA:
            out.append(c)
    return ''.join(out)


# ══════════════════════════════════════════════════════════════════════════════
# Поиск фото по артикулу
# ══════════════════════════════════════════════════════════════════════════════

def get_photo_urls(article: str,
                   index: dict[str, str]) -> tuple[str | None, list[str]]:
    """
    Главное фото:    стем содержит '1_info'
    Дополнительные: все остальные стемы, содержащие артикул (любой суффикс)
    """
    art    = article.lower()
    main   = None
    extras: list[tuple[str, str]] = []

    for stem, url in index.items():
        if art not in stem:
            continue
        if '1_info' in stem:
            main = url
        else:
            extras.append((stem, url))

    extras.sort(key=lambda x: x[0])
    return main, [u for _, u in extras]


# ══════════════════════════════════════════════════════════════════════════════
# Основная логика
# ══════════════════════════════════════════════════════════════════════════════

def parse_applicability(text: str) -> list[tuple[str, str, str]]:
    if not text:
        return []
    result = []
    for part in re.split(r';\s*', str(text).strip()):
        part = part.strip()
        if not part:
            continue
        p = part.find('(')
        if p == -1:
            w = part.split()
            result.append((w[0] if w else '', ' '.join(w[1:]), ''))
        else:
            w = part[:p].strip().split()
            result.append((
                w[0] if w else '',
                ' '.join(w[1:]) if len(w) > 1 else '',
                part[p:].strip(),
            ))
    return result


def find_matching_files() -> dict[str, tuple[str, str]]:
    source_map: dict[str, str] = {}
    for fname in os.listdir(SOURCE_DIR):
        if fname.lower().endswith('.xlsx'):
            source_map[fname[:-5]] = os.path.join(SOURCE_DIR, fname)

    upload_map: dict[str, str] = {}
    for fname in os.listdir(UPLOAD_DIR):
        if fname.lower().endswith('.xlsx'):
            cat = re.sub(r'_\d{2}\.\d{2}\.\d{4}$', '', fname[:-5])
            upload_map[cat] = os.path.join(UPLOAD_DIR, fname)

    matched = {c: (source_map[c], upload_map[c])
               for c in sorted(source_map) if c in upload_map}

    if PROCESS_ONLY:
        matched = {k: v for k, v in matched.items() if k in PROCESS_ONLY}

    return matched


def read_source_rows(path: str) -> list[tuple]:
    """Читает строки, пропуская позиции без цены продажи."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        article    = row[S['article']]
        price_sell = row[S['price_sell']]
        if article is None:
            continue
        if not price_sell:          # пустая цена продажи — пропускаем
            skipped += 1
            continue
        rows.append(row)
    wb.close()
    if skipped:
        print(f"  Пропущено без цены продажи: {skipped}")
    return rows


def safe_get(row: tuple, key: str):
    idx = S.get(key)
    return row[idx] if idx is not None and idx < len(row) else None


def clear_data_rows(ws) -> None:
    for row in ws.iter_rows(min_row=DATA_START):
        for cell in row:
            cell.value = None


def fill_template(
    category:        str,
    source_rows:     list[tuple],
    upload_path:     str,
    output_path:     str,
    file_index:      dict[str, str],
    local_photo_set: set[str] | None = None,
) -> tuple[int, int, list[str], list[str]]:
    wb      = openpyxl.load_workbook(upload_path)
    ws_main = wb['Шаблон']
    ws_app  = wb['Для чего подходит']

    # Определяем столбцы динамически из заголовка шаблона
    t = detect_template_columns(ws_main)

    clear_data_rows(ws_main)
    clear_data_rows(ws_app)

    app_rows:   list[tuple] = []
    miss_main:  list[str]   = []
    skipped_no_photo: int   = 0
    row_out = DATA_START

    for row in source_rows:
        article      = safe_get(row, 'article')
        article_con  = f"{article}-con" if article else article
        name         = normalize_name(safe_get(row, 'name'))
        brand        = safe_get(row, 'brand')
        price_sell   = safe_get(row, 'price_sell')
        oem          = safe_get(row, 'oem')
        applicability= safe_get(row, 'applicability')
        alt_articles = safe_get(row, 'alt_articles')
        description  = clean_description(safe_get(row, 'description'))
        weight       = safe_get(row, 'weight')
        length       = safe_get(row, 'length')
        width        = safe_get(row, 'width')
        height       = safe_get(row, 'height')

        # ── Фото ──────────────────────────────────────────────────────────
        main_url:  str | None = None
        extra_urls: list[str] = []

        # Локальный фильтр: нет фото в img/<category>/ → пропускаем, Диск не смотрим
        if local_photo_set is not None and article:
            art_lower = str(article).lower()
            if not any(art_lower in stem for stem in local_photo_set):
                miss_main.append(str(article))
                skipped_no_photo += 1
                continue

        if file_index and article:
            main_url, extra_urls = get_photo_urls(article, file_index)
            if not main_url:
                # Нет главного фото — позицию не выгружаем
                miss_main.append(article)
                skipped_no_photo += 1
                continue

        # ── Запись в шаблон ───────────────────────────────────────────────
        r = row_out
        row_out += 1

        ws_main.cell(r, t['article']).value      = article_con
        ws_main.cell(r, t['name']).value         = name
        ws_main.cell(r, t['price']).value        = price_sell
        ws_main.cell(r, t['vat']).value          = VAT_VALUE
        ws_main.cell(r, t['weight_pkg']).value   = weight
        ws_main.cell(r, t['width_pkg']).value    = width
        ws_main.cell(r, t['height_pkg']).value   = height
        ws_main.cell(r, t['length_pkg']).value   = length
        ws_main.cell(r, t['brand']).value        = brand
        ws_main.cell(r, t['type_']).value        = category
        ws_main.cell(r, t['part_number']).value  = article
        ws_main.cell(r, t['description']).value  = description
        ws_main.cell(r, t['oem']).value          = oem
        ws_main.cell(r, t['alt_articles']).value = alt_articles

        if main_url:
            ws_main.cell(r, t['photo_main']).value = main_url
        if extra_urls:
            ws_main.cell(r, t['photo_add']).value = '\n'.join(extra_urls)

        # ── Применяемость ─────────────────────────────────────────────────
        for br, md, mod in parse_applicability(applicability):
            app_rows.append((article, br, md, mod))

    for j, (art, br, md, mod) in enumerate(app_rows):
        r = DATA_START + j
        ws_app.cell(r, A['article']).value = art
        ws_app.cell(r, A['brand']).value   = br
        ws_app.cell(r, A['model']).value   = md
        ws_app.cell(r, A['mod']).value     = mod

    wb.save(output_path)
    filled = row_out - DATA_START
    return filled, len(app_rows), miss_main, skipped_no_photo


# ══════════════════════════════════════════════════════════════════════════════
# Точка входа
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ── Индекс фото ───────────────────────────────────────────────────────
    # При YADISK_HAS_SUBFOLDERS глобальный индекс не нужен —
    # для каждой категории индекс строится отдельно внутри цикла.
    file_index: dict[str, str] = {}

    if YADISK_PUBLIC_URL and not YADISK_HAS_SUBFOLDERS:
        file_index = fetch_yadisk_public(YADISK_PUBLIC_URL)
        print(f"  Проиндексировано: {len(file_index)} файлов\n")
    elif YADISK_DISK_PATH:
        file_index = fetch_yadisk_private(YADISK_DISK_PATH)
        if not file_index:
            print("  ⚠  Файловый индекс пуст — фото НЕ будут заполнены!")
            print("  Проверь: залогинен ли Яндекс в Chrome-профиле,")
            print("  и верен ли путь YADISK_DISK_PATH.\n")
        else:
            print()
    elif not YADISK_PUBLIC_URL and not YADISK_DISK_PATH:
        print("  YADISK не задан — фото будут пропущены.\n")

    # ── Категории ─────────────────────────────────────────────────────────
    matched = find_matching_files()
    if not matched:
        print("Совпадающих файлов не найдено.")
        if PROCESS_ONLY:
            print(f"  Искали: {PROCESS_ONLY}")
            print(f"  Источник: {SOURCE_DIR}")
        return

    print(f"Категорий: {len(matched)}\n")

    total_p, total_a, total_skip = 0, 0, 0
    all_mm:  list[tuple[str, str]] = []
    errors:  list[tuple[str, str]] = []

    for category, (src_path, upl_path) in matched.items():
        out_fname = os.path.basename(upl_path)
        out_path  = os.path.join(OUTPUT_DIR, out_fname)
        print(f"  ▶  {category} ...", end=' ', flush=True)
        try:
            # Локальный фильтр фото
            local_set = build_local_photo_set(category) if IMG_LOCAL_DIR else None

            # Для режима с подпапками — строим индекс фото под конкретную категорию
            if category in CATEGORY_YADISK_URLS:
                cat_index = fetch_yadisk_public(CATEGORY_YADISK_URLS[category])
            elif YADISK_PUBLIC_URL and YADISK_HAS_SUBFOLDERS:
                cat_index = fetch_yadisk_public(YADISK_PUBLIC_URL, subfolder=category)
            else:
                cat_index = file_index

            rows = read_source_rows(src_path)
            np, na, mm, skipped = fill_template(
                category, rows, upl_path, out_path, cat_index, local_set)
            total_p    += np
            total_a    += na
            total_skip += skipped
            status = f"{np} товаров, {na} применяемости"
            if skipped:
                status += f"  [пропущено без фото: {skipped}]"
            print(status)
            for art in mm:
                all_mm.append((category, art))
        except Exception as exc:
            errors.append((category, str(exc)))
            print(f"ОШИБКА: {exc}")

    # ── Отчёт ─────────────────────────────────────────────────────────────
    print(f"\n{'─' * 60}")
    print(f"Итого товаров:             {total_p}")
    print(f"Итого применяемости:       {total_a}")
    if total_skip:
        print(f"Пропущено (нет фото):      {total_skip}")
    print(f"Результат:                 {OUTPUT_DIR}")

    if all_mm:
        print(f"\n⚠  Нет главного фото — пропущено {len(all_mm)} артикулов:")
        for cat, art in all_mm:
            print(f"     {art}  ({cat})")

    if errors:
        print(f"\n✗  Ошибки ({len(errors)}):")
        for cat, msg in errors:
            print(f"     {cat}: {msg}")


if __name__ == '__main__':
    main()
